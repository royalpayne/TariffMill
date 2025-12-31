#!/usr/bin/env python3
# ==============================================================================
# TariffMill - Customs Documentation Processing System
# ==============================================================================
# Copyright (c) 2024-2025 Process Logic Labs, LLC. All Rights Reserved.
#
# PROPRIETARY AND CONFIDENTIAL
#
# This software and its source code are the exclusive property of Process Logic Labs, LLC
# and are protected by copyright law and international treaties. Unauthorized
# reproduction, distribution, or disclosure of this software, in whole or in
# part, is strictly prohibited.
#
# This software is l icensed, not sold. Use of this software is subject to the
# End User License Agreement (EULA) provided with this software.
#
# NO WARRANTY: This software is provided "as is" without warranty of any kind,
# express or implied, including but not limited to the warranties of
# merchantability, fitness for a particular purpose, and noninfringement.
#
# For licensing inquiries, contact: admin@processlogiclabs.com
# ==============================================================================
#
# Professional PyQt5 application for automating invoice processing, parts
# management, and Section 232 tariff compliance tracking.
#
# Features:
#   - Invoice processing with two-stage verification workflow
#   - Parts database management with HTS code mapping
#   - Section 232 tariff compliance calculations
#   - Excel export for customs documentation
#   - Invoice mapping profile management
#   - Real-time data validation and status feedback
# ==============================================================================

# ==============================================================================
# SPLASH SCREEN SUPPORT
# ==============================================================================
# PyInstaller native splash screen shows immediately when exe launches.
# This code updates the splash text and closes it when app is ready.

import sys
import os

# Hide console window on Windows immediately at startup
if sys.platform == 'win32':
    import ctypes
    kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
    user32 = ctypes.WinDLL('user32', use_last_error=True)
    hwnd = kernel32.GetConsoleWindow()
    if hwnd:
        user32.ShowWindow(hwnd, 0)  # SW_HIDE = 0

def update_splash(message):
    """Update PyInstaller native splash screen text."""
    # Native splash is disabled - this is now a no-op
    pass

def close_splash():
    """Close PyInstaller native splash screen."""
    # Native splash is disabled - this is now a no-op
    pass

# Update splash during module loading
if __name__ == "__main__":
    update_splash("Loading modules...")

# ==============================================================================
# Application Constants
# ==============================================================================

APP_NAME = "TariffMill"
DB_NAME = "tariffmill.db"

# Copyright and Legal Information
COPYRIGHT_YEAR = "2025-2026"
COPYRIGHT_HOLDER = "Process Logic Labs, LLC"
COPYRIGHT_EMAIL = "admin@processlogiclabs.com"
COPYRIGHT_NOTICE = f"Copyright (c) {COPYRIGHT_YEAR} {COPYRIGHT_HOLDER}. All Rights Reserved."
LICENSE_TYPE = "Proprietary"

# Import version from version.py
try:
    from TariffMill.version import get_version
    VERSION = get_version()
except ImportError:
    try:
        from version import get_version
        VERSION = get_version()
    except ImportError:
        # Fallback if version.py is not available
        VERSION = "v0.93.4"

# ==============================================================================
# Heavy Imports - These take time to load
# ==============================================================================
if __name__ == "__main__":
    update_splash("Loading standard libraries...")

import json
import time
import re
import shutil
import traceback
import subprocess
import configparser
import webbrowser
import urllib.request
import urllib.error
import xml.etree.ElementTree as ET
from xml.dom import minidom
from pathlib import Path
from datetime import datetime, timedelta
from threading import Thread

if __name__ == "__main__":
    update_splash("Loading pandas...")

import pandas as pd
import sqlite3

if __name__ == "__main__":
    update_splash("Loading PyQt5 components...")

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QMimeData, pyqtSignal, QTimer, QSize, QEventLoop, QRect, QSettings, QThread, QThreadPool, QRunnable, QObject, QUrl
from PyQt5.QtGui import QColor, QFont, QDrag, QKeySequence, QIcon, QPixmap, QPainter, QDoubleValidator, QCursor, QPen, QTextCursor, QTextCharFormat, QSyntaxHighlighter, QTextFormat, QDesktopServices
from PyQt5.QtSvg import QSvgRenderer

if __name__ == "__main__":
    update_splash("Loading Excel support...")

from openpyxl.styles import Font as ExcelFont, Alignment
import tempfile

# ==============================================================================
# Application Logger
# ==============================================================================
# In-memory logging system with timestamp, level, and message tracking.
# Maintains up to 1000 log entries for debugging and user feedback.
class ErrorLogger:
    """Centralized logging for application events, errors, and diagnostics."""

    def __init__(self):
        self.logs = []

    def log(self, level, message):
        """Record a log entry with timestamp and severity level."""
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        entry = f"[{ts}] {level.upper():7} | {message}"
        self.logs.append(entry)
        if len(self.logs) > 1000:
            self.logs = self.logs[-1000:]
        print(entry)

    def info(self, msg):
        self.log("info", msg)

    def debug(self, msg):
        self.log("debug", msg)

    def warning(self, msg):
        self.log("warning", msg)

    def error(self, msg):
        self.log("ERROR", msg)
        if hasattr(sys, 'exc_info') and sys.exc_info()[0]:
            tb = traceback.format_exc()
            for line in tb.splitlines():
                self.log("TRACE", line)

    def success(self, msg):
        self.log("success", msg)

    def get_logs(self):
        return "\n".join(self.logs)

logger = ErrorLogger()

# ==============================================================================
# Update Checker
# ==============================================================================
# Checks GitHub releases for new versions and provides update notifications.
# Uses the GitHub API to fetch the latest release information.

GITHUB_REPO = "ProcessLogicLabs/TariffMill"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
GITHUB_RELEASES_URL = f"https://github.com/{GITHUB_REPO}/releases"

class UpdateChecker:
    """Checks GitHub for application updates"""
    
    def __init__(self, current_version):
        self.current_version = current_version
        self.latest_version = None
        self.release_url = None
        self.release_notes = None
        self.download_url = None
        self.error = None
    
    def parse_version(self, version_str):
        """Parse version string to tuple for comparison (e.g., 'v0.61.0' -> (0, 61, 0))

        Also handles git describe format like 'v0.90.1-6-gaa8bef5' by extracting
        the base version (v0.90.1) and commits ahead count (6).
        """
        try:
            # Remove 'v' prefix if present
            clean = version_str.lstrip('v').lstrip('V')

            # Handle git describe format: v0.90.1-6-gaa8bef5
            # Split on '-' and check if it's a git describe format
            if '-' in clean:
                parts = clean.split('-')
                # Check if second part is a number (commits ahead) and third starts with 'g' (git hash)
                if len(parts) >= 2 and parts[1].isdigit():
                    # Git describe format - extract base version and commits ahead
                    base_version = parts[0]
                    commits_ahead = int(parts[1])
                    version_parts = base_version.split('.')
                    # Return tuple with commits ahead as 4th element for proper comparison
                    base_tuple = [int(p) for p in version_parts[:3]]
                    # Pad to 3 elements if needed
                    while len(base_tuple) < 3:
                        base_tuple.append(0)
                    # Add commits ahead - a dev version (commits > 0) is NEWER than the release
                    base_tuple.append(commits_ahead)
                    return tuple(base_tuple)

            # Standard version format: 0.90.1
            parts = clean.split('.')
            version_tuple = [int(p) for p in parts[:3]]
            # Pad to 3 elements if needed
            while len(version_tuple) < 3:
                version_tuple.append(0)
            # Add 0 for commits ahead (this is a release version)
            version_tuple.append(0)
            return tuple(version_tuple)
        except (ValueError, AttributeError):
            return (0, 0, 0, 0)
    
    def check_for_updates(self):
        """
        Check GitHub for the latest release.
        Returns: (has_update, latest_version, release_url, release_notes, download_url, error)
        """
        try:
            # Create request with user-agent header (required by GitHub API)
            request = urllib.request.Request(
                GITHUB_API_URL,
                headers={'User-Agent': f'TariffMill/{self.current_version}'}
            )
            
            with urllib.request.urlopen(request, timeout=10) as response:
                data = json.loads(response.read().decode('utf-8'))
            
            self.latest_version = data.get('tag_name', '')
            self.release_url = data.get('html_url', GITHUB_RELEASES_URL)
            self.release_notes = data.get('body', 'No release notes available.')
            
            # Find Windows installer download URL from assets
            # Prioritize Setup/Installer exe over standalone exe
            assets = data.get('assets', [])
            standalone_url = None
            for asset in assets:
                name = asset.get('name', '').lower()
                if name.endswith('.exe'):
                    url = asset.get('browser_download_url')
                    # Prefer Setup installer
                    if 'setup' in name or 'installer' in name:
                        self.download_url = url
                        break
                    else:
                        # Keep standalone as fallback
                        standalone_url = url

            # Use standalone exe if no installer found
            if not self.download_url and standalone_url:
                self.download_url = standalone_url
            
            # Compare versions
            current_tuple = self.parse_version(self.current_version)
            latest_tuple = self.parse_version(self.latest_version)
            
            has_update = latest_tuple > current_tuple
            
            logger.info(f"Update check: current={self.current_version}, latest={self.latest_version}, update_available={has_update}")
            
            return (has_update, self.latest_version, self.release_url, 
                    self.release_notes, self.download_url, None)
            
        except urllib.error.URLError as e:
            self.error = f"Network error: {str(e)}"
            logger.warning(f"Update check failed: {self.error}")
            return (False, None, None, None, None, self.error)
        except json.JSONDecodeError as e:
            self.error = f"Invalid response from GitHub: {str(e)}"
            logger.warning(f"Update check failed: {self.error}")
            return (False, None, None, None, None, self.error)
        except Exception as e:
            self.error = f"Update check failed: {str(e)}"
            logger.warning(f"Update check failed: {self.error}")
            return (False, None, None, None, None, self.error)


# ==============================================================================
# License Management
# ==============================================================================
# Handles license validation with Gumroad integration, trial period management,
# and hybrid online/offline validation.

# License Configuration - Update GUMROAD_PRODUCT_ID after creating Gumroad product
GUMROAD_PRODUCT_ID = "lRReBpPi8qMTg0cfHl2_3A=="
GUMROAD_PRODUCT_URL = "https://payne181.gumroad.com/l/ellnff"
GUMROAD_VERIFY_URL = "https://api.gumroad.com/v2/licenses/verify"
TRIAL_DAYS = 30
OFFLINE_GRACE_DAYS = 7  # Days to allow offline use before requiring re-validation

class LicenseManager:
    """Manages license validation with Gumroad integration and trial period"""

    def __init__(self, db_path):
        self.db_path = db_path
        self.license_key = None
        self.license_email = None
        self.license_status = 'unknown'  # 'trial', 'active', 'expired', 'invalid'
        self.trial_start_date = None
        self.last_verified = None
        self.error = None

    def _get_config(self, key):
        """Get a value from app_config table"""
        try:
            conn = sqlite3.connect(str(self.db_path))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", (key,))
            row = c.fetchone()
            conn.close()
            return row[0] if row else None
        except Exception as e:
            logger.warning(f"Failed to get config {key}: {e}")
            return None

    def _set_config(self, key, value):
        """Set a value in app_config table"""
        try:
            conn = sqlite3.connect(str(self.db_path))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)", (key, str(value)))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logger.warning(f"Failed to set config {key}: {e}")
            return False

    def get_machine_id(self):
        """Generate a unique machine identifier for tracking (not for locking)"""
        import hashlib
        import platform

        # Combine various system identifiers (no subprocess calls to avoid console flash)
        identifiers = [
            platform.node(),  # Computer network name
            platform.machine(),  # Machine type
            platform.processor(),  # Processor info
            platform.system(),  # OS name
            platform.release(),  # OS release
        ]

        # Try to get Windows-specific identifiers without subprocess
        if sys.platform == 'win32':
            try:
                import winreg
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                    r"SOFTWARE\Microsoft\Cryptography")
                machine_guid, _ = winreg.QueryValueEx(key, "MachineGuid")
                winreg.CloseKey(key)
                identifiers.append(machine_guid)
            except:
                pass

        # Create hash of combined identifiers
        combined = '|'.join(identifiers)
        return hashlib.sha256(combined.encode()).hexdigest()[:32]

    def get_trial_start_date(self):
        """Get the trial start date, initializing it if this is first launch"""
        stored = self._get_config('trial_start_date')
        if stored:
            try:
                return datetime.fromisoformat(stored)
            except:
                pass

        # First launch - initialize trial
        now = datetime.now()
        self._set_config('trial_start_date', now.isoformat())
        logger.info(f"Trial period started: {now.isoformat()}")
        return now

    def get_trial_days_remaining(self):
        """Calculate remaining days in trial period"""
        start_date = self.get_trial_start_date()
        elapsed = datetime.now() - start_date
        remaining = TRIAL_DAYS - elapsed.days
        return max(0, remaining)

    def is_trial_expired(self):
        """Check if trial period has ended"""
        return self.get_trial_days_remaining() <= 0

    def get_stored_license(self):
        """Retrieve stored license information"""
        self.license_key = self._get_config('license_key')
        self.license_email = self._get_config('license_email')
        last_verified = self._get_config('license_last_verified')
        if last_verified:
            try:
                self.last_verified = datetime.fromisoformat(last_verified)
            except:
                self.last_verified = None
        return self.license_key

    def store_license(self, license_key, email=None, purchase_data=None):
        """Save license information to database"""
        self._set_config('license_key', license_key)
        if email:
            self._set_config('license_email', email)
        self._set_config('license_activated_date', datetime.now().isoformat())
        self._set_config('license_last_verified', datetime.now().isoformat())
        if purchase_data:
            self._set_config('license_purchase_data', json.dumps(purchase_data))
        self.license_key = license_key
        self.license_email = email
        self.last_verified = datetime.now()
        logger.info(f"License stored successfully")

    def validate_online(self, license_key):
        """Verify license with Gumroad API"""
        if not GUMROAD_PRODUCT_ID:
            # Product ID not configured - skip online validation
            logger.warning("Gumroad product ID not configured, skipping online validation")
            return None, "Product not configured for online validation"

        try:
            data = urllib.parse.urlencode({
                'product_id': GUMROAD_PRODUCT_ID,
                'license_key': license_key,
                'increment_uses_count': 'false'
            }).encode('utf-8')

            request = urllib.request.Request(
                GUMROAD_VERIFY_URL,
                data=data,
                method='POST',
                headers={'User-Agent': f'TariffMill/{VERSION}'}
            )

            with urllib.request.urlopen(request, timeout=10) as response:
                result = json.loads(response.read().decode('utf-8'))

            if result.get('success'):
                purchase = result.get('purchase', {})

                # Check if refunded or disputed
                if purchase.get('refunded') or purchase.get('disputed'):
                    return False, "License has been refunded or disputed"

                # Check subscription status for memberships
                if purchase.get('subscription_cancelled_at') or purchase.get('subscription_failed_at'):
                    return False, "Subscription is no longer active"

                # Valid license
                email = purchase.get('email', '')
                return True, {'email': email, 'purchase': purchase}
            else:
                return False, result.get('message', 'Invalid license key')

        except urllib.error.HTTPError as e:
            if e.code == 404:
                return False, "Invalid license key"
            return None, f"Server error: {e.code}"
        except urllib.error.URLError as e:
            return None, f"Network error: {str(e)}"
        except Exception as e:
            return None, f"Validation error: {str(e)}"

    def validate_offline(self):
        """Check if stored license is still valid for offline use"""
        if not self.license_key:
            self.get_stored_license()

        if not self.license_key:
            return False, "No license key stored"

        if not self.last_verified:
            return False, "License has never been verified online"

        # Check if within offline grace period
        days_since_verified = (datetime.now() - self.last_verified).days
        if days_since_verified <= OFFLINE_GRACE_DAYS:
            return True, f"Offline mode ({OFFLINE_GRACE_DAYS - days_since_verified} days remaining)"

        return False, "Offline grace period expired, please connect to internet to re-verify"

    def validate_license(self, license_key=None):
        """
        Hybrid validation: try online first, fall back to offline.
        Returns: (is_valid, message)
        """
        key_to_check = license_key or self.license_key or self.get_stored_license()

        if not key_to_check:
            return False, "No license key provided"

        # Try online validation first
        online_result, online_data = self.validate_online(key_to_check)

        if online_result is True:
            # Valid online - update stored data
            email = online_data.get('email', '') if isinstance(online_data, dict) else None
            self.store_license(key_to_check, email, online_data)
            self.license_status = 'active'
            return True, "License validated successfully"

        elif online_result is False:
            # Explicitly invalid
            self.license_status = 'invalid'
            return False, online_data  # online_data contains error message

        else:
            # Online check failed (network issue) - try offline
            logger.info(f"Online validation unavailable: {online_data}, trying offline")
            offline_result, offline_msg = self.validate_offline()
            if offline_result:
                self.license_status = 'active'
                return True, offline_msg
            else:
                self.license_status = 'invalid'
                return False, f"Online: {online_data}. Offline: {offline_msg}"

    def activate_license(self, license_key):
        """Activate a new license key"""
        license_key = license_key.strip()
        if not license_key:
            return False, "Please enter a license key"

        # Validate the license
        is_valid, message = self.validate_license(license_key)

        if is_valid:
            logger.info(f"License activated successfully")
            return True, "License activated successfully!"
        else:
            logger.warning(f"License activation failed: {message}")
            return False, message

    def get_license_status(self):
        """
        Determine current license status.
        Returns: ('trial', days_remaining) or ('active', None) or ('expired', None)
        """
        # Check for valid license first
        stored_key = self.get_stored_license()
        if stored_key:
            is_valid, _ = self.validate_license(stored_key)
            if is_valid:
                return 'active', None

        # No valid license - check trial
        if not self.is_trial_expired():
            days = self.get_trial_days_remaining()
            return 'trial', days

        # Trial expired and no valid license
        return 'expired', None


# ==============================================================================
# Authentication System
# ==============================================================================
# Remote user authentication with GitHub-hosted user list, password hashing,
# role-based access control, and offline credential caching.

# GitHub API URL for private repo access (use API endpoint, not raw.githubusercontent)
AUTH_CONFIG_URL = "https://api.github.com/repos/process-logic-labs/TariffMill_Config/contents/auth_users.json"
# Personal Access Token for private repo (read-only, contents scope)
# Set TARIFFMILL_GITHUB_TOKEN environment variable or use local auth_users.json
# Generate at: https://github.com/settings/tokens/new?scopes=repo
AUTH_GITHUB_TOKEN = os.environ.get('TARIFFMILL_GITHUB_TOKEN', '')

class AuthenticationManager:
    """Manages user authentication with Windows domain auth, remote user list, and local caching."""

    # Allowed Windows domains for auto-authentication
    ALLOWED_DOMAINS = ['DMUSA']

    def __init__(self, db_path):
        self.db_path = db_path
        self.current_user = None
        self.current_role = None
        self.current_name = None
        self.is_authenticated = False
        self.auth_method = None  # 'windows' or 'password'

    def try_windows_auth(self) -> tuple:
        """Try to authenticate using Windows domain credentials.

        Returns (success: bool, message: str, user_data: dict or None)
        """
        import os

        try:
            username = os.environ.get('USERNAME', '')
            domain = os.environ.get('USERDOMAIN', '')

            if not username or not domain:
                return False, "Windows credentials not available", None

            # Check if domain is in allowed list
            if domain.upper() not in [d.upper() for d in self.ALLOWED_DOMAINS]:
                logger.debug(f"Domain {domain} not in allowed list")
                return False, f"Domain {domain} not authorized for auto-login", None

            # Build the Windows user identifier
            windows_user = f"{domain.upper()}\\{username.lower()}"

            # Fetch remote users
            remote_users = self._fetch_remote_users()
            if remote_users is None:
                return False, "Could not fetch user list", None

            # Look for Windows user in user list (case-insensitive)
            for user_key, user_data in remote_users.items():
                # Check if this is a Windows-style user (contains backslash)
                if '\\' in user_key:
                    if user_key.upper() == windows_user.upper():
                        # Windows user found
                        role = user_data.get('role', 'user')
                        name = user_data.get('name', username)

                        self.current_user = user_key
                        self.current_role = role
                        self.current_name = name
                        self.is_authenticated = True
                        self.auth_method = 'windows'

                        # Store last login
                        self._set_config('last_auth_user', user_key)
                        self._set_config('last_auth_time', datetime.now().isoformat())
                        self._set_config('last_auth_method', 'windows')

                        logger.info(f"Windows auth successful for {windows_user}")
                        return True, f"Welcome, {name}!", user_data

            # Windows user not in allowed list
            logger.debug(f"Windows user {windows_user} not found in user list")
            return False, f"Windows user {windows_user} not authorized", None

        except Exception as e:
            logger.warning(f"Windows auth failed: {e}")
            return False, str(e), None

    def get_windows_user_info(self) -> tuple:
        """Get current Windows domain and username.

        Returns (domain, username) or (None, None) if not available.
        """
        import os
        domain = os.environ.get('USERDOMAIN', '')
        username = os.environ.get('USERNAME', '')
        return (domain, username) if domain and username else (None, None)

    def _hash_password(self, password: str, salt: str = None) -> tuple:
        """Hash a password using SHA-256 with salt.

        Returns (hash, salt) tuple.
        """
        import hashlib
        import secrets

        if salt is None:
            salt = secrets.token_hex(16)

        # Combine password with salt and hash
        salted = f"{salt}{password}".encode('utf-8')
        password_hash = hashlib.sha256(salted).hexdigest()

        return password_hash, salt

    def _verify_password(self, password: str, stored_hash: str, salt: str) -> bool:
        """Verify a password against stored hash and salt."""
        computed_hash, _ = self._hash_password(password, salt)
        return computed_hash == stored_hash

    def _get_config(self, key: str) -> str:
        """Get a value from app_config table."""
        try:
            conn = sqlite3.connect(str(self.db_path))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", (key,))
            row = c.fetchone()
            conn.close()
            return row[0] if row else None
        except Exception as e:
            logger.warning(f"Failed to get auth config {key}: {e}")
            return None

    def _set_config(self, key: str, value: str):
        """Set a value in app_config table."""
        try:
            conn = sqlite3.connect(str(self.db_path))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)", (key, str(value)))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logger.warning(f"Failed to set auth config {key}: {e}")
            return False

    def _fetch_remote_users(self) -> dict:
        """Fetch user list from remote GitHub-hosted JSON (supports private repos).

        Expected JSON format:
        {
            "users": {
                "admin@processlogiclabs.com": {
                    "password_hash": "abc123...",
                    "salt": "def456...",
                    "role": "admin",
                    "name": "Admin User"
                },
                "customer@example.com": {
                    "password_hash": "ghi789...",
                    "salt": "jkl012...",
                    "role": "user",
                    "name": "Customer Name"
                }
            }
        }
        """
        import urllib.request
        import urllib.error
        import base64

        try:
            headers = {
                'User-Agent': 'TariffMill/1.0',
                'Accept': 'application/vnd.github.v3+json',
                'Cache-Control': 'no-cache'
            }

            # Add authorization header if token is configured
            if AUTH_GITHUB_TOKEN and not AUTH_GITHUB_TOKEN.startswith('ghp_REPLACE'):
                headers['Authorization'] = f'token {AUTH_GITHUB_TOKEN}'

            req = urllib.request.Request(AUTH_CONFIG_URL, headers=headers)

            with urllib.request.urlopen(req, timeout=10) as response:
                api_response = json.loads(response.read().decode('utf-8'))

                # GitHub API returns content as base64 encoded
                if 'content' in api_response:
                    # Decode base64 content from GitHub API response
                    content_b64 = api_response['content'].replace('\n', '')
                    content_bytes = base64.b64decode(content_b64)
                    data = json.loads(content_bytes.decode('utf-8'))
                else:
                    # Direct JSON response (for raw URLs)
                    data = api_response

                logger.info("Successfully fetched remote user list")
                return data.get('users', {})

        except urllib.error.HTTPError as e:
            if e.code == 401:
                logger.warning("GitHub authentication failed - check AUTH_GITHUB_TOKEN")
            elif e.code == 404:
                logger.warning("Auth config file not found in GitHub repo")
            else:
                logger.warning(f"GitHub API error: {e.code} {e.reason}")
            return self._load_local_auth_file()
        except urllib.error.URLError as e:
            logger.warning(f"Failed to fetch remote user list: {e}")
            return self._load_local_auth_file()
        except json.JSONDecodeError as e:
            logger.warning(f"Invalid auth users JSON: {e}")
            return self._load_local_auth_file()
        except Exception as e:
            logger.warning(f"Error fetching remote user list: {e}")
            return self._load_local_auth_file()

    def _load_local_auth_file(self) -> dict:
        """Load user list from local auth_users.json file as fallback.

        This is used when remote fetch fails or for development/testing.
        """
        try:
            # Check in project root directory (parent of Tariffmill folder)
            local_auth_path = Path(__file__).parent.parent / 'auth_users.json'
            if local_auth_path.exists():
                with open(local_auth_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    logger.info(f"Loaded local auth file: {local_auth_path}")
                    return data.get('users', {})

            # Also check in same directory as script
            alt_path = Path(__file__).parent / 'auth_users.json'
            if alt_path.exists():
                with open(alt_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    logger.info(f"Loaded local auth file: {alt_path}")
                    return data.get('users', {})

            logger.warning("No local auth_users.json found")
            return None
        except Exception as e:
            logger.warning(f"Failed to load local auth file: {e}")
            return None

    def _cache_credentials(self, email: str, password_hash: str, salt: str, role: str, name: str):
        """Cache user credentials locally for offline authentication."""
        try:
            cached_users = self._get_config('cached_auth_users')
            if cached_users:
                users = json.loads(cached_users)
            else:
                users = {}

            users[email.lower()] = {
                'password_hash': password_hash,
                'salt': salt,
                'role': role,
                'name': name,
                'cached_at': datetime.now().isoformat()
            }

            self._set_config('cached_auth_users', json.dumps(users))
            logger.debug(f"Cached credentials for {email}")
        except Exception as e:
            logger.warning(f"Failed to cache credentials: {e}")

    def _get_cached_user(self, email: str) -> dict:
        """Get cached user credentials for offline authentication."""
        try:
            cached_users = self._get_config('cached_auth_users')
            if cached_users:
                users = json.loads(cached_users)
                return users.get(email.lower())
        except Exception as e:
            logger.warning(f"Failed to get cached user: {e}")
        return None

    def authenticate(self, email: str, password: str) -> tuple:
        """Authenticate user against remote user list or cached credentials.

        Returns (success: bool, message: str, role: str or None)
        """
        email = email.strip().lower()
        if not email or not password:
            return False, "Email and password are required", None

        # Try remote authentication first
        remote_users = self._fetch_remote_users()

        if remote_users is not None:
            # Online authentication
            user_data = None
            for user_email, data in remote_users.items():
                if user_email.lower() == email:
                    user_data = data
                    break

            if user_data:
                stored_hash = user_data.get('password_hash', '')
                salt = user_data.get('salt', '')
                role = user_data.get('role', 'user')
                name = user_data.get('name', email)

                if self._verify_password(password, stored_hash, salt):
                    # Cache credentials for offline use
                    self._cache_credentials(email, stored_hash, salt, role, name)

                    self.current_user = email
                    self.current_role = role
                    self.current_name = name
                    self.is_authenticated = True
                    self.auth_method = 'password'

                    # Store last login
                    self._set_config('last_auth_user', email)
                    self._set_config('last_auth_time', datetime.now().isoformat())
                    self._set_config('last_auth_method', 'password')

                    logger.info(f"User {email} authenticated successfully (online)")
                    return True, f"Welcome, {name}!", role
                else:
                    return False, "Invalid email or password", None
            else:
                return False, "User not found", None

        else:
            # Offline authentication - use cached credentials
            logger.info("Remote auth unavailable, trying cached credentials")
            cached_user = self._get_cached_user(email)

            if cached_user:
                stored_hash = cached_user.get('password_hash', '')
                salt = cached_user.get('salt', '')
                role = cached_user.get('role', 'user')
                name = cached_user.get('name', email)

                if self._verify_password(password, stored_hash, salt):
                    self.current_user = email
                    self.current_role = role
                    self.is_authenticated = True

                    self._set_config('last_auth_user', email)
                    self._set_config('last_auth_time', datetime.now().isoformat())

                    logger.info(f"User {email} authenticated successfully (offline/cached)")
                    return True, f"Welcome, {name}! (Offline mode)", role
                else:
                    return False, "Invalid email or password", None
            else:
                return False, "Cannot authenticate: No network connection and no cached credentials", None

    def logout(self):
        """Clear current authentication state."""
        self.current_user = None
        self.current_role = None
        self.is_authenticated = False
        logger.info("User logged out")

    def is_admin(self) -> bool:
        """Check if current user has admin role."""
        return self.is_authenticated and self.current_role == 'admin'

    def get_last_user(self) -> str:
        """Get the last authenticated user email for convenience."""
        return self._get_config('last_auth_user') or ''

    @staticmethod
    def generate_password_hash(password: str) -> dict:
        """Utility to generate password hash for adding new users.

        Returns dict with 'password_hash' and 'salt' to add to auth_users.json
        """
        import hashlib
        import secrets

        salt = secrets.token_hex(16)
        salted = f"{salt}{password}".encode('utf-8')
        password_hash = hashlib.sha256(salted).hexdigest()

        return {
            'password_hash': password_hash,
            'salt': salt
        }


# ==============================================================================
# Self-Update Mechanism
# ==============================================================================
# Detects if the exe is running from a different location (e.g., Downloads)
# than the installed location, and offers to update the installed version.

INSTALL_PATH_FILE = Path(os.environ.get('APPDATA', '')) / "TariffMill" / "install_path.txt"

def get_installed_path():
    """Get the stored installation path, if any."""
    try:
        if INSTALL_PATH_FILE.exists():
            return Path(INSTALL_PATH_FILE.read_text().strip())
    except Exception:
        pass
    return None

def save_installed_path(path):
    """Save the current exe path as the installation path."""
    try:
        INSTALL_PATH_FILE.parent.mkdir(parents=True, exist_ok=True)
        INSTALL_PATH_FILE.write_text(str(path))
    except Exception as e:
        print(f"Warning: Could not save install path: {e}")

def check_and_perform_self_update():
    """
    Check if we're running from a different location than installed.
    If so, offer to update the installed version.
    Returns True if update was performed (caller should exit), False otherwise.
    """
    if not getattr(sys, 'frozen', False):
        return False  # Only for frozen exe

    current_exe = Path(sys.executable)
    installed_path = get_installed_path()

    # If no install path saved, or we're running from the installed location, continue normally
    if installed_path is None:
        return False

    # Normalize paths for comparison
    try:
        current_exe_resolved = current_exe.resolve()
        installed_path_resolved = installed_path.resolve()
    except Exception:
        return False

    # If running from the same location, continue normally
    if current_exe_resolved == installed_path_resolved:
        return False

    # Check if installed path still exists and is a TariffMill exe
    if not installed_path.exists():
        return False

    # We're running from a different location - likely an update download
    # Show a simple message box asking if user wants to update
    from PyQt5.QtWidgets import QApplication, QMessageBox

    # Need a QApplication for message boxes
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)

    # Get version info for display
    current_version = VERSION

    msg = QMessageBox()
    msg.setIcon(QMessageBox.Question)
    msg.setWindowTitle("Update TariffMill")
    msg.setText(f"A TariffMill installation was found at:\n{installed_path.parent}\n\n"
                f"Would you like to update it with this version ({current_version})?")
    msg.setInformativeText("The application will restart after updating.")
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
    msg.setDefaultButton(QMessageBox.Yes)

    # Add "Run from here" option
    run_here_btn = msg.addButton("Run from here", QMessageBox.ActionRole)

    result = msg.exec_()

    if msg.clickedButton() == run_here_btn:
        # User wants to run from current location - save this as new install path
        save_installed_path(current_exe)
        return False

    if result == QMessageBox.Yes:
        # Perform the update
        try:
            return perform_update(current_exe, installed_path)
        except Exception as e:
            QMessageBox.critical(None, "Update Failed",
                               f"Could not update the installation:\n{str(e)}\n\n"
                               "You can manually copy the new exe to replace the old one.")
            return False
    elif result == QMessageBox.No:
        # Don't update, just exit
        return True
    else:
        # Cancel - exit without doing anything
        return True

def perform_update(source_exe, target_exe):
    """
    Perform the update by copying source folder to target folder.
    For PyInstaller directory-mode, copies entire folder including _internal.
    Uses a batch script to complete the copy after we exit.
    Returns True if update initiated (caller should exit).
    """
    import tempfile

    source_exe = Path(source_exe)
    target_exe = Path(target_exe)
    source_dir = source_exe.parent
    target_dir = target_exe.parent

    # Create a batch script that will:
    # 1. Wait for this process to exit
    # 2. Copy the entire source folder to target (including _internal)
    # 3. Launch the updated exe
    # 4. Delete itself

    batch_content = f'''@echo off
:: Wait for the updater process to exit
timeout /t 2 /nobreak >nul

:: Remove old _internal folder if it exists
if exist "{target_dir}\\_internal" (
    rmdir /s /q "{target_dir}\\_internal"
)

:: Copy the new _internal folder
if exist "{source_dir}\\_internal" (
    xcopy /E /I /Y "{source_dir}\\_internal" "{target_dir}\\_internal"
    if errorlevel 1 (
        echo Update failed - could not copy _internal folder
        pause
        exit /b 1
    )
)

:: Copy the new exe
copy /Y "{source_exe}" "{target_exe}"
if errorlevel 1 (
    echo Update failed - could not copy exe file
    pause
    exit /b 1
)

:: Launch the updated application
start "" "{target_exe}"

:: Delete this batch file
del "%~f0"
'''

    # Write batch file to temp directory
    batch_path = Path(tempfile.gettempdir()) / "tariffmill_update.bat"
    batch_path.write_text(batch_content)

    # Run the batch file hidden
    import subprocess
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = 0  # SW_HIDE

    subprocess.Popen(
        ['cmd', '/c', str(batch_path)],
        startupinfo=startupinfo,
        creationflags=subprocess.CREATE_NO_WINDOW
    )

    return True

# ==============================================================================
# Application Paths
# ==============================================================================
# Handles path resolution for both PyInstaller-bundled executables and
# development/script execution modes. Ensures resource files are found
# regardless of deployment method.

if getattr(sys, 'frozen', False):
    # Running as compiled executable (PyInstaller)
    INSTALL_DIR = Path(sys.executable).parent
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller stores data files in _MEIPASS (_internal folder)
        RESOURCES_DIR = Path(sys._MEIPASS) / "Resources"
    else:
        RESOURCES_DIR = INSTALL_DIR / "Resources"
    # Use user's AppData for writable data (not Program Files)
    BASE_DIR = Path(os.environ.get('LOCALAPPDATA', Path.home())) / "TariffMill"
else:
    # Running as Python script
    INSTALL_DIR = Path(__file__).parent
    BASE_DIR = INSTALL_DIR
    RESOURCES_DIR = BASE_DIR / "Resources"

# Backward compatibility alias
TEMP_RESOURCES_DIR = RESOURCES_DIR
INPUT_DIR = BASE_DIR / "Tariffmill_Input"
OUTPUT_DIR = BASE_DIR / "Tariffmill_Output"

def get_processed_dir(input_dir: Path) -> Path:
    """Get the processed directory path based on input directory.

    Creates a subfolder named 'Processed' inside the input directory.
    e.g., /home/user/Downloads -> /home/user/Downloads/Processed
    """
    return input_dir / "Processed"

PROCESSED_DIR = get_processed_dir(INPUT_DIR)

# Configuration files for data mappings
MAPPING_FILE = BASE_DIR / "column_mapping.json"
SHIPMENT_MAPPING_FILE = BASE_DIR / "shipment_mapping.json"

# Create required directories (user data directories only, not RESOURCES_DIR)
for p in (BASE_DIR, INPUT_DIR, OUTPUT_DIR, PROCESSED_DIR):
    p.mkdir(parents=True, exist_ok=True)

# Copy bundled database to user location on first run (for frozen exe)
# This ensures the pre-populated tariff_232 and sec_232_actions tables are available
# Database is copied to AppData folder (BASE_DIR) which is user-writable
if getattr(sys, 'frozen', False):
    import shutil
    bundled_db = TEMP_RESOURCES_DIR / DB_NAME
    local_db = BASE_DIR / DB_NAME  # Store in AppData, not Program Files
    if bundled_db.exists() and not local_db.exists():
        try:
            shutil.copy2(bundled_db, local_db)
            print(f"Copied bundled database to {local_db}")
        except Exception as e:
            print(f"Warning: Could not copy bundled database: {e}")

# ==============================================================================
# Shared Configuration File
# ==============================================================================
# The config.ini file stores settings shared across all users (e.g., database path).
# User-specific settings remain in QSettings (Windows registry).

CONFIG_FILE = BASE_DIR / "config.ini"

def load_shared_config():
    """Load shared configuration from config.ini file."""
    config = configparser.ConfigParser()
    if CONFIG_FILE.exists():
        config.read(str(CONFIG_FILE))
    return config

def save_shared_config(config):
    """Save shared configuration to config.ini file."""
    with open(str(CONFIG_FILE), 'w') as f:
        config.write(f)

def get_database_path():
    """
    Get the database path from shared config or use default.
    Supports platform-specific paths (linux_path, windows_path) for cross-platform use.

    Returns:
        Path object pointing to the SQLite database file.
    """
    config = load_shared_config()

    # Check for platform-specific paths first
    is_windows = sys.platform == 'win32'
    platform_key = 'windows_path' if is_windows else 'linux_path'

    if config.has_option('Database', platform_key):
        platform_path = config.get('Database', platform_key)
        if platform_path and Path(platform_path).exists():
            return Path(platform_path)

    # Fall back to generic 'path' setting
    if config.has_option('Database', 'path'):
        custom_path = config.get('Database', 'path')
        if custom_path and Path(custom_path).exists():
            return Path(custom_path)

    # Default to user data folder (AppData on Windows when frozen)
    return BASE_DIR / DB_NAME

def set_database_path(path, platform=None):
    """
    Set a custom database path in shared config.

    Args:
        path: Path string to the database file (can be network path).
        platform: Optional - 'linux', 'windows', or None for generic path.
    """
    config = load_shared_config()
    if not config.has_section('Database'):
        config.add_section('Database')

    if platform == 'linux':
        config.set('Database', 'linux_path', str(path))
    elif platform == 'windows':
        config.set('Database', 'windows_path', str(path))
    else:
        config.set('Database', 'path', str(path))

    save_shared_config(config)

def get_platform_database_paths():
    """
    Get configured database paths for each platform.

    Returns:
        Dict with 'linux_path', 'windows_path', and 'path' (generic) values.
    """
    config = load_shared_config()
    result = {
        'linux_path': config.get('Database', 'linux_path', fallback=''),
        'windows_path': config.get('Database', 'windows_path', fallback=''),
        'path': config.get('Database', 'path', fallback=''),
    }
    return result

# Database location - reads from config.ini or defaults to local
DB_PATH = get_database_path()

# ==============================================================================
# Per-User Settings (QSettings - Windows Registry)
# ==============================================================================
# These settings are stored per-user in the Windows Registry under
# HKEY_CURRENT_USER\Software\TariffMill\TariffMill
# This allows each user to have their own personal preferences while
# sharing the same database for parts data, profiles, etc.

def get_user_settings():
    """Get QSettings object for per-user settings stored in Windows Registry."""
    return QSettings("TariffMill", "TariffMill")

def get_user_setting(key, default=None):
    """
    Get a per-user setting from Windows Registry.

    Args:
        key: Setting key (e.g., 'theme', 'font_size', 'column_widths')
        default: Default value if setting doesn't exist

    Returns:
        The stored value or default
    """
    settings = get_user_settings()
    return settings.value(key, default)

def set_user_setting(key, value):
    """
    Save a per-user setting to Windows Registry.

    Args:
        key: Setting key
        value: Value to store
    """
    settings = get_user_settings()
    settings.setValue(key, value)
    settings.sync()

def get_user_setting_bool(key, default=False):
    """Get a boolean per-user setting (handles string 'true'/'false' from registry)."""
    value = get_user_setting(key, default)
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ('true', '1', 'yes')
    return bool(value)

def get_user_setting_int(key, default=0):
    """Get an integer per-user setting."""
    value = get_user_setting(key, default)
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

def get_user_setting_float(key, default=0.0):
    """Get a float per-user setting."""
    value = get_user_setting(key, default)
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def get_db_config(key, default=None):
    """
    Get a value from app_config table in the database.

    Args:
        key: Config key (e.g., 'last_folder_profile', 'last_map_profile')
        default: Default value if key doesn't exist

    Returns:
        The stored value or default
    """
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("SELECT value FROM app_config WHERE key = ?", (key,))
        row = c.fetchone()
        conn.close()
        return row[0] if row else default
    except Exception as e:
        logger.warning(f"Failed to get db config {key}: {e}")
        return default

def set_db_config(key, value):
    """
    Set a value in app_config table in the database.

    Args:
        key: Config key
        value: Value to store

    Returns:
        True if successful, False otherwise
    """
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)", (key, str(value)))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.warning(f"Failed to set db config {key}: {e}")
        return False

def get_theme_color_key(base_key, theme_name=None):
    """
    Generate a theme-specific color settings key.

    Args:
        base_key: Base color key (e.g., 'preview_steel_color')
        theme_name: Theme name (if None, uses current saved theme)

    Returns:
        Theme-specific key (e.g., 'preview_steel_color_fusion_dark')
    """
    if theme_name is None:
        theme_name = get_user_setting('theme', 'Fusion (Light)')
    # Normalize theme name for use as key suffix
    theme_suffix = theme_name.lower().replace(' ', '_').replace('(', '').replace(')', '')
    return f"{base_key}_{theme_suffix}"

def get_theme_color(base_key, default_color, theme_name=None):
    """
    Get a color setting for the current or specified theme.

    Args:
        base_key: Base color key (e.g., 'preview_steel_color')
        default_color: Default color hex value
        theme_name: Theme name (if None, uses current saved theme)

    Returns:
        Color hex string
    """
    theme_key = get_theme_color_key(base_key, theme_name)
    return get_user_setting(theme_key, default_color)

def set_theme_color(base_key, color_value, theme_name=None):
    """
    Save a color setting for the current or specified theme.

    Args:
        base_key: Base color key (e.g., 'preview_steel_color')
        color_value: Color hex value to save
        theme_name: Theme name (if None, uses current saved theme)
    """
    theme_key = get_theme_color_key(base_key, theme_name)
    set_user_setting(theme_key, color_value)

def is_widget_valid(widget):
    """
    Check if a Qt widget is still valid (not deleted).
    Returns False if the widget's underlying C++ object has been deleted.
    """
    try:
        import sip
        return widget is not None and not sip.isdeleted(widget)
    except ImportError:
        # If sip is not available, try accessing the widget
        try:
            if widget is None:
                return False
            # Try to access a property - will raise RuntimeError if deleted
            widget.objectName()
            return True
        except RuntimeError:
            return False

def get_232_info(hts_code):
    """
    Lookup Section 232 tariff information for an HTS code.

    Args:
        hts_code: HTS code string (with or without dots)

    Returns:
        Tuple of (material, declaration_code, smelt_flag) where:
        - material: Material type (e.g., "Steel", "Aluminum") or None
        - declaration_code: Tariff code (e.g., "08" for Steel)
        - smelt_flag: "Y" for materials requiring smelting declaration, "" otherwise

    Process:
        1. Queries tariff_232 table for 10-digit and 8-digit HTS matches
        2. Falls back to hardcoded HTS prefixes for common materials
        3. Returns None if material not found
    """
    # Handle pandas NA, None, empty string, or NaN values
    try:
        if hts_code is None or pd.isna(hts_code) or str(hts_code).strip() == '':
            return None, "", ""
    except (ValueError, TypeError):
        # pd.isna() can raise ValueError for some types
        if not hts_code or str(hts_code).strip() == '':
            return None, "", ""

    # Normalize HTS code: remove dots, strip whitespace, convert to uppercase
    hts_clean = str(hts_code).replace(".", "").strip().upper()
    hts_8 = hts_clean[:8]
    hts_10 = hts_clean[:10]

    # Query tariff database
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("SELECT material, declaration_required FROM tariff_232 WHERE hts_code = ?", (hts_10,))
        row = c.fetchone()
        if not row and len(hts_clean) >= 8:
            c.execute("SELECT material, declaration_required FROM tariff_232 WHERE hts_code = ?", (hts_8,))
            row = c.fetchone()
        conn.close()
        if row:
            material = row[0]
            dec_code = row[1] if row[1] else ""
            dec_type = dec_code.split(" - ")[0] if " - " in dec_code else dec_code
            smelt_flag = "Y" if material in ["Aluminum", "Wood", "Copper"] else ""
            return material, dec_type, smelt_flag
    except Exception as e:
        logger.error(f"Error querying tariff_232 for HTS {hts_clean}: {e}")

    # No match found in tariff_232 database
    return None, "", ""

def get_hts_qty_unit(hts_code):
    """
    Lookup the quantity unit (Uom 1) for an HTS code from hts_units table.

    Args:
        hts_code: HTS code string (with or without dots)

    Returns:
        qty_unit string (e.g., "KG", "NO", "M2") or empty string if not found
    """
    if not hts_code:
        return ""

    # Normalize HTS code: remove dots, strip whitespace
    hts_clean = str(hts_code).replace(".", "").strip()

    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        # Try exact 10-digit match first
        c.execute("SELECT qty_unit FROM hts_units WHERE hts_code = ?", (hts_clean[:10],))
        row = c.fetchone()
        if not row and len(hts_clean) >= 8:
            # Try 8-digit match
            c.execute("SELECT qty_unit FROM hts_units WHERE hts_code = ?", (hts_clean[:8],))
            row = c.fetchone()
        conn.close()
        if row:
            return row[0]
    except Exception as e:
        logger.error(f"Error querying hts_units for HTS {hts_clean}: {e}")

    return ""

# ==============================================================================
# Database Initialization
# ==============================================================================
# Creates tables if they don't exist: parts_master, tariff_232, sec_232_actions,
# mid_table, mapping_profiles, and app_config.

def init_database():
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS parts_master (
            part_number TEXT PRIMARY KEY, description TEXT, hts_code TEXT, country_origin TEXT,
            mid TEXT, client_code TEXT, steel_ratio REAL DEFAULT 1.0, non_steel_ratio REAL DEFAULT 0.0, last_updated TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS tariff_232 (
            hts_code TEXT PRIMARY KEY,
            material TEXT,
            classification TEXT,
            chapter TEXT,
            chapter_description TEXT,
            declaration_required TEXT,
            notes TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS sec_232_actions (
            tariff_no TEXT PRIMARY KEY,
            action TEXT,
            description TEXT,
            advalorem_rate TEXT,
            effective_date TEXT,
            expiration_date TEXT,
            specific_rate TEXT,
            additional_declaration TEXT,
            note TEXT,
            link TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS mapping_profiles (
            profile_name TEXT PRIMARY KEY, mapping_json TEXT, created_date TEXT, header_row INTEGER DEFAULT 1
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS app_config (
            key TEXT PRIMARY KEY, value TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS output_column_mappings (
            profile_name TEXT PRIMARY KEY,
            mapping_json TEXT,
            created_date TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS mid_table (
            mid TEXT PRIMARY KEY,
            manufacturer_name TEXT,
            customer_id TEXT,
            related_parties TEXT DEFAULT 'N'
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS hts_units (
            hts_code TEXT PRIMARY KEY,
            qty_unit TEXT
        )""")

        # Create profile_links table for linking input map profiles to export profiles
        c.execute("""CREATE TABLE IF NOT EXISTS profile_links (
            input_profile_name TEXT PRIMARY KEY,
            export_profile_name TEXT
        )""")

        # Create folder_profiles table for input/output folder location profiles
        c.execute("""CREATE TABLE IF NOT EXISTS folder_profiles (
            profile_name TEXT PRIMARY KEY,
            input_folder TEXT,
            output_folder TEXT,
            created_date TEXT
        )""")

        # Create billing_records table for tracking exports and generating invoices
        c.execute("""CREATE TABLE IF NOT EXISTS billing_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_number TEXT NOT NULL,
            export_date TEXT NOT NULL,
            export_time TEXT NOT NULL,
            file_name TEXT,
            line_count INTEGER DEFAULT 0,
            total_value REAL DEFAULT 0.0,
            hts_codes_used TEXT,
            folder_profile TEXT,
            map_profile TEXT,
            mid TEXT,
            user_name TEXT,
            machine_id TEXT,
            processing_time_ms INTEGER DEFAULT 0,
            invoice_sent INTEGER DEFAULT 0,
            invoice_month TEXT
        )""")

        # Create billing_settings table for email and rate configuration
        c.execute("""CREATE TABLE IF NOT EXISTS billing_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )""")

        # Migration: Add manufacturer_name and customer_id columns to mid_table if they don't exist
        try:
            c.execute("PRAGMA table_info(mid_table)")
            columns = [col[1] for col in c.fetchall()]
            if 'manufacturer_name' not in columns:
                c.execute("ALTER TABLE mid_table ADD COLUMN manufacturer_name TEXT")
                logger.info("Added manufacturer_name column to mid_table")
            if 'customer_id' not in columns:
                c.execute("ALTER TABLE mid_table ADD COLUMN customer_id TEXT")
                logger.info("Added customer_id column to mid_table")
            if 'related_parties' not in columns:
                c.execute("ALTER TABLE mid_table ADD COLUMN related_parties TEXT DEFAULT 'N'")
                logger.info("Added related_parties column to mid_table")
        except Exception as e:
            logger.warning(f"Failed to check/add columns to mid_table: {e}")

        # Migration: Add client_code column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'client_code' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN client_code TEXT")
                logger.info("Added client_code column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add client_code column: {e}")

        # Migration: Add header_row column to mapping_profiles if it doesn't exist
        try:
            c.execute("PRAGMA table_info(mapping_profiles)")
            columns = [col[1] for col in c.fetchall()]
            if 'header_row' not in columns:
                c.execute("ALTER TABLE mapping_profiles ADD COLUMN header_row INTEGER DEFAULT 1")
                logger.info("Added header_row column to mapping_profiles")
        except Exception as e:
            logger.warning(f"Failed to check/add header_row column: {e}")

        # Migration: Add qty_unit column to parts_master (renamed from cbp_qty1)
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'qty_unit' not in columns:
                if 'cbp_qty1' in columns:
                    # Rename existing cbp_qty1 column to qty_unit
                    c.execute("ALTER TABLE parts_master RENAME COLUMN cbp_qty1 TO qty_unit")
                    logger.info("Renamed cbp_qty1 column to qty_unit in parts_master")
                else:
                    # Add new qty_unit column
                    c.execute("ALTER TABLE parts_master ADD COLUMN qty_unit TEXT")
                    logger.info("Added qty_unit column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add qty_unit column: {e}")

        # Migration: Add aluminum_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'aluminum_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN aluminum_ratio REAL DEFAULT 0.0")
                logger.info("Added aluminum_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add aluminum_ratio column: {e}")

        # Migration: Add copper_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'copper_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN copper_ratio REAL DEFAULT 0.0")
                logger.info("Added copper_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add copper_ratio column: {e}")

        # Migration: Add wood_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'wood_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN wood_ratio REAL DEFAULT 0.0")
                logger.info("Added wood_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add wood_ratio column: {e}")

        # Migration: Add auto_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'auto_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN auto_ratio REAL DEFAULT 0.0")
                logger.info("Added auto_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add auto_ratio column: {e}")

        # Migration: Add country_of_melt column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'country_of_melt' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN country_of_melt TEXT")
                logger.info("Added country_of_melt column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add country_of_melt column: {e}")

        # Migration: Add country_of_cast column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'country_of_cast' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN country_of_cast TEXT")
                logger.info("Added country_of_cast column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add country_of_cast column: {e}")

        # Migration: Add country_of_smelt column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'country_of_smelt' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN country_of_smelt TEXT")
                logger.info("Added country_of_smelt column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add country_of_smelt column: {e}")

        # Migration: Add Sec301_Exclusion_Tariff column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'Sec301_Exclusion_Tariff' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN Sec301_Exclusion_Tariff TEXT")
                logger.info("Added Sec301_Exclusion_Tariff column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add Sec301_Exclusion_Tariff column: {e}")

        # Migration: Add hts_verified column to parts_master if it doesn't exist
        # This column stores verification status and date of HTS code against hts.db
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'hts_verified' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN hts_verified TEXT")
                logger.info("Added hts_verified column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add hts_verified column: {e}")

        # Migration: Reset column visibility settings if we have outdated settings
        # (This handles upgrades from versions with fewer columns)
        try:
            c.execute("SELECT COUNT(*) FROM app_config WHERE key LIKE 'preview_col_visible_%'")
            count_row = c.fetchone()
            saved_count = count_row[0] if count_row else 0
            if 0 < saved_count < 17:
                # We have old settings - clear them to reset all columns to visible
                c.execute("DELETE FROM app_config WHERE key LIKE 'preview_col_visible_%'")
                logger.info(f"Cleared outdated column visibility settings (had {saved_count}, need 17)")
        except Exception as e:
            logger.warning(f"Failed to check/reset column visibility: {e}")

        # Migration: Clear corrupted column widths (any column with 0 width)
        try:
            import json
            c.execute("SELECT value FROM app_config WHERE key = 'column_widths'")
            row = c.fetchone()
            if row:
                widths = json.loads(row[0])
                if any(w == 0 for w in widths.values()):
                    c.execute("DELETE FROM app_config WHERE key = 'column_widths'")
                    logger.info("Cleared corrupted column widths (had 0-width columns)")
        except Exception as e:
            logger.warning(f"Failed to check/reset column widths: {e}")

        # Migration: Fix corrupted parts_master ratios
        # Due to a bug in populate_parts_table, steel_ratio and aluminum_ratio columns got swapped.
        # The aluminum_ratio column contains what should be steel_ratio values.
        # Fix: Swap steel_ratio and aluminum_ratio values, then set aluminum_ratio to 0.
        try:
            c.execute("SELECT value FROM app_config WHERE key = 'ratios_migration_v1'")
            if not c.fetchone():
                # Count affected rows - those where aluminum_ratio > 0 and steel_ratio appears to have non_steel values
                c.execute("""
                    SELECT COUNT(*) FROM parts_master 
                    WHERE aluminum_ratio > 0.0 
                    AND (copper_ratio IS NULL OR copper_ratio = 0.0)
                    AND (wood_ratio IS NULL OR wood_ratio = 0.0)
                """)
                affected = c.fetchone()[0]
                
                if affected > 0:
                    # Swap: steel_ratio should get aluminum_ratio value, 
                    # non_steel_ratio should get steel_ratio value (the non-232 portion),
                    # aluminum_ratio should be 0 (these aren't actually aluminum products)
                    c.execute("""
                        UPDATE parts_master 
                        SET steel_ratio = aluminum_ratio,
                            non_steel_ratio = steel_ratio,
                            aluminum_ratio = 0.0
                        WHERE aluminum_ratio > 0.0 
                        AND (copper_ratio IS NULL OR copper_ratio = 0.0)
                        AND (wood_ratio IS NULL OR wood_ratio = 0.0)
                    """)
                    logger.info(f"Fixed {affected} parts with swapped ratio data (steel/aluminum swap corrected)")
                
                # Mark migration as complete
                c.execute("INSERT INTO app_config (key, value) VALUES ('ratios_migration_v1', '1')")
        except Exception as e:
            logger.warning(f"Failed to fix corrupted ratios: {e}")

        # Migration: Convert ratios from 0.0-1.0 to 0-100 percentages
        # This makes the database values more intuitive for users
        try:
            c.execute("SELECT value FROM app_config WHERE key = 'ratios_to_percentage_v1'")
            if not c.fetchone():
                # Convert all ratio columns from decimal (0.0-1.0) to percentage (0-100)
                # Only convert rows where at least one non-zero ratio is in decimal format (0 < x <= 1)
                c.execute("""
                    UPDATE parts_master
                    SET steel_ratio = steel_ratio * 100,
                        aluminum_ratio = aluminum_ratio * 100,
                        copper_ratio = copper_ratio * 100,
                        wood_ratio = wood_ratio * 100,
                        auto_ratio = auto_ratio * 100,
                        non_steel_ratio = non_steel_ratio * 100
                    WHERE (steel_ratio > 0 AND steel_ratio <= 1.0)
                       OR (aluminum_ratio > 0 AND aluminum_ratio <= 1.0)
                       OR (copper_ratio > 0 AND copper_ratio <= 1.0)
                       OR (wood_ratio > 0 AND wood_ratio <= 1.0)
                       OR (auto_ratio > 0 AND auto_ratio <= 1.0)
                       OR (non_steel_ratio > 0 AND non_steel_ratio <= 1.0)
                """)
                migrated_count = c.execute("SELECT changes()").fetchone()[0]
                logger.info(f"Converted {migrated_count} parts from ratio (0-1) to percentage (0-100)")

                # Mark migration as complete
                c.execute("INSERT INTO app_config (key, value) VALUES ('ratios_to_percentage_v1', '1')")
        except Exception as e:
            logger.warning(f"Failed to convert ratios to percentages: {e}")

        # Migration: Update output mapping profiles to replace CalcWtNet/Pcs with Qty1/Qty2
        try:
            c.execute("SELECT value FROM app_config WHERE key = 'qty_columns_migration_v1'")
            if not c.fetchone():
                import json
                # Update mapping_profiles table
                c.execute("SELECT profile_name, mapping_json FROM mapping_profiles")
                profiles = c.fetchall()
                updated = 0
                for profile_name, mapping_json in profiles:
                    if mapping_json:
                        try:
                            data = json.loads(mapping_json)
                            changed = False
                            # Update column_order if present
                            if 'column_order' in data:
                                new_order = []
                                for col in data['column_order']:
                                    if col == 'CalcWtNet':
                                        new_order.append('Qty1')
                                        changed = True
                                    elif col == 'Pcs':
                                        new_order.append('Qty2')
                                        changed = True
                                    else:
                                        new_order.append(col)
                                # Add Qty1/Qty2 if not present
                                if 'Qty1' not in new_order:
                                    # Insert after MID if possible
                                    if 'MID' in new_order:
                                        mid_idx = new_order.index('MID')
                                        new_order.insert(mid_idx + 1, 'Qty1')
                                        changed = True
                                if 'Qty2' not in new_order:
                                    if 'Qty1' in new_order:
                                        qty1_idx = new_order.index('Qty1')
                                        new_order.insert(qty1_idx + 1, 'Qty2')
                                        changed = True
                                data['column_order'] = new_order
                            # Update output_columns if present
                            if 'output_columns' in data:
                                if 'CalcWtNet' in data['output_columns']:
                                    data['output_columns']['Qty1'] = data['output_columns'].pop('CalcWtNet')
                                    changed = True
                                if 'Pcs' in data['output_columns']:
                                    data['output_columns']['Qty2'] = data['output_columns'].pop('Pcs')
                                    changed = True
                                if 'Qty1' not in data['output_columns']:
                                    data['output_columns']['Qty1'] = 'Qty1'
                                    changed = True
                                if 'Qty2' not in data['output_columns']:
                                    data['output_columns']['Qty2'] = 'Qty2'
                                    changed = True
                            if changed:
                                c.execute("UPDATE mapping_profiles SET mapping_json = ? WHERE profile_name = ?",
                                         (json.dumps(data), profile_name))
                                updated += 1
                        except:
                            pass
                if updated > 0:
                    logger.info(f"Updated {updated} output mapping profiles: CalcWtNet->Qty1, Pcs->Qty2")
                c.execute("INSERT INTO app_config (key, value) VALUES ('qty_columns_migration_v1', '1')")
        except Exception as e:
            logger.warning(f"Failed to migrate output columns: {e}")

        # Migration: Update output_column_mappings table to replace CalcWtNet/Pcs with Qty1/Qty2
        try:
            c.execute("SELECT value FROM app_config WHERE key = 'output_mappings_qty_migration_v1'")
            if not c.fetchone():
                import json
                c.execute("SELECT profile_name, mapping_json FROM output_column_mappings")
                profiles = c.fetchall()
                updated = 0
                for profile_name, mapping_json in profiles:
                    if mapping_json:
                        try:
                            data = json.loads(mapping_json)
                            changed = False
                            # Update column_order if present
                            if 'column_order' in data:
                                new_order = []
                                for col in data['column_order']:
                                    if col == 'CalcWtNet':
                                        new_order.append('Qty1')
                                        changed = True
                                    elif col == 'Pcs':
                                        new_order.append('Qty2')
                                        changed = True
                                    else:
                                        new_order.append(col)
                                data['column_order'] = new_order
                            # Update column_mapping if present
                            if 'column_mapping' in data:
                                if 'CalcWtNet' in data['column_mapping']:
                                    data['column_mapping']['Qty1'] = data['column_mapping'].pop('CalcWtNet')
                                    changed = True
                                if 'Pcs' in data['column_mapping']:
                                    data['column_mapping']['Qty2'] = data['column_mapping'].pop('Pcs')
                                    changed = True
                            # Update column_visibility if present
                            if 'column_visibility' in data:
                                if 'CalcWtNet' in data['column_visibility']:
                                    data['column_visibility']['Qty1'] = data['column_visibility'].pop('CalcWtNet')
                                    changed = True
                                if 'Pcs' in data['column_visibility']:
                                    data['column_visibility']['Qty2'] = data['column_visibility'].pop('Pcs')
                                    changed = True
                            if changed:
                                c.execute("UPDATE output_column_mappings SET mapping_json = ? WHERE profile_name = ?",
                                         (json.dumps(data), profile_name))
                                updated += 1
                        except:
                            pass
                if updated > 0:
                    logger.info(f"Updated {updated} output_column_mappings profiles: CalcWtNet->Qty1, Pcs->Qty2")
                c.execute("INSERT INTO app_config (key, value) VALUES ('output_mappings_qty_migration_v1', '1')")
        except Exception as e:
            logger.warning(f"Failed to migrate output_column_mappings: {e}")

        # Migration: Add Section 232 Automotive tariff codes if not present
        # Always check if Auto tariffs exist, regardless of migration flag (fixes earlier bug)
        try:
            c.execute("SELECT COUNT(*) FROM tariff_232 WHERE material = 'Auto'")
            auto_count = c.fetchone()[0]

            if auto_count == 0:
                # Define automotive tariff codes from Attachment 2_Auto Parts HTS List
                # Reference: U.S. note 33, subchapter III of chapter 99, headings 9903.94.05 and 9903.94.06
                auto_tariffs = [
                        # Rubber parts (Chapter 40)
                        ('4009120020', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('4009220020', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('4009320020', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('4009420020', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('40111010', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('40111050', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('40112010', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('40121940', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('40121980', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('40122060', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('4013100010', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('4013100020', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('4016996010', 'Auto', 'Automotive Rubber Parts', '40', 'Chapter 40: Rubber and articles thereof', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Glass (Chapter 70)
                        ('70072151', 'Auto', 'Automotive Glass', '70', 'Chapter 70: Glass and glassware', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('70091000', 'Auto', 'Automotive Glass', '70', 'Chapter 70: Glass and glassware', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Iron/Steel parts (Chapter 73)
                        ('732010', 'Auto', 'Automotive Iron/Steel Parts', '73', 'Chapter 73: Articles of iron or steel', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Locks/Hardware (Chapter 83)
                        ('83012000', 'Auto', 'Automotive Locks/Hardware', '83', 'Chapter 83: Miscellaneous articles of base metal', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('83021030', 'Auto', 'Automotive Locks/Hardware', '83', 'Chapter 83: Miscellaneous articles of base metal', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('830230', 'Auto', 'Automotive Locks/Hardware', '83', 'Chapter 83: Miscellaneous articles of base metal', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Engines (Chapter 84)
                        ('84073100', 'Auto', 'Spark-Ignition Engines', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('840732', 'Auto', 'Spark-Ignition Engines', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('840733', 'Auto', 'Spark-Ignition Engines', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('840734', 'Auto', 'Spark-Ignition Engines', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84082020', 'Auto', 'Compression-Ignition Engines', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8409911040', 'Auto', 'Engine Parts', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8409991040', 'Auto', 'Engine Parts', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Pumps/Compressors/AC
                        ('84133010', 'Auto', 'Automotive Pumps', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84133090', 'Auto', 'Automotive Pumps', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84139110', 'Auto', 'Automotive Pumps', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8413919010', 'Auto', 'Automotive Pumps', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8414308030', 'Auto', 'Automotive Compressors', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84145930', 'Auto', 'Automotive Compressors', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8414596540', 'Auto', 'Automotive Compressors', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84148005', 'Auto', 'Automotive Compressors', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84152000', 'Auto', 'Automotive AC', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Filters/Lifting
                        ('84212300', 'Auto', 'Automotive Filters', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84213200', 'Auto', 'Automotive Filters', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84254900', 'Auto', 'Automotive Lifting Equipment', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84269100', 'Auto', 'Automotive Lifting Equipment', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8431100090', 'Auto', 'Automotive Lifting Equipment', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Computers
                        ('8471', 'Auto', 'Automotive Computers', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Bearings
                        ('84821010', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482105044', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482105048', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482200020', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482200030', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482200040', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482200061', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482200070', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8482200081', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84824000', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84825000', 'Auto', 'Automotive Bearings', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Transmission shafts
                        ('8483101030', 'Auto', 'Automotive Transmission Shafts', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('84831030', 'Auto', 'Automotive Transmission Shafts', '84', 'Chapter 84: Machinery and mechanical appliances', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Motors/Batteries (Chapter 85)
                        ('850132', 'Auto', 'Automotive Motors', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850133', 'Auto', 'Automotive Motors', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850134', 'Auto', 'Automotive Motors', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850140', 'Auto', 'Automotive Motors', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850151', 'Auto', 'Automotive Motors', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850152', 'Auto', 'Automotive Motors', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850710', 'Auto', 'Automotive Batteries', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('850760', 'Auto', 'Automotive Batteries', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85079040', 'Auto', 'Automotive Batteries', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85079080', 'Auto', 'Automotive Batteries', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Ignition equipment
                        ('8511100000', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85112000', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8511300040', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8511300080', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85114000', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85115000', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85118020', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85118060', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8511906020', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8511906040', 'Auto', 'Automotive Ignition', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Lighting/Signaling
                        ('85122020', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85122040', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85123000', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85124020', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85124040', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85129020', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85129060', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85129070', 'Auto', 'Automotive Lighting', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Audio/Radio
                        ('85198120', 'Auto', 'Automotive Audio', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8525601010', 'Auto', 'Automotive Video', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('852721', 'Auto', 'Automotive Radio', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('852729', 'Auto', 'Automotive Radio', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Electrical equipment
                        ('8536410005', 'Auto', 'Automotive Electrical', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('853710', 'Auto', 'Automotive Electrical', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('853720', 'Auto', 'Automotive Electrical', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8539100010', 'Auto', 'Automotive Electrical', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8539100050', 'Auto', 'Automotive Electrical', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('85443000', 'Auto', 'Automotive Wiring', '85', 'Chapter 85: Electrical machinery and equipment', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Chassis (Chapter 87)
                        ('87060003', 'Auto', 'Automotive Chassis', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87060005', 'Auto', 'Automotive Chassis', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87060015', 'Auto', 'Automotive Chassis', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87060025', 'Auto', 'Automotive Chassis', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Bodies
                        ('8707100020', 'Auto', 'Automotive Bodies', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8707100040', 'Auto', 'Automotive Bodies', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8707905020', 'Auto', 'Automotive Bodies', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8707905040', 'Auto', 'Automotive Bodies', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8707905060', 'Auto', 'Automotive Bodies', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('8707905080', 'Auto', 'Automotive Bodies', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Parts and accessories (8708)
                        ('87082100', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870822', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870829', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870830', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87084011', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87084070', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87084075', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870850', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870870', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870880', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870891', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87089360', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87089375', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870894', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('870895', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87089953', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87089955', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87089958', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('87089968', 'Auto', 'Automotive Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Trailer parts
                        ('87169050', 'Auto', 'Trailer Parts', '87', 'Chapter 87: Vehicles and parts', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Instruments (Chapter 90)
                        ('901510', 'Auto', 'Automotive Instruments', '90', 'Chapter 90: Optical and measuring instruments', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('902910', 'Auto', 'Automotive Instruments', '90', 'Chapter 90: Optical and measuring instruments', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        ('9029204080', 'Auto', 'Automotive Instruments', '90', 'Chapter 90: Optical and measuring instruments', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                        # Seats (Chapter 94)
                        ('94012000', 'Auto', 'Automotive Seats', '94', 'Chapter 94: Furniture', '12 - AUTO PARTS', 'Section 232 Automotive Tariff - 25% additional duty'),
                ]

                inserted = 0
                for tariff in auto_tariffs:
                    try:
                        c.execute("""INSERT OR IGNORE INTO tariff_232
                                    (hts_code, material, classification, chapter, chapter_description, declaration_required, notes)
                                    VALUES (?, ?, ?, ?, ?, ?, ?)""", tariff)
                        if c.rowcount > 0:
                            inserted += 1
                    except:
                        pass

                logger.info(f"Migration: Added {inserted} Section 232 Automotive tariff codes")
        except Exception as e:
            logger.warning(f"Failed to migrate auto tariffs: {e}")

        # Migration: Populate hts_units table from HTS_qty1.xlsx if empty
        try:
            c.execute("SELECT COUNT(*) FROM hts_units")
            hts_units_count = c.fetchone()[0]

            if hts_units_count == 0:
                # Try to load from bundled Excel file
                hts_qty_file = TEMP_RESOURCES_DIR / "References" / "HTS_qty1.xlsx"
                if not hts_qty_file.exists():
                    hts_qty_file = RESOURCES_DIR / "References" / "HTS_qty1.xlsx"

                if hts_qty_file.exists():
                    try:
                        hts_df = pd.read_excel(str(hts_qty_file))
                        inserted = 0
                        for _, row in hts_df.iterrows():
                            hts_code = str(row.get('Tariff No', '')).strip().replace(".", "")
                            qty_unit = str(row.get('Uom 1', '')).strip()
                            if hts_code and qty_unit:
                                try:
                                    c.execute("INSERT OR IGNORE INTO hts_units (hts_code, qty_unit) VALUES (?, ?)",
                                             (hts_code, qty_unit))
                                    if c.rowcount > 0:
                                        inserted += 1
                                except:
                                    pass
                        logger.info(f"Migration: Imported {inserted} HTS unit codes from HTS_qty1.xlsx")
                    except Exception as e:
                        logger.warning(f"Failed to import HTS units from Excel: {e}")
                else:
                    logger.debug("HTS_qty1.xlsx not found, skipping hts_units import")
        except Exception as e:
            logger.warning(f"Failed to migrate hts_units: {e}")

        # OCRMill: Create part_occurrences table for invoice line item history
        c.execute("""CREATE TABLE IF NOT EXISTS part_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_number TEXT NOT NULL,
            invoice_number TEXT,
            project_number TEXT,
            quantity REAL,
            total_price REAL,
            unit_price REAL,
            steel_pct REAL,
            steel_kg REAL,
            steel_value REAL,
            aluminum_pct REAL,
            aluminum_kg REAL,
            aluminum_value REAL,
            net_weight REAL,
            ncm_code TEXT,
            hts_code TEXT,
            processed_date TEXT,
            source_file TEXT,
            mid TEXT,
            client_code TEXT
        )""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_occurrences_part ON part_occurrences(part_number)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_occurrences_invoice ON part_occurrences(invoice_number)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_occurrences_project ON part_occurrences(project_number)")

        # OCRMill: Create hts_codes reference table for HTS lookup/matching
        c.execute("""CREATE TABLE IF NOT EXISTS hts_codes (
            hts_code TEXT PRIMARY KEY,
            description TEXT,
            suggested TEXT,
            last_updated TEXT
        )""")

        # OCRMill Migration: Add FSC fields to parts_master if they don't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'fsc_certified' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN fsc_certified TEXT DEFAULT 'N'")
                logger.info("Added fsc_certified column to parts_master")
            if 'fsc_certificate_code' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN fsc_certificate_code TEXT")
                logger.info("Added fsc_certificate_code column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to add FSC columns: {e}")

        # =====================================================================
        # LACEY ACT TABLES AND MIGRATIONS
        # =====================================================================

        # Create lacey_hts_codes table - HTS codes subject to Lacey Act requirements
        # Covers chapters 44 (Wood), 47 (Pulp), 48 (Paper), 94 (Furniture with wood)
        c.execute("""CREATE TABLE IF NOT EXISTS lacey_hts_codes (
            hts_code TEXT PRIMARY KEY,
            chapter TEXT,
            description TEXT,
            plant_type TEXT,
            requires_scientific_name TEXT DEFAULT 'Y',
            requires_country_harvest TEXT DEFAULT 'Y',
            notes TEXT
        )""")

        # Create lacey_species table - Common wood species and their scientific names
        c.execute("""CREATE TABLE IF NOT EXISTS lacey_species (
            species_id INTEGER PRIMARY KEY AUTOINCREMENT,
            common_name TEXT,
            scientific_name TEXT,
            cites_appendix TEXT,
            origin_countries TEXT,
            notes TEXT
        )""")

        # Migration: Add Lacey Act fields to parts_master if they don't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]

            lacey_columns = [
                ('lacey_applicable', "TEXT DEFAULT 'N'"),
                ('species_scientific_name', "TEXT"),
                ('species_common_name', "TEXT"),
                ('country_of_harvest', "TEXT"),
                ('percent_recycled', "REAL DEFAULT 0.0"),
                ('lacey_certificate', "TEXT"),
            ]

            for col_name, col_def in lacey_columns:
                if col_name not in columns:
                    c.execute(f"ALTER TABLE parts_master ADD COLUMN {col_name} {col_def}")
                    logger.info(f"Added {col_name} column to parts_master for Lacey Act")
        except Exception as e:
            logger.warning(f"Failed to add Lacey Act columns to parts_master: {e}")

        # Populate lacey_hts_codes with common wood/paper HTS chapters if empty
        try:
            c.execute("SELECT COUNT(*) FROM lacey_hts_codes")
            if c.fetchone()[0] == 0:
                lacey_hts_data = [
                    # Chapter 44 - Wood and articles of wood
                    ('44', '44', 'Wood and articles of wood; wood charcoal', 'Wood', 'Y', 'Y', 'Full chapter 44 coverage'),
                    # Chapter 47 - Pulp of wood
                    ('47', '47', 'Pulp of wood or other fibrous cellulosic material', 'Wood Pulp', 'Y', 'Y', 'Wood pulp products'),
                    # Chapter 48 - Paper and paperboard
                    ('48', '48', 'Paper and paperboard; articles of paper pulp', 'Paper', 'Y', 'Y', 'Paper products from wood'),
                    # Chapter 94 - Furniture (wood furniture)
                    ('9401', '94', 'Seats (wood frames)', 'Furniture', 'Y', 'Y', 'Wood frame seats'),
                    ('9403', '94', 'Other furniture (wood)', 'Furniture', 'Y', 'Y', 'Wood furniture'),
                ]
                c.executemany("""INSERT OR IGNORE INTO lacey_hts_codes
                    (hts_code, chapter, description, plant_type, requires_scientific_name, requires_country_harvest, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)""", lacey_hts_data)
                logger.info("Populated lacey_hts_codes with default HTS chapters")
        except Exception as e:
            logger.warning(f"Failed to populate lacey_hts_codes: {e}")

        # Populate lacey_species with common wood species if empty
        try:
            c.execute("SELECT COUNT(*) FROM lacey_species")
            if c.fetchone()[0] == 0:
                species_data = [
                    ('Oak', 'Quercus spp.', None, 'US, EU, CN', 'Common hardwood'),
                    ('Pine', 'Pinus spp.', None, 'US, CA, EU, CN', 'Common softwood'),
                    ('Maple', 'Acer spp.', None, 'US, CA, EU', 'Hardwood'),
                    ('Birch', 'Betula spp.', None, 'US, CA, EU, RU', 'Hardwood'),
                    ('Walnut', 'Juglans spp.', None, 'US, EU', 'Premium hardwood'),
                    ('Cherry', 'Prunus spp.', None, 'US, EU', 'Hardwood'),
                    ('Ash', 'Fraxinus spp.', None, 'US, EU', 'Hardwood'),
                    ('Beech', 'Fagus spp.', None, 'EU, US', 'Hardwood'),
                    ('Spruce', 'Picea spp.', None, 'US, CA, EU, RU', 'Softwood'),
                    ('Fir', 'Abies spp.', None, 'US, CA, EU', 'Softwood'),
                    ('Cedar', 'Cedrus spp.', None, 'US, CA', 'Softwood'),
                    ('Mahogany', 'Swietenia spp.', 'II', 'MX, BR, PE', 'CITES listed - tropical hardwood'),
                    ('Teak', 'Tectona grandis', None, 'MM, ID, IN', 'Tropical hardwood'),
                    ('Rosewood', 'Dalbergia spp.', 'II', 'BR, IN, MG', 'CITES listed - tropical hardwood'),
                    ('Ebony', 'Diospyros spp.', 'II', 'MG, IN, LK', 'CITES listed - tropical hardwood'),
                    ('Eucalyptus', 'Eucalyptus spp.', None, 'AU, BR, CL', 'Fast-growing hardwood'),
                    ('Bamboo', 'Bambusoideae', None, 'CN, VN, ID', 'Grass - may be exempt'),
                    ('Poplar', 'Populus spp.', None, 'US, CA, EU, CN', 'Fast-growing hardwood'),
                    ('MDF/Particleboard', 'Mixed species', None, 'Various', 'Composite - list primary species'),
                    ('Plywood', 'Mixed species', None, 'Various', 'Composite - list face/core species'),
                ]
                c.executemany("""INSERT OR IGNORE INTO lacey_species
                    (common_name, scientific_name, cites_appendix, origin_countries, notes)
                    VALUES (?, ?, ?, ?, ?)""", species_data)
                logger.info("Populated lacey_species with common wood species")
        except Exception as e:
            logger.warning(f"Failed to populate lacey_species: {e}")

        # =====================================================================
        # MSI-SIGMA PART NUMBER LOOKUP TABLE
        # =====================================================================
        # Maps MSI (Masonry Supply Inc.) part numbers to Sigma Corporation part numbers
        # Used for converting OCR-extracted MSI part numbers to Sigma format for export

        c.execute("""CREATE TABLE IF NOT EXISTS msi_sigma_parts (
            msi_part_number TEXT PRIMARY KEY,
            sigma_part_number TEXT NOT NULL,
            material TEXT,
            country_origin TEXT,
            hts_type TEXT,
            hts_code TEXT,
            manufacturing_class_id TEXT,
            steel_ratio REAL DEFAULT 0.0,
            last_updated TEXT
        )""")

        # Create index for faster Sigma part lookups
        c.execute("""CREATE INDEX IF NOT EXISTS idx_sigma_part ON msi_sigma_parts(sigma_part_number)""")

        conn.commit()
        conn.close()
        logger.success("Database initialized")
    except Exception as e:
        logger.error(f"Database init failed: {e}")

init_database()

# ----------------------------------------------------------------------
# Drag & Drop Components
# ----------------------------------------------------------------------
class DraggableLabel(QLabel):
    def __init__(self, text):
        super().__init__(text)
        self.setStyleSheet("background:#6b6b6b;border:2px solid #aaa;border-radius:8px;padding:12px;font-weight:bold;color:#ffffff;")
        self.setAlignment(Qt.AlignCenter)
        self.setCursor(QCursor(Qt.OpenHandCursor))  # Show hand cursor
    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            drag = QDrag(self)
            mime = QMimeData()
            mime.setText(self.text())
            drag.setMimeData(mime)
            pixmap = QPixmap(self.size())
            pixmap.fill(Qt.transparent)
            drag.setPixmap(pixmap)
            drag.exec_(Qt.CopyAction)
    def mouseMoveEvent(self, e):
        if e.buttons() == Qt.LeftButton:
            # This helps with drag detection
            pass

class DropTarget(QLabel):
    dropped = pyqtSignal(str, str)
    def __init__(self, field_key, field_name, drop_label=None):
        # Use custom drop_label if provided, otherwise use field_name
        label_text = drop_label if drop_label else field_name
        super().__init__(f"Drop {label_text} here")
        self.field_key = field_key
        # Compact style for better fit in dialog
        self.setStyleSheet("padding: 4px 8px; background: #f8f8f8; border: 1px solid #bbb; border-radius: 4px; color: #222;")
        self.setAlignment(Qt.AlignCenter)
        self.setAcceptDrops(True)
        self.setWordWrap(False)
        self.column_name = None
        self.setFixedHeight(28)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
    def dragEnterEvent(self, e): 
        if e.mimeData().hasText(): e.accept()
    def dropEvent(self, e):
        col = e.mimeData().text()
        self.column_name = col
        self.setText("✓")
        self.setStyleSheet("padding: 4px 8px; background: #d4edda; border: 1px solid #28a745; border-radius: 4px; color: #28a745; font-size: 16px; font-weight: bold;")
        self.setProperty("occupied", True)
        self.style().unpolish(self); self.style().polish(self)
        self.dropped.emit(self.field_key, col)
        e.accept()

class ForceEditableLineEdit(QLineEdit):
    """QLineEdit that forces itself to remain editable no matter what"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._force_editable = True
        # Set properties immediately
        self.setReadOnly(False)
        self.setEnabled(True)
        self.setFocusPolicy(Qt.StrongFocus)

    def setReadOnly(self, readonly):
        """Override setReadOnly to always stay editable"""
        if self._force_editable and readonly:
            return  # Ignore attempts to make readonly
        super().setReadOnly(readonly)

    def setEnabled(self, enabled):
        """Override setEnabled to always stay enabled"""
        if self._force_editable and not enabled:
            return  # Ignore attempts to disable
        super().setEnabled(enabled)

    def mousePressEvent(self, event):
        """Force editable on mouse click"""
        super(ForceEditableLineEdit, self).setReadOnly(False)
        super(ForceEditableLineEdit, self).setEnabled(True)
        self.setFocus()
        super().mousePressEvent(event)

    def focusInEvent(self, event):
        """Accept all focus events and force editable state"""
        super(ForceEditableLineEdit, self).setReadOnly(False)
        super(ForceEditableLineEdit, self).setEnabled(True)
        super().focusInEvent(event)

    def keyPressEvent(self, event):
        """Force editable before processing any key event"""
        from PyQt5.QtCore import Qt

        # For Tab keys, manually move focus to next/previous widget
        if event.key() == Qt.Key_Tab:
            # Move focus to next widget in tab order
            self.focusNextChild()
            event.accept()
            return
        elif event.key() == Qt.Key_Backtab:
            # Move focus to previous widget in tab order
            self.focusPreviousChild()
            event.accept()
            return

        super(ForceEditableLineEdit, self).setReadOnly(False)
        super(ForceEditableLineEdit, self).setEnabled(True)
        super().keyPressEvent(event)

class AutoSelectListWidget(QListWidget):
    """QListWidget that auto-selects first item when receiving focus via Tab"""
    def focusInEvent(self, event):
        """Select first item when receiving focus if nothing is selected"""
        super().focusInEvent(event)
        # Only auto-select if no item is currently selected and list has items
        if not self.currentItem() and self.count() > 0:
            self.setCurrentRow(0)

class FileDropZone(QLabel):
    """Drag-and-drop zone for importing CSV/Excel files"""
    file_dropped = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.setText("📁 Drag & Drop CSV/Excel File Here\n\nor click to browse")
        self.setAlignment(Qt.AlignCenter)
        self.setWordWrap(True)
        self.setMinimumHeight(120)
        self.setAcceptDrops(True)
        self.setCursor(Qt.PointingHandCursor)
        self.update_style(False)
        
    def update_style(self, hover=False):
        if hover:
            self.setStyleSheet("""
                QLabel {
                    background: #e3f2fd;
                    border: 3px dashed #2196F3;
                    border-radius: 10px;
                    font-weight: bold;
                    color: #1976D2;
                    padding: 20px;
                }
            """)
        else:
            self.setStyleSheet("""
                QLabel {
                    background: #f5f5f5;
                    border: 3px dashed #999;
                    border-radius: 10px;
                    font-weight: bold;
                    color: #666;
                    padding: 20px;
                }
            """)
    
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
            self.update_style(True)
        else:
            event.ignore()
    
    def dragLeaveEvent(self, event):
        self.update_style(False)
    
    def dropEvent(self, event):
        self.update_style(False)
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.lower().endswith(('.csv', '.xlsx', '.xls')):
                self.file_dropped.emit(file_path)
                event.accept()
            else:
                QMessageBox.warning(self, "Invalid File", 
                    "Please drop a CSV or Excel file (.csv, .xlsx, .xls)")
                event.ignore()
    
    def mousePressEvent(self, event):
        # Clicking opens file dialog
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select CSV/Excel File", str(INPUT_DIR),
            "CSV/Excel Files (*.csv *.xlsx *.xls)"
        )
        if file_path:
            self.file_dropped.emit(file_path)


class PDFDropZone(QLabel):
    """Drag-and-drop zone for PDF invoice files in OCRMill"""
    files_dropped = pyqtSignal(list)  # Emits list of file paths

    def __init__(self, browse_folder=None):
        super().__init__()
        self.browse_folder = browse_folder or str(Path.home())
        self.setText("Drop PDF Invoice(s) Here or click to browse")
        self.setAlignment(Qt.AlignCenter)
        self.setWordWrap(True)
        self.setAcceptDrops(True)
        self.setCursor(Qt.PointingHandCursor)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.update_style(False)

    def update_style(self, hover=False):
        if hover:
            self.setStyleSheet("""
                QLabel {
                    background: #e8f5e9;
                    border: 3px dashed #4CAF50;
                    border-radius: 10px;
                    font-weight: bold;
                    color: #2E7D32;
                    padding: 15px;
                    font-size: 13px;
                }
            """)
        else:
            self.setStyleSheet("""
                QLabel {
                    background: #fafafa;
                    border: 3px dashed #bdbdbd;
                    border-radius: 10px;
                    font-weight: bold;
                    color: #757575;
                    padding: 15px;
                    font-size: 13px;
                }
            """)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            # Check if any URL is a PDF
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith('.pdf'):
                    event.accept()
                    self.update_style(True)
                    return
        event.ignore()

    def dragLeaveEvent(self, event):
        self.update_style(False)

    def dropEvent(self, event):
        self.update_style(False)
        urls = event.mimeData().urls()
        pdf_files = []
        for url in urls:
            file_path = url.toLocalFile()
            if file_path.lower().endswith('.pdf'):
                pdf_files.append(file_path)

        if pdf_files:
            self.files_dropped.emit(pdf_files)
            event.accept()
        else:
            QMessageBox.warning(self, "Invalid File",
                "Please drop PDF file(s) only.")
            event.ignore()

    def mousePressEvent(self, event):
        # Clicking opens file dialog for multiple files
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Select PDF Invoice(s)", self.browse_folder,
            "PDF Files (*.pdf)"
        )
        if file_paths:
            self.files_dropped.emit(file_paths)

    def set_browse_folder(self, folder):
        """Update the default browse folder"""
        self.browse_folder = str(folder)


# ----------------------------------------------------------------------
# PYTHON CODE EDITOR WITH SYNTAX HIGHLIGHTING
# ----------------------------------------------------------------------
class PythonSyntaxHighlighter(QSyntaxHighlighter):
    """Syntax highlighter for Python code, similar to VS Code's dark theme."""

    def __init__(self, document):
        super().__init__(document)
        self._init_formats()
        self._init_rules()

    def _init_formats(self):
        """Initialize text formats for different syntax elements."""
        # Keywords (purple/magenta) - control flow
        self.keyword_format = QTextCharFormat()
        self.keyword_format.setForeground(QColor("#c586c0"))
        self.keyword_format.setFontWeight(QFont.Bold)

        # Built-in functions (yellow)
        self.builtin_format = QTextCharFormat()
        self.builtin_format.setForeground(QColor("#dcdcaa"))

        # Class/function definitions (blue)
        self.definition_format = QTextCharFormat()
        self.definition_format.setForeground(QColor("#4ec9b0"))

        # Strings (orange/brown)
        self.string_format = QTextCharFormat()
        self.string_format.setForeground(QColor("#ce9178"))

        # Comments (green)
        self.comment_format = QTextCharFormat()
        self.comment_format.setForeground(QColor("#6a9955"))
        self.comment_format.setFontItalic(True)

        # Numbers (light green)
        self.number_format = QTextCharFormat()
        self.number_format.setForeground(QColor("#b5cea8"))

        # Decorators (yellow)
        self.decorator_format = QTextCharFormat()
        self.decorator_format.setForeground(QColor("#dcdcaa"))

        # Self/cls (blue)
        self.self_format = QTextCharFormat()
        self.self_format.setForeground(QColor("#9cdcfe"))

        # Operators
        self.operator_format = QTextCharFormat()
        self.operator_format.setForeground(QColor("#d4d4d4"))

    def _init_rules(self):
        """Initialize highlighting rules."""
        self.rules = []

        # Keywords
        keywords = [
            'and', 'as', 'assert', 'async', 'await', 'break', 'class', 'continue',
            'def', 'del', 'elif', 'else', 'except', 'finally', 'for', 'from',
            'global', 'if', 'import', 'in', 'is', 'lambda', 'nonlocal', 'not',
            'or', 'pass', 'raise', 'return', 'try', 'while', 'with', 'yield',
            'True', 'False', 'None'
        ]
        for word in keywords:
            self.rules.append((rf'\b{word}\b', self.keyword_format))

        # Built-in functions
        builtins = [
            'abs', 'all', 'any', 'bin', 'bool', 'bytes', 'callable', 'chr',
            'classmethod', 'compile', 'complex', 'dict', 'dir', 'divmod',
            'enumerate', 'eval', 'exec', 'filter', 'float', 'format', 'frozenset',
            'getattr', 'globals', 'hasattr', 'hash', 'help', 'hex', 'id', 'input',
            'int', 'isinstance', 'issubclass', 'iter', 'len', 'list', 'locals',
            'map', 'max', 'min', 'next', 'object', 'oct', 'open', 'ord', 'pow',
            'print', 'property', 'range', 'repr', 'reversed', 'round', 'set',
            'setattr', 'slice', 'sorted', 'staticmethod', 'str', 'sum', 'super',
            'tuple', 'type', 'vars', 'zip', 're', 'Dict', 'List', 'Optional', 'Tuple'
        ]
        for word in builtins:
            self.rules.append((rf'\b{word}\b', self.builtin_format))

        # Self and cls
        self.rules.append((r'\bself\b', self.self_format))
        self.rules.append((r'\bcls\b', self.self_format))

        # Class and function definitions
        self.rules.append((r'\bclass\s+(\w+)', self.definition_format))
        self.rules.append((r'\bdef\s+(\w+)', self.definition_format))

        # Decorators
        self.rules.append((r'@\w+', self.decorator_format))

        # Numbers
        self.rules.append((r'\b[0-9]+\.?[0-9]*\b', self.number_format))

        # Single-line strings
        self.rules.append((r'"[^"\\]*(\\.[^"\\]*)*"', self.string_format))
        self.rules.append((r"'[^'\\]*(\\.[^'\\]*)*'", self.string_format))

        # Comments
        self.rules.append((r'#[^\n]*', self.comment_format))

    def highlightBlock(self, text):
        """Apply syntax highlighting to a block of text."""
        import re

        for pattern, fmt in self.rules:
            for match in re.finditer(pattern, text):
                start = match.start()
                length = match.end() - start
                # For class/def, highlight the name part
                if match.lastindex:
                    start = match.start(1)
                    length = match.end(1) - start
                self.setFormat(start, length, fmt)

        # Handle multi-line strings (triple quotes)
        self._highlight_multiline_strings(text)

    def _highlight_multiline_strings(self, text):
        """Handle multi-line string highlighting."""
        import re

        # Triple double quotes
        in_multiline = self.previousBlockState() == 1

        start_index = 0
        if not in_multiline:
            match = re.search(r'"""', text)
            if match:
                start_index = match.start()
            else:
                start_index = -1

        while start_index >= 0:
            end_match = re.search(r'"""', text[start_index + 3:])
            if end_match:
                end_index = start_index + 3 + end_match.end()
                self.setCurrentBlockState(0)
            else:
                end_index = len(text)
                self.setCurrentBlockState(1)

            self.setFormat(start_index, end_index - start_index, self.string_format)

            match = re.search(r'"""', text[end_index:])
            if match:
                start_index = end_index + match.start()
            else:
                start_index = -1


class LineNumberArea(QWidget):
    """Line number area for the code editor."""

    def __init__(self, editor):
        super().__init__(editor)
        self.editor = editor

    def sizeHint(self):
        return QSize(self.editor.line_number_area_width(), 0)

    def paintEvent(self, event):
        self.editor.line_number_area_paint_event(event)


class PythonCodeEditor(QPlainTextEdit):
    """Enhanced code editor with line numbers and auto-indentation."""

    def __init__(self, parent=None):
        super().__init__(parent)

        # Line number area
        self.line_number_area = LineNumberArea(self)

        # Connect signals
        self.blockCountChanged.connect(self.update_line_number_area_width)
        self.updateRequest.connect(self.update_line_number_area)
        self.cursorPositionChanged.connect(self.highlight_current_line)

        self.update_line_number_area_width(0)
        self.highlight_current_line()

        # Apply syntax highlighter
        self.highlighter = PythonSyntaxHighlighter(self.document())

        # Set tab width
        font_metrics = self.fontMetrics()
        self.setTabStopDistance(4 * font_metrics.horizontalAdvance(' '))

    def line_number_area_width(self):
        """Calculate the width needed for line numbers."""
        digits = len(str(max(1, self.blockCount())))
        space = 10 + self.fontMetrics().horizontalAdvance('9') * digits
        return space

    def update_line_number_area_width(self, _):
        """Update the viewport margins to make room for line numbers."""
        self.setViewportMargins(self.line_number_area_width(), 0, 0, 0)

    def update_line_number_area(self, rect, dy):
        """Update the line number area when scrolling."""
        if dy:
            self.line_number_area.scroll(0, dy)
        else:
            self.line_number_area.update(0, rect.y(), self.line_number_area.width(), rect.height())

        if rect.contains(self.viewport().rect()):
            self.update_line_number_area_width(0)

    def resizeEvent(self, event):
        """Handle resize to adjust line number area."""
        super().resizeEvent(event)
        cr = self.contentsRect()
        self.line_number_area.setGeometry(QRect(cr.left(), cr.top(), self.line_number_area_width(), cr.height()))

    def line_number_area_paint_event(self, event):
        """Paint line numbers."""
        painter = QPainter(self.line_number_area)
        painter.fillRect(event.rect(), QColor("#1e1e1e"))

        block = self.firstVisibleBlock()
        block_number = block.blockNumber()
        top = round(self.blockBoundingGeometry(block).translated(self.contentOffset()).top())
        bottom = top + round(self.blockBoundingRect(block).height())

        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                number = str(block_number + 1)
                painter.setPen(QColor("#858585"))
                painter.drawText(0, top, self.line_number_area.width() - 5,
                               self.fontMetrics().height(), Qt.AlignRight, number)

            block = block.next()
            top = bottom
            bottom = top + round(self.blockBoundingRect(block).height())
            block_number += 1

    def highlight_current_line(self):
        """Highlight the current line."""
        extra_selections = []

        if not self.isReadOnly():
            selection = QTextEdit.ExtraSelection()
            line_color = QColor("#2d2d2d")
            selection.format.setBackground(line_color)
            selection.format.setProperty(QTextFormat.FullWidthSelection, True)
            selection.cursor = self.textCursor()
            selection.cursor.clearSelection()
            extra_selections.append(selection)

        self.setExtraSelections(extra_selections)

    def keyPressEvent(self, event):
        """Handle key presses for auto-indentation."""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            # Auto-indent on Enter
            cursor = self.textCursor()
            line = cursor.block().text()

            # Get current indentation
            indent = ""
            for char in line:
                if char in (' ', '\t'):
                    indent += char
                else:
                    break

            # Add extra indent after colon
            stripped = line.rstrip()
            if stripped.endswith(':'):
                indent += "    "

            # Insert newline with indent
            super().keyPressEvent(event)
            self.insertPlainText(indent)

        elif event.key() == Qt.Key_Tab:
            # Insert 4 spaces instead of tab
            self.insertPlainText("    ")

        elif event.key() == Qt.Key_Backtab:
            # Shift+Tab - decrease indentation
            cursor = self.textCursor()
            cursor.movePosition(QTextCursor.StartOfLine)
            cursor.movePosition(QTextCursor.Right, QTextCursor.KeepAnchor, 4)
            if cursor.selectedText() == "    ":
                cursor.removeSelectedText()

        else:
            super().keyPressEvent(event)

    def format_code(self):
        """Auto-format the Python code (basic formatting)."""
        code = self.toPlainText()
        try:
            # Try to use black if available
            import black
            formatted = black.format_str(code, mode=black.Mode())
            self.setPlainText(formatted)
            return True, "Code formatted successfully"
        except ImportError:
            # Fall back to basic formatting
            return self._basic_format(code)
        except Exception as e:
            return False, f"Format error: {e}"

    def _basic_format(self, code):
        """Basic code formatting without external dependencies."""
        import re
        lines = code.split('\n')
        formatted_lines = []
        indent_level = 0

        for line in lines:
            stripped = line.strip()
            if not stripped:
                formatted_lines.append("")
                continue

            # Decrease indent for these keywords
            if stripped.startswith(('else:', 'elif ', 'except:', 'except ', 'finally:', 'elif:')):
                indent_level = max(0, indent_level - 1)

            # Apply current indent
            formatted_lines.append("    " * indent_level + stripped)

            # Increase indent after colon (but not for comments)
            if stripped.endswith(':') and not stripped.startswith('#'):
                indent_level += 1

            # Decrease indent after return/break/continue/pass/raise
            if stripped.startswith(('return ', 'return', 'break', 'continue', 'pass', 'raise ')):
                indent_level = max(0, indent_level - 1)

        self.setPlainText('\n'.join(formatted_lines))
        return True, "Basic formatting applied"


class ChatMessageInput(QPlainTextEdit):
    """Message input that sends on Enter, newline on Shift+Enter."""

    sendRequested = pyqtSignal()

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            if event.modifiers() & Qt.ShiftModifier:
                # Shift+Enter: insert newline
                super().keyPressEvent(event)
            else:
                # Enter: send message
                self.sendRequested.emit()
        else:
            super().keyPressEvent(event)


# ----------------------------------------------------------------------
# LOGIN DIALOG
# ----------------------------------------------------------------------
class LoginDialog(QDialog):
    """Login dialog for user authentication."""

    def __init__(self, auth_manager: AuthenticationManager, parent=None):
        super().__init__(parent)
        self.auth_manager = auth_manager
        self.authenticated = False
        self.user_role = None

        self.setWindowTitle(f"{APP_NAME} - Login")
        self.setFixedSize(400, 360)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        # Set dark background for the dialog
        self.setStyleSheet("""
            QDialog {
                background-color: #2d2d2d;
            }
        """)

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(30, 30, 30, 30)

        # Logo/Title
        title_label = QLabel(f"Welcome to {APP_NAME}")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #ffffff;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        subtitle_label = QLabel("Please sign in to continue")
        subtitle_label.setStyleSheet("font-size: 12px; color: #a0a0a0; margin-bottom: 10px;")
        subtitle_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(subtitle_label)

        # Email field
        email_label = QLabel("Email:")
        email_label.setStyleSheet("font-weight: bold; color: #e0e0e0;")
        layout.addWidget(email_label)

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("your@email.com")
        self.email_input.setMinimumHeight(40)
        self.email_input.setStyleSheet("""
            QLineEdit {
                padding: 8px 10px;
                border: 1px solid #555;
                border-radius: 5px;
                font-size: 14px;
                background-color: #3c3c3c;
                color: #ffffff;
            }
            QLineEdit:focus {
                border-color: #0078d4;
            }
        """)
        # Pre-fill with last user
        last_user = self.auth_manager.get_last_user()
        if last_user:
            self.email_input.setText(last_user)
        layout.addWidget(self.email_input)

        # Password field
        password_label = QLabel("Password:")
        password_label.setStyleSheet("font-weight: bold; color: #e0e0e0;")
        layout.addWidget(password_label)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Enter your password")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setMinimumHeight(40)
        self.password_input.setStyleSheet("""
            QLineEdit {
                padding: 8px 10px;
                border: 1px solid #555;
                border-radius: 5px;
                font-size: 14px;
                background-color: #3c3c3c;
                color: #ffffff;
            }
            QLineEdit:focus {
                border-color: #0078d4;
            }
        """)
        self.password_input.returnPressed.connect(self._on_login)
        layout.addWidget(self.password_input)

        # Error message label (hidden by default)
        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color: #dc3545; font-size: 12px;")
        self.error_label.setAlignment(Qt.AlignCenter)
        self.error_label.setVisible(False)
        layout.addWidget(self.error_label)

        layout.addStretch()

        # Buttons
        btn_layout = QHBoxLayout()

        self.login_btn = QPushButton("Sign In")
        self.login_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                padding: 10px 30px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
            QPushButton:pressed {
                background-color: #005a9e;
            }
        """)
        self.login_btn.clicked.connect(self._on_login)
        btn_layout.addWidget(self.login_btn)

        self.cancel_btn = QPushButton("Exit")
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 10px 30px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)

        layout.addLayout(btn_layout)

        # Focus on password if email is pre-filled, otherwise on email
        if last_user:
            self.password_input.setFocus()
        else:
            self.email_input.setFocus()

    def _on_login(self):
        """Handle login button click."""
        email = self.email_input.text().strip()
        password = self.password_input.text()

        if not email:
            self._show_error("Please enter your email address")
            self.email_input.setFocus()
            return

        if not password:
            self._show_error("Please enter your password")
            self.password_input.setFocus()
            return

        # Disable buttons during authentication
        self.login_btn.setEnabled(False)
        self.login_btn.setText("Signing in...")
        QApplication.processEvents()

        try:
            success, message, role = self.auth_manager.authenticate(email, password)

            if success:
                self.authenticated = True
                self.user_role = role
                self.accept()
            else:
                self._show_error(message)
                self.password_input.clear()
                self.password_input.setFocus()
        finally:
            self.login_btn.setEnabled(True)
            self.login_btn.setText("Sign In")

    def _show_error(self, message: str):
        """Display error message."""
        self.error_label.setText(message)
        self.error_label.setVisible(True)


# ----------------------------------------------------------------------
# VISUAL PDF PATTERN TRAINER WITH DRAWING CANVAS
# ----------------------------------------------------------------------
class TariffMill(QMainWindow):
    def eventFilter(self, obj, event):
        """Application-level event filter - intercepts ALL events before any widget processing"""
        from PyQt5.QtCore import QEvent, Qt

        # Intercept ALL keyboard events at application level
        if event.type() == QEvent.KeyPress:
            focused_widget = QApplication.focusWidget()

            # CRITICAL FIX: Forward ALL keyboard events (including Tab) to ci_input/wt_input
            # The ForceEditableLineEdit.keyPressEvent will handle Tab specially with focusNextChild()
            if hasattr(self, 'ci_input') and hasattr(self, 'wt_input'):
                if focused_widget in [self.ci_input, self.wt_input]:
                    # Only process once per event - when it comes from QWindow
                    if obj.__class__.__name__ == 'QWindow':
                        # Manually send the event to the focused widget (including Tab keys)
                        QApplication.sendEvent(focused_widget, event)
                        return True  # We handled it, block further propagation

        return False  # Let ALL other events continue normally

    def setup_tab_by_index(self, index):
        """Initialize tab by index using existing setup methods."""
        tab_setup_methods = {
            1: self.setup_master_tab
            # Invoice Mapping, Output Mapping, and Parts Import moved to Configuration menu
            # Customs Config and Section 232 Actions moved to References menu
        }
        if index in tab_setup_methods:
            tab_setup_methods[index]()
    def __init__(self):
        super().__init__()
        # Hide window during initialization to prevent ghost window flash
        self.setAttribute(Qt.WA_DontShowOnScreen, True)
        self.setWindowTitle(APP_NAME)
        # Set window size to 100% of available screen dimensions
        screen = QApplication.primaryScreen().availableGeometry()
        self.setGeometry(screen)
        self.setMinimumSize(800, 500)

        # Install application-level event filter to intercept ALL keyboard events
        QApplication.instance().installEventFilter(self)

        # Track processed events to prevent duplicates
        self._processed_events = set()
        
        # Set window icon (use TEMP_RESOURCES_DIR for bundled resources)
        # Prefer SVG icon for scalability, fall back to ICO for Windows taskbar
        icon_path = TEMP_RESOURCES_DIR / "tariffmill_icon_hybrid_2.svg"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "icon.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        
        self.current_csv = None
        self.shipment_mapping = {}
        self.output_column_mapping = None  # Will be initialized in setup_output_mapping_tab
        self.output_column_order = None  # Will be initialized in setup_output_mapping_tab
        self.profile_header_row = 1  # Default header row (1 = first row)
        self.selected_mid = ""
        self.current_worker = None
        self.missing_df = None
        self.csv_total_value = 0.0
        self.last_processed_df = None
        self.last_output_filename = None
        self.shipment_targets = {}  # Prevent attribute error before tab setup

        # Load output font color from settings
        self.output_font_color = '#000000'  # Default black
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", ('output_font_color',))
            row = c.fetchone()
            conn.close()
            if row:
                self.output_font_color = row[0]
        except:
            pass

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        # Modern header with watermill logo on left, text in center
        header_container = QWidget()
        header_container.setStyleSheet("background: transparent; border: none;")
        header_container.setFixedHeight(48)
        header_layout = QHBoxLayout(header_container)
        header_layout.setContentsMargins(20, 6, 20, 6)
        header_layout.setSpacing(10)

        # TariffMill logo with wordmark
        logo_path = TEMP_RESOURCES_DIR / "tariffmill_logo_small.svg"
        fixed_header_height = 60
        if logo_path.exists():
            logo_label = QLabel()
            pixmap = QPixmap(str(logo_path))
            # Scale to fit header height while maintaining aspect ratio
            scaled_pixmap = pixmap.scaledToHeight(fixed_header_height, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
            logo_label.setStyleSheet("background: transparent;")
            logo_label.setFixedHeight(fixed_header_height)
            self.header_logo_label = logo_label
        else:
            # Fallback to text if logo not found
            self.header_logo_label = QLabel(f"{APP_NAME}")
            self.header_logo_label.setStyleSheet("""
                font-size: 22px;
                font-weight: bold;
                color: #555555;
                font-family: 'Impact', 'Arial Black', sans-serif;
            """)

        # Add logo to header layout
        header_layout.addWidget(self.header_logo_label, 0, Qt.AlignVCenter)
        header_layout.addStretch(1)


        layout.addWidget(header_container)



        # Add a native menu bar with Settings and Log View actions
        menubar = QMenuBar(self)
        settings_menu = menubar.addMenu("Settings")
        # Use a standard gear icon from QStyle
        gear_icon = self.style().standardIcon(QStyle.SP_FileDialogDetailedView)
        settings_action = QAction(gear_icon, "Settings", self)
        settings_action.triggered.connect(self.show_settings_dialog)
        settings_menu.addAction(settings_action)

        # Add Account menu
        account_menu = menubar.addMenu("Account")
        # Show current user
        self.account_user_action = QAction("", self)
        self.account_user_action.setEnabled(False)
        account_menu.addAction(self.account_user_action)
        account_menu.addSeparator()
        # Sign out action
        signout_icon = self.style().standardIcon(QStyle.SP_DialogCloseButton)
        signout_action = QAction(signout_icon, "Sign Out", self)
        signout_action.triggered.connect(self._sign_out)
        account_menu.addAction(signout_action)

        # Add Log View menu
        log_menu = menubar.addMenu("Log View")
        log_icon = self.style().standardIcon(QStyle.SP_FileDialogContentsView)
        log_action = QAction(log_icon, "View Log", self)
        log_action.triggered.connect(self.show_log_dialog)
        log_menu.addAction(log_action)

        # Add References menu
        references_menu = menubar.addMenu("References")
        references_icon = self.style().standardIcon(QStyle.SP_FileDialogInfoView)
        references_action = QAction(references_icon, "References...", self)
        references_action.triggered.connect(self.show_references_dialog)
        references_menu.addAction(references_action)

        # Add Configuration menu
        config_menu = menubar.addMenu("Configuration")
        config_icon = self.style().standardIcon(QStyle.SP_FileDialogDetailedView)
        config_action = QAction(config_icon, "Configuration...", self)
        config_action.triggered.connect(self.show_configuration_dialog)
        config_menu.addAction(config_action)

        # TODO: Export menu - To be implemented at a later date
        # export_menu = menubar.addMenu("Export")
        #
        # # TODO: XML Export - To be implemented at a later date
        # # Export to XML action
        # xml_icon = self.style().standardIcon(QStyle.SP_FileIcon)
        # xml_export_action = QAction(xml_icon, "Export to XML...", self)
        # xml_export_action.triggered.connect(self.export_to_xml)
        # xml_export_action.setToolTip("Export processed invoice data to XML format for e2Open Customs Management")
        # export_menu.addAction(xml_export_action)
        #
        # # TODO: Lacey Act Export - To be implemented at a later date
        # lacey_icon = self.style().standardIcon(QStyle.SP_FileDialogContentsView)
        # lacey_export_action = QAction(lacey_icon, "Export Lacey Act (PPQ 505)...", self)
        # lacey_export_action.triggered.connect(self.export_lacey_act_ppq505)
        # lacey_export_action.setToolTip("Export items requiring Lacey Act declaration to PPQ Form 505 format")
        # export_menu.addAction(lacey_export_action)

        # Add Help menu
        help_menu = menubar.addMenu("Help")

        # License & Activation action
        license_icon = self.style().standardIcon(QStyle.SP_DialogApplyButton)
        license_action = QAction(license_icon, "License && Activation...", self)
        license_action.triggered.connect(self.show_license_dialog)
        help_menu.addAction(license_action)

        help_menu.addSeparator()

        # Check for Updates action
        update_icon = self.style().standardIcon(QStyle.SP_BrowserReload)
        update_action = QAction(update_icon, "Check for Updates...", self)
        update_action.triggered.connect(self.check_for_updates_manual)
        help_menu.addAction(update_action)

        help_menu.addSeparator()
        
        # About action
        about_icon = self.style().standardIcon(QStyle.SP_MessageBoxInformation)
        about_action = QAction(about_icon, "About", self)
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)
        
        layout.setMenuBar(menubar)
        self.settings_action = settings_action

        # Top status bar removed per user request
        # Create a dummy status object that ignores all calls
        class DummyStatus:
            def setText(self, text): pass
            def setStyleSheet(self, style): pass
            def setAlignment(self, align): pass
        self.status = DummyStatus()

        # Add spacing between header and tabs
        layout.addSpacing(20)

        self.tabs = QTabWidget()
        self.tab_process = QWidget()
        self.tab_shipment_map = QWidget()
        self.tab_output_map = QWidget()
        self.tab_import = QWidget()
        self.tab_master = QWidget()
        self.tab_log = QWidget()  # Keep widget for log functionality
        self.tab_config = QWidget()
        self.tab_actions = QWidget()
        self.tab_ocrmill = QWidget()
        self.tabs.addTab(self.tab_process, "Invoice Processing")
        self.tabs.addTab(self.tab_ocrmill, "PDF Processing")
        self.tabs.addTab(self.tab_master, "Parts View")
        # Invoice Mapping, Output Mapping, and Parts Import moved to Configuration menu
        # Customs Config and Section 232 Actions moved to References menu
        
        # Only tabs (no settings icon here)
        tabs_container = QWidget()
        tabs_layout = QHBoxLayout(tabs_container)
        tabs_layout.setContentsMargins(0, 0, 0, 0)
        tabs_layout.setSpacing(10)
        tabs_layout.addWidget(self.tabs)
        layout.addWidget(tabs_container, 1)
        
        # Bottom status bar with export progress indicator
        bottom_bar = QWidget()
        bottom_bar.setMinimumHeight(18)
        bottom_bar_layout = QHBoxLayout(bottom_bar)
        bottom_bar_layout.setContentsMargins(10, 2, 10, 2)
        bottom_bar_layout.setSpacing(10)

        self.bottom_status = QLabel("Ready")
        self.bottom_status.setStyleSheet("font-size: 9px;")
        bottom_bar_layout.addWidget(self.bottom_status, 1)
        
        # Export progress indicator (hidden by default)
        self.export_progress_widget = QWidget()
        export_progress_layout = QHBoxLayout(self.export_progress_widget)
        export_progress_layout.setContentsMargins(0, 0, 0, 0)
        export_progress_layout.setSpacing(5)
        
        self.export_status_label = QLabel("")
        self.export_status_label.setStyleSheet("color: #666666;")
        export_progress_layout.addWidget(self.export_status_label)
        
        self.export_progress_bar = QProgressBar()
        self.export_progress_bar.setMaximum(100)
        self.export_progress_bar.setFixedWidth(120)
        self.export_progress_bar.setFixedHeight(14)
        self.export_progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #b0b0b0;
                border-radius: 3px;
                background-color: #f5f5f5;
                text-align: center;
                font-size: 7pt;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
                border-radius: 2px;
            }
        """)
        export_progress_layout.addWidget(self.export_progress_bar)
        
        bottom_bar_layout.addWidget(self.export_progress_widget)
        self.export_progress_widget.hide()

        # Add version label in bottom right corner
        self.version_label = QLabel(VERSION)
        self.version_label.setStyleSheet("font-size: 7pt; color: #999999;")
        self.version_label.setAlignment(Qt.AlignRight)
        bottom_bar_layout.addWidget(self.version_label)

        self.bottom_bar = bottom_bar  # Store reference for theme updates
        bottom_bar.setFixedHeight(20)
        layout.addWidget(bottom_bar)

        # Track which tabs have been initialized (lazy loading for performance)
        self.tabs_initialized = set()
        

        # Only set up the first tab immediately for faster startup
        self.setup_process_tab()
        self.tabs_initialized.add(0)

        # Connect to tab change signal for lazy initialization
        self.tabs.currentChanged.connect(self.on_tab_changed)



        logger.success(f"{APP_NAME} {VERSION} GUI ready")

    # Removed deferred_initialization; tabs are now lazily loaded only when selected
    
    def initialize_data(self, splash=None, progress_callback=None):
        """Initialize database and load all data before showing window"""
        steps = [
            ("Loading configuration...", self.load_config_paths),
            ("Applying theme...", self.apply_saved_theme),
            ("Applying font size...", self.apply_saved_font_size),
            ("Loading MIDs...", self.load_available_mids),
            ("Loading profiles...", self.load_mapping_profiles),
            ("Loading export profiles...", self.load_output_mapping_profiles),
            # Removed output file scanning on startup
            ("Scanning input files...", self.refresh_input_files),
            ("Starting auto-refresh...", self.setup_auto_refresh),
            ("Finalizing...", self.update_status_bar_styles),
        ]
        
        total_steps = len(steps)
        for i, (message, func) in enumerate(steps):
            if splash:
                splash.setText(f"{message}\nPlease wait...")
            if progress_callback:
                progress_callback(int((i / total_steps) * 100))
            QApplication.processEvents()
            
            try:
                func()
            except Exception as e:
                logger.error(f"Error during {message}: {e}")
        
        if progress_callback:
            progress_callback(100)

        # Ensure input fields are enabled after all initialization
        if hasattr(self, '_enable_input_fields'):
            self._enable_input_fields()

        logger.success(f"{APP_NAME} {VERSION} loaded successfully")
    
    def on_tab_changed(self, index):
        """Initialize tabs lazily when they are first accessed"""
        if index in self.tabs_initialized:
            return

        # Map tab index to setup method
        # Tab order: 0=Invoice Processing, 1=PDF Processing, 2=Parts View
        tab_setup_methods = {
            1: self.setup_ocrmill_tab,
            2: self.setup_master_tab
            # Invoice Mapping, Output Mapping, and Parts Import moved to Configuration menu
            # Customs Config and Section 232 Actions moved to References menu
        }
        
        # Initialize the tab
        if index in tab_setup_methods:
            tab_setup_methods[index]()
            self.tabs_initialized.add(index)
            logger.debug(f"Initialized tab {index}")
    
    def apply_saved_theme(self):
        """Load and apply the saved theme preference on startup (per-user setting)"""
        theme = get_user_setting('theme', 'Fusion (Light)')
        self.apply_theme(theme)

    def apply_saved_font_size(self):
        """Load and apply the saved font size preference on startup (per-user setting)"""
        font_size = get_user_setting_int('font_size', 9)
        logger.info(f"Loading saved font size: {font_size}pt")
        self.apply_font_size_without_save(font_size)
        logger.info(f"Applied font size: {font_size}pt")

    def load_config_paths(self):
        try:
            self.bottom_status.setText("Loading Directory location...")
            QApplication.processEvents()
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = 'input_dir'")
            row = c.fetchone()
            global INPUT_DIR, PROCESSED_DIR
            if row:
                INPUT_DIR = Path(row[0])
                PROCESSED_DIR = get_processed_dir(INPUT_DIR)
                PROCESSED_DIR.mkdir(exist_ok=True)
                QApplication.processEvents()
            c.execute("SELECT value FROM app_config WHERE key = 'output_dir'")
            row = c.fetchone()
            if row:
                global OUTPUT_DIR
                OUTPUT_DIR = Path(row[0])
                OUTPUT_DIR.mkdir(exist_ok=True)
                QApplication.processEvents()
            conn.close()
            self.bottom_status.setText("Ready")
            QApplication.processEvents()
        except Exception as e:
            logger.error(f"Config load failed: {e}")
            self.bottom_status.setText("Config load failed")
            QApplication.processEvents()

    def setup_process_tab(self):
        layout = QVBoxLayout(self.tab_process)

        # MAIN CONTAINER: Left side (controls) + Right side (preview) with splitter
        main_container = QHBoxLayout()
        
        # Create splitter for resizable panels
        splitter = QSplitter(Qt.Horizontal)

        # LEFT SIDE: Single outer box containing all controls with scroll area
        left_outer_box = QGroupBox("Controls")
        left_outer_layout = QVBoxLayout(left_outer_box)
        left_outer_layout.setContentsMargins(0, 0, 0, 0)

        # Create scroll area for left side content
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        left_scroll.setFrameShape(QFrame.NoFrame)

        left_scroll_widget = QWidget()
        left_side = QVBoxLayout(left_scroll_widget)
        left_side.setSpacing(10)
        left_side.setContentsMargins(10, 10, 10, 10)

        # INPUT FILES LIST — now inside Shipment File group
        self.input_files_list = AutoSelectListWidget()
        self.input_files_list.setSelectionMode(QListWidget.SingleSelection)
        self.input_files_list.itemClicked.connect(self.load_selected_input_file)
        # Connect itemActivated for Enter key and double-click support
        self.input_files_list.itemActivated.connect(self.load_selected_input_file)
        # Allow focus for tab navigation
        self.input_files_list.setFocusPolicy(Qt.StrongFocus)
        # Limit height to show ~4-5 files to save vertical space
        self.input_files_list.setFixedHeight(75)
        self.refresh_input_btn = QPushButton("Refresh")
        self.refresh_input_btn.setFixedHeight(25)
        self.refresh_input_btn.clicked.connect(self.refresh_input_files)

        self.delete_input_btn = QPushButton("Delete")
        self.delete_input_btn.setFixedHeight(25)
        self.delete_input_btn.clicked.connect(self.delete_selected_input_file)
        self.delete_input_btn.setToolTip("Delete selected input file")

        # INVOICE VALUES
        values_group = QGroupBox("Invoice Values")
        values_layout = QFormLayout()
        values_layout.setLabelAlignment(Qt.AlignRight)

        self.ci_input = ForceEditableLineEdit("")
        self.ci_input.setObjectName("ci_input")
        self.ci_input.setPlaceholderText("Enter CI value...")
        self.ci_input.textChanged.connect(self.update_invoice_check)

        self.wt_input = ForceEditableLineEdit("")
        self.wt_input.setObjectName("wt_input")
        self.wt_input.setPlaceholderText("Enter net weight...")
        self.wt_input.textChanged.connect(self.update_invoice_check)

        values_layout.addRow("CI Value (USD):", self.ci_input)
        values_layout.addRow("Net Weight (kg):", self.wt_input)

        # MID selector (moved above Invoice Check)
        self.mid_label = QLabel("MID:")
        self.mid_combo = QComboBox()
        self.mid_combo.setFocusPolicy(Qt.StrongFocus)  # Ensure combo accepts keyboard focus
        self.mid_combo.currentTextChanged.connect(self.on_mid_changed)
        values_layout.addRow(self.mid_label, self.mid_combo)

        # File Number (mandatory billing reference)
        self.file_number_input = ForceEditableLineEdit("")
        self.file_number_input.setObjectName("file_number_input")
        self.file_number_input.setPlaceholderText("Enter file number (required)...")
        self.file_number_input.setToolTip("Billing reference number - required for export")
        self.file_number_input.textChanged.connect(self.update_invoice_check)
        values_layout.addRow("File Number *:", self.file_number_input)

        # TODO: Customer Reference - To be implemented with XML export at a later date
        # # Customer Reference Number (for XML export)
        # self.customer_ref_input = ForceEditableLineEdit("")
        # self.customer_ref_input.setObjectName("customer_ref_input")
        # self.customer_ref_input.setPlaceholderText("Optional - for XML export")
        # self.customer_ref_input.setToolTip("Customer reference number included in XML export header")
        # values_layout.addRow("Customer Ref:", self.customer_ref_input)
        self.customer_ref_input = None  # Placeholder for XML export feature

        # Removed broken setTabOrder calls - they were causing Qt warnings and possibly blocking keyboard input

        # Invoice check label and Edit Values button
        self.invoice_check_label = QLabel("No file loaded")
        self.invoice_check_label.setWordWrap(False)  # Don't wrap - keep invoice total on one line
        font_size = get_user_setting_int('font_size', 9)
        self.invoice_check_label.setStyleSheet(f"font-size: {font_size}pt;")
        self.invoice_check_label.setAlignment(Qt.AlignCenter)
        self.invoice_check_label.setMinimumWidth(200)  # Wider minimum to fit invoice totals
        self.invoice_check_label.setMaximumWidth(280)  # Allow more space for larger amounts

        vbox_check = QVBoxLayout()
        vbox_check.setSpacing(12)
        vbox_check.setContentsMargins(0, 10, 0, 0)

        vbox_check.addWidget(self.invoice_check_label, alignment=Qt.AlignCenter)

        # Edit Values button (initially hidden, shown when values don't match)
        self.edit_values_btn = QPushButton("Edit Values")
        self.edit_values_btn.setFixedSize(120, 30)
        self.edit_values_btn.setStyleSheet(self.get_button_style("warning"))
        self.edit_values_btn.setVisible(False)
        self.edit_values_btn.setAutoDefault(True)
        self.edit_values_btn.setDefault(True)
        self.edit_values_btn.clicked.connect(self.start_processing_with_editable_preview)
        vbox_check.addWidget(self.edit_values_btn, alignment=Qt.AlignCenter)
        
        vbox_check.addStretch()

        values_layout.addRow("Invoice Check:", vbox_check)
        values_group.setLayout(values_layout)

        # SHIPMENT FILE (merged with Saved Profiles and Input Files)
        file_group = QGroupBox("Shipment File")
        file_group.setObjectName("SavedProfilesGroup")
        file_layout = QFormLayout()
        file_layout.setLabelAlignment(Qt.AlignRight)

        # Folder profile selector with manage button
        folder_profile_row = QHBoxLayout()
        self.folder_profile_combo = QComboBox()
        self.folder_profile_combo.setMinimumWidth(150)
        self.folder_profile_combo.currentTextChanged.connect(self.load_folder_profile)
        folder_profile_row.addWidget(self.folder_profile_combo)
        self.manage_folder_profiles_btn = QPushButton()
        self.manage_folder_profiles_btn.setFixedWidth(30)
        self.manage_folder_profiles_btn.setToolTip("Manage folder profiles")
        self.manage_folder_profiles_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        self.manage_folder_profiles_btn.clicked.connect(self.show_folder_profile_dialog)
        folder_profile_row.addWidget(self.manage_folder_profiles_btn)
        file_layout.addRow("Folder Profile:", folder_profile_row)

        # Profile selector
        self.profile_combo = QComboBox()
        self.profile_combo.currentTextChanged.connect(self.load_selected_profile)
        file_layout.addRow("Map Profile:", self.profile_combo)
        # Input files list and refresh/delete buttons (moved here)
        file_layout.addRow("Input Files:", self.input_files_list)
        input_btn_layout = QHBoxLayout()
        input_btn_layout.setSpacing(5)
        input_btn_layout.addWidget(self.refresh_input_btn)
        input_btn_layout.addWidget(self.delete_input_btn)
        input_btn_layout.addStretch()
        file_layout.addRow("", input_btn_layout)
        # File display (read-only, shows selected file from Input Files list)
        self.file_label = QLabel("No file selected")
        self.file_label.setWordWrap(True)
        self.update_file_label_style()  # Set initial style based on theme
        file_layout.addRow("Selected File:", self.file_label)
        file_group.setLayout(file_layout)
        left_side.addWidget(file_group)
        left_side.addWidget(values_group)

        # ACTIONS GROUP — Process/Export + Reprocess + Clear All buttons
        actions_group = QGroupBox("Actions")
        actions_layout = QHBoxLayout()
        actions_layout.setContentsMargins(5, 5, 5, 5)
        actions_layout.setSpacing(5)

        self.process_btn = QPushButton("Process Invoice")
        self.process_btn.setEnabled(False)
        self.process_btn.setFixedHeight(28)
        self.process_btn.setStyleSheet(self.get_button_style("success"))
        self.process_btn.clicked.connect(self._process_or_export)
        # Make button respond to Enter/Return key when focused
        self.process_btn.setAutoDefault(True)
        self.process_btn.setDefault(False)  # Don't make it the default for the whole window

        self.reprocess_btn = QPushButton("Reprocess")
        self.reprocess_btn.setEnabled(False)
        self.reprocess_btn.setFixedHeight(28)
        self.reprocess_btn.setStyleSheet(self.get_button_style("info"))
        self.reprocess_btn.clicked.connect(self.reprocess_invoice)
        self.reprocess_btn.setToolTip("Re-process invoice to pick up database changes")

        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setFixedHeight(28)
        self.clear_btn.setStyleSheet(self.get_button_style("danger"))
        self.clear_btn.clicked.connect(self.clear_all)

        actions_layout.addWidget(self.process_btn)
        actions_layout.addWidget(self.reprocess_btn)
        actions_layout.addWidget(self.clear_btn)
        actions_group.setLayout(actions_layout)
        left_side.addWidget(actions_group)

        # EXPORTED FILES GROUP — shows recent exports
        exports_group = QGroupBox("Exported Files")
        exports_layout = QVBoxLayout()

        self.exports_list = AutoSelectListWidget()
        self.exports_list.setSelectionMode(QListWidget.SingleSelection)
        self.exports_list.itemDoubleClicked.connect(self.open_exported_file)
        # Connect itemActivated for Enter key support
        self.exports_list.itemActivated.connect(self.open_exported_file)
        # Allow focus for tab navigation
        self.exports_list.setFocusPolicy(Qt.StrongFocus)
        # Limit height to save vertical space
        self.exports_list.setFixedHeight(75)
        exports_layout.addWidget(self.exports_list)

        self.refresh_exports_btn = QPushButton("Refresh")
        self.refresh_exports_btn.setFixedHeight(25)
        self.refresh_exports_btn.clicked.connect(self.refresh_exported_files)
        exports_layout.addWidget(self.refresh_exports_btn)

        # Invoice total display for selected file
        total_layout = QHBoxLayout()
        total_layout.addWidget(QLabel("Invoice Total:"))
        self.export_invoice_total = QLineEdit()
        self.export_invoice_total.setReadOnly(True)
        self.export_invoice_total.setPlaceholderText("Select a file")
        total_layout.addWidget(self.export_invoice_total)
        exports_layout.addLayout(total_layout)

        # Connect selection change to update invoice total
        self.exports_list.itemSelectionChanged.connect(self.update_export_invoice_total)

        exports_group.setLayout(exports_layout)
        left_side.addWidget(exports_group)

        # Finish scroll area setup
        left_scroll.setWidget(left_scroll_widget)
        left_outer_layout.addWidget(left_scroll)

        # Set min/max width for left controls
        left_outer_box.setMinimumWidth(300)
        left_outer_box.setMaximumWidth(400)

        # Add left_outer_box to splitter
        splitter.addWidget(left_outer_box)

        # RIGHT SIDE: Preview table in a widget
        right_widget = QWidget()
        right_side = QVBoxLayout(right_widget)
        right_side.setContentsMargins(0, 0, 0, 0)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        right_side.addWidget(self.progress)

        preview_group = QGroupBox("Result Preview")
        preview_layout = QVBoxLayout()

        self.table = QTableWidget()
        self.table.setColumnCount(20)
        self.table.setHorizontalHeaderLabels([
            "Product No","Value","HTS","MID","Qty1","Qty2","Qty Unit","Dec","Melt","Cast","Smelt","Flag","Steel%","Al%","Cu%","Wood%","Auto%","Non-232%","232 Status","Cust Ref"
        ])
        # Make columns manually resizable instead of auto-stretch
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setSelectionBehavior(QTableWidget.SelectItems)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSortingEnabled(False)  # Disabled for better performance
        # Set row height from saved preference (per-user setting)
        saved_row_height = get_user_setting_int('preview_row_height', 22)
        self.table.verticalHeader().setDefaultSectionSize(saved_row_height)
        # Configure vertical header (row numbers) for compact display
        self.table.verticalHeader().setMinimumSectionSize(14)  # Allow small rows
        self.table.verticalHeader().setFixedWidth(30)  # Fixed width for row numbers
        self.table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        # Match body font size to the header font size and make non-bold
        header_font = self.table.horizontalHeader().font()
        header_font.setBold(False)
        self.table.horizontalHeader().setFont(header_font)
        self.table.setFont(header_font)
        # Use smaller font for row numbers to prevent clipping
        vh_font = self.table.verticalHeader().font()
        vh_font.setPointSize(8)
        self.table.verticalHeader().setFont(vh_font)

        # Enable clicking header to select entire column
        self.table.horizontalHeader().sectionClicked.connect(self.select_column)

        # Connect signal to save column widths when they change
        self.table.horizontalHeader().sectionResized.connect(self.save_column_widths)

        # Handle selection changes to make selected text readable
        self.table.itemSelectionChanged.connect(self.on_table_selection_changed)
        self._previous_selection = set()  # Track previously selected items

        # Load saved column widths
        self.load_column_widths()

        # Apply green focus color stylesheet
        self.update_table_stylesheet()

        # Add context menu for column auto-fit
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_preview_context_menu)

        preview_layout.addWidget(self.table)
        preview_group.setLayout(preview_layout)
        right_side.addWidget(preview_group, 1)

        # Add right widget to splitter
        splitter.addWidget(right_widget)

        # Set initial splitter sizes (left: 350, right: remaining)
        splitter.setSizes([350, 850])

        # Make the splitter collapsible on the left side
        splitter.setCollapsible(0, False)  # Don't allow full collapse
        splitter.setCollapsible(1, False)
        
        # Add splitter to main container
        main_container.addWidget(splitter)

        layout.addLayout(main_container, 1)

        # Set up tab order for keyboard navigation through controls
        # Order: Map Profile → Input Files → Refresh (Shipment) → CI Value → Net Weight →
        #        MID → File Number → Process Invoice → Edit Values → Clear All → Exported Files → Refresh (Exports)
        self.setTabOrder(self.profile_combo, self.input_files_list)
        self.setTabOrder(self.input_files_list, self.refresh_input_btn)
        self.setTabOrder(self.refresh_input_btn, self.ci_input)
        self.setTabOrder(self.ci_input, self.wt_input)
        self.setTabOrder(self.wt_input, self.mid_combo)
        self.setTabOrder(self.mid_combo, self.file_number_input)
        self.setTabOrder(self.file_number_input, self.process_btn)
        self.setTabOrder(self.process_btn, self.edit_values_btn)
        self.setTabOrder(self.edit_values_btn, self.clear_btn)
        self.setTabOrder(self.clear_btn, self.exports_list)
        self.setTabOrder(self.exports_list, self.refresh_exports_btn)

        self.tab_process.setLayout(layout)
        self._install_preview_shortcuts()

        # Ensure input fields are enabled on startup
        self._enable_input_fields()

        # Create a timer to continuously force fields to be editable
        # This works around whatever is locking the fields
        from PyQt5.QtCore import QTimer
        self._field_watchdog_timer = QTimer()
        self._field_watchdog_timer.timeout.connect(self._force_fields_editable)
        self._field_watchdog_timer.start(100)  # Check every 100ms

        # Event filter already installed in __init__, don't install again
    
    def _force_fields_editable(self):
        """Watchdog timer callback that forces fields to stay editable"""
        if hasattr(self, 'ci_input'):
            # Bypass the override and force the parent class method
            QLineEdit.setReadOnly(self.ci_input, False)
            QLineEdit.setEnabled(self.ci_input, True)
        if hasattr(self, 'wt_input'):
            QLineEdit.setReadOnly(self.wt_input, False)
            QLineEdit.setEnabled(self.wt_input, True)

    def _enable_input_fields(self):
        """Ensure CI and Weight input fields are enabled and ready for input"""
        # Block signals to prevent recursion during enable
        if hasattr(self, 'ci_input'):
            self.ci_input.blockSignals(True)
            self.ci_input.setReadOnly(False)
            self.ci_input.setEnabled(True)
            self.ci_input.setFocusPolicy(Qt.StrongFocus)
            # Force immediate visual update
            self.ci_input.update()
            self.ci_input.blockSignals(False)

        if hasattr(self, 'wt_input'):
            self.wt_input.blockSignals(True)
            self.wt_input.setReadOnly(False)
            self.wt_input.setEnabled(True)
            self.wt_input.setFocusPolicy(Qt.StrongFocus)
            # Force immediate visual update
            self.wt_input.update()
            self.wt_input.blockSignals(False)

    def show_log_dialog(self):
        """Show the application log in a dialog window"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Application Log")
        dialog.resize(900, 600)
        layout = QVBoxLayout(dialog)

        # Title
        title = QLabel("<h2>Application Log</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Log text area
        log_text = QTextEdit()
        log_text.setReadOnly(True)
        log_text.setFont(QFont("Consolas", 9))
        log_text.setPlainText(logger.get_logs())
        layout.addWidget(log_text)

        # Button row
        btn_layout = QHBoxLayout()
        
        btn_refresh = QPushButton("Refresh")
        btn_refresh.setStyleSheet("background:#28a745; color:white; font-weight:bold;")
        btn_refresh.clicked.connect(lambda: log_text.setPlainText(logger.get_logs()))
        
        btn_copy = QPushButton("Copy to Clipboard")
        btn_copy.setStyleSheet("background:#0078D7; color:white; font-weight:bold;")
        btn_copy.clicked.connect(lambda: QApplication.clipboard().setText(log_text.toPlainText()))
        
        btn_clear = QPushButton("Clear Log")
        btn_clear.setStyleSheet("background:#dc3545; color:white; font-weight:bold;")
        btn_clear.clicked.connect(lambda: (logger.logs.clear(), log_text.clear()))
        
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        
        btn_layout.addWidget(btn_refresh)
        btn_layout.addWidget(btn_copy)
        btn_layout.addWidget(btn_clear)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        # Auto-refresh timer
        refresh_timer = QTimer(dialog)
        refresh_timer.timeout.connect(lambda: log_text.setPlainText(logger.get_logs()))
        refresh_timer.start(1000)

        self.center_dialog(dialog)
        dialog.exec_()

    def show_references_dialog(self):
        """Show the References dialog with Customs Config, Section 232 Actions, and HTS Database tabs"""
        dialog = QDialog()  # No parent - allows independent movement from main window
        dialog.setWindowTitle("References")
        dialog.resize(2000, 700)
        layout = QVBoxLayout(dialog)

        # Create tab widget
        tabs = QTabWidget()

        # Create new tab widgets for the dialog
        tab_config = QWidget()
        tab_actions = QWidget()
        tab_hts = QWidget()

        # Temporarily swap the instance variables so setup methods populate the new widgets
        original_tab_config = self.tab_config
        original_tab_actions = self.tab_actions

        self.tab_config = tab_config
        self.tab_actions = tab_actions

        # Setup the tabs
        self.setup_config_tab()
        self.setup_actions_tab()
        self.setup_hts_database_tab(tab_hts)

        # Restore original references (though they may be deleted)
        self.tab_config = original_tab_config
        self.tab_actions = original_tab_actions

        # Add the new tabs to the dialog
        tabs.addTab(tab_config, "Customs Config")
        tabs.addTab(tab_actions, "Section 232 Actions")
        tabs.addTab(tab_hts, "HTS Database")

        layout.addWidget(tabs)

        # Close button
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        btn_close.setStyleSheet(self.get_button_style("default"))
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        self.center_dialog(dialog)
        dialog.exec_()

    def show_mid_management_dialog(self):
        """Show the MID Management dialog - redirects to Configuration dialog MID tab"""
        self.show_configuration_dialog(initial_tab=3)
        # Refresh MID combo after dialog closes
        self.load_available_mids()

    def browse_mid_import_file(self):
        """Browse for MID import file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select MID List File", "",
            "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.mid_import_file_path = file_path
            self.mid_import_path_label.setText(Path(file_path).name)
            self.mid_import_path_label.setStyleSheet("color: black;")

    def import_mid_file(self):
        """Import MID list from Excel/CSV file"""
        if not hasattr(self, 'mid_import_file_path') or not self.mid_import_file_path:
            QMessageBox.warning(self, "No File", "Please select a file to import first.")
            return

        try:
            file_path = self.mid_import_file_path
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_excel(file_path, dtype=str, keep_default_na=False)

            df = df.fillna("").rename(columns=str.strip)

            # Map column names (case-insensitive)
            col_map = {}
            for col in df.columns:
                col_lower = col.lower().replace('_', ' ').replace('-', ' ')
                if 'manufacturer' in col_lower and 'name' in col_lower:
                    col_map[col] = 'manufacturer_name'
                elif col_lower == 'mid' or col_lower == 'manufacturer id':
                    col_map[col] = 'mid'
                elif 'customer' in col_lower and 'id' in col_lower:
                    col_map[col] = 'customer_id'
                elif 'related' in col_lower or 'parties' in col_lower:
                    col_map[col] = 'related_parties'

            df = df.rename(columns=col_map)

            # Check for required MID column
            if 'mid' not in df.columns:
                QMessageBox.critical(self, "Error", "File must contain a 'MID' column.")
                return

            # Ask user if they want to append or replace
            existing_count = self.mid_table_widget.rowCount()
            if existing_count > 0:
                reply = QMessageBox.question(
                    self, "Import Mode",
                    f"There are {existing_count} existing MID records.\n\n"
                    "Do you want to ADD to the existing list?\n\n"
                    "Click 'Yes' to append new records\n"
                    "Click 'No' to replace all records",
                    QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
                )
                if reply == QMessageBox.Cancel:
                    return
                if reply == QMessageBox.No:
                    self.mid_table_widget.setRowCount(0)

            # Get existing MIDs to avoid duplicates
            existing_mids = set()
            for row in range(self.mid_table_widget.rowCount()):
                mid_item = self.mid_table_widget.item(row, 1)
                if mid_item:
                    existing_mids.add(mid_item.text().strip().upper())

            # Populate table (append mode)
            imported = 0
            skipped = 0
            for _, row in df.iterrows():
                mid = str(row.get('mid', '')).strip()
                if not mid:
                    continue

                # Skip duplicates
                if mid.upper() in existing_mids:
                    skipped += 1
                    continue

                manufacturer_name = str(row.get('manufacturer_name', '')).strip()
                customer_id = str(row.get('customer_id', '')).strip()
                related_parties = str(row.get('related_parties', 'N')).strip().upper()
                if related_parties not in ('Y', 'N'):
                    related_parties = 'N'

                row_idx = self.mid_table_widget.rowCount()
                self.mid_table_widget.insertRow(row_idx)
                self.mid_table_widget.setItem(row_idx, 0, QTableWidgetItem(manufacturer_name))
                self.mid_table_widget.setItem(row_idx, 1, QTableWidgetItem(mid))
                self.mid_table_widget.setItem(row_idx, 2, QTableWidgetItem(customer_id))

                # Related parties as combo box
                combo = QComboBox()
                combo.addItems(['N', 'Y'])
                combo.setCurrentText(related_parties)
                self.mid_table_widget.setCellWidget(row_idx, 3, combo)
                imported += 1
                existing_mids.add(mid.upper())  # Track for subsequent duplicates in same file

            msg = f"Imported {imported} MID records."
            if skipped > 0:
                msg += f"\nSkipped {skipped} duplicate MIDs."
            msg += "\n\nClick 'Save Changes' to save to database."
            QMessageBox.information(self, "Import Complete", msg)

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import file:\n{str(e)}")
            logger.error(f"MID import error: {e}")

    def add_mid_row(self):
        """Add a new empty row to the MID table"""
        row_idx = self.mid_table_widget.rowCount()
        self.mid_table_widget.insertRow(row_idx)
        self.mid_table_widget.setItem(row_idx, 0, QTableWidgetItem(""))
        self.mid_table_widget.setItem(row_idx, 1, QTableWidgetItem(""))
        self.mid_table_widget.setItem(row_idx, 2, QTableWidgetItem(""))

        combo = QComboBox()
        combo.addItems(['N', 'Y'])
        self.mid_table_widget.setCellWidget(row_idx, 3, combo)

        # Focus on the new row
        self.mid_table_widget.setCurrentCell(row_idx, 0)
        self.mid_table_widget.editItem(self.mid_table_widget.item(row_idx, 0))

    def delete_selected_mid(self):
        """Delete selected MID rows"""
        selected_rows = set(item.row() for item in self.mid_table_widget.selectedItems())
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select rows to delete.")
            return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete {len(selected_rows)} selected MID(s)?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            for row in sorted(selected_rows, reverse=True):
                self.mid_table_widget.removeRow(row)

    def clear_all_mids(self):
        """Clear all MIDs from the table"""
        if self.mid_table_widget.rowCount() == 0:
            return

        reply = QMessageBox.question(
            self, "Confirm Clear",
            "Clear all MIDs from the table?\n\nThis will not delete from database until you click 'Save Changes'.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.mid_table_widget.setRowCount(0)

    def save_mid_table(self):
        """Save MID table data to database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Clear existing data
            c.execute("DELETE FROM mid_table")

            # Insert new data
            saved = 0
            for row in range(self.mid_table_widget.rowCount()):
                manufacturer_name = self.mid_table_widget.item(row, 0)
                manufacturer_name = manufacturer_name.text().strip() if manufacturer_name else ""

                mid = self.mid_table_widget.item(row, 1)
                mid = mid.text().strip() if mid else ""

                if not mid:
                    continue

                customer_id = self.mid_table_widget.item(row, 2)
                customer_id = customer_id.text().strip() if customer_id else ""

                combo = self.mid_table_widget.cellWidget(row, 3)
                related_parties = combo.currentText() if combo else 'N'

                c.execute(
                    "INSERT OR REPLACE INTO mid_table (mid, manufacturer_name, customer_id, related_parties) VALUES (?, ?, ?, ?)",
                    (mid, manufacturer_name, customer_id, related_parties)
                )
                saved += 1

            conn.commit()
            conn.close()

            QMessageBox.information(self, "Saved", f"Saved {saved} MID records to database.")
            self.load_available_mids()

        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save:\n{str(e)}")
            logger.error(f"MID save error: {e}")

    def load_mid_table_data(self):
        """Load MID data from database into table widget"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT manufacturer_name, mid, customer_id, related_parties FROM mid_table ORDER BY manufacturer_name, mid")
            rows = c.fetchall()
            conn.close()

            self.mid_table_widget.setRowCount(0)
            for manufacturer_name, mid, customer_id, related_parties in rows:
                row_idx = self.mid_table_widget.rowCount()
                self.mid_table_widget.insertRow(row_idx)
                self.mid_table_widget.setItem(row_idx, 0, QTableWidgetItem(manufacturer_name or ""))
                self.mid_table_widget.setItem(row_idx, 1, QTableWidgetItem(mid or ""))
                self.mid_table_widget.setItem(row_idx, 2, QTableWidgetItem(customer_id or ""))

                combo = QComboBox()
                combo.addItems(['N', 'Y'])
                combo.setCurrentText(related_parties if related_parties in ('Y', 'N') else 'N')
                self.mid_table_widget.setCellWidget(row_idx, 3, combo)

        except Exception as e:
            logger.error(f"Failed to load MID table: {e}")

    def filter_mid_table(self):
        """Filter the MID table based on Customer ID, MID, and Manufacturer search fields"""
        customer_filter = self.mid_customer_filter.text().strip().upper() if hasattr(self, 'mid_customer_filter') else ''
        mid_filter = self.mid_search_filter.text().strip().upper() if hasattr(self, 'mid_search_filter') else ''
        manufacturer_filter = self.mid_manufacturer_filter.text().strip().upper() if hasattr(self, 'mid_manufacturer_filter') else ''

        for row in range(self.mid_table_widget.rowCount()):
            # Get values from each column
            manufacturer_item = self.mid_table_widget.item(row, 0)
            mid_item = self.mid_table_widget.item(row, 1)
            customer_id_item = self.mid_table_widget.item(row, 2)

            manufacturer = manufacturer_item.text().upper() if manufacturer_item else ''
            mid = mid_item.text().upper() if mid_item else ''
            customer_id = customer_id_item.text().upper() if customer_id_item else ''

            # Determine if row should be visible (all filters must match)
            show_row = True
            if customer_filter and customer_filter not in customer_id:
                show_row = False
            if mid_filter and mid_filter not in mid:
                show_row = False
            if manufacturer_filter and manufacturer_filter not in manufacturer:
                show_row = False

            self.mid_table_widget.setRowHidden(row, not show_row)

    def clear_mid_filters(self):
        """Clear all MID table filters"""
        if hasattr(self, 'mid_customer_filter'):
            self.mid_customer_filter.clear()
        if hasattr(self, 'mid_search_filter'):
            self.mid_search_filter.clear()
        if hasattr(self, 'mid_manufacturer_filter'):
            self.mid_manufacturer_filter.clear()
        # Show all rows
        for row in range(self.mid_table_widget.rowCount()):
            self.mid_table_widget.setRowHidden(row, False)

    def show_configuration_dialog(self, initial_tab=0):
        """Show the Configuration dialog with Invoice Mapping, Output Mapping, Parts Import, and MID Management tabs"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Configuration")
        dialog.resize(1000, 700)
        layout = QVBoxLayout(dialog)

        # Create tab widget
        tabs = QTabWidget()

        # Create new tab widgets for the dialog
        tab_shipment_map = QWidget()
        tab_output_map = QWidget()
        tab_import = QWidget()
        tab_mid_management = QWidget()

        # Temporarily swap the instance variables so setup methods populate the new widgets
        original_tab_shipment_map = self.tab_shipment_map
        original_tab_output_map = self.tab_output_map
        original_tab_import = self.tab_import

        self.tab_shipment_map = tab_shipment_map
        self.tab_output_map = tab_output_map
        self.tab_import = tab_import

        # Setup the tabs
        self.setup_shipment_mapping_tab()
        self.setup_output_mapping_tab()
        self.setup_import_tab()
        self.setup_mid_management_tab(tab_mid_management)

        # Restore original references (though they may be deleted)
        self.tab_shipment_map = original_tab_shipment_map
        self.tab_output_map = original_tab_output_map
        self.tab_import = original_tab_import

        # Add the new tabs to the dialog
        tabs.addTab(tab_shipment_map, "Invoice Mapping Profiles")
        tabs.addTab(tab_output_map, "Output Mapping")
        tabs.addTab(tab_import, "Parts Import")
        tabs.addTab(tab_mid_management, "MID Management")

        # Add Billing tab (only visible to admin)
        if self._is_billing_admin():
            tab_billing = QWidget()
            self.setup_billing_tab(tab_billing)
            tabs.addTab(tab_billing, "Billing")

        # Add AI Agents tab
        tab_ai_agents = QWidget()
        self.setup_ai_agents_tab(tab_ai_agents)
        tabs.addTab(tab_ai_agents, "AI Agents")

        # Set initial tab if specified
        if initial_tab > 0 and initial_tab < tabs.count():
            tabs.setCurrentIndex(initial_tab)

        layout.addWidget(tabs)

        # Close button
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        btn_close.setStyleSheet(self.get_button_style("default"))
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        self.center_dialog(dialog)
        dialog.exec_()

    def setup_mid_management_tab(self, tab_widget):
        """Setup the MID Management tab for the Configuration dialog"""
        layout = QVBoxLayout(tab_widget)

        # Title
        title = QLabel("<h2>Manufacturer ID (MID) Management</h2>")
        layout.addWidget(title)

        # Import section
        import_group = QGroupBox("Import MID List")
        import_layout = QHBoxLayout(import_group)

        self.mid_import_path_label = QLabel("No file selected")
        self.mid_import_path_label.setStyleSheet("color: gray;")
        import_layout.addWidget(self.mid_import_path_label, 1)

        btn_browse = QPushButton("Browse...")
        btn_browse.clicked.connect(self.browse_mid_import_file)
        import_layout.addWidget(btn_browse)

        btn_import = QPushButton("Import")
        btn_import.setStyleSheet(self.get_button_style("primary"))
        btn_import.clicked.connect(self.import_mid_file)
        import_layout.addWidget(btn_import)

        layout.addWidget(import_group)

        # Info label
        info_label = QLabel("Expected Excel columns: <b>Manufacturer Name</b>, <b>MID</b>, <b>Customer ID</b>, <b>Related Parties</b> (Y/N)")
        info_label.setStyleSheet("color: #666; margin: 5px;")
        layout.addWidget(info_label)

        # MID Table
        table_group = QGroupBox("Current MID List")
        table_layout = QVBoxLayout(table_group)

        # Filter/Search row
        filter_layout = QHBoxLayout()

        filter_layout.addWidget(QLabel("Customer ID:"))
        self.mid_customer_filter = QLineEdit()
        self.mid_customer_filter.setPlaceholderText("Filter...")
        self.mid_customer_filter.setMaximumWidth(150)
        self.mid_customer_filter.returnPressed.connect(self.filter_mid_table)
        filter_layout.addWidget(self.mid_customer_filter)

        filter_layout.addWidget(QLabel("MID:"))
        self.mid_search_filter = QLineEdit()
        self.mid_search_filter.setPlaceholderText("Search...")
        self.mid_search_filter.setMaximumWidth(180)
        self.mid_search_filter.returnPressed.connect(self.filter_mid_table)
        filter_layout.addWidget(self.mid_search_filter)

        filter_layout.addWidget(QLabel("Manufacturer:"))
        self.mid_manufacturer_filter = QLineEdit()
        self.mid_manufacturer_filter.setPlaceholderText("Search...")
        self.mid_manufacturer_filter.returnPressed.connect(self.filter_mid_table)
        filter_layout.addWidget(self.mid_manufacturer_filter)

        btn_search = QPushButton("Search")
        btn_search.setStyleSheet(self.get_button_style("primary"))
        btn_search.clicked.connect(self.filter_mid_table)
        filter_layout.addWidget(btn_search)

        btn_clear_filter = QPushButton("Clear Filters")
        btn_clear_filter.clicked.connect(self.clear_mid_filters)
        filter_layout.addWidget(btn_clear_filter)

        table_layout.addLayout(filter_layout)

        self.mid_table_widget = QTableWidget()
        self.mid_table_widget.setColumnCount(4)
        self.mid_table_widget.setHorizontalHeaderLabels(["Manufacturer Name", "MID", "Customer ID", "Related Parties"])
        self.mid_table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.mid_table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.mid_table_widget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.mid_table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.mid_table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.mid_table_widget.setAlternatingRowColors(True)
        table_layout.addWidget(self.mid_table_widget)

        # Table buttons
        table_btn_layout = QHBoxLayout()

        btn_add = QPushButton("Add MID")
        btn_add.clicked.connect(self.add_mid_row)
        table_btn_layout.addWidget(btn_add)

        btn_delete = QPushButton("Delete Selected")
        btn_delete.clicked.connect(self.delete_selected_mid)
        table_btn_layout.addWidget(btn_delete)

        btn_clear = QPushButton("Clear All")
        btn_clear.clicked.connect(self.clear_all_mids)
        table_btn_layout.addWidget(btn_clear)

        table_btn_layout.addStretch()

        btn_save = QPushButton("Save Changes")
        btn_save.setStyleSheet(self.get_button_style("primary"))
        btn_save.clicked.connect(self.save_mid_table)
        table_btn_layout.addWidget(btn_save)

        table_layout.addLayout(table_btn_layout)
        layout.addWidget(table_group)

        # Load current MID data
        self.load_mid_table_data()

    def show_preview_context_menu(self, pos):
        """Show context menu for the preview table"""
        menu = QMenu(self)

        # Get the column under the cursor
        col = self.table.columnAt(pos.x())

        autofit_all = menu.addAction("Auto-fit All Column Widths")
        autofit_all.triggered.connect(self.autofit_preview_columns)

        if col >= 0:
            col_name = self.table.horizontalHeaderItem(col).text() if self.table.horizontalHeaderItem(col) else f"Column {col}"
            autofit_single = menu.addAction(f"Auto-fit '{col_name}' Column")
            autofit_single.triggered.connect(lambda: self.autofit_single_column(col))

        menu.addSeparator()

        reset_widths = menu.addAction("Reset Column Widths")
        reset_widths.triggered.connect(self.reset_preview_column_widths)

        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def autofit_preview_columns(self):
        """Auto-fit all column widths based on content"""
        if not hasattr(self, 'table'):
            return

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        QApplication.processEvents()
        # Switch back to interactive mode to preserve the sizes
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        # Save the new widths
        self.save_column_widths()
        self.bottom_status.setText("Column widths auto-fitted")

    def autofit_single_column(self, col):
        """Auto-fit a single column width based on content"""
        if not hasattr(self, 'table') or col < 0:
            return

        self.table.resizeColumnToContents(col)
        self.save_column_widths()
        col_name = self.table.horizontalHeaderItem(col).text() if self.table.horizontalHeaderItem(col) else f"Column {col}"
        self.bottom_status.setText(f"Auto-fitted column: {col_name}")

    def reset_preview_column_widths(self):
        """Reset all column widths to default"""
        if not hasattr(self, 'table'):
            return

        default_width = 80
        for col in range(self.table.columnCount()):
            self.table.setColumnWidth(col, default_width)

        # Clear saved widths from database
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM app_config WHERE key LIKE 'preview_col_width_%'")
            conn.commit()
            conn.close()
        except Exception as e:
            logger.warning(f"Failed to clear saved column widths: {e}")

        self.bottom_status.setText("Column widths reset to default")

    def show_settings_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Settings")
        dialog.resize(700, 750)  # Increased size for better layout
        layout = QVBoxLayout(dialog)

        # Determine theme-aware colors for dialog elements
        is_dark = hasattr(self, 'current_theme') and self.current_theme in ["Fusion (Dark)", "Ocean"]
        info_text_color = "#aaa" if is_dark else "#666"

        # Create tab widget for better organization
        tabs = QTabWidget()

        # ===== TAB 1: APPEARANCE =====
        appearance_widget = QWidget()
        appearance_layout = QVBoxLayout(appearance_widget)

        # Theme Settings Group
        theme_group = QGroupBox("Appearance")
        theme_layout = QFormLayout()
        
        # Use local variable instead of instance variable
        theme_combo = QComboBox()
        theme_combo.addItems(["System Default", "Fusion (Light)", "macOS", "Fusion (Dark)", "Ocean", "Light Cyan"])
        
        # Load saved theme preference from per-user settings
        saved_theme = get_user_setting('theme', 'Fusion (Light)')
        index = theme_combo.findText(saved_theme)
        if index >= 0:
            # Block signals to prevent double-applying theme
            theme_combo.blockSignals(True)
            theme_combo.setCurrentIndex(index)
            theme_combo.blockSignals(False)
        
        theme_combo.currentTextChanged.connect(self.apply_theme)
        theme_layout.addRow("Application Theme:", theme_combo)

        theme_info = QLabel("<small>Theme changes apply immediately. System Default uses your Windows theme settings.</small>")
        theme_info.setWordWrap(True)
        theme_info.setStyleSheet(f"color:{info_text_color}; padding:5px;")
        theme_layout.addRow("", theme_info)

        # Font Size Slider
        font_size_layout = QHBoxLayout()
        font_size_slider = QSlider(Qt.Horizontal)
        font_size_slider.setMinimum(8)
        font_size_slider.setMaximum(16)
        font_size_slider.setValue(9)  # Default
        font_size_slider.setTickPosition(QSlider.TicksBelow)
        font_size_slider.setTickInterval(1)

        font_size_value_label = QLabel("9pt")
        font_size_value_label.setMinimumWidth(40)

        # Load saved font size preference from per-user settings
        saved_font_size = get_user_setting_int('font_size', 9)
        font_size_slider.setValue(saved_font_size)
        font_size_value_label.setText(f"{saved_font_size}pt")

        # Connect slider to update label and apply font size
        def update_font_size(value):
            font_size_value_label.setText(f"{value}pt")
            self.apply_font_size(value)

        font_size_slider.valueChanged.connect(update_font_size)

        font_size_layout.addWidget(font_size_slider)
        font_size_layout.addWidget(font_size_value_label)

        theme_layout.addRow("Font Size:", font_size_layout)

        # Row Height Slider for Result Preview table
        row_height_layout = QHBoxLayout()
        row_height_slider = QSlider(Qt.Horizontal)
        row_height_slider.setMinimum(22)
        row_height_slider.setMaximum(40)
        row_height_slider.setValue(22)  # Default
        row_height_slider.setTickPosition(QSlider.TicksBelow)
        row_height_slider.setTickInterval(4)

        row_height_value_label = QLabel("22px")
        row_height_value_label.setMinimumWidth(40)

        # Load saved row height preference from per-user settings
        saved_row_height = get_user_setting_int('preview_row_height', 22)
        row_height_slider.setValue(saved_row_height)
        row_height_value_label.setText(f"{saved_row_height}px")

        # Connect slider to update label and apply row height
        def update_row_height(value):
            row_height_value_label.setText(f"{value}px")
            self.apply_row_height(value)

        row_height_slider.valueChanged.connect(update_row_height)

        row_height_layout.addWidget(row_height_slider)
        row_height_layout.addWidget(row_height_value_label)

        theme_layout.addRow("Preview Row Height:", row_height_layout)

        theme_group.setLayout(theme_layout)
        appearance_layout.addWidget(theme_group)

        # Excel Viewer Settings Group
        viewer_group = QGroupBox("Excel File Viewer")
        viewer_layout = QFormLayout()

        # Excel viewer combo box
        viewer_combo = QComboBox()
        if sys.platform == 'linux':
            viewer_combo.addItems(["System Default", "Gnumeric"])
        else:
            viewer_combo.addItems(["System Default"])
            viewer_combo.setEnabled(False)  # Only relevant on Linux

        # Load saved preference (per-user setting)
        saved_viewer = get_user_setting('excel_viewer', 'System Default')
        index = viewer_combo.findText(saved_viewer)
        if index >= 0:
            viewer_combo.setCurrentIndex(index)

        # Save preference when changed (per-user setting)
        def save_viewer_preference(viewer):
            set_user_setting('excel_viewer', viewer)
            logger.info(f"Excel viewer preference changed to: {viewer}")

        viewer_combo.currentTextChanged.connect(save_viewer_preference)
        viewer_layout.addRow("Open Exported Files With:", viewer_combo)

        viewer_info = QLabel("<small>Choose which application opens exported Excel files. (Linux only)</small>")
        viewer_info.setWordWrap(True)
        viewer_info.setStyleSheet(f"color:{info_text_color}; padding:5px;")
        viewer_layout.addRow("", viewer_info)

        viewer_group.setLayout(viewer_layout)
        appearance_layout.addWidget(viewer_group)

        # Preview Table Colors Group - Professional styled color swatches
        # Colors are saved per-theme so each theme can have its own color scheme
        colors_group = QGroupBox("Preview Table Row Colors")
        colors_main_layout = QVBoxLayout()
        colors_main_layout.setSpacing(12)
        colors_main_layout.setContentsMargins(15, 15, 15, 15)

        # Helper function to create a compact color swatch with label
        def create_color_swatch(label_text, config_key, default_color, label_width=70):
            """Create a label with small color swatch button (theme-specific)"""
            container = QWidget()
            container.setMinimumHeight(28)
            layout = QHBoxLayout(container)
            layout.setContentsMargins(0, 2, 8, 2)
            layout.setSpacing(6)

            # Text label with fixed width for alignment
            label = QLabel(label_text + ":")
            label.setFixedWidth(label_width)
            layout.addWidget(label)

            # Small color swatch button
            button = QPushButton()
            button.setFixedSize(20, 20)
            button.setCursor(QCursor(Qt.PointingHandCursor))

            # Load saved color from per-user settings (theme-specific) or use default
            saved_color = get_theme_color(config_key, default_color)

            def update_button_style(color_hex):
                button.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {color_hex};
                        border: 1px solid #555;
                        border-radius: 3px;
                    }}
                    QPushButton:hover {{
                        border: 2px solid #888;
                    }}
                    QPushButton:pressed {{
                        border: 2px solid #aaa;
                    }}
                """)

            update_button_style(saved_color)

            def pick_color():
                current_color = get_theme_color(config_key, default_color)
                color = QColorDialog.getColor(QColor(current_color), dialog, f"Choose {label_text} Color")
                if color.isValid():
                    color_hex = color.name()
                    update_button_style(color_hex)
                    # Save to per-user settings (theme-specific)
                    set_theme_color(config_key, color_hex)
                    logger.info(f"Saved color preference {config_key} for current theme: {color_hex}")
                    # Refresh the preview table if it exists
                    if hasattr(self, 'table') and self.table.rowCount() > 0:
                        self.refresh_preview_colors()
                    # If this is the highlight color, apply it to the application palette
                    if config_key == 'preview_highlight_color':
                        self.apply_highlight_color(color_hex)

            button.clicked.connect(pick_color)
            layout.addWidget(button)
            return container

        # Section 232 Materials header
        sec232_label = QLabel("Section 232 Materials")
        sec232_label.setStyleSheet(f"font-weight: bold; color: {info_text_color}; margin-bottom: 4px; padding: 4px 0;")
        sec232_label.setMinimumHeight(24)
        colors_main_layout.addWidget(sec232_label)

        # First row: Steel, Aluminum, Copper
        row1_layout = QHBoxLayout()
        row1_layout.setSpacing(20)
        row1_layout.addWidget(create_color_swatch("Steel", 'preview_steel_color', '#4a4a4a'))
        row1_layout.addWidget(create_color_swatch("Aluminum", 'preview_aluminum_color', '#3498db'))
        row1_layout.addWidget(create_color_swatch("Copper", 'preview_copper_color', '#e67e22'))
        row1_layout.addStretch()
        colors_main_layout.addLayout(row1_layout)

        # Second row: Wood, Auto, Non-232
        row2_layout = QHBoxLayout()
        row2_layout.setSpacing(20)
        row2_layout.addWidget(create_color_swatch("Wood", 'preview_wood_color', '#27ae60'))
        row2_layout.addWidget(create_color_swatch("Auto", 'preview_auto_color', '#9b59b6'))
        row2_layout.addWidget(create_color_swatch("Non-232", 'preview_non232_color', '#ff0000'))
        row2_layout.addStretch()
        colors_main_layout.addLayout(row2_layout)

        # Separator line
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setStyleSheet("margin: 8px 0;")
        colors_main_layout.addWidget(separator)

        # Other Indicators header
        other_label = QLabel("Other Indicators")
        other_label.setStyleSheet(f"font-weight: bold; color: {info_text_color}; margin-bottom: 4px; padding: 4px 0;")
        other_label.setMinimumHeight(24)
        colors_main_layout.addWidget(other_label)

        # Not Found, Incomplete, and Sec301 rows
        row3_layout = QHBoxLayout()
        row3_layout.setSpacing(20)
        row3_layout.addWidget(create_color_swatch("Not Found", 'preview_notfound_color', '#f39c12'))
        row3_layout.addWidget(create_color_swatch("Incomplete", 'preview_incomplete_color', '#e91e63'))
        row3_layout.addWidget(create_color_swatch("Sec 301 (BG)", 'preview_sec301_bg_color', '#ffc8c8'))
        row3_layout.addStretch()
        colors_main_layout.addLayout(row3_layout)

        # Cell Selection Highlight row
        row4_layout = QHBoxLayout()
        row4_layout.setSpacing(20)
        # Default highlight color depends on theme - use a dark blue that works well with white text
        row4_layout.addWidget(create_color_swatch("Cell Highlight", 'preview_highlight_color', '#1e3c64', label_width=85))
        row4_layout.addStretch()
        colors_main_layout.addLayout(row4_layout)

        colors_group.setLayout(colors_main_layout)
        appearance_layout.addWidget(colors_group)

        # Preview Column Visibility Group
        columns_group = QGroupBox("Result Preview Column Visibility")
        columns_layout = QVBoxLayout()

        # Column names and their default visibility
        column_names = [
            "Product No", "Value", "HTS", "MID", "Qty1", "Qty2", "Qty Unit", "Dec",
            "Melt", "Cast", "Smelt", "Flag", "Steel%", "Al%", "Cu%", "Wood%", "Auto%", "Non-232%", "232 Status", "Cust Ref"
        ]

        # Create checkboxes in a grid layout
        columns_grid = QGridLayout()
        self.preview_column_checkboxes = []  # Store as instance variable

        for i, col_name in enumerate(column_names):
            checkbox = QCheckBox(col_name)
            checkbox.setChecked(True)  # Default to visible

            # Load saved visibility preference
            config_key = f'preview_col_visible_{i}'
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("SELECT value FROM app_config WHERE key = ?", (config_key,))
                row = c.fetchone()
                conn.close()
                if row:
                    checkbox.setChecked(row[0] == '1')
            except:
                pass

            # Save preference and apply when changed
            def make_toggle_handler(col_idx, cb):
                def handler(state):
                    try:
                        conn = sqlite3.connect(str(DB_PATH))
                        c = conn.cursor()
                        c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                                  (f'preview_col_visible_{col_idx}', '1' if state else '0'))
                        conn.commit()
                        conn.close()
                        # Apply visibility to table
                        if hasattr(self, 'table'):
                            self.table.setColumnHidden(col_idx, not state)
                        logger.info(f"Column {col_idx} visibility set to {'visible' if state else 'hidden'}")
                    except Exception as e:
                        logger.error(f"Failed to save column visibility: {e}")
                return handler

            checkbox.stateChanged.connect(make_toggle_handler(i, checkbox))
            self.preview_column_checkboxes.append(checkbox)

            # Arrange in 5 columns
            row_num = i // 5
            col_num = i % 5
            columns_grid.addWidget(checkbox, row_num, col_num)
        
        columns_layout.addLayout(columns_grid)

        # Store column count for reference
        self.preview_column_count = len(column_names)

        columns_info = QLabel("<small>Toggle columns to show or hide them in the Result Preview table.</small>")
        columns_info.setWordWrap(True)
        columns_info.setStyleSheet(f"color:{info_text_color}; padding:5px;")
        columns_layout.addWidget(columns_info)
        
        columns_group.setLayout(columns_layout)
        appearance_layout.addWidget(columns_group)

        # Add stretch to appearance tab
        appearance_layout.addStretch()
        tabs.addTab(appearance_widget, "Appearance")
        # ===== TAB 2: DATABASE =====
        database_widget = QWidget()
        database_layout = QVBoxLayout(database_widget)

        # Current Database Info
        db_info_group = QGroupBox("Current Database")
        db_info_layout = QFormLayout()

        db_path_label = QLabel(str(DB_PATH))
        db_path_label.setWordWrap(True)
        db_path_label.setStyleSheet("font-family: monospace;")

        # Check if using shared or local database
        config = load_shared_config()
        is_windows_platform = sys.platform == 'win32'
        platform_key = 'windows_path' if is_windows_platform else 'linux_path'
        if config.has_option('Database', platform_key):
            platform_name = "Windows" if is_windows_platform else "Linux"
            db_type_text = f"Shared ({platform_name})"
        elif config.has_option('Database', 'path'):
            db_type_text = "Shared (Network)"
        else:
            db_type_text = "Local"
        db_type_label = QLabel(db_type_text)

        db_info_layout.addRow("Type:", db_type_label)
        db_info_layout.addRow("Location:", db_path_label)

        db_info_group.setLayout(db_info_layout)
        database_layout.addWidget(db_info_group)

        # Shared Database Configuration
        shared_group = QGroupBox("Shared Database (Multi-User)")
        shared_layout = QVBoxLayout()

        # Get current platform paths
        platform_paths = get_platform_database_paths()
        is_windows = sys.platform == 'win32'
        current_platform = "Windows" if is_windows else "Linux"

        shared_info = QLabel(
            "Configure platform-specific database paths for cross-platform use.\n"
            f"Current platform: {current_platform}\n\n"
            "When running on Linux, the Linux path is used. When running on Windows, the Windows path is used.\n"
            "This allows the same config.ini to work on both platforms."
        )
        shared_info.setWordWrap(True)
        shared_layout.addWidget(shared_info)

        # Linux path input row
        linux_row = QHBoxLayout()
        linux_label = QLabel("Linux Path:")
        linux_label.setFixedWidth(85)
        if not is_windows:
            linux_label.setStyleSheet("font-weight: bold;")
        linux_row.addWidget(linux_label)

        linux_path_input = QLineEdit()
        linux_path_input.setPlaceholderText("e.g., /home/shared/tariffmill.db")
        linux_path_input.setText(platform_paths.get('linux_path', ''))
        linux_row.addWidget(linux_path_input)

        linux_browse_btn = QPushButton("Browse...")
        def browse_linux_database():
            path, _ = QFileDialog.getOpenFileName(
                dialog, "Select Linux Database File",
                str(Path.home()),
                "SQLite Database (*.db);;All Files (*.*)"
            )
            if path:
                linux_path_input.setText(path)
        linux_browse_btn.clicked.connect(browse_linux_database)
        linux_row.addWidget(linux_browse_btn)
        shared_layout.addLayout(linux_row)

        # Windows path input row
        windows_row = QHBoxLayout()
        windows_label = QLabel("Windows Path:")
        windows_label.setFixedWidth(85)
        if is_windows:
            windows_label.setStyleSheet("font-weight: bold;")
        windows_row.addWidget(windows_label)

        windows_path_input = QLineEdit()
        windows_path_input.setPlaceholderText("e.g., \\\\server\\share\\tariffmill.db or Z:\\shared\\tariffmill.db")
        windows_path_input.setText(platform_paths.get('windows_path', ''))
        windows_row.addWidget(windows_path_input)

        windows_browse_btn = QPushButton("Browse...")
        def browse_windows_database():
            path, _ = QFileDialog.getOpenFileName(
                dialog, "Select Windows Database File",
                str(Path.home()),
                "SQLite Database (*.db);;All Files (*.*)"
            )
            if path:
                windows_path_input.setText(path)
        windows_browse_btn.clicked.connect(browse_windows_database)
        windows_row.addWidget(windows_browse_btn)
        shared_layout.addLayout(windows_row)

        # Action buttons
        btn_row = QHBoxLayout()

        apply_btn = QPushButton("Apply Platform Paths")
        apply_btn.setStyleSheet(self.get_button_style("success"))
        def apply_platform_paths():
            linux_path = linux_path_input.text().strip()
            windows_path = windows_path_input.text().strip()

            if not linux_path and not windows_path:
                QMessageBox.warning(dialog, "No Paths", "Please enter at least one database path.")
                return

            # Validate current platform's path exists
            current_path = linux_path if not is_windows else windows_path
            if current_path:
                path_obj = Path(current_path)
                if not path_obj.exists():
                    reply = QMessageBox.question(dialog, "Database Not Found",
                        f"The file for your current platform does not exist:\n{current_path}\n\n"
                        "Would you like to create a new database at this location?\n"
                        "(A copy of your current database will be created)",
                        QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        try:
                            path_obj.parent.mkdir(parents=True, exist_ok=True)
                            shutil.copy2(str(DB_PATH), str(path_obj))
                        except Exception as e:
                            QMessageBox.critical(dialog, "Error", f"Failed to create database:\n{e}")
                            return
                    else:
                        return

            # Save both paths to config.ini
            if linux_path:
                set_database_path(linux_path, platform='linux')
            if windows_path:
                set_database_path(windows_path, platform='windows')

            # Update display
            active_path = linux_path if not is_windows else windows_path
            if active_path:
                db_path_label.setText(active_path)
                db_type_label.setText(f"Shared ({current_platform})")

            QMessageBox.information(dialog, "Success",
                f"Platform-specific database paths saved.\n\n"
                f"Linux: {linux_path or '(not set)'}\n"
                f"Windows: {windows_path or '(not set)'}\n\n"
                "Restart the application to use the new database.")
        apply_btn.clicked.connect(apply_platform_paths)
        btn_row.addWidget(apply_btn)

        reset_btn = QPushButton("Use Local Database")
        reset_btn.setStyleSheet(self.get_button_style("warning"))
        def reset_to_local():
            config = load_shared_config()
            # Remove all database path options
            for opt in ['path', 'linux_path', 'windows_path']:
                if config.has_option('Database', opt):
                    config.remove_option('Database', opt)
            save_shared_config(config)
            linux_path_input.clear()
            windows_path_input.clear()
            local_path = RESOURCES_DIR / DB_NAME
            db_path_label.setText(str(local_path))
            db_type_label.setText("Local")
            QMessageBox.information(dialog, "Reset",
                "Database reset to local.\n\nRestart the application to apply changes.")
        reset_btn.clicked.connect(reset_to_local)
        btn_row.addWidget(reset_btn)

        shared_layout.addLayout(btn_row)

        # Warning about concurrent access
        warning_label = QLabel(
            "<small><b>Note:</b> SQLite on network shares works best for sequential access. "
            "Avoid having multiple users edit the same record simultaneously.</small>"
        )
        warning_label.setWordWrap(True)
        warning_label.setStyleSheet(f"color:{info_text_color}; padding:5px;")
        shared_layout.addWidget(warning_label)

        shared_group.setLayout(shared_layout)
        database_layout.addWidget(shared_group)

        database_layout.addStretch()
        tabs.addTab(database_widget, "Database")

        # ===== TAB 3: UPDATES =====
        updates_widget = QWidget()
        updates_layout = QVBoxLayout(updates_widget)

        # Update Settings Group
        update_settings_group = QGroupBox("Automatic Update Checks")
        update_settings_layout = QVBoxLayout()

        # Checkbox for startup update check
        startup_check_cb = QCheckBox("Check for updates when application starts")

        # Load saved preference from per-user settings (default True)
        startup_check_cb.setChecked(get_user_setting_bool('check_updates_on_startup', True))

        def save_startup_check_preference(checked):
            set_user_setting('check_updates_on_startup', '1' if checked else '0')
            logger.info(f"Startup update check preference: {'enabled' if checked else 'disabled'}")
        
        startup_check_cb.stateChanged.connect(lambda state: save_startup_check_preference(state == Qt.Checked))
        update_settings_layout.addWidget(startup_check_cb)

        update_info = QLabel(
            "<small>When enabled, the application will check for new releases on GitHub when it starts. "
            "No personal data is sent - only a simple API request to check the latest version.</small>"
        )
        update_info.setWordWrap(True)
        update_info.setStyleSheet(f"color:{info_text_color}; padding:5px;")
        update_settings_layout.addWidget(update_info)

        update_settings_group.setLayout(update_settings_layout)
        updates_layout.addWidget(update_settings_group)

        # Version Info Group
        version_group = QGroupBox("Version Information")
        version_layout = QFormLayout()

        current_version_label = QLabel(f"<b>{VERSION}</b>")
        version_layout.addRow("Current Version:", current_version_label)

        github_link = QLabel(f'<a href="{GITHUB_RELEASES_URL}">View all releases on GitHub</a>')
        github_link.setOpenExternalLinks(True)
        version_layout.addRow("Releases:", github_link)

        version_group.setLayout(version_layout)
        updates_layout.addWidget(version_group)

        # Check Now Button
        check_now_btn = QPushButton("Check for Updates Now")
        check_now_btn.setStyleSheet(self.get_button_style("success"))
        check_now_btn.clicked.connect(lambda: (dialog.close(), self.check_for_updates_manual()))
        updates_layout.addWidget(check_now_btn)

        updates_layout.addStretch()
        tabs.addTab(updates_widget, "Updates")

        # OCRMill settings moved to OCRMill tab's Settings sub-tab

        # Add tabs to main dialog layout
        layout.addWidget(tabs)
        self.center_dialog(dialog)
        dialog.exec_()

    def apply_theme(self, theme_name):
        """Apply the selected theme to the application"""
        app = QApplication.instance()
        
        # Store current theme name
        self.current_theme = theme_name
        
        if theme_name == "System Default":
            app.setStyle("")
            app.setPalette(app.style().standardPalette())
        elif theme_name == "Fusion (Light)":
            app.setStyle("Fusion")
            app.setPalette(app.style().standardPalette())
        elif theme_name == "macOS":
            app.setStyle("Fusion")
            macos_palette = self.get_macos_palette()
            app.setPalette(macos_palette)
        elif theme_name == "Fusion (Dark)":
            app.setStyle("Fusion")
            dark_palette = self.get_dark_palette()
            app.setPalette(dark_palette)
        elif theme_name == "Ocean":
            app.setStyle("Fusion")
            ocean_palette = self.get_ocean_palette()
            app.setPalette(ocean_palette)
        elif theme_name == "Light Cyan":
            app.setStyle("Fusion")
            teal_palette = self.get_teal_professional_palette()
            app.setPalette(teal_palette)
        
        # Apply global stylesheet for QGroupBox and other widgets that don't fully respect QPalette
        is_dark = theme_name in ["Fusion (Dark)", "Ocean"]
        if theme_name == "Ocean":
            # Ocean theme - professional deep blue with gradients and depth
            app.setStyleSheet("""
                QGroupBox {
                    font-weight: normal;
                    border: 1px solid #3a6a9a;
                    border-radius: 6px;
                    margin-top: 12px;
                    padding-top: 10px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #243d5c, stop:1 #1a3050);
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    left: 12px;
                    padding: 2px 8px;
                    color: #7ec8e3;
                    background: #1a3050;
                    border-radius: 3px;
                }
                QTabWidget::pane {
                    border: 1px solid #3a6a9a;
                    border-radius: 6px;
                    background: #1e3a55;
                }
                QTabBar::tab {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #2a4a6a, stop:1 #1e3a55);
                    color: #8ac4e0;
                    padding: 8px 16px;
                    border: 1px solid #3a6a9a;
                    border-bottom: none;
                    border-top-left-radius: 6px;
                    border-top-right-radius: 6px;
                    margin-right: 2px;
                }
                QTabBar::tab:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3a6a9a, stop:1 #2a5070);
                    color: #ffffff;
                    border-bottom: 2px solid #00a8cc;
                }
                QTabBar::tab:hover:!selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #325878, stop:1 #264560);
                    color: #c0e0f0;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1a3550, stop:1 #152a42);
                    color: #e0f0ff;
                    border: 1px solid #3a6a9a;
                    border-radius: 4px;
                    padding: 5px 8px;
                    selection-background-color: #0096b4;
                }
                QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                    border: 1px solid #00a8cc;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1e3a55, stop:1 #183048);
                }
                QComboBox {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1a3550, stop:1 #152a42);
                    color: #e0f0ff;
                    border: 1px solid #3a6a9a;
                    border-radius: 4px;
                    padding: 5px 8px;
                }
                QComboBox::drop-down {
                    border: none;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3a6a9a, stop:1 #2a5070);
                    border-top-right-radius: 4px;
                    border-bottom-right-radius: 4px;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 5px solid transparent;
                    border-right: 5px solid transparent;
                    border-top: 6px solid #a0d0f0;
                    margin-right: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: #1a3550;
                    color: #e0f0ff;
                    selection-background-color: #00a8cc;
                    border: 1px solid #3a6a9a;
                    border-radius: 4px;
                }
                QListWidget {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1a3550, stop:1 #152a42);
                    color: #e0f0ff;
                    border: 1px solid #3a6a9a;
                    border-radius: 4px;
                    alternate-background-color: #1e3a55;
                }
                QListWidget::item {
                    padding: 4px;
                    border-radius: 3px;
                }
                QListWidget::item:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #00b8d4, stop:1 #0096b4);
                    color: #ffffff;
                }
                QListWidget::item:hover:!selected {
                    background: #2a4a6a;
                }
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3a7ca5, stop:1 #2a5a80);
                    color: #ffffff;
                    border: 1px solid #4a8cb5;
                    border-radius: 5px;
                    padding: 6px 14px;
                    font-weight: normal;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #4a8cb5, stop:1 #3a7095);
                    border: 1px solid #5a9cc5;
                }
                QPushButton:pressed {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #2a5a80, stop:1 #1a4a70);
                }
                QScrollBar:vertical {
                    background: #1a3050;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #3a6a9a, stop:1 #4a7aaa);
                    border-radius: 5px;
                    min-height: 30px;
                    margin: 2px;
                }
                QScrollBar::handle:vertical:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4a7aaa, stop:1 #5a8aba);
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar:horizontal {
                    background: #1a3050;
                    height: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:horizontal {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3a6a9a, stop:1 #4a7aaa);
                    border-radius: 5px;
                    min-width: 30px;
                    margin: 2px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #4a7aaa, stop:1 #5a8aba);
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QHeaderView::section {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #2a5070, stop:1 #1e3a55);
                    color: #a0d0f0;
                    padding: 6px;
                    border: none;
                    border-right: 1px solid #3a6a9a;
                    border-bottom: 2px solid #00a8cc;
                    font-weight: normal;
                }
                QTableWidget {
                    background-color: #152a42;
                    alternate-background-color: #1a3050;
                    gridline-color: #2a4a6a;
                    color: #e0f0ff;
                    border: 1px solid #3a6a9a;
                    border-radius: 4px;
                }
                QTableWidget::item:selected {
                    background-color: #1e3c64;
                    color: #ffffff;
                }
                QLabel {
                    color: #c0e0f0;
                }
                QMenu {
                    background-color: #1e3a55;
                    color: #e0f0ff;
                    border: 1px solid #3a6a9a;
                    border-radius: 4px;
                }
                QMenu::item:selected {
                    background-color: #0096b4;
                }
                QMenuBar {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #243d5c, stop:1 #1a3050);
                    color: #c0e0f0;
                }
                QMenuBar::item:selected {
                    background: #3a6a9a;
                    border-radius: 4px;
                }
            """)
        elif is_dark:
            # Fusion Dark theme - professional dark gray with gradients and depth
            app.setStyleSheet("""
                QGroupBox {
                    font-weight: normal;
                    border: 1px solid #4a4a4a;
                    border-radius: 6px;
                    margin-top: 12px;
                    padding-top: 10px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3d3d3d, stop:1 #2d2d2d);
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    left: 12px;
                    padding: 2px 8px;
                    color: #b0b0b0;
                    background: #2d2d2d;
                    border-radius: 3px;
                }
                QTabWidget::pane {
                    border: 1px solid #4a4a4a;
                    border-radius: 6px;
                    background: #353535;
                }
                QTabBar::tab {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #454545, stop:1 #353535);
                    color: #a0a0a0;
                    padding: 8px 16px;
                    border: 1px solid #4a4a4a;
                    border-bottom: none;
                    border-top-left-radius: 6px;
                    border-top-right-radius: 6px;
                    margin-right: 2px;
                }
                QTabBar::tab:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #505050, stop:1 #404040);
                    color: #ffffff;
                    border-bottom: 2px solid #5a6a7a;
                }
                QTabBar::tab:hover:!selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #4a4a4a, stop:1 #3a3a3a);
                    color: #d0d0d0;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #353535, stop:1 #2a2a2a);
                    color: #e0e0e0;
                    border: 1px solid #4a4a4a;
                    border-radius: 4px;
                    padding: 5px 8px;
                    selection-background-color: #4a5a6a;
                }
                QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                    border: 1px solid #5a6a7a;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3a3a3a, stop:1 #2f2f2f);
                }
                QComboBox {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #353535, stop:1 #2a2a2a);
                    color: #e0e0e0;
                    border: 1px solid #4a4a4a;
                    border-radius: 4px;
                    padding: 5px 8px;
                }
                QComboBox::drop-down {
                    border: none;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #505050, stop:1 #404040);
                    border-top-right-radius: 4px;
                    border-bottom-right-radius: 4px;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 5px solid transparent;
                    border-right: 5px solid transparent;
                    border-top: 6px solid #a0a0a0;
                    margin-right: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: #353535;
                    color: #e0e0e0;
                    selection-background-color: #4a5a6a;
                    border: 1px solid #4a4a4a;
                    border-radius: 4px;
                }
                QListWidget {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #353535, stop:1 #2a2a2a);
                    color: #e0e0e0;
                    border: 1px solid #4a4a4a;
                    border-radius: 4px;
                    alternate-background-color: #3a3a3a;
                }
                QListWidget::item {
                    padding: 4px;
                    border-radius: 3px;
                }
                QListWidget::item:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #5a6a7a, stop:1 #4a5a6a);
                    color: #ffffff;
                }
                QListWidget::item:hover:!selected {
                    background: #454545;
                }
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #505050, stop:1 #3a3a3a);
                    color: #e0e0e0;
                    border: 1px solid #555555;
                    border-radius: 5px;
                    padding: 6px 14px;
                    font-weight: normal;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #5a5a5a, stop:1 #454545);
                    border: 1px solid #666666;
                }
                QPushButton:pressed {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3a3a3a, stop:1 #2d2d2d);
                }
                QScrollBar:vertical {
                    background: #2d2d2d;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #505050, stop:1 #5a5a5a);
                    border-radius: 5px;
                    min-height: 30px;
                    margin: 2px;
                }
                QScrollBar::handle:vertical:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #5a5a5a, stop:1 #666666);
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar:horizontal {
                    background: #2d2d2d;
                    height: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:horizontal {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #505050, stop:1 #5a5a5a);
                    border-radius: 5px;
                    min-width: 30px;
                    margin: 2px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #5a5a5a, stop:1 #666666);
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QHeaderView::section {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #454545, stop:1 #353535);
                    color: #c0c0c0;
                    padding: 6px;
                    border: none;
                    border-right: 1px solid #4a4a4a;
                    border-bottom: 2px solid #5a6a7a;
                    font-weight: normal;
                }
                QTableWidget {
                    background-color: #383838;
                    alternate-background-color: #424242;
                    gridline-color: #4a4a4a;
                    color: #e0e0e0;
                    border: 1px solid #505050;
                    border-radius: 4px;
                }
                QTableWidget::item:selected {
                    background-color: #4a6080;
                    color: #ffffff;
                }
                QLabel {
                    color: #c0c0c0;
                }
                QMenu {
                    background-color: #353535;
                    color: #e0e0e0;
                    border: 1px solid #4a4a4a;
                    border-radius: 4px;
                }
                QMenu::item:selected {
                    background-color: #4a6080;
                    color: #ffffff;
                }
                QMenuBar {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3d3d3d, stop:1 #2d2d2d);
                    color: #c0c0c0;
                }
                QMenuBar::item:selected {
                    background: #505050;
                    border-radius: 4px;
                }
            """)
        elif theme_name == "macOS":
            # macOS-inspired theme - clean, minimal Apple aesthetic
            app.setStyleSheet("""
                QGroupBox {
                    font-weight: normal;
                    border: 1px solid #d1d1d6;
                    border-radius: 10px;
                    margin-top: 12px;
                    padding-top: 10px;
                    background-color: #ffffff;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    left: 12px;
                    padding: 2px 8px;
                    color: #1d1d1f;
                    background: #f6f6f6;
                    border-radius: 5px;
                }
                QTabWidget::pane {
                    border: 1px solid #d1d1d6;
                    border-radius: 10px;
                    background: #ffffff;
                }
                QTabBar::tab {
                    background: #f5f5f7;
                    color: #1d1d1f;
                    padding: 8px 16px;
                    border: 1px solid #d1d1d6;
                    border-bottom: none;
                    border-top-left-radius: 8px;
                    border-top-right-radius: 8px;
                    margin-right: 2px;
                }
                QTabBar::tab:selected {
                    background: #ffffff;
                    color: #007aff;
                    border-bottom: 2px solid #007aff;
                }
                QTabBar::tab:hover:!selected {
                    background: #e8e8ed;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox {
                    background: #ffffff;
                    color: #1d1d1f;
                    border: 1px solid #d1d1d6;
                    border-radius: 6px;
                    padding: 6px 10px;
                    selection-background-color: #007aff;
                    selection-color: #ffffff;
                }
                QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                    border: 2px solid #007aff;
                }
                QComboBox {
                    background: #ffffff;
                    color: #1d1d1f;
                    border: 1px solid #d1d1d6;
                    border-radius: 6px;
                    padding: 6px 10px;
                    padding-right: 25px;
                }
                QComboBox:focus {
                    border: 2px solid #007aff;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 5px solid transparent;
                    border-right: 5px solid transparent;
                    border-top: 6px solid #8e8e93;
                    margin-right: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: #ffffff;
                    color: #1d1d1f;
                    selection-background-color: #007aff;
                    selection-color: #ffffff;
                    border: 1px solid #d1d1d6;
                    border-radius: 8px;
                }
                QListWidget {
                    background: #ffffff;
                    color: #1d1d1f;
                    border: 1px solid #d1d1d6;
                    border-radius: 8px;
                    alternate-background-color: #f9f9f9;
                }
                QListWidget::item {
                    padding: 6px;
                    border-radius: 6px;
                }
                QListWidget::item:selected {
                    background: #007aff;
                    color: #ffffff;
                }
                QListWidget::item:hover:!selected {
                    background: #f5f5f7;
                }
                QPushButton {
                    background: #ffffff;
                    color: #007aff;
                    border: 1px solid #007aff;
                    border-radius: 6px;
                    padding: 6px 16px;
                    font-weight: normal;
                }
                QPushButton:hover {
                    background: #007aff;
                    color: #ffffff;
                }
                QPushButton:pressed {
                    background: #005ec4;
                    color: #ffffff;
                }
                QScrollBar:vertical {
                    background: #f6f6f6;
                    width: 10px;
                    border-radius: 5px;
                }
                QScrollBar::handle:vertical {
                    background: #c7c7cc;
                    border-radius: 4px;
                    min-height: 30px;
                    margin: 2px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #aeaeb2;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar:horizontal {
                    background: #f6f6f6;
                    height: 10px;
                    border-radius: 5px;
                }
                QScrollBar::handle:horizontal {
                    background: #c7c7cc;
                    border-radius: 4px;
                    min-width: 30px;
                    margin: 2px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: #aeaeb2;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QHeaderView::section {
                    background: #f5f5f7;
                    color: #1d1d1f;
                    padding: 8px;
                    border: none;
                    border-right: 1px solid #e5e5ea;
                    border-bottom: 1px solid #d1d1d6;
                    font-weight: 500;
                }
                QTableWidget {
                    background-color: #ffffff;
                    alternate-background-color: #f9f9f9;
                    gridline-color: #e5e5ea;
                    color: #1d1d1f;
                    border: 1px solid #d1d1d6;
                    border-radius: 8px;
                }
                QTableWidget::item:selected {
                    background-color: #007aff;
                    color: #ffffff;
                }
                QLabel {
                    color: #1d1d1f;
                }
                QMenu {
                    background-color: #ffffff;
                    color: #1d1d1f;
                    border: 1px solid #d1d1d6;
                    border-radius: 8px;
                    padding: 4px;
                }
                QMenu::item {
                    padding: 6px 20px;
                    border-radius: 4px;
                }
                QMenu::item:selected {
                    background-color: #007aff;
                    color: #ffffff;
                }
                QMenuBar {
                    background: #f6f6f6;
                    color: #1d1d1f;
                    border-bottom: 1px solid #d1d1d6;
                }
                QMenuBar::item:selected {
                    background: #007aff;
                    color: #ffffff;
                    border-radius: 4px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                    border-radius: 4px;
                    border: 1px solid #d1d1d6;
                    background: #ffffff;
                }
                QCheckBox::indicator:checked {
                    background: #007aff;
                    border: 1px solid #007aff;
                }
                QRadioButton::indicator {
                    width: 18px;
                    height: 18px;
                    border-radius: 9px;
                    border: 1px solid #d1d1d6;
                    background: #ffffff;
                }
                QRadioButton::indicator:checked {
                    background: #007aff;
                    border: 5px solid #ffffff;
                    outline: 1px solid #007aff;
                }
            """)
        else:
            # Clear any dark theme stylesheet for light themes
            app.setStyleSheet("")

        # Refresh button styles to match new theme
        self.refresh_button_styles()

        # Refresh text input styles to match new theme
        self.refresh_input_styles()

        # Update file label style for new theme
        if hasattr(self, 'file_label'):
            self.update_file_label_style()

        # Update status bar styles for new theme
        self.update_status_bar_styles()

        # Update table stylesheet for new theme
        self.update_table_stylesheet()

        # Save theme preference (per-user setting)
        set_user_setting('theme', theme_name)
        logger.info(f"Theme changed to: {theme_name}")

        # Re-apply saved font size (theme changes can reset font)
        font_size = get_user_setting_int('font_size', 9)
        self.apply_font_size_without_save(font_size)

        # Refresh preview colors for the new theme (colors are stored per-theme)
        if hasattr(self, 'table') and self.table.rowCount() > 0:
            self.refresh_preview_colors()

        # Apply saved highlight color for this theme
        self.apply_highlight_color()

        # Refresh AI template tab styling for new theme
        if hasattr(self, 'ocrmill_templates_list'):
            self._update_ai_template_styles()

        # Refresh Statistics tab styling for new theme
        if hasattr(self, 'stats_summary_frame'):
            self._update_stats_tab_styles()

        # Update logo for theme
        self.update_logo_for_theme(is_dark)

    def update_logo_for_theme(self, is_dark):
        """Update the header logo based on current theme (dark/light)"""
        if not hasattr(self, 'header_logo_label'):
            return

        if is_dark:
            logo_path = TEMP_RESOURCES_DIR / "tariffmill_logo_small_dark.svg"
        else:
            logo_path = TEMP_RESOURCES_DIR / "tariffmill_logo_small.svg"

        if logo_path.exists():
            pixmap = QPixmap(str(logo_path))
            scaled_pixmap = pixmap.scaledToHeight(60, Qt.SmoothTransformation)
            self.header_logo_label.setPixmap(scaled_pixmap)
        else:
            # Fallback to text with appropriate color
            text_color = "#e0e0e0" if is_dark else "#555555"
            self.header_logo_label.setText(f"{APP_NAME}")
            self.header_logo_label.setStyleSheet(f"""
                font-size: 22px;
                font-weight: bold;
                color: {text_color};
                font-family: 'Impact', 'Arial Black', sans-serif;
            """)

    def apply_font_size_without_save(self, size):
        """Apply font size to application without saving (used internally)"""
        app = QApplication.instance()

        # Get current font and update size
        font = app.font()
        font.setPointSize(size)
        app.setFont(font)

        # Apply font to all existing widgets (app.setFont only affects new widgets)
        def apply_font_recursive(widget):
            widget.setFont(font)
            for child in widget.findChildren(QWidget):
                # Skip widgets with explicit font-size in stylesheet
                style = child.styleSheet()
                if 'font-size' not in style:
                    child.setFont(font)

        apply_font_recursive(self)

    def apply_font_size(self, size):
        """Apply the selected font size to the application and save preference"""
        self.apply_font_size_without_save(size)
        # Save font size preference (per-user setting)
        set_user_setting('font_size', size)
        logger.info(f"Font size changed to: {size}pt")

    def apply_row_height(self, height):
        """Apply the selected row height to the Result Preview table"""
        if hasattr(self, 'table'):
            self.table.verticalHeader().setDefaultSectionSize(height)
            self.table.verticalHeader().setMinimumSectionSize(14)  # Allow small rows
            # Also update existing rows
            for row in range(self.table.rowCount()):
                self.table.setRowHeight(row, height)

        # Save row height preference (per-user setting)
        set_user_setting('preview_row_height', height)
        logger.info(f"Preview row height changed to: {height}px")

    def center_dialog(self, dialog):
        """Center a dialog on the main window"""
        # Standard modal dialog
        dialog.setWindowModality(Qt.WindowModal)

        # Ensure dialog geometry is calculated
        dialog.adjustSize()
        parent_geo = self.frameGeometry()
        dialog_geo = dialog.geometry()

        # Calculate centered position within parent
        x = parent_geo.x() + (parent_geo.width() - dialog_geo.width()) // 2
        y = parent_geo.y() + (parent_geo.height() - dialog_geo.height()) // 2

        dialog.move(x, y)

    def check_for_updates_manual(self):
        """Manually check for updates and show result dialog"""
        self.bottom_status.setText("Checking for updates...")
        QApplication.processEvents()
        
        checker = UpdateChecker(VERSION)
        has_update, latest, url, notes, download_url, error = checker.check_for_updates()
        
        if error:
            QMessageBox.warning(self, "Update Check Failed",
                f"Could not check for updates.\n\n{error}\n\n"
                f"You can check manually at:\n{GITHUB_RELEASES_URL}")
            self.bottom_status.setText("Ready")
            return
        
        if has_update:
            self.show_update_available_dialog(latest, url, notes, download_url)
        else:
            QMessageBox.information(self, "No Updates Available",
                f"You are running the latest version.\n\n"
                f"Current version: {VERSION}\n"
                f"Latest version: {latest}")
        
        self.bottom_status.setText("Ready")

    def show_update_available_dialog(self, latest_version, release_url, release_notes, download_url):
        """Show dialog when an update is available"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Update Available")
        dialog.resize(500, 400)
        layout = QVBoxLayout(dialog)

        # Header
        header = QLabel(f"<h2>🎉 New Version Available!</h2>")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        # Version info
        version_info = QLabel(
            f"<p><b>Current version:</b> {VERSION}<br>"
            f"<b>New version:</b> {latest_version}</p>"
        )
        version_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_info)

        # Release notes
        notes_group = QGroupBox("Release Notes")
        notes_layout = QVBoxLayout()
        notes_text = QTextEdit()
        notes_text.setPlainText(release_notes if release_notes else "No release notes available.")
        notes_text.setReadOnly(True)
        notes_layout.addWidget(notes_text)
        notes_group.setLayout(notes_layout)
        layout.addWidget(notes_group)

        # Buttons
        btn_layout = QHBoxLayout()

        if download_url:
            download_btn = QPushButton("Download Update")
            download_btn.setStyleSheet(self.get_button_style("success"))
            download_btn.clicked.connect(lambda: self._download_and_install_update(download_url, dialog))
            btn_layout.addWidget(download_btn)

        view_btn = QPushButton("View on GitHub")
        view_btn.clicked.connect(lambda: (webbrowser.open(release_url), dialog.accept()))
        btn_layout.addWidget(view_btn)

        later_btn = QPushButton("Remind Me Later")
        later_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(later_btn)

        layout.addLayout(btn_layout)
        self.center_dialog(dialog)
        dialog.exec_()

    def check_for_updates_startup(self):
        """Check for updates on startup (runs in background thread)"""
        # Check if startup update check is enabled (per-user setting, default True)
        if not get_user_setting_bool('check_updates_on_startup', True):
            logger.debug("Startup update check disabled")
            return

        logger.info("Checking for updates on startup...")
        
        def check_thread():
            checker = UpdateChecker(VERSION)
            has_update, latest, url, notes, download_url, error = checker.check_for_updates()
            if has_update and not error:
                # Schedule dialog to be shown on main thread
                QTimer.singleShot(0, lambda: self.show_update_available_dialog(
                    latest, url, notes, download_url))
        
        # Run in background thread to not block startup
        thread = Thread(target=check_thread, daemon=True)
        thread.start()

    def _download_and_install_update(self, download_url: str, dialog: QDialog):
        """Download the update installer and run it."""
        import tempfile
        import subprocess
        import urllib.request

        dialog.accept()

        # Create progress dialog
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("Downloading Update")
        progress_dialog.setFixedSize(400, 120)
        progress_layout = QVBoxLayout(progress_dialog)

        progress_label = QLabel("Downloading update...")
        progress_layout.addWidget(progress_label)

        progress_bar = QProgressBar()
        progress_bar.setRange(0, 100)
        progress_layout.addWidget(progress_bar)

        cancel_btn = QPushButton("Cancel")
        progress_layout.addWidget(cancel_btn)

        cancelled = [False]
        cancel_btn.clicked.connect(lambda: cancelled.__setitem__(0, True))
        cancel_btn.clicked.connect(progress_dialog.reject)

        progress_dialog.show()
        QApplication.processEvents()

        try:
            # Extract filename from URL
            filename = download_url.split('/')[-1]
            if not filename.endswith('.exe'):
                filename = 'TariffMill_Setup.exe'

            # Download to temp directory
            temp_dir = tempfile.gettempdir()
            temp_path = Path(temp_dir) / filename

            # Download with progress
            request = urllib.request.Request(
                download_url,
                headers={'User-Agent': f'TariffMill/{VERSION}'}
            )

            with urllib.request.urlopen(request, timeout=60) as response:
                total_size = int(response.headers.get('content-length', 0))
                downloaded = 0
                block_size = 8192

                with open(temp_path, 'wb') as f:
                    while True:
                        if cancelled[0]:
                            progress_dialog.close()
                            return

                        chunk = response.read(block_size)
                        if not chunk:
                            break

                        f.write(chunk)
                        downloaded += len(chunk)

                        if total_size > 0:
                            percent = int(downloaded * 100 / total_size)
                            progress_bar.setValue(percent)
                            progress_label.setText(f"Downloading... {downloaded // 1024 // 1024} MB / {total_size // 1024 // 1024} MB")

                        QApplication.processEvents()

            progress_dialog.close()

            # Ask user to confirm installation
            reply = QMessageBox.question(
                self, "Install Update",
                f"Download complete!\n\nDo you want to install the update now?\n\n"
                f"The application will close and the installer will start.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )

            if reply == QMessageBox.Yes:
                # Run the installer and exit
                logger.info(f"Starting installer: {temp_path}")

                # Use os.startfile on Windows for reliable launching
                if sys.platform == 'win32':
                    os.startfile(str(temp_path))
                else:
                    subprocess.Popen([str(temp_path)], shell=True)

                # Give the installer time to start, then exit
                QTimer.singleShot(500, QApplication.quit)

        except Exception as e:
            progress_dialog.close()
            QMessageBox.critical(
                self, "Download Failed",
                f"Failed to download update:\n\n{str(e)}\n\n"
                f"Please download manually from GitHub."
            )
            logger.error(f"Update download failed: {e}")

    def show_license_dialog(self):
        """Show the license activation dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle("License & Activation")
        dialog.resize(500, 400)
        dialog.setMinimumWidth(450)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Get current license status
        license_mgr = LicenseManager(DB_PATH)
        status, days = license_mgr.get_license_status()

        # Status display
        status_group = QGroupBox("License Status")
        status_layout = QVBoxLayout()
        status_layout.setSpacing(10)

        if status == 'active':
            status_icon = "✅"
            status_text = "Licensed"
            status_color = "#27ae60"
            status_detail = f"Email: {license_mgr.license_email or 'N/A'}"
        elif status == 'trial':
            status_icon = "⏳"
            status_text = f"Trial Period - {days} days remaining"
            status_color = "#f39c12"
            status_detail = "Full functionality available during trial"
        else:  # expired
            status_icon = "❌"
            status_text = "Trial Expired"
            status_color = "#e74c3c"
            status_detail = "Please activate a license to continue using the software"

        status_label = QLabel(f"<h2 style='color: {status_color};'>{status_icon} {status_text}</h2>")
        status_label.setAlignment(Qt.AlignCenter)
        status_layout.addWidget(status_label)

        detail_label = QLabel(status_detail)
        detail_label.setAlignment(Qt.AlignCenter)
        detail_label.setStyleSheet("color: #666;")
        status_layout.addWidget(detail_label)

        status_group.setLayout(status_layout)
        layout.addWidget(status_group)

        # License key input
        activate_group = QGroupBox("Activate License")
        activate_layout = QVBoxLayout()
        activate_layout.setSpacing(10)

        key_label = QLabel("Enter your license key:")
        activate_layout.addWidget(key_label)

        key_input = QLineEdit()
        key_input.setPlaceholderText("XXXXXXXX-XXXXXXXX-XXXXXXXX-XXXXXXXX")
        key_input.setMinimumHeight(35)
        key_input.setStyleSheet("padding: 5px;")
        activate_layout.addWidget(key_input)

        # Message label for feedback
        message_label = QLabel("")
        message_label.setWordWrap(True)
        message_label.setAlignment(Qt.AlignCenter)
        activate_layout.addWidget(message_label)

        # Activate button
        activate_btn = QPushButton("Activate License")
        activate_btn.setMinimumHeight(35)
        activate_btn.setStyleSheet(self.get_button_style("success"))

        def activate_license():
            key = key_input.text().strip()
            if not key:
                message_label.setText("<span style='color: #e74c3c;'>Please enter a license key</span>")
                return

            activate_btn.setEnabled(False)
            activate_btn.setText("Validating...")
            QApplication.processEvents()

            success, message = license_mgr.activate_license(key)

            if success:
                message_label.setText(f"<span style='color: #27ae60;'>{message}</span>")
                # Update status display
                status_label.setText("<h2 style='color: #27ae60;'>✅ Licensed</h2>")
                detail_label.setText(f"Email: {license_mgr.license_email or 'N/A'}")
                # Update window title
                self.update_license_status_title()
                QMessageBox.information(dialog, "License Activated",
                    "Your license has been activated successfully!\n\n"
                    "Thank you for purchasing TariffMill.")
            else:
                message_label.setText(f"<span style='color: #e74c3c;'>{message}</span>")

            activate_btn.setEnabled(True)
            activate_btn.setText("Activate License")

        activate_btn.clicked.connect(activate_license)
        activate_layout.addWidget(activate_btn)

        activate_group.setLayout(activate_layout)
        layout.addWidget(activate_group)

        # Purchase section
        purchase_group = QGroupBox("Don't have a license?")
        purchase_layout = QVBoxLayout()

        purchase_info = QLabel(
            "Purchase a license to support development and unlock permanent access."
        )
        purchase_info.setWordWrap(True)
        purchase_info.setAlignment(Qt.AlignCenter)
        purchase_layout.addWidget(purchase_info)

        buy_btn = QPushButton("Buy License")
        buy_btn.setMinimumHeight(35)
        buy_btn.setStyleSheet(self.get_button_style("default"))
        buy_btn.clicked.connect(lambda: webbrowser.open(GUMROAD_PRODUCT_URL))
        purchase_layout.addWidget(buy_btn)

        purchase_group.setLayout(purchase_layout)
        layout.addWidget(purchase_group)

        # Close button
        close_btn = QPushButton("Close")
        close_btn.setMinimumHeight(35)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)

        self.center_dialog(dialog)
        dialog.exec_()

    def show_trial_expired_dialog(self):
        """Show modal dialog when trial has expired - blocks app until licensed"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Trial Expired")
        dialog.resize(450, 350)
        dialog.setModal(True)
        # Prevent closing without activating license
        dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowCloseButtonHint)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)

        # Warning header
        header = QLabel("<h2 style='color: #e74c3c;'>⚠️ Trial Period Expired</h2>")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        # Message
        message = QLabel(
            "<p style='font-size: 13px;'>Your 30-day trial of TariffMill has ended.</p>"
            "<p style='font-size: 13px;'>Please purchase and activate a license to continue "
            "using the software.</p>"
        )
        message.setWordWrap(True)
        message.setAlignment(Qt.AlignCenter)
        layout.addWidget(message)

        layout.addSpacing(10)

        # License key input
        key_label = QLabel("Enter your license key:")
        layout.addWidget(key_label)

        key_input = QLineEdit()
        key_input.setPlaceholderText("XXXXXXXX-XXXXXXXX-XXXXXXXX-XXXXXXXX")
        key_input.setMinimumHeight(35)
        key_input.setStyleSheet("padding: 5px;")
        layout.addWidget(key_input)

        # Message label for feedback
        feedback_label = QLabel("")
        feedback_label.setWordWrap(True)
        feedback_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(feedback_label)

        # Buttons
        btn_layout = QHBoxLayout()

        activate_btn = QPushButton("Activate License")
        activate_btn.setMinimumHeight(35)
        activate_btn.setStyleSheet(self.get_button_style("success"))

        def try_activate():
            key = key_input.text().strip()
            if not key:
                feedback_label.setText("<span style='color: #e74c3c;'>Please enter a license key</span>")
                return

            activate_btn.setEnabled(False)
            activate_btn.setText("Validating...")
            QApplication.processEvents()

            license_mgr = LicenseManager(DB_PATH)
            success, msg = license_mgr.activate_license(key)

            if success:
                feedback_label.setText(f"<span style='color: #27ae60;'>{msg}</span>")
                QMessageBox.information(dialog, "License Activated",
                    "Your license has been activated successfully!\n\n"
                    "Thank you for purchasing TariffMill.")
                self.update_license_status_title()
                dialog.accept()
            else:
                feedback_label.setText(f"<span style='color: #e74c3c;'>{msg}</span>")
                activate_btn.setEnabled(True)
                activate_btn.setText("Activate License")

        activate_btn.clicked.connect(try_activate)
        btn_layout.addWidget(activate_btn)

        buy_btn = QPushButton("Buy License")
        buy_btn.setMinimumHeight(35)
        buy_btn.clicked.connect(lambda: webbrowser.open(GUMROAD_PRODUCT_URL))
        btn_layout.addWidget(buy_btn)

        layout.addLayout(btn_layout)

        # Exit button (closes the application)
        exit_btn = QPushButton("Exit Application")
        exit_btn.setMinimumHeight(30)
        exit_btn.setStyleSheet(self.get_button_style("danger"))
        exit_btn.clicked.connect(lambda: (dialog.reject(), QApplication.quit()))
        layout.addWidget(exit_btn)

        # Show dialog - if rejected (exit clicked), close app
        self.center_dialog(dialog)
        result = dialog.exec_()
        if result == QDialog.Rejected:
            QApplication.quit()
            sys.exit(0)

    def check_license_status(self):
        """Check license status at startup and show appropriate dialog if needed"""
        license_mgr = LicenseManager(DB_PATH)
        status, days = license_mgr.get_license_status()

        logger.info(f"License status: {status}, days remaining: {days}")

        # Update window title with license status
        self.update_license_status_title()

        if status == 'expired':
            # Show trial expired dialog (blocks until licensed or exit)
            self.show_trial_expired_dialog()
        elif status == 'trial':
            # Show trial reminder if less than 7 days remaining
            if days <= 7:
                QMessageBox.information(self, "Trial Reminder",
                    f"Your trial period will expire in {days} day{'s' if days != 1 else ''}.\n\n"
                    "Please consider purchasing a license to continue using TariffMill.\n\n"
                    "Go to Help → License & Activation to enter your license key.")

    def update_license_status_title(self):
        """Update the window title to reflect license status"""
        license_mgr = LicenseManager(DB_PATH)
        status, days = license_mgr.get_license_status()

        if status == 'active':
            title_suffix = "(Licensed)"
        elif status == 'trial':
            title_suffix = f"(Trial: {days} days left)"
        else:
            title_suffix = "(Unlicensed)"

        self.setWindowTitle(f"{APP_NAME} {VERSION} {title_suffix}")

    def show_about_dialog(self):
        """Show the About dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"About {APP_NAME}")
        dialog.resize(400, 300)
        layout = QVBoxLayout(dialog)

        # App icon and name
        header_layout = QHBoxLayout()
        
        # Try to load app icon - prefer SVG for higher quality display
        icon_path = TEMP_RESOURCES_DIR / "tariffmill_icon_hybrid_2.svg"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "icon.ico"

        if icon_path.exists():
            icon_label = QLabel()
            pixmap = QPixmap(str(icon_path))
            icon_label.setPixmap(pixmap.scaled(64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            header_layout.addWidget(icon_label)
        
        title_layout = QVBoxLayout()
        title_label = QLabel(f"<h2>{APP_NAME}</h2>")
        title_layout.addWidget(title_label)
        version_label = QLabel(f"Version {VERSION}")
        version_label.setStyleSheet("color: #666;")
        title_layout.addWidget(version_label)
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        
        layout.addLayout(header_layout)
        layout.addSpacing(20)

        # Description
        desc_label = QLabel(
            "<p>Professional customs documentation processing system for "
            "invoice processing, parts management, and Section 232 tariff "
            "compliance tracking.</p>"
        )
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)

        layout.addSpacing(10)

        # Links
        github_label = QLabel(
            f'<p>GitHub: <a href="{GITHUB_RELEASES_URL}">{GITHUB_REPO}</a></p>'
        )
        github_label.setOpenExternalLinks(True)
        layout.addWidget(github_label)

        layout.addSpacing(10)

        # Copyright notice
        copyright_label = QLabel(
            "<p style='color: #888; font-size: 10px;'>"
            f"Copyright (c) {COPYRIGHT_YEAR} {COPYRIGHT_HOLDER}. All Rights Reserved.<br>"
            "This software is proprietary and confidential.<br>"
            "Unauthorized copying or distribution is prohibited."
            "</p>"
        )
        copyright_label.setWordWrap(True)
        copyright_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(copyright_label)

        layout.addStretch()

        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignCenter)

        self.center_dialog(dialog)
        dialog.exec_()

    def update_file_label_style(self):
        """Update file label background based on current theme"""
        from PyQt5.QtGui import QPalette

        if not hasattr(self, 'current_theme'):
            self.current_theme = "System Default"

        # Use palette Base color for consistent styling with result preview
        app = QApplication.instance()
        palette = app.palette()
        base_color = palette.color(QPalette.Base)
        text_color = palette.color(QPalette.Text)

        bg_color = base_color.name()
        fg_color = text_color.name()

        self.file_label.setStyleSheet(f"background:{bg_color}; padding:5px; border:1px solid #555; color:{fg_color};")
    
    def update_status_bar_styles(self):
        """Update status bar backgrounds based on current theme"""
        if not hasattr(self, 'current_theme'):
            self.current_theme = "System Default"

        if self.current_theme == "Ocean":
            # Ocean theme status bars - professional deep blue with gradient
            self.status.setStyleSheet("font-size:14pt; padding:8px; background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #243d5c, stop:1 #1a3050); color:#c0e0f0;")
            self.bottom_status.setStyleSheet("font-size:9px; color:#8ac4e0;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #1e3a55, stop:1 #152a42);
                        border-top: 2px solid #00a8cc;
                    }
                """)
        elif self.current_theme == "Fusion (Dark)":
            # Fusion Dark theme status bars - professional with gradient
            self.status.setStyleSheet("font-size:14pt; padding:8px; background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3d3d3d, stop:1 #2d2d2d); color:#e0e0e0;")
            self.bottom_status.setStyleSheet("font-size:9px; color:#a0a0a0;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #353535, stop:1 #2a2a2a);
                        border-top: 2px solid #5a6a7a;
                    }
                """)
        else:
            # Light theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0; color:#000000;")
            self.bottom_status.setStyleSheet("font-size:9px; color:#555555;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #e8e8e8;
                        border-top: 1px solid #d0d0d0;
                    }
                """)

    def update_status_bar_styles(self):
        """Update status bar backgrounds based on current theme"""
        if not hasattr(self, 'current_theme'):
            self.current_theme = "System Default"

        is_dark = self.current_theme in ["Fusion (Dark)", "Ocean"]

        if is_dark:
            # Dark theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#2d2d2d; color:#e0e0e0;")
            self.bottom_status.setStyleSheet("font-size:9px; color:#b0b0b0;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #2d2d2d;
                        border-top: 1px solid #404040;
                    }
                """)
        else:
            # Light theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0; color:#000000;")
            self.bottom_status.setStyleSheet("font-size:9px; color:#555555;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #e8e8e8;
                        border-top: 1px solid #d0d0d0;
                    }
                """)

    def update_table_stylesheet(self):
        """Update table stylesheet with green focus color for current theme"""
        if not hasattr(self, 'table'):
            return

        # Set green focus color and reduced cell padding
        self.table.setStyleSheet("""
            QTableWidget::item {
                padding: 1px 3px;
            }
            QTableWidget::item:focus {
                background-color: #90EE90;
                border: 2px solid #228B22;
            }
        """)

    def get_preview_row_color(self, material_flag):
        """Get the color for preview table rows based on Section 232 material type

        Args:
            material_flag: String like '232_Steel', '232_Aluminum', '232_Copper', '232_Wood', '232_Auto', 'Non_232', or boolean for backward compatibility

        Returns:
            QColor: Color for the row based on material type (theme-specific)
        """
        # Default colors for each material type
        default_colors = {
            '232_Steel': '#4a4a4a',      # Dark gray
            '232_Aluminum': '#3498db',   # Blue
            '232_Copper': '#e67e22',     # Orange
            '232_Wood': '#27ae60',       # Green
            '232_Auto': '#9b59b6',       # Purple
            'Non_232': '#ff0000',        # Red
            'Not_Found': '#f39c12',      # Yellow/Gold - indicates part not in database
            'Incomplete': '#e91e63'      # Pink - part in DB but missing HTS/ratios
        }

        # Handle backward compatibility - if passed a boolean
        if isinstance(material_flag, bool):
            material_flag = '232_Steel' if material_flag else 'Non_232'

        # Determine material type from flag
        if not material_flag or material_flag == 'Non_232':
            color_key = 'preview_non232_color'
            default_color = default_colors['Non_232']
        elif material_flag == 'Not_Found':
            color_key = 'preview_notfound_color'
            default_color = default_colors['Not_Found']
        elif material_flag == 'Incomplete':
            color_key = 'preview_incomplete_color'
            default_color = default_colors['Incomplete']
        elif material_flag.startswith('232_'):
            # Map flag to color key
            material = material_flag  # e.g., '232_Steel'
            color_key = f'preview_{material.lower().replace("232_", "")}_color'  # e.g., 'preview_steel_color'
            default_color = default_colors.get(material, default_colors['232_Steel'])
        else:
            # Unknown flag, treat as non-232
            color_key = 'preview_non232_color'
            default_color = default_colors['Non_232']

        # Get color from per-user settings (theme-specific)
        saved_color = get_theme_color(color_key, default_color)
        return QColor(saved_color)

    def get_sec301_bg_color(self):
        """Get the background color for Section 301 exclusion rows (theme-specific)

        Returns:
            QColor: Background color for Sec301 exclusion rows
        """
        default_color = '#ffc8c8'  # Light red background
        saved_color = get_theme_color('preview_sec301_bg_color', default_color)
        return QColor(saved_color)

    def refresh_preview_colors(self):
        """Refresh all row colors in the preview table based on current settings"""
        if not hasattr(self, 'table') or self.table.rowCount() == 0:
            return

        try:
            # Temporarily disconnect itemChanged signal to avoid triggering edits
            self.table.blockSignals(True)

            # Get the current Sec301 background color setting
            sec301_bg_color = self.get_sec301_bg_color()

            for row in range(self.table.rowCount()):
                # Check the 232 Status column (index 18) to determine material type
                # Column order: 0=Product No, 1=Value, 2=HTS, 3=MID, 4=Qty1, 5=Qty2, 6=Qty Unit, 7=Dec,
                # 8=Melt, 9=Cast, 10=Smelt, 11=Flag, 12=Steel%, 13=Al%, 14=Cu%, 15=Wood%, 16=Auto%, 17=Non-232%, 18=232 Status
                status_item = self.table.item(row, 18)
                status_text = status_item.text() if status_item else ''

                # Get color based on material flag (232_Steel, 232_Aluminum, etc.)
                row_color = self.get_preview_row_color(status_text)

                # Check if this row has Sec301 exclusion (stored in UserRole + 1)
                first_item = self.table.item(row, 0)
                has_sec301 = first_item and first_item.data(Qt.UserRole + 1) == 'sec301_exclusion'

                # Update color for all items in this row
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item:
                        item.setForeground(row_color)
                        # Apply Sec301 background color if applicable
                        if has_sec301:
                            item.setBackground(sec301_bg_color)

            self.table.blockSignals(False)
        except Exception as e:
            logger.error(f"Error refreshing preview colors: {e}")
            self.table.blockSignals(False)

    def apply_highlight_color(self, color_hex=None):
        """Apply cell selection highlight color to the application palette and table.

        Args:
            color_hex: Hex color string (e.g., '#1e3c64'). If None, loads from settings.
        """
        if color_hex is None:
            # Get default based on current theme
            theme = get_user_setting('theme', 'Fusion (Light)')
            if 'ocean' in theme.lower():
                default_highlight = '#1e3c64'
            elif 'dark' in theme.lower():
                default_highlight = '#4a6080'
            else:
                default_highlight = '#3399ff'
            color_hex = get_theme_color('preview_highlight_color', default_highlight)

        try:
            from PyQt5.QtGui import QPalette, QColor

            # Update the application palette
            app = QApplication.instance()
            if app:
                palette = app.palette()
                palette.setColor(QPalette.Highlight, QColor(color_hex))
                palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
                app.setPalette(palette)

                # Also update the app stylesheet's QTableWidget::item:selected color
                current_style = app.styleSheet()
                if current_style:
                    import re
                    # Replace any existing item:selected background-color in app stylesheet
                    new_style = re.sub(
                        r'(QTableWidget::item:selected\s*\{[^}]*background-color:)\s*[^;]*(;)',
                        f'\\1 {color_hex}\\2',
                        current_style
                    )
                    if new_style != current_style:
                        app.setStyleSheet(new_style)

            # Also update the table's own stylesheet if it has item:selected
            if hasattr(self, 'table'):
                current_style = self.table.styleSheet()
                if 'item:selected' in current_style:
                    import re
                    new_style = re.sub(
                        r'(QTableWidget::item:selected\s*\{[^}]*background-color:)\s*[^;]*(;)',
                        f'\\1 {color_hex}\\2',
                        current_style
                    )
                    self.table.setStyleSheet(new_style)
                else:
                    # Add item:selected style to table
                    new_style = current_style + f"""
                        QTableWidget::item:selected {{
                            background-color: {color_hex};
                            color: #ffffff;
                        }}
                    """
                    self.table.setStyleSheet(new_style)
                self.table.viewport().update()

            logger.info(f"Applied highlight color: {color_hex}")
        except Exception as e:
            logger.error(f"Error applying highlight color: {e}")

    def apply_column_visibility(self):
        """Apply saved column visibility settings to the preview table"""
        if not hasattr(self, 'table'):
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Apply saved settings for each column
            for col_idx in range(self.table.columnCount()):
                config_key = f'preview_col_visible_{col_idx}'
                c.execute("SELECT value FROM app_config WHERE key = ?", (config_key,))
                row = c.fetchone()
                # Default to visible if no setting saved
                is_visible = True if row is None else (row[0] == '1')
                self.table.setColumnHidden(col_idx, not is_visible)

            conn.close()
        except Exception as e:
            logger.error(f"Error applying column visibility: {e}")

    def refresh_button_styles(self):
        """Refresh all button styles to match current theme"""
        # Process tab buttons
        if hasattr(self, 'clear_btn'):
            self.clear_btn.setStyleSheet(self.get_button_style("danger"))
        if hasattr(self, 'process_btn'):
            self.process_btn.setStyleSheet(self.get_button_style("success"))
        if hasattr(self, 'edit_values_btn'):
            self.edit_values_btn.setStyleSheet(self.get_button_style("warning"))
        
        # Import tab - need to find and update buttons in the tab
        if hasattr(self, 'tab_import'):
            for btn in self.tab_import.findChildren(QPushButton):
                if btn.text() == "Load CSV File":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Reset Mapping":
                    btn.setStyleSheet(self.get_button_style("danger"))
                elif btn.text() == "IMPORT NOW":
                    btn.setStyleSheet(self.get_button_style("success") + "QPushButton { font-size:16pt; padding:15px; }")
        
        # Shipment Mapping tab
        if hasattr(self, 'tab_shipment_map'):
            for btn in self.tab_shipment_map.findChildren(QPushButton):
                if btn.text() == "Save Current Mapping As...":
                    btn.setStyleSheet(self.get_button_style("success"))
                elif btn.text() == "Delete Profile":
                    btn.setStyleSheet(self.get_button_style("danger"))
                elif btn.text() == "Load CSV to Map":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Reset Current":
                    btn.setStyleSheet(self.get_button_style("danger"))
        
        # Master/Parts View tab
        if hasattr(self, 'tab_master'):
            for btn in self.tab_master.findChildren(QPushButton):
                if btn.text() == "Run Query":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Execute":
                    btn.setStyleSheet(self.get_button_style("success"))
        
        # Config tab
        if hasattr(self, 'tab_config'):
            for btn in self.tab_config.findChildren(QPushButton):
                if btn.text() == "Import Section 232 Tariffs (CSV/Excel)":
                    btn.setStyleSheet(self.get_button_style("success"))
                elif btn.text() == "Import from CSV":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Refresh View":
                    btn.setStyleSheet(self.get_button_style("info"))

    def refresh_input_styles(self):
        """Refresh all text input styles to match current theme"""
        input_style = self.get_input_style()

        # Parts View tab inputs - Parts Search
        if hasattr(self, 'search_value_input'):
            self.search_value_input.setStyleSheet(input_style)
        if hasattr(self, 'custom_sql_input'):
            self.custom_sql_input.setStyleSheet(input_style)
        if hasattr(self, 'search_input'):
            self.search_input.setStyleSheet(input_style)

        # Customs Config tab inputs
        if hasattr(self, 'tariff_filter'):
            self.tariff_filter.setStyleSheet(input_style)

        # Section 232 Actions tab inputs
        if hasattr(self, 'actions_filter'):
            self.actions_filter.setStyleSheet(input_style)

    def get_dark_palette(self):
        """Create a Windows 11 dark mode inspired theme"""
        from PyQt5.QtGui import QPalette, QColor

        # Get user's saved highlight color for this theme
        highlight_color = get_theme_color('preview_highlight_color', '#1678d4', 'Fusion (Dark)')

        palette = QPalette()
        # Windows 11 dark theme colors
        palette.setColor(QPalette.Window, QColor(45, 45, 45))  # Main background
        palette.setColor(QPalette.WindowText, QColor(243, 243, 243))  # Primary text
        palette.setColor(QPalette.Base, QColor(45, 45, 45))  # Text box background (matches main background)
        palette.setColor(QPalette.AlternateBase, QColor(115, 115, 115))  # Tertiary background for alternating rows
        palette.setColor(QPalette.ToolTipBase, QColor(45, 45, 45))  # Tertiary background
        palette.setColor(QPalette.ToolTipText, QColor(243, 243, 243))  # Primary text
        palette.setColor(QPalette.Text, QColor(243, 243, 243))  # Primary text in text boxes
        palette.setColor(QPalette.Button, QColor(45, 45, 45))  # Tertiary background for buttons
        palette.setColor(QPalette.ButtonText, QColor(243, 243, 243))  # Primary text on buttons
        palette.setColor(QPalette.BrightText, QColor(164, 38, 44))  # Danger/error red
        palette.setColor(QPalette.Link, QColor(0, 120, 212))  # Accent blue
        palette.setColor(QPalette.Highlight, QColor(highlight_color))  # User's saved highlight color
        palette.setColor(QPalette.HighlightedText, QColor(243, 243, 243))  # Primary text
        return palette

    def get_ocean_palette(self):
        """Create an ocean-themed color palette with deep blues and teals - professional look"""
        from PyQt5.QtGui import QPalette, QColor

        # Get user's saved highlight color for this theme
        highlight_color = get_theme_color('preview_highlight_color', '#1e3c64', 'Ocean')

        palette = QPalette()
        # Deep ocean blue backgrounds with more contrast
        palette.setColor(QPalette.Window, QColor(26, 48, 80))  # Main window background
        palette.setColor(QPalette.WindowText, QColor(192, 224, 240))  # Soft blue-white text
        palette.setColor(QPalette.Base, QColor(21, 42, 66))  # Input/table background (darker)
        palette.setColor(QPalette.AlternateBase, QColor(26, 48, 80))  # Alternating rows
        palette.setColor(QPalette.ToolTipBase, QColor(30, 58, 85))  # Tooltip background
        palette.setColor(QPalette.ToolTipText, QColor(224, 240, 255))  # Tooltip text
        palette.setColor(QPalette.Text, QColor(224, 240, 255))  # Input text color
        palette.setColor(QPalette.Button, QColor(58, 106, 154))  # Button background
        palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))  # Button text
        palette.setColor(QPalette.BrightText, QColor(0, 200, 220))  # Bright accent
        palette.setColor(QPalette.Link, QColor(0, 168, 204))  # Link color (teal)
        palette.setColor(QPalette.Highlight, QColor(highlight_color))  # User's saved highlight color
        palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))  # Selected text
        palette.setColor(QPalette.Light, QColor(58, 106, 154))  # Light shade
        palette.setColor(QPalette.Midlight, QColor(42, 80, 112))  # Mid-light shade
        palette.setColor(QPalette.Mid, QColor(58, 106, 154))  # Mid shade (for headers)
        palette.setColor(QPalette.Dark, QColor(18, 36, 56))  # Dark shade
        palette.setColor(QPalette.Shadow, QColor(12, 24, 40))  # Shadow
        return palette
    
    def get_teal_professional_palette(self):
        """Create a Light Cyan palette based on Fusion Light with custom colors"""
        from PyQt5.QtGui import QPalette, QColor
        from PyQt5.QtWidgets import QApplication

        # Start with standard Fusion Light palette
        app = QApplication.instance()
        app.setStyle("Fusion")
        palette = app.style().standardPalette()

        # Override button and column header colors with custom teal-cyan
        palette.setColor(QPalette.Button, QColor(224, 246, 247))  # Light cyan for buttons (#E0F6F7)
        palette.setColor(QPalette.ButtonText, QColor(33, 33, 33))  # Dark text on buttons
        palette.setColor(QPalette.Window, QColor(224, 246, 247))  # Light cyan background (#E0F6F7)
        palette.setColor(QPalette.Base, QColor(239, 249, 249))  # Result preview background (#EFF9F9)
        palette.setColor(QPalette.Mid, QColor(206, 243, 245))  # Light cyan for column headers (#CEF3F5)

        return palette

    def get_macos_palette(self):
        """Create a macOS-inspired light palette with Apple's signature aesthetic"""
        from PyQt5.QtGui import QPalette, QColor

        # Get user's saved highlight color for this theme (Apple blue default)
        highlight_color = get_theme_color('preview_highlight_color', '#007AFF', 'macOS')

        palette = QPalette()
        # macOS light mode colors - clean, minimal, professional
        palette.setColor(QPalette.Window, QColor(246, 246, 246))  # Light gray background
        palette.setColor(QPalette.WindowText, QColor(29, 29, 31))  # Near-black text
        palette.setColor(QPalette.Base, QColor(255, 255, 255))  # Pure white for inputs
        palette.setColor(QPalette.AlternateBase, QColor(244, 244, 244))  # Subtle alternating
        palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 255))  # White tooltips
        palette.setColor(QPalette.ToolTipText, QColor(29, 29, 31))  # Dark tooltip text
        palette.setColor(QPalette.Text, QColor(29, 29, 31))  # Primary text
        palette.setColor(QPalette.Button, QColor(255, 255, 255))  # White buttons
        palette.setColor(QPalette.ButtonText, QColor(29, 29, 31))  # Dark button text
        palette.setColor(QPalette.BrightText, QColor(255, 59, 48))  # Apple red for alerts
        palette.setColor(QPalette.Link, QColor(0, 122, 255))  # Apple blue links
        palette.setColor(QPalette.Highlight, QColor(highlight_color))  # Apple blue selection
        palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))  # White on selection
        palette.setColor(QPalette.Light, QColor(255, 255, 255))  # White
        palette.setColor(QPalette.Midlight, QColor(235, 235, 235))  # Light gray
        palette.setColor(QPalette.Mid, QColor(209, 209, 214))  # macOS separator gray
        palette.setColor(QPalette.Dark, QColor(199, 199, 204))  # Darker gray
        palette.setColor(QPalette.Shadow, QColor(174, 174, 178))  # Shadow gray
        return palette

    def _detect_current_theme(self) -> str:
        """Detect which theme is currently active based on palette colors."""
        from PyQt5.QtGui import QPalette
        from PyQt5.QtWidgets import QApplication

        palette = QApplication.palette()
        window_color = palette.color(QPalette.Window)

        # Get RGB values
        r, g, b = window_color.red(), window_color.green(), window_color.blue()

        # Light Cyan: Window = (224, 246, 247) - very light, cyan tinted
        if r > 200 and g > 230 and b > 230:
            return "light_cyan"

        # Ocean: Window = (26, 48, 80) - dark blue, blue > red significantly
        if b > r + 30 and window_color.lightness() < 80:
            return "ocean"

        # Dark: Window = (45, 45, 45) - neutral gray
        return "dark"

    def _apply_templates_list_style(self):
        """Apply theme-specific styling to the templates table widget."""
        theme = self._detect_current_theme()

        if theme == "light_cyan":
            self.ocrmill_templates_table.setStyleSheet("""
                QTableWidget {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #ffffff, stop:1 #e8f4f6);
                    border: 1px solid #b8d4dc;
                    border-radius: 8px;
                    gridline-color: #d0e8ec;
                    font-size: 13px;
                }
                QTableWidget::item {
                    padding: 8px;
                    background-color: #f0f9fa;
                }
                QTableWidget::item:alternate {
                    background-color: #e0f2f4;
                }
                QTableWidget::item:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #00b4d8, stop:1 #0096c7);
                    color: #ffffff;
                }
                QTableWidget::item:hover:!selected {
                    background-color: #c8e8f0;
                }
                QHeaderView::section {
                    background-color: #d0e8ec;
                    padding: 6px;
                    border: 1px solid #b8d4dc;
                    font-weight: bold;
                }
            """)
        elif theme == "ocean":
            self.ocrmill_templates_table.setStyleSheet("""
                QTableWidget {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1a3652, stop:1 #142840);
                    border: 1px solid #2a5070;
                    border-radius: 8px;
                    gridline-color: #285070;
                    font-size: 13px;
                }
                QTableWidget::item {
                    padding: 8px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1e4060, stop:1 #183450);
                }
                QTableWidget::item:alternate {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #1a3856, stop:1 #152e46);
                }
                QTableWidget::item:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #00b4d8, stop:1 #0090b0);
                    color: #ffffff;
                }
                QTableWidget::item:hover:!selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #245070, stop:1 #1e4260);
                }
                QHeaderView::section {
                    background-color: #1e4060;
                    padding: 6px;
                    border: 1px solid #285070;
                    font-weight: bold;
                    color: #e0e0e0;
                }
            """)
        else:  # Dark theme (neutral grays)
            self.ocrmill_templates_table.setStyleSheet("""
                QTableWidget {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #383838, stop:1 #2a2a2a);
                    border: 1px solid #505050;
                    border-radius: 8px;
                    gridline-color: #505050;
                    font-size: 13px;
                }
                QTableWidget::item {
                    padding: 8px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #454545, stop:1 #3a3a3a);
                }
                QTableWidget::item:alternate {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #404040, stop:1 #353535);
                }
                QTableWidget::item:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #0078d4, stop:1 #0060b0);
                    color: #ffffff;
                }
                QTableWidget::item:hover:!selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #505050, stop:1 #454545);
                }
                QHeaderView::section {
                    background-color: #454545;
                    padding: 6px;
                    border: 1px solid #505050;
                    font-weight: bold;
                    color: #e0e0e0;
                }
            """)

    def get_button_style(self, button_type="default"):
        """
        Generate theme-aware button styles using current palette colors.

        Args:
            button_type: "default", "primary", "danger", "info", "warning", "success"

        Returns:
            CSS stylesheet string that adapts to current theme
        """
        from PyQt5.QtGui import QPalette, QColor
        from PyQt5.QtWidgets import QApplication

        palette = QApplication.palette()

        # Get base colors from theme
        base_bg = palette.color(QPalette.Button)
        base_text = palette.color(QPalette.ButtonText)
        highlight = palette.color(QPalette.Highlight)

        # All buttons use logo blue color across all themes
        bg = QColor(52, 152, 219)  # Logo Blue
        hover_bg = QColor(41, 128, 185)  # Darker Logo Blue
        disabled_bg = QColor(160, 160, 160)  # Grey

        # Text color - white for dark buttons, black for light buttons
        text_color = QColor(255, 255, 255) if bg.lightness() < 128 else QColor(0, 0, 0)
        
        return f"""
            QPushButton {{
                background-color: rgb({bg.red()}, {bg.green()}, {bg.blue()});
                color: rgb({text_color.red()}, {text_color.green()}, {text_color.blue()});
                font-weight: normal;
            }}
            QPushButton:hover {{
                background-color: rgb({hover_bg.red()}, {hover_bg.green()}, {hover_bg.blue()});
            }}
            QPushButton:disabled {{
                background-color: rgb({disabled_bg.red()}, {disabled_bg.green()}, {disabled_bg.blue()});
            }}
        """

    def get_input_style(self):
        """
        Generate theme-aware text input (QLineEdit) styles.

        Returns:
            CSS stylesheet string that adapts background color to current theme
        """
        # Check current theme
        current_theme = getattr(self, 'current_theme', 'Fusion (Light)')

        if current_theme == "Ocean":
            # Ocean theme: deep blue background with light text
            return "QLineEdit { color: #e6f5ff; background-color: #1c3856; padding: 5px; border: 1px solid #2a5a8a; }"
        elif current_theme == "Fusion (Dark)":
            # Fusion Dark theme: dark gray background with light text
            return "QLineEdit { color: #e0e0e0; background-color: #2d2d2d; padding: 5px; border: 1px solid #555; }"
        else:
            # Light themes: light background with dark text
            return "QLineEdit { color: #000000; background-color: #f5f5f5; padding: 5px; border: 1px solid #ccc; }"

    def clear_all(self):
        self.current_csv = None
        self.file_label.setText("No file selected")
        self.ci_input.clear()
        self.wt_input.clear()
        if hasattr(self, 'file_number_input') and self.file_number_input:
            self.file_number_input.clear()
        if hasattr(self, 'customer_ref_input') and self.customer_ref_input:
            self.customer_ref_input.clear()
        self.mid_combo.setCurrentIndex(-1)
        self.selected_mid = ""
        self.table.setRowCount(0)
        self.process_btn.setEnabled(False)
        self.process_btn.setText("Process Invoice")  # Reset button text
        self.reprocess_btn.setEnabled(False)  # Disable reprocess button
        self.progress.setVisible(False)
        self.invoice_check_label.setText("No file loaded")
        self.csv_total_value = 0.0
        self.edit_values_btn.setVisible(False)
        self.bottom_status.setText("Cleared")
        self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0;")
        self.last_processed_df = None  # Clear cached processed data
        logger.info("Process tab cleared")

    def browse_file(self):
        # Simple profile check without focus manipulation
        current_profile = self.profile_combo.currentText()
        if not current_profile or current_profile == "-- Select Profile --":
            QMessageBox.warning(
                self,
                "Mapping Profile Required",
                "<b>You must select a mapping profile before loading a shipment file.</b><br><br>"
                "Please choose one from the <b>Saved Profiles</b> dropdown.",
                QMessageBox.Ok
            )
            # Re-enable input fields after modal dialog
            self._enable_input_fields()
            return
        # -------------------------------------------------------------------------

        path, _ = QFileDialog.getOpenFileName(
            self, "Select Shipment File", str(INPUT_DIR), "CSV/Excel (*.csv *.xlsx)"
        )
        if not path:
            return

        self.current_csv = path
        self.file_label.setText(Path(path).name)

        # Clear previous processing state when loading new file
        self.last_processed_df = None
        self.table.setRowCount(0)

        # Clear file number for new invoice
        if hasattr(self, 'file_number_input') and self.file_number_input:
            self.file_number_input.clear()

        try:
            # Get header row value from input field
            header_row = 0  # Default: first row is header
            if hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
                try:
                    header_row_value = int(self.header_row_input.text().strip())
                    # Convert from 1-based to 0-based indexing
                    # If user enters 1, header is at row 0
                    # If user enters 2, skip 1 row, header is at row 1
                    header_row = max(0, header_row_value - 1)
                except ValueError:
                    header_row = 0

            col_map = {v: k for k, v in self.shipment_mapping.items()}
            if Path(path).suffix.lower() == ".xlsx":
                df = pd.read_excel(path, dtype=str, header=header_row)
            else:
                df = pd.read_csv(path, dtype=str, header=header_row)

            # Calculate total before renaming (using original column names)
            # Only sum rows that have a part number to exclude total/subtotal rows
            value_column = None
            part_number_column = None

            if 'value_usd' in self.shipment_mapping:
                # Get the original column name mapped to value_usd
                original_col_name = self.shipment_mapping['value_usd']
                logger.info(f"[INVOICE TOTAL] Looking for value column: '{original_col_name}' in columns: {df.columns.tolist()}")
                if original_col_name in df.columns:
                    value_column = original_col_name
                    logger.info(f"[INVOICE TOTAL] Found value column: '{value_column}'")
                else:
                    logger.warning(f"[INVOICE TOTAL] Value column '{original_col_name}' not found in file columns: {df.columns.tolist()}")
            else:
                logger.warning(f"[INVOICE TOTAL] 'value_usd' not mapped in shipment_mapping: {self.shipment_mapping}")

            # Get part number column to filter rows
            if 'part_number' in self.shipment_mapping:
                part_number_col_name = self.shipment_mapping['part_number']
                if part_number_col_name in df.columns:
                    part_number_column = part_number_col_name
                else:
                    logger.warning(f"Part number column '{part_number_col_name}' not found in file columns: {df.columns.tolist()}")

            # If we found the value column, calculate total
            if value_column:
                # Filter to only rows that have a part number (exclude total/subtotal rows)
                if part_number_column:
                    df_filtered = df[df[part_number_column].notna() & (df[part_number_column].astype(str).str.strip() != '')]
                    logger.debug(f"Filtered {len(df)} rows to {len(df_filtered)} rows with part numbers")
                    total = pd.to_numeric(df_filtered[value_column], errors='coerce').sum()
                else:
                    # If no part number column, sum all rows (old behavior)
                    total = pd.to_numeric(df[value_column], errors='coerce').sum()

                self.csv_total_value = round(total, 2)
                logger.info(f"Calculated invoice total: ${self.csv_total_value:,.2f}")

            # Now rename columns for other uses
            df = df.rename(columns=col_map)

            # Update invoice check - this will display the total and control button state
            if value_column:
                self.update_invoice_check()
        except Exception as e:
            logger.error(f"browse_file value read failed: {e}")
            self.invoice_check_label.setText("Could not read value column")

        logger.info(f"Loaded: {Path(path).name}")

        # Force focus to ci_input and ensure keyboard input works
        QApplication.processEvents()  # Process any pending events first
        self.ci_input.setFocus(Qt.OtherFocusReason)
        self.ci_input.activateWindow()
        logger.info(f"Set focus to ci_input: hasFocus={self.ci_input.hasFocus()}")
        
    def update_invoice_check(self):
        # Guard against re-entrancy
        if getattr(self, '_updating_invoice_check', False):
            return
        self._updating_invoice_check = True

        try:
            if not self.current_csv:
                self.invoice_check_label.setWordWrap(False)  # Single line display
                self.invoice_check_label.setText("No file loaded")
                # Gold color in dark theme (text-shadow not supported in Qt)
                font_size = get_user_setting_int('font_size', 9)
                if hasattr(self, 'current_theme') and self.current_theme in ["Fusion (Dark)", "Ocean"]:
                    self.invoice_check_label.setStyleSheet(f"background:transparent; color: gold; font-weight:bold; font-size:{font_size}pt; padding:3px;")
                else:
                    self.invoice_check_label.setStyleSheet(f"background:transparent; color: #A4262C; font-weight:bold; font-size:{font_size}pt; padding:3px;")
                self.edit_values_btn.setVisible(False)
                return

            user_text = self.ci_input.text().replace(',', '').strip()
            try:
                user_val = float(user_text) if user_text else 0.0
            except:
                user_val = 0.0

            diff = abs(user_val - self.csv_total_value)
            threshold = 0.01

            # Update the invoice check label and Edit Values button
            font_size = get_user_setting_int('font_size', 9)
            if user_val == 0.0:
                self.invoice_check_label.setWordWrap(False)  # Single line display
                self.invoice_check_label.setText(f"Total: ${self.csv_total_value:,.2f}")
                self.invoice_check_label.setStyleSheet(f"background:#0078D4; color:white; font-weight:bold; font-size:{font_size}pt; padding:3px;")
                self.edit_values_btn.setVisible(False)
            elif diff <= threshold:
                self.invoice_check_label.setWordWrap(False)  # Single line display
                self.invoice_check_label.setText(f"✓ Match: ${self.csv_total_value:,.2f}")
                self.invoice_check_label.setStyleSheet(f"background:#107C10; color:white; font-weight:bold; font-size:{font_size}pt; padding:3px;")
                self.edit_values_btn.setVisible(False)
            else:
                # Values don't match - show comparison and Edit Values button (two lines)
                self.invoice_check_label.setWordWrap(True)  # Allow two lines for mismatch display
                self.invoice_check_label.setText(
                    f"Total: ${self.csv_total_value:,.2f}\n"
                    f"Difference: ${diff:,.2f}"
                )
                self.invoice_check_label.setStyleSheet(f"background:#ff9800; color:white; font-weight:bold; font-size:{font_size}pt; padding:3px;")
                # Show Edit Values button only if haven't processed yet
                if self.last_processed_df is None:
                    self.edit_values_btn.setVisible(True)
                else:
                    self.edit_values_btn.setVisible(False)
            
            # Button state control - require invoice check match before processing
            has_weight = bool(self.wt_input.text().strip())
            has_ci_value = bool(user_text)
            has_mid = bool(self.selected_mid)
            values_match = diff <= threshold

            # Changed from >= 2 to >= 1 to allow profiles with minimal column mappings
            if self.current_csv and len(self.shipment_mapping) >= 1:
                if self.last_processed_df is None:
                    # Haven't processed yet - require weight, CI value, AND invoice values must match
                    # MID can be selected later, so not required for initial processing
                    if has_weight and has_ci_value and values_match:
                        self.process_btn.setEnabled(True)
                        self.process_btn.setText("Process Invoice")
                    else:
                        self.process_btn.setEnabled(False)
                        self.process_btn.setText("Process Invoice")
                else:
                    # Already processed - button becomes Export, only enabled when values match
                    if values_match:
                        self.process_btn.setEnabled(True)
                        self.process_btn.setText("Export Worksheet")
                    else:
                        self.process_btn.setEnabled(False)
                        self.process_btn.setText("Export Worksheet")
                    # Enable reprocess button when data has been processed
                    self.reprocess_btn.setEnabled(True)

            # Always ensure input fields stay enabled
            self._enable_input_fields()
        finally:
            self._updating_invoice_check = False
    def select_input_folder(self, display_widget=None):
        global INPUT_DIR, PROCESSED_DIR
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder", str(INPUT_DIR))
        if folder:
            INPUT_DIR = Path(folder)
            PROCESSED_DIR = get_processed_dir(INPUT_DIR)
            PROCESSED_DIR.mkdir(exist_ok=True)
            if display_widget:
                if isinstance(display_widget, QPlainTextEdit):
                    display_widget.setPlainText(str(INPUT_DIR))
                else:
                    display_widget.setText(str(INPUT_DIR))
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config VALUES ('input_dir', ?)", (str(INPUT_DIR),))
            conn.commit()
            conn.close()
            self.status.setText(f"Input folder: {INPUT_DIR}")
            self.refresh_input_files()

    def select_output_folder(self, display_widget=None):
        global OUTPUT_DIR
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", str(OUTPUT_DIR))
        if folder:
            OUTPUT_DIR = Path(folder)
            OUTPUT_DIR.mkdir(exist_ok=True)
            if display_widget:
                if isinstance(display_widget, QPlainTextEdit):
                    display_widget.setPlainText(str(OUTPUT_DIR))
                else:
                    display_widget.setText(str(OUTPUT_DIR))
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config VALUES ('output_dir', ?)", (str(OUTPUT_DIR),))
            conn.commit()
            conn.close()
            self.status.setText(f"Output folder: {OUTPUT_DIR}")
            self.refresh_exported_files()

    def move_pdf_to_processed(self, pdf_path):
        """
        Move a processed PDF file to the Processed PDF folder.

        Args:
            pdf_path (str or Path): Path to the PDF file to move

        Returns:
            bool: True if move was successful, False otherwise
        """
        try:
            pdf_file = Path(pdf_path)
            if not pdf_file.exists():
                logger.warning(f"PDF file not found for processing: {pdf_path}")
                return False

            # Get destination path
            dest_path = PROCESSED_DIR / pdf_file.name

            # Handle duplicate filenames
            if dest_path.exists():
                base_name = pdf_file.stem
                ext = pdf_file.suffix
                counter = 1
                while dest_path.exists():
                    dest_path = PROCESSED_DIR / f"{base_name}_{counter}{ext}"
                    counter += 1

            # Move the file
            shutil.move(str(pdf_file), str(dest_path))
            logger.info(f"Moved processed PDF to: {dest_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to move PDF to processed folder: {e}")
            return False

    def load_file_as_dataframe(self, file_path):
        """Load CSV or Excel file and return as DataFrame"""
        # Get header row value from profile or input field
        header_row = 0  # Default: first row is header
        # First check if there's a profile header row loaded
        if hasattr(self, 'profile_header_row') and self.profile_header_row:
            header_row = max(0, self.profile_header_row - 1)
        # Otherwise check input field (for Invoice Mapping Profiles tab)
        elif hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
            try:
                header_row_value = int(self.header_row_input.text().strip())
                # Convert from 1-based to 0-based indexing
                # If user enters 1, header is at row 0
                # If user enters 2, skip 1 row, header is at row 1
                header_row = max(0, header_row_value - 1)
            except ValueError:
                header_row = 0

        logger.info(f"[LOAD DATAFRAME] Loading {file_path} with header_row={header_row}")
        file_path_str = str(file_path)
        if file_path_str.lower().endswith('.xlsx') or file_path_str.lower().endswith('.xls'):
            return pd.read_excel(file_path_str, dtype=str, keep_default_na=False, header=header_row).fillna("")
        else:
            return pd.read_csv(file_path_str, dtype=str, keep_default_na=False, header=header_row).fillna("")

    def start_processing_with_editable_preview(self):
        if not self.current_csv:
            return
        # Process with user's entered CI value, generate preview for editing
        # Do NOT change the CI input - keep user's entered value
        self.process_btn.setText("Export Worksheet")
        self.process_btn.setEnabled(False)
        self.start_processing()

    def start_processing(self, part_number_overrides=None):
        """
        Start processing the invoice file.

        Args:
            part_number_overrides: Optional dict mapping row index to new part number.
                                   Used during reprocess to apply user edits to part numbers.
        """
        if not self.current_csv:
            QMessageBox.critical(self, "Error", "Please select a file")
            return
        if len(self.shipment_mapping) < 2:
            QMessageBox.critical(self, "Error", "Map Part Number and Value USD")
            return
        
        # Verify Net Weight is entered
        wt_text = self.wt_input.text().strip()
        if not wt_text:
            self.wt_input.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.wt_input.setStyleSheet(""))
            self.wt_input.setFocus()
            QMessageBox.warning(self, "Net Weight Required", "Please enter the Net Weight (kg) before processing.")
            return
        
        try:
            wt_val = float(wt_text.replace(',', ''))
            if wt_val <= 0:
                raise ValueError()
        except:
            self.wt_input.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.wt_input.setStyleSheet(""))
            self.wt_input.setFocus()
            QMessageBox.warning(self, "Invalid Net Weight", "Please enter a valid Net Weight greater than zero.")
            return
        
        # Verify MID is selected
        if not self.selected_mid:
            self.mid_combo.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.mid_combo.setStyleSheet(""))
            self.mid_combo.setFocus()
            QMessageBox.warning(self, "MID Required", "Please select a MID (Manufacturer ID) before processing.")
            return

        # Verify File Number is entered (mandatory for billing)
        file_number = self.file_number_input.text().strip() if hasattr(self, 'file_number_input') else ""
        if not file_number:
            self.file_number_input.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.file_number_input.setStyleSheet(""))
            self.file_number_input.setFocus()
            QMessageBox.warning(self, "File Number Required", "Please enter a File Number before processing.\n\nThis is required for billing purposes.")
            return

        self.process_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.status.setText("Processing...")

        # Directly process the CSV/Excel file synchronously (single-threaded)
        try:
            self.status.setText("Loading file...")
            df = self.load_file_as_dataframe(self.current_csv)
            vr = Path(self.current_csv).stem
            col_map = {v:k for k,v in self.shipment_mapping.items()}
            # Before renaming, drop columns that would create duplicates
            # (e.g., if mapping sigma_part_number -> part_number but part_number already exists)
            for source_col, target_col in col_map.items():
                if source_col in df.columns and target_col in df.columns and source_col != target_col:
                    logger.info(f"[PROCESS] Dropping original '{target_col}' column to avoid duplicate after renaming '{source_col}' -> '{target_col}'")
                    df = df.drop(columns=[target_col])
            df = df.rename(columns=col_map)
            if not {'part_number','value_usd'}.issubset(df.columns):
                self.status.setText("Missing Part Number or Value USD")
                return

            # Filter out rows without part numbers (excludes total/subtotal rows)
            initial_row_count = len(df)
            df = df[df['part_number'].notna() & (df['part_number'].astype(str).str.strip() != '')]
            filtered_row_count = len(df)
            if filtered_row_count < initial_row_count:
                logger.info(f"[PROCESS] Filtered {initial_row_count - filtered_row_count} rows without part numbers (total/subtotal rows)")
                logger.info(f"[PROCESS] Processing {filtered_row_count} data rows")
            def safe_float(text):
                if pd.isna(text) or text == "": return 0.0
                try:
                    return float(str(text).replace(',', '').strip())
                except:
                    return 0.0
            df['value_usd'] = pd.to_numeric(df['value_usd'], errors='coerce').fillna(0)
            csv_total = df['value_usd'].sum()
            user_ci = safe_float(self.ci_input.text())
            wt = safe_float(self.wt_input.text())
            if wt <= 0:
                self.status.setText("Net Weight must be greater than zero")
                return
            # Invoice diff
            self.handle_invoice_diff(csv_total, user_ci)
            conn = sqlite3.connect(str(DB_PATH))
            parts = pd.read_sql("SELECT part_number, hts_code, steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio, non_steel_ratio, qty_unit, country_of_melt, country_of_cast, country_of_smelt, Sec301_Exclusion_Tariff FROM parts_master", conn)
            conn.close()
            # Normalize part numbers for matching (strip whitespace, uppercase)
            df['part_number'] = df['part_number'].astype(str).str.strip().str.upper()
            parts['part_number'] = parts['part_number'].astype(str).str.strip().str.upper()

            # Apply part number overrides from reprocess (user-edited part numbers in preview)
            if part_number_overrides:
                for row_idx, new_part_number in part_number_overrides.items():
                    if row_idx < len(df):
                        old_part = df.iloc[row_idx]['part_number']
                        df.iloc[row_idx, df.columns.get_loc('part_number')] = new_part_number.strip().upper()
                        logger.info(f"Applied part number override: row {row_idx}: '{old_part}' -> '{new_part_number}'")

            df = df.merge(parts, on='part_number', how='left', suffixes=('', '_master'), indicator=True)
            # Track parts not found in the database
            df['_not_in_db'] = df['_merge'] == 'left_only'
            df = df.drop(columns=['_merge'])

            # MSI-to-Sigma fallback: For parts not found by Sigma part number,
            # try looking up by the original MSI part number (if available)
            # This handles cases where parts_master has MSI parts but we want Sigma output
            if 'msi_part_number' in df.columns:
                not_found_mask = df['_not_in_db'] == True
                if not_found_mask.any():
                    # Get rows where Sigma lookup failed
                    for idx in df[not_found_mask].index:
                        msi_part = str(df.loc[idx, 'msi_part_number']).strip().upper()
                        sigma_part = str(df.loc[idx, 'part_number']).strip().upper()
                        if msi_part and msi_part != sigma_part:
                            # Try to find the MSI part in parts_master
                            match = parts[parts['part_number'] == msi_part]
                            if not match.empty:
                                # Found by MSI - merge the data but keep Sigma as the display part number
                                master_row = match.iloc[0]
                                for col in ['hts_code', 'steel_ratio', 'aluminum_ratio', 'copper_ratio',
                                           'wood_ratio', 'auto_ratio', 'non_steel_ratio', 'qty_unit',
                                           'country_of_melt', 'country_of_cast', 'country_of_smelt',
                                           'Sec301_Exclusion_Tariff']:
                                    master_col = f'{col}_master'
                                    if master_col in df.columns:
                                        df.loc[idx, master_col] = master_row.get(col, df.loc[idx, master_col])
                                    elif col in parts.columns:
                                        df.loc[idx, col] = master_row.get(col, df.loc[idx, col])
                                df.loc[idx, '_not_in_db'] = False
                                logger.info(f"MSI fallback: Found '{msi_part}' in parts_master, using Sigma '{sigma_part}' for output")

            # Merge strategy: Prefer database (master) values over invoice values
            # Database values ALWAYS take precedence; invoice values are only used as fallback when DB is empty
            merge_fields = ['hts_code', 'steel_ratio', 'aluminum_ratio', 'copper_ratio', 'wood_ratio', 'auto_ratio', 'non_steel_ratio', 'qty_unit']
            for field in merge_fields:
                master_col = f'{field}_master'
                if master_col in df.columns:
                    # Database has this field - database value ALWAYS takes precedence
                    # Only fall back to invoice value if database value is empty/NA
                    if field in ['steel_ratio', 'aluminum_ratio', 'copper_ratio', 'wood_ratio', 'auto_ratio', 'non_steel_ratio']:
                        master_vals = pd.to_numeric(df[master_col], errors='coerce')
                        invoice_vals = pd.to_numeric(df[field], errors='coerce') if field in df.columns else pd.Series([pd.NA] * len(df))
                        # Use master value if available and not NaN, otherwise invoice value
                        df[field] = master_vals.combine_first(invoice_vals)
                    else:
                        # For text fields like hts_code: database value takes precedence
                        master_series = df[master_col].replace('', pd.NA)
                        invoice_series = df[field].replace('', pd.NA) if field in df.columns else pd.Series([pd.NA] * len(df))
                        # combine_first: use master, fill gaps with invoice
                        df[field] = master_series.combine_first(invoice_series)
                elif field not in df.columns:
                    # Neither invoice nor database has it - set default
                    if field in ['steel_ratio', 'aluminum_ratio', 'copper_ratio', 'wood_ratio', 'auto_ratio', 'non_steel_ratio']:
                        df[field] = 0.0
                    else:
                        df[field] = ''

            # Convert ratio fields to numeric (values are percentages 0-100)
            # Note: fillna(0.0) for all ratios - the later processing will determine
            # material type from HTS code if no ratios are set
            df['steel_ratio'] = pd.to_numeric(df['steel_ratio'], errors='coerce').fillna(0.0)
            df['aluminum_ratio'] = pd.to_numeric(df['aluminum_ratio'], errors='coerce').fillna(0.0)
            df['copper_ratio'] = pd.to_numeric(df['copper_ratio'], errors='coerce').fillna(0.0)
            df['wood_ratio'] = pd.to_numeric(df['wood_ratio'], errors='coerce').fillna(0.0)
            df['auto_ratio'] = pd.to_numeric(df['auto_ratio'], errors='coerce').fillna(0.0)
            df['non_steel_ratio'] = pd.to_numeric(df['non_steel_ratio'], errors='coerce').fillna(0.0)
            missing = df[
                (df['hts_code'].isnull() | (df['hts_code'] == '')) |
                (df['value_usd'] == 0) |
                (df['steel_ratio'].isnull())
            ].copy()
            if not missing.empty:
                missing = missing[['part_number', 'hts_code', 'value_usd', 'steel_ratio']].copy()
                missing.columns = ['Part Number', 'HTS Code', 'Value USD', 'Sec 232 %']
                missing = missing.fillna('')
                self.log_missing_data_warning(missing)
            self._process_with_complete_data(df, vr, user_ci, wt)
        except Exception as e:
            logger.error(f"Processing failed: {e}")
            self.status.setText(f"Processing failed: {str(e)}")

    def _process_with_complete_data(self, df, vr, user_ci, wt):
        """
        Process the DataFrame with complete data, calculate required fields, and update the preview table.
        Handles multi-content items (steel, aluminum, copper, wood, non-232).
        """
        df = df.copy()

        # Steel/Aluminum/Copper/Wood/Auto/NonSteel ratios BEFORE calculating weight
        df['SteelRatio'] = pd.to_numeric(df.get('steel_ratio', 0.0), errors='coerce').fillna(0.0)
        df['AluminumRatio'] = pd.to_numeric(df.get('aluminum_ratio', 0.0), errors='coerce').fillna(0.0)
        df['CopperRatio'] = pd.to_numeric(df.get('copper_ratio', 0.0), errors='coerce').fillna(0.0)
        df['WoodRatio'] = pd.to_numeric(df.get('wood_ratio', 0.0), errors='coerce').fillna(0.0)
        df['AutoRatio'] = pd.to_numeric(df.get('auto_ratio', 0.0), errors='coerce').fillna(0.0)
        df['NonSteelRatio'] = pd.to_numeric(df.get('non_steel_ratio', 0.0), errors='coerce').fillna(0.0)

        # Split rows by steel/aluminum/copper/wood/non-232 content BEFORE calculating CalcWtNet
        # Note: Ratios are stored as percentages (0-100) in the database
        original_row_count = len(df)
        expanded_rows = []
        for _, row in df.iterrows():
            steel_pct = row['SteelRatio']
            aluminum_pct = row['AluminumRatio']
            copper_pct = row['CopperRatio']
            wood_pct = row['WoodRatio']
            auto_pct = row['AutoRatio']
            non_steel_pct = row['NonSteelRatio']
            original_value = row['value_usd']

            # If no percentages are set, use HTS lookup to determine material type
            # BUT only for parts that ARE in the database - unfound parts should NOT get defaults
            not_in_db = row.get('_not_in_db', False)

            # Check for incomplete data: part in DB but missing required fields (HTS or qty_unit)
            hts = row.get('hts_code', '')
            hts_clean = str(hts).strip() if pd.notna(hts) else ''
            qty_unit = row.get('qty_unit', '')
            qty_unit_clean = str(qty_unit).strip() if pd.notna(qty_unit) else ''

            # Mark as incomplete if in database but missing HTS or qty_unit
            if not not_in_db and (not hts_clean or not qty_unit_clean):
                row['_incomplete_data'] = True

            if steel_pct == 0 and aluminum_pct == 0 and copper_pct == 0 and wood_pct == 0 and auto_pct == 0 and non_steel_pct == 0:
                if not_in_db:
                    # Part not in database - leave all ratios at 0 to show "Not Found" status
                    pass
                elif hts_clean:
                    # Part is in database with HTS code - look up material type
                    material, _, _ = get_232_info(hts)
                    if material == 'Aluminum':
                        aluminum_pct = 100.0
                    elif material == 'Copper':
                        copper_pct = 100.0
                    elif material == 'Wood':
                        wood_pct = 100.0
                    elif material == 'Auto':
                        auto_pct = 100.0
                    elif material == 'Steel':
                        steel_pct = 100.0
                    # If material is None/unknown, _incomplete_data is already set above

            # Validate that percentages sum to 100% - recalculate non_steel_pct if needed
            # This fixes database entries where non_steel_ratio was incorrectly set to 100%
            total_232_pct = steel_pct + aluminum_pct + copper_pct + wood_pct + auto_pct
            if total_232_pct > 0:
                # If we have any 232 materials, non_steel should be the remainder
                non_steel_pct = max(0.0, 100.0 - total_232_pct)

            # Final validation: ensure total is 100%
            total_pct = steel_pct + aluminum_pct + copper_pct + wood_pct + auto_pct + non_steel_pct
            if total_pct > 100.01:  # Allow small floating point tolerance
                # Normalize all percentages to sum to 100%
                scale = 100.0 / total_pct
                steel_pct *= scale
                aluminum_pct *= scale
                copper_pct *= scale
                wood_pct *= scale
                auto_pct *= scale
                non_steel_pct *= scale

            # Create derivative rows in order: Non-232 first, then Steel, Aluminum, Copper, Wood, Auto
            # This ensures 232 materials appear BELOW their non-232 counterparts in the preview
            # Track allocated value to ensure total equals original (avoid rounding errors)
            allocated_value = 0.0
            row_start_idx = len(expanded_rows)

            # If part is marked as incomplete, create incomplete row and skip normal row creation
            if row.get('_incomplete_data', False):
                incomplete_row = row.copy()
                incomplete_row['value_usd'] = original_value
                incomplete_row['SteelRatio'] = 0.0
                incomplete_row['AluminumRatio'] = 0.0
                incomplete_row['CopperRatio'] = 0.0
                incomplete_row['WoodRatio'] = 0.0
                incomplete_row['AutoRatio'] = 0.0
                incomplete_row['NonSteelRatio'] = 0.0
                incomplete_row['_content_type'] = 'incomplete'
                expanded_rows.append(incomplete_row)
                continue  # Skip normal row creation for incomplete parts

            # Create non-232 portion row first (if non_steel_pct > 0)
            if non_steel_pct > 0:
                non_232_row = row.copy()
                portion_value = round(original_value * non_steel_pct / 100.0, 2)
                non_232_row['value_usd'] = portion_value
                allocated_value += portion_value
                non_232_row['SteelRatio'] = 0.0
                non_232_row['AluminumRatio'] = 0.0
                non_232_row['CopperRatio'] = 0.0
                non_232_row['WoodRatio'] = 0.0
                non_232_row['AutoRatio'] = 0.0
                non_232_row['NonSteelRatio'] = non_steel_pct
                non_232_row['_content_type'] = 'non_232'
                expanded_rows.append(non_232_row)

            # Create steel portion row (if steel_pct > 0)
            if steel_pct > 0:
                steel_row = row.copy()
                portion_value = round(original_value * steel_pct / 100.0, 2)
                steel_row['value_usd'] = portion_value
                allocated_value += portion_value
                steel_row['SteelRatio'] = steel_pct
                steel_row['AluminumRatio'] = 0.0
                steel_row['CopperRatio'] = 0.0
                steel_row['WoodRatio'] = 0.0
                steel_row['AutoRatio'] = 0.0
                steel_row['NonSteelRatio'] = 0.0
                steel_row['_content_type'] = 'steel'
                expanded_rows.append(steel_row)

            # Create aluminum portion row (if aluminum_pct > 0)
            if aluminum_pct > 0:
                aluminum_row = row.copy()
                portion_value = round(original_value * aluminum_pct / 100.0, 2)
                aluminum_row['value_usd'] = portion_value
                allocated_value += portion_value
                aluminum_row['SteelRatio'] = 0.0
                aluminum_row['AluminumRatio'] = aluminum_pct
                aluminum_row['CopperRatio'] = 0.0
                aluminum_row['WoodRatio'] = 0.0
                aluminum_row['AutoRatio'] = 0.0
                aluminum_row['NonSteelRatio'] = 0.0
                aluminum_row['_content_type'] = 'aluminum'
                expanded_rows.append(aluminum_row)

            # Create copper portion row (if copper_pct > 0)
            if copper_pct > 0:
                copper_row = row.copy()
                portion_value = round(original_value * copper_pct / 100.0, 2)
                copper_row['value_usd'] = portion_value
                allocated_value += portion_value
                copper_row['SteelRatio'] = 0.0
                copper_row['AluminumRatio'] = 0.0
                copper_row['CopperRatio'] = copper_pct
                copper_row['WoodRatio'] = 0.0
                copper_row['AutoRatio'] = 0.0
                copper_row['NonSteelRatio'] = 0.0
                copper_row['_content_type'] = 'copper'
                expanded_rows.append(copper_row)

            # Create wood portion row (if wood_pct > 0)
            if wood_pct > 0:
                wood_row = row.copy()
                portion_value = round(original_value * wood_pct / 100.0, 2)
                wood_row['value_usd'] = portion_value
                allocated_value += portion_value
                wood_row['SteelRatio'] = 0.0
                wood_row['AluminumRatio'] = 0.0
                wood_row['CopperRatio'] = 0.0
                wood_row['WoodRatio'] = wood_pct
                wood_row['AutoRatio'] = 0.0
                wood_row['NonSteelRatio'] = 0.0
                wood_row['_content_type'] = 'wood'
                expanded_rows.append(wood_row)

            # Create auto portion row (if auto_pct > 0)
            if auto_pct > 0:
                auto_row = row.copy()
                portion_value = round(original_value * auto_pct / 100.0, 2)
                auto_row['value_usd'] = portion_value
                allocated_value += portion_value
                auto_row['SteelRatio'] = 0.0
                auto_row['AluminumRatio'] = 0.0
                auto_row['CopperRatio'] = 0.0
                auto_row['WoodRatio'] = 0.0
                auto_row['AutoRatio'] = auto_pct
                auto_row['NonSteelRatio'] = 0.0
                auto_row['_content_type'] = 'auto'
                expanded_rows.append(auto_row)

            # Handle parts not in database (all percentages are 0) - create single row as "not_found"
            if len(expanded_rows) == row_start_idx and not_in_db:
                # No rows were created because all percentages are 0 - this is a "not found" part
                not_found_row = row.copy()
                not_found_row['value_usd'] = original_value
                not_found_row['SteelRatio'] = 0.0
                not_found_row['AluminumRatio'] = 0.0
                not_found_row['CopperRatio'] = 0.0
                not_found_row['WoodRatio'] = 0.0
                not_found_row['AutoRatio'] = 0.0
                not_found_row['NonSteelRatio'] = 0.0
                not_found_row['_content_type'] = 'not_found'
                expanded_rows.append(not_found_row)

            # Note: Incomplete parts are handled earlier with 'continue', so this block is no longer needed

            # Fix rounding errors: adjust the last created row to ensure total matches original
            if len(expanded_rows) > row_start_idx:
                remainder = round(original_value - allocated_value, 2)
                content_type = expanded_rows[-1].get('_content_type', '')
                if abs(remainder) > 0.001 and content_type not in ['not_found', 'incomplete']:
                    # Add remainder to the last row created for this item (skip for not_found/incomplete rows)
                    expanded_rows[-1]['value_usd'] = round(expanded_rows[-1]['value_usd'] + remainder, 2)

        # Rebuild dataframe from expanded rows
        df = pd.DataFrame(expanded_rows).reset_index(drop=True)
        logger.info(f"Row expansion: {original_row_count} → {len(expanded_rows)} rows (steel/aluminum/copper/wood/auto/non-232 split)")

        # Now calculate CalcWtNet based on expanded rows
        # If net_weight column exists and has values, use those directly
        # Otherwise, calculate proportionally based on value
        total_value = df['value_usd'].sum()
        logger.info(f"[CalcWtNet DEBUG] total_value={total_value}, wt={wt}, rows={len(df)}")
        logger.info(f"[CalcWtNet DEBUG] value_usd column sample: {df['value_usd'].head(5).tolist()}")

        # Check if net_weight column exists and has valid per-row weights
        # Skip if all values are the same (indicates total weight copied to all rows, not per-row weights)
        use_net_weight_column = False
        if 'net_weight' in df.columns:
            net_weights = df['net_weight'].dropna()
            if len(net_weights) > 0:
                # Convert to numeric and check if values vary
                numeric_weights = pd.to_numeric(net_weights.astype(str).str.replace(',', ''), errors='coerce').dropna()
                if len(numeric_weights) > 1:
                    unique_weights = numeric_weights.unique()
                    if len(unique_weights) > 1:
                        use_net_weight_column = True
                        logger.info(f"[CalcWtNet DEBUG] Using per-row net_weight column ({len(unique_weights)} unique values)")
                    else:
                        logger.info(f"[CalcWtNet DEBUG] Skipping net_weight column - all values are the same ({unique_weights[0]}), using proportional calculation")
                elif len(numeric_weights) == 1:
                    # Single row, use the value
                    use_net_weight_column = True

        if use_net_weight_column:
            # Use invoice-provided net_weight where available, fall back to calculated
            def get_weight(row):
                nw = row.get('net_weight', '')
                if pd.notna(nw) and str(nw).strip():
                    try:
                        return float(str(nw).replace(',', '').strip())
                    except (ValueError, TypeError):
                        pass
                # Fall back to proportional calculation
                if total_value > 0:
                    return (row['value_usd'] / total_value) * wt
                return 0.0
            df['CalcWtNet'] = df.apply(get_weight, axis=1)
        elif total_value == 0:
            df['CalcWtNet'] = 0.0
        else:
            df['CalcWtNet'] = (df['value_usd'] / total_value) * wt

        logger.info(f"[CalcWtNet DEBUG] CalcWtNet sample: {df['CalcWtNet'].head(5).tolist()}")

        # Calculate Qty1 and Qty2 based on qty_unit type from HTS database
        # Categories:
        # - Weight-only: KG, G, T -> Qty1 = CalcWtNet, Qty2 = empty
        # - Count-only: NO, PCS, DOZ, etc. -> Qty1 = quantity (pieces), Qty2 = empty
        # - Dual (count + weight): NO. AND KG, XX KG, XX G -> Qty1 = quantity, Qty2 = CalcWtNet
        # - Other units (volume, area, length): Use quantity if available, else empty

        # Weight-only units (Qty1 = weight in KG)
        WEIGHT_UNITS = {'KG', 'G', 'T', 'T ADW', 'T DWB'}

        # Count-only units (Qty1 = piece count)
        COUNT_UNITS = {'NO', 'PCS', 'DOZ', 'DOZ. PRS', 'DZ PCS', 'GROSS', 'HUNDREDS',
                       'THOUSANDS', 'PRS', 'PACK', 'DOSES', 'CARAT'}

        # Dual units: first quantity is count, second is weight (Qty1 = count, Qty2 = weight)
        # Includes NO. AND KG and metal+weight combinations
        DUAL_UNITS = {'NO. AND KG', 'NO/KG', 'NO\\KG', 'NO., KG', 'NO. KG', 'NO KG',
                      'CU KG', 'CY KG', 'NI KG', 'PB KG', 'ZN KG', 'KG AMC',
                      'AG G', 'AU G', 'IR G', 'OS G', 'PD G', 'PT G', 'RH G', 'RU G',
                      'DOZ., KG', 'DOZ. KG', 'DOZ KG', 'PRS., KG', 'PRS. KG', 'PRS KG'}

        # Volume/Area/Length units (use quantity from invoice)
        MEASURE_UNITS = {'LITERS', 'PF.LITERS', 'BBL', 'M', 'LIN. M', 'M2', 'CM2', 'M3',
                         'SQUARE', 'FIBER M', 'GBQ', 'MWH', 'THOUSAND M', 'THOUSAND M3'}

        # Units that should have BOTH Qty1 and Qty2 empty (measurement-only units per CBP requirements)
        NO_QTY_UNITS = {'M', 'M2', 'M3', 'DOZ', 'DPR', 'PRS', 'DOZ. PRS'}

        def get_qty1(row):
            qty_unit = str(row.get('qty_unit', '')).strip().upper() if pd.notna(row.get('qty_unit')) else ''
            if qty_unit == '':
                return ''

            # If qty_unit is in NO_QTY_UNITS, leave Qty1 empty
            if qty_unit in NO_QTY_UNITS:
                return ''

            # Weight-only units: Qty1 is net weight (minimum 1 KG)
            if qty_unit in WEIGHT_UNITS:
                wt = int(round(row['CalcWtNet']))
                return str(max(wt, 1))  # Minimum 1 KG

            # Count-only units: Qty1 is piece count from invoice
            if qty_unit in COUNT_UNITS:
                qty = row.get('quantity', '')
                if pd.notna(qty) and str(qty).strip():
                    try:
                        return str(int(float(str(qty).replace(',', '').strip())))
                    except (ValueError, TypeError):
                        return ''
                return ''

            # Dual units: Qty1 is piece count
            if qty_unit in DUAL_UNITS:
                qty = row.get('quantity', '')
                if pd.notna(qty) and str(qty).strip():
                    try:
                        qty_val = float(str(qty).replace(',', '').strip())
                        # Preserve 2 decimal places for dozen-based units (DOZ., KG etc.)
                        if 'DOZ' in qty_unit:
                            return f"{qty_val:.2f}"
                        # For NO-based units, if qty appears to be in dozens (< 100 and has decimals),
                        # it may need conversion. Check quantity_unit from invoice to determine.
                        qty_unit_invoice = str(row.get('quantity_unit', '')).strip().upper()
                        if 'NO' in qty_unit and qty_unit_invoice == 'DOZ':
                            # Invoice reported in dozens, convert to pieces
                            qty_val = qty_val * 12
                        result = max(int(qty_val), 1)  # Minimum 1 piece
                        return str(result)
                    except (ValueError, TypeError):
                        return ''
                # If no quantity, return minimum of 1 for NO-based units
                if 'NO' in qty_unit:
                    return '1'
                return ''

            # Measure units: Use quantity from invoice if available
            if qty_unit in MEASURE_UNITS:
                qty = row.get('quantity', '')
                if pd.notna(qty) and str(qty).strip():
                    try:
                        return str(int(float(str(qty).replace(',', '').strip())))
                    except (ValueError, TypeError):
                        return ''
                return ''

            # Unknown unit type - try quantity first, fall back to empty
            qty = row.get('quantity', '')
            if pd.notna(qty) and str(qty).strip():
                try:
                    return str(int(float(str(qty).replace(',', '').strip())))
                except (ValueError, TypeError):
                    return ''
            return ''

        def get_qty2(row):
            qty_unit = str(row.get('qty_unit', '')).strip().upper() if pd.notna(row.get('qty_unit')) else ''

            # If qty_unit is in NO_QTY_UNITS, leave Qty2 empty
            if qty_unit in NO_QTY_UNITS:
                return ''

            # IMPORTANT: For single units (weight-only or count-only), Qty2 should always be empty
            # Only dual units (like NO/KG) should have both Qty1 and Qty2
            # Single weight units: KG, G, T - only Qty1 (weight)
            # Single count units: NO, PCS, etc. - only Qty1 (count)
            if qty_unit in WEIGHT_UNITS or qty_unit in COUNT_UNITS:
                return ''

            # Dual units: Qty2 is net weight (minimum 1 KG)
            if qty_unit in DUAL_UNITS:
                calc_wt = row.get('CalcWtNet', 0)
                if pd.isna(calc_wt):
                    calc_wt = 0
                wt = int(round(calc_wt))
                return str(max(wt, 1))  # Minimum 1 KG

            # All other cases: Qty2 is empty
            return ''

        df['Qty1'] = df.apply(get_qty1, axis=1)
        df['Qty2'] = df.apply(get_qty2, axis=1)

        # Keep cbp_qty for backward compatibility (uses Qty1 logic)
        df['cbp_qty'] = df['Qty1']

        # Set HTSCode and MID (convert NaN to empty string)
        df['HTSCode'] = df['hts_code'].fillna('').astype(str).replace('nan', '')
        mid = self.selected_mid if hasattr(self, 'selected_mid') else ''
        df['MID'] = mid
        melt = str(mid)[:2] if mid else ''

        # Derivative fields (exact 8-digit match, flag logic)
        dec_type_list = []
        country_melt_list = []
        country_cast_list = []
        prim_country_smelt_list = []
        prim_smelt_flag_list = []
        flag_list = []
        for _, r in df.iterrows():
            content_type = r.get('_content_type', '')
            hts = r.get('hts_code', '')
            hts_clean = str(hts).replace('.', '').strip().upper()
            hts_8 = hts_clean[:8]
            hts_10 = hts_clean[:10]
            material, dec_type, smelt_flag = get_232_info(hts)

            # Set flag based on content type, but use consistent declaration code from HTS lookup
            # All derivative rows with the same HTS code get the same declaration code
            if content_type == 'not_found':
                flag = 'Not_Found'
            elif content_type == 'incomplete':
                flag = 'Incomplete'
            elif content_type == 'steel':
                flag = '232_Steel'
            elif content_type == 'aluminum':
                flag = '232_Aluminum'
            elif content_type == 'copper':
                flag = '232_Copper'
            elif content_type == 'wood':
                flag = '232_Wood'
            elif content_type == 'auto':
                flag = '232_Auto'
            elif content_type == 'non_232':
                flag = 'Non_232'
            else:
                # Fallback for backward compatibility
                flag = f"232_{material}" if material else ''

            # All rows with the same HTS code use the same declaration code from tariff_232 lookup
            dec_type_list.append(dec_type)
            
            # Use imported country codes if available, otherwise fall back to MID-based default
            country_of_melt = r.get('country_of_melt', '')
            country_of_cast = r.get('country_of_cast', '')
            country_of_smelt = r.get('country_of_smelt', '')

            # Check for NaN/null values and use default if empty
            melt_code = country_of_melt if pd.notna(country_of_melt) and str(country_of_melt).strip() else melt
            cast_code = country_of_cast if pd.notna(country_of_cast) and str(country_of_cast).strip() else melt
            smelt_code = country_of_smelt if pd.notna(country_of_smelt) and str(country_of_smelt).strip() else melt

            country_melt_list.append(melt_code)
            country_cast_list.append(cast_code)
            prim_country_smelt_list.append(smelt_code)
            prim_smelt_flag_list.append(smelt_flag)
            flag_list.append(flag)

        df['DecTypeCd'] = dec_type_list
        df['CountryofMelt'] = country_melt_list
        df['CountryOfCast'] = country_cast_list
        df['PrimCountryOfSmelt'] = prim_country_smelt_list
        df['DeclarationFlag'] = prim_smelt_flag_list
        df['_232_flag'] = flag_list

        # TODO: Lacey Act Detection - To be implemented at a later date
        # =====================================================================
        # LACEY ACT DETECTION
        # Check if HTS codes fall under Lacey Act requirements (Chapters 44, 47, 48, 94)
        # =====================================================================
        # lacey_flag_list = []
        # lacey_species_list = []
        # lacey_harvest_country_list = []
        # lacey_recycled_list = []
        #
        # for _, r in df.iterrows():
        #     hts = str(r.get('hts_code', '')).replace('.', '').strip()
        #     part_no = r.get('part_number', '')
        #     wood_ratio = float(r.get('WoodRatio', 0) or 0)
        #
        #     # Check if Lacey Act applies based on HTS chapter
        #     lacey_required = False
        #     if hts:
        #         chapter = hts[:2]
        #         # Chapters subject to Lacey Act: 44 (Wood), 47 (Pulp), 48 (Paper)
        #         if chapter in ('44', '47', '48'):
        #             lacey_required = True
        #         # Chapter 94 furniture - check for wood furniture (9401, 9403)
        #         elif hts[:4] in ('9401', '9403'):
        #             lacey_required = True
        #
        #     # Also flag if wood_ratio > 0 (product contains wood content)
        #     if wood_ratio > 0:
        #         lacey_required = True
        #
        #     # Look up Lacey data from parts_master if available
        #     species_name = ''
        #     harvest_country = ''
        #     recycled_pct = 0.0
        #
        #     if part_no:
        #         try:
        #             conn = sqlite3.connect(str(DB_PATH))
        #             c = conn.cursor()
        #             c.execute("""SELECT species_scientific_name, species_common_name, country_of_harvest, percent_recycled
        #                          FROM parts_master WHERE part_number = ?""", (part_no,))
        #             row = c.fetchone()
        #             conn.close()
        #             if row:
        #                 species_name = row[0] or row[1] or ''  # Prefer scientific name
        #                 harvest_country = row[2] or ''
        #                 recycled_pct = float(row[3] or 0)
        #         except:
        #             pass
        #
        #     lacey_flag_list.append('Y' if lacey_required else 'N')
        #     lacey_species_list.append(species_name)
        #     lacey_harvest_country_list.append(harvest_country)
        #     lacey_recycled_list.append(recycled_pct)
        #
        # df['_lacey_required'] = lacey_flag_list
        # df['LaceySpecies'] = lacey_species_list
        # df['LaceyHarvestCountry'] = lacey_harvest_country_list
        # df['LaceyRecycledPct'] = lacey_recycled_list
        #
        # # Log Lacey Act summary
        # lacey_count = sum(1 for f in lacey_flag_list if f == 'Y')
        # if lacey_count > 0:
        #     logger.info(f"Lacey Act: {lacey_count} items require PPQ Form 505 declaration")

        # Rename columns for preview
        df['Product No'] = df['part_number']
        df['ValueUSD'] = df['value_usd']

        # Ensure quantity column exists (may not be mapped)
        if 'quantity' not in df.columns:
            df['quantity'] = ''

        # Include invoice_number if mapped (for split by invoice export feature)
        base_preview_cols = [
            'Product No','ValueUSD','HTSCode','MID','CalcWtNet','quantity','qty_unit','Qty1','Qty2','cbp_qty','DecTypeCd',
            'CountryofMelt','CountryOfCast','PrimCountryOfSmelt','DeclarationFlag',
            'SteelRatio','AluminumRatio','CopperRatio','WoodRatio','AutoRatio','NonSteelRatio','_232_flag','_not_in_db','Sec301_Exclusion_Tariff'
            # TODO: Lacey columns - To be implemented at a later date
            # '_lacey_required','LaceySpecies','LaceyHarvestCountry','LaceyRecycledPct'
        ]
        preview_cols = base_preview_cols.copy()
        if 'invoice_number' in df.columns:
            preview_cols.append('invoice_number')
        preview_df = df[preview_cols].copy()
        self.on_done(preview_df, vr, None)
    
    def start_processing_with_editable_preview(self):
        """Open the CSV file in default editor for user to edit directly"""
        if not self.current_csv:
            return
        
        try:
            # Open the CSV file with the default system application
            import subprocess
            if sys.platform == 'win32':
                os.startfile(self.current_csv)
            elif sys.platform == 'darwin':  # macOS
                subprocess.run(['open', self.current_csv])
            else:  # linux
                subprocess.run(['xdg-open', self.current_csv])
            
            # Show message to user
            QMessageBox.information(
                self, 
                "Edit File", 
                f"Opening file for editing:\n{Path(self.current_csv).name}\n\n"
                "Edit the values, save the file, then return here.\n"
                "The CI Value input will be updated when you reload the file."
            )
            
            # Reload the file to get updated values
            self.reload_csv_values()
            
        except Exception as e:
            logger.error(f"Failed to open file: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open file:\n{e}")
    
    def reload_csv_values(self):
        """Reload CSV to recalculate total after external edits"""
        if not self.current_csv:
            return
        
        try:
            col_map = {v: k for k, v in self.shipment_mapping.items()}
            if Path(self.current_csv).suffix.lower() == ".xlsx":
                df = pd.read_excel(self.current_csv, dtype=str)
            else:
                df = pd.read_csv(self.current_csv, dtype=str)
            # Drop columns that would create duplicates after renaming
            for source_col, target_col in col_map.items():
                if source_col in df.columns and target_col in df.columns and source_col != target_col:
                    df = df.drop(columns=[target_col])
            df = df.rename(columns=col_map)

            if 'value_usd' in df.columns:
                # Check for rows where value_usd is blank, empty, or zero
                original_count = len(df)
                df['value_usd'] = pd.to_numeric(df['value_usd'], errors='coerce')
                zero_rows_df = df[df['value_usd'].isna() | (df['value_usd'] == 0)]
                zero_count = len(zero_rows_df)
                
                removed_count = 0
                if zero_count > 0:
                    # Prompt user to confirm deletion
                    reply = QMessageBox.question(
                        self,
                        "Remove Zero Value Rows",
                        f"Found {zero_count} row(s) with blank or zero values.\n\n"
                        f"Do you want to remove these rows?\n\n"
                        f"• Yes: Remove rows and continue processing\n"
                        f"• No: Keep all rows and process as is",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes
                    )
                    
                    if reply == QMessageBox.Yes:
                        # Remove the zero value rows
                        df = df[df['value_usd'].notna() & (df['value_usd'] != 0)]
                        removed_count = original_count - len(df)
                        
                        # Save cleaned data back to file
                        if removed_count > 0:
                            # Rename back to original columns for saving
                            reverse_map = {k: v for k, v in self.shipment_mapping.items()}
                            save_df = df.rename(columns=reverse_map)
                            
                            if Path(self.current_csv).suffix.lower() == ".xlsx":
                                save_df.to_excel(self.current_csv, index=False)
                            else:
                                save_df.to_csv(self.current_csv, index=False)
                            
                            logger.info(f"Removed {removed_count} rows with blank/zero values")
                    else:
                        # User chose No - keep all rows
                        logger.info(f"User chose to keep {zero_count} row(s) with blank/zero values")
                
                # Calculate total
                total = df['value_usd'].sum()
                self.csv_total_value = round(total, 2)
                # Don't overwrite user's CI input - just update the check
                self.update_invoice_check()
                
                if removed_count > 0:
                    self.status.setText(f"File reloaded - Removed {removed_count} blank/zero rows")
                    self.status.setStyleSheet("background:#ff9800; color:white; font-weight:bold; padding:8px;")
                elif zero_count > 0:
                    self.status.setText(f"File reloaded - Kept {zero_count} blank/zero rows")
                    self.status.setStyleSheet("background:#2196F3; color:white; font-weight:bold; padding:8px;")
                else:
                    self.status.setText("File reloaded - Check invoice values")
                    self.status.setStyleSheet("background:#2196F3; color:white; font-weight:bold; padding:8px;")
        except Exception as e:
            logger.error(f"reload_csv_values failed: {e}")
            QMessageBox.warning(self, "Reload Error", f"Failed to reload values:\n{e}")

    def handle_invoice_diff(self, csv_sum, user_entered):
        # Display-only; enablement handled by update/check methods
        diff = abs(csv_sum - user_entered)
        threshold = 0.01
        font_size = get_user_setting_int('font_size', 9)
        if diff > threshold:
            self.invoice_check_label.setText(
                f": CSV = ${csv_sum:,.2f} | "
                f"Entered = ${user_entered:,.2f} | Diff = ${diff:,.2f}"
            )
            self.invoice_check_label.setStyleSheet(f"background:#A4262C; color:white; font-weight:bold; font-size:{font_size}pt; padding:3px;")
        else:
            self.invoice_check_label.setText(f"Match: ${csv_sum:,.2f}")
            self.invoice_check_label.setStyleSheet(f"background:#107C10; color:white; font-weight:bold; font-size:{font_size}pt; padding:3px;")

    def save_edited_values_and_process(self):
        if not hasattr(self, 'editable_invoice_df'):
            return

        try:
            updated = 0
            for row in range(self.missing_table.rowCount()):
                value_item = self.missing_table.item(row, 1)
                if value_item:
                    clean_val = value_item.text().replace('$', '').replace(',', '').strip()
                    try:
                        new_val = float(clean_val)
                        self.editable_invoice_df.at[row, 'value_usd'] = new_val
                        updated += 1
                    except:
                        self.editable_invoice_df.at[row, 'value_usd'] = 0.0

            # Save back to original file
            if self.editable_invoice_path.suffix.lower() == ".xlsx":
                self.editable_invoice_df.to_excel(self.editable_invoice_path, index=False)
            else:
                self.editable_invoice_df.to_csv(self.editable_invoice_path, index=False)

            # Update totals and UI
            new_total = float(self.editable_invoice_df['value_usd'].sum())
            self.ci_input.setText(f"{new_total:,.2f}")
            self.csv_total_value = round(new_total, 2)
            self.update_invoice_check()

            QMessageBox.information(self, "Success", 
                f"Updated {updated} line values!\n\n"
                f"New invoice total: ${new_total:,.2f}\n\n"
                "Starting processing...")

            # Hide button and reprocess
            self.save_and_process_btn.setVisible(False)
            self.start_processing()

        except Exception as e:
            logger.error(f"save_edited_values_and_process failed: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save edits:\n{e}")

    # ====================== MISSING DATA HANDLER ======================
    def log_missing_data_warning(self, missing_df):
        """Log missing data as warning but allow processing to continue"""
        # Note: missing_df has 'Part Number' column (with capital letters)
        part_col = 'Part Number' if 'Part Number' in missing_df.columns else 'part_number'
        parts_list = ", ".join(str(p) for p in missing_df[part_col].tolist()[:10])
        if len(missing_df) > 10:
            parts_list += f" ... and {len(missing_df) - 10} more"
        
        logger.warning(f"Found {len(missing_df)} parts with missing data: {parts_list}")
        self.status.setText(f"⚠ Warning: {len(missing_df)} parts have missing data - review in preview")
        self.status.setStyleSheet("background:#f7bfa1; color:white; font-weight:bold; padding:8px;")

    def on_worker_error(self, msg):
        self.progress.setVisible(False)
        self.process_btn.setEnabled(True)
        QMessageBox.critical(self, "Error", msg)
        self.status.setText("Error")
        self.status.setStyleSheet("background:#dd6e74; color:white; font-weight:bold;")

    def on_done(self, df, vr, fname):
        # Populate preview with editable Value column; export later when totals match
        self.progress.setVisible(False)
        self.last_processed_df = df.copy()
        self.last_output_filename = f"Upload_Sheet_{vr}_{datetime.now():%Y%m%d_%H%M}.xlsx"

        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)  # Disable sorting while populating
        self.table.setRowCount(len(df))
        has_232 = False
        for i, (idx, r) in enumerate(df.iterrows()):
            flag = r.get('_232_flag', '')
            if flag:
                has_232 = True

            value_item = QTableWidgetItem(f"{r['ValueUSD']:,.2f}")
            value_item.setData(Qt.UserRole, r['ValueUSD'])

            # Get ratio values for display
            steel_ratio_val = r.get('SteelRatio', 0.0) or 0.0
            aluminum_ratio_val = r.get('AluminumRatio', 0.0) or 0.0
            copper_ratio_val = r.get('CopperRatio', 0.0) or 0.0
            wood_ratio_val = r.get('WoodRatio', 0.0) or 0.0
            auto_ratio_val = r.get('AutoRatio', 0.0) or 0.0
            non_steel_ratio_val = r.get('NonSteelRatio', 0.0) or 0.0
            is_232_row = steel_ratio_val > 0.0 or aluminum_ratio_val > 0.0 or copper_ratio_val > 0.0 or wood_ratio_val > 0.0 or auto_ratio_val > 0.0

            # Check if part is not in database or has incomplete data
            not_in_db = r.get('_not_in_db', False)
            is_incomplete = flag == 'Incomplete'

            # Set display status - Not Found takes precedence, then Incomplete, then flag
            if not_in_db:
                status_display = "Not Found"
            elif is_incomplete:
                status_display = "Incomplete"
            else:
                status_display = flag

            # Display percentages (empty for "Not Found" and "Incomplete" rows)
            if not_in_db or is_incomplete:
                steel_display = ""
                aluminum_display = ""
                copper_display = ""
                wood_display = ""
                auto_display = ""
                non_steel_display = ""
            else:
                # Values are now stored as percentages (0-100) in the database
                steel_display = f"{steel_ratio_val:.1f}%" if steel_ratio_val > 0 else ""
                aluminum_display = f"{aluminum_ratio_val:.1f}%" if aluminum_ratio_val > 0 else ""
                copper_display = f"{copper_ratio_val:.1f}%" if copper_ratio_val > 0 else ""
                wood_display = f"{wood_ratio_val:.1f}%" if wood_ratio_val > 0 else ""
                auto_display = f"{auto_ratio_val:.1f}%" if auto_ratio_val > 0 else ""
                non_steel_display = f"{non_steel_ratio_val:.1f}%" if non_steel_ratio_val > 0 else ""

            # Get Sec301 exclusion tariff value
            sec301_exclusion = str(r.get('Sec301_Exclusion_Tariff', '')).strip()
            has_sec301_exclusion = bool(sec301_exclusion and sec301_exclusion not in ['', 'nan', 'None'])

            # Create HTS item with tooltip if Sec301 exclusion exists
            hts_item = QTableWidgetItem(str(r.get('HTSCode','')))
            if has_sec301_exclusion:
                hts_item.setToolTip(f"Sec301 Exclusion Tariff: {sec301_exclusion}")

            product_no = str(r['Product No'])

            # Get Qty1 and Qty2 from calculated values (these were computed during processing)
            qty1_val = r.get('Qty1', '')
            qty1_display = str(qty1_val) if pd.notna(qty1_val) and str(qty1_val).strip() not in ['', 'nan', 'None'] else ""
            qty2_val = r.get('Qty2', '')
            qty2_display = str(qty2_val) if pd.notna(qty2_val) and str(qty2_val).strip() not in ['', 'nan', 'None'] else ""

            # Get qty_unit from database (KG, NO, etc.) for display
            qty_unit_display = str(r.get('qty_unit', '')).strip().upper() if pd.notna(r.get('qty_unit')) else ""

            # Get customer reference from input field
            customer_ref_display = self.customer_ref_input.text().strip() if hasattr(self, 'customer_ref_input') and self.customer_ref_input else ""

            # TODO: Lacey Act status - To be implemented at a later date
            # lacey_required = r.get('_lacey_required', 'N')
            # lacey_display = "Y" if lacey_required == 'Y' else ""

            items = [
                QTableWidgetItem(product_no),                        # 0: Product No
                value_item,                                          # 1: Value
                hts_item,                                            # 2: HTS
                QTableWidgetItem(str(r.get('MID',''))),              # 3: MID
                QTableWidgetItem(qty1_display),                      # 4: Qty1
                QTableWidgetItem(qty2_display),                      # 5: Qty2
                QTableWidgetItem(qty_unit_display),                  # 6: Qty Unit
                QTableWidgetItem(str(r.get('DecTypeCd',''))),        # 7: Dec
                QTableWidgetItem(str(r.get('CountryofMelt',''))),    # 8: Melt
                QTableWidgetItem(str(r.get('CountryOfCast',''))),    # 9: Cast
                QTableWidgetItem(str(r.get('PrimCountryOfSmelt',''))), # 10: Smelt
                QTableWidgetItem(str(r.get('DeclarationFlag',''))),    # 11: Flag
                QTableWidgetItem(steel_display),                     # 12: Steel%
                QTableWidgetItem(aluminum_display),                  # 13: Al%
                QTableWidgetItem(copper_display),                    # 14: Cu%
                QTableWidgetItem(wood_display),                      # 15: Wood%
                QTableWidgetItem(auto_display),                      # 16: Auto%
                QTableWidgetItem(non_steel_display),                 # 17: Non-232%
                QTableWidgetItem(status_display),                    # 18: 232 Status
                QTableWidgetItem(customer_ref_display)               # 19: Cust Ref
                # TODO: Lacey column - To be implemented at a later date
                # QTableWidgetItem(lacey_display)                      # 20: Lacey
            ]

            # Make all items editable except Qty1, Qty2, Steel%, Al%, Cu%, Wood%, Auto%, Non-232%, 232 Status
            for idx, item in enumerate(items):
                if idx not in [4, 5, 12, 13, 14, 15, 16, 17, 18]:  # Not Qty1, Qty2, Steel%, Al%, Cu%, Wood%, Auto%, Non-232%, 232 Status
                    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)

            # Set font colors based on Section 232 material type
            # Use "Not_Found" color when part is not in database, "Incomplete" when data is missing
            if not_in_db:
                color_flag = 'Not_Found'
            elif is_incomplete:
                color_flag = 'Incomplete'
            else:
                color_flag = flag
            row_color = self.get_preview_row_color(color_flag)
            for item in items:
                item.setForeground(row_color)
                f = item.font()
                f.setBold(False)
                item.setFont(f)
                item.setTextAlignment(Qt.AlignCenter)  # Center text in all columns

                # Add red border for rows with Sec301 exclusion
                if has_sec301_exclusion:
                    item.setData(Qt.UserRole + 1, 'sec301_exclusion')  # Mark for border styling

            for j, it in enumerate(items):
                self.table.setItem(i, j, it)

            # Apply background color to entire row if Sec301 exclusion exists
            if has_sec301_exclusion:
                # Get user-configured Sec301 background color
                sec301_bg_color = self.get_sec301_bg_color()
                for j in range(len(items)):
                    cell_item = self.table.item(i, j)
                    if cell_item:
                        # Set background color using saved preference
                        cell_item.setBackground(sec301_bg_color)

        self.table.setSortingEnabled(True)  # Re-enable sorting after populating
        self.table.blockSignals(False)
        self.table.itemChanged.connect(self.on_preview_value_edited)
        self.recalculate_total_and_check_match()
        self.apply_column_visibility()  # Apply saved column visibility settings

        # Enable reprocess button after data has been populated
        self.reprocess_btn.setEnabled(True)

        # Reset invoice check display after processing is complete
        self.invoice_check_label.setText("No file loaded")
        self.csv_total_value = 0.0

        # if has_232:
        #     self.status.setText("SECTION 232 ITEMS • EDIT VALUES • EXPORT WHEN READY")
        #     self.status.setStyleSheet("background:#A4262C; color:white; font-weight:bold; font-size:16pt; padding:10px;")
        # else:
        #     self.status.setText("Edit values • Export when total matches")
        #     self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0;")

    def setup_import_tab(self):
        from PyQt5.QtGui import QPalette

        layout = QVBoxLayout(self.tab_import)
        title = QLabel("<h2>Parts Import from CSV/Excel</h2><p>Drag & drop columns to map fields</p>")
        title.setAlignment(Qt.AlignCenter)
        # Use palette text color for consistent title styling
        app = QApplication.instance()
        palette = app.palette()
        text_color = palette.color(QPalette.Text)
        title.setStyleSheet(f"color: {text_color.name()};")
        layout.addWidget(title)

        # Buttons at top
        button_widget = QWidget()
        btn_layout = QHBoxLayout(button_widget)
        btn_load = QPushButton("Load CSV/Excel File")
        btn_load.setStyleSheet(self.get_button_style("info"))
        btn_load.clicked.connect(self.load_csv_for_import)
        btn_reset = QPushButton("Reset Mapping")
        btn_reset.setStyleSheet(self.get_button_style("danger"))
        btn_reset.clicked.connect(self.reset_import_mapping)
        btn_import = QPushButton("Import Now")
        btn_import.setFixedSize(100, 28)
        btn_import.setStyleSheet(self.get_button_style("success") + "QPushButton { padding:4px; }")
        btn_import.clicked.connect(self.start_parts_import)

        btn_update_sec301 = QPushButton("Update Sec301 Exclusion")
        btn_update_sec301.setStyleSheet(self.get_button_style("info"))
        btn_update_sec301.clicked.connect(self.update_sec301_single)

        btn_import_sec301 = QPushButton("Import Sec301 CSV")
        btn_import_sec301.setStyleSheet(self.get_button_style("info"))
        btn_import_sec301.clicked.connect(self.import_sec301_csv)

        btn_layout.addWidget(btn_load)
        btn_layout.addWidget(btn_reset)
        btn_layout.addWidget(btn_update_sec301)
        btn_layout.addWidget(btn_import_sec301)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_import)
        layout.addWidget(button_widget)

        # Main drag/drop area with scroll
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0,0,0,0)
        main_layout.setSpacing(20)

        left = QGroupBox("CSV/Excel Columns - Drag")
        left_outer_layout = QVBoxLayout()
        
        # Add scroll area for columns
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        scroll_widget = QWidget()
        left_layout = QVBoxLayout(scroll_widget)
        left_layout.setContentsMargins(5, 5, 5, 5)
        left_layout.setSpacing(5)
        self.drag_labels = []
        # Don't add stretch here - it will be added after labels are loaded
        
        scroll_area.setWidget(scroll_widget)
        left_outer_layout.addWidget(scroll_area)
        left.setLayout(left_outer_layout)
        
        # Store reference to left_layout for adding labels later
        self.import_left_layout = left_layout

        right = QGroupBox("Available Fields - Drop Here")
        right_outer_layout = QVBoxLayout()
        
        # Add scroll area for drop targets
        right_scroll_area = QScrollArea()
        right_scroll_area.setWidgetResizable(True)
        right_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        right_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        right_scroll_widget = QWidget()
        right_layout = QFormLayout(right_scroll_widget)
        right_layout.setLabelAlignment(Qt.AlignRight)
        right_layout.setContentsMargins(5, 5, 5, 5)
        
        self.import_targets = {}
        fields = {
            "part_number": "Part Number *",
            "hts_code": "HTS Code *",
            "mid": "MID",
            "steel_ratio": "Steel %",
            "aluminum_ratio": "Aluminum %",
            "copper_ratio": "Copper %",
            "wood_ratio": "Wood %",
            "auto_ratio": "Auto %",
            "qty_unit": "Qty Unit",
            "country_of_melt": "Country of Melt",
            "country_of_cast": "Country of Cast",
            "country_of_smelt": "Country of Smelt",
            "Sec301_Exclusion_Tariff": "Sec301 Exclusion Tariff",
            "client_code": "Client Code",
            "description": "Description",
            "country_origin": "Country of Origin"
        }
        drop_labels = {
            "steel_ratio": "Steel%",
            "aluminum_ratio": "Aluminum%",
            "copper_ratio": "Copper%",
            "wood_ratio": "Wood%",
            "auto_ratio": "Auto%"
        }
        for key, name in fields.items():
            drop_label = drop_labels.get(key)
            target = DropTarget(key, name, drop_label)
            target.dropped.connect(self.on_import_drop)
            is_required = "*" in name
            label_text = name.replace(" *", "")
            if is_required:
                label = QLabel(f"{label_text}: <span style='color:red;'>*</span>")
            else:
                label = QLabel(f"{label_text}:")
            right_layout.addRow(label, target)
            self.import_targets[key] = target
        
        right_scroll_area.setWidget(right_scroll_widget)
        right_outer_layout.addWidget(right_scroll_area)
        right.setLayout(right_outer_layout)

        main_layout.addWidget(left, 1)
        main_layout.addWidget(right, 2)
        layout.addWidget(main_widget, 1)
        self.import_widget = main_widget

        self.import_csv_path = None
        self.tab_import.setLayout(layout)

    def load_csv_for_import_from_path(self, path):
        """Load CSV/Excel file from a given path (used by drag-and-drop)"""
        if not path:
            return
        self.import_csv_path = path
        try:
            # Handle both CSV and Excel files
            if path.lower().endswith('.xlsx'):
                df = pd.read_excel(path, nrows=0, dtype=str)
            else:
                df = pd.read_csv(path, nrows=0, dtype=str)
            cols = list(df.columns)

            # Clear previous mappings when loading new file
            for target in self.import_targets.values():
                target.column_name = None
                target.setText(f"Drop {target.field_key} here")
                target.setProperty("occupied", False)
                target.style().unpolish(target)
                target.style().polish(target)

            # Clear existing labels
            for label in self.drag_labels:
                label.setParent(None)
                label.deleteLater()
            self.drag_labels = []
            
            # Add new labels directly to the layout
            for col in cols:
                lbl = DraggableLabel(str(col))
                self.import_left_layout.addWidget(lbl)
                self.drag_labels.append(lbl)

            # Add stretch at the end to push labels to the top
            self.import_left_layout.addStretch()

            # Try to restore saved mappings if they match columns in the new file
            if MAPPING_FILE.exists():
                try:
                    saved_mapping = json.loads(MAPPING_FILE.read_text())
                    # Restore mappings only if the column exists in the new file
                    for field_key, column_name in saved_mapping.items():
                        if column_name in cols and field_key in self.import_targets:
                            target = self.import_targets[field_key]
                            target.column_name = column_name
                            target.setText(f"{field_key} <- {column_name}")
                            target.setProperty("occupied", True)
                            target.style().unpolish(target)
                            target.style().polish(target)
                            logger.info(f"Restored mapping: {field_key} <- {column_name}")
                except Exception as e:
                    logger.warning(f"Failed to restore saved mappings: {e}")

            logger.info(f"Loaded CSV for import (drag-drop): {Path(path).name}")
            self.bottom_status.setText(f"CSV loaded: {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot read CSV:\n{e}")
    
    def load_csv_for_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select CSV/Excel", str(INPUT_DIR), "CSV/Excel Files (*.csv *.xlsx)")
        if not path: return
        self.load_csv_for_import_from_path(path)
    
    def load_csv_for_import_old(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select CSV/Excel", str(INPUT_DIR), "CSV/Excel Files (*.csv *.xlsx)")
        if not path: return
        self.import_csv_path = path
        try:
            # Handle both CSV and Excel files
            if path.lower().endswith('.xlsx'):
                df = pd.read_excel(path, nrows=0, dtype=str)
            else:
                df = pd.read_csv(path, nrows=0, dtype=str)
            cols = list(df.columns)
            for label in self.drag_labels:
                label.setParent(None)
            self.drag_labels = []
            left_layout = self.import_widget.layout().itemAt(0).widget().layout()
            for col in cols:
                lbl = DraggableLabel(str(col))
                left_layout.insertWidget(left_layout.count()-1, lbl)
                self.drag_labels.append(lbl)
            logger.info(f"Loaded CSV for import: {Path(path).name}")
            self.bottom_status.setText(f"CSV loaded: {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot read CSV:\n{e}")

    def reset_import_mapping(self):
        if QMessageBox.question(self, "Reset", "Clear all field mappings and column list?") != QMessageBox.Yes:
            return
        
        # Clear drop targets (right side)
        for target in self.import_targets.values():
            target.column_name = None
            target.setText(f"Drop {target.field_key} here")
            target.setProperty("occupied", False)
            target.style().unpolish(target); target.style().polish(target)
        
        # Clear drag labels (left side - CSV/Excel columns)
        for label in self.drag_labels:
            label.setParent(None)
            label.deleteLater()
        self.drag_labels = []
        
        # Clear the file path
        self.import_csv_path = None
        
        # Delete mapping file if it exists
        if MAPPING_FILE.exists():
            MAPPING_FILE.unlink()
        
        logger.info("Import mapping and column list reset")
        self.bottom_status.setText("Import mapping cleared")

    def on_import_drop(self, field_key, column_name):
        for k, t in self.import_targets.items():
            if t.column_name == column_name and k != field_key:
                t.column_name = None
                t.setText(f"Drop {t.field_key} here")
                t.setProperty("occupied", False)
                t.style().unpolish(t); t.style().polish(t)

        # Update the target that received the drop
        target = self.import_targets[field_key]
        target.column_name = column_name
        target.setText(f"{field_key} <- {column_name}")
        target.setProperty("occupied", True)
        target.style().unpolish(target); target.style().polish(target)

        self.current_mapping = getattr(self, 'current_mapping', {})
        self.current_mapping[field_key] = column_name
        MAPPING_FILE.write_text(json.dumps(self.current_mapping, indent=2))

    def start_parts_import(self):
        if not self.import_csv_path:
            QMessageBox.warning(self, "No File", "Load a CSV or Excel file first")
            return
        mapping = {k: t.column_name for k, t in self.import_targets.items() if t.column_name}
        # Only Part Number and HTS Code are required
        required_fields = ['part_number','hts_code']
        missing = [f for f in required_fields if f not in mapping]
        if missing:
            fields = {
                'part_number': 'Part Number',
                'description': 'Description',
                'quantity': 'Quantity',
                'net_weight': 'Net Weight',
                'value_usd': 'Value (USD)',
                'hts_code': 'HTS Code',
                'mid': 'MID',
                'qty_unit': 'Qty Unit',
                'steel_ratio': 'Steel %',
                'aluminum_ratio': 'Aluminum %',
                'copper_ratio': 'Copper %',
                'wood_ratio': 'Wood %',
                'auto_ratio': 'Auto %',
                'non_steel_ratio': 'Non-Steel %',
            }
        try:
            # Create progress dialog
            progress = QProgressDialog("Loading file...", "Cancel", 0, 100, self)
            progress.setWindowTitle("Parts Import Progress")
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setMinimumWidth(400)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            # Load file
            progress.setLabelText("Reading file...")
            if self.import_csv_path.lower().endswith('.xlsx'):
                df = pd.read_excel(self.import_csv_path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_csv(self.import_csv_path, dtype=str, keep_default_na=False)

            if progress.wasCanceled():
                return
            progress.setValue(10)

            df = df.fillna("").rename(columns=str.strip)
            col_map = {v: k for k, v in mapping.items()}
            df = df.rename(columns=col_map)

            # Only Part Number and HTS Code are required
            required = ['part_number','hts_code']
            missing = [f for f in required if f not in df.columns]
            if missing:
                progress.close()
                QMessageBox.critical(self, "Error", f"Missing required fields: {', '.join(missing)}")
                self.status.setText("Import failed")
                return

            total_rows = len(df)
            progress.setLabelText(f"Validating {total_rows} rows...")
            progress.setValue(15)
            QApplication.processEvents()

            # First pass: validate percentages and collect any rows with total > 100%
            invalid_ratio_rows = []
            for idx, r in df.iterrows():
                if progress.wasCanceled():
                    return

                part = str(r.get('part_number', '')).strip()
                if not part:
                    continue

                # Helper function to parse percentage values for validation
                # Database now stores percentages (0-100) instead of ratios (0-1)
                def parse_percentage_val(value_str):
                    try:
                        if value_str:
                            pct = float(value_str)
                            # If value is <= 1, assume it's a ratio and convert to percentage
                            if 0 < pct <= 1.0:
                                pct *= 100.0
                            return max(0.0, pct)
                        return 0.0
                    except:
                        return 0.0

                steel_val = parse_percentage_val(str(r.get('steel_ratio', r.get('Sec 232 Content Ratio', r.get('Steel %', '')))).strip())
                aluminum_val = parse_percentage_val(str(r.get('aluminum_ratio', r.get('Aluminum %', ''))).strip())
                copper_val = parse_percentage_val(str(r.get('copper_ratio', r.get('Copper %', ''))).strip())
                wood_val = parse_percentage_val(str(r.get('wood_ratio', r.get('Wood %', ''))).strip())
                auto_val = parse_percentage_val(str(r.get('auto_ratio', r.get('Auto %', ''))).strip())
                non_steel_val = parse_percentage_val(str(r.get('non_steel_ratio', r.get('Non-Steel %', ''))).strip())

                total_pct = steel_val + aluminum_val + copper_val + wood_val + auto_val + non_steel_val
                if total_pct > 101.0:  # Allow small floating point tolerance
                    invalid_ratio_rows.append((part, total_pct))

            progress.setValue(25)

            # If there are invalid rows, reject the import
            if invalid_ratio_rows:
                progress.close()
                msg = f"Import failed. The following {len(invalid_ratio_rows)} part(s) have total percentages exceeding 100%:\n\n"
                for part, total in invalid_ratio_rows[:15]:  # Show first 15
                    msg += f"  {part}: {total:.1f}%\n"
                if len(invalid_ratio_rows) > 15:
                    msg += f"  ... and {len(invalid_ratio_rows) - 15} more\n"
                msg += "\nPlease correct these rows in your import file and try again."

                QMessageBox.critical(self, "Invalid Percentages Detected", msg)
                self.status.setText("Import failed - invalid percentages")
                return

            progress.setLabelText(f"Importing {total_rows} parts to database...")
            progress.setValue(30)
            QApplication.processEvents()

            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            updated = inserted = 0
            now = datetime.now().isoformat()

            # Calculate progress increment per row (30% to 90% = 60% for import)
            progress_per_row = 60.0 / max(total_rows, 1)

            for row_idx, (_, r) in enumerate(df.iterrows()):
                if progress.wasCanceled():
                    conn.close()
                    return

                # Update progress every 100 rows or at least every 1%
                if row_idx % 100 == 0 or row_idx == total_rows - 1:
                    current_progress = 30 + int(row_idx * progress_per_row)
                    progress.setValue(min(current_progress, 90))
                    progress.setLabelText(f"Importing parts... ({row_idx + 1}/{total_rows})")
                    QApplication.processEvents()

                part = str(r.get('part_number', '')).strip()
                if not part: continue
                desc = str(r.get('description', r.get('Description', ''))).strip()
                hts = str(r.get('hts_code', '')).strip()
                origin = str(r.get('country_origin', '')).strip().upper()[:2]
                mid = str(r.get('mid', '')).strip()
                # Get client_code if it was mapped, otherwise empty string
                client_code = str(r.get('client_code', '')).strip() if 'client_code' in df.columns else ""
                # Get qty_unit if it was mapped, otherwise try to auto-lookup from hts_units table
                qty_unit = str(r.get('qty_unit', '')).strip() if 'qty_unit' in df.columns else ""
                if not qty_unit and hts:
                    qty_unit = get_hts_qty_unit(hts)
                # Get country codes if mapped, otherwise empty string
                country_of_melt = str(r.get('country_of_melt', '')).strip().upper()[:2] if 'country_of_melt' in df.columns else ""
                country_of_cast = str(r.get('country_of_cast', '')).strip().upper()[:2] if 'country_of_cast' in df.columns else ""
                country_of_smelt = str(r.get('country_of_smelt', '')).strip().upper()[:2] if 'country_of_smelt' in df.columns else ""
                # Get Sec301_Exclusion_Tariff if mapped, otherwise empty string
                sec301_exclusion = str(r.get('Sec301_Exclusion_Tariff', '')).strip() if 'Sec301_Exclusion_Tariff' in df.columns else ""

                # Helper function to parse percentage values (database stores 0-100)
                def parse_percentage(value_str):
                    try:
                        if value_str:
                            pct = float(value_str)
                            # If value is <= 1, assume it's a ratio and convert to percentage
                            if 0 < pct <= 1.0:
                                pct *= 100.0
                            return max(0.0, min(100.0, pct))
                        return 0.0
                    except:
                        return 0.0

                # Parse all percentage fields
                steel_str = str(r.get('steel_ratio', r.get('Sec 232 Content Ratio', r.get('Steel %', '')))).strip()
                steel_ratio = parse_percentage(steel_str)

                aluminum_str = str(r.get('aluminum_ratio', r.get('Aluminum %', ''))).strip()
                aluminum_ratio = parse_percentage(aluminum_str)

                copper_str = str(r.get('copper_ratio', r.get('Copper %', ''))).strip()
                copper_ratio = parse_percentage(copper_str)

                wood_str = str(r.get('wood_ratio', r.get('Wood %', ''))).strip()
                wood_ratio = parse_percentage(wood_str)

                auto_str = str(r.get('auto_ratio', r.get('Auto %', ''))).strip()
                auto_ratio = parse_percentage(auto_str)

                # Calculate non_steel_ratio as remainder (100 minus all 232 percentages)
                total_232 = steel_ratio + aluminum_ratio + copper_ratio + wood_ratio + auto_ratio
                non_steel_ratio = max(0.0, 100.0 - total_232)

                c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                          steel_ratio, non_steel_ratio, last_updated, qty_unit, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                          country_of_melt, country_of_cast, country_of_smelt, Sec301_Exclusion_Tariff)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                          ON CONFLICT(part_number) DO UPDATE SET
                          description=excluded.description, hts_code=excluded.hts_code,
                          country_origin=excluded.country_origin, mid=excluded.mid,
                          client_code=excluded.client_code, steel_ratio=excluded.steel_ratio,
                          non_steel_ratio=excluded.non_steel_ratio, last_updated=excluded.last_updated,
                          qty_unit=excluded.qty_unit, aluminum_ratio=excluded.aluminum_ratio,
                          copper_ratio=excluded.copper_ratio, wood_ratio=excluded.wood_ratio,
                          auto_ratio=excluded.auto_ratio, country_of_melt=excluded.country_of_melt,
                          country_of_cast=excluded.country_of_cast, country_of_smelt=excluded.country_of_smelt,
                          Sec301_Exclusion_Tariff=excluded.Sec301_Exclusion_Tariff""",
                          (part, desc, hts, origin, mid, client_code, steel_ratio, non_steel_ratio, now,
                           qty_unit, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                           country_of_melt, country_of_cast, country_of_smelt, sec301_exclusion))
                if c.rowcount:
                    inserted += 1 if conn.total_changes > updated+inserted else 0
                    updated += 1 if conn.total_changes == updated+inserted else 0

            progress.setLabelText("Committing to database...")
            progress.setValue(92)
            QApplication.processEvents()

            conn.commit()
            conn.close()

            progress.setLabelText("Refreshing data...")
            progress.setValue(95)
            QApplication.processEvents()

            # Only refresh parts table if Parts View tab has been initialized
            if hasattr(self, 'parts_table'):
                self.refresh_parts_table()

            self.load_available_mids()

            progress.setValue(100)
            progress.close()

            QMessageBox.information(self, "Success", f"Import Complete!\n\nTotal rows processed: {total_rows}\nParts imported/updated: {inserted + updated}")
            self.bottom_status.setText("Import complete")
            logger.info(f"Parts import complete: {total_rows} rows, {inserted + updated} parts imported/updated")

        except Exception as e:
            if 'progress' in locals():
                progress.close()
            logger.error(f"Import failed: {e}")
            QMessageBox.critical(self, "Error", f"Import failed: {str(e)}")
            self.status.setText("Import failed")

    def update_sec301_single(self):
        """Update Section 301 Exclusion Tariff for a single part number"""
        # Create dialog for input
        dialog = QDialog(self)
        dialog.setWindowTitle("Update Sec301 Exclusion Tariff")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout(dialog)

        # Instructions
        instructions = QLabel("Enter the part number and Section 301 Exclusion Tariff to update:")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        # Form layout
        form_layout = QFormLayout()

        part_number_input = QLineEdit()
        part_number_input.setPlaceholderText("e.g., PART12345")
        form_layout.addRow("Part Number:", part_number_input)

        sec301_input = QLineEdit()
        sec301_input.setPlaceholderText("e.g., 9903.88.15")
        form_layout.addRow("Sec301 Exclusion Tariff:", sec301_input)

        layout.addLayout(form_layout)

        # Buttons
        button_box = QHBoxLayout()
        btn_update = QPushButton("Update")
        btn_update.setStyleSheet(self.get_button_style("success"))
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setStyleSheet(self.get_button_style("danger"))

        btn_update.clicked.connect(dialog.accept)
        btn_cancel.clicked.connect(dialog.reject)

        button_box.addWidget(btn_update)
        button_box.addWidget(btn_cancel)
        layout.addLayout(button_box)

        # Show dialog
        self.center_dialog(dialog)
        if dialog.exec_() == QDialog.Accepted:
            part_number = part_number_input.text().strip()
            sec301_tariff = sec301_input.text().strip()

            if not part_number:
                QMessageBox.warning(self, "Input Required", "Please enter a part number.")
                return

            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()

                # Check if part exists
                c.execute("SELECT part_number FROM parts_master WHERE part_number=?", (part_number,))
                exists = c.fetchone()

                if not exists:
                    QMessageBox.warning(self, "Part Not Found",
                                      f"Part number '{part_number}' not found in database.\n\n"
                                      "Please import the part first using the regular import function.")
                    conn.close()
                    return

                # Update the Sec301_Exclusion_Tariff
                now = datetime.now().isoformat()
                c.execute("""UPDATE parts_master
                            SET Sec301_Exclusion_Tariff=?, last_updated=?
                            WHERE part_number=?""",
                         (sec301_tariff, now, part_number))
                conn.commit()
                conn.close()

                # Refresh parts table if visible
                if hasattr(self, 'parts_table'):
                    self.refresh_parts_table()

                QMessageBox.information(self, "Success",
                                      f"Updated Sec301 Exclusion Tariff for part '{part_number}'")
                logger.info(f"Updated Sec301 Exclusion Tariff for {part_number}: {sec301_tariff}")
                self.status.setText(f"Updated Sec301 for {part_number}")

            except Exception as e:
                logger.error(f"Failed to update Sec301 exclusion: {e}")
                QMessageBox.critical(self, "Error", f"Failed to update:\n{e}")

    def import_sec301_csv(self):
        """Import Section 301 Exclusion Tariffs from CSV file"""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Sec301 Exclusions CSV File",
            "",
            "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*.*)"
        )
        if not path:
            return

        try:
            # Read the file
            if path.lower().endswith('.xlsx'):
                df = pd.read_excel(path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_csv(path, dtype=str, keep_default_na=False)

            df = df.fillna("").rename(columns=str.strip)

            # Show column mapping dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Map CSV Columns")
            dialog.setMinimumWidth(500)
            layout = QVBoxLayout(dialog)

            instructions = QLabel(
                "<b>Map your CSV columns to the required fields:</b><br><br>"
                "Select which columns from your CSV contain the Part Number and Sec301 Exclusion Tariff."
            )
            instructions.setWordWrap(True)
            layout.addWidget(instructions)

            form_layout = QFormLayout()

            part_combo = QComboBox()
            part_combo.addItems(["-- Select Column --"] + list(df.columns))
            form_layout.addRow("Part Number Column:", part_combo)

            sec301_combo = QComboBox()
            sec301_combo.addItems(["-- Select Column --"] + list(df.columns))
            form_layout.addRow("Sec301 Exclusion Tariff Column:", sec301_combo)

            layout.addLayout(form_layout)

            # Buttons
            button_box = QHBoxLayout()
            btn_import = QPushButton("Import")
            btn_import.setStyleSheet(self.get_button_style("success"))
            btn_cancel = QPushButton("Cancel")
            btn_cancel.setStyleSheet(self.get_button_style("danger"))

            btn_import.clicked.connect(dialog.accept)
            btn_cancel.clicked.connect(dialog.reject)

            button_box.addWidget(btn_import)
            button_box.addWidget(btn_cancel)
            layout.addLayout(button_box)

            # Show dialog
            self.center_dialog(dialog)
            if dialog.exec_() == QDialog.Accepted:
                part_col = part_combo.currentText()
                sec301_col = sec301_combo.currentText()

                if part_col == "-- Select Column --" or sec301_col == "-- Select Column --":
                    QMessageBox.warning(self, "Selection Required",
                                      "Please select both columns before importing.")
                    return

                # Process the import
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                updated = 0
                not_found = []
                now = datetime.now().isoformat()

                for _, row in df.iterrows():
                    part_number = str(row.get(part_col, '')).strip()
                    sec301_tariff = str(row.get(sec301_col, '')).strip()

                    if not part_number:
                        continue

                    # Check if part exists
                    c.execute("SELECT part_number FROM parts_master WHERE part_number=?", (part_number,))
                    exists = c.fetchone()

                    if exists:
                        c.execute("""UPDATE parts_master
                                    SET Sec301_Exclusion_Tariff=?, last_updated=?
                                    WHERE part_number=?""",
                                 (sec301_tariff, now, part_number))
                        updated += 1
                    else:
                        not_found.append(part_number)

                conn.commit()
                conn.close()

                # Refresh parts table if visible
                if hasattr(self, 'parts_table'):
                    self.refresh_parts_table()

                # Show results
                result_msg = f"Updated {updated} part(s) with Sec301 Exclusion Tariffs."
                if not_found:
                    result_msg += f"\n\n{len(not_found)} part(s) not found in database:\n"
                    result_msg += "\n".join(not_found[:10])  # Show first 10
                    if len(not_found) > 10:
                        result_msg += f"\n... and {len(not_found) - 10} more"

                QMessageBox.information(self, "Import Complete", result_msg)
                logger.info(f"Imported Sec301 exclusions: {updated} updated, {len(not_found)} not found")
                self.status.setText(f"Imported {updated} Sec301 exclusions")

        except Exception as e:
            logger.error(f"Failed to import Sec301 CSV: {e}")
            QMessageBox.critical(self, "Error", f"Failed to import CSV:\n{e}")

    def setup_shipment_mapping_tab(self):
        logger.debug(f"setup_shipment_mapping_tab called - tab_shipment_map={self.tab_shipment_map}")
        layout = QVBoxLayout(self.tab_shipment_map)
        title = QLabel("<h2>Invoice Mapping Profiles</h2><p>Save and load column mappings</p>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Buttons at top - wrap in widget for proper rendering
        top_bar_widget = QWidget()
        top_bar = QHBoxLayout(top_bar_widget)
        self.profile_combo_map = QComboBox()
        self.profile_combo_map.setMinimumWidth(300)
        self.profile_combo_map.currentTextChanged.connect(self.load_selected_profile_full)
        top_bar.addWidget(QLabel("Saved Profiles:"))
        top_bar.addWidget(self.profile_combo_map)

        # Load profiles immediately after creating the combo box
        self.load_mapping_profiles()

        # Add header row input
        top_bar.addWidget(QLabel("Header Row:"))
        self.header_row_input = QLineEdit()
        self.header_row_input.setPlaceholderText("1")
        self.header_row_input.setMaximumWidth(50)
        self.header_row_input.setToolTip("Row number containing column headers (1 = first row, 2 = second row, etc.)")
        top_bar.addWidget(self.header_row_input)

        btn_save = QPushButton("Save Current Mapping As...")
        btn_save.setStyleSheet(self.get_button_style("success"))
        btn_save.clicked.connect(self.save_mapping_profile)
        btn_delete = QPushButton("Delete Profile")
        btn_delete.setStyleSheet(self.get_button_style("danger"))
        btn_delete.clicked.connect(self.delete_mapping_profile)
        btn_load_csv = QPushButton("Load Invoice File")
        btn_load_csv.setStyleSheet(self.get_button_style("info"))
        btn_load_csv.clicked.connect(self.load_csv_for_shipment_mapping)
        btn_reset = QPushButton("Reset Current")
        btn_reset.setStyleSheet(self.get_button_style("danger"))
        btn_reset.clicked.connect(self.reset_current_mapping)
        top_bar.addWidget(btn_load_csv)
        top_bar.addWidget(btn_reset)
        top_bar.addWidget(btn_save)
        top_bar.addWidget(btn_delete)
        top_bar.addStretch()
        layout.addWidget(top_bar_widget)

        # Profile linking row - link input profile to export profile
        link_bar_widget = QWidget()
        link_bar = QHBoxLayout(link_bar_widget)
        link_bar.addWidget(QLabel("Linked Export Profile:"))
        self.linked_export_combo = QComboBox()
        self.linked_export_combo.setMinimumWidth(250)
        self.linked_export_combo.addItem("(None)")
        # Populate with export profiles
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT DISTINCT profile_name FROM output_column_mappings ORDER BY profile_name")
            for row in c.fetchall():
                self.linked_export_combo.addItem(row[0])
            conn.close()
        except Exception as e:
            logger.warning(f"Failed to load export profiles for linking: {e}")
        link_bar.addWidget(self.linked_export_combo)

        btn_save_link = QPushButton("Save Link")
        btn_save_link.setStyleSheet(self.get_button_style("success"))
        btn_save_link.clicked.connect(self.save_profile_link)
        btn_save_link.setToolTip("Save the link between the current input profile and selected export profile")
        link_bar.addWidget(btn_save_link)

        btn_clear_link = QPushButton("Clear Link")
        btn_clear_link.setStyleSheet(self.get_button_style("secondary"))
        btn_clear_link.clicked.connect(self.clear_profile_link)
        btn_clear_link.setToolTip("Remove the link for the current input profile")
        link_bar.addWidget(btn_clear_link)

        link_bar.addStretch()
        layout.addWidget(link_bar_widget)

        # Main container with three columns side by side
        self.shipment_widget = QWidget()
        self.shipment_layout = QHBoxLayout(self.shipment_widget)
        self.shipment_layout.setSpacing(10)

        # LEFT: CSV Columns - Drag (with independent scroll)
        left = QGroupBox("Your CSV Columns - Drag")
        left_main_layout = QVBoxLayout()
        left_main_layout.setContentsMargins(5, 5, 5, 5)

        # Scroll area for drag labels
        self.left_scroll = QScrollArea()
        self.left_scroll.setWidgetResizable(True)
        self.left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_scroll.setMinimumHeight(200)
        left_scroll_widget = QWidget()
        self.left_scroll_layout = QVBoxLayout(left_scroll_widget)
        self.left_scroll_layout.setSpacing(4)
        self.shipment_drag_labels = []
        self.left_scroll_layout.addStretch()
        self.left_scroll.setWidget(left_scroll_widget)
        left_main_layout.addWidget(self.left_scroll)
        left.setLayout(left_main_layout)

        # CENTER: Required Fields - Drop (with independent scroll)
        right = QGroupBox("Required Fields - Drop")
        right_main_layout = QVBoxLayout()
        right_main_layout.setContentsMargins(5, 5, 5, 5)

        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        right_scroll.setMinimumHeight(200)
        right_scroll_widget = QWidget()
        right_layout = QFormLayout(right_scroll_widget)
        right_layout.setLabelAlignment(Qt.AlignRight)
        self.shipment_targets = {}
        required_fields = {
            "part_number": "Part Number *",
            "value_usd": "Value USD *"
        }
        for key, name in required_fields.items():
            target = DropTarget(key, name)
            target.dropped.connect(self.on_shipment_drop)
            right_layout.addRow(f"{name}:", target)
            self.shipment_targets[key] = target
        right_scroll.setWidget(right_scroll_widget)
        right_main_layout.addWidget(right_scroll)
        right.setLayout(right_main_layout)

        # RIGHT: Optional Fields - Drop (with independent scroll)
        optional = QGroupBox("Optional Fields - Drop")
        optional_main_layout = QVBoxLayout()
        optional_main_layout.setContentsMargins(5, 5, 5, 5)

        optional_scroll = QScrollArea()
        optional_scroll.setWidgetResizable(True)
        optional_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        optional_scroll.setMinimumHeight(200)
        optional_scroll_widget = QWidget()
        optional_layout = QFormLayout(optional_scroll_widget)
        optional_layout.setLabelAlignment(Qt.AlignRight)
        optional_fields = {
            "invoice_number": "Invoice Number",
            "quantity": "Quantity",
            "quantity_unit": "Quantity Unit",
            "net_weight": "Net Weight",
            "hts_code": "HTS Code",
            "qty_unit": "Qty Unit"
        }
        for key, name in optional_fields.items():
            target = DropTarget(key, name)
            target.dropped.connect(self.on_shipment_drop)
            optional_layout.addRow(f"{name}:", target)
            self.shipment_targets[key] = target
        optional_scroll.setWidget(optional_scroll_widget)
        optional_main_layout.addWidget(optional_scroll)
        optional.setLayout(optional_main_layout)

        self.shipment_layout.addWidget(left, 1)
        self.shipment_layout.addWidget(right, 1)
        self.shipment_layout.addWidget(optional, 1)
        layout.addWidget(self.shipment_widget, 1)
        self.tab_shipment_map.setLayout(layout)

    def setup_output_mapping_tab(self):
        """Setup the Output XLSX Mapping tab for customizing export column names and order"""
        layout = QVBoxLayout(self.tab_output_map)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)

        title = QLabel("<h2>Output Column Mapping</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Top bar with profile management
        top_bar_widget = QWidget()
        top_bar = QHBoxLayout(top_bar_widget)
        top_bar.setContentsMargins(0, 0, 0, 0)
        self.output_profile_combo = QComboBox()
        self.output_profile_combo.setMinimumWidth(250)
        self.output_profile_combo.currentTextChanged.connect(self.load_output_mapping_profile)
        top_bar.addWidget(QLabel("Saved Profiles:"))
        top_bar.addWidget(self.output_profile_combo)

        # Load existing profiles
        self.load_output_mapping_profiles()

        btn_reset_output = QPushButton("Reset to Default")
        btn_reset_output.setStyleSheet(self.get_button_style("info"))
        btn_reset_output.clicked.connect(self.reset_output_mapping)

        btn_save_output = QPushButton("Save As New...")
        btn_save_output.setStyleSheet(self.get_button_style("success"))
        btn_save_output.clicked.connect(self.save_output_mapping_profile)

        btn_edit_output = QPushButton("Update Profile")
        btn_edit_output.setStyleSheet(self.get_button_style("warning"))
        btn_edit_output.clicked.connect(self.update_output_mapping_profile)

        btn_delete_output = QPushButton("Delete Profile")
        btn_delete_output.setStyleSheet(self.get_button_style("danger"))
        btn_delete_output.clicked.connect(self.delete_output_mapping_profile)

        top_bar.addWidget(btn_reset_output)
        top_bar.addWidget(btn_save_output)
        top_bar.addWidget(btn_edit_output)
        top_bar.addWidget(btn_delete_output)
        top_bar.addStretch()
        layout.addWidget(top_bar_widget)

        # Helper function to create compact export color swatch with label
        def create_export_color_swatch(label_text, config_key, default_color):
            """Create a label with small color swatch button (per-user setting)"""
            container = QWidget()
            container.setMinimumHeight(28)
            layout = QHBoxLayout(container)
            layout.setContentsMargins(0, 2, 8, 2)
            layout.setSpacing(6)

            # Text label - let it size naturally based on content
            label = QLabel(label_text + ":")
            layout.addWidget(label)

            # Small color swatch button
            button = QPushButton()
            button.setFixedSize(20, 20)
            button.setCursor(QCursor(Qt.PointingHandCursor))

            # Load saved color from per-user settings or use default
            saved_color = get_user_setting(config_key, default_color)

            def update_button_style(color_hex):
                button.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {color_hex};
                        border: 1px solid #555;
                        border-radius: 3px;
                    }}
                    QPushButton:hover {{
                        border: 2px solid #888;
                    }}
                    QPushButton:pressed {{
                        border: 2px solid #aaa;
                    }}
                """)

            update_button_style(saved_color)

            def pick_color():
                current_color = get_user_setting(config_key, default_color)
                color = QColorDialog.getColor(QColor(current_color), self, f"Choose {label_text} Color")
                if color.isValid():
                    color_hex = color.name()
                    update_button_style(color_hex)
                    set_user_setting(config_key, color_hex)
                    logger.info(f"Saved export color preference {config_key}: {color_hex}")
                    self.bottom_status.setText(f"{label_text} export color set to {color_hex}")

            button.clicked.connect(pick_color)
            layout.addWidget(button)
            return container

        # === COLORS AND OPTIONS SECTION (horizontal layout) ===
        options_widget = QWidget()
        options_layout = QHBoxLayout(options_widget)
        options_layout.setContentsMargins(0, 5, 0, 5)
        options_layout.setSpacing(15)

        # --- Left: Export Text Colors ---
        colors_group = QGroupBox("Export Text Colors")
        colors_main_layout = QVBoxLayout(colors_group)
        colors_main_layout.setSpacing(8)
        colors_main_layout.setContentsMargins(10, 10, 10, 10)

        # Load saved font color (per-user setting)
        self.output_font_color = get_user_setting('output_font_color', '#000000')

        # Create font color swatch container
        font_color_container = QWidget()
        font_color_container.setMinimumHeight(28)
        font_color_layout = QHBoxLayout(font_color_container)
        font_color_layout.setContentsMargins(0, 2, 8, 2)
        font_color_layout.setSpacing(6)
        font_color_label = QLabel("Default:")
        font_color_layout.addWidget(font_color_label)
        self.output_font_color_btn = QPushButton()
        self.output_font_color_btn.setFixedSize(20, 20)
        self.output_font_color_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.output_font_color_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.output_font_color};
                border: 1px solid #555;
                border-radius: 3px;
            }}
            QPushButton:hover {{
                border: 2px solid #888;
            }}
        """)
        self.output_font_color_btn.clicked.connect(self.pick_output_font_color)
        font_color_layout.addWidget(self.output_font_color_btn)

        # Row 1: Default, Steel, Aluminum
        row1_layout = QHBoxLayout()
        row1_layout.setSpacing(20)
        row1_layout.addWidget(font_color_container)
        row1_layout.addWidget(create_export_color_swatch("Steel", 'export_steel_color', '#4a4a4a'))
        row1_layout.addWidget(create_export_color_swatch("Aluminum", 'export_aluminum_color', '#6495ED'))
        row1_layout.addStretch()
        colors_main_layout.addLayout(row1_layout)

        # Row 2: Copper, Wood, Auto
        row2_layout = QHBoxLayout()
        row2_layout.setSpacing(20)
        row2_layout.addWidget(create_export_color_swatch("Copper", 'export_copper_color', '#B87333'))
        row2_layout.addWidget(create_export_color_swatch("Wood", 'export_wood_color', '#8B4513'))
        row2_layout.addWidget(create_export_color_swatch("Auto", 'export_automotive_color', '#2F4F4F'))
        row2_layout.addStretch()
        colors_main_layout.addLayout(row2_layout)

        # Row 3: Non-232
        row3_layout = QHBoxLayout()
        row3_layout.setSpacing(20)
        row3_layout.addWidget(create_export_color_swatch("Non-232", 'export_non232_color', '#FF0000'))
        row3_layout.addStretch()
        colors_main_layout.addLayout(row3_layout)

        options_layout.addWidget(colors_group)

        # --- Middle: Column Visibility ---
        visibility_group = QGroupBox("Column Visibility")
        visibility_layout = QGridLayout(visibility_group)
        visibility_layout.setSpacing(5)
        visibility_layout.setContentsMargins(10, 10, 10, 10)

        self.output_column_visibility = {}
        ratio_columns = ['SteelRatio', 'AluminumRatio', 'CopperRatio', 'WoodRatio', 'AutoRatio', 'NonSteelRatio']

        for idx, col in enumerate(ratio_columns):
            is_visible = True
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("SELECT value FROM app_config WHERE key = ?", (f'export_col_visible_{col}',))
                row = c.fetchone()
                conn.close()
                if row:
                    is_visible = row[0] == 'True'
            except:
                pass

            checkbox = QCheckBox(col.replace('Ratio', '%'))
            checkbox.setChecked(is_visible)
            checkbox.stateChanged.connect(lambda state, col=col: self.update_column_visibility(col, state))
            self.output_column_visibility[col] = checkbox
            # 2 columns layout
            visibility_layout.addWidget(checkbox, idx // 2, idx % 2)

        options_layout.addWidget(visibility_group)

        # --- Right: Export Options ---
        export_options_group = QGroupBox("Export Options")
        export_options_layout = QVBoxLayout(export_options_group)
        export_options_layout.setContentsMargins(10, 10, 10, 10)

        self.split_by_invoice = False
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", ('export_split_by_invoice',))
            row = c.fetchone()
            conn.close()
            if row:
                self.split_by_invoice = row[0] == 'True'
        except:
            pass

        self.split_by_invoice_checkbox = QCheckBox("Split by Invoice Number")
        self.split_by_invoice_checkbox.setChecked(self.split_by_invoice)
        self.split_by_invoice_checkbox.stateChanged.connect(self.update_split_by_invoice_setting)
        export_options_layout.addWidget(self.split_by_invoice_checkbox)

        split_note = QLabel("Creates separate files per invoice.\nRequires Invoice Number mapping.")
        split_note.setStyleSheet("color: gray;")
        split_note.setWordWrap(True)
        export_options_layout.addWidget(split_note)
        export_options_layout.addStretch()

        options_layout.addWidget(export_options_group)
        options_layout.addStretch()

        layout.addWidget(options_widget)

        # === COLUMN MAPPING SECTION ===
        mapping_group = QGroupBox("Column Name Mapping (drag to reorder)")
        mapping_layout = QVBoxLayout(mapping_group)
        mapping_layout.setContentsMargins(10, 10, 10, 10)

        # Scrollable area for column mappings
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(300)
        self.column_mapping_scroll_widget = QWidget()
        self.column_mapping_scroll_layout = QVBoxLayout(self.column_mapping_scroll_widget)
        self.column_mapping_scroll_layout.setSpacing(2)
        self.column_mapping_scroll_layout.setContentsMargins(5, 5, 5, 5)

        # Default column order and mappings
        # Qty1/Qty2 are conditional: based on qty_unit (KG, NO, NO/KG)
        self.default_output_column_order = [
            'Product No', 'ValueUSD', 'HTSCode', 'MID', 'Qty1', 'Qty2',
            'DecTypeCd', 'CountryofMelt', 'CountryOfCast', 'PrimCountryOfSmelt',
            'DeclarationFlag', 'SteelRatio', 'AluminumRatio', 'CopperRatio',
            'WoodRatio', 'AutoRatio', 'NonSteelRatio', '232_Status', 'CustomerRef'
        ]

        # Default column mappings (internal_name: display_name)
        self.default_output_columns = {name: name for name in self.default_output_column_order}

        # Initialize current mapping and order if not exists or is None
        if not hasattr(self, 'output_column_mapping') or self.output_column_mapping is None:
            self.output_column_mapping = self.default_output_columns.copy()
        if not hasattr(self, 'output_column_order') or self.output_column_order is None:
            self.output_column_order = self.default_output_column_order.copy()

        # Build the column rows UI
        self.rebuild_column_mapping_ui()

        scroll.setWidget(self.column_mapping_scroll_widget)
        mapping_layout.addWidget(scroll)
        layout.addWidget(mapping_group, 1)
        self.tab_output_map.setLayout(layout)

    def rebuild_column_mapping_ui(self):
        """Rebuild the column mapping UI based on current order"""
        # Clear existing widgets
        while self.column_mapping_scroll_layout.count():
            item = self.column_mapping_scroll_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        self.output_column_inputs = {}
        self.output_column_rows = {}

        for idx, internal_name in enumerate(self.output_column_order):
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(5)

            # Up button
            btn_up = QPushButton("▲")
            btn_up.setFixedSize(25, 25)
            btn_up.setEnabled(idx > 0)
            btn_up.clicked.connect(lambda checked, name=internal_name: self.move_column_up(name))
            btn_up.setToolTip("Move column up")

            # Down button
            btn_down = QPushButton("▼")
            btn_down.setFixedSize(25, 25)
            btn_down.setEnabled(idx < len(self.output_column_order) - 1)
            btn_down.clicked.connect(lambda checked, name=internal_name: self.move_column_down(name))
            btn_down.setToolTip("Move column down")

            # Internal name label
            label = QLabel(f"{internal_name}:")
            label.setFixedWidth(140)
            label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

            # Display name input
            display_name = self.output_column_mapping.get(internal_name, internal_name)
            line_edit = QLineEdit(display_name)
            line_edit.setMinimumWidth(150)
            line_edit.textChanged.connect(lambda text, key=internal_name: self.update_output_column_name(key, text))
            self.output_column_inputs[internal_name] = line_edit

            row_layout.addWidget(btn_up)
            row_layout.addWidget(btn_down)
            row_layout.addWidget(label)
            row_layout.addWidget(line_edit)
            row_layout.addStretch()

            self.output_column_rows[internal_name] = row_widget
            self.column_mapping_scroll_layout.addWidget(row_widget)

        # Add stretch at the end
        self.column_mapping_scroll_layout.addStretch()

    def move_column_up(self, internal_name):
        """Move a column up in the order"""
        if internal_name not in self.output_column_order:
            return
        idx = self.output_column_order.index(internal_name)
        if idx > 0:
            # Swap with previous
            self.output_column_order[idx], self.output_column_order[idx - 1] = \
                self.output_column_order[idx - 1], self.output_column_order[idx]
            self.rebuild_column_mapping_ui()
            logger.info(f"Moved column '{internal_name}' up to position {idx}")

    def move_column_down(self, internal_name):
        """Move a column down in the order"""
        if internal_name not in self.output_column_order:
            return
        idx = self.output_column_order.index(internal_name)
        if idx < len(self.output_column_order) - 1:
            # Swap with next
            self.output_column_order[idx], self.output_column_order[idx + 1] = \
                self.output_column_order[idx + 1], self.output_column_order[idx]
            self.rebuild_column_mapping_ui()
            logger.info(f"Moved column '{internal_name}' down to position {idx + 2}")

    def update_output_column_name(self, internal_name, new_name):
        """Update the output column mapping when user changes a name"""
        self.output_column_mapping[internal_name] = new_name

    def update_column_visibility(self, col_name, state):
        """Save column visibility setting to database"""
        is_visible = state == 2  # Qt.Checked = 2
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                      (f'export_col_visible_{col_name}', str(is_visible)))
            conn.commit()
            conn.close()
            logger.info(f"Column visibility updated: {col_name} = {is_visible}")
            self.bottom_status.setText(f"{col_name} export visibility: {'visible' if is_visible else 'hidden'}")
        except Exception as e:
            logger.error(f"Failed to save column visibility: {e}")

    def update_split_by_invoice_setting(self, state):
        """Save split by invoice setting to database"""
        self.split_by_invoice = state == 2  # Qt.Checked = 2
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                      ('export_split_by_invoice', str(self.split_by_invoice)))
            conn.commit()
            conn.close()
            logger.info(f"Split by invoice setting updated: {self.split_by_invoice}")
            self.bottom_status.setText(f"Split by invoice: {'enabled' if self.split_by_invoice else 'disabled'}")
        except Exception as e:
            logger.error(f"Failed to save split by invoice setting: {e}")

    def pick_output_font_color(self):
        """Open color picker for output font color (per-user setting)"""
        color = QColorDialog.getColor(QColor(self.output_font_color), self, "Choose Export Font Color")
        if color.isValid():
            color_hex = color.name()
            self.output_font_color = color_hex
            if is_widget_valid(self.output_font_color_btn):
                self.output_font_color_btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {color_hex};
                        border: 1px solid #555;
                        border-radius: 3px;
                    }}
                    QPushButton:hover {{
                        border: 2px solid #888;
                    }}
                """)
            # Save to per-user settings
            set_user_setting('output_font_color', color_hex)
            logger.info(f"Saved output font color: {color_hex}")
            self.bottom_status.setText(f"Output font color set to {color_hex}")

    def reset_output_mapping(self):
        """Reset output column mapping to default values and order"""
        self.output_column_mapping = self.default_output_columns.copy()
        self.output_column_order = self.default_output_column_order.copy()
        # Rebuild the UI with default order
        if hasattr(self, 'column_mapping_scroll_layout'):
            self.rebuild_column_mapping_ui()
        self.bottom_status.setText("Output mapping reset to default")
        logger.info("Output column mapping reset to default")

    def load_output_mapping_profiles(self):
        """Load output mapping profiles from database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("SELECT profile_name FROM output_column_mappings ORDER BY created_date DESC", conn)
            conn.close()

            profile_names = df['profile_name'].tolist()

            # Update the Configuration dialog combo box (if still valid)
            if hasattr(self, 'output_profile_combo') and is_widget_valid(self.output_profile_combo):
                self.output_profile_combo.blockSignals(True)
                self.output_profile_combo.clear()
                self.output_profile_combo.addItem("-- Select Profile --")
                for name in profile_names:
                    self.output_profile_combo.addItem(name)
                self.output_profile_combo.blockSignals(False)

            logger.info(f"Loaded {len(df)} output mapping profiles")
        except Exception as e:
            logger.error(f"Failed to load output mapping profiles: {e}")

    def on_quick_export_profile_changed(self, profile_name):
        """Handle export profile change from Process Shipment tab"""
        if not profile_name or profile_name == "-- Select Profile --":
            return

        # Load the profile (reuse existing method)
        self.load_output_mapping_profile(profile_name)

        # Sync the Configuration dialog combo if it exists and is still valid
        if hasattr(self, 'output_profile_combo') and is_widget_valid(self.output_profile_combo):
            self.output_profile_combo.blockSignals(True)
            self.output_profile_combo.setCurrentText(profile_name)
            self.output_profile_combo.blockSignals(False)

        self.bottom_status.setText(f"Export profile: {profile_name}")

    def load_output_mapping_profile(self, profile_name):
        """Load selected output mapping profile"""
        if not profile_name or profile_name == "-- Select Profile --":
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT mapping_json FROM output_column_mappings WHERE profile_name=?", (profile_name,))
            row = c.fetchone()
            conn.close()

            if row:
                profile_data = json.loads(row[0])

                # Check if this is the new format (with nested structure) or old format (just column mapping)
                if 'column_mapping' in profile_data:
                    # New format
                    self.output_column_mapping = profile_data.get('column_mapping', {})
                    self.output_column_order = profile_data.get('column_order', self.default_output_column_order)
                    column_visibility = profile_data.get('column_visibility', {})
                    split_by_invoice = profile_data.get('split_by_invoice', False)
                else:
                    # Old format - just column mapping
                    self.output_column_mapping = profile_data
                    self.output_column_order = self.default_output_column_order.copy()
                    column_visibility = {}
                    split_by_invoice = False

                # Rebuild column mapping UI with new order (if Configuration dialog is still open)
                if hasattr(self, 'column_mapping_scroll_layout'):
                    self.rebuild_column_mapping_ui()

                # Update column visibility checkboxes (if Configuration dialog is still open)
                if hasattr(self, 'output_column_visibility'):
                    for col_name, checkbox in self.output_column_visibility.items():
                        if is_widget_valid(checkbox):
                            checkbox.blockSignals(True)
                            # Default to True if not specified in profile
                            is_visible = column_visibility.get(col_name, True)
                            checkbox.setChecked(is_visible)
                            checkbox.blockSignals(False)
                        # Save to database regardless of widget state
                        is_visible = column_visibility.get(col_name, True)
                        self.update_column_visibility(col_name, 2 if is_visible else 0)

                # Update split by invoice checkbox (if Configuration dialog is still open)
                if hasattr(self, 'split_by_invoice_checkbox') and is_widget_valid(self.split_by_invoice_checkbox):
                    self.split_by_invoice_checkbox.blockSignals(True)
                    self.split_by_invoice_checkbox.setChecked(split_by_invoice)
                    self.split_by_invoice_checkbox.blockSignals(False)
                self.split_by_invoice = split_by_invoice
                # Save to database regardless of widget state
                self.update_split_by_invoice_setting(2 if split_by_invoice else 0)

                self.bottom_status.setText(f"Loaded output mapping profile: {profile_name}")
                logger.info(f"Loaded output mapping profile: {profile_name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load profile:\n{e}")
            logger.error(f"Failed to load output mapping profile: {e}")

    def save_output_mapping_profile(self):
        """Save current output column mapping as a named profile"""
        name, ok = QInputDialog.getText(self, "Save Output Mapping Profile", "Enter profile name:")
        if not ok or not name.strip():
            return
        name = name.strip()

        # Check if profile exists
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT profile_name FROM output_column_mappings WHERE profile_name=?", (name,))
            exists = c.fetchone()

            if exists:
                reply = QMessageBox.question(self, "Overwrite?", f"Profile '{name}' exists. Overwrite?")
                if reply != QMessageBox.Yes:
                    conn.close()
                    return

            # Build column visibility dict from checkboxes
            column_visibility = {}
            if hasattr(self, 'output_column_visibility'):
                for col_name, checkbox in self.output_column_visibility.items():
                    column_visibility[col_name] = checkbox.isChecked()

            # Get split by invoice setting
            split_by_invoice = self.split_by_invoice if hasattr(self, 'split_by_invoice') else False

            # Get column order
            column_order = self.output_column_order if hasattr(self, 'output_column_order') else self.default_output_column_order

            # Save all settings in new format
            profile_data = {
                'column_mapping': self.output_column_mapping,
                'column_order': column_order,
                'column_visibility': column_visibility,
                'split_by_invoice': split_by_invoice
            }
            mapping_str = json.dumps(profile_data)
            now = datetime.now().isoformat()

            c.execute("""INSERT OR REPLACE INTO output_column_mappings (profile_name, mapping_json, created_date)
                         VALUES (?,?,?)""", (name, mapping_str, now))
            conn.commit()
            conn.close()

            self.load_output_mapping_profiles()
            if is_widget_valid(self.output_profile_combo):
                self.output_profile_combo.setCurrentText(name)
            # Also refresh the linked export combo in Invoice Mapping tab
            self.refresh_linked_export_combo()
            self.bottom_status.setText(f"Saved output mapping profile: {name}")
            logger.info(f"Saved output mapping profile: {name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save profile:\n{e}")
            logger.error(f"Failed to save output mapping profile: {e}")

    def update_output_mapping_profile(self):
        """Update the currently selected output mapping profile with current settings"""
        if not is_widget_valid(self.output_profile_combo):
            return
        profile_name = self.output_profile_combo.currentText()
        if not profile_name or profile_name == "-- Select Profile --":
            QMessageBox.information(self, "No Profile Selected", "Please select a profile to update.")
            return

        reply = QMessageBox.question(self, "Confirm Update",
                                     f"Update profile '{profile_name}' with current settings?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Build column visibility dict from checkboxes
            column_visibility = {}
            if hasattr(self, 'output_column_visibility'):
                for col_name, checkbox in self.output_column_visibility.items():
                    column_visibility[col_name] = checkbox.isChecked()

            # Get split by invoice setting
            split_by_invoice = self.split_by_invoice if hasattr(self, 'split_by_invoice') else False

            # Get column order
            column_order = self.output_column_order if hasattr(self, 'output_column_order') else self.default_output_column_order

            # Save all settings in new format
            profile_data = {
                'column_mapping': self.output_column_mapping,
                'column_order': column_order,
                'column_visibility': column_visibility,
                'split_by_invoice': split_by_invoice
            }
            mapping_str = json.dumps(profile_data)
            now = datetime.now().isoformat()

            c.execute("""UPDATE output_column_mappings SET mapping_json=?, created_date=?
                         WHERE profile_name=?""", (mapping_str, now, profile_name))
            conn.commit()
            conn.close()

            self.bottom_status.setText(f"Updated output mapping profile: {profile_name}")
            logger.info(f"Updated output mapping profile: {profile_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update profile:\n{e}")
            logger.error(f"Failed to update output mapping profile: {e}")

    def delete_output_mapping_profile(self):
        """Delete selected output mapping profile"""
        if not is_widget_valid(self.output_profile_combo):
            return
        profile_name = self.output_profile_combo.currentText()
        if not profile_name or profile_name == "-- Select Profile --":
            QMessageBox.information(self, "No Profile Selected", "Please select a profile to delete.")
            return

        reply = QMessageBox.question(self, "Confirm Delete",
                                     f"Delete profile '{profile_name}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM output_column_mappings WHERE profile_name=?", (profile_name,))
            conn.commit()
            conn.close()

            self.load_output_mapping_profiles()
            # Also refresh the linked export combo in Invoice Mapping tab
            self.refresh_linked_export_combo()
            self.bottom_status.setText(f"Deleted output mapping profile: {profile_name}")
            logger.info(f"Deleted output mapping profile: {profile_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete profile:\n{e}")
            logger.error(f"Failed to delete output mapping profile: {e}")

    def load_csv_for_shipment_mapping(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Invoice File",
            str(INPUT_DIR),
            "All Supported (*.csv *.xlsx);;CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        if not path: return
        try:
            # Get header row value from input field
            header_row = 0  # Default: first row is header
            if hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
                try:
                    header_row_value = int(self.header_row_input.text().strip())
                    # Convert from 1-based to 0-based indexing
                    header_row = max(0, header_row_value - 1)
                except ValueError:
                    header_row = 0

            # Determine file type and extract data accordingly
            file_ext = Path(path).suffix.lower()

            if file_ext == '.pdf':
                # Extract using pdfplumber table extraction
                df = self.extract_pdf_table(path)
                logger.info(f"PDF detected: {Path(path).name} - using table extraction")
                # Move PDF to processed folder after successful extraction
                self.move_pdf_to_processed(path)
            elif file_ext == '.xlsx':
                df = pd.read_excel(path, nrows=0, dtype=str, header=header_row)
            else:  # .csv
                df = pd.read_csv(path, nrows=0, dtype=str, header=header_row)

            cols = list(df.columns)

            # Clear existing labels
            for label in self.shipment_drag_labels:
                label.setParent(None)
            self.shipment_drag_labels = []

            # Add new labels from extracted columns to the scroll layout
            for col in cols:
                lbl = DraggableLabel(str(col))
                # Insert before the stretch at the end
                self.left_scroll_layout.insertWidget(self.left_scroll_layout.count()-1, lbl)
                self.shipment_drag_labels.append(lbl)

            # Determine file type for status message
            file_type = "PDF" if file_ext == '.pdf' else ("Excel" if file_ext == '.xlsx' else "CSV")
            logger.info(f"{file_type} file loaded for mapping: {Path(path).name}")
            self.status.setText(f"{file_type} file loaded: {Path(path).name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot read file:\n{e}")
            logger.error(f"File loading failed: {str(e)}")

    def extract_pdf_table(self, pdf_path):
        """
        Extract tabular data from PDF invoices using pdfplumber.

        Attempts to find and extract the first valid table from the PDF.
        If no table found, tries text-based extraction as fallback.
        Returns a DataFrame with the extracted data.

        Args:
            pdf_path (str): Path to PDF file

        Returns:
            pd.DataFrame: DataFrame with extracted table data

        Raises:
            Exception: If PDF cannot be processed or no table found
        """
        raise Exception("PDF support for invoice extraction has been removed in this version. Please use CSV or Excel files.")

        # PDF support removed; all code below is unreachable

    def _extract_pdf_text_fallback(self, pdf_path):
        """
        Fallback method to extract text-based data from PDFs without tables.

        Attempts supplier-specific extraction (e.g., AROMATE invoices).
        Falls back to generic text extraction if no supplier pattern matches.

        Args:
            pdf_path (str): Path to PDF file

        Returns:
            pd.DataFrame: DataFrame with extracted data or text lines

        Raises:
            Exception: If PDF cannot be read
        """
        raise Exception("PDF support for invoice extraction has been removed in this version. Please use CSV or Excel files.")
        # PDF support removed; all code below is unreachable

    def _extract_aromate_invoice(self, text):
        """
        Extract AROMATE invoice data using regex pattern matching.

        Extracts SKU, quantity, unit price, and total price from AROMATE invoices.

        Args:
            text (str): Extracted PDF text

        Returns:
            pd.DataFrame: DataFrame with part_number, quantity, unit_price, total_price
        """
        import re

        # Pattern matches both formats:
        # Format 1: SKU# 1562485 76,080 PCS USD 0.6580 USD 50,060.64
        # Format 2: SKU# 2641486 15,120 PCS 0.7140 10,795.68
        pattern = r'SKU#\s*(\d+)\s+(\d+(?:,\d{3})*)\s+PCS\s+(?:USD\s+)?([\d.]+)\s+(?:USD\s+)?([\d,]+\.\d{2})'

        matches = re.findall(pattern, text)

        if not matches:
            logger.warning("AROMATE pattern not found, falling back to generic text extraction")
            all_text = [line.strip() for line in text.split('\n') if line.strip()]
            return pd.DataFrame({'text_line': all_text})

        # Convert matches to DataFrame
        data = []
        for sku, qty, unit_price, total_price in matches:
            data.append({
                'part_number': sku,
                'quantity': int(qty.replace(',', '')),
                'unit_price': float(unit_price),
                'total_price': float(total_price.replace(',', ''))
            })

        df = pd.DataFrame(data)
        logger.info(f"AROMATE invoice extraction successful: {len(df)} items extracted")
        return df

    def on_shipment_drop(self, field_key, column_name):
        # Check if shipment_targets is valid
        if hasattr(self, 'shipment_targets') and self.shipment_targets:
            try:
                for k, t in self.shipment_targets.items():
                    if t and hasattr(t, 'setText'):
                        if t.column_name == column_name and k != field_key:
                            t.column_name = None
                            t.setText(f"Drop {t.field_key} here")
                            t.setProperty("occupied", False)
                            t.style().unpolish(t); t.style().polish(t)
            except (RuntimeError, AttributeError):
                pass  # Widgets have been deleted

        self.shipment_mapping[field_key] = column_name
        SHIPMENT_MAPPING_FILE.write_text(json.dumps(self.shipment_mapping, indent=2))
        logger.info(f"Shipment mapping saved: {field_key} to {column_name}")

    def reset_current_mapping(self):
        self.shipment_mapping = {}

        # Clear drop targets (right side) - only if they exist and are valid
        if hasattr(self, 'shipment_targets') and self.shipment_targets:
            try:
                for target in self.shipment_targets.values():
                    if target and hasattr(target, 'setText'):
                        target.column_name = None
                        target.setText(f"Drop {target.field_key} here")
                        target.setProperty("occupied", False)
                        target.style().unpolish(target); target.style().polish(target)
            except (RuntimeError, AttributeError):
                pass  # Widgets have been deleted

        # Clear CSV columns drag labels (left side) - only if they exist and are valid
        if hasattr(self, 'shipment_drag_labels') and self.shipment_drag_labels:
            try:
                for label in self.shipment_drag_labels:
                    if label and hasattr(label, 'setParent'):
                        label.setParent(None)
                        label.deleteLater()
                self.shipment_drag_labels = []
            except (RuntimeError, AttributeError):
                pass  # Widgets have been deleted

        self.status.setText("Current mapping reset")

    def load_mapping_profiles(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("SELECT profile_name FROM mapping_profiles ORDER BY created_date DESC", conn)
            conn.close()
            
            # Update Process tab combo (tab 0 - always initialized)
            if hasattr(self, 'profile_combo'):
                self.profile_combo.blockSignals(True)
                self.profile_combo.clear()
                self.profile_combo.addItem("-- Select Profile --")
                for name in df['profile_name'].tolist():
                    self.profile_combo.addItem(name)
                self.profile_combo.blockSignals(False)
            
            # Update Invoice Mapping Profiles tab combo (tab 1 - may not be initialized yet)
            if hasattr(self, 'profile_combo_map'):
                self.profile_combo_map.blockSignals(True)
                self.profile_combo_map.clear()
                self.profile_combo_map.addItem("-- Select Profile --")
                for name in df['profile_name'].tolist():
                    self.profile_combo_map.addItem(name)
                self.profile_combo_map.blockSignals(False)
                
            logger.info(f"Loaded {len(df)} mapping profiles")
        except Exception as e:
            logger.error(f"Failed to load mapping profiles: {e}")

    def save_mapping_profile(self):
        name, ok = QInputDialog.getText(self, "Save Mapping Profile", "Enter profile name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        if self.profile_combo.findText(name) != -1:
            if QMessageBox.question(self, "Overwrite?", f"Profile '{name}' exists. Overwrite?") != QMessageBox.Yes:
                return

        mapping_str = json.dumps(self.shipment_mapping)

        # Get header row value from input field
        header_row_value = 1  # Default
        if hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
            try:
                header_row_value = int(self.header_row_input.text().strip())
            except ValueError:
                header_row_value = 1

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO mapping_profiles (profile_name, mapping_json, header_row) VALUES (?, ?, ?)",
                      (name, mapping_str, header_row_value))
            conn.commit()
            conn.close()
            self.load_mapping_profiles()
            # Only update the combo on the Invoice Mapping Profiles tab (where save button is)
            self.profile_combo_map.setCurrentText(name)
            logger.success(f"Mapping profile saved: {name} (header_row={header_row_value})")
            self.status.setText(f"Profile saved: {name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save failed: {e}")

    def load_selected_profile(self, name):
        if not name or name == "-- Select Profile --":
            return
        self.load_selected_profile_full(name)

    def load_selected_profile_full(self, name):
        if not name or name == "-- Select Profile --":
            # Clear all mappings and reset UI
            self.shipment_mapping = {}
            self.profile_header_row = 1  # Reset to default

            # Clear drop targets (only if they exist and are valid)
            if hasattr(self, 'shipment_targets') and self.shipment_targets:
                try:
                    for target in self.shipment_targets.values():
                        if target and hasattr(target, 'setText'):
                            target.column_name = None
                            target.setText(f"Drop {target.field_key.replace('_', ' ')} here")
                            target.setProperty("occupied", False)
                            target.style().unpolish(target)
                            target.style().polish(target)
                except (RuntimeError, AttributeError):
                    pass  # Widgets have been deleted

            # Clear draggable CSV columns (only if they exist and are valid)
            if hasattr(self, 'shipment_drag_labels') and self.shipment_drag_labels:
                try:
                    for label in self.shipment_drag_labels:
                        if label and hasattr(label, 'deleteLater'):
                            label.deleteLater()
                    self.shipment_drag_labels.clear()
                except (RuntimeError, AttributeError):
                    pass  # Widgets have been deleted

            # Reset header row to default (only if widget exists and is valid)
            if hasattr(self, 'header_row_input') and self.header_row_input:
                try:
                    if hasattr(self.header_row_input, 'setText'):
                        self.header_row_input.setText("1")
                except (RuntimeError, AttributeError):
                    pass  # Widget has been deleted

            self.bottom_status.setText("Profile cleared")
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT mapping_json, header_row FROM mapping_profiles WHERE profile_name = ?", (name,))
            row = c.fetchone()
            conn.close()
            if row:
                self.shipment_mapping = json.loads(row[0])
                # Restore header row value
                header_row_value = row[1] if len(row) > 1 and row[1] is not None else 1
                # Store as instance variable for use in Process Shipments tab
                self.profile_header_row = header_row_value
                if hasattr(self, 'header_row_input') and self.header_row_input:
                    try:
                        if hasattr(self.header_row_input, 'setText'):
                            self.header_row_input.setText(str(header_row_value))
                    except (RuntimeError, AttributeError):
                        pass  # Widget has been deleted
                self.apply_current_mapping()
                logger.info(f"Profile loaded: {name} (header_row={header_row_value})")
                self.bottom_status.setText(f"Loaded profile: {name}")
                # Save as last used map profile
                set_db_config('last_map_profile', name)
                # Load linked export profile
                self.load_profile_link(name)
                # Apply linked export profile settings
                self.apply_linked_export_profile(name)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Load failed: {e}")

    def delete_mapping_profile(self):
        # Get profile name from Invoice Mapping Profiles tab combo (where delete button is)
        name = self.profile_combo_map.currentText()
        if not name or name == "-- Select Profile --":
            return
        if QMessageBox.question(self, "Delete", f"Delete profile '{name}'?") != QMessageBox.Yes:
            return
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM mapping_profiles WHERE profile_name = ?", (name,))
            # Also delete any profile link
            c.execute("DELETE FROM profile_links WHERE input_profile_name = ?", (name,))
            conn.commit()
            conn.close()
            self.load_mapping_profiles()
            # Reset both combos to default after deletion
            self.profile_combo.setCurrentIndex(0)
            self.profile_combo_map.setCurrentIndex(0)
            logger.info(f"Profile deleted: {name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Delete failed: {e}")

    def save_profile_link(self):
        """Save link between current input profile and selected export profile"""
        input_profile = self.profile_combo_map.currentText()
        if not input_profile or input_profile == "-- Select Profile --":
            QMessageBox.warning(self, "No Profile", "Please select an input profile first")
            return

        export_profile = self.linked_export_combo.currentText()
        if export_profile == "(None)":
            export_profile = None

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            if export_profile:
                c.execute("INSERT OR REPLACE INTO profile_links (input_profile_name, export_profile_name) VALUES (?, ?)",
                         (input_profile, export_profile))
                logger.info(f"Linked profile '{input_profile}' to export profile '{export_profile}'")
                self.bottom_status.setText(f"Linked to export profile: {export_profile}")
            else:
                c.execute("DELETE FROM profile_links WHERE input_profile_name = ?", (input_profile,))
                logger.info(f"Removed link for profile '{input_profile}'")
                self.bottom_status.setText("Export profile link removed")
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to save profile link: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save link: {e}")

    def clear_profile_link(self):
        """Clear the export profile link for current input profile"""
        input_profile = self.profile_combo_map.currentText()
        if not input_profile or input_profile == "-- Select Profile --":
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM profile_links WHERE input_profile_name = ?", (input_profile,))
            conn.commit()
            conn.close()
            self.linked_export_combo.setCurrentIndex(0)  # Set to "(None)"
            logger.info(f"Cleared link for profile '{input_profile}'")
            self.bottom_status.setText("Export profile link cleared")
        except Exception as e:
            logger.error(f"Failed to clear profile link: {e}")

    def load_profile_link(self, input_profile_name):
        """Load the linked export profile for an input profile"""
        if not hasattr(self, 'linked_export_combo') or not is_widget_valid(self.linked_export_combo):
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT export_profile_name FROM profile_links WHERE input_profile_name = ?", (input_profile_name,))
            row = c.fetchone()
            conn.close()

            if row and row[0]:
                # Find and select the export profile in combo
                idx = self.linked_export_combo.findText(row[0])
                if idx >= 0:
                    self.linked_export_combo.setCurrentIndex(idx)
                else:
                    self.linked_export_combo.setCurrentIndex(0)
            else:
                self.linked_export_combo.setCurrentIndex(0)
        except (RuntimeError, AttributeError):
            pass  # Widget has been deleted
        except Exception as e:
            logger.warning(f"Failed to load profile link: {e}")

    def apply_linked_export_profile(self, input_profile_name):
        """Apply the linked export profile settings when an input profile is loaded"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT export_profile_name FROM profile_links WHERE input_profile_name = ?", (input_profile_name,))
            row = c.fetchone()
            conn.close()

            if row and row[0]:
                export_profile = row[0]
                # Load the export profile settings
                if hasattr(self, 'output_profile_combo') and is_widget_valid(self.output_profile_combo):
                    idx = self.output_profile_combo.findText(export_profile)
                    if idx >= 0:
                        self.output_profile_combo.setCurrentIndex(idx)
                        logger.info(f"Auto-loaded linked export profile: {export_profile}")
        except (RuntimeError, AttributeError):
            pass  # Widget has been deleted
        except Exception as e:
            logger.warning(f"Failed to apply linked export profile: {e}")

    def refresh_linked_export_combo(self):
        """Refresh the linked export profile dropdown with current output profiles"""
        if not hasattr(self, 'linked_export_combo') or not is_widget_valid(self.linked_export_combo):
            return

        try:
            # Remember current selection
            current_text = self.linked_export_combo.currentText()

            # Clear and repopulate
            self.linked_export_combo.clear()
            self.linked_export_combo.addItem("(None)")

            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT DISTINCT profile_name FROM output_column_mappings ORDER BY profile_name")
            for row in c.fetchall():
                self.linked_export_combo.addItem(row[0])
            conn.close()

            # Restore selection if it still exists
            idx = self.linked_export_combo.findText(current_text)
            if idx >= 0:
                self.linked_export_combo.setCurrentIndex(idx)
        except (RuntimeError, AttributeError):
            pass  # Widget has been deleted
        except Exception as e:
            logger.warning(f"Failed to refresh linked export combo: {e}")

    # ========== FOLDER PROFILE FUNCTIONS ==========

    def load_folder_profiles(self):
        """Load folder profiles into the dropdown"""
        if not hasattr(self, 'folder_profile_combo'):
            return

        try:
            current_text = self.folder_profile_combo.currentText()
            self.folder_profile_combo.blockSignals(True)
            self.folder_profile_combo.clear()
            self.folder_profile_combo.addItem("-- Select Folder Profile --")

            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT profile_name FROM folder_profiles ORDER BY profile_name")
            for row in c.fetchall():
                self.folder_profile_combo.addItem(row[0])
            conn.close()

            # Restore selection if it exists
            if current_text and current_text != "-- Select Folder Profile --":
                idx = self.folder_profile_combo.findText(current_text)
                if idx >= 0:
                    self.folder_profile_combo.setCurrentIndex(idx)

            self.folder_profile_combo.blockSignals(False)
        except Exception as e:
            logger.warning(f"Failed to load folder profiles: {e}")
            self.folder_profile_combo.blockSignals(False)

    def load_folder_profile(self, name):
        """Load a folder profile and update INPUT_DIR, OUTPUT_DIR, and PROCESSED_DIR"""
        global INPUT_DIR, OUTPUT_DIR, PROCESSED_DIR

        if not name or name == "-- Select Folder Profile --":
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT input_folder, output_folder FROM folder_profiles WHERE profile_name = ?", (name,))
            row = c.fetchone()
            conn.close()

            if row:
                input_folder, output_folder = row
                if input_folder:
                    input_folder = os.path.normpath(input_folder)
                    INPUT_DIR = Path(input_folder)
                    PROCESSED_DIR = get_processed_dir(INPUT_DIR)
                    PROCESSED_DIR.mkdir(exist_ok=True)
                    set_user_setting('input_folder', input_folder)
                if output_folder:
                    output_folder = os.path.normpath(output_folder)
                    OUTPUT_DIR = Path(output_folder)
                    set_user_setting('output_folder', output_folder)

                # Refresh the input files list
                self.refresh_input_files()
                self.refresh_exported_files()

                # Save as last used folder profile
                set_db_config('last_folder_profile', name)

                logger.info(f"Loaded folder profile: {name}")
                self.bottom_status.setText(f"Folder profile loaded: {name}")
        except Exception as e:
            logger.error(f"Failed to load folder profile: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load folder profile: {e}")

    def restore_last_used_settings(self):
        """Restore last used folder profile and map profile from database on startup"""
        try:
            # Restore last used folder profile
            last_folder_profile = get_db_config('last_folder_profile')
            if last_folder_profile and hasattr(self, 'folder_profile_combo'):
                idx = self.folder_profile_combo.findText(last_folder_profile)
                if idx >= 0:
                    self.folder_profile_combo.setCurrentIndex(idx)
                    logger.info(f"Restored last folder profile: {last_folder_profile}")

            # Restore last used map profile
            last_map_profile = get_db_config('last_map_profile')
            if last_map_profile:
                # Update Process tab combo
                if hasattr(self, 'profile_combo'):
                    idx = self.profile_combo.findText(last_map_profile)
                    if idx >= 0:
                        self.profile_combo.setCurrentIndex(idx)
                # Update Invoice Mapping Profiles tab combo
                if hasattr(self, 'profile_combo_map'):
                    idx = self.profile_combo_map.findText(last_map_profile)
                    if idx >= 0:
                        self.profile_combo_map.setCurrentIndex(idx)
                logger.info(f"Restored last map profile: {last_map_profile}")

        except Exception as e:
            logger.warning(f"Failed to restore last used settings: {e}")

    # ========== BILLING FUNCTIONS ==========

    def record_billing_event(self, file_name: str, line_count: int, total_value: float,
                            hts_codes: list, processing_time_ms: int = 0):
        """Record a billing event when a file is exported.

        Only records if the file number has not been billed before (allows re-exports
        without additional billing charges).

        Args:
            file_name: Name of the exported file
            line_count: Number of line items in the export
            total_value: Total value of items processed
            hts_codes: List of HTS codes used in the export
            processing_time_ms: Processing time in milliseconds
        """
        try:
            import os
            import getpass

            file_number = self.file_number_input.text().strip() if hasattr(self, 'file_number_input') else ""

            # Skip if no file number provided
            if not file_number:
                logger.debug("No file number provided, skipping billing record")
                return

            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Check if this file number has already been billed
            c.execute("SELECT COUNT(*) FROM billing_records WHERE file_number = ?", (file_number,))
            existing_count = c.fetchone()[0]

            if existing_count > 0:
                conn.close()
                logger.info(f"File #{file_number} already billed, skipping duplicate billing record")
                return

            folder_profile = self.folder_profile_combo.currentText() if hasattr(self, 'folder_profile_combo') else ""
            map_profile = self.profile_combo.currentText() if hasattr(self, 'profile_combo') else ""
            mid = self.selected_mid if hasattr(self, 'selected_mid') else ""
            user_name = getpass.getuser()

            # Get machine ID
            machine_id = ""
            try:
                import hashlib
                import platform
                machine_info = f"{platform.node()}-{platform.machine()}"
                machine_id = hashlib.md5(machine_info.encode()).hexdigest()[:12]
            except:
                pass

            now = datetime.now()
            export_date = now.strftime("%Y-%m-%d")
            export_time = now.strftime("%H:%M:%S")
            invoice_month = now.strftime("%Y-%m")

            # Convert HTS codes list to comma-separated string
            hts_codes_str = ",".join(set(str(h) for h in hts_codes if h)) if hts_codes else ""

            c.execute("""INSERT INTO billing_records
                        (file_number, export_date, export_time, file_name, line_count, total_value,
                         hts_codes_used, folder_profile, map_profile, mid, user_name, machine_id,
                         processing_time_ms, invoice_sent, invoice_month)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)""",
                     (file_number, export_date, export_time, file_name, line_count, total_value,
                      hts_codes_str, folder_profile, map_profile, mid, user_name, machine_id,
                      processing_time_ms, invoice_month))
            conn.commit()
            conn.close()

            logger.info(f"Billing record created: File #{file_number}, {line_count} lines, ${total_value:,.2f}")

        except Exception as e:
            logger.warning(f"Failed to record billing event: {e}")

    def get_billing_summary(self, month: str = None):
        """Get billing summary for a specific month.

        Args:
            month: Month in YYYY-MM format, defaults to current month

        Returns:
            dict with billing summary data
        """
        try:
            if month is None:
                month = datetime.now().strftime("%Y-%m")

            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Get summary statistics
            c.execute("""SELECT
                            COUNT(*) as export_count,
                            SUM(line_count) as total_lines,
                            SUM(total_value) as total_value,
                            COUNT(DISTINCT file_number) as unique_files
                        FROM billing_records
                        WHERE invoice_month = ?""", (month,))
            row = c.fetchone()

            # Get detailed records
            c.execute("""SELECT * FROM billing_records
                        WHERE invoice_month = ?
                        ORDER BY export_date, export_time""", (month,))
            records = c.fetchall()
            columns = [desc[0] for desc in c.description]

            conn.close()

            return {
                'month': month,
                'export_count': row[0] or 0,
                'total_lines': row[1] or 0,
                'total_value': row[2] or 0.0,
                'unique_files': row[3] or 0,
                'records': [dict(zip(columns, r)) for r in records]
            }

        except Exception as e:
            logger.warning(f"Failed to get billing summary: {e}")
            return {'month': month, 'export_count': 0, 'total_lines': 0, 'total_value': 0.0, 'unique_files': 0, 'records': []}

    def get_billing_setting(self, key: str, default: str = ""):
        """Get a billing setting value."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM billing_settings WHERE key = ?", (key,))
            row = c.fetchone()
            conn.close()
            return row[0] if row else default
        except:
            return default

    def set_billing_setting(self, key: str, value: str):
        """Set a billing setting value."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO billing_settings (key, value) VALUES (?, ?)", (key, value))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logger.warning(f"Failed to set billing setting: {e}")
            return False

    def _is_billing_admin(self) -> bool:
        """Check if the current user has admin role (can access billing settings)."""
        try:
            # Primary check: Use authentication manager
            if hasattr(self, 'auth_manager') and self.auth_manager:
                return self.auth_manager.is_admin()

            # Fallback for legacy: check specific admin email
            ADMIN_EMAIL = "admin@processlogiclabs.com"
            if hasattr(self, 'license_mgr') and self.license_mgr:
                if self.license_mgr.license_email:
                    return self.license_mgr.license_email.lower() == ADMIN_EMAIL.lower()
        except Exception as e:
            logger.warning(f"Error checking billing admin status: {e}")
        return False

    def _update_account_menu(self):
        """Update the Account menu to show current user."""
        if hasattr(self, 'account_user_action') and hasattr(self, 'auth_manager'):
            if self.auth_manager and self.auth_manager.is_authenticated:
                user_display = self.auth_manager.current_user
                role_badge = " [Admin]" if self.auth_manager.current_role == 'admin' else ""
                self.account_user_action.setText(f"Signed in as: {user_display}{role_badge}")
            else:
                self.account_user_action.setText("Not signed in")

    def _sign_out(self):
        """Sign out current user and show login dialog."""
        if hasattr(self, 'auth_manager') and self.auth_manager:
            # Confirm sign out
            reply = QMessageBox.question(
                self, "Sign Out",
                "Are you sure you want to sign out?\n\nThe application will close and you'll need to sign in again.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                self.auth_manager.logout()
                logger.info("User signed out, closing application")
                # Close the application - user will need to restart and log in again
                self.close()

    def _fetch_remote_billing_config(self, customer_email: str) -> dict:
        """Fetch billing configuration from remote GitHub-hosted JSON.

        Expected JSON format at GitHub:
        {
            "customers": {
                "customer@example.com": {
                    "rate_per_file": "2.50",
                    "billing_enabled": true,
                    "admin_email": "admin@processlogiclabs.com"
                }
            },
            "default": {
                "rate_per_file": "0",
                "billing_enabled": false
            }
        }
        """
        import urllib.request
        import urllib.error

        BILLING_CONFIG_URL = "https://raw.githubusercontent.com/ProcessLogicLabs/TariffMill/main/billing_config.json"

        try:
            req = urllib.request.Request(
                BILLING_CONFIG_URL,
                headers={'User-Agent': 'TariffMill/1.0'}
            )
            with urllib.request.urlopen(req, timeout=10) as response:
                data = json.loads(response.read().decode('utf-8'))

                # Look up customer-specific config
                customers = data.get('customers', {})
                if customer_email and customer_email.lower() in {k.lower() for k in customers}:
                    # Find the matching key (case-insensitive)
                    for key in customers:
                        if key.lower() == customer_email.lower():
                            logger.info(f"Found remote billing config for {customer_email}")
                            return customers[key]

                # Return default config if no customer-specific config found
                return data.get('default', {})

        except urllib.error.URLError as e:
            logger.warning(f"Failed to fetch remote billing config: {e}")
            return {}
        except json.JSONDecodeError as e:
            logger.warning(f"Invalid billing config JSON: {e}")
            return {}
        except Exception as e:
            logger.warning(f"Error fetching remote billing config: {e}")
            return {}

    def _sync_billing_from_remote(self):
        """Sync billing settings from remote configuration."""
        # Get customer email - prefer license email, fallback to stored customer email
        customer_email = None
        if hasattr(self, 'license_mgr') and self.license_mgr and self.license_mgr.license_email:
            customer_email = self.license_mgr.license_email
        else:
            customer_email = self.get_billing_setting('customer_email', '')

        if not customer_email:
            logger.debug("No customer email configured, skipping remote billing sync")
            return

        remote_config = self._fetch_remote_billing_config(customer_email)
        if remote_config:
            # Apply remote settings (admin can override locally)
            if 'rate_per_file' in remote_config:
                self.set_billing_setting('rate_per_file', str(remote_config['rate_per_file']))
                logger.info(f"Updated billing rate from remote: {remote_config['rate_per_file']}")
            if 'admin_email' in remote_config:
                self.set_billing_setting('admin_email', remote_config['admin_email'])
            if 'billing_enabled' in remote_config:
                self.set_billing_setting('billing_enabled', str(remote_config['billing_enabled']))

            # Store last sync time
            self.set_billing_setting('last_remote_sync', datetime.now().isoformat())
            logger.info("Billing settings synced from remote configuration")

    def setup_billing_tab(self, tab_widget):
        """Setup the Billing tab for the Configuration dialog."""
        layout = QVBoxLayout(tab_widget)

        # Title
        title = QLabel("<h2>Billing Configuration</h2>")
        layout.addWidget(title)

        # Billing Rate Group
        rate_group = QGroupBox("Billing Rate")
        rate_layout = QFormLayout()

        self.billing_rate_spin = QDoubleSpinBox()
        self.billing_rate_spin.setRange(0, 1000)
        self.billing_rate_spin.setDecimals(2)
        self.billing_rate_spin.setPrefix("$")
        self.billing_rate_spin.setValue(float(self.get_billing_setting('rate_per_file', '0')))
        rate_layout.addRow("Rate per Export:", self.billing_rate_spin)

        rate_info = QLabel("<small>Amount to charge per exported file. Set to 0 to disable billing.</small>")
        rate_info.setWordWrap(True)
        rate_info.setStyleSheet("color:#666; padding:5px;")
        rate_layout.addRow("", rate_info)

        rate_group.setLayout(rate_layout)
        layout.addWidget(rate_group)

        # Email Configuration Group
        email_group = QGroupBox("Email Configuration")
        email_layout = QFormLayout()

        self.billing_customer_email = QLineEdit()
        self.billing_customer_email.setPlaceholderText("customer@example.com")
        self.billing_customer_email.setText(self.get_billing_setting('customer_email', ''))
        email_layout.addRow("Customer Email:", self.billing_customer_email)

        self.billing_admin_email = QLineEdit()
        self.billing_admin_email.setPlaceholderText("your@email.com")
        self.billing_admin_email.setText(self.get_billing_setting('admin_email', ''))
        email_layout.addRow("Admin Email:", self.billing_admin_email)

        self.billing_smtp_server = QLineEdit()
        self.billing_smtp_server.setPlaceholderText("smtp.gmail.com")
        self.billing_smtp_server.setText(self.get_billing_setting('smtp_server', ''))
        email_layout.addRow("SMTP Server:", self.billing_smtp_server)

        self.billing_smtp_port = QSpinBox()
        self.billing_smtp_port.setRange(1, 65535)
        self.billing_smtp_port.setValue(int(self.get_billing_setting('smtp_port', '587')))
        email_layout.addRow("SMTP Port:", self.billing_smtp_port)

        self.billing_smtp_user = QLineEdit()
        self.billing_smtp_user.setPlaceholderText("username@gmail.com")
        self.billing_smtp_user.setText(self.get_billing_setting('smtp_user', ''))
        email_layout.addRow("SMTP Username:", self.billing_smtp_user)

        self.billing_smtp_password = QLineEdit()
        self.billing_smtp_password.setEchoMode(QLineEdit.Password)
        self.billing_smtp_password.setText(self.get_billing_setting('smtp_password', ''))
        email_layout.addRow("SMTP Password:", self.billing_smtp_password)

        email_info = QLabel("<small>Configure SMTP settings to enable automatic monthly billing reports.</small>")
        email_info.setWordWrap(True)
        email_info.setStyleSheet("color:#666; padding:5px;")
        email_layout.addRow("", email_info)

        email_group.setLayout(email_layout)
        layout.addWidget(email_group)

        # Buttons row
        btn_layout = QHBoxLayout()

        btn_test_email = QPushButton("Test Email")
        btn_test_email.clicked.connect(self._test_billing_email_from_tab)
        btn_layout.addWidget(btn_test_email)

        btn_layout.addStretch()

        btn_view_report = QPushButton("View Billing Report")
        btn_view_report.clicked.connect(self.show_billing_report_dialog)
        btn_layout.addWidget(btn_view_report)

        btn_save = QPushButton("Save Settings")
        btn_save.setStyleSheet(self.get_button_style("primary"))
        btn_save.clicked.connect(self._save_billing_settings_from_tab)
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

        # User Management Section
        user_group = QGroupBox("User Management")
        user_layout = QVBoxLayout()

        # User table
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(4)
        self.user_table.setHorizontalHeaderLabels(["Email", "Name", "Role", "Actions"])
        self.user_table.horizontalHeader().setStretchLastSection(True)
        self.user_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.user_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.user_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.user_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.user_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.user_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.user_table.setMinimumHeight(150)
        user_layout.addWidget(self.user_table)

        # User management buttons
        user_btn_layout = QHBoxLayout()

        btn_refresh_users = QPushButton("Refresh")
        btn_refresh_users.clicked.connect(self._refresh_user_list)
        user_btn_layout.addWidget(btn_refresh_users)

        btn_add_user = QPushButton("Add User")
        btn_add_user.setStyleSheet(self.get_button_style("primary"))
        btn_add_user.clicked.connect(self._add_user_dialog)
        user_btn_layout.addWidget(btn_add_user)

        user_btn_layout.addStretch()

        user_info = QLabel("<small>Changes are synced to the private GitHub config repository.</small>")
        user_info.setStyleSheet("color:#666;")
        user_btn_layout.addWidget(user_info)

        user_layout.addLayout(user_btn_layout)
        user_group.setLayout(user_layout)
        layout.addWidget(user_group)

        # Load users on tab creation
        QTimer.singleShot(100, self._refresh_user_list)

        layout.addStretch()

    def _save_billing_settings_from_tab(self):
        """Save billing settings from the Configuration dialog tab."""
        self.set_billing_setting('rate_per_file', str(self.billing_rate_spin.value()))
        self.set_billing_setting('customer_email', self.billing_customer_email.text())
        self.set_billing_setting('admin_email', self.billing_admin_email.text())
        self.set_billing_setting('smtp_server', self.billing_smtp_server.text())
        self.set_billing_setting('smtp_port', str(self.billing_smtp_port.value()))
        self.set_billing_setting('smtp_user', self.billing_smtp_user.text())
        self.set_billing_setting('smtp_password', self.billing_smtp_password.text())
        QMessageBox.information(self, "Saved", "Billing settings saved successfully.")

    def _test_billing_email_from_tab(self):
        """Test email configuration from the Configuration dialog tab."""
        self._test_billing_email(
            self.billing_smtp_server.text(),
            self.billing_smtp_port.value(),
            self.billing_smtp_user.text(),
            self.billing_smtp_password.text(),
            self.billing_admin_email.text()
        )

    # =========================================================================
    # User Management Methods
    # =========================================================================

    def _refresh_user_list(self):
        """Refresh the user list table from local auth file."""
        if not hasattr(self, 'user_table'):
            return

        self.user_table.setRowCount(0)

        # Load from local auth file
        users = self._load_auth_users()
        if not users:
            return

        for user_key, data in users.items():
            row = self.user_table.rowCount()
            self.user_table.insertRow(row)

            # Determine if this is a Windows user
            is_windows_user = '\\' in user_key or data.get('auth_type') == 'windows'

            # User identifier (email or DOMAIN\username)
            user_item = QTableWidgetItem(user_key)
            if is_windows_user:
                user_item.setToolTip("Windows Domain User - Auto-login enabled")
                user_item.setForeground(QColor("#28a745"))  # Green for Windows users
            self.user_table.setItem(row, 0, user_item)

            # Name
            self.user_table.setItem(row, 1, QTableWidgetItem(data.get('name', '')))

            # Role with auth type indicator
            role_text = data.get('role', 'user')
            if is_windows_user:
                role_text += " (Win)"
            role_item = QTableWidgetItem(role_text)
            if data.get('role') == 'admin':
                role_item.setForeground(QColor("#0078d4"))
            self.user_table.setItem(row, 2, role_item)

            # Actions buttons
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(2, 2, 2, 2)
            actions_layout.setSpacing(4)

            btn_edit = QPushButton("Edit")
            btn_edit.setFixedWidth(55)
            btn_edit.clicked.connect(lambda checked, e=user_key: self._edit_user_dialog(e))
            actions_layout.addWidget(btn_edit)

            # Only show Reset PW for password-based users
            if not is_windows_user:
                btn_reset = QPushButton("Reset PW")
                btn_reset.setFixedWidth(80)
                btn_reset.clicked.connect(lambda checked, e=user_key: self._reset_user_password(e))
                actions_layout.addWidget(btn_reset)

            # Don't allow deleting yourself
            if hasattr(self, 'auth_manager') and self.auth_manager:
                current_user = self.auth_manager.current_user or ''
                if user_key.lower() != current_user.lower():
                    btn_delete = QPushButton("Del")
                    btn_delete.setFixedWidth(45)
                    btn_delete.setStyleSheet("color: #dc3545;")
                    btn_delete.clicked.connect(lambda checked, e=user_key: self._delete_user(e))
                    actions_layout.addWidget(btn_delete)

            self.user_table.setCellWidget(row, 3, actions_widget)

    def _load_auth_users(self) -> dict:
        """Load users from local auth_users.json file."""
        try:
            # Check in project root directory
            local_auth_path = Path(__file__).parent.parent / 'auth_users.json'
            if local_auth_path.exists():
                with open(local_auth_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('users', {})

            # Also check in TariffMill_Config repo
            config_path = Path.home() / 'TariffMill_Config' / 'auth_users.json'
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('users', {})

            return {}
        except Exception as e:
            logger.warning(f"Failed to load auth users: {e}")
            return {}

    def _save_auth_users(self, users: dict) -> bool:
        """Save users to local auth_users.json file and sync to GitHub."""
        try:
            data = {
                "_comment": "TariffMill User Authentication Configuration - DO NOT COMMIT REAL PASSWORDS TO PUBLIC REPO",
                "_instructions": "To add users: 1) Run scripts/generate_password_hash.py 2) Add the output to the users object below",
                "users": users
            }

            # Save to local file first
            local_auth_path = Path(__file__).parent.parent / 'auth_users.json'
            with open(local_auth_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4)

            # Also save to TariffMill_Config repo if it exists
            config_path = Path.home() / 'TariffMill_Config' / 'auth_users.json'
            if config_path.parent.exists():
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=4)

                # Git commit and push
                self._sync_auth_to_github(config_path)

            return True
        except Exception as e:
            logger.error(f"Failed to save auth users: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save users: {e}")
            return False

    def _sync_auth_to_github(self, config_path: Path):
        """Sync auth_users.json to GitHub repository."""
        import subprocess
        try:
            repo_dir = config_path.parent
            # Git add, commit, push
            subprocess.run(['git', 'add', 'auth_users.json'], cwd=repo_dir, check=True, capture_output=True)
            subprocess.run(['git', 'commit', '-m', 'Update user credentials'], cwd=repo_dir, check=True, capture_output=True)
            subprocess.run(['git', 'push'], cwd=repo_dir, check=True, capture_output=True)
            logger.info("Successfully synced auth_users.json to GitHub")
        except subprocess.CalledProcessError as e:
            # Commit might fail if no changes - that's OK
            if b'nothing to commit' not in e.stderr:
                logger.warning(f"Failed to sync to GitHub: {e.stderr.decode() if e.stderr else e}")
        except Exception as e:
            logger.warning(f"Failed to sync to GitHub: {e}")

    def _generate_password_hash(self, password: str) -> tuple:
        """Generate a salted SHA-256 hash for a password."""
        import hashlib
        import secrets
        salt = secrets.token_hex(16)
        salted = f"{salt}{password}".encode('utf-8')
        password_hash = hashlib.sha256(salted).hexdigest()
        return password_hash, salt

    def _add_user_dialog(self):
        """Show dialog to add a new user."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Add User")
        dialog.setFixedSize(450, 340)

        layout = QVBoxLayout(dialog)
        form = QFormLayout()

        # User type selection
        user_type_combo = QComboBox()
        user_type_combo.addItems(["Email/Password User", "Windows Domain User"])
        form.addRow("User Type:", user_type_combo)

        # Email/Username input
        email_input = QLineEdit()
        email_input.setPlaceholderText("user@example.com")
        email_label = QLabel("Email:")
        form.addRow(email_label, email_input)

        # Windows domain input (for domain users)
        domain_input = QLineEdit()
        domain_input.setPlaceholderText("DMUSA")
        domain_input.setText("DMUSA")
        domain_label = QLabel("Domain:")
        domain_input.setVisible(False)
        domain_label.setVisible(False)
        form.addRow(domain_label, domain_input)

        name_input = QLineEdit()
        name_input.setPlaceholderText("Display Name")
        form.addRow("Name:", name_input)

        role_combo = QComboBox()
        role_combo.addItems(["user", "admin"])
        form.addRow("Role:", role_combo)

        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.Password)
        password_input.setPlaceholderText("Password")
        password_label = QLabel("Password:")
        form.addRow(password_label, password_input)

        confirm_input = QLineEdit()
        confirm_input.setEchoMode(QLineEdit.Password)
        confirm_input.setPlaceholderText("Confirm Password")
        confirm_label = QLabel("Confirm:")
        form.addRow(confirm_label, confirm_input)

        # Toggle password fields visibility based on user type
        def on_user_type_changed(index):
            is_windows = index == 1
            domain_input.setVisible(is_windows)
            domain_label.setVisible(is_windows)
            password_input.setVisible(not is_windows)
            password_label.setVisible(not is_windows)
            confirm_input.setVisible(not is_windows)
            confirm_label.setVisible(not is_windows)
            if is_windows:
                email_label.setText("Username:")
                email_input.setPlaceholderText("username (without domain)")
            else:
                email_label.setText("Email:")
                email_input.setPlaceholderText("user@example.com")

        user_type_combo.currentIndexChanged.connect(on_user_type_changed)

        layout.addLayout(form)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_cancel)

        btn_save = QPushButton("Add User")
        btn_save.setStyleSheet(self.get_button_style("primary"))
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

        def save_user():
            is_windows_user = user_type_combo.currentIndex() == 1
            name = name_input.text().strip()
            role = role_combo.currentText()

            if is_windows_user:
                # Windows domain user
                username = email_input.text().strip().lower()
                domain = domain_input.text().strip().upper()

                if not username:
                    QMessageBox.warning(dialog, "Error", "Please enter a username.")
                    return
                if not domain:
                    QMessageBox.warning(dialog, "Error", "Please enter a domain.")
                    return

                # Build Windows user key: DOMAIN\username
                user_key = f"{domain}\\{username}"
                name = name or username

                users = self._load_auth_users()
                if user_key in users or user_key.upper() in [k.upper() for k in users.keys()]:
                    QMessageBox.warning(dialog, "Error", "User already exists.")
                    return

                # Windows users don't need password hash
                users[user_key] = {
                    "role": role,
                    "name": name,
                    "auth_type": "windows"
                }

                if self._save_auth_users(users):
                    QMessageBox.information(dialog, "Success", f"Windows user {user_key} added successfully.\n\nThis user will auto-login when logged into the {domain} domain.")
                    dialog.accept()
                    self._refresh_user_list()
            else:
                # Email/password user
                email = email_input.text().strip().lower()
                password = password_input.text()
                confirm = confirm_input.text()
                name = name or email

                if not email or '@' not in email:
                    QMessageBox.warning(dialog, "Error", "Please enter a valid email address.")
                    return
                if not password:
                    QMessageBox.warning(dialog, "Error", "Please enter a password.")
                    return
                if password != confirm:
                    QMessageBox.warning(dialog, "Error", "Passwords do not match.")
                    return

                users = self._load_auth_users()
                if email in users:
                    QMessageBox.warning(dialog, "Error", "User already exists.")
                    return

                password_hash, salt = self._generate_password_hash(password)
                users[email] = {
                    "password_hash": password_hash,
                    "salt": salt,
                    "role": role,
                    "name": name,
                    "auth_type": "password"
                }

                if self._save_auth_users(users):
                    QMessageBox.information(dialog, "Success", f"User {email} added successfully.")
                    dialog.accept()
                    self._refresh_user_list()

        btn_save.clicked.connect(save_user)
        dialog.exec_()

    def _edit_user_dialog(self, email: str):
        """Show dialog to edit a user."""
        users = self._load_auth_users()
        if email not in users:
            QMessageBox.warning(self, "Error", "User not found.")
            return

        user_data = users[email]

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edit User: {email}")
        dialog.setFixedSize(400, 200)

        layout = QVBoxLayout(dialog)
        form = QFormLayout()

        email_label = QLabel(email)
        email_label.setStyleSheet("font-weight: bold;")
        form.addRow("Email:", email_label)

        name_input = QLineEdit()
        name_input.setText(user_data.get('name', ''))
        form.addRow("Name:", name_input)

        role_combo = QComboBox()
        role_combo.addItems(["user", "admin"])
        role_combo.setCurrentText(user_data.get('role', 'user'))
        form.addRow("Role:", role_combo)

        layout.addLayout(form)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_cancel)

        btn_save = QPushButton("Save Changes")
        btn_save.setStyleSheet(self.get_button_style("primary"))
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

        def save_changes():
            users[email]['name'] = name_input.text().strip() or email
            users[email]['role'] = role_combo.currentText()

            if self._save_auth_users(users):
                QMessageBox.information(dialog, "Success", "User updated successfully.")
                dialog.accept()
                self._refresh_user_list()

        btn_save.clicked.connect(save_changes)
        dialog.exec_()

    def _reset_user_password(self, email: str):
        """Reset a user's password."""
        users = self._load_auth_users()
        if email not in users:
            QMessageBox.warning(self, "Error", "User not found.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Reset Password: {email}")
        dialog.setFixedSize(350, 180)

        layout = QVBoxLayout(dialog)
        form = QFormLayout()

        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.Password)
        password_input.setPlaceholderText("New Password")
        form.addRow("New Password:", password_input)

        confirm_input = QLineEdit()
        confirm_input.setEchoMode(QLineEdit.Password)
        confirm_input.setPlaceholderText("Confirm Password")
        form.addRow("Confirm:", confirm_input)

        layout.addLayout(form)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_cancel)

        btn_save = QPushButton("Reset Password")
        btn_save.setStyleSheet(self.get_button_style("primary"))
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

        def do_reset():
            password = password_input.text()
            confirm = confirm_input.text()

            if not password:
                QMessageBox.warning(dialog, "Error", "Please enter a password.")
                return
            if password != confirm:
                QMessageBox.warning(dialog, "Error", "Passwords do not match.")
                return

            password_hash, salt = self._generate_password_hash(password)
            users[email]['password_hash'] = password_hash
            users[email]['salt'] = salt

            if self._save_auth_users(users):
                QMessageBox.information(dialog, "Success", f"Password for {email} has been reset.")
                dialog.accept()

        btn_save.clicked.connect(do_reset)
        dialog.exec_()

    def _delete_user(self, email: str):
        """Delete a user after confirmation."""
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete user:\n{email}?\n\nThis cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            users = self._load_auth_users()
            if email in users:
                del users[email]
                if self._save_auth_users(users):
                    QMessageBox.information(self, "Success", f"User {email} deleted.")
                    self._refresh_user_list()

    def setup_ai_agents_tab(self, tab_widget):
        """Setup the AI Agents tab for the Configuration dialog."""
        main_layout = QVBoxLayout(tab_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Create scroll area for the content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)

        # Create container widget for scroll area
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)

        # Title
        title = QLabel("<h2>AI Agent Configuration</h2>")
        layout.addWidget(title)

        description = QLabel(
            "Configure AI providers and API keys for the AI Template Generator and other AI features. "
            "API keys are stored securely in the local database."
        )
        description.setWordWrap(True)
        description.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(description)

        # OpenAI Configuration
        openai_group = QGroupBox("OpenAI")
        openai_layout = QFormLayout()

        # Package status row
        openai_pkg_layout = QHBoxLayout()
        self.openai_pkg_indicator = QLabel("●")
        self.openai_pkg_label = QLabel("Checking...")
        openai_pkg_layout.addWidget(self.openai_pkg_indicator)
        openai_pkg_layout.addWidget(self.openai_pkg_label)
        openai_pkg_layout.addStretch()
        self.btn_install_openai = QPushButton("Install Package")
        self.btn_install_openai.clicked.connect(lambda: self._install_ai_package('openai'))
        self.btn_install_openai.setVisible(False)
        openai_pkg_layout.addWidget(self.btn_install_openai)
        openai_layout.addRow("Package:", openai_pkg_layout)

        self.ai_openai_api_key = QLineEdit()
        self.ai_openai_api_key.setEchoMode(QLineEdit.Password)
        self.ai_openai_api_key.setPlaceholderText("sk-...")
        saved_openai_key = self._get_ai_api_key('openai')
        if saved_openai_key:
            self.ai_openai_api_key.setText(saved_openai_key)
        openai_layout.addRow("API Key:", self.ai_openai_api_key)

        self.ai_openai_model = QComboBox()
        self.ai_openai_model.setEditable(True)
        self.ai_openai_model.addItems(["gpt-4o", "gpt-4-turbo", "gpt-4", "gpt-3.5-turbo", "gpt-4o-mini"])
        saved_openai_model = self._get_ai_setting('openai_default_model')
        if saved_openai_model:
            idx = self.ai_openai_model.findText(saved_openai_model)
            if idx >= 0:
                self.ai_openai_model.setCurrentIndex(idx)
            else:
                self.ai_openai_model.setCurrentText(saved_openai_model)
        openai_layout.addRow("Default Model:", self.ai_openai_model)

        # OpenAI status indicator
        openai_status_layout = QHBoxLayout()
        self.openai_status_indicator = QLabel("●")
        self.openai_status_label = QLabel("Not configured")
        openai_status_layout.addWidget(self.openai_status_indicator)
        openai_status_layout.addWidget(self.openai_status_label)
        openai_status_layout.addStretch()

        btn_test_openai = QPushButton("Test Connection")
        btn_test_openai.clicked.connect(lambda: self._test_ai_connection('openai'))
        openai_status_layout.addWidget(btn_test_openai)
        openai_layout.addRow("Status:", openai_status_layout)

        openai_group.setLayout(openai_layout)
        layout.addWidget(openai_group)

        # Anthropic Configuration
        anthropic_group = QGroupBox("Anthropic (Claude)")
        anthropic_layout = QFormLayout()

        # Package status row
        anthropic_pkg_layout = QHBoxLayout()
        self.anthropic_pkg_indicator = QLabel("●")
        self.anthropic_pkg_label = QLabel("Checking...")
        anthropic_pkg_layout.addWidget(self.anthropic_pkg_indicator)
        anthropic_pkg_layout.addWidget(self.anthropic_pkg_label)
        anthropic_pkg_layout.addStretch()
        self.btn_install_anthropic = QPushButton("Install Package")
        self.btn_install_anthropic.clicked.connect(lambda: self._install_ai_package('anthropic'))
        self.btn_install_anthropic.setVisible(False)
        anthropic_pkg_layout.addWidget(self.btn_install_anthropic)
        anthropic_layout.addRow("Package:", anthropic_pkg_layout)

        self.ai_anthropic_api_key = QLineEdit()
        self.ai_anthropic_api_key.setEchoMode(QLineEdit.Password)
        self.ai_anthropic_api_key.setPlaceholderText("sk-ant-...")
        saved_anthropic_key = self._get_ai_api_key('anthropic')
        if saved_anthropic_key:
            self.ai_anthropic_api_key.setText(saved_anthropic_key)
        anthropic_layout.addRow("API Key:", self.ai_anthropic_api_key)

        self.ai_anthropic_model = QComboBox()
        self.ai_anthropic_model.setEditable(True)
        self.ai_anthropic_model.addItems(["claude-sonnet-4-20250514", "claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022", "claude-3-opus-20240229"])
        saved_anthropic_model = self._get_ai_setting('anthropic_default_model')
        if saved_anthropic_model:
            idx = self.ai_anthropic_model.findText(saved_anthropic_model)
            if idx >= 0:
                self.ai_anthropic_model.setCurrentIndex(idx)
            else:
                self.ai_anthropic_model.setCurrentText(saved_anthropic_model)
        anthropic_layout.addRow("Default Model:", self.ai_anthropic_model)

        # Anthropic status indicator
        anthropic_status_layout = QHBoxLayout()
        self.anthropic_status_indicator = QLabel("●")
        self.anthropic_status_label = QLabel("Not configured")
        anthropic_status_layout.addWidget(self.anthropic_status_indicator)
        anthropic_status_layout.addWidget(self.anthropic_status_label)
        anthropic_status_layout.addStretch()

        btn_test_anthropic = QPushButton("Test Connection")
        btn_test_anthropic.clicked.connect(lambda: self._test_ai_connection('anthropic'))
        anthropic_status_layout.addWidget(btn_test_anthropic)
        anthropic_layout.addRow("Status:", anthropic_status_layout)

        anthropic_group.setLayout(anthropic_layout)
        layout.addWidget(anthropic_group)

        # Google Gemini Configuration
        gemini_group = QGroupBox("Google Gemini")
        gemini_layout = QFormLayout()

        # Package status row
        gemini_pkg_layout = QHBoxLayout()
        self.gemini_pkg_indicator = QLabel("●")
        self.gemini_pkg_label = QLabel("Checking...")
        gemini_pkg_layout.addWidget(self.gemini_pkg_indicator)
        gemini_pkg_layout.addWidget(self.gemini_pkg_label)
        gemini_pkg_layout.addStretch()
        self.btn_install_gemini = QPushButton("Install Package")
        self.btn_install_gemini.clicked.connect(lambda: self._install_ai_package('google-generativeai'))
        self.btn_install_gemini.setVisible(False)
        gemini_pkg_layout.addWidget(self.btn_install_gemini)
        gemini_layout.addRow("Package:", gemini_pkg_layout)

        self.ai_gemini_api_key = QLineEdit()
        self.ai_gemini_api_key.setEchoMode(QLineEdit.Password)
        self.ai_gemini_api_key.setPlaceholderText("AI...")
        saved_gemini_key = self._get_ai_api_key('gemini')
        if saved_gemini_key:
            self.ai_gemini_api_key.setText(saved_gemini_key)
        gemini_layout.addRow("API Key:", self.ai_gemini_api_key)

        self.ai_gemini_model = QComboBox()
        self.ai_gemini_model.setEditable(True)
        self.ai_gemini_model.addItems(["gemini-1.5-pro", "gemini-1.5-flash", "gemini-1.0-pro"])
        saved_gemini_model = self._get_ai_setting('gemini_default_model')
        if saved_gemini_model:
            idx = self.ai_gemini_model.findText(saved_gemini_model)
            if idx >= 0:
                self.ai_gemini_model.setCurrentIndex(idx)
            else:
                self.ai_gemini_model.setCurrentText(saved_gemini_model)
        gemini_layout.addRow("Default Model:", self.ai_gemini_model)

        # Gemini status indicator
        gemini_status_layout = QHBoxLayout()
        self.gemini_status_indicator = QLabel("●")
        self.gemini_status_label = QLabel("Not configured")
        gemini_status_layout.addWidget(self.gemini_status_indicator)
        gemini_status_layout.addWidget(self.gemini_status_label)
        gemini_status_layout.addStretch()

        btn_test_gemini = QPushButton("Test Connection")
        btn_test_gemini.clicked.connect(lambda: self._test_ai_connection('gemini'))
        gemini_status_layout.addWidget(btn_test_gemini)
        gemini_layout.addRow("Status:", gemini_status_layout)

        gemini_group.setLayout(gemini_layout)
        layout.addWidget(gemini_group)

        # Groq Configuration
        groq_group = QGroupBox("Groq")
        groq_layout = QFormLayout()

        # Package status row
        groq_pkg_layout = QHBoxLayout()
        self.groq_pkg_indicator = QLabel("●")
        self.groq_pkg_label = QLabel("Checking...")
        groq_pkg_layout.addWidget(self.groq_pkg_indicator)
        groq_pkg_layout.addWidget(self.groq_pkg_label)
        groq_pkg_layout.addStretch()
        self.btn_install_groq = QPushButton("Install Package")
        self.btn_install_groq.clicked.connect(lambda: self._install_ai_package('groq'))
        self.btn_install_groq.setVisible(False)
        groq_pkg_layout.addWidget(self.btn_install_groq)
        groq_layout.addRow("Package:", groq_pkg_layout)

        self.ai_groq_api_key = QLineEdit()
        self.ai_groq_api_key.setEchoMode(QLineEdit.Password)
        self.ai_groq_api_key.setPlaceholderText("gsk_...")
        saved_groq_key = self._get_ai_api_key('groq')
        if saved_groq_key:
            self.ai_groq_api_key.setText(saved_groq_key)
        groq_layout.addRow("API Key:", self.ai_groq_api_key)

        self.ai_groq_model = QComboBox()
        self.ai_groq_model.setEditable(True)
        self.ai_groq_model.addItems(["llama-3.3-70b-versatile", "llama-3.1-8b-instant", "mixtral-8x7b-32768", "gemma2-9b-it"])
        saved_groq_model = self._get_ai_setting('groq_default_model')
        if saved_groq_model:
            idx = self.ai_groq_model.findText(saved_groq_model)
            if idx >= 0:
                self.ai_groq_model.setCurrentIndex(idx)
            else:
                self.ai_groq_model.setCurrentText(saved_groq_model)
        groq_layout.addRow("Default Model:", self.ai_groq_model)

        # Groq status indicator
        groq_status_layout = QHBoxLayout()
        self.groq_status_indicator = QLabel("●")
        self.groq_status_label = QLabel("Not configured")
        groq_status_layout.addWidget(self.groq_status_indicator)
        groq_status_layout.addWidget(self.groq_status_label)
        groq_status_layout.addStretch()

        btn_test_groq = QPushButton("Test Connection")
        btn_test_groq.clicked.connect(lambda: self._test_ai_connection('groq'))
        groq_status_layout.addWidget(btn_test_groq)
        groq_layout.addRow("Status:", groq_status_layout)

        groq_group.setLayout(groq_layout)
        layout.addWidget(groq_group)

        # Default Provider Selection
        default_group = QGroupBox("Default Settings")
        default_layout = QFormLayout()

        self.ai_default_provider = QComboBox()
        self.ai_default_provider.addItems(["OpenAI", "Anthropic", "Google Gemini", "Groq"])
        saved_default_provider = self._get_ai_setting('default_provider')
        if saved_default_provider:
            idx = self.ai_default_provider.findText(saved_default_provider)
            if idx >= 0:
                self.ai_default_provider.setCurrentIndex(idx)
        default_layout.addRow("Default Provider:", self.ai_default_provider)

        default_group.setLayout(default_layout)
        layout.addWidget(default_group)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_save = QPushButton("Save AI Settings")
        btn_save.setStyleSheet(self.get_button_style("primary"))
        btn_save.clicked.connect(self._save_ai_settings)
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)
        layout.addStretch()

        # Set the scroll content and add to main layout
        scroll_area.setWidget(scroll_content)
        main_layout.addWidget(scroll_area)

        # Update status indicators
        self._update_ai_status_indicators()
        self._update_package_indicators()

    def _get_ai_api_key(self, provider: str) -> str:
        """Get saved AI API key from database."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", (f'api_key_{provider}',))
            row = c.fetchone()
            conn.close()
            return row[0] if row else ""
        except Exception:
            return ""

    def _save_ai_api_key(self, provider: str, api_key: str):
        """Save AI API key to database."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS app_config (
                key TEXT PRIMARY KEY,
                value TEXT
            )""")
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                     (f'api_key_{provider}', api_key))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to save AI API key: {e}")

    def _get_ai_setting(self, key: str) -> str:
        """Get AI setting from database."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", (f'ai_{key}',))
            row = c.fetchone()
            conn.close()
            return row[0] if row else ""
        except Exception:
            return ""

    def _save_ai_setting(self, key: str, value: str):
        """Save AI setting to database."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS app_config (
                key TEXT PRIMARY KEY,
                value TEXT
            )""")
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                     (f'ai_{key}', value))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to save AI setting: {e}")

    def _test_ai_connection(self, provider: str):
        """Test connection to AI provider."""
        try:
            if provider == 'openai':
                api_key = self.ai_openai_api_key.text().strip()
                if not api_key:
                    QMessageBox.warning(self, "Test Failed", "Please enter an OpenAI API key.")
                    return
                # Test with a simple API call
                import urllib.request
                import json
                request = urllib.request.Request(
                    "https://api.openai.com/v1/models",
                    headers={
                        'Authorization': f'Bearer {api_key}',
                        'Content-Type': 'application/json'
                    }
                )
                with urllib.request.urlopen(request, timeout=10) as response:
                    data = json.loads(response.read().decode())
                    model_count = len(data.get('data', []))
                    QMessageBox.information(self, "Connection Successful",
                        f"Successfully connected to OpenAI API.\n{model_count} models available.")
                    self.openai_status_indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    self.openai_status_label.setText("Connected")
                    self.openai_status_label.setStyleSheet("color: #27ae60;")

            elif provider == 'anthropic':
                api_key = self.ai_anthropic_api_key.text().strip()
                if not api_key:
                    QMessageBox.warning(self, "Test Failed", "Please enter an Anthropic API key.")
                    return
                # Test by checking key format (Anthropic doesn't have a simple models endpoint)
                if api_key.startswith('sk-ant-'):
                    QMessageBox.information(self, "API Key Valid",
                        "Anthropic API key format is valid.\nThe key will be tested when generating templates.")
                    self.anthropic_status_indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    self.anthropic_status_label.setText("Key configured")
                    self.anthropic_status_label.setStyleSheet("color: #27ae60;")
                else:
                    QMessageBox.warning(self, "Invalid Key Format",
                        "Anthropic API keys should start with 'sk-ant-'")

            elif provider == 'gemini':
                api_key = self.ai_gemini_api_key.text().strip()
                if not api_key:
                    QMessageBox.warning(self, "Test Failed", "Please enter a Google AI API key.")
                    return
                # Test with a simple API call to list models
                import urllib.request
                import json
                request = urllib.request.Request(
                    f"https://generativelanguage.googleapis.com/v1/models?key={api_key}",
                    headers={'Content-Type': 'application/json'}
                )
                with urllib.request.urlopen(request, timeout=10) as response:
                    data = json.loads(response.read().decode())
                    model_count = len(data.get('models', []))
                    QMessageBox.information(self, "Connection Successful",
                        f"Successfully connected to Google AI API.\n{model_count} models available.")
                    self.gemini_status_indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    self.gemini_status_label.setText("Connected")
                    self.gemini_status_label.setStyleSheet("color: #27ae60;")

            elif provider == 'groq':
                api_key = self.ai_groq_api_key.text().strip()
                if not api_key:
                    QMessageBox.warning(self, "Test Failed", "Please enter a Groq API key.")
                    return
                # Test with a simple API call to list models
                import urllib.request
                import json
                request = urllib.request.Request(
                    "https://api.groq.com/openai/v1/models",
                    headers={
                        'Authorization': f'Bearer {api_key}',
                        'Content-Type': 'application/json'
                    }
                )
                with urllib.request.urlopen(request, timeout=10) as response:
                    data = json.loads(response.read().decode())
                    model_count = len(data.get('data', []))
                    QMessageBox.information(self, "Connection Successful",
                        f"Successfully connected to Groq API.\n{model_count} models available.")
                    self.groq_status_indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    self.groq_status_label.setText("Connected")
                    self.groq_status_label.setStyleSheet("color: #27ae60;")

        except urllib.error.HTTPError as e:
            QMessageBox.warning(self, "Connection Failed", f"HTTP Error: {e.code}\n{e.reason}")
            self._set_provider_status_error(provider)
        except urllib.error.URLError as e:
            QMessageBox.warning(self, "Connection Failed", f"Could not connect: {e.reason}")
            self._set_provider_status_error(provider)
        except Exception as e:
            QMessageBox.warning(self, "Connection Failed", f"Error: {str(e)}")
            self._set_provider_status_error(provider)

    def _set_provider_status_error(self, provider: str):
        """Set provider status indicator to error state."""
        if provider == 'openai':
            self.openai_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.openai_status_label.setText("Connection failed")
            self.openai_status_label.setStyleSheet("color: #e74c3c;")
        elif provider == 'anthropic':
            self.anthropic_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.anthropic_status_label.setText("Connection failed")
            self.anthropic_status_label.setStyleSheet("color: #e74c3c;")
        elif provider == 'gemini':
            self.gemini_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.gemini_status_label.setText("Connection failed")
            self.gemini_status_label.setStyleSheet("color: #e74c3c;")
        elif provider == 'groq':
            self.groq_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.groq_status_label.setText("Connection failed")
            self.groq_status_label.setStyleSheet("color: #e74c3c;")

    def _update_ai_status_indicators(self):
        """Update all AI status indicators based on current configuration."""
        # OpenAI
        openai_key = self.ai_openai_api_key.text().strip()
        if openai_key and openai_key.startswith('sk-'):
            self.openai_status_indicator.setStyleSheet("color: #f39c12; font-size: 16px; font-weight: bold;")
            self.openai_status_label.setText("Key configured - test to verify")
            self.openai_status_label.setStyleSheet("color: #f39c12;")
        elif openai_key:
            self.openai_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.openai_status_label.setText("Invalid key format")
            self.openai_status_label.setStyleSheet("color: #e74c3c;")
        else:
            self.openai_status_indicator.setStyleSheet("color: #95a5a6; font-size: 16px; font-weight: bold;")
            self.openai_status_label.setText("Not configured")
            self.openai_status_label.setStyleSheet("color: #95a5a6;")

        # Anthropic
        anthropic_key = self.ai_anthropic_api_key.text().strip()
        if anthropic_key and anthropic_key.startswith('sk-ant-'):
            self.anthropic_status_indicator.setStyleSheet("color: #f39c12; font-size: 16px; font-weight: bold;")
            self.anthropic_status_label.setText("Key configured - test to verify")
            self.anthropic_status_label.setStyleSheet("color: #f39c12;")
        elif anthropic_key:
            self.anthropic_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.anthropic_status_label.setText("Invalid key format (should start with sk-ant-)")
            self.anthropic_status_label.setStyleSheet("color: #e74c3c;")
        else:
            self.anthropic_status_indicator.setStyleSheet("color: #95a5a6; font-size: 16px; font-weight: bold;")
            self.anthropic_status_label.setText("Not configured")
            self.anthropic_status_label.setStyleSheet("color: #95a5a6;")

        # Google Gemini
        gemini_key = self.ai_gemini_api_key.text().strip()
        if gemini_key and gemini_key.startswith('AI'):
            self.gemini_status_indicator.setStyleSheet("color: #f39c12; font-size: 16px; font-weight: bold;")
            self.gemini_status_label.setText("Key configured - test to verify")
            self.gemini_status_label.setStyleSheet("color: #f39c12;")
        elif gemini_key:
            self.gemini_status_indicator.setStyleSheet("color: #f39c12; font-size: 16px; font-weight: bold;")
            self.gemini_status_label.setText("Key configured - test to verify")
            self.gemini_status_label.setStyleSheet("color: #f39c12;")
        else:
            self.gemini_status_indicator.setStyleSheet("color: #95a5a6; font-size: 16px; font-weight: bold;")
            self.gemini_status_label.setText("Not configured")
            self.gemini_status_label.setStyleSheet("color: #95a5a6;")

        # Groq
        groq_key = self.ai_groq_api_key.text().strip()
        if groq_key and groq_key.startswith('gsk_'):
            self.groq_status_indicator.setStyleSheet("color: #f39c12; font-size: 16px; font-weight: bold;")
            self.groq_status_label.setText("Key configured - test to verify")
            self.groq_status_label.setStyleSheet("color: #f39c12;")
        elif groq_key:
            self.groq_status_indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
            self.groq_status_label.setText("Invalid key format (should start with gsk_)")
            self.groq_status_label.setStyleSheet("color: #e74c3c;")
        else:
            self.groq_status_indicator.setStyleSheet("color: #95a5a6; font-size: 16px; font-weight: bold;")
            self.groq_status_label.setText("Not configured")
            self.groq_status_label.setStyleSheet("color: #95a5a6;")

    def _save_ai_settings(self):
        """Save all AI settings from the Configuration dialog tab."""
        # Save API keys
        self._save_ai_api_key('openai', self.ai_openai_api_key.text().strip())
        self._save_ai_api_key('anthropic', self.ai_anthropic_api_key.text().strip())
        self._save_ai_api_key('gemini', self.ai_gemini_api_key.text().strip())
        self._save_ai_api_key('groq', self.ai_groq_api_key.text().strip())

        # Save default models
        self._save_ai_setting('openai_default_model', self.ai_openai_model.currentText())
        self._save_ai_setting('anthropic_default_model', self.ai_anthropic_model.currentText())
        self._save_ai_setting('gemini_default_model', self.ai_gemini_model.currentText())
        self._save_ai_setting('groq_default_model', self.ai_groq_model.currentText())

        # Save default provider
        self._save_ai_setting('default_provider', self.ai_default_provider.currentText())

        QMessageBox.information(self, "Settings Saved", "AI Agent settings have been saved successfully.")

        # Update status indicators
        self._update_ai_status_indicators()

    def _update_package_indicators(self):
        """Update package availability indicators for all AI providers."""
        packages = {
            'openai': ('openai', self.openai_pkg_indicator, self.openai_pkg_label, self.btn_install_openai),
            'anthropic': ('anthropic', self.anthropic_pkg_indicator, self.anthropic_pkg_label, self.btn_install_anthropic),
            'gemini': ('google-generativeai', self.gemini_pkg_indicator, self.gemini_pkg_label, self.btn_install_gemini),
            'groq': ('groq', self.groq_pkg_indicator, self.groq_pkg_label, self.btn_install_groq),
        }

        for provider, (pkg_name, indicator, label, btn) in packages.items():
            try:
                if provider == 'openai':
                    import openai
                    version = getattr(openai, '__version__', 'installed')
                    indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    label.setText(f"openai v{version} installed")
                    label.setStyleSheet("color: #27ae60;")
                    btn.setVisible(False)
                elif provider == 'anthropic':
                    import anthropic
                    version = getattr(anthropic, '__version__', 'installed')
                    indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    label.setText(f"anthropic v{version} installed")
                    label.setStyleSheet("color: #27ae60;")
                    btn.setVisible(False)
                elif provider == 'gemini':
                    import google.generativeai as genai
                    indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    label.setText("google-generativeai installed")
                    label.setStyleSheet("color: #27ae60;")
                    btn.setVisible(False)
                elif provider == 'groq':
                    from groq import Groq
                    indicator.setStyleSheet("color: #27ae60; font-size: 16px; font-weight: bold;")
                    label.setText("groq installed")
                    label.setStyleSheet("color: #27ae60;")
                    btn.setVisible(False)
            except ImportError:
                indicator.setStyleSheet("color: #e74c3c; font-size: 16px; font-weight: bold;")
                label.setText(f"{pkg_name} not installed")
                label.setStyleSheet("color: #e74c3c;")
                btn.setVisible(True)

    def _install_ai_package(self, package_name: str):
        """Install an AI package using pip."""
        reply = QMessageBox.question(
            self,
            "Install Package",
            f"Do you want to install the '{package_name}' package?\n\n"
            f"This will run: pip install {package_name}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )

        if reply != QMessageBox.Yes:
            return

        # Create progress dialog
        progress = QProgressDialog(f"Installing {package_name}...", None, 0, 0, self)
        progress.setWindowTitle("Installing Package")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        QApplication.processEvents()

        try:
            import subprocess
            import sys

            # Run pip install
            result = subprocess.run(
                [sys.executable, '-m', 'pip', 'install', package_name],
                capture_output=True,
                text=True,
                timeout=120
            )

            progress.close()

            if result.returncode == 0:
                QMessageBox.information(
                    self,
                    "Installation Complete",
                    f"Successfully installed '{package_name}'.\n\n"
                    "Please restart the application to use this provider."
                )
                # Update indicators
                self._update_package_indicators()
            else:
                QMessageBox.warning(
                    self,
                    "Installation Failed",
                    f"Failed to install '{package_name}'.\n\n"
                    f"Error: {result.stderr[:500] if result.stderr else 'Unknown error'}"
                )
        except subprocess.TimeoutExpired:
            progress.close()
            QMessageBox.warning(
                self,
                "Installation Timeout",
                f"Installation of '{package_name}' timed out.\n"
                "Please try installing manually using:\n"
                f"pip install {package_name}"
            )
        except Exception as e:
            progress.close()
            QMessageBox.warning(
                self,
                "Installation Error",
                f"An error occurred during installation:\n{str(e)}"
            )

    def show_billing_settings_dialog(self):
        """Show dialog to configure billing settings."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Billing Settings")
        dialog.setMinimumSize(450, 350)

        layout = QVBoxLayout(dialog)

        # Billing Rate Group
        rate_group = QGroupBox("Billing Rate")
        rate_layout = QFormLayout()

        rate_per_file = QDoubleSpinBox()
        rate_per_file.setRange(0, 1000)
        rate_per_file.setDecimals(2)
        rate_per_file.setPrefix("$")
        rate_per_file.setValue(float(self.get_billing_setting('rate_per_file', '0')))
        rate_layout.addRow("Rate per Export:", rate_per_file)

        rate_info = QLabel("<small>Amount to charge per exported file. Set to 0 to disable billing.</small>")
        rate_info.setWordWrap(True)
        rate_info.setStyleSheet("color:#666; padding:5px;")
        rate_layout.addRow("", rate_info)

        rate_group.setLayout(rate_layout)
        layout.addWidget(rate_group)

        # Email Configuration Group
        email_group = QGroupBox("Email Configuration")
        email_layout = QFormLayout()

        customer_email = QLineEdit()
        customer_email.setPlaceholderText("customer@example.com")
        customer_email.setText(self.get_billing_setting('customer_email', ''))
        email_layout.addRow("Customer Email:", customer_email)

        admin_email = QLineEdit()
        admin_email.setPlaceholderText("your@email.com")
        admin_email.setText(self.get_billing_setting('admin_email', ''))
        email_layout.addRow("Admin Email:", admin_email)

        smtp_server = QLineEdit()
        smtp_server.setPlaceholderText("smtp.gmail.com")
        smtp_server.setText(self.get_billing_setting('smtp_server', ''))
        email_layout.addRow("SMTP Server:", smtp_server)

        smtp_port = QSpinBox()
        smtp_port.setRange(1, 65535)
        smtp_port.setValue(int(self.get_billing_setting('smtp_port', '587')))
        email_layout.addRow("SMTP Port:", smtp_port)

        smtp_user = QLineEdit()
        smtp_user.setPlaceholderText("username@gmail.com")
        smtp_user.setText(self.get_billing_setting('smtp_user', ''))
        email_layout.addRow("SMTP Username:", smtp_user)

        smtp_password = QLineEdit()
        smtp_password.setEchoMode(QLineEdit.Password)
        smtp_password.setText(self.get_billing_setting('smtp_password', ''))
        email_layout.addRow("SMTP Password:", smtp_password)

        email_info = QLabel("<small>Configure SMTP settings to enable automatic monthly billing reports.</small>")
        email_info.setWordWrap(True)
        email_info.setStyleSheet("color:#666; padding:5px;")
        email_layout.addRow("", email_info)

        email_group.setLayout(email_layout)
        layout.addWidget(email_group)

        # Buttons
        btn_layout = QHBoxLayout()

        test_btn = QPushButton("Test Email")
        test_btn.clicked.connect(lambda: self._test_billing_email(
            smtp_server.text(), smtp_port.value(), smtp_user.text(), smtp_password.text(), admin_email.text()
        ))
        btn_layout.addWidget(test_btn)

        btn_layout.addStretch()

        save_btn = QPushButton("Save")
        save_btn.clicked.connect(lambda: self._save_billing_settings(
            dialog, rate_per_file.value(), customer_email.text(), admin_email.text(),
            smtp_server.text(), smtp_port.value(), smtp_user.text(), smtp_password.text()
        ))
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

        dialog.exec_()

    def _save_billing_settings(self, dialog, rate, customer_email, admin_email, smtp_server, smtp_port, smtp_user, smtp_password):
        """Save billing settings to database."""
        self.set_billing_setting('rate_per_file', str(rate))
        self.set_billing_setting('customer_email', customer_email)
        self.set_billing_setting('admin_email', admin_email)
        self.set_billing_setting('smtp_server', smtp_server)
        self.set_billing_setting('smtp_port', str(smtp_port))
        self.set_billing_setting('smtp_user', smtp_user)
        self.set_billing_setting('smtp_password', smtp_password)

        QMessageBox.information(self, "Saved", "Billing settings saved successfully.")
        dialog.accept()

    def _test_billing_email(self, smtp_server, smtp_port, smtp_user, smtp_password, to_email):
        """Test email configuration by sending a test email."""
        if not all([smtp_server, smtp_user, smtp_password, to_email]):
            QMessageBox.warning(self, "Missing Settings", "Please fill in all SMTP settings and admin email to test.")
            return

        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = to_email
            msg['Subject'] = "TariffMill Billing - Test Email"

            body = "This is a test email from TariffMill billing system.\n\nIf you received this, your email configuration is working correctly."
            msg.attach(MIMEText(body, 'plain'))

            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            server.quit()

            QMessageBox.information(self, "Success", f"Test email sent successfully to {to_email}!")

        except Exception as e:
            QMessageBox.critical(self, "Email Failed", f"Failed to send test email:\n{str(e)}")

    def show_billing_report_dialog(self):
        """Show dialog with billing report for the selected month."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Billing Report")
        dialog.setMinimumSize(800, 600)

        layout = QVBoxLayout(dialog)

        # Month selection
        month_layout = QHBoxLayout()
        month_layout.addWidget(QLabel("Month:"))

        month_combo = QComboBox()
        # Generate list of last 12 months
        now = datetime.now()
        for i in range(12):
            month_date = now - timedelta(days=i * 30)
            month_str = month_date.strftime("%Y-%m")
            month_display = month_date.strftime("%B %Y")
            month_combo.addItem(month_display, month_str)

        month_layout.addWidget(month_combo)
        month_layout.addStretch()

        refresh_btn = QPushButton("Refresh")
        month_layout.addWidget(refresh_btn)

        generate_btn = QPushButton("Generate Report")
        month_layout.addWidget(generate_btn)

        send_btn = QPushButton("Send Invoice")
        month_layout.addWidget(send_btn)

        layout.addLayout(month_layout)

        # Summary
        summary_group = QGroupBox("Summary")
        summary_layout = QFormLayout()

        export_count_label = QLabel("0")
        summary_layout.addRow("Total Exports:", export_count_label)

        total_lines_label = QLabel("0")
        summary_layout.addRow("Total Lines:", total_lines_label)

        total_value_label = QLabel("$0.00")
        summary_layout.addRow("Total Value Processed:", total_value_label)

        rate_label = QLabel("$0.00")
        summary_layout.addRow("Rate per Export:", rate_label)

        amount_due_label = QLabel("$0.00")
        amount_due_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        summary_layout.addRow("Amount Due:", amount_due_label)

        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)

        # Details table
        details_group = QGroupBox("Export Details")
        details_layout = QVBoxLayout()

        details_table = QTableWidget()
        details_table.setColumnCount(7)
        details_table.setHorizontalHeaderLabels([
            "File #", "Date", "Time", "File Name", "Lines", "Value", "MID"
        ])
        details_table.horizontalHeader().setStretchLastSection(True)
        details_table.setAlternatingRowColors(True)

        details_layout.addWidget(details_table)
        details_group.setLayout(details_layout)
        layout.addWidget(details_group)

        # Function to refresh the report
        def refresh_report():
            selected_month = month_combo.currentData()
            summary = self.get_billing_summary(selected_month)
            rate = float(self.get_billing_setting('rate_per_file', '0'))

            export_count_label.setText(str(summary['export_count']))
            total_lines_label.setText(str(summary['total_lines']))
            total_value_label.setText(f"${summary['total_value']:,.2f}")
            rate_label.setText(f"${rate:.2f}")
            amount_due_label.setText(f"${summary['export_count'] * rate:,.2f}")

            # Populate table
            details_table.setRowCount(len(summary['records']))
            for row_idx, record in enumerate(summary['records']):
                details_table.setItem(row_idx, 0, QTableWidgetItem(record.get('file_number', '')))
                details_table.setItem(row_idx, 1, QTableWidgetItem(record.get('export_date', '')))
                details_table.setItem(row_idx, 2, QTableWidgetItem(record.get('export_time', '')))
                details_table.setItem(row_idx, 3, QTableWidgetItem(record.get('file_name', '')))
                details_table.setItem(row_idx, 4, QTableWidgetItem(str(record.get('line_count', 0))))
                details_table.setItem(row_idx, 5, QTableWidgetItem(f"${record.get('total_value', 0):,.2f}"))
                details_table.setItem(row_idx, 6, QTableWidgetItem(record.get('mid', '')))

            details_table.resizeColumnsToContents()

        # Connect signals
        month_combo.currentIndexChanged.connect(refresh_report)
        refresh_btn.clicked.connect(refresh_report)
        generate_btn.clicked.connect(lambda: self._generate_billing_report(month_combo.currentData()))
        send_btn.clicked.connect(lambda: self._send_billing_invoice(month_combo.currentData()))

        # Initial load
        refresh_report()

        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)

        dialog.exec_()

    def _generate_billing_report(self, month: str):
        """Generate a billing report file for the specified month."""
        summary = self.get_billing_summary(month)
        rate = float(self.get_billing_setting('rate_per_file', '0'))
        amount_due = summary['export_count'] * rate

        # Prompt for save location
        default_name = f"TariffMill_Invoice_{month}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Billing Report",
            str(OUTPUT_DIR / default_name),
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            # Create workbook
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice"

            # Header
            ws['A1'] = "TariffMill Billing Invoice"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Period: {month}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

            # Summary section
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = "Total Exports:"
            ws['B6'] = summary['export_count']
            ws['A7'] = "Total Lines Processed:"
            ws['B7'] = summary['total_lines']
            ws['A8'] = "Total Value Processed:"
            ws['B8'] = f"${summary['total_value']:,.2f}"
            ws['A9'] = "Rate per Export:"
            ws['B9'] = f"${rate:.2f}"
            ws['A10'] = "Amount Due:"
            ws['B10'] = f"${amount_due:,.2f}"
            ws['A10'].font = Font(bold=True)
            ws['B10'].font = Font(bold=True)

            # Details section
            ws['A12'] = "Export Details"
            ws['A12'].font = Font(bold=True)

            headers = ["File #", "Date", "Time", "File Name", "Lines", "Value", "MID", "Folder Profile"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=13, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            for row_idx, record in enumerate(summary['records'], 14):
                ws.cell(row=row_idx, column=1).value = record.get('file_number', '')
                ws.cell(row=row_idx, column=2).value = record.get('export_date', '')
                ws.cell(row=row_idx, column=3).value = record.get('export_time', '')
                ws.cell(row=row_idx, column=4).value = record.get('file_name', '')
                ws.cell(row=row_idx, column=5).value = record.get('line_count', 0)
                ws.cell(row=row_idx, column=6).value = f"${record.get('total_value', 0):,.2f}"
                ws.cell(row=row_idx, column=7).value = record.get('mid', '')
                ws.cell(row=row_idx, column=8).value = record.get('folder_profile', '')

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(file_path)
            QMessageBox.information(self, "Report Generated", f"Billing report saved to:\n{file_path}")
            logger.info(f"Generated billing report: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report:\n{str(e)}")
            logger.error(f"Failed to generate billing report: {e}")

    def _send_billing_invoice(self, month: str):
        """Send billing invoice email for the specified month."""
        customer_email = self.get_billing_setting('customer_email', '')
        admin_email = self.get_billing_setting('admin_email', '')
        smtp_server = self.get_billing_setting('smtp_server', '')
        smtp_port = int(self.get_billing_setting('smtp_port', '587'))
        smtp_user = self.get_billing_setting('smtp_user', '')
        smtp_password = self.get_billing_setting('smtp_password', '')

        if not all([customer_email, smtp_server, smtp_user, smtp_password]):
            QMessageBox.warning(self, "Missing Settings",
                              "Please configure billing email settings first (Billing > Billing Settings).")
            return

        summary = self.get_billing_summary(month)
        rate = float(self.get_billing_setting('rate_per_file', '0'))
        amount_due = summary['export_count'] * rate

        if summary['export_count'] == 0:
            QMessageBox.information(self, "No Exports", f"No exports found for {month}. No invoice to send.")
            return

        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = customer_email
            if admin_email:
                msg['Cc'] = admin_email
            msg['Subject'] = f"TariffMill Invoice - {month}"

            # Build email body
            body = f"""TariffMill Billing Invoice

Period: {month}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

SUMMARY
-------
Total Exports: {summary['export_count']}
Total Lines Processed: {summary['total_lines']}
Total Value Processed: ${summary['total_value']:,.2f}
Rate per Export: ${rate:.2f}

AMOUNT DUE: ${amount_due:,.2f}

EXPORT DETAILS
--------------
"""
            for record in summary['records']:
                body += f"File #{record.get('file_number', 'N/A')} - {record.get('export_date', '')} - {record.get('file_name', '')} ({record.get('line_count', 0)} lines, ${record.get('total_value', 0):,.2f})\n"

            body += f"\n\nThank you for using TariffMill."

            msg.attach(MIMEText(body, 'plain'))

            # Send email
            recipients = [customer_email]
            if admin_email:
                recipients.append(admin_email)

            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            server.quit()

            # Mark records as invoiced
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("UPDATE billing_records SET invoice_sent = 1 WHERE invoice_month = ?", (month,))
                conn.commit()
                conn.close()
            except:
                pass

            QMessageBox.information(self, "Invoice Sent", f"Invoice sent successfully to {customer_email}!")
            logger.info(f"Sent billing invoice for {month} to {customer_email}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to send invoice:\n{str(e)}")
            logger.error(f"Failed to send billing invoice: {e}")

    def show_folder_profile_dialog(self):
        """Show dialog to manage folder profiles"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Manage Folder Profiles")
        dialog.setMinimumSize(500, 400)

        layout = QVBoxLayout(dialog)

        # Profile list
        list_group = QGroupBox("Saved Folder Profiles")
        list_layout = QVBoxLayout()

        self.folder_profile_list = QListWidget()
        self.folder_profile_list.itemClicked.connect(self._folder_profile_selected)
        list_layout.addWidget(self.folder_profile_list)

        list_group.setLayout(list_layout)
        layout.addWidget(list_group)

        # Edit group
        edit_group = QGroupBox("Profile Details")
        edit_layout = QFormLayout()

        self.folder_profile_name_edit = QLineEdit()
        self.folder_profile_name_edit.setPlaceholderText("Enter profile name...")
        edit_layout.addRow("Profile Name:", self.folder_profile_name_edit)

        # Input folder
        input_row = QHBoxLayout()
        self.folder_profile_input_edit = QLineEdit()
        self.folder_profile_input_edit.setPlaceholderText("Select input folder...")
        input_browse_btn = QPushButton("Browse...")
        input_browse_btn.clicked.connect(self._browse_folder_profile_input)
        input_row.addWidget(self.folder_profile_input_edit)
        input_row.addWidget(input_browse_btn)
        edit_layout.addRow("Input Folder:", input_row)

        # Output folder
        output_row = QHBoxLayout()
        self.folder_profile_output_edit = QLineEdit()
        self.folder_profile_output_edit.setPlaceholderText("Select output folder...")
        output_browse_btn = QPushButton("Browse...")
        output_browse_btn.clicked.connect(self._browse_folder_profile_output)
        output_row.addWidget(self.folder_profile_output_edit)
        output_row.addWidget(output_browse_btn)
        edit_layout.addRow("Output Folder:", output_row)

        edit_group.setLayout(edit_layout)
        layout.addWidget(edit_group)

        # Buttons
        btn_layout = QHBoxLayout()

        save_btn = QPushButton("Save Profile")
        save_btn.setStyleSheet(self.get_button_style("success"))
        save_btn.clicked.connect(lambda: self._save_folder_profile(dialog))
        btn_layout.addWidget(save_btn)

        delete_btn = QPushButton("Delete Profile")
        delete_btn.setStyleSheet(self.get_button_style("danger"))
        delete_btn.clicked.connect(lambda: self._delete_folder_profile(dialog))
        btn_layout.addWidget(delete_btn)

        btn_layout.addStretch()

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

        # Load existing profiles
        self._refresh_folder_profile_list()

        # Pre-fill with current folders
        self.folder_profile_input_edit.setText(str(INPUT_DIR))
        self.folder_profile_output_edit.setText(str(OUTPUT_DIR))

        dialog.exec_()

        # Refresh dropdown after dialog closes
        self.load_folder_profiles()

    def _refresh_folder_profile_list(self):
        """Refresh the folder profile list in the dialog"""
        self.folder_profile_list.clear()
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT profile_name, input_folder, output_folder FROM folder_profiles ORDER BY profile_name")
            for row in c.fetchall():
                item = QListWidgetItem(row[0])
                item.setData(Qt.UserRole, {'input': row[1], 'output': row[2]})
                self.folder_profile_list.addItem(item)
            conn.close()
        except Exception as e:
            logger.warning(f"Failed to refresh folder profile list: {e}")

    def _folder_profile_selected(self, item):
        """Handle folder profile selection in dialog"""
        self.folder_profile_name_edit.setText(item.text())
        data = item.data(Qt.UserRole)
        if data:
            self.folder_profile_input_edit.setText(self._normalize_path(data.get('input', '')))
            self.folder_profile_output_edit.setText(self._normalize_path(data.get('output', '')))

    def _normalize_path(self, path):
        """Normalize path for the current operating system"""
        if not path:
            return path
        # Use os.path.normpath to convert slashes appropriately for the OS
        return os.path.normpath(path)

    def _browse_folder_profile_input(self):
        """Browse for input folder in profile dialog"""
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder", self.folder_profile_input_edit.text())
        if folder:
            self.folder_profile_input_edit.setText(self._normalize_path(folder))

    def _browse_folder_profile_output(self):
        """Browse for output folder in profile dialog"""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", self.folder_profile_output_edit.text())
        if folder:
            self.folder_profile_output_edit.setText(self._normalize_path(folder))

    def _save_folder_profile(self, dialog):
        """Save the folder profile"""
        name = self.folder_profile_name_edit.text().strip()
        if not name:
            QMessageBox.warning(dialog, "Missing Name", "Please enter a profile name.")
            return

        input_folder = self._normalize_path(self.folder_profile_input_edit.text().strip())
        output_folder = self._normalize_path(self.folder_profile_output_edit.text().strip())

        if not input_folder or not output_folder:
            QMessageBox.warning(dialog, "Missing Folders", "Please select both input and output folders.")
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("""INSERT OR REPLACE INTO folder_profiles
                        (profile_name, input_folder, output_folder, created_date)
                        VALUES (?, ?, ?, ?)""",
                     (name, input_folder, output_folder, datetime.now().isoformat()))
            conn.commit()
            conn.close()

            logger.success(f"Folder profile saved: {name}")
            QMessageBox.information(dialog, "Saved", f"Folder profile '{name}' saved successfully.")
            self._refresh_folder_profile_list()
        except Exception as e:
            logger.error(f"Failed to save folder profile: {e}")
            QMessageBox.critical(dialog, "Error", f"Failed to save folder profile: {e}")

    def _delete_folder_profile(self, dialog):
        """Delete the selected folder profile"""
        current_item = self.folder_profile_list.currentItem()
        if not current_item:
            QMessageBox.information(dialog, "No Selection", "Please select a profile to delete.")
            return

        name = current_item.text()
        reply = QMessageBox.question(dialog, "Confirm Delete",
                                    f"Are you sure you want to delete the folder profile '{name}'?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM folder_profiles WHERE profile_name = ?", (name,))
                conn.commit()
                conn.close()

                logger.info(f"Folder profile deleted: {name}")
                self._refresh_folder_profile_list()
                self.folder_profile_name_edit.clear()
            except Exception as e:
                logger.error(f"Failed to delete folder profile: {e}")
                QMessageBox.critical(dialog, "Error", f"Failed to delete folder profile: {e}")

    def apply_current_mapping(self):
        # Check if shipment_targets is valid (not deleted after dialog close)
        if not hasattr(self, 'shipment_targets') or not self.shipment_targets:
            return

        # Check if the widgets in shipment_targets are still valid
        try:
            # Try to actually access the widget's properties to see if it's been deleted
            for target in self.shipment_targets.values():
                if not target:
                    return
                # Try to access the widget - this will raise RuntimeError if deleted
                _ = target.isVisible()
                break  # Just check the first one
        except (RuntimeError, AttributeError):
            # Widgets have been deleted
            logger.debug("apply_current_mapping: shipment_targets widgets have been deleted, skipping")
            return

        # Batch UI updates to prevent GUI freezing
        try:
            for key, target in self.shipment_targets.items():
                # Additional safety check per widget
                try:
                    _ = target.isVisible()
                except (RuntimeError, AttributeError):
                    logger.debug(f"apply_current_mapping: target widget for {key} has been deleted, skipping")
                    continue

                col = self.shipment_mapping.get(key)
                if col:
                    target.column_name = col
                    target.setText(f"{key} <- {col}")
                    target.setProperty("occupied", True)
                else:
                    target.column_name = None
                    target.setText(f"Drop {key.replace('_', ' ')} here")
                    target.setProperty("occupied", False)

            # Apply all style updates at once after setting properties
            for target in self.shipment_targets.values():
                try:
                    _ = target.isVisible()
                    target.style().unpolish(target)
                    target.style().polish(target)
                except (RuntimeError, AttributeError):
                    continue
        except (RuntimeError, AttributeError) as e:
            logger.debug(f"apply_current_mapping: Error during mapping update: {e}")
            return

    def setup_master_tab(self):
        layout = QVBoxLayout(self.tab_master)
        title = QLabel("<h2>Parts View - Click any cell to edit</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        edit_box = QHBoxLayout()
        btn_add = QPushButton("Add Row")
        btn_add.setStyleSheet(self.get_button_style("success"))
        btn_del = QPushButton("Delete Selected")
        btn_del.setStyleSheet(self.get_button_style("danger"))
        btn_save = QPushButton("Save Changes")
        btn_save.setStyleSheet(self.get_button_style("success"))
        btn_refresh = QPushButton("Refresh")
        btn_refresh.setStyleSheet(self.get_button_style("info"))
        btn_export_missing = QPushButton("Export Missing HTS")
        btn_export_missing.setStyleSheet(self.get_button_style("secondary"))
        btn_export_missing.setToolTip("Export HTS codes missing CBP Qty1 to the reference file for lookup")
        btn_verify_hts = QPushButton("Verify HTS")
        btn_verify_hts.setStyleSheet(self.get_button_style("info"))
        btn_verify_hts.setToolTip("Verify all HTS codes against hts.db and update qty_unit values")
        btn_export_by_client = QPushButton("Export by Client")
        btn_export_by_client.setStyleSheet(self.get_button_style("primary"))
        btn_export_by_client.setToolTip("Export parts list filtered by client code to Excel")
        btn_add.clicked.connect(self.add_part_row)
        btn_del.clicked.connect(self.delete_selected_parts)
        btn_save.clicked.connect(self.save_parts_table)
        btn_refresh.clicked.connect(self.refresh_parts_table)
        btn_export_missing.clicked.connect(self.export_missing_hts_codes)
        btn_verify_hts.clicked.connect(self.verify_hts_codes_with_dialog)
        btn_export_by_client.clicked.connect(self.export_parts_by_client)
        edit_box.addWidget(QLabel("Edit:"))
        edit_box.addWidget(btn_add); edit_box.addWidget(btn_del); edit_box.addWidget(btn_save); edit_box.addWidget(btn_refresh)
        edit_box.addWidget(btn_export_missing)
        edit_box.addWidget(btn_verify_hts)
        edit_box.addWidget(btn_export_by_client)
        edit_box.addStretch()
        layout.addLayout(edit_box)

        # Parts Search - User-friendly search interface
        search_group = QGroupBox("Parts Search")
        search_main_layout = QVBoxLayout()

        # Row 1: Quick search presets
        presets_layout = QHBoxLayout()
        presets_layout.addWidget(QLabel("Quick Search:"))

        btn_search_all = QPushButton("Show All Parts")
        btn_search_all.setStyleSheet(self.get_button_style("default"))
        btn_search_all.setToolTip("Display all parts in the database")
        btn_search_all.clicked.connect(self.refresh_parts_table)
        presets_layout.addWidget(btn_search_all)

        btn_search_missing_hts = QPushButton("Missing HTS Codes")
        btn_search_missing_hts.setStyleSheet(self.get_button_style("warning"))
        btn_search_missing_hts.setToolTip("Find parts without HTS codes assigned")
        btn_search_missing_hts.clicked.connect(self.search_missing_hts)
        presets_layout.addWidget(btn_search_missing_hts)

        btn_search_invalid_hts = QPushButton("Invalid HTS Codes")
        btn_search_invalid_hts.setStyleSheet(self.get_button_style("danger"))
        btn_search_invalid_hts.setToolTip("Find parts with HTS codes not found in HTS database")
        btn_search_invalid_hts.clicked.connect(self.search_invalid_hts)
        presets_layout.addWidget(btn_search_invalid_hts)

        btn_search_steel = QPushButton("Steel Parts")
        btn_search_steel.setStyleSheet(self.get_button_style("info"))
        btn_search_steel.setToolTip("Find parts with steel content > 0%")
        btn_search_steel.clicked.connect(lambda: self.search_by_material("steel"))
        presets_layout.addWidget(btn_search_steel)

        btn_search_aluminum = QPushButton("Aluminum Parts")
        btn_search_aluminum.setStyleSheet(self.get_button_style("info"))
        btn_search_aluminum.setToolTip("Find parts with aluminum content > 0%")
        btn_search_aluminum.clicked.connect(lambda: self.search_by_material("aluminum"))
        presets_layout.addWidget(btn_search_aluminum)

        presets_layout.addStretch()
        search_main_layout.addLayout(presets_layout)

        # Row 2: Search by field
        field_search_layout = QHBoxLayout()

        field_search_layout.addWidget(QLabel("Search by:"))

        self.search_field_type = QComboBox()
        self.search_field_type.setMinimumWidth(150)
        # User-friendly field names with mapping to database columns
        self.search_field_map = {
            "Part Number": "part_number",
            "Description": "description",
            "HTS Code": "hts_code",
            "Country of Origin": "country_origin",
            "Manufacturer ID": "mid",
            "Client": "client_code",
            "Steel %": "steel_ratio",
            "Aluminum %": "aluminum_ratio",
            "Copper %": "copper_ratio",
            "Wood %": "wood_ratio",
            "Non-Steel %": "non_steel_ratio",
            "Unit of Measure": "qty_unit",
            "Sec 301 Exclusion": "Sec301_Exclusion_Tariff"
        }
        self.search_field_type.addItems(list(self.search_field_map.keys()))
        field_search_layout.addWidget(self.search_field_type)

        self.search_match_type = QComboBox()
        self.search_match_type.setMinimumWidth(120)
        self.search_match_type.addItems(["Contains", "Equals", "Starts with", "Ends with", "Greater than", "Less than"])
        self.search_match_type.setToolTip("How to match the search value")
        field_search_layout.addWidget(self.search_match_type)

        self.search_value_input = QLineEdit()
        self.search_value_input.setPlaceholderText("Enter search value...")
        self.search_value_input.setStyleSheet(self.get_input_style())
        self.search_value_input.returnPressed.connect(self.run_parts_search)
        field_search_layout.addWidget(self.search_value_input, 1)

        btn_search = QPushButton("Search")
        btn_search.setStyleSheet(self.get_button_style("success"))
        btn_search.setToolTip("Search parts with the specified criteria")
        btn_search.clicked.connect(self.run_parts_search)
        field_search_layout.addWidget(btn_search)

        search_main_layout.addLayout(field_search_layout)

        # Row 3: Filter by Client and Country (common filters)
        filter_layout = QHBoxLayout()

        filter_layout.addWidget(QLabel("Filter by Client:"))
        self.filter_client_combo = QComboBox()
        self.filter_client_combo.setMinimumWidth(150)
        self.filter_client_combo.addItem("All Clients", "")
        self.filter_client_combo.currentIndexChanged.connect(self.apply_combined_filters)
        filter_layout.addWidget(self.filter_client_combo)

        filter_layout.addWidget(QLabel("Country:"))
        self.filter_country_combo = QComboBox()
        self.filter_country_combo.setMinimumWidth(120)
        self.filter_country_combo.addItem("All Countries", "")
        self.filter_country_combo.currentIndexChanged.connect(self.apply_combined_filters)
        filter_layout.addWidget(self.filter_country_combo)

        btn_clear_filters = QPushButton("Clear Filters")
        btn_clear_filters.setStyleSheet(self.get_button_style("default"))
        btn_clear_filters.clicked.connect(self.clear_search_filters)
        filter_layout.addWidget(btn_clear_filters)

        filter_layout.addStretch()

        # Advanced SQL toggle
        self.show_advanced_sql = QCheckBox("Show Advanced SQL")
        self.show_advanced_sql.setToolTip("Show custom SQL input for advanced queries")
        self.show_advanced_sql.toggled.connect(self.toggle_advanced_sql)
        filter_layout.addWidget(self.show_advanced_sql)

        search_main_layout.addLayout(filter_layout)

        # Row 4: Advanced SQL (hidden by default)
        self.advanced_sql_widget = QWidget()
        advanced_sql_layout = QHBoxLayout(self.advanced_sql_widget)
        advanced_sql_layout.setContentsMargins(0, 5, 0, 0)

        advanced_sql_layout.addWidget(QLabel("Custom SQL:"))
        self.custom_sql_input = QLineEdit()
        self.custom_sql_input.setPlaceholderText("SELECT * FROM parts_master WHERE part_number LIKE '%ABC%'")
        self.custom_sql_input.setStyleSheet(self.get_input_style())
        self.custom_sql_input.returnPressed.connect(self.run_custom_sql)
        advanced_sql_layout.addWidget(self.custom_sql_input, 1)

        btn_run_sql = QPushButton("Execute SQL")
        btn_run_sql.setStyleSheet(self.get_button_style("info"))
        btn_run_sql.clicked.connect(self.run_custom_sql)
        advanced_sql_layout.addWidget(btn_run_sql)

        self.advanced_sql_widget.setVisible(False)
        search_main_layout.addWidget(self.advanced_sql_widget)

        # Results status bar
        self.search_result_label = QLabel("Ready - Click 'Show All Parts' or enter search criteria")
        self.search_result_label.setStyleSheet("padding: 8px; background: #e8f4f8; border-radius: 4px; color: #2c3e50;")
        search_main_layout.addWidget(self.search_result_label)

        search_group.setLayout(search_main_layout)
        layout.addWidget(search_group)

        # Populate filter dropdowns after UI is set up
        self._populate_search_filters()

        # Table filter - filters displayed rows without querying database
        filter_box = QHBoxLayout()
        filter_box.addWidget(QLabel("Table Filter:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Type to filter displayed rows...")
        self.search_input.setStyleSheet(self.get_input_style())
        self.search_input.textChanged.connect(self.filter_parts_table)
        filter_box.addWidget(self.search_input, 1)
        layout.addLayout(filter_box)

        table_box = QGroupBox("Parts Master Table")
        tl = QVBoxLayout()
        self.parts_table = QTableWidget()
        self.parts_table.setColumnCount(16)
        self.parts_table.setHorizontalHeaderLabels([
            "part_number", "description", "hts_code", "country_origin", "mid", "client_code", "steel_%", "aluminum_%", "copper_%", "wood_%", "auto_%", "non_steel_%", "qty_unit", "hts_verified", "Sec301_Exclusion_Tariff", "updated_date"
        ])
        self.parts_table.setEditTriggers(QTableWidget.AllEditTriggers)
        self.parts_table.setSelectionBehavior(QTableWidget.SelectRows)
        # Allow user to resize columns by dragging, with last column stretching to fill
        self.parts_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.parts_table.horizontalHeader().setStretchLastSection(True)
        self.parts_table.setSortingEnabled(False)  # Disabled for better performance
        # Set reasonable default column widths
        default_widths = [120, 200, 100, 50, 120, 80, 60, 60, 60, 60, 60, 60, 60, 120, 120, 150]
        for i, width in enumerate(default_widths):
            self.parts_table.setColumnWidth(i, width)
        tl.addWidget(self.parts_table)
        table_box.setLayout(tl)
        layout.addWidget(table_box, 1)

        self.refresh_parts_table()
        self.tab_master.setLayout(layout)

    def setup_ocrmill_tab(self):
        """Setup the OCRMill tab for OCR invoice processing"""
        from ocrmill_database import OCRMillDatabase
        from ocrmill_processor import ProcessorEngine, OCRMillConfig
        from ocrmill_worker import OCRMillWorker, SingleFileWorker, MultiFileWorker, ParallelFolderWorker

        # Use QVBoxLayout directly on the tab widget (like Process Shipment tab)
        layout = QVBoxLayout(self.tab_ocrmill)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Initialize OCRMill components
        self.ocrmill_db = OCRMillDatabase(DB_PATH)
        self.ocrmill_config = OCRMillConfig()

        # Load saved settings (normalize paths to OS-native format)
        input_setting = get_user_setting('ocrmill_input_folder', str(BASE_DIR / "OCR_Input"))
        output_setting = get_user_setting('ocrmill_output_folder', str(BASE_DIR / "OCR_Output"))
        self.ocrmill_config.input_folder = Path(input_setting)
        self.ocrmill_config.output_folder = Path(output_setting)
        # Update registry with normalized paths if they were stored with wrong separators
        set_user_setting('ocrmill_input_folder', str(self.ocrmill_config.input_folder))
        set_user_setting('ocrmill_output_folder', str(self.ocrmill_config.output_folder))
        self.ocrmill_config.poll_interval = get_user_setting_int('ocrmill_poll_interval', 60)
        self.ocrmill_config.consolidate_multi_invoice = get_user_setting_bool('ocrmill_consolidate', False)

        self.ocrmill_processor = ProcessorEngine(self.ocrmill_db, self.ocrmill_config, log_callback=self.ocrmill_log)
        self.ocrmill_worker = OCRMillWorker(self.ocrmill_processor)

        # Connect worker signals
        self.ocrmill_worker.log_message.connect(self.ocrmill_log)
        self.ocrmill_worker.processing_finished.connect(self.ocrmill_on_processing_finished)
        self.ocrmill_worker.error.connect(lambda e: self.ocrmill_log(f"Error: {e}"))
        self.ocrmill_worker.items_extracted.connect(self.ocrmill_on_items_extracted)

        # Create sub-tabs
        self.ocrmill_tabs = QTabWidget()

        # ===== TAB 1: INVOICE PROCESSING =====
        processing_widget = QWidget()
        processing_layout = QVBoxLayout(processing_widget)

        # Create splitter for resizable panels (like Process Shipment tab)
        splitter = QSplitter(Qt.Horizontal)

        # LEFT SIDE: Controls box with scroll area
        left_outer_box = QGroupBox("Controls")
        left_outer_layout = QVBoxLayout(left_outer_box)
        left_outer_layout.setContentsMargins(0, 0, 0, 0)

        # Create scroll area for left side content
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        left_scroll.setFrameShape(QFrame.NoFrame)

        left_scroll_widget = QWidget()
        left_side = QVBoxLayout(left_scroll_widget)
        left_side.setSpacing(10)
        left_side.setContentsMargins(10, 10, 10, 10)

        # Folder configuration
        folder_group = QGroupBox("Folder Configuration")
        folder_layout = QFormLayout()

        # Input folder
        input_row = QHBoxLayout()
        self.ocrmill_input_edit = QLineEdit(str(self.ocrmill_config.input_folder))
        self.ocrmill_input_edit.setReadOnly(True)
        input_browse_btn = QPushButton("...")
        input_browse_btn.setFixedWidth(30)
        input_browse_btn.clicked.connect(self.ocrmill_browse_input_folder)
        input_row.addWidget(self.ocrmill_input_edit)
        input_row.addWidget(input_browse_btn)
        folder_layout.addRow("Input:", input_row)

        # Output folder
        output_row = QHBoxLayout()
        self.ocrmill_output_edit = QLineEdit(str(self.ocrmill_config.output_folder))
        self.ocrmill_output_edit.setReadOnly(True)
        output_browse_btn = QPushButton("...")
        output_browse_btn.setFixedWidth(30)
        output_browse_btn.clicked.connect(self.ocrmill_browse_output_folder)
        output_row.addWidget(self.ocrmill_output_edit)
        output_row.addWidget(output_browse_btn)
        folder_layout.addRow("Output:", output_row)

        folder_group.setLayout(folder_layout)
        left_side.addWidget(folder_group)

        # Input files list
        input_files_group = QGroupBox("Input Files (PDFs)")
        input_files_layout = QVBoxLayout()
        self.ocrmill_input_files_list = QListWidget()
        self.ocrmill_input_files_list.setSelectionMode(QListWidget.SingleSelection)
        self.ocrmill_input_files_list.itemDoubleClicked.connect(self.ocrmill_open_input_file)
        self.ocrmill_input_files_list.setFixedHeight(75)
        input_files_layout.addWidget(self.ocrmill_input_files_list)
        refresh_input_btn = QPushButton("Refresh")
        refresh_input_btn.setFixedHeight(25)
        refresh_input_btn.clicked.connect(self.ocrmill_refresh_input_files)
        input_files_layout.addWidget(refresh_input_btn)
        input_files_group.setLayout(input_files_layout)
        left_side.addWidget(input_files_group)

        # Output files list
        output_files_group = QGroupBox("Output Files (CSVs)")
        output_files_layout = QVBoxLayout()
        self.ocrmill_output_files_list = QListWidget()
        self.ocrmill_output_files_list.setSelectionMode(QListWidget.SingleSelection)
        self.ocrmill_output_files_list.itemClicked.connect(self.ocrmill_preview_output_file)
        self.ocrmill_output_files_list.itemDoubleClicked.connect(self.ocrmill_open_output_file)
        self.ocrmill_output_files_list.setFixedHeight(75)
        output_files_layout.addWidget(self.ocrmill_output_files_list)
        refresh_output_btn = QPushButton("Refresh")
        refresh_output_btn.setFixedHeight(25)
        refresh_output_btn.clicked.connect(self.ocrmill_refresh_output_files)
        output_files_layout.addWidget(refresh_output_btn)
        output_files_group.setLayout(output_files_layout)
        left_side.addWidget(output_files_group)

        # Actions group
        actions_group = QGroupBox("Actions")
        actions_layout = QVBoxLayout()

        # Monitoring row with button and auto-start checkbox
        monitor_row = QHBoxLayout()
        self.ocrmill_monitor_btn = QPushButton("Start Monitoring")
        self.ocrmill_monitor_btn.setCheckable(True)
        self.ocrmill_monitor_btn.setFixedHeight(28)
        self.ocrmill_monitor_btn.clicked.connect(self.ocrmill_toggle_monitoring)
        monitor_row.addWidget(self.ocrmill_monitor_btn)

        self.ocrmill_autostart_check = QCheckBox("Auto-start")
        self.ocrmill_autostart_check.setToolTip("Auto-start monitoring on application launch")
        self.ocrmill_autostart_check.setChecked(get_user_setting_bool('ocrmill_autostart', False))
        self.ocrmill_autostart_check.stateChanged.connect(lambda s: set_user_setting('ocrmill_autostart', 'true' if s else 'false'))
        monitor_row.addWidget(self.ocrmill_autostart_check)
        actions_layout.addLayout(monitor_row)

        process_file_btn = QPushButton("Process PDF File...")
        process_file_btn.setFixedHeight(28)
        process_file_btn.clicked.connect(self.ocrmill_process_single_file)
        actions_layout.addWidget(process_file_btn)

        process_folder_btn = QPushButton("Process Folder Now")
        process_folder_btn.setFixedHeight(28)
        process_folder_btn.clicked.connect(self.ocrmill_process_folder_now)
        actions_layout.addWidget(process_folder_btn)

        self.ocrmill_send_btn = QPushButton("Send to Tariffmill")
        self.ocrmill_send_btn.setFixedHeight(28)
        self.ocrmill_send_btn.setEnabled(False)
        self.ocrmill_send_btn.clicked.connect(self.ocrmill_send_to_tariffmill)
        self.ocrmill_send_btn.setToolTip("Move CSV files from OCR_Output to Tariffmill_Input")
        actions_layout.addWidget(self.ocrmill_send_btn)

        # Output mode for multi-invoice PDFs
        output_mode_label = QLabel("Multi-invoice PDFs:")
        output_mode_label.setStyleSheet("font-weight: bold; margin-top: 5px;")
        actions_layout.addWidget(output_mode_label)

        self.ocrmill_output_mode_group = QButtonGroup(self)

        self.ocrmill_split_output_radio = QRadioButton("Split (one CSV per invoice)")
        self.ocrmill_split_output_radio.setToolTip("Each invoice in the PDF will be saved as a separate CSV file")
        self.ocrmill_output_mode_group.addButton(self.ocrmill_split_output_radio, 0)
        actions_layout.addWidget(self.ocrmill_split_output_radio)

        self.ocrmill_single_output_radio = QRadioButton("Combine (all invoices together)")
        self.ocrmill_single_output_radio.setToolTip("All invoices in the PDF will be combined into one CSV file")
        self.ocrmill_output_mode_group.addButton(self.ocrmill_single_output_radio, 1)
        actions_layout.addWidget(self.ocrmill_single_output_radio)

        # Set initial state from config
        if get_user_setting_bool('ocrmill_consolidate', False):
            self.ocrmill_single_output_radio.setChecked(True)
        else:
            self.ocrmill_split_output_radio.setChecked(True)

        self.ocrmill_output_mode_group.buttonClicked.connect(self.ocrmill_output_mode_changed)

        actions_group.setLayout(actions_layout)
        left_side.addWidget(actions_group)

        # Finish scroll area setup
        left_scroll.setWidget(left_scroll_widget)
        left_outer_layout.addWidget(left_scroll)

        # Set min/max width for left controls
        left_outer_box.setMinimumWidth(300)
        left_outer_box.setMaximumWidth(400)

        # Add left_outer_box to splitter
        splitter.addWidget(left_outer_box)

        # RIGHT SIDE: Activity Log, Results Preview, and Drop Zone
        right_widget = QWidget()
        right_side = QVBoxLayout(right_widget)
        right_side.setContentsMargins(0, 0, 0, 0)

        # PDF Drop Zone at top
        self.ocrmill_drop_zone = PDFDropZone(str(self.ocrmill_config.input_folder))
        self.ocrmill_drop_zone.files_dropped.connect(self.ocrmill_process_dropped_files)
        self.ocrmill_drop_zone.setFixedHeight(80)
        right_side.addWidget(self.ocrmill_drop_zone)

        # Vertical splitter for Activity Log and Results Preview
        right_splitter = QSplitter(Qt.Vertical)
        right_splitter.setChildrenCollapsible(False)

        # Activity log
        log_group = QGroupBox("Activity Log")
        log_layout = QVBoxLayout()
        self.ocrmill_log_text = QPlainTextEdit()
        self.ocrmill_log_text.setReadOnly(True)
        self.ocrmill_log_text.setMaximumBlockCount(1000)
        log_layout.addWidget(self.ocrmill_log_text)

        log_btn_layout = QHBoxLayout()
        clear_log_btn = QPushButton("Clear Log")
        clear_log_btn.clicked.connect(self.ocrmill_log_text.clear)
        log_btn_layout.addStretch()
        log_btn_layout.addWidget(clear_log_btn)
        log_layout.addLayout(log_btn_layout)

        log_group.setLayout(log_layout)
        right_splitter.addWidget(log_group)

        # Results Preview
        preview_group = QGroupBox("Results Preview")
        preview_layout = QVBoxLayout()

        self.ocrmill_preview_table = QTableWidget()
        self.ocrmill_preview_table.setColumnCount(5)
        self.ocrmill_preview_table.setHorizontalHeaderLabels([
            "Part Number", "Description", "Quantity", "Unit Price", "Total"
        ])
        self.ocrmill_preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.ocrmill_preview_table.horizontalHeader().setStretchLastSection(True)
        self.ocrmill_preview_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.ocrmill_preview_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.ocrmill_preview_table.setAlternatingRowColors(True)
        self.ocrmill_preview_table.verticalHeader().setDefaultSectionSize(22)
        self.ocrmill_preview_table.verticalHeader().setFixedWidth(30)
        # Enable edit tracking for learning from corrections
        self.ocrmill_preview_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.ocrmill_preview_table.itemChanged.connect(self._on_ocrmill_item_changed)
        self._ocrmill_editing = False  # Flag to prevent recursive signals during population
        preview_layout.addWidget(self.ocrmill_preview_table)

        preview_group.setLayout(preview_layout)
        right_splitter.addWidget(preview_group)

        # Set initial splitter sizes (log: 200, preview: 300)
        right_splitter.setSizes([200, 300])

        right_side.addWidget(right_splitter, 1)

        # Add right widget to splitter
        splitter.addWidget(right_widget)

        # Set initial splitter sizes (left: 350, right: remaining)
        splitter.setSizes([350, 650])

        processing_layout.addWidget(splitter, 1)

        self.ocrmill_tabs.addTab(processing_widget, "Invoice Processing")

        # ===== TAB 2: AI TEMPLATE GENERATOR (integrated editor) =====
        self.setup_ai_template_tab()
        self.ocrmill_tabs.addTab(self.ai_template_widget, "AI Templates")

        # ===== TAB 3: PROCESSING STATISTICS =====
        self.setup_processing_stats_tab()
        self.ocrmill_tabs.addTab(self.processing_stats_widget, "Statistics")

        layout.addWidget(self.ocrmill_tabs, 1)

        # Store extracted items for "Send to Tariffmill"
        self.ocrmill_last_items = []

        # Initial log message
        self.ocrmill_log("OCRMill initialized. Ready to process PDF invoices.")

        # Initial file list refresh
        self.ocrmill_refresh_input_files()
        self.ocrmill_refresh_output_files()

    def ocrmill_log(self, message: str):
        """Add a message to the OCRMill activity log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.ocrmill_log_text.appendPlainText(f"[{timestamp}] {message}")

    def setup_ai_template_tab(self):
        """Setup the integrated AI Template Generator tab with template list, editor, and chat."""
        self.ai_template_widget = QWidget()
        layout = QVBoxLayout(self.ai_template_widget)
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        # Main horizontal splitter: Left (template list) | Right (editor + chat)
        main_splitter = QSplitter(Qt.Horizontal)

        # === LEFT PANEL: Template List ===
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(5)

        # Template list header
        self.ai_templates_header = QLabel("Templates")
        self.ai_templates_header.setStyleSheet("font-weight: bold; font-size: 12px;")
        left_layout.addWidget(self.ai_templates_header)

        # Template list
        self.ocrmill_templates_list = QListWidget()
        self.ocrmill_templates_list.setStyleSheet("""
            QListWidget {
                background-color: #252526;
                color: #cccccc;
                border: 1px solid #3c3c3c;
                border-radius: 4px;
                padding: 4px;
            }
            QListWidget::item {
                padding: 8px 10px;
                border-bottom: 1px solid #3c3c3c;
            }
            QListWidget::item:selected {
                background-color: #094771;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #2a2d2e;
            }
        """)
        self.ocrmill_templates_list.itemClicked.connect(self._on_template_selected)
        self.ocrmill_templates_list.itemDoubleClicked.connect(self._on_template_double_clicked)
        left_layout.addWidget(self.ocrmill_templates_list, 1)

        # Template list buttons
        list_btn_layout = QHBoxLayout()
        list_btn_layout.setSpacing(5)

        btn_new = QPushButton("New")
        btn_new.setToolTip("Create a new template using AI")
        btn_new.setStyleSheet(self._get_small_button_style("success"))
        btn_new.clicked.connect(self._ai_template_new)
        list_btn_layout.addWidget(btn_new)

        btn_duplicate = QPushButton("Duplicate")
        btn_duplicate.setToolTip("Create a copy of the selected template")
        btn_duplicate.setStyleSheet(self._get_small_button_style("default"))
        btn_duplicate.clicked.connect(self._duplicate_template)
        list_btn_layout.addWidget(btn_duplicate)

        btn_delete = QPushButton("Delete")
        btn_delete.setToolTip("Delete selected template")
        btn_delete.setStyleSheet(self._get_small_button_style("danger"))
        btn_delete.clicked.connect(self.ocrmill_delete_template)
        list_btn_layout.addWidget(btn_delete)

        btn_refresh = QPushButton()
        btn_refresh.setToolTip("Refresh template list")
        btn_refresh.setFixedWidth(30)
        btn_refresh.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        btn_refresh.setStyleSheet(self._get_small_button_style("default"))
        btn_refresh.clicked.connect(self.ocrmill_refresh_templates)
        list_btn_layout.addWidget(btn_refresh)

        left_layout.addLayout(list_btn_layout)

        # Store template data
        self.ocrmill_templates_data = []

        main_splitter.addWidget(left_panel)

        # === RIGHT PANEL: Stacked Widget (Edit Mode / Create Mode) ===
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(5)

        # Header bar with template info and AI provider
        self.ai_template_header = QWidget()
        self.ai_template_header.setStyleSheet("background-color: #2d2d30; border-radius: 4px;")
        header_layout = QHBoxLayout(self.ai_template_header)
        header_layout.setContentsMargins(10, 6, 10, 6)

        self.ai_template_name_label = QLabel("Select a template")
        self.ai_template_name_label.setStyleSheet("font-weight: bold; color: #dcdcdc;")
        header_layout.addWidget(self.ai_template_name_label)

        header_layout.addStretch()

        # AI Provider selection
        self.ai_provider_label = QLabel("AI:")
        header_layout.addWidget(self.ai_provider_label)
        self.ai_provider_combo = QComboBox()
        self.ai_provider_combo.addItems(["Anthropic", "OpenAI", "Google Gemini", "Groq"])
        self.ai_provider_combo.setFixedWidth(120)
        self.ai_provider_combo.currentTextChanged.connect(self._on_ai_provider_changed)
        header_layout.addWidget(self.ai_provider_combo)

        self.ai_model_combo = QComboBox()
        self.ai_model_combo.setFixedWidth(180)
        self._on_ai_provider_changed(self.ai_provider_combo.currentText())
        header_layout.addWidget(self.ai_model_combo)

        right_layout.addWidget(self.ai_template_header)

        # Stacked widget for Edit/Create modes
        self.ai_mode_stack = QStackedWidget()

        # === PAGE 0: EDIT MODE (Editor + Chat) ===
        edit_mode_widget = QWidget()
        edit_mode_layout = QVBoxLayout(edit_mode_widget)
        edit_mode_layout.setContentsMargins(0, 0, 0, 0)
        edit_mode_layout.setSpacing(5)

        # Splitter for code editor and chat
        editor_chat_splitter = QSplitter(Qt.Horizontal)

        # Code editor
        code_widget = QWidget()
        code_layout = QVBoxLayout(code_widget)
        code_layout.setContentsMargins(0, 0, 0, 0)
        code_layout.setSpacing(3)

        code_header = QHBoxLayout()
        self.ai_code_label = QLabel("Template Code")
        code_header.addWidget(self.ai_code_label)
        code_header.addStretch()
        self.ai_syntax_indicator = QLabel("✓ Valid")
        self.ai_syntax_indicator.setStyleSheet("color: #4ec9b0; font-size: 10px;")
        code_header.addWidget(self.ai_syntax_indicator)
        code_layout.addLayout(code_header)

        self.ai_code_edit = PythonCodeEditor()
        self.ai_code_edit.setFont(QFont("Consolas", 10))
        self.ai_code_edit.setStyleSheet("""
            QPlainTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                border: 1px solid #3c3c3c;
                border-radius: 4px;
                padding: 6px;
                selection-background-color: #264f78;
            }
        """)
        self.ai_code_edit.setLineWrapMode(QPlainTextEdit.NoWrap)
        self.ai_code_edit.textChanged.connect(self._ai_validate_syntax)
        self.ai_code_edit.setPlaceholderText("Select a template from the list to edit, or click 'New' to create one...")
        code_layout.addWidget(self.ai_code_edit, 1)

        editor_chat_splitter.addWidget(code_widget)

        # Chat panel
        self.ai_chat_widget = QWidget()
        self.ai_chat_widget.setStyleSheet("background-color: #252526; border-radius: 4px;")
        chat_layout = QVBoxLayout(self.ai_chat_widget)
        chat_layout.setContentsMargins(8, 8, 8, 8)
        chat_layout.setSpacing(5)

        self.ai_chat_label = QLabel("AI Assistant")
        chat_layout.addWidget(self.ai_chat_label)

        self.ai_chat_display = QTextEdit()
        self.ai_chat_display.setReadOnly(True)
        self.ai_chat_display.setFont(QFont("Segoe UI", 9))
        self.ai_chat_display.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #cccccc;
                border: 1px solid #3c3c3c;
                border-radius: 4px;
                padding: 6px;
            }
            QScrollBar:vertical {
                background-color: #1e1e1e;
                width: 8px;
            }
            QScrollBar::handle:vertical {
                background-color: #5a5a5a;
                border-radius: 4px;
            }
        """)
        self._show_ai_welcome_message()
        chat_layout.addWidget(self.ai_chat_display, 1)

        # Context indicator - shows which template is being used
        self.ai_context_label = QLabel()
        self.ai_context_label.setStyleSheet("""
            QLabel {
                color: #888888;
                font-size: 9pt;
                padding: 2px 4px;
            }
        """)
        self._update_ai_context_label()
        chat_layout.addWidget(self.ai_context_label)

        # Message input (Enter to send, Shift+Enter for newline)
        self.ai_message_input = ChatMessageInput()
        self.ai_message_input.setMaximumHeight(60)
        self.ai_message_input.setPlaceholderText("Ask the AI to modify the template... (Enter to send)")
        self.ai_message_input.sendRequested.connect(self._ai_send_message)
        self.ai_message_input.setStyleSheet("""
            QPlainTextEdit {
                background-color: #3c3c3c;
                color: #cccccc;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 6px;
                font-size: 10pt;
            }
            QPlainTextEdit:focus {
                border: 1px solid #007acc;
            }
        """)
        chat_layout.addWidget(self.ai_message_input)

        # Chat buttons
        chat_btn_layout = QHBoxLayout()
        chat_btn_layout.setSpacing(5)

        self.ai_send_btn = QPushButton("Send")
        self.ai_send_btn.setStyleSheet(self._get_small_button_style("primary"))
        self.ai_send_btn.clicked.connect(self._ai_send_message)
        chat_btn_layout.addWidget(self.ai_send_btn)

        self.ai_cancel_btn = QPushButton("Cancel")
        self.ai_cancel_btn.setStyleSheet(self._get_small_button_style("default"))
        self.ai_cancel_btn.setEnabled(False)
        self.ai_cancel_btn.clicked.connect(self._ai_cancel_request)
        chat_btn_layout.addWidget(self.ai_cancel_btn)

        chat_btn_layout.addStretch()

        self.ai_apply_btn = QPushButton("Apply Code")
        self.ai_apply_btn.setStyleSheet(self._get_small_button_style("success"))
        self.ai_apply_btn.setEnabled(False)
        self.ai_apply_btn.clicked.connect(self._ai_apply_changes)
        chat_btn_layout.addWidget(self.ai_apply_btn)

        chat_layout.addLayout(chat_btn_layout)

        editor_chat_splitter.addWidget(self.ai_chat_widget)
        editor_chat_splitter.setSizes([500, 350])

        edit_mode_layout.addWidget(editor_chat_splitter, 1)

        # Bottom action bar for edit mode
        action_layout = QHBoxLayout()
        action_layout.setSpacing(8)

        btn_test = QPushButton("Test Template")
        btn_test.setStyleSheet(self._get_small_button_style("info"))
        btn_test.clicked.connect(self._ai_test_template)
        action_layout.addWidget(btn_test)

        btn_format = QPushButton("Format Code")
        btn_format.setStyleSheet(self._get_small_button_style("default"))
        btn_format.clicked.connect(self._ai_format_code)
        action_layout.addWidget(btn_format)

        action_layout.addStretch()

        self.ai_save_btn = QPushButton("Save Template")
        self.ai_save_btn.setStyleSheet(self._get_small_button_style("success"))
        self.ai_save_btn.clicked.connect(self._ai_save_template)
        action_layout.addWidget(self.ai_save_btn)

        edit_mode_layout.addLayout(action_layout)

        self.ai_mode_stack.addWidget(edit_mode_widget)  # Index 0: Edit mode

        # === PAGE 1: CREATE MODE (New Template from Sample) ===
        create_mode_widget = QWidget()
        create_mode_layout = QVBoxLayout(create_mode_widget)
        create_mode_layout.setContentsMargins(0, 0, 0, 0)
        create_mode_layout.setSpacing(8)

        # Sample Invoice Input
        sample_group = QGroupBox("Sample Invoice")
        sample_layout = QVBoxLayout()

        # File selection row
        file_row = QHBoxLayout()
        self.ai_new_pdf_path = QLineEdit()
        self.ai_new_pdf_path.setPlaceholderText("Select a PDF invoice or paste text below...")
        self.ai_new_pdf_path.setReadOnly(True)
        file_row.addWidget(self.ai_new_pdf_path, stretch=1)

        btn_browse_pdf = QPushButton("Browse PDF...")
        btn_browse_pdf.setStyleSheet(self._get_small_button_style("primary"))
        btn_browse_pdf.clicked.connect(self._ai_browse_sample_pdf)
        file_row.addWidget(btn_browse_pdf)

        sample_layout.addLayout(file_row)

        self.ai_new_invoice_text = QPlainTextEdit()
        self.ai_new_invoice_text.setPlaceholderText(
            "Paste invoice text here, or load from PDF above.\n\n"
            "The AI will analyze this text to create extraction patterns."
        )
        self.ai_new_invoice_text.setFont(QFont("Consolas", 9))
        sample_layout.addWidget(self.ai_new_invoice_text)

        sample_group.setLayout(sample_layout)
        create_mode_layout.addWidget(sample_group, 1)

        # Template Settings
        settings_group = QGroupBox("Template Settings")
        settings_layout = QFormLayout()

        self.ai_new_template_name = QLineEdit()
        self.ai_new_template_name.setPlaceholderText("e.g., acme_corp (lowercase with underscores)")
        settings_layout.addRow("Template Name:", self.ai_new_template_name)

        self.ai_new_supplier_name = QLineEdit()
        self.ai_new_supplier_name.setPlaceholderText("e.g., Acme Corporation")
        settings_layout.addRow("Supplier Name:", self.ai_new_supplier_name)

        self.ai_new_client = QLineEdit()
        self.ai_new_client.setPlaceholderText("e.g., Sigma Corporation")
        settings_layout.addRow("Client:", self.ai_new_client)

        self.ai_new_country = QLineEdit()
        self.ai_new_country.setPlaceholderText("e.g., CHINA, INDIA, USA")
        settings_layout.addRow("Country of Origin:", self.ai_new_country)

        settings_group.setLayout(settings_layout)
        create_mode_layout.addWidget(settings_group)

        # Progress bar
        self.ai_new_progress = QProgressBar()
        self.ai_new_progress.setVisible(False)
        self.ai_new_progress.setTextVisible(True)
        create_mode_layout.addWidget(self.ai_new_progress)

        # Generated Code Preview
        preview_group = QGroupBox("Generated Template (Preview)")
        preview_layout = QVBoxLayout()

        self.ai_new_code_preview = QPlainTextEdit()
        self.ai_new_code_preview.setReadOnly(True)
        self.ai_new_code_preview.setFont(QFont("Consolas", 9))
        self.ai_new_code_preview.setPlaceholderText("Generated template code will appear here...")
        preview_layout.addWidget(self.ai_new_code_preview)

        preview_group.setLayout(preview_layout)
        create_mode_layout.addWidget(preview_group, 1)

        # Create mode action buttons
        create_action_layout = QHBoxLayout()
        create_action_layout.setSpacing(8)

        self.ai_generate_btn = QPushButton("Generate Template")
        self.ai_generate_btn.setStyleSheet(self._get_small_button_style("primary"))
        self.ai_generate_btn.clicked.connect(self._ai_generate_new_template)
        create_action_layout.addWidget(self.ai_generate_btn)

        self.ai_cancel_gen_btn = QPushButton("Cancel")
        self.ai_cancel_gen_btn.setStyleSheet(self._get_small_button_style("danger"))
        self.ai_cancel_gen_btn.setVisible(False)
        self.ai_cancel_gen_btn.clicked.connect(self._ai_cancel_generation)
        create_action_layout.addWidget(self.ai_cancel_gen_btn)

        self.ai_save_new_btn = QPushButton("Save Template")
        self.ai_save_new_btn.setStyleSheet(self._get_small_button_style("success"))
        self.ai_save_new_btn.setEnabled(False)
        self.ai_save_new_btn.clicked.connect(self._ai_save_new_template)
        create_action_layout.addWidget(self.ai_save_new_btn)

        create_action_layout.addStretch()

        btn_back_to_edit = QPushButton("Back to Templates")
        btn_back_to_edit.setStyleSheet(self._get_small_button_style("default"))
        btn_back_to_edit.clicked.connect(self._ai_switch_to_edit_mode)
        create_action_layout.addWidget(btn_back_to_edit)

        create_mode_layout.addLayout(create_action_layout)

        self.ai_mode_stack.addWidget(create_mode_widget)  # Index 1: Create mode

        right_layout.addWidget(self.ai_mode_stack, 1)

        main_splitter.addWidget(right_panel)

        # Set splitter proportions (20% list, 80% editor+chat)
        main_splitter.setSizes([200, 800])

        layout.addWidget(main_splitter, 1)

        # Initialize state
        self.ai_current_template_path = None
        self.ai_conversation_history = []
        self.ai_chat_thread = None
        self.ai_pending_code = None

        # Load templates
        self.ocrmill_refresh_templates()

        # Apply theme-aware styles
        self._update_ai_template_styles()

    def _get_ai_theme_colors(self) -> dict:
        """Get theme-aware colors for the AI template tab."""
        theme = getattr(self, 'current_theme', 'System Default')

        if theme == "Ocean":
            return {
                'bg_primary': '#1a3050',
                'bg_secondary': '#243d5c',
                'bg_input': '#152a42',
                'bg_code': '#0d1f30',
                'border': '#3a6a9a',
                'border_focus': '#00a8cc',
                'text': '#e0f0ff',
                'text_muted': '#8ac4e0',
                'text_header': '#7ec8e3',
                'accent': '#00a8cc',
                'success': '#2e8b57',
                'danger': '#c44536',
                'warning': '#d4a574',
                'selection': '#0096b4',
                'user_bubble': '#1a5a7a',
                'user_bubble_text': '#ffffff',
                'ai_bubble': '#0d2840',
                'ai_bubble_text': '#e0f0ff',
            }
        elif theme == "Fusion (Dark)":
            return {
                'bg_primary': '#2d2d2d',
                'bg_secondary': '#353535',
                'bg_input': '#3c3c3c',
                'bg_code': '#1e1e1e',
                'border': '#555555',
                'border_focus': '#0078d4',
                'text': '#e0e0e0',
                'text_muted': '#a0a0a0',
                'text_header': '#ffffff',
                'accent': '#0078d4',
                'success': '#388a34',
                'danger': '#c42b1c',
                'warning': '#ce9178',
                'selection': '#264f78',
                'user_bubble': '#264f78',
                'user_bubble_text': '#ffffff',
                'ai_bubble': '#2d2d2d',
                'ai_bubble_text': '#e0e0e0',
            }
        elif theme == "Light Cyan":
            return {
                'bg_primary': '#e0f6f7',
                'bg_secondary': '#c8e8ec',
                'bg_input': '#ffffff',
                'bg_code': '#f0fafa',
                'border': '#90c8d0',
                'border_focus': '#00a8b8',
                'text': '#1a3a3a',
                'text_muted': '#4a7a7a',
                'text_header': '#0a2a2a',
                'accent': '#00a8b8',
                'success': '#107c50',
                'danger': '#c43838',
                'warning': '#b87000',
                'selection': '#a0d8e0',
                'user_bubble': '#00a8b8',
                'user_bubble_text': '#ffffff',
                'ai_bubble': '#d0ecf0',
                'ai_bubble_text': '#1a3a3a',
            }
        elif theme == "macOS":
            return {
                'bg_primary': '#f6f6f6',
                'bg_secondary': '#ffffff',
                'bg_input': '#ffffff',
                'bg_code': '#f5f5f7',
                'border': '#d1d1d6',
                'border_focus': '#007aff',
                'text': '#1d1d1f',
                'text_muted': '#8e8e93',
                'text_header': '#1d1d1f',
                'accent': '#007aff',
                'success': '#34c759',
                'danger': '#ff3b30',
                'warning': '#ff9500',
                'selection': '#007aff',
                'user_bubble': '#007aff',
                'user_bubble_text': '#ffffff',
                'ai_bubble': '#e5e5ea',
                'ai_bubble_text': '#1d1d1f',
            }
        else:  # Light themes (System Default, Fusion Light)
            return {
                'bg_primary': '#ffffff',
                'bg_secondary': '#f5f5f5',
                'bg_input': '#ffffff',
                'bg_code': '#fafafa',
                'border': '#d0d0d0',
                'border_focus': '#0078d4',
                'text': '#1a1a1a',
                'text_muted': '#666666',
                'text_header': '#333333',
                'accent': '#0078d4',
                'success': '#107c10',
                'danger': '#d13438',
                'warning': '#b85c00',
                'selection': '#cce8ff',
                'user_bubble': '#0078d4',
                'user_bubble_text': '#ffffff',
                'ai_bubble': '#f0f0f0',
                'ai_bubble_text': '#1a1a1a',
            }

    def _get_small_button_style(self, style_type: str) -> str:
        """Get button stylesheet for the AI template tab."""
        colors = self._get_ai_theme_colors()
        is_dark = getattr(self, 'current_theme', '') in ["Fusion (Dark)", "Ocean"]

        button_colors = {
            "primary": (colors['accent'], self._lighten_color(colors['accent']), self._darken_color(colors['accent'])),
            "success": (colors['success'], self._lighten_color(colors['success']), self._darken_color(colors['success'])),
            "danger": (colors['danger'], self._lighten_color(colors['danger']), self._darken_color(colors['danger'])),
            "info": ("#6c5ce7", "#7d6ee8", "#5b4ccc"),
            "default": (colors['bg_input'], colors['bg_secondary'], colors['bg_primary']),
        }
        bg, hover, pressed = button_colors.get(style_type, button_colors["default"])
        text_color = "#ffffff" if style_type != "default" else colors['text']
        disabled_bg = colors['bg_primary'] if is_dark else "#e0e0e0"
        disabled_text = colors['text_muted']

        return f"""
            QPushButton {{
                background-color: {bg};
                color: {text_color};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                padding: 6px 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{ background-color: {hover}; }}
            QPushButton:pressed {{ background-color: {pressed}; }}
            QPushButton:disabled {{ background-color: {disabled_bg}; color: {disabled_text}; }}
        """

    def _lighten_color(self, hex_color: str) -> str:
        """Lighten a hex color by 15%."""
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        r = min(255, int(r + (255 - r) * 0.15))
        g = min(255, int(g + (255 - g) * 0.15))
        b = min(255, int(b + (255 - b) * 0.15))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _darken_color(self, hex_color: str) -> str:
        """Darken a hex color by 15%."""
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        r = max(0, int(r * 0.85))
        g = max(0, int(g * 0.85))
        b = max(0, int(b * 0.85))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _show_ai_welcome_message(self):
        """Show welcome message in AI chat."""
        colors = self._get_ai_theme_colors()
        base_font_size = get_user_setting_int('font_size', 9)
        font_size = base_font_size + 1  # Chat text slightly larger for readability
        header_size = font_size + 2
        self.ai_chat_display.setHtml(f'''
        <div style="font-family: Segoe UI; padding: 8px;">
            <div style="background-color: {colors['bg_secondary']}; border-radius: 6px; padding: 12px; border: 1px solid {colors['border']};">
                <div style="color: {colors['accent']}; font-size: {header_size}pt; font-weight: bold; margin-bottom: 8px;">
                    👋 AI Template Assistant
                </div>
                <div style="color: {colors['text']}; line-height: 1.5; font-size: {font_size}pt;">
                    <p>Select a template to edit, or create a new one.</p>
                    <p>I can help you:</p>
                    <ul style="margin: 5px 0; padding-left: 15px;">
                        <li>Fix regex patterns</li>
                        <li>Add new extraction fields</li>
                        <li>Handle new invoice formats</li>
                        <li>Explain template logic</li>
                    </ul>
                </div>
            </div>
        </div>
        ''')

    def _update_ai_context_label(self):
        """Update the context indicator label to show current template."""
        if not hasattr(self, 'ai_context_label'):
            return

        colors = self._get_ai_theme_colors()

        if getattr(self, 'ai_current_template_path', None):
            template_name = self.ai_current_template_path.stem
            self.ai_context_label.setText(f"📎 Context: {template_name}")
            self.ai_context_label.setStyleSheet(f"""
                QLabel {{
                    color: {colors['accent']};
                    font-size: 9pt;
                    padding: 2px 4px;
                    background-color: {colors['bg_secondary']};
                    border-radius: 3px;
                }}
            """)
            self.ai_context_label.setToolTip(f"AI has access to template: {self.ai_current_template_path}")
        else:
            self.ai_context_label.setText("📎 No template selected")
            self.ai_context_label.setStyleSheet(f"""
                QLabel {{
                    color: {colors['text_muted']};
                    font-size: 9pt;
                    padding: 2px 4px;
                }}
            """)
            self.ai_context_label.setToolTip("Select a template to provide context for the AI")

    def setup_processing_stats_tab(self):
        """Setup the Processing Statistics tab to display template usage and processing metrics."""
        self.processing_stats_widget = QWidget()
        layout = QVBoxLayout(self.processing_stats_widget)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Header with refresh button
        header_layout = QHBoxLayout()
        self.stats_header_label = QLabel("Processing Statistics")
        header_layout.addWidget(self.stats_header_label)
        header_layout.addStretch()

        self.stats_refresh_btn = QPushButton("Refresh")
        self.stats_refresh_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        self.stats_refresh_btn.clicked.connect(self._refresh_processing_stats)
        header_layout.addWidget(self.stats_refresh_btn)
        layout.addLayout(header_layout)

        # Summary cards
        self.stats_summary_frame = QFrame()
        summary_layout = QHBoxLayout(self.stats_summary_frame)

        # Today's stats
        self.stats_today_label = self._create_stat_card("Today", "0 files", "0 items")
        summary_layout.addWidget(self.stats_today_label)

        # This week's stats
        self.stats_week_label = self._create_stat_card("This Week", "0 files", "0 items")
        summary_layout.addWidget(self.stats_week_label)

        # All time stats
        self.stats_total_label = self._create_stat_card("All Time", "0 files", "0 items")
        summary_layout.addWidget(self.stats_total_label)

        # Success rate
        self.stats_success_label = self._create_stat_card("Success Rate", "0%", "0 templates")
        summary_layout.addWidget(self.stats_success_label)

        layout.addWidget(self.stats_summary_frame)

        # Template statistics table
        self.stats_table_label = QLabel("Template Usage Statistics")
        layout.addWidget(self.stats_table_label)

        self.template_stats_table = QTableWidget()
        self.template_stats_table.setColumnCount(7)
        self.template_stats_table.setHorizontalHeaderLabels([
            "Template", "Total Uses", "Successful", "Items Extracted",
            "Avg Confidence", "Avg Time (ms)", "Last Used"
        ])
        self.template_stats_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(1, 7):
            self.template_stats_table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.template_stats_table.setAlternatingRowColors(True)
        self.template_stats_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.template_stats_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.template_stats_table.verticalHeader().setVisible(False)
        layout.addWidget(self.template_stats_table, 1)

        # Recent activity section
        self.stats_recent_label = QLabel("Recent Processing Activity")
        layout.addWidget(self.stats_recent_label)

        self.recent_activity_table = QTableWidget()
        self.recent_activity_table.setColumnCount(6)
        self.recent_activity_table.setHorizontalHeaderLabels([
            "Date/Time", "Template", "PDF File", "Items", "Confidence", "Status"
        ])
        self.recent_activity_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.recent_activity_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.recent_activity_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        for i in range(3, 6):
            self.recent_activity_table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.recent_activity_table.setAlternatingRowColors(True)
        self.recent_activity_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.recent_activity_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.recent_activity_table.verticalHeader().setVisible(False)
        self.recent_activity_table.setMaximumHeight(200)
        layout.addWidget(self.recent_activity_table)

        # Learned Corrections section
        self.stats_corrections_label = QLabel("Learned Corrections")
        layout.addWidget(self.stats_corrections_label)

        corrections_info_layout = QHBoxLayout()
        self.corrections_count_label = QLabel("0 corrections recorded")
        corrections_info_layout.addWidget(self.corrections_count_label)
        corrections_info_layout.addStretch()

        self.stats_view_corrections_btn = QPushButton("View Details")
        self.stats_view_corrections_btn.clicked.connect(self._show_corrections_dialog)
        corrections_info_layout.addWidget(self.stats_view_corrections_btn)
        layout.addLayout(corrections_info_layout)

        # Apply theme-aware styles
        self._update_stats_tab_styles()

        # Initial data load
        QTimer.singleShot(500, self._refresh_processing_stats)

    def _create_stat_card(self, title: str, value: str, subtitle: str) -> QFrame:
        """Create a statistics card widget (styles applied via _update_stats_tab_styles)."""
        card = QFrame()
        card.setObjectName("stat_card")
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(2)
        card_layout.setContentsMargins(12, 8, 12, 8)

        title_lbl = QLabel(title)
        title_lbl.setObjectName("title")
        card_layout.addWidget(title_lbl)

        value_lbl = QLabel(value)
        value_lbl.setObjectName("value")
        card_layout.addWidget(value_lbl)

        subtitle_lbl = QLabel(subtitle)
        subtitle_lbl.setObjectName("subtitle")
        card_layout.addWidget(subtitle_lbl)

        return card

    def _update_stat_card(self, card: QFrame, value: str, subtitle: str):
        """Update a statistics card's values."""
        value_lbl = card.findChild(QLabel, "value")
        subtitle_lbl = card.findChild(QLabel, "subtitle")
        if value_lbl:
            value_lbl.setText(value)
        if subtitle_lbl:
            subtitle_lbl.setText(subtitle)

    def _update_stats_tab_styles(self):
        """Update Statistics tab styles based on current theme."""
        colors = self._get_ai_theme_colors()

        # Summary frame background
        self.stats_summary_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {colors['bg_secondary']};
                border-radius: 6px;
                padding: 10px;
            }}
        """)

        # Stat cards
        for card in [self.stats_today_label, self.stats_week_label,
                     self.stats_total_label, self.stats_success_label]:
            card.setStyleSheet(f"""
                QFrame {{
                    background-color: {colors['bg_input']};
                    border-radius: 6px;
                    padding: 8px;
                    border: 1px solid {colors['border']};
                }}
            """)
            # Update child labels
            title_lbl = card.findChild(QLabel, "title")
            value_lbl = card.findChild(QLabel, "value")
            subtitle_lbl = card.findChild(QLabel, "subtitle")
            if title_lbl:
                title_lbl.setStyleSheet(f"color: {colors['text_muted']}; font-size: 10px;")
            if value_lbl:
                value_lbl.setStyleSheet(f"color: {colors['accent']}; font-size: 18px; font-weight: bold;")
            if subtitle_lbl:
                subtitle_lbl.setStyleSheet(f"color: {colors['text_muted']}; font-size: 10px;")

        # Header and section labels
        self.stats_header_label.setStyleSheet(f"font-weight: bold; font-size: 14px; color: {colors['text_header']};")
        self.stats_table_label.setStyleSheet(f"font-weight: bold; font-size: 12px; margin-top: 10px; color: {colors['text_header']};")
        self.stats_recent_label.setStyleSheet(f"font-weight: bold; font-size: 12px; margin-top: 10px; color: {colors['text_header']};")
        self.stats_corrections_label.setStyleSheet(f"font-weight: bold; font-size: 12px; margin-top: 10px; color: {colors['text_header']};")
        self.corrections_count_label.setStyleSheet(f"color: {colors['text_muted']};")

        # Table styles
        table_style = f"""
            QTableWidget {{
                background-color: {colors['bg_code']};
                color: {colors['text']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                gridline-color: {colors['border']};
            }}
            QTableWidget::item {{
                padding: 5px;
            }}
            QTableWidget::item:selected {{
                background-color: {colors['selection']};
            }}
            QHeaderView::section {{
                background-color: {colors['bg_secondary']};
                color: {colors['text']};
                padding: 6px;
                border: none;
                border-bottom: 1px solid {colors['border']};
            }}
        """
        self.template_stats_table.setStyleSheet(table_style)
        self.recent_activity_table.setStyleSheet(table_style)

    def _refresh_processing_stats(self):
        """Refresh all processing statistics."""
        try:
            # Get summary stats
            summary = self.ocrmill_db.get_processing_stats_summary()

            # Update summary cards
            self._update_stat_card(
                self.stats_today_label,
                str(summary.get('processed_today', 0) or 0),
                f"{summary.get('items_today', 0) or 0} items"
            )
            self._update_stat_card(
                self.stats_week_label,
                str(summary.get('processed_week', 0) or 0),
                f"{summary.get('items_week', 0) or 0} items"
            )
            self._update_stat_card(
                self.stats_total_label,
                str(summary.get('total_processed', 0) or 0),
                f"{summary.get('total_items', 0) or 0} items"
            )

            # Calculate success rate
            total = summary.get('total_processed', 0) or 0
            successful = summary.get('successful', 0) or 0
            success_rate = (successful / total * 100) if total > 0 else 0
            self._update_stat_card(
                self.stats_success_label,
                f"{success_rate:.1f}%",
                f"{summary.get('templates_used', 0) or 0} templates"
            )

            # Get template statistics
            template_stats = self.ocrmill_db.get_template_statistics()
            self.template_stats_table.setRowCount(len(template_stats))

            for row, stat in enumerate(template_stats):
                self.template_stats_table.setItem(row, 0, QTableWidgetItem(stat.get('template_name', '')))
                self.template_stats_table.setItem(row, 1, QTableWidgetItem(str(stat.get('total_uses', 0))))
                self.template_stats_table.setItem(row, 2, QTableWidgetItem(str(stat.get('successful_uses', 0))))
                self.template_stats_table.setItem(row, 3, QTableWidgetItem(str(stat.get('total_items', 0) or 0)))
                self.template_stats_table.setItem(row, 4, QTableWidgetItem(f"{stat.get('avg_confidence', 0) or 0:.2f}"))
                self.template_stats_table.setItem(row, 5, QTableWidgetItem(str(int(stat.get('avg_time_ms', 0) or 0))))

                # Format last used date
                last_used = stat.get('last_used', '')
                if last_used:
                    try:
                        dt = datetime.fromisoformat(last_used)
                        last_used = dt.strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        pass
                self.template_stats_table.setItem(row, 6, QTableWidgetItem(last_used))

            # Get recent activity (get the most recent 20 entries from all templates)
            conn = self.ocrmill_db._get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM template_stats
                ORDER BY processed_date DESC
                LIMIT 20
            """)
            recent = [dict(row) for row in cursor.fetchall()]
            conn.close()

            self.recent_activity_table.setRowCount(len(recent))
            for row, activity in enumerate(recent):
                # Format date
                date_str = activity.get('processed_date', '')
                if date_str:
                    try:
                        dt = datetime.fromisoformat(date_str)
                        date_str = dt.strftime("%m/%d %H:%M:%S")
                    except Exception:
                        pass

                self.recent_activity_table.setItem(row, 0, QTableWidgetItem(date_str))
                self.recent_activity_table.setItem(row, 1, QTableWidgetItem(activity.get('template_name', '')))
                self.recent_activity_table.setItem(row, 2, QTableWidgetItem(activity.get('pdf_file', '') or ''))
                self.recent_activity_table.setItem(row, 3, QTableWidgetItem(str(activity.get('items_extracted', 0))))
                self.recent_activity_table.setItem(row, 4, QTableWidgetItem(f"{activity.get('confidence_score', 0) or 0:.2f}"))

                # Status with color
                status_item = QTableWidgetItem("Success" if activity.get('success') else "Failed")
                if activity.get('success'):
                    status_item.setForeground(QColor("#4caf50"))
                else:
                    status_item.setForeground(QColor("#f44336"))
                    if activity.get('error_message'):
                        status_item.setToolTip(activity.get('error_message'))
                self.recent_activity_table.setItem(row, 5, status_item)

            # Update corrections count
            try:
                correction_stats = self.ocrmill_db.get_correction_stats()
                count = correction_stats.get('total_corrections', 0)
                if count > 0:
                    self.corrections_count_label.setText(
                        f"{count} corrections recorded ({correction_stats.get('templates_with_corrections', 0)} templates)"
                    )
                else:
                    self.corrections_count_label.setText("No corrections recorded yet")
            except Exception:
                pass

        except Exception as e:
            logger.error(f"Error refreshing processing stats: {e}")

    def _show_corrections_dialog(self):
        """Show dialog with learned corrections details."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Learned Corrections")
        dialog.setMinimumSize(700, 500)

        layout = QVBoxLayout(dialog)

        # Info text
        info_label = QLabel(
            "The system learns from corrections you make to extracted data. "
            "When the same error pattern is seen multiple times, the system can "
            "suggest corrections automatically."
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #888888; margin-bottom: 10px;")
        layout.addWidget(info_label)

        # Stats summary
        try:
            stats = self.ocrmill_db.get_correction_stats()
            stats_text = (
                f"Total Corrections: {stats.get('total_corrections', 0)}  |  "
                f"Templates: {stats.get('templates_with_corrections', 0)}  |  "
                f"Fields: {stats.get('fields_corrected', 0)}"
            )
            stats_label = QLabel(stats_text)
            stats_label.setStyleSheet("font-weight: bold; margin-bottom: 5px;")
            layout.addWidget(stats_label)
        except Exception:
            pass

        # Corrections table
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels([
            "Template", "Field", "Original", "Corrected", "Frequency"
        ])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)

        try:
            corrections = self.ocrmill_db.get_common_corrections(limit=100)
            table.setRowCount(len(corrections))
            for row, corr in enumerate(corrections):
                table.setItem(row, 0, QTableWidgetItem(corr.get('template_name', '')))
                table.setItem(row, 1, QTableWidgetItem(corr.get('field_name', '')))
                table.setItem(row, 2, QTableWidgetItem(corr.get('original_value', '') or ''))
                table.setItem(row, 3, QTableWidgetItem(corr.get('corrected_value', '') or ''))
                table.setItem(row, 4, QTableWidgetItem(str(corr.get('frequency', 0))))
        except Exception as e:
            logger.error(f"Error loading corrections: {e}")

        layout.addWidget(table, 1)

        # Close button
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)

        dialog.exec_()

    def _on_ocrmill_item_changed(self, item: QTableWidgetItem):
        """Handle item changes in the OCRMill preview table for learning from corrections."""
        if self._ocrmill_editing:
            return

        row = item.row()
        col = item.column()
        new_value = item.text()

        # Get original value
        original_value = self._ocrmill_original_values.get((row, col), '')

        # Only record if value actually changed
        if original_value == new_value:
            return

        # Get field name from header
        field_name = ''
        if hasattr(self, '_ocrmill_preview_headers') and col < len(self._ocrmill_preview_headers):
            field_name = self._ocrmill_preview_headers[col]

        # Get part number for this row (usually first column)
        part_number = ''
        if self.ocrmill_preview_table.item(row, 0):
            part_number = self.ocrmill_preview_table.item(row, 0).text()

        # Get current file name (to extract template name)
        pdf_file = getattr(self, '_ocrmill_preview_file', '')

        # Determine template name from file pattern or recent processing
        template_name = 'unknown'
        if hasattr(self, 'ocrmill_processor') and hasattr(self.ocrmill_processor, 'templates'):
            # Try to get the most recently used template
            try:
                stats = self.ocrmill_db.get_template_statistics()
                if stats:
                    template_name = stats[0].get('template_name', 'unknown')
            except Exception:
                pass

        # Record the correction
        try:
            self.ocrmill_db.record_correction(
                template_name=template_name,
                pdf_file=pdf_file,
                field_name=field_name,
                original_value=original_value,
                corrected_value=new_value,
                part_number=part_number
            )
            self.ocrmill_log(f"Correction recorded: {field_name} '{original_value}' -> '{new_value}'")

            # Update original value tracking
            self._ocrmill_original_values[(row, col)] = new_value
        except Exception as e:
            logger.error(f"Failed to record correction: {e}")

    def _update_ai_template_styles(self):
        """Update AI template tab styles based on current theme."""
        colors = self._get_ai_theme_colors()

        # Template list
        self.ocrmill_templates_list.setStyleSheet(f"""
            QListWidget {{
                background-color: {colors['bg_primary']};
                color: {colors['text']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                padding: 4px;
            }}
            QListWidget::item {{
                padding: 8px 10px;
                border-bottom: 1px solid {colors['border']};
            }}
            QListWidget::item:selected {{
                background-color: {colors['selection']};
                color: {colors['text_header']};
            }}
            QListWidget::item:hover {{
                background-color: {colors['bg_secondary']};
            }}
        """)

        # Templates list header
        if hasattr(self, 'ai_templates_header'):
            self.ai_templates_header.setStyleSheet(f"font-weight: bold; font-size: 12px; color: {colors['text_header']};")

        # Header widget
        if hasattr(self, 'ai_template_header'):
            self.ai_template_header.setStyleSheet(f"background-color: {colors['bg_secondary']}; border-radius: 4px;")

        # Template name label (bold header)
        if hasattr(self, 'ai_template_name_label'):
            self.ai_template_name_label.setStyleSheet(f"font-weight: bold; color: {colors['text_header']};")

        # AI provider label
        if hasattr(self, 'ai_provider_label'):
            self.ai_provider_label.setStyleSheet(f"color: {colors['text_header']};")

        # Template Code label
        if hasattr(self, 'ai_code_label'):
            self.ai_code_label.setStyleSheet(f"color: {colors['text_header']};")

        # AI Assistant label
        if hasattr(self, 'ai_chat_label'):
            self.ai_chat_label.setStyleSheet(f"color: {colors['text_header']};")

        # Chat panel background
        if hasattr(self, 'ai_chat_widget'):
            self.ai_chat_widget.setStyleSheet(f"background-color: {colors['bg_secondary']}; border-radius: 4px;")

        # Code editor
        self.ai_code_edit.setStyleSheet(f"""
            QPlainTextEdit {{
                background-color: {colors['bg_code']};
                color: {colors['text']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                padding: 6px;
                selection-background-color: {colors['selection']};
            }}
        """)

        # Chat display
        self.ai_chat_display.setStyleSheet(f"""
            QTextEdit {{
                background-color: {colors['bg_code']};
                color: {colors['text']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                padding: 6px;
            }}
            QScrollBar:vertical {{
                background-color: {colors['bg_primary']};
                width: 8px;
            }}
            QScrollBar::handle:vertical {{
                background-color: {colors['border']};
                border-radius: 4px;
            }}
        """)

        # Message input
        self.ai_message_input.setStyleSheet(f"""
            QPlainTextEdit {{
                background-color: {colors['bg_input']};
                color: {colors['text']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                padding: 6px;
                font-size: 10pt;
            }}
            QPlainTextEdit:focus {{
                border: 1px solid {colors['border_focus']};
            }}
        """)

        # Context label (shows which template AI is using)
        self._update_ai_context_label()

        # Update syntax indicator color
        self.ai_syntax_indicator.setStyleSheet(f"color: {colors['success']}; font-size: 10px;")

        # Update buttons
        self.ai_send_btn.setStyleSheet(self._get_small_button_style("primary"))
        self.ai_cancel_btn.setStyleSheet(self._get_small_button_style("default"))
        self.ai_apply_btn.setStyleSheet(self._get_small_button_style("success"))
        self.ai_save_btn.setStyleSheet(self._get_small_button_style("success"))

        # Create mode widgets styling
        if hasattr(self, 'ai_new_invoice_text'):
            self.ai_new_invoice_text.setStyleSheet(f"""
                QPlainTextEdit {{
                    background-color: {colors['bg_code']};
                    color: {colors['text']};
                    border: 1px solid {colors['border']};
                    border-radius: 4px;
                    padding: 6px;
                }}
            """)

        if hasattr(self, 'ai_new_code_preview'):
            self.ai_new_code_preview.setStyleSheet(f"""
                QPlainTextEdit {{
                    background-color: {colors['bg_code']};
                    color: {colors['text']};
                    border: 1px solid {colors['border']};
                    border-radius: 4px;
                    padding: 6px;
                }}
            """)

        if hasattr(self, 'ai_generate_btn'):
            self.ai_generate_btn.setStyleSheet(self._get_small_button_style("primary"))
        if hasattr(self, 'ai_cancel_gen_btn'):
            self.ai_cancel_gen_btn.setStyleSheet(self._get_small_button_style("danger"))
        if hasattr(self, 'ai_save_new_btn'):
            self.ai_save_new_btn.setStyleSheet(self._get_small_button_style("success"))

        # Update welcome message if showing
        if not self.ai_conversation_history:
            self._show_ai_welcome_message()

    def _on_ai_provider_changed(self, provider: str):
        """Update model list when AI provider changes."""
        self.ai_model_combo.clear()
        if provider == "OpenAI":
            self.ai_model_combo.addItems(["gpt-4o", "gpt-4-turbo", "gpt-4", "gpt-3.5-turbo"])
        elif provider == "Anthropic":
            self.ai_model_combo.addItems(["claude-sonnet-4-20250514", "claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022"])
        elif provider == "Google Gemini":
            self.ai_model_combo.addItems(["gemini-1.5-pro", "gemini-1.5-flash", "gemini-1.0-pro"])
        elif provider == "Groq":
            self.ai_model_combo.addItems(["llama-3.3-70b-versatile", "llama-3.1-8b-instant", "mixtral-8x7b-32768", "gemma2-9b-it"])

    def _on_template_selected(self, item):
        """Handle template selection from list."""
        if not item:
            return

        # Switch to edit mode if in create mode
        if hasattr(self, 'ai_mode_stack') and self.ai_mode_stack.currentIndex() == 1:
            self.ai_mode_stack.setCurrentIndex(0)

        # Find template data - use row index for reliable matching
        row = self.ocrmill_templates_list.row(item)
        if 0 <= row < len(self.ocrmill_templates_data):
            self._load_template_for_editing(self.ocrmill_templates_data[row])

    def _on_template_double_clicked(self, item):
        """Handle double-click on template (same as single click for now)."""
        self._on_template_selected(item)

    def _load_template_for_editing(self, template_info):
        """Load a template into the editor."""
        template_path = Path(template_info['file_path'])
        if not template_path.exists():
            QMessageBox.warning(self, "Error", f"Template file not found: {template_path}")
            return

        try:
            with open(template_path, 'r', encoding='utf-8') as f:
                code = f.read()

            self.ai_code_edit.setPlainText(code)
            self.ai_current_template_path = template_path
            self.ai_template_name_label.setText(template_info['name'])
            self._update_ai_context_label()

            # Load AI metadata if exists
            ai_metadata_path = template_path.with_suffix('.ai_meta.json')
            if ai_metadata_path.exists():
                import json
                with open(ai_metadata_path, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                self.ai_conversation_history = metadata.get('conversation_history', [])
                # Set provider/model if saved
                provider = metadata.get('provider', '')
                if provider:
                    idx = self.ai_provider_combo.findText(provider)
                    if idx >= 0:
                        self.ai_provider_combo.setCurrentIndex(idx)
            else:
                self.ai_conversation_history = []

            # Update chat display
            self._ai_display_chat_history()

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load template: {e}")

    def _ai_template_new(self):
        """Create a new template using AI - switch to create mode."""
        # Clear the create mode fields
        self.ai_new_pdf_path.clear()
        self.ai_new_invoice_text.clear()
        self.ai_new_template_name.clear()
        self.ai_new_supplier_name.clear()
        self.ai_new_client.clear()
        self.ai_new_country.clear()
        self.ai_new_code_preview.clear()
        self.ai_new_progress.setVisible(False)
        self.ai_generate_btn.setVisible(True)
        self.ai_cancel_gen_btn.setVisible(False)
        self.ai_save_new_btn.setEnabled(False)

        # Clear template context for new template mode
        self.ai_current_template_path = None
        self._update_ai_context_label()

        # Update header
        self.ai_template_name_label.setText("New Template")

        # Switch to create mode
        self.ai_mode_stack.setCurrentIndex(1)

    def _ai_switch_to_edit_mode(self):
        """Switch back to edit mode from create mode."""
        self.ai_mode_stack.setCurrentIndex(0)
        # Reset header if no template is selected
        if not self.ai_current_template_path:
            self.ai_template_name_label.setText("Select a template")
        self._update_ai_context_label()

    def _ai_browse_sample_pdf(self):
        """Browse for a sample PDF invoice."""
        try:
            import pdfplumber
        except ImportError:
            QMessageBox.warning(
                self, "Missing Dependency",
                "pdfplumber is required to load PDFs.\n\n"
                "Install with: pip install pdfplumber"
            )
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Select Invoice PDF",
            str(Path.home()),
            "PDF Files (*.pdf);;All Files (*.*)"
        )

        if not path:
            return

        try:
            self.ai_new_pdf_path.setText(path)

            # Extract text from PDF
            text_parts = []
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages[:5]:  # First 5 pages
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)

            full_text = '\n\n'.join(text_parts)
            self.ai_new_invoice_text.setPlainText(full_text)

            # Auto-detect supplier name
            self._ai_auto_detect_supplier(full_text)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to extract PDF text:\n{e}")

    def _ai_auto_detect_supplier(self, text: str):
        """Try to auto-detect supplier name from invoice text."""
        import re
        lines = text.split('\n')

        for line in lines[:10]:
            line = line.strip()
            if len(line) > 5 and len(line) < 60:
                if re.match(r'^[\d\s\-/]+$', line):
                    continue
                if re.search(r'\d{4}', line):
                    continue
                if '@' in line or 'www.' in line.lower():
                    continue

                if not self.ai_new_supplier_name.text():
                    self.ai_new_supplier_name.setText(line)
                    template_name = re.sub(r'[^a-z0-9]+', '_', line.lower()).strip('_')
                    self.ai_new_template_name.setText(template_name[:30])
                break

    def _ai_generate_new_template(self):
        """Generate a new template from sample invoice using AI."""
        import re

        # Validate inputs
        invoice_text = self.ai_new_invoice_text.toPlainText().strip()
        if not invoice_text:
            QMessageBox.warning(self, "Missing Input", "Please provide invoice text or load a PDF.")
            return

        template_name = self.ai_new_template_name.text().strip()
        if not template_name:
            QMessageBox.warning(self, "Missing Input", "Please enter a template name.")
            return

        if not re.match(r'^[a-z][a-z0-9_]*$', template_name):
            QMessageBox.warning(
                self, "Invalid Name",
                "Template name must be lowercase, start with a letter, "
                "and contain only letters, numbers, and underscores."
            )
            return

        supplier_name = self.ai_new_supplier_name.text().strip()
        if not supplier_name:
            QMessageBox.warning(self, "Missing Input", "Please enter a supplier name.")
            return

        provider = self.ai_provider_combo.currentText()

        # Check if required package is installed
        if provider == "OpenAI":
            if not self._check_and_install_ai_package("openai"):
                return
        elif provider == "Anthropic":
            if not self._check_and_install_ai_package("anthropic"):
                return

        api_key = self._ai_get_api_key(provider)

        if provider in ["OpenAI", "Anthropic"] and not api_key:
            QMessageBox.warning(
                self, "Missing API Key",
                f"No API key found for {provider}.\n"
                "Configure it in Configuration > AI Agents tab."
            )
            return

        # Show progress
        self.ai_new_progress.setVisible(True)
        self.ai_new_progress.setRange(0, 0)
        self.ai_new_progress.setFormat("Generating template...")
        self.ai_generate_btn.setVisible(False)
        self.ai_cancel_gen_btn.setVisible(True)
        self.ai_save_new_btn.setEnabled(False)

        # Start generation thread
        from ai_template_generator import AIGeneratorThread
        self.ai_generator_thread = AIGeneratorThread(
            provider=provider,
            model=self.ai_model_combo.currentText(),
            api_key=api_key,
            invoice_text=invoice_text,
            template_name=template_name,
            supplier_name=supplier_name,
            country=self.ai_new_country.text().strip() or "UNKNOWN",
            client=self.ai_new_client.text().strip() or "Universal"
        )
        self.ai_generator_thread.finished.connect(self._ai_on_generation_complete)
        self.ai_generator_thread.error.connect(self._ai_on_generation_error)
        self.ai_generator_thread.progress.connect(self._ai_on_generation_progress)
        self.ai_generator_thread.stream_update.connect(self._ai_on_stream_update)
        self.ai_generator_thread.cancelled.connect(self._ai_on_generation_cancelled)
        self.ai_generator_thread.start()

    def _ai_on_generation_progress(self, message: str):
        """Update progress message during generation."""
        self.ai_new_progress.setFormat(message)

    def _ai_on_stream_update(self, text: str):
        """Update preview with streaming text."""
        self.ai_new_code_preview.setPlainText(text)
        scrollbar = self.ai_new_code_preview.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def _ai_on_generation_complete(self, code: str):
        """Handle successful template generation."""
        self.ai_new_progress.setVisible(False)
        self.ai_generate_btn.setVisible(True)
        self.ai_cancel_gen_btn.setVisible(False)
        self.ai_save_new_btn.setEnabled(True)

        self.ai_new_code_preview.setPlainText(code)

        QMessageBox.information(
            self, "Generation Complete",
            "Template generated successfully!\n\n"
            "Review the code in the preview, then click 'Save Template' to save it."
        )

    def _ai_on_generation_error(self, error: str):
        """Handle generation error."""
        self.ai_new_progress.setVisible(False)
        self.ai_generate_btn.setVisible(True)
        self.ai_cancel_gen_btn.setVisible(False)

        QMessageBox.critical(
            self, "Generation Error",
            f"Failed to generate template:\n\n{error}"
        )

    def _ai_cancel_generation(self):
        """Cancel the ongoing template generation."""
        if hasattr(self, 'ai_generator_thread') and self.ai_generator_thread.isRunning():
            self.ai_new_progress.setFormat("Cancelling...")
            self.ai_cancel_gen_btn.setEnabled(False)
            self.ai_generator_thread.cancel()
            if not self.ai_generator_thread.wait(2000):
                self.ai_generator_thread.terminate()
                self.ai_generator_thread.wait()
            self._ai_on_generation_cancelled()

    def _ai_on_generation_cancelled(self):
        """Handle cancelled generation."""
        self.ai_new_progress.setVisible(False)
        self.ai_generate_btn.setVisible(True)
        self.ai_cancel_gen_btn.setVisible(False)
        self.ai_cancel_gen_btn.setEnabled(True)

    def _ai_save_new_template(self):
        """Save the newly generated template."""
        code = self.ai_new_code_preview.toPlainText().strip()
        if not code:
            QMessageBox.warning(self, "No Code", "No template code to save.")
            return

        template_name = self.ai_new_template_name.text().strip()

        # Determine templates directory
        templates_dir = Path(__file__).parent / 'templates'
        if not templates_dir.exists():
            templates_dir.mkdir(parents=True, exist_ok=True)

        file_path = templates_dir / f"{template_name}.py"

        # Check if file already exists
        if file_path.exists():
            reply = QMessageBox.question(
                self, "File Exists",
                f"Template '{template_name}' already exists.\n\nOverwrite?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        # Save the file
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(code)

            # Save AI metadata alongside the template
            import json
            from datetime import datetime
            metadata_path = file_path.with_suffix('.ai_meta.json')
            metadata = {
                'provider': self.ai_provider_combo.currentText(),
                'model': self.ai_model_combo.currentText(),
                'template_name': template_name,
                'supplier_name': self.ai_new_supplier_name.text().strip(),
                'country': self.ai_new_country.text().strip(),
                'client': self.ai_new_client.text().strip(),
                'invoice_text': self.ai_new_invoice_text.toPlainText().strip()[:5000],
                'created_at': datetime.now().isoformat(),
                'conversation_history': []
            }
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2)

            QMessageBox.information(
                self, "Template Saved",
                f"Template saved successfully!\n\n"
                f"File: {file_path}"
            )

            # Refresh templates list and switch to edit mode
            self.ocrmill_refresh_templates()
            self._ai_switch_to_edit_mode()

            # Select the new template in the list
            for i in range(self.ocrmill_templates_list.count()):
                item = self.ocrmill_templates_list.item(i)
                if item and item.text() == template_name.replace('_', ' ').title():
                    self.ocrmill_templates_list.setCurrentItem(item)
                    self._on_template_selected(item)
                    break

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save template:\n{e}")

    def _ai_validate_syntax(self):
        """Validate Python syntax of the template code."""
        colors = self._get_ai_theme_colors()
        code = self.ai_code_edit.toPlainText()
        if not code.strip():
            self.ai_syntax_indicator.setText("")
            return
        try:
            compile(code, '<template>', 'exec')
            self.ai_syntax_indicator.setText("✓ Valid")
            self.ai_syntax_indicator.setStyleSheet(f"color: {colors['success']}; font-size: 10px;")
        except SyntaxError as e:
            self.ai_syntax_indicator.setText(f"✗ Line {e.lineno}")
            self.ai_syntax_indicator.setStyleSheet(f"color: {colors['danger']}; font-size: 10px;")

    def _ai_test_template(self):
        """Test the current template."""
        code = self.ai_code_edit.toPlainText()
        if not code.strip():
            QMessageBox.warning(self, "No Code", "Please enter or load a template first.")
            return

        try:
            # First check syntax
            compile(code, '<template>', 'exec')

            # Convert relative imports to absolute imports for testing
            # Replace "from .base_template" with "from templates.base_template"
            test_code = code.replace('from .base_template', 'from templates.base_template')
            test_code = test_code.replace('from .', 'from templates.')

            # Add templates directory to path if needed
            import sys
            templates_dir = str(Path(__file__).parent / 'templates')
            if templates_dir not in sys.path:
                sys.path.insert(0, templates_dir)

            # Execute with proper context
            exec(test_code, {'__name__': '__main__'})
            QMessageBox.information(self, "Test Passed", "Template syntax is valid and imports work correctly!")
        except SyntaxError as e:
            error_msg = f"Line {e.lineno}: {e.msg}"
            self._handle_template_error("Syntax Error", error_msg)
        except Exception as e:
            import traceback
            error_msg = str(e)
            full_traceback = traceback.format_exc()
            self._handle_template_error("Template Error", error_msg, full_traceback)

    def _handle_template_error(self, error_type: str, error_msg: str, traceback_info: str = None):
        """Handle template test errors and offer to send to AI for correction."""
        # Create custom dialog with option to send to AI
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle("Test Failed")

        display_msg = f"{error_type}:\n\n{error_msg}"
        if traceback_info:
            # Show abbreviated traceback
            tb_lines = traceback_info.strip().split('\n')
            if len(tb_lines) > 6:
                display_msg += f"\n\n...{chr(10).join(tb_lines[-4:])}"

        msg_box.setText(display_msg)

        # Add buttons
        fix_btn = msg_box.addButton("Send to AI to Fix", QMessageBox.AcceptRole)
        msg_box.addButton("OK", QMessageBox.RejectRole)

        msg_box.exec_()

        if msg_box.clickedButton() == fix_btn:
            self._send_error_to_ai(error_type, error_msg, traceback_info)

    def _send_error_to_ai(self, error_type: str, error_msg: str, traceback_info: str = None):
        """Send the template error to AI Assistant for correction."""
        # Build the error correction prompt
        current_code = self.ai_code_edit.toPlainText()

        error_prompt = f"""The template code has an error that needs to be fixed.

ERROR TYPE: {error_type}
ERROR MESSAGE: {error_msg}
"""
        if traceback_info:
            error_prompt += f"""
FULL TRACEBACK:
{traceback_info}
"""

        error_prompt += """
Please fix this error in the template code. Return the complete corrected template code."""

        # Set the prompt in the chat input
        self.ai_message_input.setPlainText(error_prompt)

        # Automatically send to AI
        self._ai_send_message()

    def _ai_format_code(self):
        """Format the template code."""
        if not self.ai_code_edit.toPlainText().strip():
            return
        success, message = self.ai_code_edit.format_code()
        if not success:
            QMessageBox.warning(self, "Format Error", message)

    def _ai_get_api_key(self, provider: str) -> str:
        """Get API key for the provider from the AI Agents configuration."""
        # Use the same method as AI Agents tab for consistency
        key_name = {'OpenAI': 'openai', 'Anthropic': 'anthropic', 'Google Gemini': 'gemini', 'Groq': 'groq'}.get(provider, 'openai')
        api_key = self._get_ai_api_key(key_name)
        if api_key:
            return api_key
        # Fallback to environment variables
        if provider == "OpenAI":
            return os.environ.get('OPENAI_API_KEY', '')
        elif provider == "Anthropic":
            return os.environ.get('ANTHROPIC_API_KEY', '')
        elif provider == "Google Gemini":
            return os.environ.get('GOOGLE_API_KEY', '')
        elif provider == "Groq":
            return os.environ.get('GROQ_API_KEY', '')
        return ""

    def _check_and_install_ai_package(self, package_name: str) -> bool:
        """
        Check if a package is installed, and offer to install if not.
        Returns True if package is available, False otherwise.
        """
        import importlib
        import importlib.util

        # Method 1: Try direct import (most reliable - works for bundled PyInstaller apps too)
        try:
            importlib.import_module(package_name)
            logger.debug(f"Package {package_name} imported successfully")
            return True
        except (ImportError, ModuleNotFoundError) as e:
            logger.debug(f"Direct import of {package_name} failed: {e}")

        # Method 2: Clear import caches and try find_spec
        try:
            importlib.invalidate_caches()
            spec = importlib.util.find_spec(package_name)
            if spec is not None:
                logger.debug(f"Package {package_name} found via find_spec")
                return True
        except (ImportError, ModuleNotFoundError, ValueError) as e:
            logger.debug(f"find_spec for {package_name} failed: {e}")

        # If running as a frozen PyInstaller app, we can't install packages at runtime
        # The package should have been bundled - if import failed, it's not available
        if getattr(sys, 'frozen', False):
            logger.warning(f"Package {package_name} not bundled in frozen app")
            QMessageBox.warning(
                self, "Package Not Available",
                f"The {package_name} package is not available in this build.\n\n"
                "Please contact support or use a development version."
            )
            return False

        # Method 3: For non-frozen apps, use pip to check if package is installed
        import subprocess
        try:
            startupinfo = None
            creationflags = 0
            if sys.platform == 'win32':
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = subprocess.SW_HIDE
                creationflags = subprocess.CREATE_NO_WINDOW

            result = subprocess.run(
                [sys.executable, "-m", "pip", "show", package_name],
                capture_output=True,
                text=True,
                startupinfo=startupinfo,
                creationflags=creationflags,
                timeout=10
            )
            if result.returncode == 0 and package_name.lower() in result.stdout.lower():
                # Package is installed but Python hasn't loaded it yet
                # Try importing again after cache invalidation
                importlib.invalidate_caches()
                try:
                    importlib.import_module(package_name)
                    return True
                except:
                    # Package installed but can't be imported - may need app restart
                    logger.info(f"Package {package_name} is installed but requires app restart to use")
                    return True  # Return True since it IS installed
        except Exception as e:
            logger.debug(f"pip show check failed: {e}")

        # Track packages currently being installed
        if not hasattr(self, '_installing_packages'):
            self._installing_packages = set()

        if package_name in self._installing_packages:
            QMessageBox.information(
                self, "Installation In Progress",
                f"The {package_name} package is currently being installed.\n"
                "Please wait for installation to complete."
            )
            return False

        # Ask user if they want to install
        reply = QMessageBox.question(
            self, "Package Not Installed",
            f"The {package_name} package is required but not installed.\n\n"
            f"Would you like to install it now?\n"
            f"(Installation runs in background)",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )

        if reply == QMessageBox.Yes:
            self._install_package_background(package_name)

        return False

    def _install_package_background(self, package_name: str):
        """Install a package in a hidden background process with progress dialog."""
        import subprocess
        import threading

        self._installing_packages.add(package_name)

        # Create progress dialog
        progress = QProgressDialog(
            f"Installing {package_name}...\nThis may take a minute.",
            None,  # No cancel button
            0, 0,  # Indeterminate progress (0, 0)
            self
        )
        progress.setWindowTitle("Installing AI Package")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)  # Show immediately
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.setCancelButton(None)  # Remove cancel button
        progress.setMinimumWidth(350)

        # Store reference to close later
        self._install_progress_dialog = progress
        progress.show()
        QApplication.processEvents()

        def do_install():
            try:
                python_exe = sys.executable
                # Use CREATE_NO_WINDOW flag on Windows for hidden console
                startupinfo = None
                creationflags = 0
                if sys.platform == 'win32':
                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                    startupinfo.wShowWindow = subprocess.SW_HIDE
                    creationflags = subprocess.CREATE_NO_WINDOW

                result = subprocess.run(
                    [python_exe, "-m", "pip", "install", package_name],
                    capture_output=True,
                    text=True,
                    startupinfo=startupinfo,
                    creationflags=creationflags
                )

                # Signal completion back to main thread
                QTimer.singleShot(0, lambda: self._on_package_install_complete(
                    package_name, result.returncode == 0, result.stderr
                ))
            except Exception as e:
                QTimer.singleShot(0, lambda: self._on_package_install_complete(
                    package_name, False, str(e)
                ))

        # Show installing message in status bar as backup
        self.statusBar().showMessage(f"Installing {package_name}...", 0)

        # Run in background thread
        thread = threading.Thread(target=do_install, daemon=True)
        thread.start()

    def _on_package_install_complete(self, package_name: str, success: bool, error_msg: str):
        """Handle package installation completion."""
        self._installing_packages.discard(package_name)
        self.statusBar().clearMessage()

        # Close the progress dialog
        if hasattr(self, '_install_progress_dialog') and self._install_progress_dialog:
            self._install_progress_dialog.close()
            self._install_progress_dialog = None

        if success:
            QMessageBox.information(
                self, "Installation Complete",
                f"The {package_name} package has been installed successfully!\n\n"
                "You can now use this feature."
            )
        else:
            QMessageBox.warning(
                self, "Installation Failed",
                f"Failed to install {package_name}.\n\n"
                f"Error: {error_msg}\n\n"
                f"Please try manually:\n  pip install {package_name}"
            )

    def _ai_format_message_html(self, role: str, content: str) -> str:
        """Format a chat message as HTML with theme-aware colors."""
        import html as html_module
        import re

        # Get theme colors
        colors = self._get_ai_theme_colors()
        code_bg = colors['bg_code']
        code_text = colors['text']
        code_accent = colors['accent']
        inline_code_bg = colors['bg_secondary']
        inline_code_text = colors['warning']

        # Get font size from settings and scale up for chat readability
        base_font_size = get_user_setting_int('font_size', 9)
        font_size = base_font_size + 1  # Chat text slightly larger for readability
        label_size = max(base_font_size - 1, 7)  # Label slightly smaller, minimum 7pt
        code_size = base_font_size  # Code same size as base setting

        def replace_code_block(match):
            code = html_module.escape(match.group(1))
            return f'</p><pre style="background-color: {code_bg}; color: {code_text}; padding: 8px; border-radius: 4px; font-family: Consolas; font-size: {code_size}pt; margin: 5px 0; border-left: 2px solid {code_accent}; white-space: pre-wrap;">{code}</pre><p style="margin: 0;">'

        formatted = re.sub(r'```(?:python)?\s*(.*?)\s*```', replace_code_block, content, flags=re.DOTALL)

        parts = re.split(r'(<pre.*?</pre>)', formatted, flags=re.DOTALL)
        escaped_parts = []
        for part in parts:
            if part.startswith('<pre'):
                escaped_parts.append(part)
            else:
                escaped_parts.append(html_module.escape(part))
        formatted = ''.join(escaped_parts)

        formatted = re.sub(r'`([^`]+)`', rf'<code style="background-color: {inline_code_bg}; color: {inline_code_text}; padding: 1px 4px; border-radius: 2px;">\1</code>', formatted)

        parts = re.split(r'(<pre.*?</pre>)', formatted, flags=re.DOTALL)
        for i, part in enumerate(parts):
            if not part.startswith('<pre'):
                parts[i] = part.replace('\n', '<br>')
        formatted = ''.join(parts)

        # Theme-aware message bubble colors
        user_bubble_bg = colors['user_bubble']
        user_bubble_text = colors['user_bubble_text']
        ai_bubble_bg = colors['ai_bubble']
        ai_bubble_text = colors['ai_bubble_text']
        ai_bubble_border = colors['border']
        label_color = colors['text_muted']

        if role == 'user':
            return f'''<div style="margin: 8px 0; text-align: right;">
                <div style="display: inline-block; max-width: 80%; text-align: left;">
                    <div style="color: {label_color}; font-size: {label_size}pt;">You</div>
                    <div style="background-color: {user_bubble_bg}; color: {user_bubble_text}; padding: 8px 10px; border-radius: 8px 8px 2px 8px;">
                        <p style="margin: 0; font-size: {font_size}pt;">{formatted}</p>
                    </div>
                </div>
            </div>'''
        else:
            return f'''<div style="margin: 8px 0;">
                <div style="display: inline-block; max-width: 80%;">
                    <div style="color: {label_color}; font-size: {label_size}pt;">AI</div>
                    <div style="background-color: {ai_bubble_bg}; color: {ai_bubble_text}; padding: 8px 10px; border-radius: 8px 8px 8px 2px; border: 1px solid {ai_bubble_border};">
                        <p style="margin: 0; font-size: {font_size}pt;">{formatted}</p>
                    </div>
                </div>
            </div>'''

    def _ai_display_chat_history(self, show_thinking: bool = False):
        """Display chat history with optional thinking indicator."""
        if not self.ai_conversation_history:
            self._show_ai_welcome_message()
            return

        colors = self._get_ai_theme_colors()
        base_font_size = get_user_setting_int('font_size', 9)
        font_size = base_font_size + 1  # Chat text slightly larger for readability
        label_size = max(base_font_size - 1, 7)  # Label slightly smaller, minimum 7pt

        html = '<div style="font-family: Segoe UI;">'
        for msg in self.ai_conversation_history:
            html += self._ai_format_message_html(msg.get('role', 'user'), msg.get('content', ''))

        # Add thinking indicator if AI is processing
        if show_thinking:
            thinking_dots = getattr(self, '_ai_thinking_dots', 1)
            dots = '.' * thinking_dots
            html += f'''<div style="margin: 8px 0;">
                <div style="display: inline-block; max-width: 80%;">
                    <div style="color: {colors['text_muted']}; font-size: {label_size}pt;">AI</div>
                    <div style="background-color: {colors['ai_bubble']}; color: {colors['text_muted']}; padding: 8px 10px; border-radius: 8px 8px 8px 2px; border: 1px solid {colors['border']};">
                        <p style="margin: 0; font-size: {font_size}pt; font-style: italic;">Thinking{dots}</p>
                    </div>
                </div>
            </div>'''

        html += '</div>'
        self.ai_chat_display.setHtml(html)

    def _ai_start_thinking_animation(self):
        """Start the thinking dots animation."""
        self._ai_thinking_dots = 1
        if not hasattr(self, '_ai_thinking_timer'):
            self._ai_thinking_timer = QTimer()
            self._ai_thinking_timer.timeout.connect(self._ai_update_thinking_animation)
        self._ai_thinking_timer.start(400)  # Update every 400ms
        self._ai_display_chat_history(show_thinking=True)

    def _ai_update_thinking_animation(self):
        """Update the thinking dots animation."""
        self._ai_thinking_dots = (self._ai_thinking_dots % 3) + 1
        self._ai_display_chat_history(show_thinking=True)
        # Scroll to bottom
        sb = self.ai_chat_display.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _ai_stop_thinking_animation(self):
        """Stop the thinking dots animation."""
        if hasattr(self, '_ai_thinking_timer'):
            self._ai_thinking_timer.stop()
        self._ai_display_chat_history(show_thinking=False)

    def _ai_send_message(self):
        """Send a message to the AI."""
        message = self.ai_message_input.toPlainText().strip()
        if not message:
            return

        provider = self.ai_provider_combo.currentText()
        api_key = self._ai_get_api_key(provider)

        # Check if required package is installed
        if provider == "OpenAI":
            if not self._check_and_install_ai_package("openai"):
                return
        elif provider == "Anthropic":
            if not self._check_and_install_ai_package("anthropic"):
                return

        if provider in ["OpenAI", "Anthropic"] and not api_key:
            QMessageBox.warning(self, "Missing API Key",
                              f"No API key found for {provider}.\n"
                              "Configure it in Configuration > AI Agents tab.")
            return

        # Check for PDF file paths in the message and extract text automatically
        invoice_text = ""
        enhanced_message = message
        pdf_paths = self._ai_extract_pdf_paths(message)
        if pdf_paths:
            extracted_texts = []
            for pdf_path in pdf_paths:
                text = self._ai_extract_pdf_text(pdf_path)
                if text:
                    extracted_texts.append(f"--- Content from {Path(pdf_path).name} ---\n{text}")
            if extracted_texts:
                invoice_text = "\n\n".join(extracted_texts)
                enhanced_message = f"{message}\n\n[PDF text extracted automatically - {len(extracted_texts)} file(s)]"

        # Add to history and display (show enhanced message to user)
        self.ai_conversation_history.append({"role": "user", "content": enhanced_message})
        self._ai_display_chat_history()
        self.ai_message_input.clear()

        # Scroll to bottom
        sb = self.ai_chat_display.verticalScrollBar()
        sb.setValue(sb.maximum())

        # Disable send, enable cancel, start thinking animation
        self.ai_send_btn.setEnabled(False)
        self.ai_cancel_btn.setEnabled(True)
        self._ai_start_thinking_animation()

        # Build the actual message to send to AI (with full PDF content)
        ai_message = message
        if invoice_text:
            ai_message = f"{message}\n\nHere is the extracted text from the PDF file(s):\n\n{invoice_text}"

        # Start AI request
        from ai_template_generator import AITemplateChatThread
        self.ai_chat_thread = AITemplateChatThread(
            provider=provider,
            model=self.ai_model_combo.currentText(),
            api_key=api_key,
            current_code=self.ai_code_edit.toPlainText(),
            user_message=ai_message,
            conversation_history=self.ai_conversation_history[:-1],
            invoice_text=invoice_text
        )
        self.ai_chat_thread.finished.connect(self._ai_on_response)
        self.ai_chat_thread.error.connect(self._ai_on_error)
        self.ai_chat_thread.start()

    def _ai_extract_pdf_paths(self, message: str) -> list:
        """Extract PDF file paths from a message."""
        import re
        paths = []

        # Pattern to match Windows paths (with or without quotes)
        # Matches: C:\path\to\file.pdf, "C:\path\to\file.pdf", 'C:\path\to\file.pdf'
        patterns = [
            r'["\']([A-Za-z]:\\[^"\']+\.pdf)["\']',  # Quoted paths
            r'([A-Za-z]:\\(?:[^\s<>:"|?*]+\\)*[^\s<>:"|?*]+\.pdf)',  # Unquoted paths
        ]

        for pattern in patterns:
            matches = re.findall(pattern, message, re.IGNORECASE)
            for match in matches:
                path = match.strip('"\'')
                if path not in paths and Path(path).exists():
                    paths.append(path)

        return paths

    def _ai_extract_pdf_text(self, pdf_path: str) -> str:
        """Extract text from a PDF file."""
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(pdf_path) as pdf:
                for i, page in enumerate(pdf.pages[:10]):  # Limit to first 10 pages
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(f"[Page {i+1}]\n{page_text}")
            return "\n\n".join(text_parts)
        except ImportError:
            self.ocrmill_log("pdfplumber not installed - cannot extract PDF text")
            return ""
        except Exception as e:
            self.ocrmill_log(f"Error extracting PDF text: {e}")
            return ""

    def _ai_cancel_request(self):
        """Cancel the current AI request."""
        self._ai_stop_thinking_animation()
        if self.ai_chat_thread:
            self.ai_chat_thread.cancel()
            self.ai_chat_thread = None
        self.ai_send_btn.setEnabled(True)
        self.ai_cancel_btn.setEnabled(False)

    def _ai_on_response(self, response: str):
        """Handle AI response."""
        self._ai_stop_thinking_animation()
        self.ai_send_btn.setEnabled(True)
        self.ai_cancel_btn.setEnabled(False)

        self.ai_conversation_history.append({"role": "assistant", "content": response})
        self._ai_display_chat_history()

        sb = self.ai_chat_display.verticalScrollBar()
        sb.setValue(sb.maximum())

        # Check for code block and auto-apply like VS Code
        import re
        code_match = re.search(r'```python\s*(.*?)\s*```', response, re.DOTALL)
        if code_match:
            new_code = code_match.group(1).strip()
            # Auto-apply the code to the editor (VS Code style)
            self.ai_code_edit.setPlainText(new_code)
            self.ai_pending_code = None
            self.ai_apply_btn.setEnabled(False)

    def _ai_on_error(self, error: str):
        """Handle AI error."""
        self._ai_stop_thinking_animation()
        self.ai_send_btn.setEnabled(True)
        self.ai_cancel_btn.setEnabled(False)
        QMessageBox.warning(self, "AI Error", f"Error: {error}")

    def _ai_apply_changes(self):
        """Apply AI-suggested code changes."""
        if self.ai_pending_code:
            self.ai_code_edit.setPlainText(self.ai_pending_code)
            self.ai_apply_btn.setEnabled(False)
            self.ai_pending_code = None

    def _ai_save_template(self):
        """Save the current template."""
        if not self.ai_current_template_path:
            QMessageBox.warning(self, "No Template", "No template selected. Create a new one first.")
            return

        code = self.ai_code_edit.toPlainText()
        if not code.strip():
            QMessageBox.warning(self, "Empty Code", "Cannot save empty template.")
            return

        # Validate syntax
        try:
            compile(code, '<template>', 'exec')
        except SyntaxError as e:
            reply = QMessageBox.question(
                self, "Syntax Error",
                f"Template has syntax errors:\nLine {e.lineno}: {e.msg}\n\nSave anyway?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        try:
            import json
            # Save code
            with open(self.ai_current_template_path, 'w', encoding='utf-8') as f:
                f.write(code)

            # Save metadata
            metadata = {
                'provider': self.ai_provider_combo.currentText(),
                'model': self.ai_model_combo.currentText(),
                'conversation_history': self.ai_conversation_history,
                'last_modified': datetime.now().isoformat()
            }
            metadata_path = self.ai_current_template_path.with_suffix('.ai_meta.json')
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2)

            QMessageBox.information(self, "Saved", "Template saved successfully!")
            self.ocrmill_refresh_templates()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save: {e}")

    def ocrmill_browse_input_folder(self):
        """Browse for input folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder", str(self.ocrmill_config.input_folder))
        if folder:
            # Normalize path to OS-native format
            folder_path = Path(folder)
            folder_str = str(folder_path)
            self.ocrmill_config.input_folder = folder_path
            self.ocrmill_input_edit.setText(folder_str)
            self.ocrmill_drop_zone.set_browse_folder(folder_str)
            set_user_setting('ocrmill_input_folder', folder_str)
            self.ocrmill_log(f"Input folder set to: {folder_str}")
            self.ocrmill_refresh_input_files()

    def ocrmill_browse_output_folder(self):
        """Browse for output folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", str(self.ocrmill_config.output_folder))
        if folder:
            # Normalize path to OS-native format
            folder_path = Path(folder)
            folder_str = str(folder_path)
            self.ocrmill_config.output_folder = folder_path
            self.ocrmill_output_edit.setText(folder_str)
            set_user_setting('ocrmill_output_folder', folder_str)
            self.ocrmill_log(f"Output folder set to: {folder_str}")
            self.ocrmill_refresh_output_files()

    def ocrmill_refresh_input_files(self):
        """Refresh the input files list with PDFs from the input folder."""
        try:
            self.ocrmill_input_files_list.clear()
            input_folder = self.ocrmill_config.input_folder
            if input_folder.exists():
                # Get PDF files sorted by modification time (newest first)
                pdf_files = sorted(
                    input_folder.glob("*.pdf"),
                    key=lambda f: f.stat().st_mtime,
                    reverse=True
                )
                for f in pdf_files:
                    self.ocrmill_input_files_list.addItem(f.name)
        except Exception as e:
            logger.warning(f"Failed to refresh OCRMill input files: {e}")

    def ocrmill_refresh_output_files(self):
        """Refresh the output files list with CSVs from the output folder."""
        try:
            self.ocrmill_output_files_list.clear()
            output_folder = self.ocrmill_config.output_folder
            if output_folder.exists():
                # Get CSV files sorted by modification time (newest first)
                csv_files = sorted(
                    output_folder.glob("*.csv"),
                    key=lambda f: f.stat().st_mtime,
                    reverse=True
                )
                for f in csv_files:
                    self.ocrmill_output_files_list.addItem(f.name)
                # Enable Send button if there are CSV files to send
                self.ocrmill_send_btn.setEnabled(len(csv_files) > 0)
            else:
                self.ocrmill_send_btn.setEnabled(False)
        except Exception as e:
            logger.warning(f"Failed to refresh OCRMill output files: {e}")

    def ocrmill_open_input_file(self, item):
        """Open the selected input PDF file."""
        if not item:
            return
        file_path = self.ocrmill_config.input_folder / item.text()
        if file_path.exists():
            try:
                import subprocess
                import sys
                if sys.platform == 'win32':
                    os.startfile(str(file_path))
                elif sys.platform == 'darwin':
                    subprocess.run(['open', str(file_path)])
                else:
                    subprocess.run(['xdg-open', str(file_path)])
                self.ocrmill_log(f"Opened: {item.text()}")
            except Exception as e:
                self.ocrmill_log(f"Failed to open file: {e}")

    def ocrmill_open_output_file(self, item):
        """Open the selected output CSV file."""
        if not item:
            return
        file_path = self.ocrmill_config.output_folder / item.text()
        if file_path.exists():
            try:
                import subprocess
                import sys
                if sys.platform == 'win32':
                    os.startfile(str(file_path))
                elif sys.platform == 'darwin':
                    subprocess.run(['open', str(file_path)])
                else:
                    subprocess.run(['xdg-open', str(file_path)])
                self.ocrmill_log(f"Opened: {item.text()}")
            except Exception as e:
                self.ocrmill_log(f"Failed to open file: {e}")

    def ocrmill_preview_output_file(self, item):
        """Load the selected output CSV file into the Results Preview table."""
        if not item:
            return
        file_path = self.ocrmill_config.output_folder / item.text()
        if not file_path.exists():
            return

        try:
            import csv
            with open(file_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                self.ocrmill_preview_table.setRowCount(0)
                return

            # First row is header
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []

            # Store headers and current file for correction tracking
            self._ocrmill_preview_headers = headers
            self._ocrmill_preview_file = item.text()

            # Block signals during population
            self._ocrmill_editing = True

            # Update table columns based on CSV headers
            self.ocrmill_preview_table.setColumnCount(len(headers))
            self.ocrmill_preview_table.setHorizontalHeaderLabels(headers)
            self.ocrmill_preview_table.setRowCount(len(data_rows))

            # Populate table and store original values
            self._ocrmill_original_values = {}
            for row_idx, row_data in enumerate(data_rows):
                for col_idx, cell_value in enumerate(row_data):
                    item_widget = QTableWidgetItem(str(cell_value))
                    self.ocrmill_preview_table.setItem(row_idx, col_idx, item_widget)
                    # Store original value for correction tracking
                    self._ocrmill_original_values[(row_idx, col_idx)] = str(cell_value)

            # Re-enable signals
            self._ocrmill_editing = False

            # Resize columns to content
            self.ocrmill_preview_table.resizeColumnsToContents()

            # Limit Description column (index 1) to ~50 characters wide but keep it expandable
            if len(headers) > 1:
                desc_col_idx = None
                for i, h in enumerate(headers):
                    if h.lower() == 'description':
                        desc_col_idx = i
                        break
                if desc_col_idx is not None:
                    # Approximate 50 characters at ~7 pixels per character
                    max_width = 350
                    if self.ocrmill_preview_table.columnWidth(desc_col_idx) > max_width:
                        self.ocrmill_preview_table.setColumnWidth(desc_col_idx, max_width)

            self.ocrmill_log(f"Preview loaded: {item.text()} ({len(data_rows)} rows)")

        except Exception as e:
            logger.error(f"Failed to preview file {file_path}: {e}")
            self.ocrmill_log(f"Failed to preview file: {e}")


    def ocrmill_output_mode_changed(self, button):
        """Handle output mode radio button change."""
        # Button ID 0 = split files, Button ID 1 = single file
        consolidate = self.ocrmill_output_mode_group.id(button) == 1
        self.ocrmill_config.consolidate_multi_invoice = consolidate
        set_user_setting('ocrmill_consolidate', 'true' if consolidate else 'false')
        mode_text = "single combined file" if consolidate else "separate files per invoice"
        self.ocrmill_log(f"Output mode changed to: {mode_text}")

    def ocrmill_toggle_monitoring(self):
        """Toggle folder monitoring on/off."""
        if self.ocrmill_monitor_btn.isChecked():
            self.ocrmill_worker.start_monitoring()
            self.ocrmill_monitor_btn.setText("Stop Monitoring")
            self.ocrmill_log(f"Started monitoring: {self.ocrmill_config.input_folder}")
        else:
            self.ocrmill_worker.stop_monitoring()
            self.ocrmill_monitor_btn.setText("Start Monitoring")
            self.ocrmill_log("Stopped monitoring")

    def ocrmill_process_single_file(self):
        """Process a single PDF file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select PDF Invoice",
            str(self.ocrmill_config.input_folder),
            "PDF Files (*.pdf)"
        )
        if file_path:
            self.ocrmill_log(f"Processing file: {file_path}")
            from ocrmill_worker import SingleFileWorker
            self.ocrmill_single_worker = SingleFileWorker(
                self.ocrmill_processor,
                Path(file_path),
                self.ocrmill_config.output_folder
            )
            self.ocrmill_single_worker.log_message.connect(self.ocrmill_log)
            self.ocrmill_single_worker.finished.connect(self.ocrmill_on_single_file_finished)
            self.ocrmill_single_worker.error.connect(lambda e: self.ocrmill_log(f"Error: {e}"))
            self.ocrmill_single_worker.start()

    def ocrmill_on_single_file_finished(self, items: list):
        """Handle completion of single file processing."""
        if items:
            self.ocrmill_last_items = items
            self.ocrmill_send_btn.setEnabled(True)
            self.ocrmill_log(f"Extracted {len(items)} items. Ready to send to Tariffmill.")
        else:
            self.ocrmill_log("No items extracted from file.")

    def ocrmill_process_dropped_files(self, file_paths: list):
        """Process PDF files dropped onto the drop zone using parallel processing."""
        if not file_paths:
            return

        # Use parallel processing for multiple files
        from ocrmill_worker import MultiFileWorker

        self.ocrmill_log(f"Processing {len(file_paths)} dropped file(s) in parallel...")

        # Create and start the parallel worker
        self.ocrmill_multi_worker = MultiFileWorker(
            self.ocrmill_processor,
            file_paths,
            self.ocrmill_config.output_folder
        )

        # Connect signals
        self.ocrmill_multi_worker.log_message.connect(self.ocrmill_log)
        self.ocrmill_multi_worker.progress.connect(self._ocrmill_on_multi_progress)
        self.ocrmill_multi_worker.all_finished.connect(self._ocrmill_on_multi_finished)
        self.ocrmill_multi_worker.error.connect(lambda e: self.ocrmill_log(f"Error: {e}"))

        # Start processing
        self.ocrmill_multi_worker.start()

    def _ocrmill_on_multi_progress(self, completed: int, total: int):
        """Handle progress updates from parallel processing."""
        # Could update a progress bar here if desired
        pass

    def _ocrmill_on_multi_finished(self, all_items: list):
        """Handle completion of parallel multi-file processing."""
        if all_items:
            self.ocrmill_last_items = all_items
            self.ocrmill_send_btn.setEnabled(True)

    def ocrmill_process_folder_now(self):
        """Process all PDFs in the input folder immediately using parallel processing."""
        from ocrmill_worker import ParallelFolderWorker

        self.ocrmill_log("Processing folder in parallel...")
        input_folder = self.ocrmill_config.input_folder
        output_folder = self.ocrmill_config.output_folder

        # Create and start the parallel folder worker
        self.ocrmill_folder_worker = ParallelFolderWorker(
            self.ocrmill_processor,
            input_folder,
            output_folder
        )

        # Connect signals
        self.ocrmill_folder_worker.log_message.connect(self.ocrmill_log)
        self.ocrmill_folder_worker.progress.connect(self._ocrmill_on_folder_progress)
        self.ocrmill_folder_worker.all_finished.connect(self._ocrmill_on_folder_finished)
        self.ocrmill_folder_worker.error.connect(lambda e: self.ocrmill_log(f"Error: {e}"))

        # Start processing
        self.ocrmill_folder_worker.start()

    def _ocrmill_on_folder_progress(self, completed: int, total: int):
        """Handle progress updates from parallel folder processing."""
        # Could update a progress bar here if desired
        pass

    def _ocrmill_on_folder_finished(self, total_items: int):
        """Handle completion of parallel folder processing."""
        self.ocrmill_log(f"Folder processing complete: {total_items} total items")

    def ocrmill_on_processing_finished(self, item_count: int):
        """Handle completion of batch processing."""
        self.ocrmill_log(f"Processing complete: {item_count} items extracted")

    def ocrmill_on_items_extracted(self, items: list):
        """Handle extracted items from worker."""
        if items:
            self.ocrmill_last_items = items
            self.ocrmill_send_btn.setEnabled(True)

    def ocrmill_send_to_tariffmill(self):
        """Move CSV files from OCR_Output to Tariffmill_Input folder."""
        # Get the OCRMill output folder (where CSV files are saved)
        ocr_output = self.ocrmill_config.output_folder
        tariffmill_input = INPUT_DIR

        # Find all CSV files in OCR_Output
        csv_files = list(ocr_output.glob("*.csv"))

        if not csv_files:
            QMessageBox.information(self, "No Files", "No CSV files found in OCR_Output folder.")
            return

        # Ensure Tariffmill_Input exists
        tariffmill_input.mkdir(parents=True, exist_ok=True)

        moved_count = 0
        errors = []

        for csv_file in csv_files:
            try:
                dest_path = tariffmill_input / csv_file.name
                # If file exists at destination, add a number suffix
                if dest_path.exists():
                    base = csv_file.stem
                    ext = csv_file.suffix
                    counter = 1
                    while dest_path.exists():
                        dest_path = tariffmill_input / f"{base}_{counter}{ext}"
                        counter += 1

                shutil.move(str(csv_file), str(dest_path))
                moved_count += 1
                self.ocrmill_log(f"Moved: {csv_file.name} -> {dest_path.name}")
            except Exception as e:
                errors.append(f"{csv_file.name}: {e}")
                logger.error(f"Failed to move {csv_file}: {e}")

        # Refresh file lists
        self.ocrmill_refresh_output_files()
        self.refresh_input_files()

        # Switch to Process Shipment tab
        self.tabs.setCurrentIndex(0)

        if errors:
            QMessageBox.warning(self, "Files Moved with Errors",
                f"Moved {moved_count} of {len(csv_files)} files to Tariffmill_Input.\n\n"
                f"Errors:\n" + "\n".join(errors))
        else:
            QMessageBox.information(self, "Files Moved",
                f"Moved {moved_count} CSV file(s) to Tariffmill_Input.\n\n"
                f"Destination: {tariffmill_input}")

    def ocrmill_refresh_templates(self):
        """Refresh the templates list by re-scanning the templates directory."""
        import re as regex_module

        # Re-discover templates from disk
        try:
            from templates import refresh_templates
            refresh_templates()
        except Exception as e:
            self.ocrmill_log(f"Warning: Could not refresh templates module: {e}")

        # Reload processor's template list
        self.ocrmill_processor.reload_templates()

        # Clear list and data
        self.ocrmill_templates_list.clear()
        self.ocrmill_templates_data = []

        # Templates directory
        templates_dir = BASE_DIR / "templates"
        if not templates_dir.exists():
            return

        # Excluded files
        excluded = {'__init__.py', 'base_template.py', 'sample_template.py', '__pycache__'}

        for file_path in sorted(templates_dir.glob("*.py")):
            if file_path.name in excluded:
                continue

            template_info = self._extract_template_info_from_file(file_path)
            if template_info:
                self.ocrmill_templates_data.append(template_info)

                # Create list item with template name
                display_text = template_info['name']

                item = QListWidgetItem(display_text)
                item.setToolTip(f"Supplier: {template_info['supplier']}\nClient: {template_info['client']}\nCountry: {template_info['country']}")
                self.ocrmill_templates_list.addItem(item)

    def _extract_template_info_from_file(self, file_path: Path) -> dict:
        """Extract template metadata from a template file."""
        import re as regex_module
        try:
            content = file_path.read_text(encoding='utf-8')

            info = {
                'file_path': str(file_path),
                'file_name': file_path.stem,
                'name': file_path.stem.replace('_', ' ').title(),
                'supplier': '',
                'client': '',
                'country': '',
                'ai_agent': ''
            }

            # Extract name
            name_match = regex_module.search(r'^\s*name\s*=\s*["\'](.+?)["\']', content, regex_module.MULTILINE)
            if name_match:
                info['name'] = name_match.group(1)

            # Extract description (often contains supplier info)
            desc_match = regex_module.search(r'^\s*description\s*=\s*["\'](.+?)["\']', content, regex_module.MULTILINE)
            if desc_match:
                info['supplier'] = desc_match.group(1)

            # Extract client
            client_match = regex_module.search(r'^\s*client\s*=\s*["\'](.+?)["\']', content, regex_module.MULTILINE)
            if client_match:
                info['client'] = client_match.group(1)

            # Try to extract country from docstring or content
            docstring_match = regex_module.search(r'"""[\s\S]*?"""', content)
            if docstring_match:
                docstring = docstring_match.group(0).lower()
                for pattern in ['china', 'india', 'usa', 'mexico', 'brazil', 'czech republic',
                               'el salvador', 'taiwan', 'japan', 'korea', 'vietnam']:
                    if pattern in docstring:
                        info['country'] = pattern.upper()
                        break

            # Check for AI metadata file
            ai_metadata_path = file_path.with_suffix('.ai_meta.json')
            if ai_metadata_path.exists():
                try:
                    import json
                    with open(ai_metadata_path, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                        info['ai_agent'] = metadata.get('provider', '')
                        if metadata.get('model'):
                            info['ai_agent'] += f" ({metadata.get('model')})"
                except:
                    info['ai_agent'] = 'AI'

            return info

        except Exception:
            return None

    def ocrmill_open_ai_generator(self):
        """Open the AI Template Generator dialog."""
        try:
            from ai_template_generator import AITemplateGeneratorDialog
            dialog = AITemplateGeneratorDialog(self)
            dialog.template_created.connect(self.ocrmill_on_template_created)
            dialog.exec_()
        except ImportError as e:
            QMessageBox.warning(
                self, "Import Error",
                f"Failed to load AI Template Generator: {e}\n\n"
                "Make sure ai_template_generator.py exists.\n"
                "For AI features, install: pip install openai anthropic"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open AI Generator: {e}")

    def ocrmill_on_template_created(self, template_name: str, file_path: str):
        """Handle new template creation."""
        self.ocrmill_log(f"New template created: {template_name} at {file_path}")
        self.ocrmill_log("Note: Restart the application to use the new template, or manually register it in templates/__init__.py")
        self.ocrmill_refresh_templates()

    def ocrmill_create_new_template(self):
        """Create a new invoice template from sample_template.py"""
        # Get template name from user
        name, ok = QInputDialog.getText(
            self, "Create New Template",
            "Enter template name (e.g., 'acme_corp'):\n\n"
            "Use lowercase with underscores, no spaces.",
            QLineEdit.Normal, ""
        )
        if not ok or not name:
            return

        # Validate name
        name = name.strip().lower().replace(' ', '_').replace('-', '_')
        if not name.replace('_', '').isalnum():
            QMessageBox.warning(self, "Invalid Name", "Template name must contain only letters, numbers, and underscores.")
            return

        # Check if template already exists
        templates_dir = BASE_DIR / "templates"
        new_template_path = templates_dir / f"{name}.py"
        if new_template_path.exists():
            QMessageBox.warning(self, "Template Exists", f"A template named '{name}' already exists.")
            return

        # Read sample template
        sample_path = templates_dir / "sample_template.py"
        if not sample_path.exists():
            QMessageBox.critical(self, "Error", "sample_template.py not found in templates folder.")
            return

        try:
            with open(sample_path, 'r', encoding='utf-8') as f:
                sample_content = f.read()

            # Create class name from template name
            class_name = ''.join(word.capitalize() for word in name.split('_')) + 'Template'

            # Replace sample values with new template name
            new_content = sample_content.replace('SampleTemplate', class_name)
            new_content = new_content.replace('Sample Template', name.replace('_', ' ').title())
            new_content = new_content.replace('Sample Client', 'Client Name')
            new_content = new_content.replace('enabled = False', 'enabled = True')

            # Write new template
            with open(new_template_path, 'w', encoding='utf-8') as f:
                f.write(new_content)

            # Ask user if they want to edit the template now
            reply = QMessageBox.question(
                self, "Template Created",
                f"Template '{name}' created successfully!\n\n"
                f"File: {new_template_path}\n\n"
                f"Would you like to open it for editing?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                self.ocrmill_open_template_file(new_template_path)

            # Show instructions
            QMessageBox.information(
                self, "Next Steps",
                f"To complete your template:\n\n"
                f"1. Edit {name}.py to customize extraction logic\n"
                f"2. Register in templates/__init__.py:\n\n"
                f"   from .{name} import {class_name}\n\n"
                f"   Add to TEMPLATE_REGISTRY:\n"
                f"   '{name}': {class_name},\n\n"
                f"3. Restart TariffMill or click Refresh"
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create template: {e}")

    def ocrmill_edit_template(self):
        """Edit the selected template file using AI Template Editor"""
        current_row = self.ocrmill_templates_table.currentRow()
        if current_row < 0:
            QMessageBox.information(self, "No Selection", "Please select a template to edit.")
            return

        # Get template info from stored data
        if current_row >= len(self.ocrmill_templates_data):
            QMessageBox.warning(self, "Error", "Template data not found.")
            return

        template_info = self.ocrmill_templates_data[current_row]
        template_path = Path(template_info['file_path'])

        logger.debug(f"Edit template - row: {current_row}, path: {template_path}")

        if template_path and template_path.exists():
            # Always open AI Template Editor for all templates
            self.ocrmill_open_ai_template_editor(template_path, template_info)
        else:
            QMessageBox.warning(self, "File Not Found", f"Template file not found: {template_path}")

    def ocrmill_open_ai_template_editor(self, template_path, template_info):
        """Open the AI Template Editor for editing templates."""
        try:
            import json
            from ai_template_generator import AITemplateEditorDialog

            template_path = Path(template_path)

            # Check if AI metadata exists
            ai_metadata_path = template_path.with_suffix('.ai_meta.json')
            if ai_metadata_path.exists():
                # Load existing AI metadata
                with open(ai_metadata_path, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
            else:
                # Create metadata from template_info for non-AI templates
                metadata = {
                    'provider': '',
                    'model': '',
                    'invoice_text': '',
                    'conversation_history': [],
                    'name': template_info.get('name', template_path.stem),
                    'supplier': template_info.get('supplier', ''),
                    'client': template_info.get('client', ''),
                    'country': template_info.get('country', '')
                }

            dialog = AITemplateEditorDialog(str(template_path), metadata, self)
            dialog.template_modified.connect(lambda: self.ocrmill_refresh_templates())
            dialog.exec_()

        except ImportError as e:
            QMessageBox.warning(
                self, "Import Error",
                f"Failed to load AI Template Editor: {e}\n\n"
                "Falling back to regular file editor."
            )
            self.ocrmill_open_template_file(template_path)
        except Exception as e:
            QMessageBox.warning(
                self, "Error",
                f"Failed to open AI Template Editor: {e}\n\n"
                "Falling back to regular file editor."
            )
            self.ocrmill_open_template_file(template_path)

    def ocrmill_open_template_file(self, file_path):
        """Open a template file in the default editor"""
        import os
        import subprocess
        import platform

        file_path = Path(file_path)
        if not file_path.exists():
            QMessageBox.warning(self, "File Not Found", f"Template file not found:\n{file_path}")
            return

        logger.debug(f"Attempting to open template file: {file_path}")

        try:
            if platform.system() == 'Windows':
                # Use os.startfile which opens with the default .py editor
                os.startfile(str(file_path))
                self.ocrmill_log(f"Opened template: {file_path.name}")
                logger.info(f"Opened template file with default editor: {file_path}")
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(file_path)])
                self.ocrmill_log(f"Opened template: {file_path.name}")
            else:  # Linux
                subprocess.run(['xdg-open', str(file_path)])
                self.ocrmill_log(f"Opened template: {file_path.name}")
        except Exception as e:
            logger.error(f"Failed to open template file: {e}")
            # Fallback: show path to user
            QMessageBox.information(
                self, "Open Template",
                f"Please open the following file in your text editor:\n\n{file_path}"
            )

    def ocrmill_delete_template(self):
        """Delete the selected template file"""
        current_row = self.ocrmill_templates_list.currentRow()
        if current_row < 0:
            QMessageBox.information(self, "No Selection", "Please select a template to delete.")
            return

        # Get template info from stored data
        if current_row >= len(self.ocrmill_templates_data):
            QMessageBox.warning(self, "Error", "Template data not found.")
            return

        template_info = self.ocrmill_templates_data[current_row]
        template_path = Path(template_info['file_path'])
        template_name = template_info['name']

        if template_path and template_path.exists():
            # Confirm deletion
            reply = QMessageBox.question(
                self, "Confirm Delete",
                f"Are you sure you want to delete the template:\n\n{template_name}\n({template_path.name})\n\nThis action cannot be undone.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                try:
                    template_path.unlink()
                    # Also delete AI metadata file if exists
                    ai_metadata_path = template_path.with_suffix('.ai_meta.json')
                    if ai_metadata_path.exists():
                        ai_metadata_path.unlink()
                    self.ocrmill_log(f"Deleted template: {template_path.name}")
                    self.ocrmill_refresh_templates()
                    QMessageBox.information(self, "Deleted", f"Template '{template_name}' has been deleted.")
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to delete template:\n{str(e)}")
        else:
            QMessageBox.warning(self, "File Not Found", f"Template file not found: {template_path}")

    def _duplicate_template(self):
        """Create a copy of the selected template with a new name."""
        current_row = self.ocrmill_templates_list.currentRow()
        if current_row < 0:
            QMessageBox.information(self, "No Selection", "Please select a template to duplicate.")
            return

        if current_row >= len(self.ocrmill_templates_data):
            QMessageBox.warning(self, "Error", "Template data not found.")
            return

        template_info = self.ocrmill_templates_data[current_row]
        template_path = Path(template_info['file_path'])
        original_name = template_info['name']

        if not template_path.exists():
            QMessageBox.warning(self, "File Not Found", f"Template file not found: {template_path}")
            return

        # Prompt for new template name
        new_name, ok = QInputDialog.getText(
            self, "Duplicate Template",
            f"Enter a name for the new template:\n\n(copying from: {original_name})",
            QLineEdit.Normal,
            f"{original_name}_copy"
        )

        if not ok or not new_name.strip():
            return

        new_name = new_name.strip()

        # Validate name (lowercase, underscores only)
        import re
        safe_name = re.sub(r'[^a-z0-9_]', '_', new_name.lower())

        # Create new filename
        new_filename = f"{safe_name}.py"
        new_path = template_path.parent / new_filename

        if new_path.exists():
            QMessageBox.warning(
                self, "File Exists",
                f"A template with this name already exists:\n{new_filename}\n\nPlease choose a different name."
            )
            return

        try:
            # Read original template
            with open(template_path, 'r', encoding='utf-8') as f:
                original_content = f.read()

            # Replace class name
            original_class = template_info.get('class_name', f"{original_name.title().replace('_', '')}Template")
            new_class = f"{safe_name.title().replace('_', '')}Template"

            # Replace the class name in content
            new_content = original_content.replace(original_class, new_class)

            # Also update any name = "..." or name = '...' references
            new_content = re.sub(
                r'(name\s*=\s*["\'])' + re.escape(original_name) + r'(["\'])',
                r'\g<1>' + safe_name + r'\g<2>',
                new_content
            )

            # Write new template
            with open(new_path, 'w', encoding='utf-8') as f:
                f.write(new_content)

            # Copy AI metadata if exists
            original_meta = template_path.with_suffix('.ai_meta.json')
            if original_meta.exists():
                import json
                with open(original_meta, 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                # Update metadata
                meta['template_name'] = safe_name
                meta['class_name'] = new_class
                meta['duplicated_from'] = original_name
                meta['duplicated_date'] = datetime.now().isoformat()
                new_meta_path = new_path.with_suffix('.ai_meta.json')
                with open(new_meta_path, 'w', encoding='utf-8') as f:
                    json.dump(meta, f, indent=2)

            self.ocrmill_log(f"Duplicated template: {original_name} -> {safe_name}")

            # Refresh templates and select new one
            self.ocrmill_processor.reload_templates()
            self.ocrmill_refresh_templates()

            QMessageBox.information(
                self, "Template Duplicated",
                f"Template duplicated successfully!\n\nNew template: {safe_name}\nFile: {new_filename}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to duplicate template:\n{str(e)}")

    def refresh_parts_table(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            # Use explicit column ordering to match header labels
            df = pd.read_sql("""
                SELECT part_number, description, hts_code, country_origin, mid, client_code,
                       steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio, non_steel_ratio,
                       qty_unit, hts_verified, Sec301_Exclusion_Tariff, last_updated as updated_date
                FROM parts_master ORDER BY part_number
            """, conn)
            conn.close()
            self.populate_parts_table(df)
            self._update_search_result(f"Showing all {len(df)} parts", "info")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot load parts:\n{e}")

    def import_hts_units_silent(self, part_numbers=None):
        """
        Silently import CBP Qty1 units from hts.db for specific parts.
        Called automatically after saving new parts or reprocessing parts.

        Args:
            part_numbers: List of part numbers to update (if None, updates all parts)

        Returns:
            Number of parts updated, or -1 if hts.db not found
        """
        hts_db_path = RESOURCES_DIR / "References" / "hts.db"

        if not hts_db_path.exists():
            logger.debug("hts.db not found, skipping HTS units import")
            return -1

        try:
            # Clean up HTS codes (remove dots for matching)
            def normalize_hts(hts):
                if pd.isna(hts) or hts is None:
                    return ""
                return str(hts).replace(".", "").strip()

            # Load unit_of_quantity lookup from hts.db
            hts_conn = sqlite3.connect(str(hts_db_path))
            hts_cursor = hts_conn.cursor()
            hts_cursor.execute("SELECT full_code, unit_of_quantity FROM hts_codes WHERE unit_of_quantity IS NOT NULL AND unit_of_quantity != ''")
            hts_units = {row[0]: row[1] for row in hts_cursor.fetchall()}
            hts_conn.close()

            if not hts_units:
                logger.debug("No unit_of_quantity data found in hts.db")
                return 0

            # Update parts_master database
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Get parts with HTS codes (filter by part_numbers if provided)
            if part_numbers:
                placeholders = ','.join(['?' for _ in part_numbers])
                c.execute(f"SELECT part_number, hts_code FROM parts_master WHERE hts_code IS NOT NULL AND hts_code != '' AND part_number IN ({placeholders})", part_numbers)
            else:
                c.execute("SELECT part_number, hts_code FROM parts_master WHERE hts_code IS NOT NULL AND hts_code != ''")
            parts = c.fetchall()

            updated = 0
            for part_number, hts_code in parts:
                normalized = normalize_hts(hts_code)
                if normalized in hts_units:
                    c.execute("UPDATE parts_master SET qty_unit=? WHERE part_number=?",
                              (hts_units[normalized], part_number))
                    updated += c.rowcount

            conn.commit()
            conn.close()

            if updated > 0:
                logger.info(f"Silently updated {updated} parts with Qty Unit values from hts.db")

            return updated

        except Exception as e:
            logger.error(f"Silent HTS units import failed: {e}")
            return -1

    def verify_hts_codes_in_parts_master(self, part_numbers=None):
        """
        Verify HTS codes in parts_master against hts.db and update hts_verified column.

        For each part:
        - If HTS found in hts.db: sets hts_verified to "Verified YYYY-MM-DD"
        - If HTS NOT found in hts.db: sets hts_verified to "Invalid HTS - YYYY-MM-DD"
        - Also updates qty_unit from hts.db if HTS is valid

        Args:
            part_numbers: List of part numbers to verify (if None, verifies all parts)

        Returns:
            Tuple of (verified_count, invalid_count, total_checked)
        """
        from datetime import datetime

        hts_db_path = RESOURCES_DIR / "References" / "hts.db"
        today = datetime.now().strftime("%Y-%m-%d")

        if not hts_db_path.exists():
            logger.warning("hts.db not found, cannot verify HTS codes")
            return (0, 0, 0)

        try:
            # Clean up HTS codes (remove dots for matching)
            def normalize_hts(hts):
                if pd.isna(hts) or hts is None:
                    return ""
                return str(hts).replace(".", "").strip()

            # Load all valid HTS codes and their units from hts.db
            hts_conn = sqlite3.connect(str(hts_db_path))
            hts_cursor = hts_conn.cursor()
            hts_cursor.execute("SELECT full_code, unit_of_quantity FROM hts_codes")
            hts_data = {row[0]: row[1] for row in hts_cursor.fetchall()}
            hts_conn.close()

            if not hts_data:
                logger.warning("No HTS codes found in hts.db")
                return (0, 0, 0)

            # Update parts_master database
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()

            # Get parts with HTS codes (filter by part_numbers if provided)
            if part_numbers:
                placeholders = ','.join(['?' for _ in part_numbers])
                c.execute(f"SELECT part_number, hts_code FROM parts_master WHERE hts_code IS NOT NULL AND hts_code != '' AND part_number IN ({placeholders})", part_numbers)
            else:
                c.execute("SELECT part_number, hts_code FROM parts_master WHERE hts_code IS NOT NULL AND hts_code != ''")
            parts = c.fetchall()

            verified_count = 0
            invalid_count = 0

            for part_number, hts_code in parts:
                normalized = normalize_hts(hts_code)

                if normalized in hts_data:
                    # HTS is valid - update verified status and qty_unit
                    qty_unit = hts_data[normalized] or ""
                    c.execute("""UPDATE parts_master
                                SET hts_verified=?, qty_unit=?
                                WHERE part_number=?""",
                              (f"Verified {today}", qty_unit, part_number))
                    verified_count += 1
                else:
                    # HTS is invalid - mark as invalid
                    c.execute("""UPDATE parts_master
                                SET hts_verified=?
                                WHERE part_number=?""",
                              (f"Invalid HTS - {today}", part_number))
                    invalid_count += 1

            conn.commit()
            conn.close()

            total_checked = verified_count + invalid_count
            logger.info(f"HTS verification complete: {verified_count} verified, {invalid_count} invalid, {total_checked} total")

            return (verified_count, invalid_count, total_checked)

        except Exception as e:
            logger.error(f"HTS verification failed: {e}")
            return (0, 0, 0)

    def verify_hts_codes_with_dialog(self):
        """Verify HTS codes in parts_master with progress dialog and results message."""
        from PyQt5.QtWidgets import QProgressDialog
        from PyQt5.QtCore import Qt

        # Confirm action
        reply = QMessageBox.question(
            self,
            "Verify HTS Codes",
            "This will verify all HTS codes in the Parts Master against the HTS database.\n\n"
            "- Valid HTS codes will be marked as 'Verified' with today's date\n"
            "- Invalid HTS codes will be marked as 'Invalid HTS' with today's date\n"
            "- Qty Unit values will be updated for valid HTS codes\n\n"
            "Do you want to continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )

        if reply != QMessageBox.Yes:
            return

        # Show progress dialog
        progress = QProgressDialog("Verifying HTS codes...", None, 0, 0, self)
        progress.setWindowTitle("HTS Verification")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setMinimumWidth(300)
        progress.setValue(0)
        QApplication.processEvents()

        # Run verification
        verified, invalid, total = self.verify_hts_codes_in_parts_master()

        progress.close()

        # Show results
        if total == 0:
            QMessageBox.information(
                self, "Verification Complete",
                "No parts with HTS codes found to verify."
            )
        else:
            QMessageBox.information(
                self, "Verification Complete",
                f"HTS Code Verification Results:\n\n"
                f"Total parts checked: {total:,}\n"
                f"Valid HTS codes: {verified:,}\n"
                f"Invalid HTS codes: {invalid:,}\n\n"
                f"The 'hts_verified' column has been updated.\n"
                f"Qty Unit values have been updated for valid codes."
            )

        # Refresh the parts table to show updated values
        self.refresh_parts_table()

    def export_missing_hts_codes(self):
        """Export HTS codes that are missing from hts.db or missing Qty Unit values, including part numbers."""
        hts_db_path = RESOURCES_DIR / "References" / "hts.db"

        try:
            def normalize_hts(hts):
                if pd.isna(hts) or hts is None:
                    return ""
                return str(hts).replace(".", "").strip()

            # Load all valid HTS codes from hts.db
            all_valid_hts = set()
            if hts_db_path.exists():
                hts_conn = sqlite3.connect(str(hts_db_path))
                hts_cursor = hts_conn.cursor()
                hts_cursor.execute("SELECT full_code FROM hts_codes")
                all_valid_hts = {row[0] for row in hts_cursor.fetchall()}
                hts_conn.close()

            # Query database for all parts with HTS codes (including part_number and client_code)
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("""
                SELECT part_number, hts_code, qty_unit, client_code
                FROM parts_master
                WHERE hts_code IS NOT NULL
                AND hts_code != ''
                ORDER BY client_code, hts_code, part_number
            """)
            all_parts = c.fetchall()
            conn.close()

            # Categorize parts
            invalid_parts = []  # Parts with HTS not found in hts.db
            missing_unit_parts = []  # Parts with valid HTS but missing qty_unit

            for part_number, hts_code, qty_unit, client_code in all_parts:
                normalized = normalize_hts(hts_code)
                if not normalized:
                    continue

                if normalized not in all_valid_hts:
                    invalid_parts.append((part_number, hts_code, client_code, "INVALID - Not in HTS Database"))
                elif not qty_unit or pd.isna(qty_unit) or str(qty_unit).strip() == '':
                    missing_unit_parts.append((part_number, hts_code, client_code, "Missing Qty Unit"))

            if not invalid_parts and not missing_unit_parts:
                QMessageBox.information(self, "No Issues Found",
                    "All HTS codes are valid and have Qty Unit values assigned.")
                return

            # Create export data with part numbers and client codes
            export_data = []
            for part_number, hts, client_code, status in invalid_parts:
                export_data.append({'Part Number': part_number, 'Client Code': client_code or '', 'HTS Code': hts, 'Status': status, 'Qty Unit': ''})
            for part_number, hts, client_code, status in missing_unit_parts:
                export_data.append({'Part Number': part_number, 'Client Code': client_code or '', 'HTS Code': hts, 'Status': status, 'Qty Unit': ''})

            df_export = pd.DataFrame(export_data)

            # Ask user where to save
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export HTS Issues",
                str(BASE_DIR / "Resources" / "References" / "HTS_Issues.xlsx"),
                "Excel Files (*.xlsx);;CSV Files (*.csv)"
            )

            if not file_path:
                return

            # Save the file
            if file_path.endswith('.csv'):
                df_export.to_csv(file_path, index=False)
            else:
                df_export.to_excel(file_path, index=False)

            # Show summary
            msg = f"Exported {len(export_data)} parts with HTS issues:\n\n"
            if invalid_parts:
                msg += f"  • {len(invalid_parts)} parts with INVALID HTS (not found in hts.db)\n"
            if missing_unit_parts:
                msg += f"  • {len(missing_unit_parts)} parts missing Qty Unit\n"
            msg += f"\nFile: {file_path}"

            QMessageBox.information(self, "Export Complete", msg)

            # Ask if user wants to open the file
            reply = QMessageBox.question(self, "Open File?",
                "Would you like to open the exported file now?",
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                import os
                os.startfile(str(file_path))

        except Exception as e:
            logger.error(f"Failed to export HTS issues: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export HTS issues:\n{e}")

    def export_parts_by_client(self):
        """Export parts list filtered by client code to Excel."""
        try:
            # Get list of unique client codes from database
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("""
                SELECT DISTINCT client_code
                FROM parts_master
                WHERE client_code IS NOT NULL AND client_code != ''
                ORDER BY client_code
            """)
            client_codes = [row[0] for row in c.fetchall()]
            conn.close()

            if not client_codes:
                QMessageBox.warning(self, "No Client Codes",
                    "No parts have client codes assigned.\n\n"
                    "Add client codes to parts in the Parts Master table first.")
                return

            # Show dialog to select client code
            dialog = QDialog(self)
            dialog.setWindowTitle("Export Parts by Client")
            dialog.resize(400, 300)
            layout = QVBoxLayout(dialog)

            layout.addWidget(QLabel("<h3>Select Client Code to Export</h3>"))

            # Client code selection
            client_list = QListWidget()
            client_list.addItems(client_codes)
            client_list.setSelectionMode(QListWidget.SingleSelection)
            if client_codes:
                client_list.setCurrentRow(0)
            layout.addWidget(client_list)

            # Export all option
            export_all_cb = QCheckBox("Export ALL parts (ignore client filter)")
            layout.addWidget(export_all_cb)

            # Buttons
            btn_layout = QHBoxLayout()
            btn_export = QPushButton("Export to Excel")
            btn_export.setStyleSheet(self.get_button_style("primary"))
            btn_cancel = QPushButton("Cancel")
            btn_layout.addStretch()
            btn_layout.addWidget(btn_export)
            btn_layout.addWidget(btn_cancel)
            layout.addLayout(btn_layout)

            btn_cancel.clicked.connect(dialog.reject)

            def do_export():
                if export_all_cb.isChecked():
                    selected_client = None
                else:
                    selected_items = client_list.selectedItems()
                    if not selected_items:
                        QMessageBox.warning(dialog, "No Selection", "Please select a client code or check 'Export ALL'.")
                        return
                    selected_client = selected_items[0].text()

                # Query parts
                conn = sqlite3.connect(str(DB_PATH))
                if selected_client:
                    df = pd.read_sql("""
                        SELECT part_number, description, hts_code, country_origin, mid, client_code,
                               steel_ratio as 'steel_%', aluminum_ratio as 'aluminum_%',
                               copper_ratio as 'copper_%', wood_ratio as 'wood_%',
                               auto_ratio as 'auto_%', non_steel_ratio as 'non_steel_%',
                               qty_unit, Sec301_Exclusion_Tariff, last_updated
                        FROM parts_master
                        WHERE client_code = ?
                        ORDER BY part_number
                    """, conn, params=[selected_client])
                else:
                    df = pd.read_sql("""
                        SELECT part_number, description, hts_code, country_origin, mid, client_code,
                               steel_ratio as 'steel_%', aluminum_ratio as 'aluminum_%',
                               copper_ratio as 'copper_%', wood_ratio as 'wood_%',
                               auto_ratio as 'auto_%', non_steel_ratio as 'non_steel_%',
                               qty_unit, Sec301_Exclusion_Tariff, last_updated
                        FROM parts_master
                        ORDER BY client_code, part_number
                    """, conn)
                conn.close()

                if df.empty:
                    QMessageBox.warning(dialog, "No Parts", "No parts found for the selected criteria.")
                    return

                # Choose save location
                default_name = f"parts_{selected_client}.xlsx" if selected_client else "parts_all.xlsx"
                save_path, _ = QFileDialog.getSaveFileName(
                    dialog, "Save Parts Export", str(OUTPUT_DIR / default_name),
                    "Excel Files (*.xlsx)"
                )

                if not save_path:
                    return

                # Export to Excel
                try:
                    df.to_excel(save_path, index=False)
                    dialog.accept()

                    QMessageBox.information(self, "Export Complete",
                        f"Exported {len(df)} parts to:\n{save_path}")

                    # Ask to open file
                    reply = QMessageBox.question(self, "Open File?",
                        "Would you like to open the exported file?",
                        QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        import os
                        os.startfile(save_path)

                except Exception as e:
                    QMessageBox.critical(dialog, "Export Error", f"Failed to save file:\n{e}")

            btn_export.clicked.connect(do_export)
            self.center_dialog(dialog)
            dialog.exec_()

        except Exception as e:
            logger.error(f"Failed to export parts by client: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export parts:\n{e}")

    def add_part_row(self):
        row = self.parts_table.rowCount()
        self.parts_table.insertRow(row)
        self.parts_table.setItem(row, 0, QTableWidgetItem("NEW_PART"))

    def delete_selected_parts(self):
        rows = sorted(set(index.row() for index in self.parts_table.selectedIndexes()), reverse=True)
        if not rows:
            QMessageBox.information(self, "Info", "Select rows to delete")
            return
        if QMessageBox.question(self, "Confirm", f"Delete {len(rows)} parts?") != QMessageBox.Yes:
            return
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        deleted = 0
        for row in rows:
            part = self.parts_table.item(row, 0).text().strip()
            if part and part != "NEW_PART":
                c.execute("DELETE FROM parts_master WHERE part_number=?", (part,))
                deleted += c.rowcount
                self.parts_table.removeRow(row)
        conn.commit(); conn.close()
        QMessageBox.information(self, "Success", f"Deleted {deleted} parts")
        self.load_available_mids()

    def save_parts_table(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            now = datetime.now().isoformat()
            saved = 0
            renamed = 0
            inserted = 0
            updated = 0
            unchanged = 0
            for row in range(self.parts_table.rowCount()):
                items = [self.parts_table.item(row, col) for col in range(15)]
                if not items[0] or not items[0].text().strip(): continue
                part = items[0].text().strip()

                # Check if part number was renamed (original stored in UserRole)
                original_part = items[0].data(Qt.UserRole)
                if original_part and original_part != part:
                    # Part number was renamed - delete old record first
                    c.execute("DELETE FROM parts_master WHERE part_number = ?", (original_part,))
                    if c.rowcount:
                        renamed += 1
                        logger.info(f"Renamed part: '{original_part}' -> '{part}'")
                    # Update UserRole to new part number for future saves
                    items[0].setData(Qt.UserRole, part)

                desc = items[1].text() if items[1] else ""
                hts = items[2].text() if items[2] else ""
                origin = (items[3].text() or "").upper()[:2]
                mid = items[4].text() if items[4] else ""
                client_code = items[5].text() if items[5] else ""
                # Parse percentage values (0-100 format)
                try:
                    steel = float(items[6].text()) if items[6] and items[6].text() else 0.0
                    steel = max(0.0, min(100.0, steel))
                except:
                    steel = 0.0
                try:
                    aluminum = float(items[7].text()) if items[7] and items[7].text() else 0.0
                    aluminum = max(0.0, min(100.0, aluminum))
                except:
                    aluminum = 0.0
                try:
                    copper = float(items[8].text()) if items[8] and items[8].text() else 0.0
                    copper = max(0.0, min(100.0, copper))
                except:
                    copper = 0.0
                try:
                    wood = float(items[9].text()) if items[9] and items[9].text() else 0.0
                    wood = max(0.0, min(100.0, wood))
                except:
                    wood = 0.0
                try:
                    auto = float(items[10].text()) if items[10] and items[10].text() else 0.0
                    auto = max(0.0, min(100.0, auto))
                except:
                    auto = 0.0
                # Non-232 percentage is remainder after all Section 232 materials
                non_steel = max(0.0, 100.0 - steel - aluminum - copper - wood - auto)
                qty_unit = items[12].text() if items[12] else ""
                # Auto-lookup qty_unit from hts_units table if not set but HTS exists
                if not qty_unit and hts:
                    qty_unit = get_hts_qty_unit(hts)
                sec301_exclusion = items[13].text() if items[13] else ""

                # Check if this part exists and if data has changed
                c.execute("""SELECT description, hts_code, country_origin, mid, client_code,
                          steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                          non_steel_ratio, qty_unit, Sec301_Exclusion_Tariff
                          FROM parts_master WHERE part_number = ?""", (part,))
                existing = c.fetchone()

                if existing is None:
                    # New part - insert with current timestamp
                    c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                              steel_ratio, non_steel_ratio, last_updated, qty_unit, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                              Sec301_Exclusion_Tariff)
                              VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                              (part, desc, hts, origin, mid, client_code, steel, non_steel, now, qty_unit, aluminum, copper, wood, auto, sec301_exclusion))
                    inserted += 1
                    saved += 1
                else:
                    # Part exists - check if any field changed
                    # existing: (desc, hts, origin, mid, client, steel, al, cu, wood, auto, non_steel, qty_unit, sec301)
                    old_desc, old_hts, old_origin, old_mid, old_client, old_steel, old_al, old_cu, old_wood, old_auto, old_non, old_qty, old_sec301 = existing
                    # Normalize for comparison (handle None values)
                    old_desc = old_desc or ""
                    old_hts = old_hts or ""
                    old_origin = old_origin or ""
                    old_mid = old_mid or ""
                    old_client = old_client or ""
                    old_steel = old_steel or 0.0
                    old_al = old_al or 0.0
                    old_cu = old_cu or 0.0
                    old_wood = old_wood or 0.0
                    old_auto = old_auto or 0.0
                    old_non = old_non or 0.0
                    old_qty = old_qty or ""
                    old_sec301 = old_sec301 or ""

                    # Compare values (use small epsilon for float comparison)
                    eps = 0.001
                    has_changes = (
                        desc != old_desc or
                        hts != old_hts or
                        origin != old_origin or
                        mid != old_mid or
                        client_code != old_client or
                        abs(steel - old_steel) > eps or
                        abs(aluminum - old_al) > eps or
                        abs(copper - old_cu) > eps or
                        abs(wood - old_wood) > eps or
                        abs(auto - old_auto) > eps or
                        abs(non_steel - old_non) > eps or
                        qty_unit != old_qty or
                        sec301_exclusion != old_sec301
                    )

                    if has_changes:
                        # Update only if data changed
                        c.execute("""UPDATE parts_master SET
                                  description=?, hts_code=?, country_origin=?, mid=?, client_code=?,
                                  steel_ratio=?, non_steel_ratio=?, last_updated=?, qty_unit=?,
                                  aluminum_ratio=?, copper_ratio=?, wood_ratio=?, auto_ratio=?,
                                  Sec301_Exclusion_Tariff=?
                                  WHERE part_number=?""",
                                  (desc, hts, origin, mid, client_code, steel, non_steel, now, qty_unit, aluminum, copper, wood, auto, sec301_exclusion, part))
                        updated += 1
                        saved += 1
                    else:
                        unchanged += 1

            conn.commit(); conn.close()
            # Build success message with breakdown
            if saved > 0:
                details = []
                if inserted > 0:
                    details.append(f"{inserted} new")
                if updated > 0:
                    details.append(f"{updated} updated")
                if renamed > 0:
                    details.append(f"{renamed} renamed")
                msg = f"Saved {saved} parts!"
                if details:
                    msg += f"\n({', '.join(details)})"
            else:
                msg = "No changes to save."
            QMessageBox.information(self, "Success", msg)
            self.bottom_status.setText("Database saved")
            self.load_available_mids()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save failed:\n{e}")

    def add_not_found_parts_to_db(self):
        """
        Add parts that were not found in the database during processing.
        Uses the part number and MID from the current preview table.
        Returns the count of parts added.
        """
        if self.last_processed_df is None:
            return 0

        # Get rows where _not_in_db is True
        not_found_df = self.last_processed_df[self.last_processed_df['_not_in_db'] == True]
        if not_found_df.empty:
            return 0

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            now = datetime.now().isoformat()
            added_count = 0
            added_parts = []

            for idx, row in not_found_df.iterrows():
                part_number = str(row.get('Product No', '')).strip()
                if not part_number:
                    continue

                # Check if part already exists (might have been added manually since processing)
                c.execute("SELECT 1 FROM parts_master WHERE part_number = ?", (part_number,))
                if c.fetchone():
                    continue  # Part already exists, skip

                # Get MID from the table at the corresponding row index
                # Find the table row that matches this part number
                table_row = None
                for i in range(self.table.rowCount()):
                    item = self.table.item(i, 0)  # Column 0 is Product No
                    if item and item.text() == part_number:
                        table_row = i
                        break

                if table_row is None:
                    continue

                # Get values from the preview table
                mid = self.table.item(table_row, 3).text() if self.table.item(table_row, 3) else ""
                hts_code = self.table.item(table_row, 2).text() if self.table.item(table_row, 2) else ""

                # Auto-lookup qty_unit from hts_units table based on HTS code
                qty_unit = get_hts_qty_unit(hts_code) if hts_code else ""

                # Insert the part with minimal information (part_number and MID)
                # Percentages are in 0-100 format; default to 100% non-232
                c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                          steel_ratio, non_steel_ratio, last_updated, qty_unit, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                          Sec301_Exclusion_Tariff)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                          (part_number, '', hts_code, '', mid, '', 0.0, 100.0, now, qty_unit, 0.0, 0.0, 0.0, 0.0, ''))

                if c.rowcount:
                    added_count += 1
                    added_parts.append(part_number)

            conn.commit()
            conn.close()

            if added_count > 0:
                logger.info(f"Added {added_count} new parts to database: {added_parts}")
                # Refresh the MID dropdown in case new MIDs were added
                self.load_available_mids()

            return added_count

        except Exception as e:
            logger.error(f"Failed to add not-found parts to database: {e}")
            return 0

    def filter_parts_table(self, text):
        text = text.lower().strip()
        if not text:
            for row in range(self.parts_table.rowCount()):
                self.parts_table.setRowHidden(row, False)
            return
        for row in range(self.parts_table.rowCount()):
            match = any(text in (self.parts_table.item(row, col).text() if self.parts_table.item(row, col) else "").lower() 
                       for col in range(self.parts_table.columnCount()))
            self.parts_table.setRowHidden(row, not match)

    def run_custom_sql(self):
        """Execute custom SQL query"""
        try:
            sql = self.custom_sql_input.text().strip()

            if not sql:
                QMessageBox.warning(self, "Query Error", "Please enter a SQL query.")
                return

            # Basic validation - must be SELECT and FROM parts_master
            if not sql.upper().startswith("SELECT"):
                QMessageBox.warning(self, "Query Error", "Only SELECT queries are allowed.")
                return

            if "parts_master" not in sql.lower():
                QMessageBox.warning(self, "Query Error", "Query must reference 'parts_master' table.")
                return

            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn)
            conn.close()

            self.populate_parts_table(df)
            self._update_search_result(f"Custom SQL returned {len(df)} parts", "success")
            logger.info(f"Custom SQL executed: {sql}")

        except Exception as e:
            logger.error(f"Custom SQL execution failed: {e}")
            self._update_search_result(f"SQL Error: {str(e)}", "error")
            QMessageBox.critical(self, "SQL Error", f"Failed to execute SQL:\n{e}")

    # ===== New Parts Search Methods =====

    def _populate_search_filters(self):
        """Populate the client and country filter dropdowns from database."""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            cursor = conn.cursor()

            # Get distinct clients
            cursor.execute("SELECT DISTINCT client_code FROM parts_master WHERE client_code IS NOT NULL AND client_code != '' ORDER BY client_code")
            clients = cursor.fetchall()
            for (client,) in clients:
                self.filter_client_combo.addItem(client, client)

            # Get distinct countries
            cursor.execute("SELECT DISTINCT country_origin FROM parts_master WHERE country_origin IS NOT NULL AND country_origin != '' ORDER BY country_origin")
            countries = cursor.fetchall()
            for (country,) in countries:
                self.filter_country_combo.addItem(country, country)

            conn.close()
        except Exception as e:
            logger.error(f"Failed to populate search filters: {e}")

    def _update_search_result(self, message: str, status: str = "info"):
        """Update the search result label with appropriate styling."""
        styles = {
            "success": "padding: 8px; background: #d4edda; border-radius: 4px; color: #155724;",
            "error": "padding: 8px; background: #f8d7da; border-radius: 4px; color: #721c24;",
            "warning": "padding: 8px; background: #fff3cd; border-radius: 4px; color: #856404;",
            "info": "padding: 8px; background: #e8f4f8; border-radius: 4px; color: #2c3e50;",
        }
        self.search_result_label.setText(message)
        self.search_result_label.setStyleSheet(styles.get(status, styles["info"]))

    def toggle_advanced_sql(self, checked: bool):
        """Show or hide the advanced SQL input."""
        self.advanced_sql_widget.setVisible(checked)

    def clear_search_filters(self):
        """Clear all search filters and show all parts."""
        self.filter_client_combo.setCurrentIndex(0)
        self.filter_country_combo.setCurrentIndex(0)
        self.search_value_input.clear()
        self.refresh_parts_table()

    def search_missing_hts(self):
        """Find parts with missing or empty HTS codes."""
        try:
            sql = """
                SELECT part_number, description, hts_code, country_origin, mid, client_code,
                       steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio, non_steel_ratio,
                       qty_unit, hts_verified, Sec301_Exclusion_Tariff, last_updated as updated_date
                FROM parts_master
                WHERE hts_code IS NULL OR hts_code = '' OR hts_code = 'UNKNOWN'
                ORDER BY part_number
            """
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn)
            conn.close()

            self.populate_parts_table(df)
            self._update_search_result(f"Found {len(df)} parts with missing HTS codes", "warning" if len(df) > 0 else "success")

        except Exception as e:
            logger.error(f"Search failed: {e}")
            self._update_search_result(f"Search error: {e}", "error")

    def search_invalid_hts(self):
        """Find parts with invalid HTS codes (not found in hts.db)."""
        try:
            sql = """
                SELECT part_number, description, hts_code, country_origin, mid, client_code,
                       steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio, non_steel_ratio,
                       qty_unit, hts_verified, Sec301_Exclusion_Tariff, last_updated as updated_date
                FROM parts_master
                WHERE hts_verified LIKE 'Invalid HTS%'
                ORDER BY part_number
            """
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn)
            conn.close()

            self.populate_parts_table(df)
            if len(df) > 0:
                self._update_search_result(f"Found {len(df)} parts with invalid HTS codes", "error")
            else:
                self._update_search_result("No invalid HTS codes found. Run 'Verify HTS' to check all parts.", "success")

        except Exception as e:
            logger.error(f"Search failed: {e}")
            self._update_search_result(f"Search error: {e}", "error")

    def search_by_material(self, material: str):
        """Search for parts containing specific material (steel, aluminum, etc.)."""
        try:
            column_map = {
                "steel": "steel_ratio",
                "aluminum": "aluminum_ratio",
                "copper": "copper_ratio",
                "wood": "wood_ratio",
            }
            column = column_map.get(material.lower())
            if not column:
                self._update_search_result(f"Unknown material: {material}", "error")
                return

            sql = f"""
                SELECT * FROM parts_master
                WHERE {column} IS NOT NULL AND {column} > 0
                ORDER BY {column} DESC, part_number
            """
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn)
            conn.close()

            self.populate_parts_table(df)
            self._update_search_result(f"Found {len(df)} parts with {material} content", "success")

        except Exception as e:
            logger.error(f"Material search failed: {e}")
            self._update_search_result(f"Search error: {e}", "error")

    def run_parts_search(self):
        """Execute the main parts search based on user input."""
        try:
            field_display = self.search_field_type.currentText()
            field = self.search_field_map.get(field_display, "part_number")
            match_type = self.search_match_type.currentText()
            value = self.search_value_input.text().strip()

            if not value:
                QMessageBox.warning(self, "Search", "Please enter a search value.")
                return

            # Build the WHERE clause based on match type
            if match_type == "Contains":
                where_clause = f"{field} LIKE ?"
                params = (f"%{value}%",)
            elif match_type == "Equals":
                where_clause = f"{field} = ?"
                params = (value,)
            elif match_type == "Starts with":
                where_clause = f"{field} LIKE ?"
                params = (f"{value}%",)
            elif match_type == "Ends with":
                where_clause = f"{field} LIKE ?"
                params = (f"%{value}",)
            elif match_type == "Greater than":
                where_clause = f"CAST({field} AS REAL) > ?"
                params = (value,)
            elif match_type == "Less than":
                where_clause = f"CAST({field} AS REAL) < ?"
                params = (value,)
            else:
                where_clause = f"{field} LIKE ?"
                params = (f"%{value}%",)

            sql = f"SELECT * FROM parts_master WHERE {where_clause} ORDER BY part_number"

            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn, params=params)
            conn.close()

            self.populate_parts_table(df)
            self._update_search_result(
                f"Found {len(df)} parts where {field_display} {match_type.lower()} '{value}'",
                "success" if len(df) > 0 else "warning"
            )
            logger.info(f"Parts search: {field} {match_type} '{value}' - {len(df)} results")

        except Exception as e:
            logger.error(f"Parts search failed: {e}")
            self._update_search_result(f"Search error: {e}", "error")
            QMessageBox.critical(self, "Search Error", f"Failed to search:\n{e}")

    def apply_combined_filters(self):
        """Apply client and country filters together."""
        try:
            client = self.filter_client_combo.currentData()
            country = self.filter_country_combo.currentData()

            # If both are empty, show all
            if not client and not country:
                return  # Don't auto-refresh on initial load

            conditions = []
            params = []

            if client:
                conditions.append("client_code = ?")
                params.append(client)

            if country:
                conditions.append("country_origin = ?")
                params.append(country)

            if conditions:
                where_clause = " AND ".join(conditions)
                sql = f"SELECT * FROM parts_master WHERE {where_clause} ORDER BY part_number"

                conn = sqlite3.connect(str(DB_PATH))
                df = pd.read_sql(sql, conn, params=params)
                conn.close()

                self.populate_parts_table(df)

                filter_desc = []
                if client:
                    filter_desc.append(f"Client: {client}")
                if country:
                    filter_desc.append(f"Country: {country}")

                self._update_search_result(
                    f"Filtered {len(df)} parts ({', '.join(filter_desc)})",
                    "success"
                )

        except Exception as e:
            logger.error(f"Filter failed: {e}")
            self._update_search_result(f"Filter error: {e}", "error")

    def populate_parts_table(self, df):
        """Populate the parts table with a dataframe"""
        self.parts_table.blockSignals(True)
        self.parts_table.setRowCount(len(df))
        # Map table column headers to dataframe columns
        # Headers: part_number, description, hts_code, country_origin, mid, client_code,
        #          steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, non_steel_ratio,
        #          qty_unit, updated_date
        for i, row in df.iterrows():
            # Column 0: part_number - store original in UserRole for rename detection
            part_number = str(row.get('part_number', '')) if pd.notna(row.get('part_number')) else ""
            part_item = QTableWidgetItem(part_number)
            part_item.setData(Qt.UserRole, part_number)  # Store original for rename tracking
            self.parts_table.setItem(i, 0, part_item)
            # Column 1: description
            self.parts_table.setItem(i, 1, QTableWidgetItem(str(row.get('description', '')) if pd.notna(row.get('description')) else ""))
            # Column 2: hts_code
            self.parts_table.setItem(i, 2, QTableWidgetItem(str(row.get('hts_code', '')) if pd.notna(row.get('hts_code')) else ""))
            # Column 3: country_origin
            self.parts_table.setItem(i, 3, QTableWidgetItem(str(row.get('country_origin', '')) if pd.notna(row.get('country_origin')) else ""))
            # Column 4: mid
            self.parts_table.setItem(i, 4, QTableWidgetItem(str(row.get('mid', '')) if pd.notna(row.get('mid')) else ""))
            # Column 5: client_code
            self.parts_table.setItem(i, 5, QTableWidgetItem(str(row.get('client_code', '')) if pd.notna(row.get('client_code')) else ""))
            # Column 6: steel_ratio
            self.parts_table.setItem(i, 6, QTableWidgetItem(str(row.get('steel_ratio', 0.0)) if pd.notna(row.get('steel_ratio')) else "0.0"))
            # Column 7: aluminum_ratio
            self.parts_table.setItem(i, 7, QTableWidgetItem(str(row.get('aluminum_ratio', 0.0)) if pd.notna(row.get('aluminum_ratio')) else "0.0"))
            # Column 8: copper_ratio
            self.parts_table.setItem(i, 8, QTableWidgetItem(str(row.get('copper_ratio', 0.0)) if pd.notna(row.get('copper_ratio')) else "0.0"))
            # Column 9: wood_ratio
            self.parts_table.setItem(i, 9, QTableWidgetItem(str(row.get('wood_ratio', 0.0)) if pd.notna(row.get('wood_ratio')) else "0.0"))
            # Column 10: auto_ratio
            self.parts_table.setItem(i, 10, QTableWidgetItem(str(row.get('auto_ratio', 0.0)) if pd.notna(row.get('auto_ratio')) else "0.0"))
            # Column 11: non_steel_ratio
            self.parts_table.setItem(i, 11, QTableWidgetItem(str(row.get('non_steel_ratio', 0.0)) if pd.notna(row.get('non_steel_ratio')) else "0.0"))
            # Column 12: qty_unit
            self.parts_table.setItem(i, 12, QTableWidgetItem(str(row.get('qty_unit', '')) if pd.notna(row.get('qty_unit')) else ""))
            # Column 13: Sec301_Exclusion_Tariff
            self.parts_table.setItem(i, 13, QTableWidgetItem(str(row.get('Sec301_Exclusion_Tariff', '')) if pd.notna(row.get('Sec301_Exclusion_Tariff')) else ""))
            # Column 14: updated_date
            self.parts_table.setItem(i, 14, QTableWidgetItem(str(row.get('updated_date', '')) if pd.notna(row.get('updated_date')) else ""))
        self.parts_table.blockSignals(False)

    # ...existing code...

    def setup_config_tab(self):
        layout = QVBoxLayout(self.tab_config)
        title = QLabel("<h2>Customs Configuration</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Buttons at top
        top_bar = QHBoxLayout()
        btn_import_excel = QPushButton("Import Section 232 Tariffs (CSV/Excel)")
        btn_import_excel.setStyleSheet(self.get_button_style("success"))
        btn_import_excel.clicked.connect(self.import_tariff_232)
        btn_refresh = QPushButton("Refresh View")
        btn_refresh.setStyleSheet(self.get_button_style("info"))
        btn_refresh.clicked.connect(self.refresh_tariff_view)
        top_bar.addWidget(btn_import_excel)
        top_bar.addWidget(btn_refresh)
        top_bar.addStretch()
        layout.addLayout(top_bar)

        # Section 232 Tariff Viewer
        group1 = QGroupBox("Section 232 Tariff List")
        g1_layout = QVBoxLayout()

        # Search and filter
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel("Filter:"))
        self.tariff_filter = QLineEdit()
        self.tariff_filter.setPlaceholderText("Search HTS code, classification, or chapter...")
        self.tariff_filter.setStyleSheet(self.get_input_style())
        self.tariff_filter.textChanged.connect(self.filter_tariff_table)
        filter_bar.addWidget(self.tariff_filter)

        self.tariff_material_filter = QComboBox()
        self.tariff_material_filter.addItems(["All", "Steel", "Aluminum", "Wood", "Copper", "Auto"])
        self.tariff_material_filter.currentTextChanged.connect(self.filter_tariff_table)
        filter_bar.addWidget(QLabel("Material:"))
        filter_bar.addWidget(self.tariff_material_filter)
        self.bottom_status.setText("Loading Exported Files...")
        QApplication.processEvents()
        # ...existing code...
        self.bottom_status.setText("Ready")
        
        # Add color toggle checkbox
        self.tariff_color_toggle = QCheckBox("Color by Material")
        self.tariff_color_toggle.setChecked(False)  # Disabled by default
        self.tariff_color_toggle.stateChanged.connect(self.filter_tariff_table)
        filter_bar.addWidget(self.tariff_color_toggle)
        
        g1_layout.addLayout(filter_bar)
        
        # Table
        self.tariff_table = QTableWidget()
        self.tariff_table.setColumnCount(7)
        self.tariff_table.setHorizontalHeaderLabels(["HTS Code", "Material", "Classification", 
                       "Chapter", "Chapter Description", 
                       "Declaration", "Notes"])
        self.tariff_table.horizontalHeader().setStretchLastSection(True)
        self.tariff_table.setAlternatingRowColors(True)
        self.tariff_table.setStyleSheet("")
        self.tariff_table.setSortingEnabled(False)  # Disabled for better performance
        g1_layout.addWidget(self.tariff_table)
        
        # Count label
        self.tariff_count_label = QLabel("Total: 0 tariff codes")
        self.tariff_count_label.setStyleSheet("font-weight:bold; padding:5px;")
        g1_layout.addWidget(self.tariff_count_label)
        
        group1.setLayout(g1_layout)
        layout.addWidget(group1)

        self.tab_config.setLayout(layout)
        
        # Load initial data
        self.refresh_tariff_view()

    def import_tariff_232(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Section 232 Tariffs CSV", "", "CSV Files (*.csv);;Excel (*.xlsx)")
        if not path:
            return
        try:
            # Read file based on extension
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_excel(path, header=0)
            
            df = df.fillna("")
            
            # Check if it's the new comprehensive format
            if 'HTS Code' in df.columns and 'Material' in df.columns and 'Classification' in df.columns:
                # New comprehensive CSV format with all columns
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM tariff_232")
                
                imported = 0
                for _, row in df.iterrows():
                    hts_code = str(row['HTS Code']).strip().replace(".", "")[:10]
                    material = str(row['Material']).strip()
                    classification = str(row['Classification']).strip()
                    chapter = str(row['Chapter']).strip()
                    chapter_desc = str(row['Chapter Description']).strip()
                    declaration = str(row['Declaration Required']).strip()
                    notes = str(row['Notes']).strip()
                    
                    if hts_code and material in ['Steel', 'Aluminum', 'Wood', 'Copper', 'Auto']:
                        c.execute("""INSERT OR REPLACE INTO tariff_232
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (hts_code, material, classification, chapter,
                                  chapter_desc, declaration, notes))
                        imported += 1

                conn.commit()
                conn.close()
                QMessageBox.information(self, "Success",
                    f"Imported {imported} Section 232 tariff codes\n\n"
                    f"Format: Comprehensive 7-column CSV")
                logger.success(f"tariff_232 table updated with {imported} codes (comprehensive format)")
                self.status.setText(f"Section 232 list imported: {imported} codes")
            elif 'HTS Code' in df.columns and 'Material' in df.columns:
                # Simple CSV format (HTS Code, Material only)
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM tariff_232")

                imported = 0
                for _, row in df.iterrows():
                    hts_code = str(row['HTS Code']).strip().replace(".", "")[:10]
                    material = str(row['Material']).strip()

                    if hts_code and material in ['Steel', 'Aluminum', 'Wood', 'Copper', 'Auto']:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (hts_code, material, '', '', '', '', ''))
                        imported += 1
                
                conn.commit()
                conn.close()
                QMessageBox.information(self, "Success", f"Imported {imported} Section 232 tariff codes")
                logger.success(f"tariff_232 table updated with {imported} codes")
                self.status.setText(f"Section 232 list imported: {imported} codes")
            else:
                # Legacy Excel format (2 columns: steel, aluminum)
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM tariff_232")
                steel_codes = [str(x).replace(".", "")[:10] for x in df.iloc[1:, 0] if pd.notna(x) and str(x).strip()]
                alum_codes = [str(x).replace(".", "")[:10] for x in df.iloc[1:, 1] if pd.notna(x) and str(x).strip()]
                for code in steel_codes:
                    if code:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (code, 'Steel', '', '', '', '', ''))
                for code in alum_codes:
                    if code:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (code, 'Aluminum', '', '', '', '', ''))
                conn.commit()
                conn.close()
                imported = len(steel_codes) + len(alum_codes)
                QMessageBox.information(self, "Success", 
                    f"Imported {imported} 232 codes\n\n"
                    f"Format: Legacy Excel (2-column)")
                logger.success("tariff_232 table updated (legacy format)")
                self.status.setText("Section 232 list imported")
            
            self.refresh_tariff_view()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Import failed: {e}")
            logger.error(f"Section 232 import error: {e}")


    def refresh_tariff_view(self):
        """Load and display all tariff codes from database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("""SELECT hts_code, material, classification, chapter, 
                                       chapter_description, declaration_required, notes 
                                FROM tariff_232 
                                ORDER BY material, chapter, hts_code""", conn)
            conn.close()
            
            self.tariff_full_data = df  # Store for filtering
            self.filter_tariff_table()
            
        except Exception as e:
            logger.error(f"Refresh tariff view failed: {e}")
            self.tariff_table.setRowCount(0)
            self.tariff_count_label.setText("Error loading tariff data")

    def filter_tariff_table(self):
        """Filter and display tariff table based on search criteria"""
        try:
            if not hasattr(self, 'tariff_full_data') or self.tariff_full_data.empty:
                self.tariff_table.setRowCount(0)
                self.tariff_count_label.setText("No tariff data")
                return
            
            df = self.tariff_full_data.copy()
            
            # Apply text filter - search across HTS code, classification, and chapter description
            search_text = self.tariff_filter.text().strip().lower()
            if search_text:
                mask = (df['hts_code'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['classification'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['chapter_description'].astype(str).str.lower().str.contains(search_text, na=False))
                df = df[mask]
            
            # Apply material filter
            material_filter = self.tariff_material_filter.currentText()
            if material_filter != "All":
                df = df[df['material'] == material_filter]
            
            # Populate table
            self.tariff_table.setSortingEnabled(False)
            self.tariff_table.setRowCount(len(df))
            
            for row_idx, (_, row) in enumerate(df.iterrows()):
                # Create items for all 7 columns
                items = [
                    QTableWidgetItem(str(row['hts_code'])),
                    QTableWidgetItem(str(row['material'])),
                    QTableWidgetItem(str(row['classification'])),
                    QTableWidgetItem(str(row['chapter'])),
                    QTableWidgetItem(str(row['chapter_description'])),
                    QTableWidgetItem(str(row['declaration_required'])),
                    QTableWidgetItem(str(row['notes']))
                ]
                
                # Make all items read-only
                for item in items:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                
                # Color code by material (if toggle is enabled)
                if self.tariff_color_toggle.isChecked():
                    material_colors = {
                        'Steel': QColor('#e3f2fd'),      # Light blue
                        'Aluminum': QColor('#fff3e0'),   # Light orange
                        'Wood': QColor('#f1f8e9'),       # Light green
                        'Copper': QColor('#ffe0b2')      # Light copper/bronze
                    }
                    
                    material = row['material']
                    if material in material_colors:
                        bg_color = material_colors[material]
                        # Apply color to entire row for better visibility
                        for item in items:
                            item.setBackground(bg_color)
                
                # Add items to table
                for col_idx, item in enumerate(items):
                    self.tariff_table.setItem(row_idx, col_idx, item)
            
            self.tariff_table.setSortingEnabled(True)
            self.tariff_count_label.setText(f"Total: {len(df)} tariff codes (filtered from {len(self.tariff_full_data)})")
            
        except Exception as e:
            logger.error(f"Filter tariff table failed: {e}")
            logger.trace(traceback.format_exc())
            self.tariff_table.setRowCount(0)

    def setup_log_tab(self):
        layout = QVBoxLayout(self.tab_log)
        title = QLabel("<h2>Log View</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        log_box = QGroupBox("Real-time Log")
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        self.log_text.setContextMenuPolicy(Qt.CustomContextMenu)
        self.log_text.customContextMenuRequested.connect(self.log_context_menu)
        log_layout.addWidget(self.log_text)

        btn_layout = QHBoxLayout()
        btn_copy = QPushButton("Copy to Clipboard")
        btn_copy.setStyleSheet("background:#0078D7; color:white; font-weight:bold;")
        btn_copy.clicked.connect(self.copy_log_to_clipboard)
        btn_clear = QPushButton("Clear Log")
        btn_clear.clicked.connect(lambda: (self.log_text.clear(), logger.logs.clear()))
        btn_layout.addWidget(btn_copy)
        btn_layout.addWidget(btn_clear)
        btn_layout.addStretch()
        log_layout.addLayout(btn_layout)
        
        log_box.setLayout(log_layout)
        layout.addWidget(log_box)

        self.log_timer = QTimer()
        self.log_timer.timeout.connect(self.update_log)
        self.log_timer.start(500)
        self.tab_log.setLayout(layout)

    def setup_actions_tab(self):
        """Section 232 Actions Reference Tab - Chapter 99 tariff actions"""
        layout = QVBoxLayout(self.tab_actions)
        
        # Title
        title = QLabel("<h2>Section 232 Tariff Actions (Chapter 99)</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Info box
        info_box = QGroupBox("Reference Information")
        info_layout = QVBoxLayout()
        info_text = QLabel(
            "This table contains the Chapter 99 tariff action codes and rates for Section 232 duties. "
            "These are the actual tariff numbers used in ACE for declaring Section 232 merchandise."
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)
        info_box.setLayout(info_layout)
        layout.addWidget(info_box)
        
        # Control buttons
        btn_layout = QHBoxLayout()
        btn_import = QPushButton("Import Actions CSV")
        btn_import.setStyleSheet(self.get_button_style("info"))
        btn_import.clicked.connect(self.import_actions_csv)
        btn_layout.addWidget(btn_import)

        btn_refresh = QPushButton("Refresh View")
        btn_refresh.setStyleSheet(self.get_button_style("default"))
        btn_refresh.clicked.connect(self.refresh_actions_view)
        btn_layout.addWidget(btn_refresh)

        # Edit mode toggle
        self.actions_edit_mode = False
        self.btn_edit_actions = QPushButton("Enable Edit Mode")
        self.btn_edit_actions.setStyleSheet(self.get_button_style("warning"))
        self.btn_edit_actions.clicked.connect(self.toggle_actions_edit_mode)
        btn_layout.addWidget(self.btn_edit_actions)

        # Save/Cancel buttons (hidden by default)
        self.btn_save_actions = QPushButton("Save Changes")
        self.btn_save_actions.setStyleSheet(self.get_button_style("success"))
        self.btn_save_actions.clicked.connect(self.save_actions_changes)
        self.btn_save_actions.setVisible(False)
        btn_layout.addWidget(self.btn_save_actions)

        self.btn_cancel_actions = QPushButton("Cancel")
        self.btn_cancel_actions.setStyleSheet(self.get_button_style("default"))
        self.btn_cancel_actions.clicked.connect(self.cancel_actions_edit)
        self.btn_cancel_actions.setVisible(False)
        btn_layout.addWidget(self.btn_cancel_actions)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Filter bar
        filter_bar = QHBoxLayout()
        self.actions_filter = QLineEdit()
        self.actions_filter.setPlaceholderText("Search tariff number, action, or description...")
        self.actions_filter.setStyleSheet(self.get_input_style())
        self.actions_filter.textChanged.connect(self.filter_actions_table)
        filter_bar.addWidget(self.actions_filter)
        
        self.actions_material_filter = QComboBox()
        self.actions_material_filter.addItems(["All", "Steel", "Aluminum", "Copper", "Wood", "Auto"])
        self.actions_material_filter.currentTextChanged.connect(self.filter_actions_table)
        filter_bar.addWidget(QLabel("Commodity:"))
        filter_bar.addWidget(self.actions_material_filter)
        
        # Add color toggle checkbox
        self.actions_color_toggle = QCheckBox("Color by Material")
        self.actions_color_toggle.setChecked(False)  # Disabled by default
        self.actions_color_toggle.stateChanged.connect(self.filter_actions_table)
        filter_bar.addWidget(self.actions_color_toggle)
        
        layout.addLayout(filter_bar)
        
        # Table
        self.actions_table = QTableWidget()
        self.actions_table.setColumnCount(10)
        self.actions_table.setHorizontalHeaderLabels([
            "Tariff No", "Action", "Description", "Ad Valorem Rate",
            "Effective Date", "Expiration Date", "Specific Rate",
            "Additional Declaration", "Note", "Link"
        ])
        self.actions_table.horizontalHeader().setStretchLastSection(True)
        self.actions_table.setAlternatingRowColors(True)
        self.actions_table.setStyleSheet("")
        self.actions_table.setSortingEnabled(True)
        layout.addWidget(self.actions_table)
        
        # Count label
        self.actions_count_label = QLabel("Total: 0 actions")
        self.actions_count_label.setStyleSheet("font-weight:bold; padding:5px;")
        layout.addWidget(self.actions_count_label)
        
        self.tab_actions.setLayout(layout)

        # Load data
        self.refresh_actions_view()

    def setup_hts_database_tab(self, tab_widget):
        """HTS Database Reference Tab - displays contents of hts.db"""
        layout = QVBoxLayout(tab_widget)

        # Title row with menu button
        title_row = QHBoxLayout()
        title_row.addStretch()

        title = QLabel("<h2>HTS Code Database</h2>")
        title.setAlignment(Qt.AlignCenter)
        title_row.addWidget(title)

        title_row.addStretch()

        # Menu button (hamburger style) in top right
        from PyQt5.QtWidgets import QToolButton, QMenu
        hts_menu_btn = QToolButton()
        hts_menu_btn.setText("☰")  # Unicode hamburger menu
        hts_menu_btn.setToolTip("HTS Database Options")
        hts_menu_btn.setStyleSheet("""
            QToolButton {
                font-size: 16px;
                font-weight: bold;
                padding: 4px 8px;
                border: 1px solid #555;
                border-radius: 4px;
                background: transparent;
            }
            QToolButton:hover {
                background: #3a3a3a;
            }
            QToolButton::menu-indicator {
                image: none;
            }
        """)
        hts_menu_btn.setPopupMode(QToolButton.InstantPopup)

        # Create menu
        hts_menu = QMenu(hts_menu_btn)
        update_action = hts_menu.addAction("Update from USITC (Auto)")
        update_action.setToolTip("Automatically download latest HTS data from the US International Trade Commission")
        update_action.triggered.connect(self.update_hts_from_usitc)

        import_action = hts_menu.addAction("Import from CSV File...")
        import_action.setToolTip("Import HTS data from a manually downloaded CSV file")
        import_action.triggered.connect(self.import_hts_from_csv)

        hts_menu.addSeparator()

        download_link_action = hts_menu.addAction("Open USITC Download Page")
        download_link_action.setToolTip("Open the USITC website to manually download HTS data")
        download_link_action.triggered.connect(lambda: QDesktopServices.openUrl(QUrl("https://hts.usitc.gov/export")))

        hts_menu_btn.setMenu(hts_menu)
        title_row.addWidget(hts_menu_btn)

        layout.addLayout(title_row)

        # Info box
        info_box = QGroupBox("Reference Information")
        info_layout = QVBoxLayout()
        info_text = QLabel(
            "This table contains HTS (Harmonized Tariff Schedule) codes with their descriptions, "
            "units of quantity, and duty rates. Use this reference to look up tariff classifications."
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)
        info_box.setLayout(info_layout)
        layout.addWidget(info_box)

        # Search/Filter bar
        filter_bar = QHBoxLayout()

        self.hts_db_search = QLineEdit()
        self.hts_db_search.setPlaceholderText("Search: multiple words (AND), word1 | word2 (OR), % wildcard...")
        self.hts_db_search.setStyleSheet(self.get_input_style())
        self.hts_db_search.returnPressed.connect(lambda: self.search_hts_database())
        filter_bar.addWidget(self.hts_db_search, 1)

        btn_search = QPushButton("Search")
        btn_search.setStyleSheet(self.get_button_style("info"))
        btn_search.clicked.connect(self.search_hts_database)
        filter_bar.addWidget(btn_search)

        btn_clear = QPushButton("Clear")
        btn_clear.setStyleSheet(self.get_button_style("default"))
        btn_clear.clicked.connect(self.clear_hts_database_search)
        filter_bar.addWidget(btn_clear)

        layout.addLayout(filter_bar)

        # Table
        self.hts_db_table = QTableWidget()
        self.hts_db_table.setColumnCount(7)
        self.hts_db_table.setHorizontalHeaderLabels([
            "HTS Code", "Description", "Unit of Qty", "General Rate",
            "Special Rate", "Column 2 Rate", "Chapter"
        ])
        self.hts_db_table.horizontalHeader().setStretchLastSection(True)
        self.hts_db_table.setAlternatingRowColors(True)
        self.hts_db_table.setSortingEnabled(True)
        self.hts_db_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.hts_db_table.setEditTriggers(QTableWidget.NoEditTriggers)

        # Set column widths
        self.hts_db_table.setColumnWidth(0, 120)  # HTS Code
        self.hts_db_table.setColumnWidth(1, 400)  # Description
        self.hts_db_table.setColumnWidth(2, 80)   # Unit
        self.hts_db_table.setColumnWidth(3, 100)  # General Rate
        self.hts_db_table.setColumnWidth(4, 100)  # Special Rate
        self.hts_db_table.setColumnWidth(5, 100)  # Column 2 Rate
        self.hts_db_table.setColumnWidth(6, 60)   # Chapter

        layout.addWidget(self.hts_db_table)

        # Count label
        self.hts_db_count_label = QLabel("Enter a search term to find HTS codes (showing first 500 results)")
        self.hts_db_count_label.setStyleSheet("font-weight:bold; padding:5px;")
        layout.addWidget(self.hts_db_count_label)

        tab_widget.setLayout(layout)

    def _build_hts_search_conditions(self, term: str) -> tuple:
        """Build SQL WHERE conditions for a single search term.

        Returns (conditions_string, params_list) tuple.
        Searches both HTS code and description.
        """
        term = term.strip()
        if not term:
            return "1=1", []

        # Check if term already has wildcards
        has_wildcard = '%' in term

        # Check if it looks like an HTS code (starts with digit)
        is_code_search = term[0].isdigit()

        if is_code_search:
            # Clean the code (remove periods)
            clean_code = term.replace('.', '')
            if has_wildcard:
                code_pattern = clean_code
            else:
                code_pattern = f"{clean_code}%"
            # Search HTS code
            return "full_code LIKE ?", [code_pattern]
        else:
            # Text search in description
            if has_wildcard:
                search_pattern = term
            else:
                search_pattern = f"%{term}%"
            return "description LIKE ? COLLATE NOCASE", [search_pattern]

    def search_hts_database(self):
        """Search the HTS database and display results.

        Search syntax:
        - Multiple words: AND search (all words must match) - e.g., "aluminum post"
        - OR search: Use | or OR between terms - e.g., "aluminum | steel" or "aluminum OR steel"
        - Wildcard: Use % for wildcards - e.g., "alum%" matches aluminum, aluminium, etc.
        - HTS code: Start with digit to search codes - e.g., "7606" or "7606.11"
        - Mixed: Searches both HTS code and description
        """
        search_term = self.hts_db_search.text().strip()

        hts_db_path = RESOURCES_DIR / "References" / "hts.db"
        logger.info(f"HTS search: term='{search_term}', db_path={hts_db_path}, exists={hts_db_path.exists()}")
        if not hts_db_path.exists():
            QMessageBox.warning(self, "Database Not Found", f"hts.db not found at: {hts_db_path}")
            return

        try:
            conn = sqlite3.connect(str(hts_db_path))
            cursor = conn.cursor()

            if search_term:
                # Check for OR search (using | or OR)
                or_terms = []
                if ' OR ' in search_term.upper() or '|' in search_term:
                    # Split by | or OR (case insensitive)
                    import re
                    or_terms = re.split(r'\s*\|\s*|\s+OR\s+', search_term, flags=re.IGNORECASE)
                    or_terms = [t.strip() for t in or_terms if t.strip()]

                if or_terms:
                    # OR search - any term can match
                    conditions = []
                    params = []
                    for term in or_terms:
                        term_conditions, term_params = self._build_hts_search_conditions(term)
                        conditions.append(f"({term_conditions})")
                        params.extend(term_params)

                    where_clause = " OR ".join(conditions)
                    query = f"""
                        SELECT full_code, description, unit_of_quantity, general_rate,
                               special_rate, column2_rate, chapter
                        FROM hts_codes
                        WHERE {where_clause}
                        ORDER BY full_code
                        LIMIT 500
                    """
                    cursor.execute(query, params)
                else:
                    # AND search - all words must match (or single term)
                    # Split by spaces, &, commas for multiple terms
                    import re
                    terms = re.split(r'[\s,&]+', search_term)
                    terms = [t.strip() for t in terms if t.strip()]

                    if len(terms) == 1:
                        # Single term search
                        conditions, params = self._build_hts_search_conditions(terms[0])
                        query = f"""
                            SELECT full_code, description, unit_of_quantity, general_rate,
                                   special_rate, column2_rate, chapter
                            FROM hts_codes
                            WHERE {conditions}
                            ORDER BY full_code
                            LIMIT 500
                        """
                        cursor.execute(query, params)
                    else:
                        # Multiple terms - all must match (AND)
                        all_conditions = []
                        all_params = []
                        for term in terms:
                            term_conditions, term_params = self._build_hts_search_conditions(term)
                            all_conditions.append(f"({term_conditions})")
                            all_params.extend(term_params)

                        where_clause = " AND ".join(all_conditions)
                        query = f"""
                            SELECT full_code, description, unit_of_quantity, general_rate,
                                   special_rate, column2_rate, chapter
                            FROM hts_codes
                            WHERE {where_clause}
                            ORDER BY full_code
                            LIMIT 500
                        """
                        cursor.execute(query, all_params)
            else:
                # Show first 500 entries if no search term
                cursor.execute("""
                    SELECT full_code, description, unit_of_quantity, general_rate,
                           special_rate, column2_rate, chapter
                    FROM hts_codes
                    ORDER BY full_code
                    LIMIT 500
                """)

            rows = cursor.fetchall()
            conn.close()
            logger.info(f"HTS search returned {len(rows)} rows")

            # Populate table
            self.hts_db_table.setRowCount(len(rows))
            for row_idx, row_data in enumerate(rows):
                for col_idx, value in enumerate(row_data):
                    display_value = str(value) if value else ""
                    # Format HTS code with periods for readability (column 0)
                    if col_idx == 0 and display_value and len(display_value) >= 4:
                        # Format as XXXX.XX.XXXX (e.g., 4009420050 -> 4009.42.0050)
                        code = display_value
                        if len(code) >= 6:
                            display_value = f"{code[:4]}.{code[4:6]}.{code[6:]}"
                        elif len(code) >= 4:
                            display_value = f"{code[:4]}.{code[4:]}"
                    item = QTableWidgetItem(display_value)
                    self.hts_db_table.setItem(row_idx, col_idx, item)

            # Update count label
            if search_term:
                self.hts_db_count_label.setText(f"Found {len(rows)} results for '{search_term}'" +
                                                (" (showing first 500)" if len(rows) == 500 else ""))
            else:
                self.hts_db_count_label.setText(f"Showing first {len(rows)} HTS codes")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to search HTS database: {e}")

    def clear_hts_database_search(self):
        """Clear HTS database search and results"""
        self.hts_db_search.clear()
        self.hts_db_table.setRowCount(0)
        self.hts_db_count_label.setText("Enter a search term to find HTS codes (showing first 500 results)")

    def update_hts_from_usitc(self):
        """Download and update HTS database from USITC website."""
        from PyQt5.QtWidgets import QProgressDialog
        from PyQt5.QtCore import Qt
        import urllib.request
        import urllib.error
        import json as json_module
        import csv
        import io

        # USITC now provides HTS data via export API at https://hts.usitc.gov/export
        # CSV format: https://hts.usitc.gov/export?format=csv&from=0101&to=9999

        reply = QMessageBox.question(
            self,
            "Update HTS Database",
            "This will download the latest HTS data from the US International Trade Commission (USITC) "
            "and replace the current database.\n\n"
            "The download may take a few minutes depending on your connection.\n\n"
            "Do you want to continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        progress = QProgressDialog("Downloading HTS data from USITC...", "Cancel", 0, 100, self)
        progress.setWindowTitle("HTS Database Update")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setMinimumWidth(400)
        progress.setValue(5)
        QApplication.processEvents()

        try:
            hts_db_path = RESOURCES_DIR / "References" / "hts.db"

            # Backup current database
            backup_path = None
            if hts_db_path.exists():
                backup_path = hts_db_path.with_suffix('.db.backup')
                import shutil
                shutil.copy2(hts_db_path, backup_path)
                logger.info(f"Backed up HTS database to {backup_path}")

            progress.setLabelText("Connecting to USITC HTS database...")
            progress.setValue(10)
            QApplication.processEvents()

            if progress.wasCanceled():
                return

            # USITC provides direct CSV download of HTS Basic Edition
            # Format: https://www.usitc.gov/tata/hts/hts_YYYY_basic_edition_csv.csv
            from datetime import datetime
            current_year = datetime.now().year

            # Try multiple URL patterns - basic edition is most reliable
            csv_urls = [
                (f"https://www.usitc.gov/tata/hts/hts_{current_year}_basic_edition_csv.csv", current_year),
                (f"https://www.usitc.gov/sites/default/files/tata/hts/hts_{current_year}_basic_edition_csv.csv", current_year),
                (f"https://www.usitc.gov/tata/hts/hts_{current_year}_csv.csv", current_year),
                (f"https://www.usitc.gov/tata/hts/hts_{current_year - 1}_basic_edition_csv.csv", current_year - 1),
            ]

            csv_data = None
            download_year = current_year

            for test_url, year in csv_urls:
                progress.setLabelText(f"Downloading HTS {year} data...")
                progress.setValue(15)
                QApplication.processEvents()

                if progress.wasCanceled():
                    return

                try:
                    logger.info(f"Trying HTS download URL: {test_url}")
                    req = urllib.request.Request(
                        test_url,
                        headers={
                            'User-Agent': 'TariffMill/1.0 (HTS Database Update)',
                            'Accept': 'text/csv,application/csv,*/*'
                        }
                    )

                    with urllib.request.urlopen(req, timeout=300) as response:  # 5 minute timeout
                        # Read with progress updates
                        total_size = int(response.headers.get('content-length', 0))
                        downloaded = 0
                        chunks = []

                        while True:
                            if progress.wasCanceled():
                                QMessageBox.information(self, "Cancelled", "HTS update was cancelled.")
                                return

                            chunk = response.read(65536)  # 64KB chunks
                            if not chunk:
                                break
                            chunks.append(chunk)
                            downloaded += len(chunk)

                            if total_size > 0:
                                pct = 15 + int((downloaded / total_size) * 45)
                                progress.setValue(pct)
                                mb_downloaded = downloaded / (1024 * 1024)
                                mb_total = total_size / (1024 * 1024)
                                progress.setLabelText(f"Downloading... {mb_downloaded:.1f} / {mb_total:.1f} MB")
                            else:
                                # Unknown size - show bytes downloaded
                                mb_downloaded = downloaded / (1024 * 1024)
                                progress.setLabelText(f"Downloading... {mb_downloaded:.1f} MB")
                            QApplication.processEvents()

                        csv_data = b''.join(chunks).decode('utf-8')
                        download_year = year
                        logger.info(f"Successfully downloaded HTS from: {test_url}")
                        break

                except urllib.error.HTTPError as e:
                    logger.info(f"HTS URL returned HTTP {e.code}: {test_url}")
                    continue
                except urllib.error.URLError as e:
                    logger.info(f"HTS URL failed: {test_url} - {e}")
                    continue
                except Exception as e:
                    logger.info(f"HTS download error: {test_url} - {e}")
                    continue

            if not csv_data:
                QMessageBox.warning(
                    self, "Download Failed",
                    "Could not download HTS data from USITC.\n\n"
                    "Please check your internet connection or try again later.\n\n"
                    "You can also manually download from:\n"
                    "https://hts.usitc.gov/export"
                )
                return

            progress.setLabelText("Processing HTS records...")
            progress.setValue(65)
            QApplication.processEvents()

            # Check if we got HTML instead of CSV (error page or redirect)
            if csv_data.strip().startswith('<!DOCTYPE') or csv_data.strip().startswith('<html'):
                logger.error("USITC returned HTML instead of CSV - may need form submission")
                logger.error(f"First 1000 chars: {csv_data[:1000]}")
                QMessageBox.warning(
                    self, "Download Failed",
                    "USITC returned a web page instead of CSV data.\n\n"
                    "The export API may require manual download.\n\n"
                    "Please visit https://hts.usitc.gov/export manually,\n"
                    "enter 0101 to 9999, and download the CSV file."
                )
                return

            # Process the CSV data
            all_records = []
            csv_reader = csv.DictReader(io.StringIO(csv_data))

            # Log the actual column names for debugging
            fieldnames = csv_reader.fieldnames
            logger.info(f"CSV columns found: {fieldnames}")

            # Map various possible column names to our expected names
            # USITC export format may use different column names
            def get_column_value(row, *possible_names):
                """Get value from row trying multiple possible column names."""
                for name in possible_names:
                    if name in row and row[name]:
                        return str(row[name]).strip()
                # Also try case-insensitive match
                row_lower = {k.lower(): v for k, v in row.items()}
                for name in possible_names:
                    if name.lower() in row_lower and row_lower[name.lower()]:
                        return str(row_lower[name.lower()]).strip()
                return ''

            for row in csv_reader:
                if progress.wasCanceled():
                    return

                # Try various possible column names for HTS code
                # USITC export may use: HTS Number, htsno, HTS, HTS Code, HTSNumber, etc.
                htsno = get_column_value(row,
                    'HTS Number', 'HTSNumber', 'HTS_Number', 'htsno', 'HTS',
                    'HTS Code', 'HTSCode', 'HTS_Code', 'hts_number', 'hts')

                if not htsno:
                    # If no HTS column found, try the first column which might be the code
                    if fieldnames and row.get(fieldnames[0]):
                        first_val = str(row[fieldnames[0]]).strip()
                        # Check if it looks like an HTS code (starts with digits, contains dots)
                        if first_val and (first_val[0].isdigit() or first_val.startswith('0')):
                            htsno = first_val
                    if not htsno:
                        continue

                # Clean the HTS code (remove periods)
                full_code = htsno.replace('.', '').strip()
                if not full_code:
                    continue

                # Extract chapter from code
                try:
                    chapter = int(full_code[:2]) if len(full_code) >= 2 else 0
                except ValueError:
                    chapter = 0

                # Get unit of quantity - try various column names
                unit_str = get_column_value(row,
                    'Unit of Quantity', 'UnitOfQuantity', 'Unit_of_Quantity',
                    'units', 'Units', 'UoQ', 'Unit')

                # Get description - try various column names
                description = get_column_value(row,
                    'Description', 'description', 'Desc', 'desc', 'Article Description')

                # Get duty rates - try various column names
                general_rate = get_column_value(row,
                    'General Rate of Duty', 'GeneralRateOfDuty', 'General_Rate',
                    'general', 'General', 'Gen Rate', 'General Rate')

                special_rate = get_column_value(row,
                    'Special Rate of Duty', 'SpecialRateOfDuty', 'Special_Rate',
                    'special', 'Special', 'Special Rate')

                column2_rate = get_column_value(row,
                    'Column 2 Rate of Duty', 'Column2RateOfDuty', 'Column_2_Rate',
                    'other', 'Other', 'Column 2', 'Col 2 Rate', 'Column2')

                # Get indent level
                indent_str = get_column_value(row, 'Indent', 'indent', 'Indentation', 'Level')
                try:
                    indent_level = int(indent_str) if indent_str else 0
                except ValueError:
                    indent_level = 0

                record = {
                    'heading': full_code[:4] if len(full_code) >= 4 else full_code,
                    'subheading': full_code[:6] if len(full_code) >= 6 else full_code,
                    'stat_suffix': full_code[6:] if len(full_code) > 6 else '',
                    'full_code': full_code,
                    'description': description,
                    'unit_of_quantity': unit_str,
                    'general_rate': general_rate,
                    'special_rate': special_rate,
                    'column2_rate': column2_rate,
                    'chapter': chapter,
                    'indent_level': indent_level,
                }
                all_records.append(record)

            if not all_records:
                # Log more details for debugging
                logger.error(f"No records found. CSV columns: {fieldnames}")
                # Show first 500 chars of CSV data for debugging
                logger.error(f"First 500 chars of CSV: {csv_data[:500] if csv_data else 'EMPTY'}")

                columns_msg = f"\n\nColumns found: {', '.join(fieldnames) if fieldnames else 'None'}"
                QMessageBox.warning(
                    self, "Processing Failed",
                    f"Downloaded file contained no valid HTS records.\n\n"
                    f"The file format may have changed.{columns_msg}"
                )
                return

            progress.setLabelText(f"Updating database with {len(all_records):,} records...")
            progress.setValue(75)
            QApplication.processEvents()

            # Update the database
            conn = sqlite3.connect(str(hts_db_path))
            cursor = conn.cursor()

            # Clear existing data
            cursor.execute("DELETE FROM hts_codes")

            # Insert new records in batches for better performance
            batch_size = 1000
            for i in range(0, len(all_records), batch_size):
                if progress.wasCanceled():
                    conn.rollback()
                    conn.close()
                    QMessageBox.information(self, "Cancelled", "HTS update was cancelled. Database unchanged.")
                    return

                batch = all_records[i:i + batch_size]
                for record in batch:
                    cursor.execute("""
                        INSERT OR REPLACE INTO hts_codes
                        (heading, subheading, stat_suffix, full_code, description,
                         unit_of_quantity, general_rate, special_rate, column2_rate,
                         chapter, indent_level, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """, (
                        record['heading'],
                        record['subheading'],
                        record['stat_suffix'],
                        record['full_code'],
                        record['description'],
                        record['unit_of_quantity'],
                        record['general_rate'],
                        record['special_rate'],
                        record['column2_rate'],
                        record['chapter'],
                        record['indent_level'],
                    ))

                pct = 75 + int((i / len(all_records)) * 20)
                progress.setValue(pct)
                progress.setLabelText(f"Inserting records... {min(i + batch_size, len(all_records)):,} / {len(all_records):,}")
                QApplication.processEvents()

            conn.commit()

            # Get final count
            cursor.execute("SELECT COUNT(*) FROM hts_codes")
            final_count = cursor.fetchone()[0]
            conn.close()

            progress.setValue(100)
            progress.close()

            backup_msg = f"\n\nA backup of the previous database was saved to:\n{backup_path}" if backup_path else ""
            QMessageBox.information(
                self, "Update Complete",
                f"HTS database updated successfully!\n\n"
                f"Year: {download_year} HTS Basic Edition\n"
                f"Total records: {final_count:,}{backup_msg}"
            )

            # Refresh the search results if any
            self.clear_hts_database_search()

        except Exception as e:
            progress.close()
            logger.error(f"HTS update failed: {e}")
            QMessageBox.critical(
                self, "Update Failed",
                f"Failed to update HTS database:\n\n{str(e)}\n\n"
                "If a backup exists, it has not been modified."
            )

    def import_hts_from_csv(self):
        """Import HTS data from a manually downloaded CSV file."""
        from PyQt5.QtWidgets import QProgressDialog, QFileDialog
        from PyQt5.QtCore import Qt
        import csv
        import io

        # Show file dialog
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select HTS CSV File",
            "",
            "CSV Files (*.csv);;All Files (*.*)"
        )

        if not file_path:
            return

        # Confirm import
        reply = QMessageBox.question(
            self,
            "Import HTS Data",
            f"This will import HTS data from:\n{file_path}\n\n"
            "This will replace the current HTS database.\n\n"
            "Do you want to continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        progress = QProgressDialog("Importing HTS data...", "Cancel", 0, 100, self)
        progress.setWindowTitle("HTS Import")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setMinimumWidth(400)
        progress.setValue(5)
        QApplication.processEvents()

        try:
            hts_db_path = RESOURCES_DIR / "References" / "hts.db"

            # Backup current database
            backup_path = None
            if hts_db_path.exists():
                backup_path = hts_db_path.with_suffix('.db.backup')
                import shutil
                shutil.copy2(hts_db_path, backup_path)
                logger.info(f"Backed up HTS database to {backup_path}")

            progress.setLabelText("Reading CSV file...")
            progress.setValue(10)
            QApplication.processEvents()

            # Read the CSV file
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                csv_data = f.read()

            if not csv_data.strip():
                QMessageBox.warning(self, "Import Failed", "The selected file is empty.")
                return

            # Check if we got HTML instead of CSV
            if csv_data.strip().startswith('<!DOCTYPE') or csv_data.strip().startswith('<html'):
                QMessageBox.warning(
                    self, "Import Failed",
                    "The selected file appears to be HTML, not CSV.\n\n"
                    "Please download the CSV format from the USITC website."
                )
                return

            progress.setLabelText("Processing HTS records...")
            progress.setValue(20)
            QApplication.processEvents()

            # Process the CSV data
            all_records = []
            csv_reader = csv.DictReader(io.StringIO(csv_data))

            # Log the actual column names for debugging
            fieldnames = csv_reader.fieldnames
            logger.info(f"Import CSV columns found: {fieldnames}")

            # Map various possible column names to our expected names
            def get_column_value(row, *possible_names):
                """Get value from row trying multiple possible column names."""
                for name in possible_names:
                    if name in row and row[name]:
                        return str(row[name]).strip()
                # Also try case-insensitive match
                row_lower = {k.lower(): v for k, v in row.items()}
                for name in possible_names:
                    if name.lower() in row_lower and row_lower[name.lower()]:
                        return str(row_lower[name.lower()]).strip()
                return ''

            row_count = 0
            for row in csv_reader:
                if progress.wasCanceled():
                    return

                row_count += 1
                if row_count % 5000 == 0:
                    progress.setLabelText(f"Processing record {row_count:,}...")
                    progress.setValue(20 + min(40, int(row_count / 1000)))
                    QApplication.processEvents()

                # Try various possible column names for HTS code
                htsno = get_column_value(row,
                    'HTS Number', 'HTSNumber', 'HTS_Number', 'htsno', 'HTS',
                    'HTS Code', 'HTSCode', 'HTS_Code', 'hts_number', 'hts')

                if not htsno:
                    # If no HTS column found, try the first column
                    if fieldnames and row.get(fieldnames[0]):
                        first_val = str(row[fieldnames[0]]).strip()
                        if first_val and (first_val[0].isdigit() or first_val.startswith('0')):
                            htsno = first_val
                    if not htsno:
                        continue

                # Clean the HTS code (remove periods)
                full_code = htsno.replace('.', '').strip()
                if not full_code:
                    continue

                # Extract chapter from code
                try:
                    chapter = int(full_code[:2]) if len(full_code) >= 2 else 0
                except ValueError:
                    chapter = 0

                # Get other fields
                unit_str = get_column_value(row,
                    'Unit of Quantity', 'UnitOfQuantity', 'Unit_of_Quantity',
                    'units', 'Units', 'UoQ', 'Unit')

                description = get_column_value(row,
                    'Description', 'description', 'Desc', 'desc', 'Article Description')

                general_rate = get_column_value(row,
                    'General Rate of Duty', 'GeneralRateOfDuty', 'General_Rate',
                    'general', 'General', 'Gen Rate', 'General Rate')

                special_rate = get_column_value(row,
                    'Special Rate of Duty', 'SpecialRateOfDuty', 'Special_Rate',
                    'special', 'Special', 'Special Rate')

                column2_rate = get_column_value(row,
                    'Column 2 Rate of Duty', 'Column2RateOfDuty', 'Column_2_Rate',
                    'other', 'Other', 'Column 2', 'Col 2 Rate', 'Column2')

                indent_str = get_column_value(row, 'Indent', 'indent', 'Indentation', 'Level')
                try:
                    indent_level = int(indent_str) if indent_str else 0
                except ValueError:
                    indent_level = 0

                record = {
                    'heading': full_code[:4] if len(full_code) >= 4 else full_code,
                    'subheading': full_code[:6] if len(full_code) >= 6 else full_code,
                    'stat_suffix': full_code[6:] if len(full_code) > 6 else '',
                    'full_code': full_code,
                    'description': description,
                    'unit_of_quantity': unit_str,
                    'general_rate': general_rate,
                    'special_rate': special_rate,
                    'column2_rate': column2_rate,
                    'chapter': chapter,
                    'indent_level': indent_level,
                }
                all_records.append(record)

            if not all_records:
                logger.error(f"No records found. CSV columns: {fieldnames}")
                columns_msg = f"\n\nColumns found: {', '.join(fieldnames) if fieldnames else 'None'}"
                QMessageBox.warning(
                    self, "Import Failed",
                    f"No valid HTS records found in the CSV file.\n\n"
                    f"Please ensure the file contains HTS data with recognizable column names.{columns_msg}"
                )
                return

            progress.setLabelText(f"Importing {len(all_records):,} records into database...")
            progress.setValue(65)
            QApplication.processEvents()

            # Update the database
            conn = sqlite3.connect(str(hts_db_path))
            cursor = conn.cursor()

            # Clear existing data
            cursor.execute("DELETE FROM hts_codes")

            # Insert new records in batches
            batch_size = 1000
            for i in range(0, len(all_records), batch_size):
                if progress.wasCanceled():
                    conn.rollback()
                    conn.close()
                    QMessageBox.information(self, "Cancelled", "HTS import was cancelled. Database unchanged.")
                    return

                batch = all_records[i:i + batch_size]
                for record in batch:
                    cursor.execute("""
                        INSERT OR REPLACE INTO hts_codes
                        (heading, subheading, stat_suffix, full_code, description,
                         unit_of_quantity, general_rate, special_rate, column2_rate,
                         chapter, indent_level, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """, (
                        record['heading'],
                        record['subheading'],
                        record['stat_suffix'],
                        record['full_code'],
                        record['description'],
                        record['unit_of_quantity'],
                        record['general_rate'],
                        record['special_rate'],
                        record['column2_rate'],
                        record['chapter'],
                        record['indent_level'],
                    ))

                pct = 65 + int((i / len(all_records)) * 30)
                progress.setValue(pct)
                progress.setLabelText(f"Importing records... {min(i + batch_size, len(all_records)):,} / {len(all_records):,}")
                QApplication.processEvents()

            conn.commit()

            # Get final count
            cursor.execute("SELECT COUNT(*) FROM hts_codes")
            final_count = cursor.fetchone()[0]
            conn.close()

            progress.setValue(100)
            progress.close()

            backup_msg = f"\n\nA backup of the previous database was saved to:\n{backup_path}" if backup_path else ""
            QMessageBox.information(
                self, "Import Complete",
                f"HTS database imported successfully!\n\n"
                f"Source: {file_path}\n"
                f"Total records: {final_count:,}{backup_msg}"
            )

            # Refresh the search results
            self.clear_hts_database_search()

        except Exception as e:
            progress.close()
            logger.error(f"HTS import failed: {e}")
            QMessageBox.critical(
                self, "Import Failed",
                f"Failed to import HTS database:\n\n{str(e)}\n\n"
                "If a backup exists, it has not been modified."
            )

    def import_actions_csv(self):
        """Import Section 232 Actions from CSV/TSV"""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Section 232 Actions File", "",
            "CSV/TSV Files (*.csv *.txt);;All Files (*.*)"
        )
        if not path:
            return
        
        try:
            # Try to detect the delimiter and read the file
            if path.lower().endswith('.txt'):
                # For .txt files, assume tab-delimited
                df = pd.read_csv(path, sep='\t', dtype=str, keep_default_na=False)
            else:
                # For .csv files, let pandas auto-detect
                df = pd.read_csv(path, dtype=str, keep_default_na=False)
            
            df = df.fillna("")
            
            # Normalize column names (strip whitespace)
            df.columns = df.columns.str.strip()
            
            # Check for required columns
            required_cols = ['Tariff No', 'Action']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                QMessageBox.warning(self, "Invalid Format", 
                    f"CSV is missing required columns: {', '.join(missing_cols)}\n\n"
                    f"Found columns: {', '.join(df.columns.tolist())}")
                return
            
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM sec_232_actions")
            
            imported = 0
            for _, row in df.iterrows():
                tariff_no = str(row.get('Tariff No', '')).strip()
                action = str(row.get('Action', '')).strip()
                description = str(row.get('Tariff Description', '')).strip()
                advalorem = str(row.get('Advalorem Rate', '')).strip()
                effective = str(row.get('Effective Date', '')).strip()
                expiration = str(row.get('Expiration Date', '')).strip()
                specific = str(row.get('Specific Rate', '')).strip()
                declaration = str(row.get('Additional Declaration Required', '')).strip()
                note = str(row.get('Note', '')).strip()
                link = str(row.get('Link', '')).strip()
                
                if tariff_no and action:
                    c.execute("""INSERT OR REPLACE INTO sec_232_actions 
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                             (tariff_no, action, description, advalorem, effective, 
                              expiration, specific, declaration, note, link))
                    imported += 1
            
            conn.commit()
            conn.close()
            
            QMessageBox.information(self, "Success", 
                f"Imported {imported} Section 232 action records")
            logger.success(f"sec_232_actions table updated with {imported} records")
            self.status.setText(f"Section 232 actions imported: {imported} records")
            self.refresh_actions_view()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Import failed: {e}")
            logger.error(f"Section 232 actions import error: {e}")

    def refresh_actions_view(self):
        """Load and display all Section 232 actions from database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("""SELECT tariff_no, action, description, advalorem_rate,
                                       effective_date, expiration_date, specific_rate,
                                       additional_declaration, note, link
                                FROM sec_232_actions
                                ORDER BY tariff_no""", conn)
            conn.close()
            
            self.actions_full_data = df
            self.filter_actions_table()
            
        except Exception as e:
            logger.error(f"Refresh actions view failed: {e}")
            self.actions_table.setRowCount(0)
            self.actions_count_label.setText("Error loading actions data")

    def filter_actions_table(self):
        """Filter and display actions table based on search criteria"""
        try:
            if not hasattr(self, 'actions_full_data') or self.actions_full_data.empty:
                self.actions_table.setRowCount(0)
                self.actions_count_label.setText("No actions data")
                return
            
            df = self.actions_full_data.copy()
            
            # Apply text filter
            search_text = self.actions_filter.text().strip().lower()
            if search_text:
                mask = (df['tariff_no'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['action'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['description'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['note'].astype(str).str.lower().str.contains(search_text, na=False))
                df = df[mask]
            
            # Apply material filter
            material_filter = self.actions_material_filter.currentText()
            if material_filter != "All":
                # Filter by action column containing material name
                df = df[df['action'].str.contains(material_filter, case=False, na=False)]
            
            # Populate table
            self.actions_table.setSortingEnabled(False)
            self.actions_table.setRowCount(len(df))
            
            for row_idx, (_, row) in enumerate(df.iterrows()):
                items = [
                    QTableWidgetItem(str(row['tariff_no'])),
                    QTableWidgetItem(str(row['action'])),
                    QTableWidgetItem(str(row['description'])),
                    QTableWidgetItem(str(row['advalorem_rate'])),
                    QTableWidgetItem(str(row['effective_date'])),
                    QTableWidgetItem(str(row['expiration_date'])),
                    QTableWidgetItem(str(row['specific_rate'])),
                    QTableWidgetItem(str(row['additional_declaration'])),
                    QTableWidgetItem(str(row['note'])),
                    QTableWidgetItem(str(row['link']))
                ]
                
                # Make items editable/read-only based on edit mode
                # Columns that can be edited: Action (1), Description (2), Note (8), Link (9)
                editable_columns = {1, 2, 8, 9}
                for col_idx, item in enumerate(items):
                    if col_idx in editable_columns and self.actions_edit_mode:
                        # Enable editing
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                    else:
                        # Read-only
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                # Apply color coding if toggle is enabled
                if self.actions_color_toggle.isChecked():
                    material_colors = {
                        'Steel': QColor('#e3f2fd'),
                        'Aluminum': QColor('#fff3e0'),
                        'Wood': QColor('#f1f8e9'),
                        'Copper': QColor('#ffe0b2')
                    }
                    
                    action_text = str(row['action']).upper()
                    bg_color = None
                    
                    if 'STEEL' in action_text:
                        bg_color = material_colors['Steel']
                    elif 'ALUMINUM' in action_text:
                        bg_color = material_colors['Aluminum']
                    elif 'COPPER' in action_text:
                        bg_color = material_colors['Copper']
                    elif 'WOOD' in action_text or 'LUMBER' in action_text or 'FURNITURE' in action_text:
                        bg_color = material_colors['Wood']
                    
                    # Apply background color to entire row
                    if bg_color:
                        for item in items:
                            item.setBackground(bg_color)
                
                # Highlight expired actions regardless of toggle state
                if 'EXPIRED' in str(row['note']).upper():
                    for item in items:
                        item.setForeground(QColor('#999999'))  # Gray out expired
                
                # Add items to table
                for col_idx, item in enumerate(items):
                    self.actions_table.setItem(row_idx, col_idx, item)
            
            self.actions_table.setSortingEnabled(True)
            self.actions_count_label.setText(
                f"Total: {len(df)} actions (filtered from {len(self.actions_full_data)})"
            )
            
        except Exception as e:
            logger.error(f"Filter actions table failed: {e}")
            self.actions_table.setRowCount(0)

    def toggle_actions_edit_mode(self):
        """Toggle edit mode for Section 232 Actions table"""
        self.actions_edit_mode = not self.actions_edit_mode

        if self.actions_edit_mode:
            # Entering edit mode
            self.btn_edit_actions.setText("Disable Edit Mode")
            self.btn_edit_actions.setStyleSheet(self.get_button_style("danger"))
            self.btn_save_actions.setVisible(True)
            self.btn_cancel_actions.setVisible(True)
            self.actions_filter.setEnabled(False)
            self.actions_material_filter.setEnabled(False)
            self.actions_color_toggle.setEnabled(False)

            # Store original data for cancel functionality
            if hasattr(self, 'actions_full_data'):
                self.actions_original_data = self.actions_full_data.copy()

            # Re-render table with editable cells
            self.filter_actions_table()
        else:
            # Exiting edit mode (cancel)
            self.cancel_actions_edit()

    def save_actions_changes(self):
        """Save changes made to Section 232 Actions table to database"""
        if not hasattr(self, 'actions_full_data'):
            QMessageBox.warning(self, "No Data", "No actions data to save")
            return

        try:
            # Collect all current table data
            updated_rows = []
            for row_idx in range(self.actions_table.rowCount()):
                row_data = []
                for col_idx in range(self.actions_table.columnCount()):
                    item = self.actions_table.item(row_idx, col_idx)
                    row_data.append(item.text() if item else "")
                updated_rows.append(row_data)

            # Create DataFrame from updated rows
            columns = ['tariff_no', 'action', 'description', 'advalorem_rate',
                      'effective_date', 'expiration_date', 'specific_rate',
                      'additional_declaration', 'note', 'link']
            df_updated = pd.DataFrame(updated_rows, columns=columns)

            # Update database
            conn = sqlite3.connect(str(DB_PATH))
            df_updated.to_sql('sec_232_actions', conn, if_exists='replace', index=False)
            conn.close()

            # Update internal data
            self.actions_full_data = df_updated

            # Exit edit mode
            self.actions_edit_mode = False
            self.btn_edit_actions.setText("Enable Edit Mode")
            self.btn_edit_actions.setStyleSheet(self.get_button_style("warning"))
            self.btn_save_actions.setVisible(False)
            self.btn_cancel_actions.setVisible(False)
            self.actions_filter.setEnabled(True)
            self.actions_material_filter.setEnabled(True)
            self.actions_color_toggle.setEnabled(True)

            # Refresh display
            self.filter_actions_table()

            QMessageBox.information(self, "Success", f"Saved {len(df_updated)} actions to database")
            logger.success(f"Saved {len(df_updated)} Section 232 actions to database")

        except Exception as e:
            logger.error(f"Save actions failed: {e}")
            QMessageBox.critical(self, "Save Failed", f"Error saving changes:\n{str(e)}")

    def cancel_actions_edit(self):
        """Cancel edit mode and discard changes"""
        self.actions_edit_mode = False
        self.btn_edit_actions.setText("Enable Edit Mode")
        self.btn_edit_actions.setStyleSheet(self.get_button_style("warning"))
        self.btn_save_actions.setVisible(False)
        self.btn_cancel_actions.setVisible(False)
        self.actions_filter.setEnabled(True)
        self.actions_material_filter.setEnabled(True)
        self.actions_color_toggle.setEnabled(True)

        # Restore original data if available
        if hasattr(self, 'actions_original_data'):
            self.actions_full_data = self.actions_original_data.copy()
            del self.actions_original_data

        # Refresh display
        self.filter_actions_table()

    def on_preview_value_edited(self, item):
        # Only update for Value column edits
        if item.column() != 1:
            return
        text = item.text().replace('$', '').replace(',', '').strip()
        try:
            new_val = float(text)
            if new_val < 0:
                raise ValueError()
            item.setData(Qt.UserRole, new_val)
            item.setText(f"{new_val:,.2f}")
        except Exception:
            old_val = item.data(Qt.UserRole) or 0.0
            item.setText(f"{old_val:,.2f}")
        self.recalculate_total_and_check_match()

    def add_preview_row(self):
        """Add a new empty row to the preview table"""
        if self.last_processed_df is None:
            QMessageBox.warning(self, "No Data", "Please process a shipment first before adding rows.")
            return
        
        # Disconnect signals while adding row
        self.table.blockSignals(True)
        
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Create default items for the new row
        default_mid = self.selected_mid or ""
        default_melt = str(default_mid)[:2] if default_mid else ""
        
        value_item = QTableWidgetItem("0.00")
        value_item.setData(Qt.UserRole, 0.0)
        value_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
        
        items = [
            QTableWidgetItem("NEW_PART"),  # 0: Product No
            value_item,                     # 1: Value
            QTableWidgetItem(""),           # 2: HTS
            QTableWidgetItem(default_mid),  # 3: MID
            QTableWidgetItem("0.00"),       # 4: Qty1
            QTableWidgetItem("0.00"),       # 5: Qty2
            QTableWidgetItem("NO"),         # 6: Qty Unit
            QTableWidgetItem("CO"),         # 7: Dec
            QTableWidgetItem(default_melt), # 8: Melt
            QTableWidgetItem(""),           # 9: Cast
            QTableWidgetItem(""),           # 10: Smelt
            QTableWidgetItem(""),           # 11: Flag
            QTableWidgetItem("100.0%"),     # 12: Steel%
            QTableWidgetItem(""),           # 13: Al%
            QTableWidgetItem(""),           # 14: Cu%
            QTableWidgetItem(""),           # 15: Wood%
            QTableWidgetItem(""),           # 16: Auto%
            QTableWidgetItem(""),           # 17: Non-232%
            QTableWidgetItem(""),           # 18: 232 Status
            QTableWidgetItem("")            # 19: Cust Ref
            # TODO: Lacey column - To be implemented at a later date
            # QTableWidgetItem("")            # 20: Lacey
        ]

        # Make all items editable except Qty1, Qty2, Steel%, Al%, Cu%, Wood%, Auto%, Non-232%, 232 Status
        for i, item in enumerate(items):
            if i not in [4, 5, 12, 13, 14, 15, 16, 17, 18]:
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self.table.setItem(row, i, item)
        
        self.table.blockSignals(False)
        self.recalculate_total_and_check_match()
        logger.info(f"Added new row at position {row + 1}")

    def delete_preview_row(self):
        """Delete selected row(s) from the preview table"""
        if self.last_processed_df is None:
            QMessageBox.warning(self, "No Data", "No preview data to delete.")
            return
        
        selected_rows = sorted(set(index.row() for index in self.table.selectedIndexes()), reverse=True)
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select row(s) to delete.")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete {len(selected_rows)} row(s)?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # Disconnect signals while deleting
        self.table.blockSignals(True)
        
        for row in selected_rows:
            self.table.removeRow(row)
            logger.info(f"Deleted row {row + 1}")
        
        self.table.blockSignals(False)
        self.recalculate_total_and_check_match()
        self.status.setText(f"Deleted {len(selected_rows)} row(s)")

    def copy_column_to_clipboard(self):
        """Copy selected column data to clipboard"""
        if self.last_processed_df is None:
            QMessageBox.warning(self, "No Data", "No preview data available.")
            return
        
        # Get selected cells
        selected = self.table.selectedIndexes()
        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select cells from a column to copy.")
            return
        
        # Determine if user selected a single column or multiple cells
        columns = set(index.column() for index in selected)
        
        if len(columns) == 1:
            # Single column selected - copy all values from that column
            col = list(columns)[0]
            column_data = []
            for row in range(self.table.rowCount()):
                item = self.table.item(row, col)
                if item:
                    # For Value column, use the stored float value
                    if col == 1:
                        value = item.data(Qt.UserRole)
                        column_data.append(str(value) if value is not None else "")
                    else:
                        column_data.append(item.text())
                else:
                    column_data.append("")
            
            # Copy to clipboard
            clipboard_text = "\n".join(column_data)
            QApplication.clipboard().setText(clipboard_text)
            
            # Get column name
            header = self.table.horizontalHeaderItem(col)
            col_name = header.text() if header else f"Column {col + 1}"
            QMessageBox.information(self, "Copied", f"Copied {len(column_data)} values from '{col_name}' to clipboard.")
            logger.info(f"Copied column '{col_name}' to clipboard ({len(column_data)} rows)")
        else:
            # Multiple columns or mixed selection - copy selected cells as tab-separated
            # Group by row
            by_row = {}
            for index in selected:
                row = index.row()
                col = index.column()
                if row not in by_row:
                    by_row[row] = {}
                by_row[row][col] = index
            
            # Build clipboard text
            rows_text = []
            for row in sorted(by_row.keys()):
                cells = []
                for col in sorted(by_row[row].keys()):
                    item = self.table.item(row, col)
                    if item:
                        if col == 1:  # Value column
                            value = item.data(Qt.UserRole)
                            cells.append(str(value) if value is not None else "")
                        else:
                            cells.append(item.text())
                    else:
                        cells.append("")
                rows_text.append("\t".join(cells))
            
            clipboard_text = "\n".join(rows_text)
            QApplication.clipboard().setText(clipboard_text)
            QMessageBox.information(self, "Copied", f"Copied {len(selected)} cells to clipboard (tab-separated).")
            logger.info(f"Copied {len(selected)} selected cells to clipboard")

    def select_column(self, column_index):
        """Select entire column when header is clicked"""
        self.table.clearSelection()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, column_index)
            if item:
                item.setSelected(True)

    def on_table_selection_changed(self):
        """Handle selection changes to make selected text readable with white foreground."""
        try:
            # Get current selection
            current_selection = set()
            for item in self.table.selectedItems():
                current_selection.add((item.row(), item.column()))

            # Restore original colors for previously selected items that are no longer selected
            for row, col in self._previous_selection - current_selection:
                item = self.table.item(row, col)
                if item:
                    # Get the original color from the 232 Status column for this row
                    status_item = self.table.item(row, 18)
                    status_text = status_item.text() if status_item else ''
                    original_color = self.get_preview_row_color(status_text)
                    item.setForeground(original_color)

            # Set white foreground for newly selected items
            for row, col in current_selection - self._previous_selection:
                item = self.table.item(row, col)
                if item:
                    # Store original color if not already stored
                    if item.data(Qt.UserRole + 2) is None:
                        item.setData(Qt.UserRole + 2, item.foreground().color().name())
                    item.setForeground(QColor(255, 255, 255))  # White text for selected

            self._previous_selection = current_selection
        except Exception as e:
            logger.debug(f"Selection change handling: {e}")

    def save_column_widths(self):
        """Save column widths to per-user settings for persistence"""
        try:
            widths = {}
            for col in range(self.table.columnCount()):
                header_text = self.table.horizontalHeaderItem(col).text()
                widths[header_text] = self.table.columnWidth(col)

            import json
            set_user_setting('column_widths', json.dumps(widths))
        except Exception as e:
            logger.debug(f"Could not save column widths: {e}")

    def load_column_widths(self):
        """Load saved column widths from per-user settings"""
        try:
            import json
            widths_json = get_user_setting('column_widths')

            if widths_json:
                widths = json.loads(widths_json)
                # Check if any width is 0 - if so, this is corrupted data, clear it
                has_zero_width = any(w == 0 for w in widths.values())
                if has_zero_width:
                    set_user_setting('column_widths', '')
                    logger.info("Cleared corrupted column widths (had 0-width columns)")
                else:
                    for col in range(self.table.columnCount()):
                        header_text = self.table.horizontalHeaderItem(col).text()
                        if header_text in widths and widths[header_text] > 20:  # Minimum 20px width
                            self.table.setColumnWidth(col, widths[header_text])
        except Exception as e:
            logger.debug(f"Could not load column widths: {e}")

    def recalculate_total_and_check_match(self):
        if self.last_processed_df is None:
            return
        total = 0.0
        for i in range(self.table.rowCount()):
            cell = self.table.item(i, 1)
            total += (cell.data(Qt.UserRole) or 0.0) if cell else 0.0

        # Don't update CI input - let user keep their target value
        # Just compare the preview total against the CI input
        ci_text = self.ci_input.text().replace(',', '').strip()
        try:
            target_value = float(ci_text) if ci_text else self.csv_total_value
        except:
            target_value = self.csv_total_value

        diff = abs(total - target_value)
        threshold = 0.01
        if diff <= threshold:
            self.process_btn.setEnabled(True)
            self.process_btn.setText("Export Worksheet")
            self.process_btn.setFocus()  # Keep focus on button so user can press Enter to export
            self.bottom_status.setText(f"VALUES MATCH - Preview: ${total:,.2f} = Target: ${target_value:,.2f}")
            self.bottom_status.setStyleSheet("background:#107C10; color:white; font-weight:bold; font-size:9px; padding:2px 3px;")
        else:
            self.process_btn.setEnabled(False)
            self.process_btn.setText("Export Worksheet (Values Don't Match)")
            diff_display = total - target_value
            sign = "+" if diff_display > 0 else ""
            self.bottom_status.setText(f"Preview: ${total:,.2f} • Target: ${target_value:,.2f} • Diff: {sign}${diff_display:,.2f}")
            self.bottom_status.setStyleSheet("background:#ff9800; color:white; font-weight:bold; font-size:9px; padding:2px 3px;")

    def _process_or_export(self):
        # If no preview yet, run processing; otherwise proceed to export
        if self.last_processed_df is None:
            self.start_processing()
        else:
            self.final_export()

    def _capture_part_number_modifications(self):
        """
        Capture any part number modifications from the preview table.

        Returns a dictionary mapping row index to the modified part number.
        This allows users to edit a part number in the Result Preview and have
        the reprocess use the edited value to look up in parts_master.
        """
        overrides = {}
        if self.table.rowCount() == 0 or self.last_processed_df is None:
            return overrides

        try:
            # Compare table part numbers to original DataFrame part numbers
            for row in range(self.table.rowCount()):
                table_item = self.table.item(row, 0)  # Column 0 = Product No
                if not table_item:
                    continue

                table_part = table_item.text().strip().upper()

                # Get the original part number from the DataFrame
                if row < len(self.last_processed_df):
                    original_part = str(self.last_processed_df.iloc[row].get('Product No', '')).strip().upper()

                    # If the part number was modified, record the override
                    if table_part and table_part != original_part:
                        overrides[row] = table_part
                        logger.info(f"Part number override detected: row {row}: '{original_part}' -> '{table_part}'")
        except Exception as e:
            logger.warning(f"Error capturing part number modifications: {e}")

        return overrides

    def reprocess_invoice(self):
        """Re-process the current invoice to pick up database changes (e.g., deleted/updated parts)."""
        if not self.current_csv:
            QMessageBox.warning(self, "No File", "No invoice file is currently loaded.")
            return

        # Capture modified part numbers from the preview table BEFORE clearing
        # This allows users to edit a part number in the table and have it re-looked up
        part_number_overrides = self._capture_part_number_modifications()
        if part_number_overrides:
            logger.info(f"Captured {len(part_number_overrides)} part number modification(s) for reprocessing")

        # First, save any "Not Found" parts that have been edited in the preview table
        # This saves HTS codes, MIDs, etc. that the user has entered for new parts
        added_count = self.save_preview_parts_to_db()
        if added_count > 0:
            self.status.setText(f"Saved {added_count} new part(s) to database, reprocessing...")
            logger.info(f"Saved {added_count} new parts to database before reprocessing")

        # Clear the cached processed data so start_processing will run fresh
        self.last_processed_df = None
        self.table.setRowCount(0)

        # Reset button states
        self.process_btn.setText("Process Invoice")
        self.process_btn.setEnabled(True)
        self.reprocess_btn.setEnabled(False)

        # Run processing with part number overrides
        self.status.setText("Reprocessing invoice...")
        self.start_processing(part_number_overrides=part_number_overrides)

        # Update qty_unit values from hts.db for all parts (in case HTS codes changed)
        units_updated = self.import_hts_units_silent()
        if units_updated > 0:
            logger.info(f"Updated qty_unit for {units_updated} parts from hts.db during reprocess")

        logger.info("Reprocessed invoice to pick up database changes")

    def save_preview_parts_to_db(self):
        """
        Save parts from the preview table to the database.
        This saves "Not Found" parts with their edited HTS codes, MIDs, etc.
        Returns the count of parts added/updated.
        """
        if self.table.rowCount() == 0:
            return 0

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            now = datetime.now().isoformat()
            added_count = 0

            # Track which parts we've already processed (avoid duplicates from derivative rows)
            processed_parts = set()
            # Track parts that were added (for HTS units import)
            saved_part_numbers = []

            for row in range(self.table.rowCount()):
                # Get part number from column 0
                part_item = self.table.item(row, 0)
                if not part_item:
                    continue
                part_number = part_item.text().strip().upper()
                if not part_number or part_number in processed_parts:
                    continue
                processed_parts.add(part_number)

                # Check if this row is marked as "Not Found" (column 18 = 232 Status)
                # Only save parts that were NOT in the database - don't overwrite existing DB values
                status_item = self.table.item(row, 18)
                status_text = status_item.text().strip() if status_item else ""
                if status_text != "Not Found":
                    # Part exists in database - skip to preserve database values
                    continue

                # Get values from the preview table (only for "Not Found" parts)
                hts_code = self.table.item(row, 2).text().strip() if self.table.item(row, 2) else ""
                mid = self.table.item(row, 3).text().strip() if self.table.item(row, 3) else ""

                # Check if part exists in database (case-insensitive) - double check
                c.execute("SELECT hts_code, mid FROM parts_master WHERE UPPER(part_number) = UPPER(?)", (part_number,))
                existing = c.fetchone()

                if existing:
                    # Part already exists in database - don't overwrite with preview table values
                    # This preserves any updates the user made directly to the database
                    continue
                else:
                    # Part doesn't exist - add it if we have at least an HTS code or MID
                    # Percentages are in 0-100 format; default to 100% non-232
                    if hts_code or mid:
                        # Auto-lookup qty_unit from hts_units table based on HTS code
                        qty_unit = get_hts_qty_unit(hts_code) if hts_code else ""
                        c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                                  steel_ratio, non_steel_ratio, last_updated, qty_unit, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                                  Sec301_Exclusion_Tariff)
                                  VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                  (part_number, '', hts_code, '', mid, '', 0.0, 100.0, now, qty_unit, 0.0, 0.0, 0.0, 0.0, ''))
                        added_count += 1
                        saved_part_numbers.append(part_number)
                        logger.info(f"Added new part to database: {part_number} (HTS: {hts_code}, MID: {mid})")

            conn.commit()
            conn.close()

            if added_count > 0:
                # Refresh the MID dropdown and parts table
                self.load_available_mids()

                # Import HTS units for the saved parts (updates CBP Qty1 field)
                if saved_part_numbers:
                    units_updated = self.import_hts_units_silent(saved_part_numbers)
                    if units_updated > 0:
                        logger.info(f"Updated CBP Qty1 units for {units_updated} saved parts")

            return added_count

        except Exception as e:
            logger.error(f"Failed to save preview parts to database: {e}")
            return 0

    def _export_single_file(self, df_out, cols, filename, is_network, steel_mask, aluminum_mask, copper_mask, wood_mask, auto_mask, non232_mask, sec301_mask):
        """Export a single Excel file with formatting. Used by both regular export and split-by-invoice export."""
        from openpyxl.styles import PatternFill
        from openpyxl.worksheet.page import PrintPageSetup

        # Helper function to get export color from per-user settings
        def get_export_color(config_key, default_color):
            return get_user_setting(config_key, default_color)

        # Get user-selected font color from per-user settings
        font_color_hex = get_user_setting('output_font_color', '#000000')
        font_color_rgb = '00' + font_color_hex.lstrip('#').upper()

        # Get Section 232 material type colors
        steel_color = get_export_color('export_steel_color', '#4a4a4a')
        aluminum_color = get_export_color('export_aluminum_color', '#6495ED')
        copper_color = get_export_color('export_copper_color', '#B87333')
        wood_color = get_export_color('export_wood_color', '#8B4513')
        auto_color = get_export_color('export_automotive_color', '#2F4F4F')
        non232_color = get_export_color('export_non232_color', '#FF0000')

        # Create fonts for each material type
        steel_font = ExcelFont(name='Arial', size=11, color='00' + steel_color.lstrip('#').upper())
        aluminum_font = ExcelFont(name='Arial', size=11, color='00' + aluminum_color.lstrip('#').upper())
        copper_font = ExcelFont(name='Arial', size=11, color='00' + copper_color.lstrip('#').upper())
        wood_font = ExcelFont(name='Arial', size=11, color='00' + wood_color.lstrip('#').upper())
        auto_font = ExcelFont(name='Arial', size=11, color='00' + auto_color.lstrip('#').upper())
        non232_font = ExcelFont(name='Arial', size=11, color='00' + non232_color.lstrip('#').upper())
        default_font = ExcelFont(name='Arial', size=11, color=font_color_rgb)

        orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

        # Build index lists for each material type
        steel_indices = [i for i, val in enumerate(steel_mask.tolist()) if val]
        aluminum_indices = [i for i, val in enumerate(aluminum_mask.tolist()) if val]
        copper_indices = [i for i, val in enumerate(copper_mask.tolist()) if val]
        wood_indices = [i for i, val in enumerate(wood_mask.tolist()) if val]
        auto_indices = [i for i, val in enumerate(auto_mask.tolist()) if val]
        non232_indices = [i for i, val in enumerate(non232_mask.tolist()) if val]
        sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

        if is_network:
            # Write to temp file then copy
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                temp_path = Path(tmp.name)

            with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                df_out[cols].to_excel(writer, index=False)
                ws = next(iter(writer.sheets.values()))

                # Apply font to header row
                for col_idx in range(1, len(cols) + 1):
                    ws.cell(row=1, column=col_idx).font = ExcelFont(name='Arial', size=11, bold=True)

                # Apply font and background to data rows
                for row_idx in range(len(df_out)):
                    row_num = row_idx + 2
                    is_sec301 = row_idx in sec301_indices

                    if row_idx in steel_indices:
                        cell_font = steel_font
                    elif row_idx in aluminum_indices:
                        cell_font = aluminum_font
                    elif row_idx in copper_indices:
                        cell_font = copper_font
                    elif row_idx in wood_indices:
                        cell_font = wood_font
                    elif row_idx in auto_indices:
                        cell_font = auto_font
                    elif row_idx in non232_indices:
                        cell_font = non232_font
                    else:
                        cell_font = default_font

                    for col_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.font = cell_font
                        if is_sec301:
                            cell.fill = orange_fill

                # Set page setup: landscape orientation, fit all columns on one page
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0  # 0 means unlimited pages vertically

            # Copy to network location
            out = OUTPUT_DIR / filename
            shutil.copy2(temp_path, out)
            temp_path.unlink()
        else:
            # Direct write for local path
            out = OUTPUT_DIR / filename
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                df_out[cols].to_excel(writer, index=False)
                ws = next(iter(writer.sheets.values()))

                center_alignment = Alignment(horizontal="center", vertical="center")

                # Apply font to header row
                for col_idx in range(1, len(cols) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = default_font
                    cell.alignment = center_alignment

                # Apply font and background to data rows
                for row_num in range(2, len(df_out) + 2):
                    row_idx = row_num - 2
                    is_sec301 = row_idx in sec301_indices

                    if row_idx in steel_indices:
                        font_to_use = steel_font
                    elif row_idx in aluminum_indices:
                        font_to_use = aluminum_font
                    elif row_idx in copper_indices:
                        font_to_use = copper_font
                    elif row_idx in wood_indices:
                        font_to_use = wood_font
                    elif row_idx in auto_indices:
                        font_to_use = auto_font
                    elif row_idx in non232_indices:
                        font_to_use = non232_font
                    else:
                        font_to_use = default_font

                    for col_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.font = font_to_use
                        cell.alignment = center_alignment
                        if is_sec301:
                            cell.fill = orange_fill

                # Auto-size columns
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Set page setup: landscape orientation, fit all columns on one page
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0  # 0 means unlimited pages vertically

        logger.info(f"Exported: {out.name}")
        return out

    def final_export(self):
        if self.last_processed_df is None:
            return
        
        # Check if table has rows before attempting export
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Empty Preview", "No data to export. Please process a shipment file first.")
            return
            
        # Ensure totals match prior to export
        running_total = 0.0
        for i in range(self.table.rowCount()):
            cell = self.table.item(i, 1)
            running_total += (cell.data(Qt.UserRole) or 0.0) if cell else 0.0
        
        # Compare against CI input value (what user entered/approved)
        ci_text = self.ci_input.text().replace(',', '').strip()
        try:
            target_value = float(ci_text)
        except:
            target_value = self.csv_total_value
        
        if abs(running_total - target_value) > 0.05:
            QMessageBox.warning(self, "Totals Mismatch", f"Values don't match invoice total.\nPreview: ${running_total:,.2f}\nTarget: ${target_value:,.2f}")
            return

        out = OUTPUT_DIR / (self.last_output_filename or f"Upload_Sheet_{datetime.now():%Y%m%d_%H%M}.xlsx")
        
        # Rebuild DataFrame from current table state (handles added/deleted/edited rows)
        export_data = []
        for i in range(self.table.rowCount()):
            value_cell = self.table.item(i, 1)
            value = value_cell.data(Qt.UserRole) if value_cell else 0.0
            
            # Get ratio percentages as floats (handle empty values)
            # Column indices: 12=Steel%, 13=Al%, 14=Cu%, 15=Wood%, 16=Auto%, 17=Non-232%, 18=232 Status
            steel_text = self.table.item(i, 12).text() if self.table.item(i, 12) else ""
            aluminum_text = self.table.item(i, 13).text() if self.table.item(i, 13) else ""
            copper_text = self.table.item(i, 14).text() if self.table.item(i, 14) else ""
            wood_text = self.table.item(i, 15).text() if self.table.item(i, 15) else ""
            auto_text = self.table.item(i, 16).text() if self.table.item(i, 16) else ""

            # Parse percentages safely (values are already in 0-100 format)
            def parse_pct(text):
                if not text or text.strip() == '':
                    return 0.0
                try:
                    return float(text.replace('%', '').strip())
                except (ValueError, TypeError):
                    return 0.0

            steel_ratio = parse_pct(steel_text)
            aluminum_ratio = parse_pct(aluminum_text)
            copper_ratio = parse_pct(copper_text)
            wood_ratio = parse_pct(wood_text)
            auto_ratio = parse_pct(auto_text)

            # Get Non-232% ratio from column 17
            non_steel_text = self.table.item(i, 17).text() if self.table.item(i, 17) else ""
            non_steel_ratio = parse_pct(non_steel_text)

            # Get Sec301 exclusion data and Qty1/Qty2 from last_processed_df if available
            sec301_exclusion = ""
            invoice_number = ""
            qty1_value = ""
            qty2_value = ""
            if self.last_processed_df is not None and i < len(self.last_processed_df):
                sec301_exclusion = str(self.last_processed_df.iloc[i].get('Sec301_Exclusion_Tariff', '')).strip()
                if sec301_exclusion in ['', 'nan', 'None']:
                    sec301_exclusion = ""
                # Get invoice number for split by invoice feature
                invoice_number = str(self.last_processed_df.iloc[i].get('invoice_number', '')).strip()
                if invoice_number in ['', 'nan', 'None']:
                    invoice_number = ""
                # Get Qty1 and Qty2 (calculated based on qty_unit during processing)
                qty1_value = str(self.last_processed_df.iloc[i].get('Qty1', '')).strip()
                if qty1_value in ['nan', 'None']:
                    qty1_value = ""
                qty2_value = str(self.last_processed_df.iloc[i].get('Qty2', '')).strip()
                if qty2_value in ['nan', 'None']:
                    qty2_value = ""

            # Column indices:
            # 0=Product No, 1=Value, 2=HTS, 3=MID, 4=Qty1, 5=Qty2, 6=Qty Unit, 7=Dec,
            # 8=Melt, 9=Cast, 10=Smelt, 11=Flag, 12=Steel%, 13=Al%, 14=Cu%, 15=Wood%, 16=Auto%, 17=Non-232%, 18=232 Status
            row_data = {
                'Product No': self.table.item(i, 0).text() if self.table.item(i, 0) else "",
                'ValueUSD': value,
                'HTSCode': self.table.item(i, 2).text() if self.table.item(i, 2) else "",
                'MID': self.table.item(i, 3).text() if self.table.item(i, 3) else "",
                'CalcWtNet': round(float(self.table.item(i, 4).text())) if self.table.item(i, 4) and self.table.item(i, 4).text() else 0,
                'Pcs': int(self.table.item(i, 5).text()) if self.table.item(i, 5) and self.table.item(i, 5).text() else 0,
                'Qty1': qty1_value if qty1_value else '',
                'Qty2': qty2_value if qty2_value else '',
                'DecTypeCd': self.table.item(i, 7).text() if self.table.item(i, 7) else "CO",
                'CountryofMelt': self.table.item(i, 8).text() if self.table.item(i, 8) else "",
                'CountryOfCast': self.table.item(i, 9).text() if self.table.item(i, 9) else "",
                'PrimCountryOfSmelt': self.table.item(i, 10).text() if self.table.item(i, 10) else "",
                'DeclarationFlag': self.table.item(i, 11).text() if self.table.item(i, 11) else "",
                'SteelRatio': steel_ratio,
                'AluminumRatio': aluminum_ratio,
                'CopperRatio': copper_ratio,
                'WoodRatio': wood_ratio,
                'AutoRatio': auto_ratio,
                'NonSteelRatio': non_steel_ratio,
                '_232_flag': self.table.item(i, 18).text() if self.table.item(i, 18) else "",  # Column 18 is 232_Status
                '_sec301_exclusion': sec301_exclusion,
                '_invoice_number': invoice_number
            }
            export_data.append(row_data)

        df_out = pd.DataFrame(export_data)

        # Add CustomerRef column from input field
        customer_ref = self.customer_ref_input.text().strip() if hasattr(self, 'customer_ref_input') and self.customer_ref_input else ""
        df_out['CustomerRef'] = customer_ref

        # Build masks for each Section 232 material type BEFORE converting to percentage strings
        steel_mask = df_out['_232_flag'].fillna('').str.contains('232_Steel', case=False, na=False)
        aluminum_mask = df_out['_232_flag'].fillna('').str.contains('232_Aluminum', case=False, na=False)
        copper_mask = df_out['_232_flag'].fillna('').str.contains('232_Copper', case=False, na=False)
        wood_mask = df_out['_232_flag'].fillna('').str.contains('232_Wood', case=False, na=False)
        auto_mask = df_out['_232_flag'].fillna('').str.contains('232_Auto', case=False, na=False)
        non232_mask = df_out['_232_flag'].fillna('').str.contains('Non_232', case=False, na=False)

        # Build mask for Sec301 exclusion rows (for light orange background)
        sec301_mask = df_out['_sec301_exclusion'].fillna('').astype(str).str.strip().ne('') & \
                      ~df_out['_sec301_exclusion'].fillna('').astype(str).str.contains('nan|None', case=False, na=False)
        
        # Convert ratio values to percentage strings for export (values are already 0-100)
        df_out['SteelRatio'] = df_out['SteelRatio'].round(1).astype(str) + "%"
        df_out['AluminumRatio'] = df_out['AluminumRatio'].round(1).astype(str) + "%"
        df_out['CopperRatio'] = df_out['CopperRatio'].round(1).astype(str) + "%"
        df_out['WoodRatio'] = df_out['WoodRatio'].round(1).astype(str) + "%"
        df_out['AutoRatio'] = df_out['AutoRatio'].round(1).astype(str) + "%"
        df_out['NonSteelRatio'] = df_out['NonSteelRatio'].round(1).astype(str) + "%"
        df_out['232_Status'] = df_out['_232_flag'].fillna('')

        # Build columns list using saved column order (or default order)
        # Qty1/Qty2 replace CalcWtNet/Pcs for conditional quantity output based on qty_unit
        if hasattr(self, 'output_column_order') and self.output_column_order:
            all_columns = self.output_column_order
        else:
            all_columns = ['Product No', 'ValueUSD', 'HTSCode', 'MID', 'Qty1', 'Qty2',
                'DecTypeCd', 'CountryofMelt', 'CountryOfCast', 'PrimCountryOfSmelt',
                'DeclarationFlag', 'SteelRatio', 'AluminumRatio', 'CopperRatio',
                'WoodRatio', 'AutoRatio', 'NonSteelRatio', '232_Status', 'CustomerRef']

        # Filter columns based on visibility settings
        cols = []
        ratio_columns = ['SteelRatio', 'AluminumRatio', 'CopperRatio', 'WoodRatio', 'AutoRatio', 'NonSteelRatio']
        for col in all_columns:
            # Check visibility for ratio columns
            if col in ratio_columns:
                is_visible = True
                try:
                    conn = sqlite3.connect(str(DB_PATH))
                    c = conn.cursor()
                    c.execute("SELECT value FROM app_config WHERE key = ?", (f'export_col_visible_{col}',))
                    row = c.fetchone()
                    conn.close()
                    if row:
                        is_visible = row[0] == 'True'
                except:
                    pass
                if is_visible:
                    cols.append(col)
            else:
                # Non-ratio columns are always included
                cols.append(col)

        # Filter out columns that don't exist in the DataFrame
        cols = [col for col in cols if col in df_out.columns]

        # Ensure we have at least one column to export
        if not cols:
            QMessageBox.critical(self, "Export Error", "No valid columns to export. Please check your column configuration.")
            return

        # Apply custom column mapping if set
        if hasattr(self, 'output_column_mapping') and self.output_column_mapping:
            # Create rename dictionary for columns that have custom names
            rename_dict = {}
            for col in cols:
                if col in self.output_column_mapping and self.output_column_mapping[col] != col:
                    rename_dict[col] = self.output_column_mapping[col]

            # Rename columns if custom names are defined
            if rename_dict:
                df_out = df_out.rename(columns=rename_dict)
                # Update cols list with new names
                cols = [self.output_column_mapping.get(col, col) for col in cols]
                logger.info(f"Applied custom column mapping: {rename_dict}")

        # Check if we should split by invoice number
        split_by_invoice = False
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", ('export_split_by_invoice',))
            row = c.fetchone()
            conn.close()
            if row:
                split_by_invoice = row[0] == 'True'
        except:
            pass

        # Get unique invoice numbers if splitting is enabled
        unique_invoices = []
        if split_by_invoice and '_invoice_number' in df_out.columns:
            unique_invoices = df_out['_invoice_number'].dropna().unique()
            unique_invoices = [inv for inv in unique_invoices if inv and str(inv).strip() not in ['', 'nan', 'None']]
            if len(unique_invoices) > 1:
                logger.info(f"Split by invoice enabled: Found {len(unique_invoices)} unique invoices")
            else:
                # Only one or no invoices, don't split
                unique_invoices = []

        try:
            t_start = time.time()

            # Show export progress indicator
            self.export_progress_widget.show()
            self.export_status_label.setText("Exporting:")
            self.export_progress_bar.setValue(0)
            QApplication.processEvents()
            
            # Check if output directory is on network (slower) or local
            output_str = str(OUTPUT_DIR)
            is_network = output_str.startswith('\\\\') or (len(output_str) > 1 and output_str[1] == ':' and not output_str.startswith('C:'))

            # Handle split by invoice if enabled
            if unique_invoices and len(unique_invoices) > 1:
                # Export multiple files, one per invoice
                exported_files = []
                total_invoices = len(unique_invoices)

                for idx, invoice_num in enumerate(unique_invoices):
                    # Filter dataframe for this invoice
                    invoice_df = df_out[df_out['_invoice_number'] == invoice_num].copy()

                    # Recalculate masks for this subset
                    inv_steel_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Steel', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_aluminum_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Aluminum', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_copper_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Copper', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_wood_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Wood', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_auto_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Auto', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_non232_mask = invoice_df['_232_flag'].fillna('').str.contains('Non_232', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_sec301_mask = invoice_df['_sec301_exclusion'].fillna('').astype(str).str.strip().ne('') & \
                                      ~invoice_df['_sec301_exclusion'].fillna('').astype(str).str.contains('nan|None', case=False, na=False) if '_sec301_exclusion' in invoice_df.columns else pd.Series([False] * len(invoice_df))

                    # Generate filename with invoice number and date
                    invoice_filename = f"{invoice_num}_{datetime.now():%Y%m%d}.xlsx"

                    # Update progress
                    progress_pct = int(10 + (idx / total_invoices) * 80)
                    self.export_progress_bar.setValue(progress_pct)
                    self.bottom_status.setText(f"Exporting invoice {idx + 1} of {total_invoices}: {invoice_num}")
                    QApplication.processEvents()

                    # Export this invoice's data
                    out_path = self._export_single_file(
                        invoice_df, cols, invoice_filename, is_network,
                        inv_steel_mask, inv_aluminum_mask, inv_copper_mask, inv_wood_mask,
                        inv_auto_mask, inv_non232_mask, inv_sec301_mask
                    )
                    exported_files.append(out_path.name)

                self.export_progress_bar.setValue(100)
                QApplication.processEvents()

                # Move processed CSV to Processed folder
                if self.current_csv and Path(self.current_csv).exists():
                    try:
                        source_file = Path(self.current_csv)
                        dest_file = PROCESSED_DIR / source_file.name
                        if dest_file.exists():
                            dest_file.unlink()
                        source_file.rename(dest_file)
                        logger.info(f"Moved processed file: {source_file.name} -> Processed/")
                        self.current_csv = None
                    except Exception as e:
                        logger.warning(f"Could not move CSV to Processed folder: {e}")

                self.refresh_exported_files()
                self.refresh_input_files()

                # Add "Not Found" parts to the database
                added_parts_count = self.add_not_found_parts_to_db()

                QTimer.singleShot(500, self.export_progress_widget.hide)

                # Build success message
                success_msg = f"Export complete!\nCreated {len(exported_files)} files:\n" + "\n".join(exported_files[:10])
                if len(exported_files) > 10:
                    success_msg += f"\n... and {len(exported_files) - 10} more"
                if added_parts_count > 0:
                    success_msg += f"\n\n{added_parts_count} new part(s) added to database."

                QMessageBox.information(self, "Success", success_msg)
                logger.success(f"Split export complete: {len(exported_files)} files created" + (f" ({added_parts_count} parts added to DB)" if added_parts_count > 0 else ""))
                self.clear_all()
                return

            # If network path, use local temp then copy (40x faster!)
            if is_network:
                self.bottom_status.setText("Generating export file...")
                self.export_progress_bar.setValue(10)
                QApplication.processEvents()
                
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    temp_path = Path(tmp.name)
                
                self.export_progress_bar.setValue(20)
                QApplication.processEvents()
                
                with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                    df_out[cols].to_excel(writer, index=False)
                    t_write = time.time() - t_start
                    
                    self.export_progress_bar.setValue(50)
                    QApplication.processEvents()
                    
                    # Apply Arial font to all cells and red font to non-steel rows
                    t_format_start = time.time()
                    ws = next(iter(writer.sheets.values()))

                    # Set Arial font for all cells (including header)
                    from openpyxl.styles import PatternFill

                    # Helper function to get export color from per-user settings
                    def get_export_color(config_key, default_color):
                        return get_user_setting(config_key, default_color)

                    # Get user-selected font color from per-user settings
                    font_color_hex = get_user_setting('output_font_color', '#000000')
                    font_color_rgb = '00' + font_color_hex.lstrip('#').upper()

                    # Get Section 232 material type colors
                    steel_color = get_export_color('export_steel_color', '#4a4a4a')
                    aluminum_color = get_export_color('export_aluminum_color', '#6495ED')
                    copper_color = get_export_color('export_copper_color', '#B87333')
                    wood_color = get_export_color('export_wood_color', '#8B4513')
                    auto_color = get_export_color('export_automotive_color', '#2F4F4F')
                    non232_color = get_export_color('export_non232_color', '#FF0000')

                    # Create fonts for each material type
                    steel_font = ExcelFont(name='Arial', size=11, color='00' + steel_color.lstrip('#').upper())
                    aluminum_font = ExcelFont(name='Arial', size=11, color='00' + aluminum_color.lstrip('#').upper())
                    copper_font = ExcelFont(name='Arial', size=11, color='00' + copper_color.lstrip('#').upper())
                    wood_font = ExcelFont(name='Arial', size=11, color='00' + wood_color.lstrip('#').upper())
                    auto_font = ExcelFont(name='Arial', size=11, color='00' + auto_color.lstrip('#').upper())
                    non232_font = ExcelFont(name='Arial', size=11, color='00' + non232_color.lstrip('#').upper())
                    default_font = ExcelFont(name='Arial', size=11, color=font_color_rgb)

                    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Light orange

                    # Apply font to header row
                    for col_idx in range(1, len(cols) + 1):
                        ws.cell(row=1, column=col_idx).font = ExcelFont(name='Arial', size=11, bold=True)

                    # Build index lists for each material type
                    steel_indices = [i for i, val in enumerate(steel_mask.tolist()) if val]
                    aluminum_indices = [i for i, val in enumerate(aluminum_mask.tolist()) if val]
                    copper_indices = [i for i, val in enumerate(copper_mask.tolist()) if val]
                    wood_indices = [i for i, val in enumerate(wood_mask.tolist()) if val]
                    auto_indices = [i for i, val in enumerate(auto_mask.tolist()) if val]
                    non232_indices = [i for i, val in enumerate(non232_mask.tolist()) if val]
                    sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

                    # Apply font and background to data rows
                    for row_idx in range(len(df_out)):
                        row_num = row_idx + 2
                        is_sec301 = row_idx in sec301_indices

                        # Determine font based on material type
                        if row_idx in steel_indices:
                            cell_font = steel_font
                        elif row_idx in aluminum_indices:
                            cell_font = aluminum_font
                        elif row_idx in copper_indices:
                            cell_font = copper_font
                        elif row_idx in wood_indices:
                            cell_font = wood_font
                        elif row_idx in auto_indices:
                            cell_font = auto_font
                        elif row_idx in non232_indices:
                            cell_font = non232_font
                        else:
                            cell_font = default_font

                        for col_idx in range(1, len(cols) + 1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.font = cell_font
                            # Apply light orange background for Sec301 exclusions
                            if is_sec301:
                                cell.fill = orange_fill
                    
                    t_format = time.time() - t_format_start
                
                self.export_progress_bar.setValue(70)
                QApplication.processEvents()
                
                # Copy to network location
                self.bottom_status.setText("Copying to network location...")
                self.export_status_label.setText("Copying:")
                t_copy_start = time.time()
                out = OUTPUT_DIR / (self.last_output_filename or f"Upload_Sheet_{datetime.now():%Y%m%d_%H%M}.xlsx")
                shutil.copy2(temp_path, out)
                temp_path.unlink()
                t_copy = time.time() - t_copy_start
                
                self.export_progress_bar.setValue(90)
                QApplication.processEvents()
                
                t_total = time.time() - t_start
                logger.info(f"Export timing - Write: {t_write:.2f}s, Format: {t_format:.2f}s, Copy: {t_copy:.2f}s, Total: {t_total:.2f}s")
            else:
                # Local path - direct write
                self.export_progress_bar.setValue(20)
                QApplication.processEvents()
                
                out = OUTPUT_DIR / (self.last_output_filename or f"Upload_Sheet_{datetime.now():%Y%m%d_%H%M}.xlsx")
                with pd.ExcelWriter(out, engine='openpyxl') as writer:
                    df_out[cols].to_excel(writer, index=False)
                    t_write = time.time() - t_start
                    
                    self.export_progress_bar.setValue(60)
                    QApplication.processEvents()
                    
                    # Apply formatting: Arial font, center alignment, auto-sized columns
                    t_format_start = time.time()
                    ws = next(iter(writer.sheets.values()))

                    # Create font, fill, and alignment styles
                    from openpyxl.styles import PatternFill

                    # Helper function to get export color from per-user settings
                    def get_export_color(config_key, default_color):
                        return get_user_setting(config_key, default_color)

                    # Get user-selected font color from per-user settings
                    font_color_hex = get_user_setting('output_font_color', '#000000')
                    font_color_rgb = '00' + font_color_hex.lstrip('#').upper()

                    # Get Section 232 material type colors
                    steel_color = get_export_color('export_steel_color', '#4a4a4a')
                    aluminum_color = get_export_color('export_aluminum_color', '#6495ED')
                    copper_color = get_export_color('export_copper_color', '#B87333')
                    wood_color = get_export_color('export_wood_color', '#8B4513')
                    auto_color = get_export_color('export_automotive_color', '#2F4F4F')
                    non232_color = get_export_color('export_non232_color', '#FF0000')

                    # Create fonts for each material type
                    steel_font = ExcelFont(name='Arial', color='00' + steel_color.lstrip('#').upper())
                    aluminum_font = ExcelFont(name='Arial', color='00' + aluminum_color.lstrip('#').upper())
                    copper_font = ExcelFont(name='Arial', color='00' + copper_color.lstrip('#').upper())
                    wood_font = ExcelFont(name='Arial', color='00' + wood_color.lstrip('#').upper())
                    auto_font = ExcelFont(name='Arial', color='00' + auto_color.lstrip('#').upper())
                    non232_font = ExcelFont(name='Arial', color='00' + non232_color.lstrip('#').upper())
                    normal_font = ExcelFont(name="Arial", color=font_color_rgb)

                    center_alignment = Alignment(horizontal="center", vertical="center")
                    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Light orange

                    # Build index lists for each material type
                    steel_indices = [i for i, val in enumerate(steel_mask.tolist()) if val]
                    aluminum_indices = [i for i, val in enumerate(aluminum_mask.tolist()) if val]
                    copper_indices = [i for i, val in enumerate(copper_mask.tolist()) if val]
                    wood_indices = [i for i, val in enumerate(wood_mask.tolist()) if val]
                    auto_indices = [i for i, val in enumerate(auto_mask.tolist()) if val]
                    non232_indices = [i for i, val in enumerate(non232_mask.tolist()) if val]
                    sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

                    # Apply font and background to data rows
                    for row_num in range(2, len(df_out) + 2):  # Start at 2 (after header)
                        row_idx = row_num - 2
                        is_sec301 = row_idx in sec301_indices

                        # Determine font based on material type
                        if row_idx in steel_indices:
                            font_to_use = steel_font
                        elif row_idx in aluminum_indices:
                            font_to_use = aluminum_font
                        elif row_idx in copper_indices:
                            font_to_use = copper_font
                        elif row_idx in wood_indices:
                            font_to_use = wood_font
                        elif row_idx in auto_indices:
                            font_to_use = auto_font
                        elif row_idx in non232_indices:
                            font_to_use = non232_font
                        else:
                            font_to_use = normal_font

                        for col_idx in range(1, len(cols) + 1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.font = font_to_use
                            cell.alignment = center_alignment
                            # Apply light orange background for Sec301 exclusions
                            if is_sec301:
                                cell.fill = orange_fill

                    # Apply Arial font and center alignment to header row
                    for col_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = normal_font
                        cell.alignment = center_alignment

                    # Auto-size columns based on content with padding
                    for col_idx, column in enumerate(ws.columns, 1):
                        max_length = 0
                        column_letter = ws.cell(row=1, column=col_idx).column_letter

                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass

                        # Add padding (2 extra characters) and set column width
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column_letter].width = adjusted_width

                    t_format = time.time() - t_format_start
                
                self.export_progress_bar.setValue(90)
                QApplication.processEvents()
                
                t_total = time.time() - t_start
                logger.info(f"Export timing - Write: {t_write:.2f}s, Format: {t_format:.2f}s, Total: {t_total:.2f}s")
            
            self.export_progress_bar.setValue(100)
            QApplication.processEvents()
            
            # Move processed CSV to Processed folder
            if self.current_csv and Path(self.current_csv).exists():
                try:
                    source_file = Path(self.current_csv)
                    dest_file = PROCESSED_DIR / source_file.name
                    
                    # If destination exists, remove it first
                    if dest_file.exists():
                        dest_file.unlink()
                    
                    source_file.rename(dest_file)
                    logger.info(f"Moved processed file: {source_file.name} -> Processed/")
                    self.current_csv = None
                except Exception as e:
                    logger.warning(f"Could not move CSV to Processed folder: {e}")
            
            self.refresh_exported_files()
            self.refresh_input_files()  # Refresh to show file moved

            # Add "Not Found" parts to the database
            added_parts_count = self.add_not_found_parts_to_db()

            # Record billing event for this export
            try:
                hts_codes = [row.get('HTSCode', '') for row in export_data if row.get('HTSCode')]
                processing_time_ms = int((time.time() - t_start) * 1000) if 't_start' in dir() else 0
                self.record_billing_event(
                    file_name=out.name,
                    line_count=len(export_data),
                    total_value=running_total,
                    hts_codes=hts_codes,
                    processing_time_ms=processing_time_ms
                )
                logger.info(f"Recorded billing event for {out.name}: {len(export_data)} lines, ${running_total:,.2f}")
            except Exception as billing_err:
                logger.warning(f"Failed to record billing event: {billing_err}")

            # Hide progress indicator after brief delay
            QTimer.singleShot(500, self.export_progress_widget.hide)

            # Build success message
            success_msg = f"Export complete!\nSaved: {out.name}"
            if added_parts_count > 0:
                success_msg += f"\n\n{added_parts_count} new part(s) added to database."

            QMessageBox.information(self, "Success", success_msg)
            logger.success(f"Export complete: {out.name}" + (f" ({added_parts_count} parts added to DB)" if added_parts_count > 0 else ""))
        except Exception as e:
            self.export_progress_widget.hide()
            QMessageBox.critical(self, "Export Failed", str(e))
            return
        self.clear_all()

    # TODO: XML Export - To be implemented at a later date
    # def export_to_xml(self):
    #     """Export processed invoice data to XML format for e2Open Customs Management."""
    #     if self.last_processed_df is None or self.table.rowCount() == 0:
    #         QMessageBox.warning(self, "No Data", "No processed data to export. Please process a shipment file first.")
    #         return
    #
    #     # Build filename from CSV name + date/time
    #     # Get customer reference from input field, fall back to CSV filename
    #     customer_ref = self.customer_ref_input.text().strip() if hasattr(self, 'customer_ref_input') else ""
    #     if self.current_csv:
    #         csv_name = Path(self.current_csv).stem  # Get filename without extension
    #     else:
    #         csv_name = "Invoice"
    #     # Use customer reference for filename if provided, otherwise use CSV name
    #     filename_base = customer_ref if customer_ref else csv_name
    #     default_filename = f"{filename_base}_{datetime.now():%Y%m%d_%H%M}.xml"
    #
    #     # Prompt user for save location
    #     file_path, _ = QFileDialog.getSaveFileName(
    #         self,
    #         "Export Commercial Invoice XML",
    #         str(OUTPUT_DIR / default_filename),
    #         "XML Files (*.xml);;All Files (*)"
    #     )
    #
    #     if not file_path:
    #         return  # User cancelled
    #
    #     try:
    #         # Build export data from current table state
    #         export_rows = []
    #         for i in range(self.table.rowCount()):
    #             value_cell = self.table.item(i, 1)
    #             value = value_cell.data(Qt.UserRole) if value_cell else 0.0
    #
    #             # Get Qty1/Qty2 from last_processed_df
    #             qty1_value = ""
    #             qty2_value = ""
    #             qty_unit = ""
    #             if self.last_processed_df is not None and i < len(self.last_processed_df):
    #                 qty1_value = str(self.last_processed_df.iloc[i].get('Qty1', '')).strip()
    #                 if qty1_value in ['nan', 'None']:
    #                     qty1_value = ""
    #                 qty2_value = str(self.last_processed_df.iloc[i].get('Qty2', '')).strip()
    #                 if qty2_value in ['nan', 'None']:
    #                     qty2_value = ""
    #                 qty_unit = str(self.last_processed_df.iloc[i].get('qty_unit', '')).strip().upper()
    #                 if qty_unit in ['nan', 'None']:
    #                     qty_unit = ""
    #
    #             row_data = {
    #                 'product_no': self.table.item(i, 0).text() if self.table.item(i, 0) else "",
    #                 'value_usd': value,
    #                 'hts_code': self.table.item(i, 2).text() if self.table.item(i, 2) else "",
    #                 'mid': self.table.item(i, 3).text() if self.table.item(i, 3) else "",
    #                 'qty1': qty1_value,
    #                 'qty2': qty2_value,
    #                 'qty_unit': qty_unit,
    #                 'dec_type_cd': self.table.item(i, 7).text() if self.table.item(i, 7) else "",
    #                 'country_of_melt': self.table.item(i, 8).text() if self.table.item(i, 8) else "",
    #                 'country_of_cast': self.table.item(i, 9).text() if self.table.item(i, 9) else "",
    #                 'country_of_smelt': self.table.item(i, 10).text() if self.table.item(i, 10) else "",
    #                 'declaration_flag': self.table.item(i, 11).text() if self.table.item(i, 11) else "",
    #                 'status_232': self.table.item(i, 18).text() if self.table.item(i, 18) else ""
    #             }
    #             export_rows.append(row_data)
    #
    #         # Get client_code (importer ID) from parts_master for any part in the invoice
    #         importer_id = ""
    #         try:
    #             part_numbers = [row['product_no'] for row in export_rows if row.get('product_no')]
    #             if part_numbers:
    #                 conn = sqlite3.connect(str(DB_PATH))
    #                 c = conn.cursor()
    #                 placeholders = ','.join(['?' for _ in part_numbers])
    #                 c.execute(f"""SELECT DISTINCT client_code FROM parts_master
    #                              WHERE part_number IN ({placeholders})
    #                              AND client_code IS NOT NULL AND client_code != ''
    #                              LIMIT 1""", part_numbers)
    #                 result = c.fetchone()
    #                 if result:
    #                     importer_id = result[0]
    #                 conn.close()
    #         except Exception as e:
    #             logger.warning(f"Could not fetch client_code for XML export: {e}")
    #
    #         # Generate XML (use customer ref if provided, otherwise CSV name; client_code as importer ID)
    #         reference_number = customer_ref if customer_ref else csv_name
    #         xml_content = self._generate_commercial_invoice_xml(export_rows, reference_number, importer_id)
    #
    #         # Write to file
    #         with open(file_path, 'w', encoding='utf-8') as f:
    #             f.write(xml_content)
    #
    #         QMessageBox.information(self, "Success", f"XML export complete!\nSaved: {Path(file_path).name}")
    #         logger.success(f"XML export complete: {file_path}")
    #
    #     except Exception as e:
    #         logger.error(f"XML export failed: {e}")
    #         QMessageBox.critical(self, "Export Failed", f"XML export failed: {str(e)}")

    # TODO: Lacey Act PPQ 505 Export - To be implemented at a later date
    # def export_lacey_act_ppq505(self):
    #     """Export items requiring Lacey Act declaration to PPQ Form 505 format (Excel)."""
    #     if self.last_processed_df is None or self.table.rowCount() == 0:
    #         QMessageBox.warning(self, "No Data", "No processed data to export. Please process a shipment file first.")
    #         return
    #
    #     # Filter for Lacey Act items only
    #     df = self.last_processed_df.copy()
    #     if '_lacey_required' not in df.columns:
    #         QMessageBox.warning(self, "No Lacey Data",
    #             "Lacey Act information not available. This may be an older processed file.\n"
    #             "Please reprocess the invoice to detect Lacey Act items.")
    #         return
    #
    #     lacey_df = df[df['_lacey_required'] == 'Y'].copy()
    #
    #     if len(lacey_df) == 0:
    #         QMessageBox.information(self, "No Lacey Items",
    #             "No items in this shipment require Lacey Act declaration.\n\n"
    #             "Lacey Act applies to:\n"
    #             "- HTS Chapters 44, 47, 48 (Wood, Pulp, Paper)\n"
    #             "- HTS 9401, 9403 (Wood furniture)\n"
    #             "- Any item with wood content > 0%")
    #         return
    #
    #     # Build filename
    #     csv_name = Path(self.current_csv).stem if self.current_csv else "Invoice"
    #     default_filename = f"{csv_name}_PPQ505_{datetime.now():%Y%m%d_%H%M}.xlsx"
    #
    #     # Prompt user for save location
    #     file_path, _ = QFileDialog.getSaveFileName(
    #         self,
    #         "Export Lacey Act PPQ Form 505",
    #         str(OUTPUT_DIR / default_filename),
    #         "Excel Files (*.xlsx);;All Files (*)"
    #     )
    #
    #     if not file_path:
    #         return  # User cancelled
    #
    #     try:
    #         # Prepare PPQ 505 columns
    #         ppq505_columns = {
    #             'HTSCode': 'HTSUS Number',
    #             'ValueUSD': 'Entered Value (USD)',
    #             'Product No': 'Article/Component',
    #             'LaceySpecies': 'Genus & Species (Scientific Name)',
    #             'LaceyHarvestCountry': 'Country of Harvest',
    #             'CalcWtNet': 'Quantity',
    #             'qty_unit': 'Unit of Measure',
    #             'LaceyRecycledPct': '% Recycled',
    #             'WoodRatio': 'Wood Content %',
    #         }
    #
    #         # Create export dataframe with PPQ 505 format
    #         export_df = pd.DataFrame()
    #         for src_col, dest_col in ppq505_columns.items():
    #             if src_col in lacey_df.columns:
    #                 export_df[dest_col] = lacey_df[src_col]
    #             else:
    #                 export_df[dest_col] = ''
    #
    #         # Add warning column for missing data
    #         warnings = []
    #         for _, row in export_df.iterrows():
    #             missing = []
    #             if not row.get('Genus & Species (Scientific Name)', ''):
    #                 missing.append('Species')
    #             if not row.get('Country of Harvest', ''):
    #                 missing.append('Harvest Country')
    #             warnings.append(', '.join(missing) if missing else '')
    #         export_df['Missing Data'] = warnings
    #
    #         # Export to Excel with formatting
    #         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    #             export_df.to_excel(writer, index=False, sheet_name='PPQ 505 Data')
    #
    #             # Access workbook for formatting
    #             workbook = writer.book
    #             worksheet = writer.sheets['PPQ 505 Data']
    #
    #             # Format header row
    #             from openpyxl.styles import Font, PatternFill, Alignment
    #             header_fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
    #             header_font = Font(bold=True, color='FFFFFF')
    #
    #             for col_num, cell in enumerate(worksheet[1], 1):
    #                 cell.fill = header_fill
    #                 cell.font = header_font
    #                 cell.alignment = Alignment(horizontal='center')
    #
    #             # Highlight rows with missing data
    #             warning_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
    #             for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
    #                 missing_data_cell = worksheet.cell(row=row_num, column=len(ppq505_columns) + 1)
    #                 if missing_data_cell.value:
    #                     for cell in row:
    #                         cell.fill = warning_fill
    #
    #             # Adjust column widths
    #             for column in worksheet.columns:
    #                 max_length = 0
    #                 column_letter = column[0].column_letter
    #                 for cell in column:
    #                     try:
    #                         if len(str(cell.value)) > max_length:
    #                             max_length = len(str(cell.value))
    #                     except:
    #                         pass
    #                 adjusted_width = min(max_length + 2, 50)
    #                 worksheet.column_dimensions[column_letter].width = adjusted_width
    #
    #         QMessageBox.information(self, "Success",
    #             f"Lacey Act PPQ 505 export complete!\n\n"
    #             f"Items exported: {len(lacey_df)}\n"
    #             f"File: {Path(file_path).name}\n\n"
    #             f"Note: Review items highlighted in orange - they have missing species or country of harvest data.")
    #         logger.success(f"Lacey Act PPQ 505 export: {len(lacey_df)} items to {file_path}")
    #
    #     except Exception as e:
    #         logger.error(f"Lacey Act export failed: {e}")
    #         QMessageBox.critical(self, "Export Failed", f"Lacey Act export failed: {str(e)}")

    # TODO: XML Generation - To be implemented at a later date
    # def _generate_commercial_invoice_xml(self, rows, customer_reference="", importer_id=""):
    #     """Generate XML content for commercial invoice in e2Open-compatible format."""
    #     # Create root element with namespace
    #     root = ET.Element('CommercialInvoice')
    #     root.set('xmlns', 'urn:customs:commercial-invoice:v1')
    #     root.set('xmlns:xsi', 'http://www.w3.org/2001/XMLSchema-instance')
    #
    #     # Add header information
    #     header = ET.SubElement(root, 'Header')
    #
    #     # Document information
    #     doc_info = ET.SubElement(header, 'DocumentInfo')
    #     ET.SubElement(doc_info, 'DocumentType').text = 'CommercialInvoice'
    #     ET.SubElement(doc_info, 'CreationDateTime').text = datetime.now().isoformat()
    #     ET.SubElement(doc_info, 'DocumentID').text = f"INV-{datetime.now():%Y%m%d%H%M%S}"
    #     if customer_reference:
    #         ET.SubElement(doc_info, 'CustomerReferenceNumber').text = customer_reference
    #
    #     # Importer information (from client_code in parts_master)
    #     if importer_id:
    #         importer = ET.SubElement(header, 'Importer')
    #         ET.SubElement(importer, 'ImporterID').text = importer_id
    #
    #     # Exporter/Shipper information (from MID if available)
    #     first_mid = rows[0].get('mid', '') if rows else ''
    #     if first_mid:
    #         shipper = ET.SubElement(header, 'Shipper')
    #         ET.SubElement(shipper, 'ManufacturerID').text = first_mid
    #         # Extract country from MID prefix (first 2 characters)
    #         if len(first_mid) >= 2:
    #             ET.SubElement(shipper, 'CountryCode').text = first_mid[:2]
    #
    #     # Invoice summary
    #     summary = ET.SubElement(header, 'InvoiceSummary')
    #     total_value = sum(row.get('value_usd', 0) for row in rows)
    #     ET.SubElement(summary, 'TotalValue').text = f"{total_value:.2f}"
    #     ET.SubElement(summary, 'CurrencyCode').text = 'USD'
    #     ET.SubElement(summary, 'TotalLineItems').text = str(len(rows))
    #
    #     # Line items
    #     line_items = ET.SubElement(root, 'LineItems')
    #
    #     for idx, row in enumerate(rows, start=1):
    #         item = ET.SubElement(line_items, 'LineItem')
    #         item.set('lineNumber', str(idx))
    #
    #         # Product identification
    #         ET.SubElement(item, 'ProductNumber').text = row.get('product_no', '')
    #
    #         # Tariff classification
    #         tariff = ET.SubElement(item, 'TariffClassification')
    #         hts_code = row.get('hts_code', '')
    #         ET.SubElement(tariff, 'HTSCode').text = hts_code
    #         # Extract chapter for material type indication
    #         if len(hts_code.replace('.', '')) >= 2:
    #             ET.SubElement(tariff, 'HTSChapter').text = hts_code.replace('.', '')[:2]
    #
    #         # Value
    #         value_elem = ET.SubElement(item, 'Value')
    #         ET.SubElement(value_elem, 'Amount').text = f"{row.get('value_usd', 0):.2f}"
    #         ET.SubElement(value_elem, 'CurrencyCode').text = 'USD'
    #
    #         # Quantities
    #         quantities = ET.SubElement(item, 'Quantities')
    #         qty_unit = row.get('qty_unit', '')
    #         if qty_unit:
    #             ET.SubElement(quantities, 'UnitOfMeasure').text = qty_unit
    #         qty1 = row.get('qty1', '')
    #         if qty1:
    #             ET.SubElement(quantities, 'Quantity1').text = str(qty1)
    #         qty2 = row.get('qty2', '')
    #         if qty2:
    #             ET.SubElement(quantities, 'Quantity2').text = str(qty2)
    #
    #         # Manufacturer ID
    #         mid = row.get('mid', '')
    #         if mid:
    #             ET.SubElement(item, 'ManufacturerID').text = mid
    #
    #         # Section 232 information
    #         dec_type_cd = row.get('dec_type_cd', '')
    #         status_232 = row.get('status_232', '')
    #
    #         if dec_type_cd or status_232:
    #             section232 = ET.SubElement(item, 'Section232')
    #
    #             if dec_type_cd:
    #                 ET.SubElement(section232, 'DeclarationTypeCode').text = dec_type_cd
    #
    #             if status_232:
    #                 ET.SubElement(section232, 'MaterialStatus').text = status_232
    #
    #             # Country of origin information (for Section 232 materials)
    #             country_melt = row.get('country_of_melt', '')
    #             country_cast = row.get('country_of_cast', '')
    #             country_smelt = row.get('country_of_smelt', '')
    #             dec_flag = row.get('declaration_flag', '')
    #
    #             if country_melt or country_cast or country_smelt:
    #                 origin = ET.SubElement(section232, 'CountryOfOrigin')
    #                 if country_melt:
    #                     ET.SubElement(origin, 'CountryOfMelt').text = country_melt
    #                 if country_cast:
    #                     ET.SubElement(origin, 'CountryOfCast').text = country_cast
    #                 if country_smelt:
    #                     ET.SubElement(origin, 'PrimaryCountryOfSmelt').text = country_smelt
    #
    #             if dec_flag:
    #                 ET.SubElement(section232, 'DeclarationFlag').text = dec_flag
    #
    #     # Convert to pretty-printed XML string
    #     xml_str = ET.tostring(root, encoding='unicode')
    #     # Use minidom for pretty printing
    #     dom = minidom.parseString(xml_str)
    #     pretty_xml = dom.toprettyxml(indent='  ', encoding=None)
    #
    #     # Remove the XML declaration line that minidom adds (we'll add our own)
    #     lines = pretty_xml.split('\n')
    #     if lines[0].startswith('<?xml'):
    #         lines = lines[1:]
    #
    #     # Add proper XML declaration
    #     xml_declaration = '<?xml version="1.0" encoding="UTF-8"?>'
    #     return xml_declaration + '\n' + '\n'.join(lines)

    def log_context_menu(self, pos):
        menu = QMenu()
        copy_action = menu.addAction("Copy")
        action = menu.exec_(self.log_text.mapToGlobal(pos))
        if action == copy_action:
            self.log_text.copy()

    def copy_log_to_clipboard(self):
        QApplication.clipboard().setText(self.log_text.toPlainText())
        QMessageBox.information(self, "Copied", "Log copied to clipboard!")

    def update_log(self):
        self.log_text.setPlainText(logger.get_logs())
        sb = self.log_text.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _install_preview_shortcuts(self):
        # Install Ctrl+B on the preview table to toggle bold on selected cells
        try:
            self._bold_shortcut = QShortcut(QKeySequence("Ctrl+B"), self.table)
            self._bold_shortcut.setContext(Qt.WidgetWithChildrenShortcut)
            self._bold_shortcut.activated.connect(self.toggle_preview_bold)
        except Exception:
            pass

    def toggle_preview_bold(self):
        items = self.table.selectedItems()
        if not items:
            return
        # Toggle based on the first selected item's current bold state
        target_bold = not items[0].font().bold()
        for it in items:
            f = it.font()
            f.setBold(target_bold)
            it.setFont(f)

    def load_available_mids(self):
        try:
            # Preserve current selection before reloading
            current_selection = self.selected_mid

            conn = sqlite3.connect(str(DB_PATH))
            # Load from mid_table - includes manufacturer name, customer_id, related_parties
            df = pd.read_sql("""
                SELECT mid, manufacturer_name, customer_id, related_parties
                FROM mid_table
                WHERE mid IS NOT NULL AND mid != ''
                ORDER BY mid
            """, conn)
            conn.close()

            # Store MID data for lookup (mid -> {manufacturer_name, customer_id, related_parties})
            self.mid_data = {}
            self.available_mids = []

            for _, row in df.iterrows():
                mid = row['mid']
                manufacturer_name = row['manufacturer_name'] or ""
                customer_id = row['customer_id'] or ""
                related_parties = row['related_parties'] or "N"

                self.mid_data[mid] = {
                    'manufacturer_name': manufacturer_name,
                    'customer_id': customer_id,
                    'related_parties': related_parties
                }
                self.available_mids.append(mid)

            self.mid_combo.blockSignals(True)  # Prevent signal during reload
            self.mid_combo.clear()
            self.mid_combo.addItem("-- Select MID --")  # Placeholder item

            if self.available_mids:
                # Show only MID in dropdown (not manufacturer name)
                self.mid_combo.addItems(self.available_mids)

                # Restore previous selection if it exists
                if current_selection:
                    # Try to find by MID value
                    found_index = -1
                    for i, mid in enumerate(self.available_mids):
                        if mid == current_selection:
                            found_index = i + 1  # +1 for placeholder
                            break

                    if found_index >= 0:
                        self.mid_combo.setCurrentIndex(found_index)
                        self.selected_mid = current_selection
                    else:
                        self.mid_combo.setCurrentIndex(0)
                        self.selected_mid = ""
                else:
                    self.mid_combo.setCurrentIndex(0)  # Start with placeholder
                    self.selected_mid = ""  # No default selection

            self.mid_combo.blockSignals(False)
        except Exception as e:
            logger.error(f"MID load failed: {e}")

    def on_mid_changed(self, text):
        """Handle MID selection change"""
        if text and text != "-- Select MID --":
            self.selected_mid = text
        else:
            self.selected_mid = ""

    def update_export_invoice_total(self):
        """Update the invoice total display when a file is selected"""
        selected_items = self.exports_list.selectedItems()
        if not selected_items:
            self.export_invoice_total.setText("")
            return

        filename = selected_items[0].text()
        filepath = OUTPUT_DIR / filename

        try:
            # Read the Excel file and sum the ValueUSD column
            df = pd.read_excel(filepath, engine='openpyxl')

            # Try to find the value column (might be named differently based on mapping)
            value_col = None
            for col in df.columns:
                if 'value' in col.lower() or 'usd' in col.lower():
                    value_col = col
                    break

            if value_col:
                total = pd.to_numeric(df[value_col], errors='coerce').sum()
                self.export_invoice_total.setText(f"${total:,.2f}")
            else:
                self.export_invoice_total.setText("N/A")
        except Exception as e:
            logger.debug(f"Could not read invoice total from {filename}: {e}")
            self.export_invoice_total.setText("Error")

    def refresh_exported_files(self):
        """Load and display exported files from Output folder"""
        try:
            # Remember current selection and focus state
            current_item = self.exports_list.currentItem()
            current_file = current_item.text() if current_item else None
            had_focus = self.exports_list.hasFocus()

            # Block signals during refresh to prevent triggering events
            self.exports_list.blockSignals(True)
            self.exports_list.clear()
            if OUTPUT_DIR.exists():
                # Include all xlsx files (Upload_Sheet_* and split-by-invoice files)
                files = sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
                for f in files[:20]:  # Show last 20 files
                    self.exports_list.addItem(f.name)

            # Restore selection if the file still exists
            if current_file:
                items = self.exports_list.findItems(current_file, Qt.MatchExactly)
                if items:
                    self.exports_list.setCurrentItem(items[0])
            # If widget had focus but no previous selection, select first item
            elif had_focus and self.exports_list.count() > 0:
                self.exports_list.setCurrentRow(0)

            # Unblock signals after refresh complete
            self.exports_list.blockSignals(False)
        except Exception as e:
            logger.error(f"Refresh exports failed: {e}")
            # Make sure to unblock signals even on error
            self.exports_list.blockSignals(False)

    def refresh_input_files(self):
        """Load and display CSV/Excel files from Input folder"""
        try:
            # Remember current selection and focus state
            current_item = self.input_files_list.currentItem()
            current_file = current_item.text() if current_item else None
            had_focus = self.input_files_list.hasFocus()

            # Block signals during refresh to prevent triggering events
            self.input_files_list.blockSignals(True)
            self.input_files_list.clear()
            if INPUT_DIR.exists():
                # Combine CSV, XLSX, and XLS files
                csv_files = list(INPUT_DIR.glob("*.csv"))
                xlsx_files = list(INPUT_DIR.glob("*.xlsx"))
                xls_files = list(INPUT_DIR.glob("*.xls"))
                files = sorted(csv_files + xlsx_files + xls_files, key=lambda p: p.stat().st_mtime, reverse=True)
                for f in files[:50]:  # Show up to 50 files
                    self.input_files_list.addItem(f.name)
            
            # Restore selection if the file still exists
            if current_file:
                items = self.input_files_list.findItems(current_file, Qt.MatchExactly)
                if items:
                    self.input_files_list.setCurrentItem(items[0])
            # If widget had focus but no previous selection, select first item
            elif had_focus and self.input_files_list.count() > 0:
                self.input_files_list.setCurrentRow(0)

            # Unblock signals after refresh complete
            self.input_files_list.blockSignals(False)
        except Exception as e:
            logger.error(f"Refresh input files failed: {e}")
            # Make sure to unblock signals even on error
            self.input_files_list.blockSignals(False)

    def delete_selected_input_file(self):
        """Delete the selected input file after confirmation."""
        current_item = self.input_files_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No Selection", "Please select a file to delete.")
            return

        file_name = current_item.text()
        file_path = INPUT_DIR / file_name

        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete:\n\n{file_name}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                if file_path.exists():
                    file_path.unlink()
                    logger.info(f"Deleted input file: {file_name}")
                    # Clear current file if it was the deleted one
                    if hasattr(self, 'current_csv') and self.current_csv == str(file_path):
                        self.current_csv = None
                        self.file_label.setText("No file selected")
                        self.process_btn.setEnabled(False)
                    # Refresh the list
                    self.refresh_input_files()
            except Exception as e:
                logger.error(f"Failed to delete file: {e}")
                QMessageBox.critical(self, "Delete Failed", f"Could not delete file:\n{e}")

    def load_selected_input_file(self, item):
        """Load the selected CSV file from Input folder"""
        # Check if a map profile is selected first
        current_profile = self.profile_combo.currentText()
        if not current_profile or current_profile == "-- Select Profile --":
            QMessageBox.warning(self, "No Profile", "Please select a Map Profile first.")
            # Re-enable input fields after modal dialog
            self._enable_input_fields()
            return
        
        try:
            file_path = INPUT_DIR / item.text()
            if file_path.exists():
                self.current_csv = str(file_path)
                self.file_label.setText(file_path.name)
                
                # Clear previous processing state when loading new file
                self.last_processed_df = None
                self.table.setRowCount(0)

                # Get header row value from profile or input field
                header_row = 0  # Default: first row is header
                # First check if there's a profile header row loaded
                if hasattr(self, 'profile_header_row') and self.profile_header_row:
                    header_row = max(0, self.profile_header_row - 1)
                # Otherwise check input field (for Invoice Mapping Profiles tab)
                elif hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
                    try:
                        header_row_value = int(self.header_row_input.text().strip())
                        header_row = max(0, header_row_value - 1)
                    except ValueError:
                        header_row = 0

                # Read total value - handle both CSV and Excel files
                col_map = {v: k for k, v in self.shipment_mapping.items()}
                if file_path.suffix.lower() == '.xlsx':
                    df = pd.read_excel(file_path, dtype=str, header=header_row)
                else:
                    df = pd.read_csv(file_path, dtype=str, header=header_row)

                # DEBUG: Log DataFrame info
                logger.info(f"[CSV TOTAL DEBUG] DataFrame columns after reading with header={header_row}: {df.columns.tolist()}")
                logger.info(f"[CSV TOTAL DEBUG] Shipment mapping: {self.shipment_mapping}")

                # Calculate total using original column name before renaming
                # Only sum rows that have a part number to exclude total/subtotal rows
                value_column = None
                part_number_column = None

                if 'value_usd' in self.shipment_mapping:
                    original_col_name = self.shipment_mapping['value_usd']
                    logger.info(f"[CSV TOTAL DEBUG] Looking for column '{original_col_name}' in DataFrame")
                    logger.info(f"[CSV TOTAL DEBUG] Column exists: {original_col_name in df.columns}")
                    if original_col_name in df.columns:
                        value_column = original_col_name
                        logger.info(f"[CSV TOTAL DEBUG] First 5 values in {original_col_name}: {df[original_col_name].head().tolist()}")
                else:
                    logger.warning("[CSV TOTAL DEBUG] 'value_usd' not found in shipment_mapping")

                # Get part number column to filter rows
                if 'part_number' in self.shipment_mapping:
                    part_number_col_name = self.shipment_mapping['part_number']
                    if part_number_col_name in df.columns:
                        part_number_column = part_number_col_name
                        logger.info(f"[CSV TOTAL DEBUG] Part number column: '{part_number_column}'")

                if value_column:
                    # Filter to only rows that have a part number (exclude total/subtotal rows)
                    if part_number_column:
                        df_filtered = df[df[part_number_column].notna() & (df[part_number_column].astype(str).str.strip() != '')]
                        logger.info(f"[CSV TOTAL DEBUG] Rows with part numbers: {len(df_filtered)} of {len(df)} total rows")
                        total = pd.to_numeric(df_filtered[value_column], errors='coerce').sum()
                    else:
                        # If no part number column, sum all rows (old behavior)
                        logger.warning("[CSV TOTAL DEBUG] No part_number column mapped, summing all rows")
                        total = pd.to_numeric(df[value_column], errors='coerce').sum()

                    self.csv_total_value = round(total, 2)
                    logger.info(f"[CSV TOTAL DEBUG] Calculated total: ${self.csv_total_value:,.2f}")

                    # Update the invoice check - this will display the total and control button state
                    self.update_invoice_check()
                else:
                    logger.warning("[CSV TOTAL DEBUG] No value column found, CSV total will be 0.00")

                # Rename columns for other uses
                df = df.rename(columns=col_map)

                self.bottom_status.setText(f"Loaded: {file_path.name}")
                logger.info(f"Loaded: {file_path.name}")

                # Ensure input fields remain editable after loading
                self._enable_input_fields()

                # CRITICAL FIX: Modal dialogs fix the keyboard lock by triggering Qt event processing
                # When "No Profile" dialog is dismissed, _enable_input_fields() is called
                # and the modal event loop refresh fixes keyboard routing
                # Simulate the same effect here by processing events + enabling fields
                QApplication.processEvents()  # Process any pending Qt events
                self._enable_input_fields()  # Re-enable fields (same as after dialog dismissal)

                # Force widget style refresh (same as reselecting profile)
                for widget in [self.ci_input, self.wt_input]:
                    widget.style().unpolish(widget)
                    widget.style().polish(widget)

                # Move keyboard focus to CI input
                self.input_files_list.clearFocus()  # Remove focus from list
                self.ci_input.setFocus(Qt.OtherFocusReason)  # Give focus to CI input
                logger.info(f"[FOCUS] ci_input.hasFocus()={self.ci_input.hasFocus()}")
        except Exception as e:
            logger.error(f"Load input file failed: {e}")
            self.status.setText(f"Error loading file: {e}")
            QMessageBox.warning(self, "Error", f"Could not load file:\n{e}")
            # Ensure fields stay enabled even after error
            self._enable_input_fields()

    def open_exported_file(self, item):
        """Open the selected exported file with user's preferred application"""
        try:
            file_path = OUTPUT_DIR / item.text()
            if file_path.exists():
                import subprocess
                if sys.platform == 'win32':
                    os.startfile(str(file_path))
                elif sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', str(file_path)])
                else:  # Linux and other Unix-like systems
                    # Check user preference for Excel viewer (per-user setting)
                    viewer_preference = get_user_setting('excel_viewer', 'System Default')

                    if viewer_preference == "Gnumeric":
                        subprocess.run(['gnumeric', str(file_path)])
                    else:
                        subprocess.run(['xdg-open', str(file_path)])
        except Exception as e:
            logger.error(f"Open file failed: {e}")
            QMessageBox.warning(self, "Error", f"Could not open file:\n{e}")

    def setup_auto_refresh(self):
        """Set up automatic refresh timers for file lists"""
        # Track last modification times to avoid unnecessary refreshes
        self.last_input_mtime = 0
        self.last_output_mtime = 0
        
        # Auto-refresh input files every 10 seconds
        self.input_refresh_timer = QTimer(self)
        self.input_refresh_timer.timeout.connect(self.refresh_input_files_light)
        self.input_refresh_timer.start(10000)  # 10000ms = 10 seconds
        
        # Auto-refresh exported files every 10 seconds
        self.export_refresh_timer = QTimer(self)
        self.export_refresh_timer.timeout.connect(self.refresh_exported_files_light)
        self.export_refresh_timer.start(10000)  # 10000ms = 10 seconds
        
        # Clean up old exported files every 30 minutes
        self.cleanup_timer = QTimer(self)
        self.cleanup_timer.timeout.connect(self.cleanup_old_exports)
        self.cleanup_timer.start(1800000)  # 1800000ms = 30 minutes
        
        # Run cleanup once on startup
        QTimer.singleShot(5000, self.cleanup_old_exports)  # Wait 5 seconds after startup

        # Auto-refresh OCRMill input files every 10 seconds
        self.ocrmill_input_refresh_timer = QTimer(self)
        self.ocrmill_input_refresh_timer.timeout.connect(self.ocrmill_refresh_input_files_light)
        self.ocrmill_input_refresh_timer.start(10000)  # 10000ms = 10 seconds

        # Auto-refresh OCRMill output files every 10 seconds
        self.ocrmill_output_refresh_timer = QTimer(self)
        self.ocrmill_output_refresh_timer.timeout.connect(self.ocrmill_refresh_output_files_light)
        self.ocrmill_output_refresh_timer.start(10000)  # 10000ms = 10 seconds

        logger.info("Auto-refresh enabled for file lists (10 second interval)")
    
    def refresh_input_files_light(self):
        """Lightweight refresh - only update if on Process Shipment tab"""
        try:
            # Only refresh if on Process Shipment tab (tab index 0)
            if self.tabs.currentIndex() != 0:
                return

            if not INPUT_DIR.exists():
                return

            # Always refresh - directory mtime is unreliable on network drives
            self.refresh_input_files()
        except:
            pass  # Silently ignore errors during auto-refresh

    def refresh_exported_files_light(self):
        """Lightweight refresh - only update if on Process Shipment tab"""
        try:
            # Only refresh if on Process Shipment tab (tab index 0)
            if self.tabs.currentIndex() != 0:
                return

            if not OUTPUT_DIR.exists():
                return

            # Always refresh - directory mtime is unreliable on network drives
            self.refresh_exported_files()
        except:
            pass  # Silently ignore errors during auto-refresh

    def ocrmill_refresh_input_files_light(self):
        """Lightweight refresh - only update if on OCRMill tab's Invoice Processing sub-tab"""
        try:
            # Only refresh if on OCRMill tab (tab index 2)
            if self.tabs.currentIndex() != 2:
                return

            # Only refresh if on Invoice Processing sub-tab (tab index 0)
            if hasattr(self, 'ocrmill_tabs') and self.ocrmill_tabs.currentIndex() != 0:
                return

            if not self.ocrmill_config.input_folder.exists():
                return

            # Always refresh - directory mtime is unreliable on network drives
            self.ocrmill_refresh_input_files()
        except:
            pass  # Silently ignore errors during auto-refresh

    def ocrmill_refresh_output_files_light(self):
        """Lightweight refresh - only update if on OCRMill tab's Invoice Processing sub-tab"""
        try:
            # Only refresh if on OCRMill tab (tab index 2)
            if self.tabs.currentIndex() != 2:
                return

            # Only refresh if on Invoice Processing sub-tab (tab index 0)
            if hasattr(self, 'ocrmill_tabs') and self.ocrmill_tabs.currentIndex() != 0:
                return

            if not self.ocrmill_config.output_folder.exists():
                return

            # Always refresh - directory mtime is unreliable on network drives
            self.ocrmill_refresh_output_files()
        except:
            pass  # Silently ignore errors during auto-refresh

    def cleanup_old_exports(self):
        """Move exported files older than 3 days to Output_Processed directory"""
        try:
            if not OUTPUT_DIR.exists():
                return

            # Get the processed directory for output (sibling folder named {OutputDir}_Processed)
            output_processed_dir = get_processed_dir(OUTPUT_DIR)
            output_processed_dir.mkdir(exist_ok=True)

            from datetime import datetime, timedelta
            cutoff_date = datetime.now() - timedelta(days=3)
            moved_count = 0

            # Process all .xlsx files in Output directory
            for file_path in OUTPUT_DIR.glob("*.xlsx"):
                try:
                    # Skip if it's a directory
                    if file_path.is_dir():
                        continue

                    # Get file modification time
                    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)

                    # Move if older than 3 days
                    if file_mtime < cutoff_date:
                        dest_path = output_processed_dir / file_path.name

                        # Handle duplicate filenames
                        if dest_path.exists():
                            base_name = file_path.stem
                            ext = file_path.suffix
                            counter = 1
                            while dest_path.exists():
                                dest_path = output_processed_dir / f"{base_name}_{counter}{ext}"
                                counter += 1

                        # Move the file
                        shutil.move(str(file_path), str(dest_path))
                        moved_count += 1
                        logger.info(f"Moved old export to Processed: {file_path.name}")
                except Exception as e:
                    logger.warning(f"Failed to move {file_path.name}: {e}")

            if moved_count > 0:
                logger.info(f"Cleanup: Moved {moved_count} exported file(s) older than 3 days to {output_processed_dir}")
                # Refresh the exported files list if we're on the Process Shipment tab
                if self.tabs.currentIndex() == 0:
                    self.refresh_exported_files()
        except Exception as e:
            logger.error(f"Cleanup old exports failed: {e}")

if __name__ == "__main__":
    # Prevent PyInstaller multiprocessing from spawning console windows on Windows
    import multiprocessing
    multiprocessing.freeze_support()

    # Update native splash
    update_splash("Checking for updates...")

    # Check for self-update scenario (exe running from different location than installed)
    if check_and_perform_self_update():
        sys.exit(0)  # Update initiated or user declined, exit

    # Save current location as install path (for future update detection)
    if getattr(sys, 'frozen', False):
        save_installed_path(Path(sys.executable))

    update_splash("Starting application...")

    import traceback
    app = QApplication(sys.argv)

    try:
        # Theme will be set by apply_saved_theme() during initialization
        icon_path = TEMP_RESOURCES_DIR / "tariffmill_icon_hybrid_2.svg"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "icon.ico"
        if icon_path.exists():
            app.setWindowIcon(QIcon(str(icon_path)))
        
        # Create and show splash screen with spinner
        # Custom spinner widget
        class SpinnerWidget(QWidget):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.angle = 0
                self.setFixedSize(50, 50)
                self.timer = QTimer(self)
                self.timer.timeout.connect(self.rotate)
                self.timer.start(50)  # Update every 50ms

            def rotate(self):
                self.angle = (self.angle + 30) % 360
                self.update()

            def paintEvent(self, event):
                from PyQt5.QtGui import QPainter, QPen, QColor
                painter = QPainter(self)
                painter.setRenderHint(QPainter.Antialiasing)
                painter.translate(self.width() / 2, self.height() / 2)
                painter.rotate(self.angle)

                # Draw spinning dots
                for i in range(12):
                    alpha = int(255 * (i + 1) / 12)
                    painter.setBrush(QColor(0, 120, 212, alpha))
                    painter.setPen(Qt.NoPen)
                    painter.drawEllipse(-3, -20, 6, 6)
                    painter.rotate(30)

            def stop(self):
                self.timer.stop()

        splash_widget = QWidget()
        splash_widget.setFixedSize(500, 320)
        splash_widget.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        splash_widget.setAttribute(Qt.WA_TranslucentBackground)
        splash_layout = QVBoxLayout(splash_widget)
        splash_layout.setContentsMargins(0, 0, 0, 0)
        splash_container = QWidget()
        splash_container.setStyleSheet("""
            QWidget {
                background-color: #333333;
                border: 3px solid #0078D4;
                border-radius: 15px;
            }
        """)
        container_layout = QVBoxLayout(splash_container)
        container_layout.setContentsMargins(30, 30, 30, 30)
        container_layout.setSpacing(15)
        title_label = QLabel(f"<h1 style='color: #0078D4;'>{APP_NAME}</h1>")
        title_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(title_label)

        # Spinner
        spinner = SpinnerWidget()
        spinner_container = QWidget()
        spinner_layout = QHBoxLayout(spinner_container)
        spinner_layout.setContentsMargins(0, 0, 0, 0)
        spinner_layout.addStretch()
        spinner_layout.addWidget(spinner)
        spinner_layout.addStretch()
        container_layout.addWidget(spinner_container)

        splash_message = QLabel("Initializing application...")
        splash_message.setStyleSheet("color: #f3f3f3; font-size: 11pt;")
        splash_message.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(splash_message)
        splash_progress = QProgressBar()
        splash_progress.setRange(0, 100)
        splash_progress.setValue(0)
        splash_progress.setTextVisible(True)
        splash_progress.setFormat("%p%")
        splash_progress.setFixedHeight(20)
        splash_progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #555;
                border-radius: 5px;
                background-color: #1e1e1e;
                text-align: center;
                color: white;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
                border-radius: 3px;
            }
        """)
        container_layout.addWidget(splash_progress)
        splash_layout.addWidget(splash_container)
        # Close the native PyInstaller splash before showing the Qt splash
        close_splash()

        splash_widget.show()
        screen_geo = app.desktop().availableGeometry()
        splash_widget.move(
            (screen_geo.width() - splash_widget.width()) // 2,
            (screen_geo.height() - splash_widget.height()) // 2
        )
        app.processEvents()

        logger.info("Application started")

        # ============================================================
        # AUTHENTICATION CHECK
        # ============================================================
        splash_message.setText("Checking authentication...")
        splash_progress.setValue(5)
        app.processEvents()

        # Initialize authentication manager
        auth_manager = AuthenticationManager(DB_PATH)

        # Try Windows domain authentication first
        windows_auth_success, windows_msg, _ = auth_manager.try_windows_auth()

        if windows_auth_success:
            # Windows auth successful - skip login dialog
            logger.info(f"Windows auth successful: {auth_manager.current_user}")
        else:
            # Windows auth failed - show login dialog
            logger.debug(f"Windows auth not available: {windows_msg}")

            # Hide splash temporarily to show login dialog
            splash_widget.hide()
            app.processEvents()

            # Show login dialog
            login_dialog = LoginDialog(auth_manager)
            login_dialog.setWindowIcon(app.windowIcon())

            if login_dialog.exec_() != QDialog.Accepted:
                # User cancelled login - exit application
                logger.info("User cancelled login, exiting application")
                spinner.stop()
                sys.exit(0)

            # Show splash again
            splash_widget.show()
            app.processEvents()

        # Authentication successful - continue with startup
        logger.info(f"User authenticated: {auth_manager.current_user} (role: {auth_manager.current_role}, method: {auth_manager.auth_method})")
        # ============================================================

        splash_message.setText("Creating main window...\nPlease wait...")
        splash_progress.setValue(10)
        app.processEvents()

        # Create main window but keep it completely hidden during initialization
        splash_message.setText("Creating main window...")
        splash_progress.setValue(20)
        app.processEvents()

        win = TariffMill()
        # Pass authentication manager to main window
        win.auth_manager = auth_manager
        win.hide()  # Explicitly hide immediately after creation
        win.setWindowTitle(APP_NAME)

        def finish_initialization():
            splash_message.setText("Loading configuration...")
            splash_progress.setValue(30)
            app.processEvents()
            win.load_config_paths()

            splash_message.setText("Applying theme...")
            splash_progress.setValue(40)
            app.processEvents()
            win.apply_saved_theme()

            splash_message.setText("Loading MIDs...")
            splash_progress.setValue(50)
            app.processEvents()
            win.load_available_mids()

            splash_message.setText("Loading profiles...")
            splash_progress.setValue(60)
            app.processEvents()
            win.load_mapping_profiles()
            win.load_folder_profiles()

            splash_message.setText("Restoring last session...")
            splash_progress.setValue(65)
            app.processEvents()
            win.restore_last_used_settings()

            splash_message.setText("Loading export profiles...")
            splash_progress.setValue(70)
            app.processEvents()
            win.load_output_mapping_profiles()

            splash_message.setText("Scanning input files...")
            splash_progress.setValue(80)
            app.processEvents()
            win.refresh_input_files()

            splash_message.setText("Starting services...")
            splash_progress.setValue(90)
            app.processEvents()
            win.setup_auto_refresh()

            # TODO: Re-enable license check when ready to sell
            # win.check_license_status()

            splash_message.setText("Ready!")
            splash_progress.setValue(100)
            app.processEvents()

            # Now close splash and show main window
            spinner.stop()
            splash_widget.close()
            # Move window to center of screen before showing
            screen_geo = app.primaryScreen().availableGeometry()
            win.move(
                (screen_geo.width() - win.width()) // 2,
                (screen_geo.height() - win.height()) // 2
            )
            # Clear the hidden attribute before showing
            win.setAttribute(Qt.WA_DontShowOnScreen, False)
            win.show()
            win.raise_()
            win.activateWindow()
            win.status.setText("Ready")

            # Update account menu to show logged-in user
            win._update_account_menu()

            # Final aggressive enable after all initialization
            QTimer.singleShot(0, win._enable_input_fields)
            QTimer.singleShot(100, win._enable_input_fields)
            QTimer.singleShot(500, win._enable_input_fields)
            QTimer.singleShot(1000, win._enable_input_fields)
            # Check for updates after a short delay (non-blocking)
            QTimer.singleShot(2000, win.check_for_updates_startup)

        # Start initialization after a brief delay to let splash render
        QTimer.singleShot(50, finish_initialization)
        sys.exit(app.exec_())
    except Exception as e:
        error_msg = f"Unhandled Exception:\n{str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        try:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(None, "Application Error", error_msg)
        except Exception:
            pass
        sys.exit(1)
