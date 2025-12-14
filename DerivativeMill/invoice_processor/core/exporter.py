"""
Excel export functionality with Section 232 styling.

This module handles exporting processed invoice data to Excel format
with color-coded rows based on material type and Section 301 indicators.
"""

import pandas as pd
from pathlib import Path
from typing import Optional, Dict, List, Union
from dataclasses import dataclass, field


@dataclass
class ExportStyle:
    """Configuration for Excel export styling."""

    # Font settings
    font_name: str = 'Arial'
    font_size: int = 11
    default_font_color: str = '#000000'

    # Material type colors (hex format)
    steel_color: str = '#4a4a4a'
    aluminum_color: str = '#6495ED'
    copper_color: str = '#B87333'
    wood_color: str = '#8B4513'
    auto_color: str = '#2F4F4F'
    non232_color: str = '#FF0000'

    # Section 301 highlight
    sec301_fill_color: str = '#FFCC99'

    # Page setup
    landscape: bool = True
    fit_to_width: bool = True
    auto_size_columns: bool = True


@dataclass
class ExportResult:
    """Result of an export operation."""

    success: bool
    file_path: Optional[Path] = None
    row_count: int = 0
    error: Optional[str] = None
    files_created: List[Path] = field(default_factory=list)

    def __repr__(self):
        if self.success:
            return f"ExportResult(success=True, rows={self.row_count}, file={self.file_path})"
        return f"ExportResult(success=False, error={self.error})"


def export_to_excel(
    df: pd.DataFrame,
    output_path: Union[str, Path],
    columns: Optional[List[str]] = None,
    style: Optional[ExportStyle] = None,
    material_column: str = '_232_flag',
    sec301_column: str = 'Sec301_Exclusion_Tariff'
) -> ExportResult:
    """
    Export DataFrame to Excel with Section 232 styling.

    This function creates a formatted Excel file with:
    - Color-coded rows based on material type (steel, aluminum, copper, wood, auto, non-232)
    - Orange background for rows with Section 301 exclusion tariffs
    - Landscape orientation and fit-to-width page setup
    - Auto-sized columns

    Args:
        df: DataFrame to export
        output_path: Path for the output Excel file
        columns: List of column names to export. If None, exports all columns.
        style: ExportStyle configuration. If None, uses defaults.
        material_column: Column name containing material type flags
        sec301_column: Column name containing Section 301 exclusion tariff values

    Returns:
        ExportResult with success status and file information

    Example:
        >>> from invoice_processor.core import export_to_excel, ExportStyle
        >>> style = ExportStyle(steel_color='#333333', aluminum_color='#0066CC')
        >>> result = export_to_excel(processed_df, 'output.xlsx', style=style)
        >>> if result.success:
        ...     print(f"Exported {result.row_count} rows to {result.file_path}")
    """
    try:
        from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
    except ImportError:
        return ExportResult(
            success=False,
            error="openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    if df.empty:
        return ExportResult(success=False, error="DataFrame is empty")

    output_path = Path(output_path)
    style = style or ExportStyle()

    # Determine columns to export
    if columns is None:
        columns = list(df.columns)
    else:
        # Filter to columns that exist in the DataFrame
        columns = [c for c in columns if c in df.columns]

    if not columns:
        return ExportResult(success=False, error="No valid columns to export")

    # Create material type masks
    def get_material_mask(flag_value: str) -> List[int]:
        if material_column not in df.columns:
            return []
        mask = df[material_column].astype(str).str.contains(flag_value, case=False, na=False)
        return [i for i, val in enumerate(mask.tolist()) if val]

    steel_indices = get_material_mask('Steel')
    aluminum_indices = get_material_mask('Aluminum')
    copper_indices = get_material_mask('Copper')
    wood_indices = get_material_mask('Wood')
    auto_indices = get_material_mask('Auto')
    non232_indices = get_material_mask('Non_232')

    # Section 301 mask
    sec301_indices = []
    if sec301_column in df.columns:
        sec301_mask = df[sec301_column].apply(
            lambda x: pd.notna(x) and str(x).strip() != ''
        )
        sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

    # Create fonts for each material type
    def hex_to_argb(hex_color: str) -> str:
        """Convert hex color to ARGB format for openpyxl."""
        return '00' + hex_color.lstrip('#').upper()

    fonts = {
        'steel': ExcelFont(name=style.font_name, size=style.font_size,
                          color=hex_to_argb(style.steel_color)),
        'aluminum': ExcelFont(name=style.font_name, size=style.font_size,
                             color=hex_to_argb(style.aluminum_color)),
        'copper': ExcelFont(name=style.font_name, size=style.font_size,
                           color=hex_to_argb(style.copper_color)),
        'wood': ExcelFont(name=style.font_name, size=style.font_size,
                         color=hex_to_argb(style.wood_color)),
        'auto': ExcelFont(name=style.font_name, size=style.font_size,
                         color=hex_to_argb(style.auto_color)),
        'non232': ExcelFont(name=style.font_name, size=style.font_size,
                           color=hex_to_argb(style.non232_color)),
        'default': ExcelFont(name=style.font_name, size=style.font_size,
                            color=hex_to_argb(style.default_font_color)),
        'header': ExcelFont(name=style.font_name, size=style.font_size, bold=True,
                           color=hex_to_argb(style.default_font_color)),
    }

    sec301_fill = PatternFill(
        start_color=style.sec301_fill_color.lstrip('#'),
        end_color=style.sec301_fill_color.lstrip('#'),
        fill_type="solid"
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df[columns].to_excel(writer, index=False)
            ws = next(iter(writer.sheets.values()))

            # Apply header formatting
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = fonts['header']
                cell.alignment = center_alignment

            # Apply row formatting
            for row_idx in range(len(df)):
                row_num = row_idx + 2  # Excel rows are 1-indexed, plus header
                is_sec301 = row_idx in sec301_indices

                # Determine font based on material type
                if row_idx in steel_indices:
                    row_font = fonts['steel']
                elif row_idx in aluminum_indices:
                    row_font = fonts['aluminum']
                elif row_idx in copper_indices:
                    row_font = fonts['copper']
                elif row_idx in wood_indices:
                    row_font = fonts['wood']
                elif row_idx in auto_indices:
                    row_font = fonts['auto']
                elif row_idx in non232_indices:
                    row_font = fonts['non232']
                else:
                    row_font = fonts['default']

                # Apply formatting to each cell in the row
                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = row_font
                    cell.alignment = center_alignment
                    if is_sec301:
                        cell.fill = sec301_fill

            # Auto-size columns
            if style.auto_size_columns:
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max(max_length + 2, 8)  # Minimum width of 8
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Page setup
            if style.landscape:
                ws.page_setup.orientation = 'landscape'
            if style.fit_to_width:
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0  # Unlimited pages vertically

        return ExportResult(
            success=True,
            file_path=output_path,
            row_count=len(df),
            files_created=[output_path]
        )

    except Exception as e:
        return ExportResult(success=False, error=str(e))


def export_split_by_invoice(
    df: pd.DataFrame,
    output_dir: Union[str, Path],
    invoice_column: str = 'invoice_number',
    filename_prefix: str = 'invoice_',
    columns: Optional[List[str]] = None,
    style: Optional[ExportStyle] = None,
    material_column: str = '_232_flag',
    sec301_column: str = 'Sec301_Exclusion_Tariff'
) -> ExportResult:
    """
    Export DataFrame split into separate Excel files by invoice number.

    Args:
        df: DataFrame to export
        output_dir: Directory for output files
        invoice_column: Column name containing invoice numbers
        filename_prefix: Prefix for output filenames
        columns: List of column names to export. If None, exports all columns.
        style: ExportStyle configuration. If None, uses defaults.
        material_column: Column name containing material type flags
        sec301_column: Column name containing Section 301 exclusion tariff values

    Returns:
        ExportResult with success status and list of created files

    Example:
        >>> result = export_split_by_invoice(df, 'output/', invoice_column='InvoiceNo')
        >>> for file in result.files_created:
        ...     print(f"Created: {file}")
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if invoice_column not in df.columns:
        return ExportResult(
            success=False,
            error=f"Invoice column '{invoice_column}' not found in DataFrame"
        )

    invoice_numbers = df[invoice_column].dropna().unique()
    if len(invoice_numbers) == 0:
        return ExportResult(success=False, error="No invoice numbers found")

    files_created = []
    total_rows = 0
    errors = []

    for invoice_num in invoice_numbers:
        invoice_df = df[df[invoice_column] == invoice_num]
        safe_invoice = str(invoice_num).replace('/', '_').replace('\\', '_')
        output_path = output_dir / f"{filename_prefix}{safe_invoice}.xlsx"

        result = export_to_excel(
            invoice_df,
            output_path,
            columns=columns,
            style=style,
            material_column=material_column,
            sec301_column=sec301_column
        )

        if result.success:
            files_created.append(result.file_path)
            total_rows += result.row_count
        else:
            errors.append(f"{invoice_num}: {result.error}")

    if files_created:
        return ExportResult(
            success=True,
            file_path=output_dir,
            row_count=total_rows,
            files_created=files_created,
            error='; '.join(errors) if errors else None
        )
    else:
        return ExportResult(
            success=False,
            error='; '.join(errors) if errors else "No files created"
        )
