#!/usr/bin/env python3
# ==============================================================================
# Derivative Mill - Customs Documentation Processing System
# ==============================================================================
# Professional PyQt5 application for automating invoice processing, parts
# management, and Section 232 tariff compliance tracking.
#
# Features:
#   - Invoice processing with two-stage verification workflow
#   - Parts database management with HTS code mapping
#   - Section 232 tariff compliance calculations
#   - Excel export for customs documentation
#   - Invoice mapping profile management
#   - Real-time data validation and status feedback
# ==============================================================================

APP_NAME = "Derivative Mill"
DB_NAME = "derivativemill.db"

# Import version from version.py
try:
    from DerivativeMill.version import get_version
    VERSION = get_version()
except ImportError:
    try:
        from version import get_version
        VERSION = get_version()
    except ImportError:
        # Fallback if version.py is not available
        VERSION = "v0.6"


import sys
import os
import json
import time
import re
import shutil
import traceback
import subprocess
import configparser
import webbrowser
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime
from threading import Thread
import pandas as pd
import sqlite3
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QMimeData, pyqtSignal, QTimer, QSize, QEventLoop, QRect, QSettings
from PyQt5.QtGui import QColor, QFont, QDrag, QKeySequence, QIcon, QPixmap, QPainter, QDoubleValidator, QCursor, QPen, QTextCursor, QTextCharFormat
from PyQt5.QtSvg import QSvgRenderer
from openpyxl.styles import Font as ExcelFont, Alignment
import tempfile

# ==============================================================================
# Application Logger
# ==============================================================================
# In-memory logging system with timestamp, level, and message tracking.
# Maintains up to 1000 log entries for debugging and user feedback.
class ErrorLogger:
    """Centralized logging for application events, errors, and diagnostics."""

    def __init__(self):
        self.logs = []

    def log(self, level, message):
        """Record a log entry with timestamp and severity level."""
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        entry = f"[{ts}] {level.upper():7} | {message}"
        self.logs.append(entry)
        if len(self.logs) > 1000:
            self.logs = self.logs[-1000:]
        print(entry)

    def info(self, msg):
        self.log("info", msg)

    def debug(self, msg):
        self.log("debug", msg)

    def warning(self, msg):
        self.log("warning", msg)

    def error(self, msg):
        self.log("ERROR", msg)
        if hasattr(sys, 'exc_info') and sys.exc_info()[0]:
            tb = traceback.format_exc()
            for line in tb.splitlines():
                self.log("TRACE", line)

    def success(self, msg):
        self.log("success", msg)

    def get_logs(self):
        return "\n".join(self.logs)

logger = ErrorLogger()

# ==============================================================================
# Update Checker
# ==============================================================================
# Checks GitHub releases for new versions and provides update notifications.
# Uses the GitHub API to fetch the latest release information.

GITHUB_REPO = "royalpayne/DerivativeMill"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
GITHUB_RELEASES_URL = f"https://github.com/{GITHUB_REPO}/releases"

class UpdateChecker:
    """Checks GitHub for application updates"""
    
    def __init__(self, current_version):
        self.current_version = current_version
        self.latest_version = None
        self.release_url = None
        self.release_notes = None
        self.download_url = None
        self.error = None
    
    def parse_version(self, version_str):
        """Parse version string to tuple for comparison (e.g., 'v0.61.0' -> (0, 61, 0))"""
        try:
            # Remove 'v' prefix if present
            clean = version_str.lstrip('v').lstrip('V')
            parts = clean.split('.')
            return tuple(int(p) for p in parts[:3])  # Take first 3 parts
        except (ValueError, AttributeError):
            return (0, 0, 0)
    
    def check_for_updates(self):
        """
        Check GitHub for the latest release.
        Returns: (has_update, latest_version, release_url, release_notes, download_url, error)
        """
        try:
            # Create request with user-agent header (required by GitHub API)
            request = urllib.request.Request(
                GITHUB_API_URL,
                headers={'User-Agent': f'DerivativeMill/{self.current_version}'}
            )
            
            with urllib.request.urlopen(request, timeout=10) as response:
                data = json.loads(response.read().decode('utf-8'))
            
            self.latest_version = data.get('tag_name', '')
            self.release_url = data.get('html_url', GITHUB_RELEASES_URL)
            self.release_notes = data.get('body', 'No release notes available.')
            
            # Find Windows executable download URL from assets
            assets = data.get('assets', [])
            for asset in assets:
                name = asset.get('name', '').lower()
                if name.endswith('.exe') or 'windows' in name:
                    self.download_url = asset.get('browser_download_url')
                    break
            
            # Compare versions
            current_tuple = self.parse_version(self.current_version)
            latest_tuple = self.parse_version(self.latest_version)
            
            has_update = latest_tuple > current_tuple
            
            logger.info(f"Update check: current={self.current_version}, latest={self.latest_version}, update_available={has_update}")
            
            return (has_update, self.latest_version, self.release_url, 
                    self.release_notes, self.download_url, None)
            
        except urllib.error.URLError as e:
            self.error = f"Network error: {str(e)}"
            logger.warning(f"Update check failed: {self.error}")
            return (False, None, None, None, None, self.error)
        except json.JSONDecodeError as e:
            self.error = f"Invalid response from GitHub: {str(e)}"
            logger.warning(f"Update check failed: {self.error}")
            return (False, None, None, None, None, self.error)
        except Exception as e:
            self.error = f"Update check failed: {str(e)}"
            logger.warning(f"Update check failed: {self.error}")
            return (False, None, None, None, None, self.error)


# ==============================================================================
# Application Paths
# ==============================================================================
# Handles path resolution for both PyInstaller-bundled executables and
# development/script execution modes. Ensures resource files are found
# regardless of deployment method.

if getattr(sys, 'frozen', False):
    # Running as compiled executable (PyInstaller)
    BASE_DIR = Path(sys.executable).parent
    if hasattr(sys, '_MEIPASS'):
        TEMP_RESOURCES_DIR = Path(sys._MEIPASS) / "Resources"
    else:
        TEMP_RESOURCES_DIR = BASE_DIR / "Resources"
else:
    # Running as Python script
    BASE_DIR = Path(__file__).parent
    TEMP_RESOURCES_DIR = BASE_DIR / "Resources"

# Directory structure for application data
RESOURCES_DIR = BASE_DIR / "Resources"
INPUT_DIR = BASE_DIR / "Input"
OUTPUT_DIR = BASE_DIR / "Output"
PROCESSED_DIR = INPUT_DIR / "Processed"
OUTPUT_PROCESSED_DIR = OUTPUT_DIR / "Processed"

# Configuration files for data mappings
MAPPING_FILE = BASE_DIR / "column_mapping.json"
SHIPMENT_MAPPING_FILE = BASE_DIR / "shipment_mapping.json"

# Create required directories
for p in (RESOURCES_DIR, INPUT_DIR, OUTPUT_DIR, PROCESSED_DIR, OUTPUT_PROCESSED_DIR):
    p.mkdir(exist_ok=True)

# ==============================================================================
# Shared Configuration File
# ==============================================================================
# The config.ini file stores settings shared across all users (e.g., database path).
# User-specific settings remain in QSettings (Windows registry).

CONFIG_FILE = BASE_DIR / "config.ini"

def load_shared_config():
    """Load shared configuration from config.ini file."""
    config = configparser.ConfigParser()
    if CONFIG_FILE.exists():
        config.read(str(CONFIG_FILE))
    return config

def save_shared_config(config):
    """Save shared configuration to config.ini file."""
    with open(str(CONFIG_FILE), 'w') as f:
        config.write(f)

def get_database_path():
    """
    Get the database path from shared config or use default.
    
    Returns:
        Path object pointing to the SQLite database file.
    """
    config = load_shared_config()
    if config.has_option('Database', 'path'):
        custom_path = config.get('Database', 'path')
        if custom_path and Path(custom_path).exists():
            return Path(custom_path)
    # Default to local Resources folder
    return RESOURCES_DIR / DB_NAME

def set_database_path(path):
    """
    Set a custom database path in shared config.
    
    Args:
        path: Path string to the database file (can be network path).
    """
    config = load_shared_config()
    if not config.has_section('Database'):
        config.add_section('Database')
    config.set('Database', 'path', str(path))
    save_shared_config(config)

# Database location - reads from config.ini or defaults to local
DB_PATH = get_database_path()

# ==============================================================================
# Per-User Settings (QSettings - Windows Registry)
# ==============================================================================
# These settings are stored per-user in the Windows Registry under
# HKEY_CURRENT_USER\Software\DerivativeMill\DerivativeMill
# This allows each user to have their own personal preferences while
# sharing the same database for parts data, profiles, etc.

def get_user_settings():
    """Get QSettings object for per-user settings stored in Windows Registry."""
    return QSettings("DerivativeMill", "DerivativeMill")

def get_user_setting(key, default=None):
    """
    Get a per-user setting from Windows Registry.

    Args:
        key: Setting key (e.g., 'theme', 'font_size', 'column_widths')
        default: Default value if setting doesn't exist

    Returns:
        The stored value or default
    """
    settings = get_user_settings()
    return settings.value(key, default)

def set_user_setting(key, value):
    """
    Save a per-user setting to Windows Registry.

    Args:
        key: Setting key
        value: Value to store
    """
    settings = get_user_settings()
    settings.setValue(key, value)
    settings.sync()

def get_user_setting_bool(key, default=False):
    """Get a boolean per-user setting (handles string 'true'/'false' from registry)."""
    value = get_user_setting(key, default)
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ('true', '1', 'yes')
    return bool(value)

def get_user_setting_int(key, default=0):
    """Get an integer per-user setting."""
    value = get_user_setting(key, default)
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

def get_user_setting_float(key, default=0.0):
    """Get a float per-user setting."""
    value = get_user_setting(key, default)
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def is_widget_valid(widget):
    """
    Check if a Qt widget is still valid (not deleted).
    Returns False if the widget's underlying C++ object has been deleted.
    """
    try:
        import sip
        return widget is not None and not sip.isdeleted(widget)
    except ImportError:
        # If sip is not available, try accessing the widget
        try:
            if widget is None:
                return False
            # Try to access a property - will raise RuntimeError if deleted
            widget.objectName()
            return True
        except RuntimeError:
            return False

def get_232_info(hts_code):
    """
    Lookup Section 232 tariff information for an HTS code.

    Args:
        hts_code: HTS code string (with or without dots)

    Returns:
        Tuple of (material, declaration_code, smelt_flag) where:
        - material: Material type (e.g., "Steel", "Aluminum") or None
        - declaration_code: Tariff code (e.g., "08" for Steel)
        - smelt_flag: "Y" for materials requiring smelting declaration, "" otherwise

    Process:
        1. Queries tariff_232 table for 10-digit and 8-digit HTS matches
        2. Falls back to hardcoded HTS prefixes for common materials
        3. Returns None if material not found
    """
    if not hts_code:
        return None, "", ""

    # Normalize HTS code: remove dots, strip whitespace, convert to uppercase
    hts_clean = str(hts_code).replace(".", "").strip().upper()
    hts_8 = hts_clean[:8]
    hts_10 = hts_clean[:10]

    # Query tariff database
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("SELECT material, declaration_required FROM tariff_232 WHERE hts_code = ?", (hts_10,))
        row = c.fetchone()
        if not row and len(hts_clean) >= 8:
            c.execute("SELECT material, declaration_required FROM tariff_232 WHERE hts_code = ?", (hts_8,))
            row = c.fetchone()
        conn.close()
        if row:
            material = row[0]
            dec_code = row[1] if row[1] else ""
            dec_type = dec_code.split(" - ")[0] if " - " in dec_code else dec_code
            smelt_flag = "Y" if material in ["Aluminum", "Wood", "Copper"] else ""
            return material, dec_type, smelt_flag
    except Exception as e:
        logger.error(f"Error querying tariff_232 for HTS {hts_clean}: {e}")

    # No match found in tariff_232 database
    return None, "", ""

# ==============================================================================
# Database Initialization
# ==============================================================================
# Creates tables if they don't exist: parts_master, tariff_232, sec_232_actions,
# mid_table, mapping_profiles, and app_config.

def init_database():
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS parts_master (
            part_number TEXT PRIMARY KEY, description TEXT, hts_code TEXT, country_origin TEXT,
            mid TEXT, client_code TEXT, steel_ratio REAL DEFAULT 1.0, non_steel_ratio REAL DEFAULT 0.0, last_updated TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS tariff_232 (
            hts_code TEXT PRIMARY KEY,
            material TEXT,
            classification TEXT,
            chapter TEXT,
            chapter_description TEXT,
            declaration_required TEXT,
            notes TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS sec_232_actions (
            tariff_no TEXT PRIMARY KEY,
            action TEXT,
            description TEXT,
            advalorem_rate TEXT,
            effective_date TEXT,
            expiration_date TEXT,
            specific_rate TEXT,
            additional_declaration TEXT,
            note TEXT,
            link TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS mapping_profiles (
            profile_name TEXT PRIMARY KEY, mapping_json TEXT, created_date TEXT, header_row INTEGER DEFAULT 1
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS app_config (
            key TEXT PRIMARY KEY, value TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS output_column_mappings (
            profile_name TEXT PRIMARY KEY,
            mapping_json TEXT,
            created_date TEXT
        )""")

        # Migration: Add client_code column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'client_code' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN client_code TEXT")
                logger.info("Added client_code column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add client_code column: {e}")

        # Migration: Add header_row column to mapping_profiles if it doesn't exist
        try:
            c.execute("PRAGMA table_info(mapping_profiles)")
            columns = [col[1] for col in c.fetchall()]
            if 'header_row' not in columns:
                c.execute("ALTER TABLE mapping_profiles ADD COLUMN header_row INTEGER DEFAULT 1")
                logger.info("Added header_row column to mapping_profiles")
        except Exception as e:
            logger.warning(f"Failed to check/add header_row column: {e}")

        # Migration: Add cbp_qty1 column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'cbp_qty1' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN cbp_qty1 TEXT")
                logger.info("Added cbp_qty1 column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add cbp_qty1 column: {e}")

        # Migration: Add aluminum_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'aluminum_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN aluminum_ratio REAL DEFAULT 0.0")
                logger.info("Added aluminum_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add aluminum_ratio column: {e}")

        # Migration: Add copper_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'copper_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN copper_ratio REAL DEFAULT 0.0")
                logger.info("Added copper_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add copper_ratio column: {e}")

        # Migration: Add wood_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'wood_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN wood_ratio REAL DEFAULT 0.0")
                logger.info("Added wood_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add wood_ratio column: {e}")

        # Migration: Add auto_ratio column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'auto_ratio' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN auto_ratio REAL DEFAULT 0.0")
                logger.info("Added auto_ratio column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add auto_ratio column: {e}")

        # Migration: Add country_of_melt column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'country_of_melt' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN country_of_melt TEXT")
                logger.info("Added country_of_melt column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add country_of_melt column: {e}")

        # Migration: Add country_of_cast column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'country_of_cast' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN country_of_cast TEXT")
                logger.info("Added country_of_cast column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add country_of_cast column: {e}")

        # Migration: Add country_of_smelt column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'country_of_smelt' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN country_of_smelt TEXT")
                logger.info("Added country_of_smelt column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add country_of_smelt column: {e}")

        # Migration: Add Sec301_Exclusion_Tariff column to parts_master if it doesn't exist
        try:
            c.execute("PRAGMA table_info(parts_master)")
            columns = [col[1] for col in c.fetchall()]
            if 'Sec301_Exclusion_Tariff' not in columns:
                c.execute("ALTER TABLE parts_master ADD COLUMN Sec301_Exclusion_Tariff TEXT")
                logger.info("Added Sec301_Exclusion_Tariff column to parts_master")
        except Exception as e:
            logger.warning(f"Failed to check/add Sec301_Exclusion_Tariff column: {e}")

        # Migration: Reset column visibility settings if we have outdated settings
        # (This handles upgrades from versions with fewer columns)
        try:
            c.execute("SELECT COUNT(*) FROM app_config WHERE key LIKE 'preview_col_visible_%'")
            count_row = c.fetchone()
            saved_count = count_row[0] if count_row else 0
            if 0 < saved_count < 17:
                # We have old settings - clear them to reset all columns to visible
                c.execute("DELETE FROM app_config WHERE key LIKE 'preview_col_visible_%'")
                logger.info(f"Cleared outdated column visibility settings (had {saved_count}, need 17)")
        except Exception as e:
            logger.warning(f"Failed to check/reset column visibility: {e}")

        # Migration: Clear corrupted column widths (any column with 0 width)
        try:
            import json
            c.execute("SELECT value FROM app_config WHERE key = 'column_widths'")
            row = c.fetchone()
            if row:
                widths = json.loads(row[0])
                if any(w == 0 for w in widths.values()):
                    c.execute("DELETE FROM app_config WHERE key = 'column_widths'")
                    logger.info("Cleared corrupted column widths (had 0-width columns)")
        except Exception as e:
            logger.warning(f"Failed to check/reset column widths: {e}")

        # Migration: Fix corrupted parts_master ratios
        # Due to a bug in populate_parts_table, steel_ratio and aluminum_ratio columns got swapped.
        # The aluminum_ratio column contains what should be steel_ratio values.
        # Fix: Swap steel_ratio and aluminum_ratio values, then set aluminum_ratio to 0.
        try:
            c.execute("SELECT value FROM app_config WHERE key = 'ratios_migration_v1'")
            if not c.fetchone():
                # Count affected rows - those where aluminum_ratio > 0 and steel_ratio appears to have non_steel values
                c.execute("""
                    SELECT COUNT(*) FROM parts_master 
                    WHERE aluminum_ratio > 0.0 
                    AND (copper_ratio IS NULL OR copper_ratio = 0.0)
                    AND (wood_ratio IS NULL OR wood_ratio = 0.0)
                """)
                affected = c.fetchone()[0]
                
                if affected > 0:
                    # Swap: steel_ratio should get aluminum_ratio value, 
                    # non_steel_ratio should get steel_ratio value (the non-232 portion),
                    # aluminum_ratio should be 0 (these aren't actually aluminum products)
                    c.execute("""
                        UPDATE parts_master 
                        SET steel_ratio = aluminum_ratio,
                            non_steel_ratio = steel_ratio,
                            aluminum_ratio = 0.0
                        WHERE aluminum_ratio > 0.0 
                        AND (copper_ratio IS NULL OR copper_ratio = 0.0)
                        AND (wood_ratio IS NULL OR wood_ratio = 0.0)
                    """)
                    logger.info(f"Fixed {affected} parts with swapped ratio data (steel/aluminum swap corrected)")
                
                # Mark migration as complete
                c.execute("INSERT INTO app_config (key, value) VALUES ('ratios_migration_v1', '1')")
        except Exception as e:
            logger.warning(f"Failed to fix corrupted ratios: {e}")

        conn.commit()
        conn.close()
        logger.success("Database initialized")
    except Exception as e:
        logger.error(f"Database init failed: {e}")

init_database()

# ----------------------------------------------------------------------
# Drag & Drop Components
# ----------------------------------------------------------------------
class DraggableLabel(QLabel):
    def __init__(self, text):
        super().__init__(text)
        self.setStyleSheet("background:#6b6b6b;border:2px solid #aaa;border-radius:8px;padding:12px;font-weight:bold;color:#ffffff;cursor:hand;")
        self.setAlignment(Qt.AlignCenter)
        self.setCursor(QCursor(Qt.OpenHandCursor))  # Show hand cursor
    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            drag = QDrag(self)
            mime = QMimeData()
            mime.setText(self.text())
            drag.setMimeData(mime)
            pixmap = QPixmap(self.size())
            pixmap.fill(Qt.transparent)
            drag.setPixmap(pixmap)
            drag.exec_(Qt.CopyAction)
    def mouseMoveEvent(self, e):
        if e.buttons() == Qt.LeftButton:
            # This helps with drag detection
            pass

class DropTarget(QLabel):
    dropped = pyqtSignal(str, str)
    def __init__(self, field_key, field_name, drop_label=None):
        # Use custom drop_label if provided, otherwise use field_name
        label_text = drop_label if drop_label else field_name
        super().__init__(f"Drop {label_text} here")
        self.field_key = field_key
        # Unified style, proportional sizing
        self.setStyleSheet("font-size: 12pt; padding: 8px; background: #f8f8f8; border: 2px solid #bbb; border-radius: 8px; color: #222;")
        self.setAlignment(Qt.AlignCenter)
        self.setAcceptDrops(True)
        self.column_name = None
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
    def dragEnterEvent(self, e): 
        if e.mimeData().hasText(): e.accept()
    def dropEvent(self, e):
        col = e.mimeData().text()
        self.column_name = col
        self.setText(f"{self.field_key}\n<- {col}")
        self.setProperty("occupied", True)
        self.style().unpolish(self); self.style().polish(self)
        self.dropped.emit(self.field_key, col)
        e.accept()

class ForceEditableLineEdit(QLineEdit):
    """QLineEdit that forces itself to remain editable no matter what"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._force_editable = True
        # Set properties immediately
        self.setReadOnly(False)
        self.setEnabled(True)
        self.setFocusPolicy(Qt.StrongFocus)

    def setReadOnly(self, readonly):
        """Override setReadOnly to always stay editable"""
        if self._force_editable and readonly:
            return  # Ignore attempts to make readonly
        super().setReadOnly(readonly)

    def setEnabled(self, enabled):
        """Override setEnabled to always stay enabled"""
        if self._force_editable and not enabled:
            return  # Ignore attempts to disable
        super().setEnabled(enabled)

    def mousePressEvent(self, event):
        """Force editable on mouse click"""
        super(ForceEditableLineEdit, self).setReadOnly(False)
        super(ForceEditableLineEdit, self).setEnabled(True)
        self.setFocus()
        super().mousePressEvent(event)

    def focusInEvent(self, event):
        """Accept all focus events and force editable state"""
        super(ForceEditableLineEdit, self).setReadOnly(False)
        super(ForceEditableLineEdit, self).setEnabled(True)
        super().focusInEvent(event)

    def keyPressEvent(self, event):
        """Force editable before processing any key event"""
        from PyQt5.QtCore import Qt

        # For Tab keys, manually move focus to next/previous widget
        if event.key() == Qt.Key_Tab:
            # Move focus to next widget in tab order
            self.focusNextChild()
            event.accept()
            return
        elif event.key() == Qt.Key_Backtab:
            # Move focus to previous widget in tab order
            self.focusPreviousChild()
            event.accept()
            return

        super(ForceEditableLineEdit, self).setReadOnly(False)
        super(ForceEditableLineEdit, self).setEnabled(True)
        super().keyPressEvent(event)

class AutoSelectListWidget(QListWidget):
    """QListWidget that auto-selects first item when receiving focus via Tab"""
    def focusInEvent(self, event):
        """Select first item when receiving focus if nothing is selected"""
        super().focusInEvent(event)
        # Only auto-select if no item is currently selected and list has items
        if not self.currentItem() and self.count() > 0:
            self.setCurrentRow(0)

class FileDropZone(QLabel):
    """Drag-and-drop zone for importing CSV/Excel files"""
    file_dropped = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.setText("📁 Drag & Drop CSV/Excel File Here\n\nor click to browse")
        self.setAlignment(Qt.AlignCenter)
        self.setWordWrap(True)
        self.setMinimumHeight(120)
        self.setAcceptDrops(True)
        self.setCursor(Qt.PointingHandCursor)
        self.update_style(False)
        
    def update_style(self, hover=False):
        if hover:
            self.setStyleSheet("""
                QLabel {
                    background: #e3f2fd;
                    border: 3px dashed #2196F3;
                    border-radius: 10px;
                    font-size: 14pt;
                    font-weight: bold;
                    color: #1976D2;
                    padding: 20px;
                }
            """)
        else:
            self.setStyleSheet("""
                QLabel {
                    background: #f5f5f5;
                    border: 3px dashed #999;
                    border-radius: 10px;
                    font-size: 14pt;
                    font-weight: bold;
                    color: #666;
                    padding: 20px;
                }
            """)
    
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
            self.update_style(True)
        else:
            event.ignore()
    
    def dragLeaveEvent(self, event):
        self.update_style(False)
    
    def dropEvent(self, event):
        self.update_style(False)
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.lower().endswith(('.csv', '.xlsx', '.xls')):
                self.file_dropped.emit(file_path)
                event.accept()
            else:
                QMessageBox.warning(self, "Invalid File", 
                    "Please drop a CSV or Excel file (.csv, .xlsx, .xls)")
                event.ignore()
    
    def mousePressEvent(self, event):
        # Clicking opens file dialog
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select CSV/Excel File", str(INPUT_DIR), 
            "CSV/Excel Files (*.csv *.xlsx *.xls)"
        )
        if file_path:
            self.file_dropped.emit(file_path)



# ----------------------------------------------------------------------
# VISUAL PDF PATTERN TRAINER WITH DRAWING CANVAS
# ----------------------------------------------------------------------
class DerivativeMill(QMainWindow):
    def eventFilter(self, obj, event):
        """Application-level event filter - intercepts ALL events before any widget processing"""
        from PyQt5.QtCore import QEvent, Qt

        # Intercept ALL keyboard events at application level
        if event.type() == QEvent.KeyPress:
            focused_widget = QApplication.focusWidget()

            # CRITICAL FIX: Forward ALL keyboard events (including Tab) to ci_input/wt_input
            # The ForceEditableLineEdit.keyPressEvent will handle Tab specially with focusNextChild()
            if hasattr(self, 'ci_input') and hasattr(self, 'wt_input'):
                if focused_widget in [self.ci_input, self.wt_input]:
                    # Only process once per event - when it comes from QWindow
                    if obj.__class__.__name__ == 'QWindow':
                        # Manually send the event to the focused widget (including Tab keys)
                        QApplication.sendEvent(focused_widget, event)
                        return True  # We handled it, block further propagation

        return False  # Let ALL other events continue normally

    def setup_tab_by_index(self, index):
        """Initialize tab by index using existing setup methods."""
        tab_setup_methods = {
            1: self.setup_master_tab
            # Invoice Mapping, Output Mapping, and Parts Import moved to Configuration menu
            # Customs Config and Section 232 Actions moved to References menu
        }
        if index in tab_setup_methods:
            tab_setup_methods[index]()
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        # Compact default size - fully scalable with no minimum constraint
        self.setGeometry(50, 50, 1200, 700)

        # Install application-level event filter to intercept ALL keyboard events
        QApplication.instance().installEventFilter(self)

        # Track processed events to prevent duplicates
        self._processed_events = set()
        
        # Set window icon (use TEMP_RESOURCES_DIR for bundled resources)
        # Prefer icon.ico which is a proper multi-size icon created from banner_bg.png
        icon_path = TEMP_RESOURCES_DIR / "icon.ico"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "banner_bg.png"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "icon.png"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        
        self.current_csv = None
        self.shipment_mapping = {}
        self.output_column_mapping = None  # Will be initialized in setup_output_mapping_tab
        self.profile_header_row = 1  # Default header row (1 = first row)
        self.selected_mid = ""
        self.current_worker = None
        self.missing_df = None
        self.csv_total_value = 0.0
        self.last_processed_df = None
        self.last_output_filename = None
        self.shipment_targets = {}  # Prevent attribute error before tab setup

        # Load output font color from settings
        self.output_font_color = '#000000'  # Default black
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", ('output_font_color',))
            row = c.fetchone()
            conn.close()
            if row:
                self.output_font_color = row[0]
        except:
            pass

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        # Modern header with watermill logo on left, text in center
        header_container = QWidget()
        header_container.setStyleSheet("background: transparent; border: none;")
        header_container.setFixedHeight(48)
        header_layout = QHBoxLayout(header_container)
        header_layout.setContentsMargins(20, 6, 20, 6)
        header_layout.setSpacing(10)

        # Watermill logo on left (larger and more opaque)
        bg_path = TEMP_RESOURCES_DIR / "banner_bg.png"
        fixed_header_height = 48
        if bg_path.exists():
            bg_label = QLabel()
            pixmap = QPixmap(str(bg_path))
            scaled_pixmap = pixmap.scaled(fixed_header_height, fixed_header_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter_pixmap = QPixmap(scaled_pixmap.size())
            painter_pixmap.fill(Qt.transparent)
            from PyQt5.QtGui import QPainter
            painter = QPainter(painter_pixmap)
            painter.setOpacity(0.85)
            painter.drawPixmap(0, 0, scaled_pixmap)
            painter.end()
            bg_label.setPixmap(painter_pixmap)
            bg_label.setStyleSheet("background: transparent;")
            bg_label.setFixedSize(fixed_header_height, fixed_header_height)
            self.header_bg_label = bg_label
        else:
            self.header_bg_label = None

        # App name centered
        app_name = QLabel(f"{APP_NAME}")
        app_name.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        app_name.setFixedHeight(fixed_header_height)
        # Set font color based on theme
        dark_mode_teal_color = "#42A0BD"  # Matches enabled Process Invoice button color in dark mode
        light_mode_color = "#555555"  # Original color for light mode
        color = dark_mode_teal_color if hasattr(self, 'current_theme') and self.current_theme in ["Fusion (Dark)", "Ocean"] else light_mode_color
        app_name.setStyleSheet(f"""
            font-size: 22px;
            font-weight: bold;
            color: {color};
            font-family: 'Impact', 'Arial Black', sans-serif;
            padding: 0px;
        """)
        from PyQt5.QtWidgets import QGraphicsDropShadowEffect
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        shadow.setColor(QColor(0, 0, 0, 120))
        shadow.setOffset(3, 3)
        app_name.setGraphicsEffect(shadow)
        # Add logo and title directly to the QHBoxLayout, both vertically centered, same height
        if self.header_bg_label:
            header_layout.addWidget(self.header_bg_label, 0, Qt.AlignVCenter)
        header_layout.addWidget(app_name, 1, Qt.AlignVCenter)


        layout.addWidget(header_container)



        # Add a native menu bar with Settings and Log View actions
        menubar = QMenuBar(self)
        settings_menu = menubar.addMenu("Settings")
        # Use a standard gear icon from QStyle
        gear_icon = self.style().standardIcon(QStyle.SP_FileDialogDetailedView)
        settings_action = QAction(gear_icon, "Settings", self)
        settings_action.triggered.connect(self.show_settings_dialog)
        settings_menu.addAction(settings_action)
        
        # Add Log View menu
        log_menu = menubar.addMenu("Log View")
        log_icon = self.style().standardIcon(QStyle.SP_FileDialogContentsView)
        log_action = QAction(log_icon, "View Log", self)
        log_action.triggered.connect(self.show_log_dialog)
        log_menu.addAction(log_action)

        # Add References menu
        references_menu = menubar.addMenu("References")
        references_icon = self.style().standardIcon(QStyle.SP_FileDialogInfoView)
        references_action = QAction(references_icon, "References...", self)
        references_action.triggered.connect(self.show_references_dialog)
        references_menu.addAction(references_action)

        # Add Configuration menu
        config_menu = menubar.addMenu("Configuration")
        config_icon = self.style().standardIcon(QStyle.SP_FileDialogDetailedView)
        config_action = QAction(config_icon, "Configuration...", self)
        config_action.triggered.connect(self.show_configuration_dialog)
        config_menu.addAction(config_action)

        # Add Help menu
        help_menu = menubar.addMenu("Help")
        
        # Check for Updates action
        update_icon = self.style().standardIcon(QStyle.SP_BrowserReload)
        update_action = QAction(update_icon, "Check for Updates...", self)
        update_action.triggered.connect(self.check_for_updates_manual)
        help_menu.addAction(update_action)
        
        help_menu.addSeparator()
        
        # About action
        about_icon = self.style().standardIcon(QStyle.SP_MessageBoxInformation)
        about_action = QAction(about_icon, "About", self)
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)
        
        layout.setMenuBar(menubar)
        self.settings_action = settings_action

        # Top status bar removed per user request
        # Create a dummy status object that ignores all calls
        class DummyStatus:
            def setText(self, text): pass
            def setStyleSheet(self, style): pass
            def setAlignment(self, align): pass
        self.status = DummyStatus()

        # Add spacing between header and tabs
        layout.addSpacing(20)

        self.tabs = QTabWidget()
        self.tab_process = QWidget()
        self.tab_shipment_map = QWidget()
        self.tab_output_map = QWidget()
        self.tab_import = QWidget()
        self.tab_master = QWidget()
        self.tab_log = QWidget()  # Keep widget for log functionality
        self.tab_config = QWidget()
        self.tab_actions = QWidget()
        self.tabs.addTab(self.tab_process, "Process Shipment")
        self.tabs.addTab(self.tab_master, "Parts View")
        # Invoice Mapping, Output Mapping, and Parts Import moved to Configuration menu
        # Customs Config and Section 232 Actions moved to References menu
        
        # Only tabs (no settings icon here)
        tabs_container = QWidget()
        tabs_layout = QHBoxLayout(tabs_container)
        tabs_layout.setContentsMargins(0, 0, 0, 0)
        tabs_layout.setSpacing(10)
        tabs_layout.addWidget(self.tabs)
        layout.addWidget(tabs_container)
        
        # Bottom status bar with export progress indicator
        bottom_bar = QWidget()
        bottom_bar_layout = QHBoxLayout(bottom_bar)
        bottom_bar_layout.setContentsMargins(10, 3, 10, 3)
        bottom_bar_layout.setSpacing(10)
        
        self.bottom_status = QLabel("Ready")
        bottom_bar_layout.addWidget(self.bottom_status, 1)
        
        # Export progress indicator (hidden by default)
        self.export_progress_widget = QWidget()
        export_progress_layout = QHBoxLayout(self.export_progress_widget)
        export_progress_layout.setContentsMargins(0, 0, 0, 0)
        export_progress_layout.setSpacing(5)
        
        self.export_status_label = QLabel("")
        self.export_status_label.setStyleSheet("font-size: 8pt; color: #666666;")
        export_progress_layout.addWidget(self.export_status_label)
        
        self.export_progress_bar = QProgressBar()
        self.export_progress_bar.setMaximum(100)
        self.export_progress_bar.setFixedWidth(120)
        self.export_progress_bar.setFixedHeight(14)
        self.export_progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #b0b0b0;
                border-radius: 3px;
                background-color: #f5f5f5;
                text-align: center;
                font-size: 7pt;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
                border-radius: 2px;
            }
        """)
        export_progress_layout.addWidget(self.export_progress_bar)
        
        bottom_bar_layout.addWidget(self.export_progress_widget)
        self.export_progress_widget.hide()

        # Add version label in bottom right corner
        self.version_label = QLabel(VERSION)
        self.version_label.setStyleSheet("font-size: 7pt; color: #999999;")
        self.version_label.setAlignment(Qt.AlignRight)
        bottom_bar_layout.addWidget(self.version_label)

        self.bottom_bar = bottom_bar  # Store reference for theme updates
        bottom_bar.setFixedHeight(20)
        layout.addWidget(bottom_bar)

        # Track which tabs have been initialized (lazy loading for performance)
        self.tabs_initialized = set()
        

        # Only set up the first tab immediately for faster startup
        self.setup_process_tab()
        self.tabs_initialized.add(0)

        # Connect to tab change signal for lazy initialization
        self.tabs.currentChanged.connect(self.on_tab_changed)



        logger.success(f"{APP_NAME} {VERSION} GUI ready")

    # Removed deferred_initialization; tabs are now lazily loaded only when selected
    
    def initialize_data(self, splash=None, progress_callback=None):
        """Initialize database and load all data before showing window"""
        steps = [
            ("Loading configuration...", self.load_config_paths),
            ("Applying theme...", self.apply_saved_theme),
            ("Applying font size...", self.apply_saved_font_size),
            ("Loading MIDs...", self.load_available_mids),
            ("Loading profiles...", self.load_mapping_profiles),
            ("Loading export profiles...", self.load_output_mapping_profiles),
            # Removed output file scanning on startup
            ("Scanning input files...", self.refresh_input_files),
            ("Starting auto-refresh...", self.setup_auto_refresh),
            ("Finalizing...", self.update_status_bar_styles),
        ]
        
        total_steps = len(steps)
        for i, (message, func) in enumerate(steps):
            if splash:
                splash.setText(f"{message}\nPlease wait...")
            if progress_callback:
                progress_callback(int((i / total_steps) * 100))
            QApplication.processEvents()
            
            try:
                func()
            except Exception as e:
                logger.error(f"Error during {message}: {e}")
        
        if progress_callback:
            progress_callback(100)

        # Ensure input fields are enabled after all initialization
        if hasattr(self, '_enable_input_fields'):
            self._enable_input_fields()

        logger.success(f"{APP_NAME} {VERSION} loaded successfully")
    
    def on_tab_changed(self, index):
        """Initialize tabs lazily when they are first accessed"""
        if index in self.tabs_initialized:
            return

        # Map tab index to setup method
        # Tab order: 0=Process, 1=Parts View
        tab_setup_methods = {
            1: self.setup_master_tab
            # Invoice Mapping, Output Mapping, and Parts Import moved to Configuration menu
            # Customs Config and Section 232 Actions moved to References menu
        }
        
        # Initialize the tab
        if index in tab_setup_methods:
            tab_setup_methods[index]()
            self.tabs_initialized.add(index)
            logger.debug(f"Initialized tab {index}")
    
    def apply_saved_theme(self):
        """Load and apply the saved theme preference on startup (per-user setting)"""
        theme = get_user_setting('theme', 'Fusion (Light)')
        self.apply_theme(theme)

    def apply_saved_font_size(self):
        """Load and apply the saved font size preference on startup (per-user setting)"""
        font_size = get_user_setting_int('font_size', 10)
        self.apply_font_size(font_size)

    def load_config_paths(self):
        try:
            self.bottom_status.setText("Loading Directory location...")
            QApplication.processEvents()
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = 'input_dir'")
            row = c.fetchone()
            global INPUT_DIR, PROCESSED_DIR
            if row:
                INPUT_DIR = Path(row[0])
                PROCESSED_DIR = INPUT_DIR / "Processed"
                PROCESSED_DIR.mkdir(exist_ok=True)
                QApplication.processEvents()
            c.execute("SELECT value FROM app_config WHERE key = 'output_dir'")
            row = c.fetchone()
            if row:
                global OUTPUT_DIR
                OUTPUT_DIR = Path(row[0])
                OUTPUT_DIR.mkdir(exist_ok=True)
                QApplication.processEvents()
            conn.close()
            self.bottom_status.setText("Ready")
            QApplication.processEvents()
        except Exception as e:
            logger.error(f"Config load failed: {e}")
            self.bottom_status.setText("Config load failed")
            QApplication.processEvents()

    def setup_process_tab(self):
        layout = QVBoxLayout(self.tab_process)

        # MAIN CONTAINER: Left side (controls) + Right side (preview) with splitter
        main_container = QHBoxLayout()
        
        # Create splitter for resizable panels
        splitter = QSplitter(Qt.Horizontal)
        
        # LEFT SIDE: Single outer box containing all controls
        left_outer_box = QGroupBox("Controls")
        left_side = QVBoxLayout(left_outer_box)
        left_side.setSpacing(10)
        left_side.setContentsMargins(10, 10, 10, 10)
        
        # INPUT FILES LIST — now inside Shipment File group
        self.input_files_list = AutoSelectListWidget()
        self.input_files_list.setSelectionMode(QListWidget.SingleSelection)
        self.input_files_list.itemClicked.connect(self.load_selected_input_file)
        # Connect itemActivated for Enter key and double-click support
        self.input_files_list.itemActivated.connect(self.load_selected_input_file)
        # Allow focus for tab navigation
        self.input_files_list.setFocusPolicy(Qt.StrongFocus)
        # Limit height to show ~6 files to save vertical space
        self.input_files_list.setFixedHeight(100)
        self.refresh_input_btn = QPushButton("Refresh")
        self.refresh_input_btn.setFixedHeight(25)
        self.refresh_input_btn.clicked.connect(self.refresh_input_files)

        # INVOICE VALUES
        values_group = QGroupBox("Invoice Values")
        values_layout = QFormLayout()
        values_layout.setLabelAlignment(Qt.AlignRight)

        self.ci_input = ForceEditableLineEdit("")
        self.ci_input.setObjectName("ci_input")
        self.ci_input.setPlaceholderText("Enter CI value...")
        self.ci_input.textChanged.connect(self.update_invoice_check)

        self.wt_input = ForceEditableLineEdit("")
        self.wt_input.setObjectName("wt_input")
        self.wt_input.setPlaceholderText("Enter net weight...")
        self.wt_input.textChanged.connect(self.update_invoice_check)

        values_layout.addRow("CI Value (USD):", self.ci_input)
        values_layout.addRow("Net Weight (kg):", self.wt_input)

        # MID selector (moved above Invoice Check)
        self.mid_label = QLabel("MID:")
        self.mid_combo = QComboBox()
        self.mid_combo.setFocusPolicy(Qt.StrongFocus)  # Ensure combo accepts keyboard focus
        self.mid_combo.currentTextChanged.connect(self.on_mid_changed)
        values_layout.addRow(self.mid_label, self.mid_combo)

        # Removed broken setTabOrder calls - they were causing Qt warnings and possibly blocking keyboard input

        # Invoice check label and Edit Values button
        self.invoice_check_label = QLabel("No file loaded")
        self.invoice_check_label.setWordWrap(False)  # Don't wrap - keep invoice total on one line
        self.invoice_check_label.setStyleSheet("font-size: 7pt;")
        self.invoice_check_label.setAlignment(Qt.AlignCenter)
        self.invoice_check_label.setMinimumWidth(180)  # Wider minimum to fit invoice totals
        self.invoice_check_label.setMaximumWidth(250)  # Allow more space for larger amounts

        vbox_check = QVBoxLayout()
        vbox_check.setSpacing(12)
        vbox_check.setContentsMargins(0, 10, 0, 0)

        vbox_check.addWidget(self.invoice_check_label, alignment=Qt.AlignCenter)

        # Edit Values button (initially hidden, shown when values don't match)
        self.edit_values_btn = QPushButton("Edit Values")
        self.edit_values_btn.setFixedSize(120, 30)
        self.edit_values_btn.setStyleSheet(self.get_button_style("warning"))
        self.edit_values_btn.setVisible(False)
        self.edit_values_btn.setAutoDefault(True)
        self.edit_values_btn.setDefault(True)
        self.edit_values_btn.clicked.connect(self.start_processing_with_editable_preview)
        vbox_check.addWidget(self.edit_values_btn, alignment=Qt.AlignCenter)
        
        vbox_check.addStretch()

        values_layout.addRow("Invoice Check:", vbox_check)
        values_group.setLayout(values_layout)

        # SHIPMENT FILE (merged with Saved Profiles and Input Files)
        file_group = QGroupBox("Shipment File")
        file_group.setObjectName("SavedProfilesGroup")
        file_layout = QFormLayout()
        file_layout.setLabelAlignment(Qt.AlignRight)
        # Profile selector
        self.profile_combo = QComboBox()
        self.profile_combo.currentTextChanged.connect(self.load_selected_profile)
        file_layout.addRow("Map Profile:", self.profile_combo)
        # Input files list and refresh button (moved here)
        file_layout.addRow("Input Files:", self.input_files_list)
        file_layout.addRow("", self.refresh_input_btn)
        # File display (read-only, shows selected file from Input Files list)
        self.file_label = QLabel("No file selected")
        self.file_label.setWordWrap(True)
        self.update_file_label_style()  # Set initial style based on theme
        file_layout.addRow("Selected File:", self.file_label)
        file_group.setLayout(file_layout)
        left_side.addWidget(file_group)
        left_side.addWidget(values_group)

        # ACTIONS GROUP — Clear All + Export Worksheet buttons (side by side)
        actions_group = QGroupBox("Actions")
        actions_layout = QHBoxLayout()
        actions_layout.setContentsMargins(5, 5, 5, 5)
        actions_layout.setSpacing(5)

        self.process_btn = QPushButton("Process Invoice")
        self.process_btn.setEnabled(False)
        self.process_btn.setFixedHeight(28)
        self.process_btn.setStyleSheet(self.get_button_style("success"))
        self.process_btn.clicked.connect(self._process_or_export)
        # Make button respond to Enter/Return key when focused
        self.process_btn.setAutoDefault(True)
        self.process_btn.setDefault(False)  # Don't make it the default for the whole window

        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setFixedHeight(28)
        self.clear_btn.setStyleSheet(self.get_button_style("danger"))
        self.clear_btn.clicked.connect(self.clear_all)

        actions_layout.addWidget(self.process_btn)
        actions_layout.addWidget(self.clear_btn)
        actions_group.setLayout(actions_layout)
        left_side.addWidget(actions_group)

        # EXPORT PROFILE GROUP — quick access to output mapping profiles
        export_profile_group = QGroupBox("Export Profile")
        export_profile_layout = QVBoxLayout()
        export_profile_layout.setContentsMargins(5, 5, 5, 5)

        self.quick_export_profile_combo = QComboBox()
        self.quick_export_profile_combo.currentTextChanged.connect(self.on_quick_export_profile_changed)
        export_profile_layout.addWidget(self.quick_export_profile_combo)

        export_profile_group.setLayout(export_profile_layout)
        left_side.addWidget(export_profile_group)

        # EXPORTED FILES GROUP — shows recent exports
        exports_group = QGroupBox("Exported Files")
        exports_layout = QVBoxLayout()

        self.exports_list = AutoSelectListWidget()
        self.exports_list.setSelectionMode(QListWidget.SingleSelection)
        self.exports_list.itemDoubleClicked.connect(self.open_exported_file)
        # Connect itemActivated for Enter key support
        self.exports_list.itemActivated.connect(self.open_exported_file)
        # Allow focus for tab navigation
        self.exports_list.setFocusPolicy(Qt.StrongFocus)
        exports_layout.addWidget(self.exports_list)

        self.refresh_exports_btn = QPushButton("Refresh")
        self.refresh_exports_btn.setFixedHeight(25)
        self.refresh_exports_btn.clicked.connect(self.refresh_exported_files)
        exports_layout.addWidget(self.refresh_exports_btn)

        # Invoice total display for selected file
        total_layout = QHBoxLayout()
        total_layout.addWidget(QLabel("Invoice Total:"))
        self.export_invoice_total = QLineEdit()
        self.export_invoice_total.setReadOnly(True)
        self.export_invoice_total.setPlaceholderText("Select a file")
        total_layout.addWidget(self.export_invoice_total)
        exports_layout.addLayout(total_layout)

        # Connect selection change to update invoice total
        self.exports_list.itemSelectionChanged.connect(self.update_export_invoice_total)

        exports_group.setLayout(exports_layout)
        left_side.addWidget(exports_group)
        
        # Set maximum width for left controls to keep it compact
        left_outer_box.setMaximumWidth(350)

        # Add left_outer_box to splitter
        splitter.addWidget(left_outer_box)

        # RIGHT SIDE: Preview table in a widget
        right_widget = QWidget()
        right_side = QVBoxLayout(right_widget)
        right_side.setContentsMargins(0, 0, 0, 0)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        right_side.addWidget(self.progress)

        preview_group = QGroupBox("Result Preview")
        preview_layout = QVBoxLayout()

        self.table = QTableWidget()
        self.table.setColumnCount(17)
        self.table.setHorizontalHeaderLabels([
            "Product No","Value","HTS","MID","CBP_qty1","Dec","Melt","Cast","Smelt","Flag","Steel%","Al%","Cu%","Wood%","Auto%","Non-232%","232 Status"
        ])
        # Make columns manually resizable instead of auto-stretch
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setSelectionBehavior(QTableWidget.SelectItems)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSortingEnabled(False)  # Disabled for better performance
        # Set row height from saved preference (per-user setting)
        saved_row_height = get_user_setting_int('preview_row_height', 20)
        self.table.verticalHeader().setDefaultSectionSize(saved_row_height)
        # Match body font size to the header font size and make non-bold
        header_font = self.table.horizontalHeader().font()
        header_font.setBold(False)
        self.table.horizontalHeader().setFont(header_font)
        self.table.setFont(header_font)

        # Enable clicking header to select entire column
        self.table.horizontalHeader().sectionClicked.connect(self.select_column)

        # Connect signal to save column widths when they change
        self.table.horizontalHeader().sectionResized.connect(self.save_column_widths)

        # Load saved column widths
        self.load_column_widths()

        # Apply green focus color stylesheet
        self.update_table_stylesheet()

        preview_layout.addWidget(self.table)
        preview_group.setLayout(preview_layout)
        right_side.addWidget(preview_group, 1)

        # Add right widget to splitter
        splitter.addWidget(right_widget)

        # Set initial splitter sizes - minimize left, maximize right (15% left, 85% right)
        splitter.setSizes([200, 1000])

        # Make the splitter collapsible on the left side
        splitter.setCollapsible(0, False)  # Don't allow full collapse
        splitter.setCollapsible(1, False)
        
        # Add splitter to main container
        main_container.addWidget(splitter)

        layout.addLayout(main_container, 1)

        # Set up tab order for keyboard navigation through controls
        # Order: Map Profile → Input Files → Refresh (Shipment) → CI Value → Net Weight →
        #        MID → Process Invoice → Edit Values → Clear All → Exported Files → Refresh (Exports)
        self.setTabOrder(self.profile_combo, self.input_files_list)
        self.setTabOrder(self.input_files_list, self.refresh_input_btn)
        self.setTabOrder(self.refresh_input_btn, self.ci_input)
        self.setTabOrder(self.ci_input, self.wt_input)
        self.setTabOrder(self.wt_input, self.mid_combo)
        self.setTabOrder(self.mid_combo, self.process_btn)
        self.setTabOrder(self.process_btn, self.edit_values_btn)
        self.setTabOrder(self.edit_values_btn, self.clear_btn)
        self.setTabOrder(self.clear_btn, self.exports_list)
        self.setTabOrder(self.exports_list, self.refresh_exports_btn)

        self.tab_process.setLayout(layout)
        self._install_preview_shortcuts()

        # Ensure input fields are enabled on startup
        self._enable_input_fields()

        # Create a timer to continuously force fields to be editable
        # This works around whatever is locking the fields
        from PyQt5.QtCore import QTimer
        self._field_watchdog_timer = QTimer()
        self._field_watchdog_timer.timeout.connect(self._force_fields_editable)
        self._field_watchdog_timer.start(100)  # Check every 100ms

        # Event filter already installed in __init__, don't install again
    
    def _force_fields_editable(self):
        """Watchdog timer callback that forces fields to stay editable"""
        if hasattr(self, 'ci_input'):
            # Bypass the override and force the parent class method
            QLineEdit.setReadOnly(self.ci_input, False)
            QLineEdit.setEnabled(self.ci_input, True)
        if hasattr(self, 'wt_input'):
            QLineEdit.setReadOnly(self.wt_input, False)
            QLineEdit.setEnabled(self.wt_input, True)

    def _enable_input_fields(self):
        """Ensure CI and Weight input fields are enabled and ready for input"""
        # Block signals to prevent recursion during enable
        if hasattr(self, 'ci_input'):
            self.ci_input.blockSignals(True)
            self.ci_input.setReadOnly(False)
            self.ci_input.setEnabled(True)
            self.ci_input.setFocusPolicy(Qt.StrongFocus)
            # Force immediate visual update
            self.ci_input.update()
            self.ci_input.blockSignals(False)

        if hasattr(self, 'wt_input'):
            self.wt_input.blockSignals(True)
            self.wt_input.setReadOnly(False)
            self.wt_input.setEnabled(True)
            self.wt_input.setFocusPolicy(Qt.StrongFocus)
            # Force immediate visual update
            self.wt_input.update()
            self.wt_input.blockSignals(False)

    def show_log_dialog(self):
        """Show the application log in a dialog window"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Application Log")
        dialog.resize(900, 600)
        layout = QVBoxLayout(dialog)

        # Title
        title = QLabel("<h2>Application Log</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Log text area
        log_text = QTextEdit()
        log_text.setReadOnly(True)
        log_text.setFont(QFont("Consolas", 9))
        log_text.setPlainText(logger.get_logs())
        layout.addWidget(log_text)

        # Button row
        btn_layout = QHBoxLayout()
        
        btn_refresh = QPushButton("Refresh")
        btn_refresh.setStyleSheet("background:#28a745; color:white; font-weight:bold;")
        btn_refresh.clicked.connect(lambda: log_text.setPlainText(logger.get_logs()))
        
        btn_copy = QPushButton("Copy to Clipboard")
        btn_copy.setStyleSheet("background:#0078D7; color:white; font-weight:bold;")
        btn_copy.clicked.connect(lambda: QApplication.clipboard().setText(log_text.toPlainText()))
        
        btn_clear = QPushButton("Clear Log")
        btn_clear.setStyleSheet("background:#dc3545; color:white; font-weight:bold;")
        btn_clear.clicked.connect(lambda: (logger.logs.clear(), log_text.clear()))
        
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        
        btn_layout.addWidget(btn_refresh)
        btn_layout.addWidget(btn_copy)
        btn_layout.addWidget(btn_clear)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        # Auto-refresh timer
        refresh_timer = QTimer(dialog)
        refresh_timer.timeout.connect(lambda: log_text.setPlainText(logger.get_logs()))
        refresh_timer.start(1000)

        dialog.exec_()

    def show_references_dialog(self):
        """Show the References dialog with Customs Config and Section 232 Actions tabs"""
        dialog = QDialog(self)
        dialog.setWindowTitle("References")
        dialog.resize(1000, 700)
        layout = QVBoxLayout(dialog)

        # Create tab widget
        tabs = QTabWidget()

        # Create new tab widgets for the dialog
        tab_config = QWidget()
        tab_actions = QWidget()

        # Temporarily swap the instance variables so setup methods populate the new widgets
        original_tab_config = self.tab_config
        original_tab_actions = self.tab_actions

        self.tab_config = tab_config
        self.tab_actions = tab_actions

        # Setup the tabs
        self.setup_config_tab()
        self.setup_actions_tab()

        # Restore original references (though they may be deleted)
        self.tab_config = original_tab_config
        self.tab_actions = original_tab_actions

        # Add the new tabs to the dialog
        tabs.addTab(tab_config, "Customs Config")
        tabs.addTab(tab_actions, "Section 232 Actions")

        layout.addWidget(tabs)

        # Close button
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        btn_close.setStyleSheet(self.get_button_style("default"))
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        dialog.exec_()

    def show_configuration_dialog(self):
        """Show the Configuration dialog with Invoice Mapping, Output Mapping, and Parts Import tabs"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Configuration")
        dialog.resize(1000, 700)
        layout = QVBoxLayout(dialog)

        # Create tab widget
        tabs = QTabWidget()

        # Create new tab widgets for the dialog
        tab_shipment_map = QWidget()
        tab_output_map = QWidget()
        tab_import = QWidget()

        # Temporarily swap the instance variables so setup methods populate the new widgets
        original_tab_shipment_map = self.tab_shipment_map
        original_tab_output_map = self.tab_output_map
        original_tab_import = self.tab_import

        self.tab_shipment_map = tab_shipment_map
        self.tab_output_map = tab_output_map
        self.tab_import = tab_import

        # Setup the tabs
        self.setup_shipment_mapping_tab()
        self.setup_output_mapping_tab()
        self.setup_import_tab()

        # Restore original references (though they may be deleted)
        self.tab_shipment_map = original_tab_shipment_map
        self.tab_output_map = original_tab_output_map
        self.tab_import = original_tab_import

        # Add the new tabs to the dialog
        tabs.addTab(tab_shipment_map, "Invoice Mapping Profiles")
        tabs.addTab(tab_output_map, "Output Mapping")
        tabs.addTab(tab_import, "Parts Import")

        layout.addWidget(tabs)

        # Close button
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        btn_close.setStyleSheet(self.get_button_style("default"))
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        dialog.exec_()

    def show_settings_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Settings")
        dialog.resize(700, 750)  # Increased size for better layout
        layout = QVBoxLayout(dialog)

        # Create tab widget for better organization
        tabs = QTabWidget()

        # ===== TAB 1: APPEARANCE =====
        appearance_widget = QWidget()
        appearance_layout = QVBoxLayout(appearance_widget)

        # Theme Settings Group
        theme_group = QGroupBox("Appearance")
        theme_layout = QFormLayout()
        
        # Use local variable instead of instance variable
        theme_combo = QComboBox()
        theme_combo.addItems(["System Default", "Fusion (Light)", "Windows", "Fusion (Dark)", "Ocean", "Light Cyan"])
        
        # Load saved theme preference from per-user settings
        saved_theme = get_user_setting('theme', 'Fusion (Light)')
        index = theme_combo.findText(saved_theme)
        if index >= 0:
            # Block signals to prevent double-applying theme
            theme_combo.blockSignals(True)
            theme_combo.setCurrentIndex(index)
            theme_combo.blockSignals(False)
        
        theme_combo.currentTextChanged.connect(self.apply_theme)
        theme_layout.addRow("Application Theme:", theme_combo)

        theme_info = QLabel("<small>Theme changes apply immediately. System Default uses your Windows theme settings.</small>")
        theme_info.setWordWrap(True)
        theme_info.setStyleSheet("color:#666; padding:5px;")
        theme_layout.addRow("", theme_info)

        # Font Size Slider
        font_size_layout = QHBoxLayout()
        font_size_slider = QSlider(Qt.Horizontal)
        font_size_slider.setMinimum(8)
        font_size_slider.setMaximum(16)
        font_size_slider.setValue(10)  # Default
        font_size_slider.setTickPosition(QSlider.TicksBelow)
        font_size_slider.setTickInterval(1)

        font_size_value_label = QLabel("10pt")
        font_size_value_label.setMinimumWidth(40)

        # Load saved font size preference from per-user settings
        saved_font_size = get_user_setting_int('font_size', 10)
        font_size_slider.setValue(saved_font_size)
        font_size_value_label.setText(f"{saved_font_size}pt")

        # Connect slider to update label and apply font size
        def update_font_size(value):
            font_size_value_label.setText(f"{value}pt")
            self.apply_font_size(value)

        font_size_slider.valueChanged.connect(update_font_size)

        font_size_layout.addWidget(font_size_slider)
        font_size_layout.addWidget(font_size_value_label)

        theme_layout.addRow("Font Size:", font_size_layout)

        # Row Height Slider for Result Preview table
        row_height_layout = QHBoxLayout()
        row_height_slider = QSlider(Qt.Horizontal)
        row_height_slider.setMinimum(16)
        row_height_slider.setMaximum(40)
        row_height_slider.setValue(20)  # Default
        row_height_slider.setTickPosition(QSlider.TicksBelow)
        row_height_slider.setTickInterval(4)

        row_height_value_label = QLabel("20px")
        row_height_value_label.setMinimumWidth(40)

        # Load saved row height preference from per-user settings
        saved_row_height = get_user_setting_int('preview_row_height', 20)
        row_height_slider.setValue(saved_row_height)
        row_height_value_label.setText(f"{saved_row_height}px")

        # Connect slider to update label and apply row height
        def update_row_height(value):
            row_height_value_label.setText(f"{value}px")
            self.apply_row_height(value)

        row_height_slider.valueChanged.connect(update_row_height)

        row_height_layout.addWidget(row_height_slider)
        row_height_layout.addWidget(row_height_value_label)

        theme_layout.addRow("Preview Row Height:", row_height_layout)

        theme_group.setLayout(theme_layout)
        appearance_layout.addWidget(theme_group)

        # Excel Viewer Settings Group
        viewer_group = QGroupBox("Excel File Viewer")
        viewer_layout = QFormLayout()

        # Excel viewer combo box
        viewer_combo = QComboBox()
        if sys.platform == 'linux':
            viewer_combo.addItems(["System Default", "Gnumeric"])
        else:
            viewer_combo.addItems(["System Default"])
            viewer_combo.setEnabled(False)  # Only relevant on Linux

        # Load saved preference (per-user setting)
        saved_viewer = get_user_setting('excel_viewer', 'System Default')
        index = viewer_combo.findText(saved_viewer)
        if index >= 0:
            viewer_combo.setCurrentIndex(index)

        # Save preference when changed (per-user setting)
        def save_viewer_preference(viewer):
            set_user_setting('excel_viewer', viewer)
            logger.info(f"Excel viewer preference changed to: {viewer}")

        viewer_combo.currentTextChanged.connect(save_viewer_preference)
        viewer_layout.addRow("Open Exported Files With:", viewer_combo)

        viewer_info = QLabel("<small>Choose which application opens exported Excel files. (Linux only)</small>")
        viewer_info.setWordWrap(True)
        viewer_info.setStyleSheet("color:#666; padding:5px;")
        viewer_layout.addRow("", viewer_info)

        viewer_group.setLayout(viewer_layout)
        appearance_layout.addWidget(viewer_group)

        # Preview Table Colors Group
        colors_group = QGroupBox("Preview Table Row Colors")
        colors_layout = QGridLayout()
        colors_layout.setSpacing(8)
        colors_layout.setContentsMargins(10, 10, 10, 10)

        # Helper function to create color picker button
        def create_color_button(config_key, default_color):
            """Create a color picker button with saved color (per-user setting)"""
            button = QPushButton()
            button.setFixedSize(60, 20)

            # Load saved color from per-user settings or use default
            saved_color = get_user_setting(config_key, default_color)

            # Set button style with current color
            button.setStyleSheet(f"background-color: {saved_color}; border: 1px solid #999;")

            def pick_color():
                current_color = get_user_setting(config_key, default_color)
                color = QColorDialog.getColor(QColor(current_color), dialog, "Choose Color")
                if color.isValid():
                    color_hex = color.name()
                    button.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999;")
                    # Save to per-user settings
                    set_user_setting(config_key, color_hex)
                    logger.info(f"Saved color preference {config_key}: {color_hex}")
                    # Refresh the preview table if it exists
                    if hasattr(self, 'table') and self.table.rowCount() > 0:
                        self.refresh_preview_colors()

            button.clicked.connect(pick_color)
            return button

        # Section 232 material type color pickers in grid layout
        # Row 0: Steel, Aluminum, Copper
        colors_layout.addWidget(QLabel("Steel:"), 0, 0, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_steel_color', '#4a4a4a'), 0, 1)
        colors_layout.addWidget(QLabel("Aluminum:"), 0, 2, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_aluminum_color', '#3498db'), 0, 3)
        colors_layout.addWidget(QLabel("Copper:"), 0, 4, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_copper_color', '#e67e22'), 0, 5)

        # Row 1: Wood, Auto, Non-232
        colors_layout.addWidget(QLabel("Wood:"), 1, 0, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_wood_color', '#27ae60'), 1, 1)
        colors_layout.addWidget(QLabel("Auto:"), 1, 2, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_auto_color', '#9b59b6'), 1, 3)
        colors_layout.addWidget(QLabel("Non-232:"), 1, 4, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_non232_color', '#ff0000'), 1, 5)

        # Row 2: Sec301 Background (spans multiple columns)
        colors_layout.addWidget(QLabel("Sec301 Background:"), 2, 0, 1, 2, Qt.AlignRight)
        colors_layout.addWidget(create_color_button('preview_sec301_bg_color', '#ffc8c8'), 2, 2)

        colors_group.setLayout(colors_layout)
        appearance_layout.addWidget(colors_group)

        # Preview Column Visibility Group
        columns_group = QGroupBox("Result Preview Column Visibility")
        columns_layout = QVBoxLayout()
        
        # Column names and their default visibility
        column_names = [
            "Product No", "Value", "HTS", "MID", "CBP_qty1", "Dec",
            "Melt", "Cast", "Smelt", "Flag", "Steel%", "Al%", "Cu%", "Wood%", "Auto%", "Non-232%", "232 Status"
        ]
        
        # Create checkboxes in a grid layout
        columns_grid = QGridLayout()
        column_checkboxes = []
        
        for i, col_name in enumerate(column_names):
            checkbox = QCheckBox(col_name)
            checkbox.setChecked(True)  # Default to visible
            
            # Load saved visibility preference
            config_key = f'preview_col_visible_{i}'
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("SELECT value FROM app_config WHERE key = ?", (config_key,))
                row = c.fetchone()
                conn.close()
                if row:
                    checkbox.setChecked(row[0] == '1')
            except:
                pass
            
            # Save preference and apply when changed
            def make_toggle_handler(col_idx, cb):
                def handler(state):
                    try:
                        conn = sqlite3.connect(str(DB_PATH))
                        c = conn.cursor()
                        c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                                  (f'preview_col_visible_{col_idx}', '1' if state else '0'))
                        conn.commit()
                        conn.close()
                        # Apply visibility to table
                        if hasattr(self, 'table'):
                            self.table.setColumnHidden(col_idx, not state)
                        logger.info(f"Column {col_idx} visibility set to {'visible' if state else 'hidden'}")
                    except Exception as e:
                        logger.error(f"Failed to save column visibility: {e}")
                return handler
            
            checkbox.stateChanged.connect(make_toggle_handler(i, checkbox))
            column_checkboxes.append(checkbox)
            
            # Arrange in 5 columns
            row_num = i // 5
            col_num = i % 5
            columns_grid.addWidget(checkbox, row_num, col_num)
        
        columns_layout.addLayout(columns_grid)
        
        columns_info = QLabel("<small>Toggle columns to show or hide them in the Result Preview table.</small>")
        columns_info.setWordWrap(True)
        columns_info.setStyleSheet("color:#666; padding:5px;")
        columns_layout.addWidget(columns_info)
        
        columns_group.setLayout(columns_layout)
        appearance_layout.addWidget(columns_group)

        # Add stretch to appearance tab
        appearance_layout.addStretch()
        tabs.addTab(appearance_widget, "Appearance")

        # ===== TAB 2: FOLDER LOCATIONS =====
        folders_widget = QWidget()
        folders_layout = QVBoxLayout(folders_widget)

        group = QGroupBox("Folder Locations")
        glayout = QFormLayout()
        
        # Input folder display and button
        global INPUT_DIR, OUTPUT_DIR
        input_dir_str = str(INPUT_DIR) if 'INPUT_DIR' in globals() and INPUT_DIR else "(not set)"
        output_dir_str = str(OUTPUT_DIR) if 'OUTPUT_DIR' in globals() and OUTPUT_DIR else "(not set)"

        # Helper function to create path display widget
        def create_path_display(path_str):
            """Create a read-only text edit for displaying file paths"""
            from PyQt5.QtGui import QPalette

            text_edit = QPlainTextEdit()
            text_edit.setPlainText(path_str)
            text_edit.setReadOnly(True)
            text_edit.setFixedHeight(45)

            # Apply theme-aware styling using result preview background color
            app = QApplication.instance()
            palette = app.palette()
            base_color = palette.color(QPalette.Base)
            text_color = palette.color(QPalette.Text)

            # Format colors for stylesheet
            bg_color = base_color.name()
            fg_color = text_color.name()

            text_edit.setStyleSheet(f"background:{bg_color}; padding:5px; border:1px solid #555; color:{fg_color}; font-family: monospace;")

            return text_edit

        input_path_display = create_path_display(input_dir_str)

        input_btn = QPushButton("Change Input Folder")
        input_btn.clicked.connect(lambda: self.select_input_folder(input_path_display))
        glayout.addRow("Input Folder:", input_path_display)
        glayout.addRow("", input_btn)

        # Output folder display and button
        output_path_display = create_path_display(output_dir_str)

        output_btn = QPushButton("Change Output Folder")
        output_btn.clicked.connect(lambda: self.select_output_folder(output_path_display))
        glayout.addRow("Output Folder:", output_path_display)
        glayout.addRow("", output_btn)

        group.setLayout(glayout)
        folders_layout.addWidget(group)

        folders_layout.addStretch()
        tabs.addTab(folders_widget, "Folders")

        # ===== TAB 3: DATABASE =====
        database_widget = QWidget()
        database_layout = QVBoxLayout(database_widget)

        # Current Database Info
        db_info_group = QGroupBox("Current Database")
        db_info_layout = QFormLayout()

        db_path_label = QLabel(str(DB_PATH))
        db_path_label.setWordWrap(True)
        db_path_label.setStyleSheet("font-family: monospace; font-size: 9pt;")

        # Check if using shared or local database
        config = load_shared_config()
        if config.has_option('Database', 'path'):
            db_type_text = "Shared (Network)"
        else:
            db_type_text = "Local"
        db_type_label = QLabel(db_type_text)

        db_info_layout.addRow("Type:", db_type_label)
        db_info_layout.addRow("Location:", db_path_label)

        db_info_group.setLayout(db_info_layout)
        database_layout.addWidget(db_info_group)

        # Shared Database Configuration
        shared_group = QGroupBox("Shared Database (Multi-User)")
        shared_layout = QVBoxLayout()

        shared_info = QLabel(
            "Configure a shared database location for multiple users.\n"
            "This setting is stored in config.ini next to the application and applies to all users.\n\n"
            "User-specific settings (window size, theme, etc.) remain personal."
        )
        shared_info.setWordWrap(True)
        shared_layout.addWidget(shared_info)

        # Path input row
        path_row = QHBoxLayout()
        shared_path_input = QLineEdit()
        shared_path_input.setPlaceholderText("e.g., \\\\server\\share\\derivativemill.db")
        if config.has_option('Database', 'path'):
            shared_path_input.setText(config.get('Database', 'path'))
        path_row.addWidget(shared_path_input)

        browse_btn = QPushButton("Browse...")
        def browse_database():
            path, _ = QFileDialog.getOpenFileName(
                dialog, "Select Database File", 
                str(Path.home()),
                "SQLite Database (*.db);;All Files (*.*)"
            )
            if path:
                shared_path_input.setText(path)
        browse_btn.clicked.connect(browse_database)
        path_row.addWidget(browse_btn)
        shared_layout.addLayout(path_row)

        # Action buttons
        btn_row = QHBoxLayout()

        apply_btn = QPushButton("Apply Shared Path")
        apply_btn.setStyleSheet(self.get_button_style("success"))
        def apply_shared_path():
            new_path = shared_path_input.text().strip()
            if not new_path:
                QMessageBox.warning(dialog, "No Path", "Please enter a database path.")
                return
            
            path_obj = Path(new_path)
            if not path_obj.exists():
                reply = QMessageBox.question(dialog, "Database Not Found",
                    f"The file does not exist:\n{new_path}\n\n"
                    "Would you like to create a new database at this location?\n"
                    "(A copy of your current database will be created)",
                    QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.Yes:
                    try:
                        # Copy current database to new location
                        path_obj.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(str(DB_PATH), str(path_obj))
                    except Exception as e:
                        QMessageBox.critical(dialog, "Error", f"Failed to create database:\n{e}")
                        return
                else:
                    return
            
            # Save to config.ini
            set_database_path(new_path)
            db_path_label.setText(new_path)
            db_type_label.setText("Shared (Network)")
            QMessageBox.information(dialog, "Success", 
                f"Database path updated to:\n{new_path}\n\n"
                "Restart the application to use the new database.")
        apply_btn.clicked.connect(apply_shared_path)
        btn_row.addWidget(apply_btn)

        reset_btn = QPushButton("Use Local Database")
        reset_btn.setStyleSheet(self.get_button_style("warning"))
        def reset_to_local():
            config = load_shared_config()
            if config.has_option('Database', 'path'):
                config.remove_option('Database', 'path')
                save_shared_config(config)
            shared_path_input.clear()
            local_path = RESOURCES_DIR / DB_NAME
            db_path_label.setText(str(local_path))
            db_type_label.setText("Local")
            QMessageBox.information(dialog, "Reset", 
                "Database reset to local.\n\nRestart the application to apply changes.")
        reset_btn.clicked.connect(reset_to_local)
        btn_row.addWidget(reset_btn)

        shared_layout.addLayout(btn_row)

        # Warning about concurrent access
        warning_label = QLabel(
            "<small><b>Note:</b> SQLite on network shares works best for sequential access. "
            "Avoid having multiple users edit the same record simultaneously.</small>"
        )
        warning_label.setWordWrap(True)
        warning_label.setStyleSheet("color:#666; padding:5px;")
        shared_layout.addWidget(warning_label)

        shared_group.setLayout(shared_layout)
        database_layout.addWidget(shared_group)

        database_layout.addStretch()
        tabs.addTab(database_widget, "Database")

        # ===== TAB 4: UPDATES =====
        updates_widget = QWidget()
        updates_layout = QVBoxLayout(updates_widget)

        # Update Settings Group
        update_settings_group = QGroupBox("Automatic Update Checks")
        update_settings_layout = QVBoxLayout()

        # Checkbox for startup update check
        startup_check_cb = QCheckBox("Check for updates when application starts")

        # Load saved preference from per-user settings
        startup_check_cb.setChecked(get_user_setting_bool('check_updates_on_startup', False))

        def save_startup_check_preference(checked):
            set_user_setting('check_updates_on_startup', '1' if checked else '0')
            logger.info(f"Startup update check preference: {'enabled' if checked else 'disabled'}")
        
        startup_check_cb.stateChanged.connect(lambda state: save_startup_check_preference(state == Qt.Checked))
        update_settings_layout.addWidget(startup_check_cb)

        update_info = QLabel(
            "<small>When enabled, the application will check for new releases on GitHub when it starts. "
            "No personal data is sent - only a simple API request to check the latest version.</small>"
        )
        update_info.setWordWrap(True)
        update_info.setStyleSheet("color:#666; padding:5px;")
        update_settings_layout.addWidget(update_info)

        update_settings_group.setLayout(update_settings_layout)
        updates_layout.addWidget(update_settings_group)

        # Version Info Group
        version_group = QGroupBox("Version Information")
        version_layout = QFormLayout()

        current_version_label = QLabel(f"<b>{VERSION}</b>")
        version_layout.addRow("Current Version:", current_version_label)

        github_link = QLabel(f'<a href="{GITHUB_RELEASES_URL}">View all releases on GitHub</a>')
        github_link.setOpenExternalLinks(True)
        version_layout.addRow("Releases:", github_link)

        version_group.setLayout(version_layout)
        updates_layout.addWidget(version_group)

        # Check Now Button
        check_now_btn = QPushButton("Check for Updates Now")
        check_now_btn.setStyleSheet(self.get_button_style("success"))
        check_now_btn.clicked.connect(lambda: (dialog.close(), self.check_for_updates_manual()))
        updates_layout.addWidget(check_now_btn)

        updates_layout.addStretch()
        tabs.addTab(updates_widget, "Updates")

        # Add tabs to main dialog layout
        layout.addWidget(tabs)
        dialog.exec_()

    def apply_theme(self, theme_name):
        """Apply the selected theme to the application"""
        app = QApplication.instance()
        
        # Store current theme name
        self.current_theme = theme_name
        
        if theme_name == "System Default":
            app.setStyle("")
            app.setPalette(app.style().standardPalette())
        elif theme_name == "Fusion (Light)":
            app.setStyle("Fusion")
            app.setPalette(app.style().standardPalette())
        elif theme_name == "Windows":
            app.setStyle("Windows")
            app.setPalette(app.style().standardPalette())
        elif theme_name == "Fusion (Dark)":
            app.setStyle("Fusion")
            dark_palette = self.get_dark_palette()
            app.setPalette(dark_palette)
        elif theme_name == "Ocean":
            app.setStyle("Fusion")
            ocean_palette = self.get_ocean_palette()
            app.setPalette(ocean_palette)
        elif theme_name == "Light Cyan":
            app.setStyle("Fusion")
            teal_palette = self.get_teal_professional_palette()
            app.setPalette(teal_palette)
        
        # Refresh button styles to match new theme
        self.refresh_button_styles()

        # Refresh text input styles to match new theme
        self.refresh_input_styles()

        # Update file label style for new theme
        if hasattr(self, 'file_label'):
            self.update_file_label_style()

        # Update status bar styles for new theme
        self.update_status_bar_styles()

        # Update table stylesheet for new theme
        self.update_table_stylesheet()
        
        # Save theme preference (per-user setting)
        set_user_setting('theme', theme_name)
        logger.info(f"Theme changed to: {theme_name}")

    def apply_font_size(self, size):
        """Apply the selected font size to the application"""
        app = QApplication.instance()

        # Get current font and update size
        font = app.font()
        font.setPointSize(size)
        app.setFont(font)

        # Save font size preference (per-user setting)
        set_user_setting('font_size', size)
        logger.info(f"Font size changed to: {size}pt")

    def apply_row_height(self, height):
        """Apply the selected row height to the Result Preview table"""
        if hasattr(self, 'table'):
            self.table.verticalHeader().setDefaultSectionSize(height)
            # Also update existing rows
            for row in range(self.table.rowCount()):
                self.table.setRowHeight(row, height)

        # Save row height preference (per-user setting)
        set_user_setting('preview_row_height', height)
        logger.info(f"Preview row height changed to: {height}px")

    def check_for_updates_manual(self):
        """Manually check for updates and show result dialog"""
        self.bottom_status.setText("Checking for updates...")
        QApplication.processEvents()
        
        checker = UpdateChecker(VERSION)
        has_update, latest, url, notes, download_url, error = checker.check_for_updates()
        
        if error:
            QMessageBox.warning(self, "Update Check Failed",
                f"Could not check for updates.\n\n{error}\n\n"
                f"You can check manually at:\n{GITHUB_RELEASES_URL}")
            self.bottom_status.setText("Ready")
            return
        
        if has_update:
            self.show_update_available_dialog(latest, url, notes, download_url)
        else:
            QMessageBox.information(self, "No Updates Available",
                f"You are running the latest version.\n\n"
                f"Current version: {VERSION}\n"
                f"Latest version: {latest}")
        
        self.bottom_status.setText("Ready")

    def show_update_available_dialog(self, latest_version, release_url, release_notes, download_url):
        """Show dialog when an update is available"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Update Available")
        dialog.resize(500, 400)
        layout = QVBoxLayout(dialog)

        # Header
        header = QLabel(f"<h2>🎉 New Version Available!</h2>")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        # Version info
        version_info = QLabel(
            f"<p><b>Current version:</b> {VERSION}<br>"
            f"<b>New version:</b> {latest_version}</p>"
        )
        version_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_info)

        # Release notes
        notes_group = QGroupBox("Release Notes")
        notes_layout = QVBoxLayout()
        notes_text = QTextEdit()
        notes_text.setPlainText(release_notes if release_notes else "No release notes available.")
        notes_text.setReadOnly(True)
        notes_layout.addWidget(notes_text)
        notes_group.setLayout(notes_layout)
        layout.addWidget(notes_group)

        # Buttons
        btn_layout = QHBoxLayout()

        if download_url:
            download_btn = QPushButton("Download Update")
            download_btn.setStyleSheet(self.get_button_style("success"))
            download_btn.clicked.connect(lambda: (webbrowser.open(download_url), dialog.accept()))
            btn_layout.addWidget(download_btn)

        view_btn = QPushButton("View on GitHub")
        view_btn.clicked.connect(lambda: (webbrowser.open(release_url), dialog.accept()))
        btn_layout.addWidget(view_btn)

        later_btn = QPushButton("Remind Me Later")
        later_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(later_btn)

        layout.addLayout(btn_layout)
        dialog.exec_()

    def check_for_updates_startup(self):
        """Check for updates on startup (runs in background thread)"""
        # Check if startup update check is enabled (per-user setting)
        if not get_user_setting_bool('check_updates_on_startup', False):
            logger.debug("Startup update check disabled")
            return

        logger.info("Checking for updates on startup...")
        
        def check_thread():
            checker = UpdateChecker(VERSION)
            has_update, latest, url, notes, download_url, error = checker.check_for_updates()
            if has_update and not error:
                # Schedule dialog to be shown on main thread
                QTimer.singleShot(0, lambda: self.show_update_available_dialog(
                    latest, url, notes, download_url))
        
        # Run in background thread to not block startup
        thread = Thread(target=check_thread, daemon=True)
        thread.start()

    def show_about_dialog(self):
        """Show the About dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"About {APP_NAME}")
        dialog.resize(400, 300)
        layout = QVBoxLayout(dialog)

        # App icon and name
        header_layout = QHBoxLayout()
        
        # Try to load app icon
        icon_path = TEMP_RESOURCES_DIR / "icon.ico"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "banner_bg.png"
        
        if icon_path.exists():
            icon_label = QLabel()
            pixmap = QPixmap(str(icon_path))
            icon_label.setPixmap(pixmap.scaled(64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            header_layout.addWidget(icon_label)
        
        title_layout = QVBoxLayout()
        title_label = QLabel(f"<h2>{APP_NAME}</h2>")
        title_layout.addWidget(title_label)
        version_label = QLabel(f"Version {VERSION}")
        version_label.setStyleSheet("color: #666;")
        title_layout.addWidget(version_label)
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        
        layout.addLayout(header_layout)
        layout.addSpacing(20)

        # Description
        desc_label = QLabel(
            "<p>Professional customs documentation processing system for "
            "invoice processing, parts management, and Section 232 tariff "
            "compliance tracking.</p>"
        )
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)

        layout.addSpacing(10)

        # Links
        github_label = QLabel(
            f'<p>GitHub: <a href="{GITHUB_RELEASES_URL}">{GITHUB_REPO}</a></p>'
        )
        github_label.setOpenExternalLinks(True)
        layout.addWidget(github_label)

        layout.addStretch()

        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignCenter)

        dialog.exec_()

    def update_file_label_style(self):
        """Update file label background based on current theme"""
        from PyQt5.QtGui import QPalette

        if not hasattr(self, 'current_theme'):
            self.current_theme = "System Default"

        # Use palette Base color for consistent styling with result preview
        app = QApplication.instance()
        palette = app.palette()
        base_color = palette.color(QPalette.Base)
        text_color = palette.color(QPalette.Text)

        bg_color = base_color.name()
        fg_color = text_color.name()

        self.file_label.setStyleSheet(f"background:{bg_color}; padding:5px; border:1px solid #555; color:{fg_color};")
    
    def update_status_bar_styles(self):
        """Update status bar backgrounds based on current theme"""
        if not hasattr(self, 'current_theme'):
            self.current_theme = "System Default"
        
        is_dark = self.current_theme in ["Fusion (Dark)", "Ocean"]
        
        if is_dark:
            # Dark theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#2d2d2d; color:#e0e0e0;")
            self.bottom_status.setStyleSheet("font-size:9pt; color:#b0b0b0;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #2d2d2d;
                        border-top: 1px solid #404040;
                    }
                """)
        else:
            # Light theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0; color:#000000;")
            self.bottom_status.setStyleSheet("font-size:9pt; color:#555555;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #e8e8e8;
                        border-top: 1px solid #d0d0d0;
                    }
                """)
    
    def update_status_bar_styles(self):
        """Update status bar backgrounds based on current theme"""
        if not hasattr(self, 'current_theme'):
            self.current_theme = "System Default"
        
        is_dark = self.current_theme in ["Fusion (Dark)", "Ocean"]
        
        if is_dark:
            # Dark theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#2d2d2d; color:#e0e0e0;")
            self.bottom_status.setStyleSheet("font-size:9pt; color:#b0b0b0;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #2d2d2d;
                        border-top: 1px solid #404040;
                    }
                """)
        else:
            # Light theme status bars
            self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0; color:#000000;")
            self.bottom_status.setStyleSheet("font-size:9pt; color:#555555;")
            if hasattr(self, 'bottom_bar'):
                self.bottom_bar.setStyleSheet("""
                    QWidget {
                        background: #e8e8e8;
                        border-top: 1px solid #d0d0d0;
                    }
                """)

    def update_table_stylesheet(self):
        """Update table stylesheet with green focus color for current theme"""
        if not hasattr(self, 'table'):
            return

        # Set green focus color and reduced cell padding
        self.table.setStyleSheet("""
            QTableWidget::item {
                padding: 1px 3px;
            }
            QTableWidget::item:focus {
                background-color: #90EE90;
                border: 2px solid #228B22;
            }
        """)

    def get_preview_row_color(self, material_flag):
        """Get the color for preview table rows based on Section 232 material type

        Args:
            material_flag: String like '232_Steel', '232_Aluminum', '232_Copper', '232_Wood', '232_Auto', 'Non_232', or boolean for backward compatibility

        Returns:
            QColor: Color for the row based on material type
        """
        # Default colors for each material type
        default_colors = {
            '232_Steel': '#4a4a4a',      # Dark gray
            '232_Aluminum': '#3498db',   # Blue
            '232_Copper': '#e67e22',     # Orange
            '232_Wood': '#27ae60',       # Green
            '232_Auto': '#9b59b6',       # Purple
            'Non_232': '#ff0000'         # Red
        }

        # Handle backward compatibility - if passed a boolean
        if isinstance(material_flag, bool):
            material_flag = '232_Steel' if material_flag else 'Non_232'

        # Determine material type from flag
        if not material_flag or material_flag == 'Non_232':
            color_key = 'preview_non232_color'
            default_color = default_colors['Non_232']
        elif material_flag.startswith('232_'):
            # Map flag to color key
            material = material_flag  # e.g., '232_Steel'
            color_key = f'preview_{material.lower().replace("232_", "")}_color'  # e.g., 'preview_steel_color'
            default_color = default_colors.get(material, default_colors['232_Steel'])
        else:
            # Unknown flag, treat as non-232
            color_key = 'preview_non232_color'
            default_color = default_colors['Non_232']

        # Get color from per-user settings
        saved_color = get_user_setting(color_key, default_color)
        return QColor(saved_color)

    def get_sec301_bg_color(self):
        """Get the background color for Section 301 exclusion rows (per-user setting)

        Returns:
            QColor: Background color for Sec301 exclusion rows
        """
        default_color = '#ffc8c8'  # Light red background
        saved_color = get_user_setting('preview_sec301_bg_color', default_color)
        return QColor(saved_color)

    def refresh_preview_colors(self):
        """Refresh all row colors in the preview table based on current settings"""
        if not hasattr(self, 'table') or self.table.rowCount() == 0:
            return

        try:
            # Temporarily disconnect itemChanged signal to avoid triggering edits
            self.table.blockSignals(True)

            for row in range(self.table.rowCount()):
                # Check the 232 Status column (index 16) to determine material type
                status_item = self.table.item(row, 16)
                status_text = status_item.text() if status_item else ''

                # Get color based on material flag (232_Steel, 232_Aluminum, etc.)
                row_color = self.get_preview_row_color(status_text)

                # Update color for all items in this row
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item:
                        item.setForeground(row_color)

            self.table.blockSignals(False)
        except Exception as e:
            logger.error(f"Error refreshing preview colors: {e}")
            self.table.blockSignals(False)

    def apply_column_visibility(self):
        """Apply saved column visibility settings to the preview table"""
        if not hasattr(self, 'table'):
            return
        
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            
            # First, check if we have all 15 column settings saved
            # If not, reset all to visible (handles version upgrades with new columns)
            c.execute("SELECT COUNT(*) FROM app_config WHERE key LIKE 'preview_col_visible_%'")
            count_row = c.fetchone()
            saved_count = count_row[0] if count_row else 0
            
            if saved_count < 17:
                # Clear old settings and reset all columns to visible
                c.execute("DELETE FROM app_config WHERE key LIKE 'preview_col_visible_%'")
                for col_idx in range(17):
                    c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                              (f'preview_col_visible_{col_idx}', '1'))
                conn.commit()
                logger.info(f"Reset column visibility settings (had {saved_count}, need 17)")
                # Show all columns
                for col_idx in range(self.table.columnCount()):
                    self.table.setColumnHidden(col_idx, False)
            else:
                # Apply saved settings
                for col_idx in range(self.table.columnCount()):
                    config_key = f'preview_col_visible_{col_idx}'
                    c.execute("SELECT value FROM app_config WHERE key = ?", (config_key,))
                    row = c.fetchone()
                    # Default to visible if no setting saved
                    is_visible = True if row is None else (row[0] == '1')
                    self.table.setColumnHidden(col_idx, not is_visible)
            
            conn.close()
            
            conn.close()
        except Exception as e:
            logger.error(f"Error applying column visibility: {e}")

    def refresh_button_styles(self):
        """Refresh all button styles to match current theme"""
        # Process tab buttons
        if hasattr(self, 'clear_btn'):
            self.clear_btn.setStyleSheet(self.get_button_style("danger"))
        if hasattr(self, 'process_btn'):
            self.process_btn.setStyleSheet(self.get_button_style("success"))
        if hasattr(self, 'edit_values_btn'):
            self.edit_values_btn.setStyleSheet(self.get_button_style("warning"))
        
        # Import tab - need to find and update buttons in the tab
        if hasattr(self, 'tab_import'):
            for btn in self.tab_import.findChildren(QPushButton):
                if btn.text() == "Load CSV File":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Reset Mapping":
                    btn.setStyleSheet(self.get_button_style("danger"))
                elif btn.text() == "IMPORT NOW":
                    btn.setStyleSheet(self.get_button_style("success") + "QPushButton { font-size:16pt; padding:15px; }")
        
        # Shipment Mapping tab
        if hasattr(self, 'tab_shipment_map'):
            for btn in self.tab_shipment_map.findChildren(QPushButton):
                if btn.text() == "Save Current Mapping As...":
                    btn.setStyleSheet(self.get_button_style("success"))
                elif btn.text() == "Delete Profile":
                    btn.setStyleSheet(self.get_button_style("danger"))
                elif btn.text() == "Load CSV to Map":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Reset Current":
                    btn.setStyleSheet(self.get_button_style("danger"))
        
        # Master/Parts View tab
        if hasattr(self, 'tab_master'):
            for btn in self.tab_master.findChildren(QPushButton):
                if btn.text() == "Run Query":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Execute":
                    btn.setStyleSheet(self.get_button_style("success"))
        
        # Config tab
        if hasattr(self, 'tab_config'):
            for btn in self.tab_config.findChildren(QPushButton):
                if btn.text() == "Import Section 232 Tariffs (CSV/Excel)":
                    btn.setStyleSheet(self.get_button_style("success"))
                elif btn.text() == "Import from CSV":
                    btn.setStyleSheet(self.get_button_style("info"))
                elif btn.text() == "Refresh View":
                    btn.setStyleSheet(self.get_button_style("info"))

    def refresh_input_styles(self):
        """Refresh all text input styles to match current theme"""
        input_style = self.get_input_style()

        # Parts View tab inputs
        if hasattr(self, 'query_value'):
            self.query_value.setStyleSheet(input_style)
        if hasattr(self, 'custom_sql_input'):
            self.custom_sql_input.setStyleSheet(input_style)
        if hasattr(self, 'search_input'):
            self.search_input.setStyleSheet(input_style)

        # Customs Config tab inputs
        if hasattr(self, 'tariff_filter'):
            self.tariff_filter.setStyleSheet(input_style)

        # Section 232 Actions tab inputs
        if hasattr(self, 'actions_filter'):
            self.actions_filter.setStyleSheet(input_style)

    def get_dark_palette(self):
        """Create a Windows 11 dark mode inspired theme"""
        from PyQt5.QtGui import QPalette, QColor
        
        palette = QPalette()
        # Windows 11 dark theme colors
        palette.setColor(QPalette.Window, QColor(41, 41, 41))  # Main background
        palette.setColor(QPalette.WindowText, QColor(243, 243, 243))  # Primary text
        palette.setColor(QPalette.Base, QColor(85, 83, 83))  # Result preview background (#555353)
        palette.setColor(QPalette.AlternateBase, QColor(115, 115, 115))  # Tertiary background for alternating rows
        palette.setColor(QPalette.ToolTipBase, QColor(45, 45, 45))  # Tertiary background
        palette.setColor(QPalette.ToolTipText, QColor(243, 243, 243))  # Primary text
        palette.setColor(QPalette.Text, QColor(243, 243, 243))  # Primary text in text boxes
        palette.setColor(QPalette.Button, QColor(45, 45, 45))  # Tertiary background for buttons
        palette.setColor(QPalette.ButtonText, QColor(243, 243, 243))  # Primary text on buttons
        palette.setColor(QPalette.BrightText, QColor(164, 38, 44))  # Danger/error red
        palette.setColor(QPalette.Link, QColor(0, 120, 212))  # Accent blue
        palette.setColor(QPalette.Highlight, QColor(22, 120, 212))  # Selection/highlight blue
        palette.setColor(QPalette.HighlightedText, QColor(243, 243, 243))  # Primary text
        return palette

    def get_ocean_palette(self):
        """Create an ocean-themed color palette with deep blues and teals"""
        from PyQt5.QtGui import QPalette, QColor

        palette = QPalette()
        # Deep ocean blue backgrounds
        palette.setColor(QPalette.Window, QColor(28, 57, 87))  # Deep ocean blue
        palette.setColor(QPalette.WindowText, QColor(230, 245, 255))  # Light blue-white text
        palette.setColor(QPalette.Base, QColor(28, 56, 86))  # Result preview background (#1C3856)
        palette.setColor(QPalette.AlternateBase, QColor(35, 65, 95))  # Lighter ocean blue
        palette.setColor(QPalette.ToolTipBase, QColor(200, 230, 255))  # Light blue
        palette.setColor(QPalette.ToolTipText, QColor(15, 35, 55))  # Dark blue
        palette.setColor(QPalette.Text, QColor(230, 245, 255))  # Light blue-white text
        palette.setColor(QPalette.Button, QColor(40, 75, 110))  # Medium ocean blue
        palette.setColor(QPalette.ButtonText, QColor(230, 245, 255))  # Light text
        palette.setColor(QPalette.BrightText, QColor(0, 255, 200))  # Bright teal
        palette.setColor(QPalette.Link, QColor(100, 200, 255))  # Bright cyan
        palette.setColor(QPalette.Highlight, QColor(0, 150, 180))  # Teal highlight
        palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))  # White text
        return palette
    
    def get_teal_professional_palette(self):
        """Create a Light Cyan palette based on Fusion Light with custom colors"""
        from PyQt5.QtGui import QPalette, QColor
        from PyQt5.QtWidgets import QApplication

        # Start with standard Fusion Light palette
        app = QApplication.instance()
        app.setStyle("Fusion")
        palette = app.style().standardPalette()

        # Override button and column header colors with custom teal-cyan
        palette.setColor(QPalette.Button, QColor(224, 246, 247))  # Light cyan for buttons (#E0F6F7)
        palette.setColor(QPalette.ButtonText, QColor(33, 33, 33))  # Dark text on buttons
        palette.setColor(QPalette.Window, QColor(224, 246, 247))  # Light cyan background (#E0F6F7)
        palette.setColor(QPalette.Base, QColor(239, 249, 249))  # Result preview background (#EFF9F9)
        palette.setColor(QPalette.Mid, QColor(206, 243, 245))  # Light cyan for column headers (#CEF3F5)

        return palette

    def get_button_style(self, button_type="default"):
        """
        Generate theme-aware button styles using current palette colors.

        Args:
            button_type: "default", "primary", "danger", "info", "warning", "success"

        Returns:
            CSS stylesheet string that adapts to current theme
        """
        from PyQt5.QtGui import QPalette, QColor
        from PyQt5.QtWidgets import QApplication

        palette = QApplication.palette()

        # Get base colors from theme
        base_bg = palette.color(QPalette.Button)
        base_text = palette.color(QPalette.ButtonText)
        highlight = palette.color(QPalette.Highlight)

        # All buttons use logo blue color across all themes
        bg = QColor(52, 152, 219)  # Logo Blue
        hover_bg = QColor(41, 128, 185)  # Darker Logo Blue
        disabled_bg = QColor(160, 160, 160)  # Grey

        # Text color - white for dark buttons, black for light buttons
        text_color = QColor(255, 255, 255) if bg.lightness() < 128 else QColor(0, 0, 0)
        
        return f"""
            QPushButton {{
                background-color: rgb({bg.red()}, {bg.green()}, {bg.blue()});
                color: rgb({text_color.red()}, {text_color.green()}, {text_color.blue()});
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: rgb({hover_bg.red()}, {hover_bg.green()}, {hover_bg.blue()});
            }}
            QPushButton:disabled {{
                background-color: rgb({disabled_bg.red()}, {disabled_bg.green()}, {disabled_bg.blue()});
            }}
        """

    def get_input_style(self):
        """
        Generate theme-aware text input (QLineEdit) styles.

        Returns:
            CSS stylesheet string that adapts background color to current theme
        """
        # Check if we're in a dark theme
        is_dark_theme = hasattr(self, 'current_theme') and self.current_theme in ["Fusion (Dark)", "Ocean"]

        if is_dark_theme:
            # Dark theme: dark background with light text
            return "QLineEdit { color: #e0e0e0; background-color: #2d2d2d; padding: 5px; border: 1px solid #555; }"
        else:
            # Light theme: light background with dark text
            return "QLineEdit { color: #000000; background-color: #f5f5f5; padding: 5px; border: 1px solid #ccc; }"

    def clear_all(self):
        self.current_csv = None
        self.file_label.setText("No file selected")
        self.ci_input.clear()
        self.wt_input.clear()
        self.mid_combo.setCurrentIndex(-1)
        self.selected_mid = ""
        self.table.setRowCount(0)
        self.process_btn.setEnabled(False)
        self.process_btn.setText("Process Invoice")  # Reset button text
        self.progress.setVisible(False)
        self.invoice_check_label.setText("No file loaded")
        self.csv_total_value = 0.0
        self.edit_values_btn.setVisible(False)
        self.bottom_status.setText("Cleared")
        self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0;")
        logger.info("Process tab cleared")

    def browse_file(self):
        # Simple profile check without focus manipulation
        current_profile = self.profile_combo.currentText()
        if not current_profile or current_profile == "-- Select Profile --":
            QMessageBox.warning(
                self,
                "Mapping Profile Required",
                "<b>You must select a mapping profile before loading a shipment file.</b><br><br>"
                "Please choose one from the <b>Saved Profiles</b> dropdown.",
                QMessageBox.Ok
            )
            # Re-enable input fields after modal dialog
            self._enable_input_fields()
            return
        # -------------------------------------------------------------------------

        path, _ = QFileDialog.getOpenFileName(
            self, "Select Shipment File", str(INPUT_DIR), "CSV/Excel (*.csv *.xlsx)"
        )
        if not path:
            return

        self.current_csv = path
        self.file_label.setText(Path(path).name)
        
        # Clear previous processing state when loading new file
        self.last_processed_df = None
        self.table.setRowCount(0)

        try:
            # Get header row value from input field
            header_row = 0  # Default: first row is header
            if hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
                try:
                    header_row_value = int(self.header_row_input.text().strip())
                    # Convert from 1-based to 0-based indexing
                    # If user enters 1, header is at row 0
                    # If user enters 2, skip 1 row, header is at row 1
                    header_row = max(0, header_row_value - 1)
                except ValueError:
                    header_row = 0

            col_map = {v: k for k, v in self.shipment_mapping.items()}
            if Path(path).suffix.lower() == ".xlsx":
                df = pd.read_excel(path, dtype=str, header=header_row)
            else:
                df = pd.read_csv(path, dtype=str, header=header_row)

            # Calculate total before renaming (using original column names)
            # Only sum rows that have a part number to exclude total/subtotal rows
            value_column = None
            part_number_column = None

            if 'value_usd' in self.shipment_mapping:
                # Get the original column name mapped to value_usd
                original_col_name = self.shipment_mapping['value_usd']
                logger.info(f"[INVOICE TOTAL] Looking for value column: '{original_col_name}' in columns: {df.columns.tolist()}")
                if original_col_name in df.columns:
                    value_column = original_col_name
                    logger.info(f"[INVOICE TOTAL] Found value column: '{value_column}'")
                else:
                    logger.warning(f"[INVOICE TOTAL] Value column '{original_col_name}' not found in file columns: {df.columns.tolist()}")
            else:
                logger.warning(f"[INVOICE TOTAL] 'value_usd' not mapped in shipment_mapping: {self.shipment_mapping}")

            # Get part number column to filter rows
            if 'part_number' in self.shipment_mapping:
                part_number_col_name = self.shipment_mapping['part_number']
                if part_number_col_name in df.columns:
                    part_number_column = part_number_col_name
                else:
                    logger.warning(f"Part number column '{part_number_col_name}' not found in file columns: {df.columns.tolist()}")

            # If we found the value column, calculate total
            if value_column:
                # Filter to only rows that have a part number (exclude total/subtotal rows)
                if part_number_column:
                    df_filtered = df[df[part_number_column].notna() & (df[part_number_column].astype(str).str.strip() != '')]
                    logger.debug(f"Filtered {len(df)} rows to {len(df_filtered)} rows with part numbers")
                    total = pd.to_numeric(df_filtered[value_column], errors='coerce').sum()
                else:
                    # If no part number column, sum all rows (old behavior)
                    total = pd.to_numeric(df[value_column], errors='coerce').sum()

                self.csv_total_value = round(total, 2)
                logger.info(f"Calculated invoice total: ${self.csv_total_value:,.2f}")

            # Now rename columns for other uses
            df = df.rename(columns=col_map)

            # Update invoice check - this will display the total and control button state
            if value_column:
                self.update_invoice_check()
        except Exception as e:
            logger.error(f"browse_file value read failed: {e}")
            self.invoice_check_label.setText("Could not read value column")

        logger.info(f"Loaded: {Path(path).name}")

        # Force focus to ci_input and ensure keyboard input works
        QApplication.processEvents()  # Process any pending events first
        self.ci_input.setFocus(Qt.OtherFocusReason)
        self.ci_input.activateWindow()
        logger.info(f"Set focus to ci_input: hasFocus={self.ci_input.hasFocus()}")
        
    def update_invoice_check(self):
        # Guard against re-entrancy
        if getattr(self, '_updating_invoice_check', False):
            return
        self._updating_invoice_check = True

        try:
            if not self.current_csv:
                self.invoice_check_label.setWordWrap(False)  # Single line display
                self.invoice_check_label.setText("No file loaded")
                # Gold color in dark theme (text-shadow not supported in Qt)
                if hasattr(self, 'current_theme') and self.current_theme in ["Fusion (Dark)", "Ocean"]:
                    self.invoice_check_label.setStyleSheet("background:transparent; color: gold; font-weight:bold; font-size:7pt; padding:3px;")
                else:
                    self.invoice_check_label.setStyleSheet("background:transparent; color: #A4262C; font-weight:bold; font-size:7pt; padding:3px;")
                self.edit_values_btn.setVisible(False)
                return

            user_text = self.ci_input.text().replace(',', '').strip()
            try:
                user_val = float(user_text) if user_text else 0.0
            except:
                user_val = 0.0

            diff = abs(user_val - self.csv_total_value)
            threshold = 0.01

            # Update the invoice check label and Edit Values button
            if user_val == 0.0:
                self.invoice_check_label.setWordWrap(False)  # Single line display
                self.invoice_check_label.setText(f"Total: ${self.csv_total_value:,.2f}")
                self.invoice_check_label.setStyleSheet("background:#0078D4; color:white; font-weight:bold; font-size:7pt; padding:3px;")
                self.edit_values_btn.setVisible(False)
            elif diff <= threshold:
                self.invoice_check_label.setWordWrap(False)  # Single line display
                self.invoice_check_label.setText(f"✓ Match: ${self.csv_total_value:,.2f}")
                self.invoice_check_label.setStyleSheet("background:#107C10; color:white; font-weight:bold; font-size:7pt; padding:3px;")
                self.edit_values_btn.setVisible(False)
            else:
                # Values don't match - show comparison and Edit Values button (two lines)
                self.invoice_check_label.setWordWrap(True)  # Allow two lines for mismatch display
                self.invoice_check_label.setText(
                    f"Total: ${self.csv_total_value:,.2f}\n"
                    f"Difference: ${diff:,.2f}"
                )
                self.invoice_check_label.setStyleSheet("background:#ff9800; color:white; font-weight:bold; font-size:7pt; padding:3px;")
                # Show Edit Values button only if haven't processed yet
                if self.last_processed_df is None:
                    self.edit_values_btn.setVisible(True)
                else:
                    self.edit_values_btn.setVisible(False)
            
            # Button state control - require invoice check match before processing
            has_weight = bool(self.wt_input.text().strip())
            has_ci_value = bool(user_text)
            has_mid = bool(self.selected_mid)
            values_match = diff <= threshold

            # Changed from >= 2 to >= 1 to allow profiles with minimal column mappings
            if self.current_csv and len(self.shipment_mapping) >= 1:
                if self.last_processed_df is None:
                    # Haven't processed yet - require weight, CI value, AND invoice values must match
                    # MID can be selected later, so not required for initial processing
                    if has_weight and has_ci_value and values_match:
                        self.process_btn.setEnabled(True)
                        self.process_btn.setText("Process Invoice")
                    else:
                        self.process_btn.setEnabled(False)
                        self.process_btn.setText("Process Invoice")
                else:
                    # Already processed - button becomes Export, only enabled when values match
                    if values_match:
                        self.process_btn.setEnabled(True)
                        self.process_btn.setText("Export Worksheet")
                    else:
                        self.process_btn.setEnabled(False)
                        self.process_btn.setText("Export Worksheet")

            # Always ensure input fields stay enabled
            self._enable_input_fields()
        finally:
            self._updating_invoice_check = False
    def select_input_folder(self, display_widget=None):
        global INPUT_DIR, PROCESSED_DIR
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder", str(INPUT_DIR))
        if folder:
            INPUT_DIR = Path(folder)
            PROCESSED_DIR = INPUT_DIR / "Processed"
            PROCESSED_DIR.mkdir(exist_ok=True)
            if display_widget:
                if isinstance(display_widget, QPlainTextEdit):
                    display_widget.setPlainText(str(INPUT_DIR))
                else:
                    display_widget.setText(str(INPUT_DIR))
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config VALUES ('input_dir', ?)", (str(INPUT_DIR),))
            conn.commit()
            conn.close()
            self.status.setText(f"Input folder: {INPUT_DIR}")
            self.refresh_input_files()

    def select_output_folder(self, display_widget=None):
        global OUTPUT_DIR
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", str(OUTPUT_DIR))
        if folder:
            OUTPUT_DIR = Path(folder)
            OUTPUT_DIR.mkdir(exist_ok=True)
            if display_widget:
                if isinstance(display_widget, QPlainTextEdit):
                    display_widget.setPlainText(str(OUTPUT_DIR))
                else:
                    display_widget.setText(str(OUTPUT_DIR))
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config VALUES ('output_dir', ?)", (str(OUTPUT_DIR),))
            conn.commit()
            conn.close()
            self.status.setText(f"Output folder: {OUTPUT_DIR}")
            self.refresh_exported_files()

    def move_pdf_to_processed(self, pdf_path):
        """
        Move a processed PDF file to the Processed PDF folder.

        Args:
            pdf_path (str or Path): Path to the PDF file to move

        Returns:
            bool: True if move was successful, False otherwise
        """
        try:
            pdf_file = Path(pdf_path)
            if not pdf_file.exists():
                logger.warning(f"PDF file not found for processing: {pdf_path}")
                return False

            # Get destination path
            dest_path = PROCESSED_DIR / pdf_file.name

            # Handle duplicate filenames
            if dest_path.exists():
                base_name = pdf_file.stem
                ext = pdf_file.suffix
                counter = 1
                while dest_path.exists():
                    dest_path = PROCESSED_DIR / f"{base_name}_{counter}{ext}"
                    counter += 1

            # Move the file
            shutil.move(str(pdf_file), str(dest_path))
            logger.info(f"Moved processed PDF to: {dest_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to move PDF to processed folder: {e}")
            return False

    def load_file_as_dataframe(self, file_path):
        """Load CSV or Excel file and return as DataFrame"""
        # Get header row value from profile or input field
        header_row = 0  # Default: first row is header
        # First check if there's a profile header row loaded
        if hasattr(self, 'profile_header_row') and self.profile_header_row:
            header_row = max(0, self.profile_header_row - 1)
        # Otherwise check input field (for Invoice Mapping Profiles tab)
        elif hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
            try:
                header_row_value = int(self.header_row_input.text().strip())
                # Convert from 1-based to 0-based indexing
                # If user enters 1, header is at row 0
                # If user enters 2, skip 1 row, header is at row 1
                header_row = max(0, header_row_value - 1)
            except ValueError:
                header_row = 0

        logger.info(f"[LOAD DATAFRAME] Loading {file_path} with header_row={header_row}")
        file_path_str = str(file_path)
        if file_path_str.lower().endswith('.xlsx') or file_path_str.lower().endswith('.xls'):
            return pd.read_excel(file_path_str, dtype=str, keep_default_na=False, header=header_row).fillna("")
        else:
            return pd.read_csv(file_path_str, dtype=str, keep_default_na=False, header=header_row).fillna("")

    def start_processing_with_editable_preview(self):
        if not self.current_csv:
            return
        # Process with user's entered CI value, generate preview for editing
        # Do NOT change the CI input - keep user's entered value
        self.process_btn.setText("Export Worksheet")
        self.process_btn.setEnabled(False)
        self.start_processing()

    def start_processing(self):
        if not self.current_csv:
            QMessageBox.critical(self, "Error", "Please select a file")
            return
        if len(self.shipment_mapping) < 2:
            QMessageBox.critical(self, "Error", "Map Part Number and Value USD")
            return
        
        # Verify Net Weight is entered
        wt_text = self.wt_input.text().strip()
        if not wt_text:
            self.wt_input.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.wt_input.setStyleSheet(""))
            self.wt_input.setFocus()
            QMessageBox.warning(self, "Net Weight Required", "Please enter the Net Weight (kg) before processing.")
            return
        
        try:
            wt_val = float(wt_text.replace(',', ''))
            if wt_val <= 0:
                raise ValueError()
        except:
            self.wt_input.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.wt_input.setStyleSheet(""))
            self.wt_input.setFocus()
            QMessageBox.warning(self, "Invalid Net Weight", "Please enter a valid Net Weight greater than zero.")
            return
        
        # Verify MID is selected
        if not self.selected_mid:
            self.mid_combo.setStyleSheet("border: 2px solid #ff4444; background-color: #ffebee;")
            QTimer.singleShot(1200, lambda: self.mid_combo.setStyleSheet(""))
            self.mid_combo.setFocus()
            QMessageBox.warning(self, "MID Required", "Please select a MID (Melt ID) before processing.")
            return

        self.process_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.status.setText("Processing...")

        # Directly process the CSV/Excel file synchronously (single-threaded)
        try:
            self.status.setText("Loading file...")
            df = self.load_file_as_dataframe(self.current_csv)
            vr = Path(self.current_csv).stem
            col_map = {v:k for k,v in self.shipment_mapping.items()}
            df = df.rename(columns=col_map)
            if not {'part_number','value_usd'}.issubset(df.columns):
                self.status.setText("Missing Part Number or Value USD")
                return

            # Filter out rows without part numbers (excludes total/subtotal rows)
            initial_row_count = len(df)
            df = df[df['part_number'].notna() & (df['part_number'].astype(str).str.strip() != '')]
            filtered_row_count = len(df)
            if filtered_row_count < initial_row_count:
                logger.info(f"[PROCESS] Filtered {initial_row_count - filtered_row_count} rows without part numbers (total/subtotal rows)")
                logger.info(f"[PROCESS] Processing {filtered_row_count} data rows")
            def safe_float(text):
                if pd.isna(text) or text == "": return 0.0
                try:
                    return float(str(text).replace(',', '').strip())
                except:
                    return 0.0
            df['value_usd'] = pd.to_numeric(df['value_usd'], errors='coerce').fillna(0)
            csv_total = df['value_usd'].sum()
            user_ci = safe_float(self.ci_input.text())
            wt = safe_float(self.wt_input.text())
            if wt <= 0:
                self.status.setText("Net Weight must be greater than zero")
                return
            # Invoice diff
            self.handle_invoice_diff(csv_total, user_ci)
            conn = sqlite3.connect(str(DB_PATH))
            parts = pd.read_sql("SELECT part_number, hts_code, steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio, non_steel_ratio, cbp_qty1, country_of_melt, country_of_cast, country_of_smelt, Sec301_Exclusion_Tariff FROM parts_master", conn)
            conn.close()
            df = df.merge(parts, on='part_number', how='left', suffixes=('', '_master'), indicator=True)
            # Track parts not found in the database
            df['_not_in_db'] = df['_merge'] == 'left_only'
            df = df.drop(columns=['_merge'])

            # Merge strategy: Prefer invoice values over database values for all optional fields
            # For each field, if invoice has it, use it; otherwise fall back to database value
            merge_fields = ['hts_code', 'steel_ratio', 'aluminum_ratio', 'copper_ratio', 'wood_ratio', 'auto_ratio', 'non_steel_ratio', 'cbp_qty1']
            for field in merge_fields:
                master_col = f'{field}_master'
                if field in df.columns and master_col in df.columns:
                    # Invoice has this column mapped - prefer invoice value, fall back to master
                    # Use combine_first to handle both NaN and empty strings
                    df[field] = df[field].replace('', pd.NA).fillna(df[master_col])
                elif field not in df.columns and master_col in df.columns:
                    # Invoice doesn't have this column - use master value
                    df[field] = df[master_col]
                elif field not in df.columns:
                    # Neither invoice nor master has it - set default
                    if field in ['steel_ratio', 'aluminum_ratio', 'copper_ratio', 'wood_ratio', 'auto_ratio', 'non_steel_ratio']:
                        df[field] = 0.0
                    else:
                        df[field] = ''

            # Convert ratio fields to numeric
            df['steel_ratio'] = pd.to_numeric(df['steel_ratio'], errors='coerce').fillna(1.0)
            df['aluminum_ratio'] = pd.to_numeric(df['aluminum_ratio'], errors='coerce').fillna(0.0)
            df['copper_ratio'] = pd.to_numeric(df['copper_ratio'], errors='coerce').fillna(0.0)
            df['wood_ratio'] = pd.to_numeric(df['wood_ratio'], errors='coerce').fillna(0.0)
            df['auto_ratio'] = pd.to_numeric(df['auto_ratio'], errors='coerce').fillna(0.0)
            df['non_steel_ratio'] = pd.to_numeric(df['non_steel_ratio'], errors='coerce').fillna(0.0)
            missing = df[
                (df['hts_code'].isnull() | (df['hts_code'] == '')) |
                (df['value_usd'] == 0) |
                (df['steel_ratio'].isnull())
            ].copy()
            if not missing.empty:
                missing = missing[['part_number', 'hts_code', 'value_usd', 'steel_ratio']].copy()
                missing.columns = ['Part Number', 'HTS Code', 'Value USD', 'Sec 232 Ratio']
                missing = missing.fillna('')
                self.log_missing_data_warning(missing)
            self._process_with_complete_data(df, vr, user_ci, wt)
        except Exception as e:
            logger.error(f"Processing failed: {e}")
            self.status.setText(f"Processing failed: {str(e)}")

    def _process_with_complete_data(self, df, vr, user_ci, wt):
        """
        Process the DataFrame with complete data, calculate required fields, and update the preview table.
        Handles multi-content items (steel, aluminum, copper, wood, non-232).
        """
        df = df.copy()

        # Steel/Aluminum/Copper/Wood/Auto/NonSteel ratios BEFORE calculating weight
        df['SteelRatio'] = pd.to_numeric(df.get('steel_ratio', 0.0), errors='coerce').fillna(0.0)
        df['AluminumRatio'] = pd.to_numeric(df.get('aluminum_ratio', 0.0), errors='coerce').fillna(0.0)
        df['CopperRatio'] = pd.to_numeric(df.get('copper_ratio', 0.0), errors='coerce').fillna(0.0)
        df['WoodRatio'] = pd.to_numeric(df.get('wood_ratio', 0.0), errors='coerce').fillna(0.0)
        df['AutoRatio'] = pd.to_numeric(df.get('auto_ratio', 0.0), errors='coerce').fillna(0.0)
        df['NonSteelRatio'] = pd.to_numeric(df.get('non_steel_ratio', 0.0), errors='coerce').fillna(0.0)

        # Split rows by steel/aluminum/copper/wood/non-232 content BEFORE calculating CalcWtNet
        original_row_count = len(df)
        expanded_rows = []
        for _, row in df.iterrows():
            steel_ratio = row['SteelRatio']
            aluminum_ratio = row['AluminumRatio']
            copper_ratio = row['CopperRatio']
            wood_ratio = row['WoodRatio']
            auto_ratio = row['AutoRatio']
            non_steel_ratio = row['NonSteelRatio']
            original_value = row['value_usd']

            # If no ratios are set, use HTS lookup to determine material type
            if steel_ratio == 0 and aluminum_ratio == 0 and copper_ratio == 0 and wood_ratio == 0 and auto_ratio == 0 and non_steel_ratio == 0:
                # Look up HTS code to determine material type
                hts = row.get('hts_code', '')
                material, _, _ = get_232_info(hts)
                if material == 'Aluminum':
                    aluminum_ratio = 1.0
                elif material == 'Copper':
                    copper_ratio = 1.0
                elif material == 'Wood':
                    wood_ratio = 1.0
                elif material == 'Auto':
                    auto_ratio = 1.0
                elif material == 'Steel':
                    steel_ratio = 1.0
                else:
                    # Default to 100% steel for backward compatibility if no material found
                    steel_ratio = 1.0

            # Create derivative rows in order: Steel first, then Aluminum, Copper, Wood, Auto, Non-232 last
            # This ensures derivatives appear BELOW the primary row in the preview

            # Create steel portion row (if steel_ratio > 0)
            if steel_ratio > 0:
                steel_row = row.copy()
                steel_row['value_usd'] = original_value * steel_ratio
                steel_row['SteelRatio'] = steel_ratio
                steel_row['AluminumRatio'] = 0.0
                steel_row['CopperRatio'] = 0.0
                steel_row['WoodRatio'] = 0.0
                steel_row['AutoRatio'] = 0.0
                steel_row['NonSteelRatio'] = 0.0
                steel_row['_content_type'] = 'steel'
                expanded_rows.append(steel_row)

            # Create aluminum portion row (if aluminum_ratio > 0)
            if aluminum_ratio > 0:
                aluminum_row = row.copy()
                aluminum_row['value_usd'] = original_value * aluminum_ratio
                aluminum_row['SteelRatio'] = 0.0
                aluminum_row['AluminumRatio'] = aluminum_ratio
                aluminum_row['CopperRatio'] = 0.0
                aluminum_row['WoodRatio'] = 0.0
                aluminum_row['AutoRatio'] = 0.0
                aluminum_row['NonSteelRatio'] = 0.0
                aluminum_row['_content_type'] = 'aluminum'
                expanded_rows.append(aluminum_row)

            # Create copper portion row (if copper_ratio > 0)
            if copper_ratio > 0:
                copper_row = row.copy()
                copper_row['value_usd'] = original_value * copper_ratio
                copper_row['SteelRatio'] = 0.0
                copper_row['AluminumRatio'] = 0.0
                copper_row['CopperRatio'] = copper_ratio
                copper_row['WoodRatio'] = 0.0
                copper_row['AutoRatio'] = 0.0
                copper_row['NonSteelRatio'] = 0.0
                copper_row['_content_type'] = 'copper'
                expanded_rows.append(copper_row)

            # Create wood portion row (if wood_ratio > 0)
            if wood_ratio > 0:
                wood_row = row.copy()
                wood_row['value_usd'] = original_value * wood_ratio
                wood_row['SteelRatio'] = 0.0
                wood_row['AluminumRatio'] = 0.0
                wood_row['CopperRatio'] = 0.0
                wood_row['WoodRatio'] = wood_ratio
                wood_row['AutoRatio'] = 0.0
                wood_row['NonSteelRatio'] = 0.0
                wood_row['_content_type'] = 'wood'
                expanded_rows.append(wood_row)

            # Create auto portion row (if auto_ratio > 0)
            if auto_ratio > 0:
                auto_row = row.copy()
                auto_row['value_usd'] = original_value * auto_ratio
                auto_row['SteelRatio'] = 0.0
                auto_row['AluminumRatio'] = 0.0
                auto_row['CopperRatio'] = 0.0
                auto_row['WoodRatio'] = 0.0
                auto_row['AutoRatio'] = auto_ratio
                auto_row['NonSteelRatio'] = 0.0
                auto_row['_content_type'] = 'auto'
                expanded_rows.append(auto_row)

            # Create non-232 portion row (if non_steel_ratio > 0)
            if non_steel_ratio > 0:
                non_232_row = row.copy()
                non_232_row['value_usd'] = original_value * non_steel_ratio
                non_232_row['SteelRatio'] = 0.0
                non_232_row['AluminumRatio'] = 0.0
                non_232_row['CopperRatio'] = 0.0
                non_232_row['WoodRatio'] = 0.0
                non_232_row['AutoRatio'] = 0.0
                non_232_row['NonSteelRatio'] = non_steel_ratio
                non_232_row['_content_type'] = 'non_232'
                expanded_rows.append(non_232_row)

        # Rebuild dataframe from expanded rows
        df = pd.DataFrame(expanded_rows).reset_index(drop=True)
        logger.info(f"Row expansion: {original_row_count} → {len(expanded_rows)} rows (steel/aluminum/copper/wood/auto/non-232 split)")

        # Now calculate CalcWtNet based on expanded rows
        total_value = df['value_usd'].sum()
        if total_value == 0:
            df['CalcWtNet'] = 0.0
        else:
            df['CalcWtNet'] = (df['value_usd'] / total_value) * wt

        # Calculate cbp_qty based on cbp_qty1 unit type
        # KG = use net weight (CalcWtNet), NO = use quantity, blank/empty/other = empty
        def get_cbp_qty(row):
            cbp_qty1 = str(row.get('cbp_qty1', '')).strip().upper() if pd.notna(row.get('cbp_qty1')) else ''
            if cbp_qty1 == '':
                # Blank/empty defaults to empty
                return ''
            elif cbp_qty1 == 'KG':
                # KG uses net weight
                return str(int(round(row['CalcWtNet'])))
            elif cbp_qty1 == 'NO':
                # NO uses quantity from invoice
                qty = row.get('quantity', '')
                if pd.notna(qty) and str(qty).strip():
                    try:
                        return str(int(float(str(qty).replace(',', '').strip())))
                    except (ValueError, TypeError):
                        return ''
                return ''
            else:
                # Unknown unit type, return empty
                return ''
        
        df['cbp_qty'] = df.apply(get_cbp_qty, axis=1)

        # Set HTSCode and MID
        df['HTSCode'] = df.get('hts_code', '')
        mid = self.selected_mid if hasattr(self, 'selected_mid') else ''
        df['MID'] = mid
        melt = str(mid)[:2] if mid else ''

        # Derivative fields (exact 8-digit match, flag logic)
        dec_type_list = []
        country_melt_list = []
        country_cast_list = []
        prim_country_smelt_list = []
        prim_smelt_flag_list = []
        flag_list = []
        for _, r in df.iterrows():
            content_type = r.get('_content_type', '')
            hts = r.get('hts_code', '')
            hts_clean = str(hts).replace('.', '').strip().upper()
            hts_8 = hts_clean[:8]
            hts_10 = hts_clean[:10]
            material, dec_type, smelt_flag = get_232_info(hts)

            # Set flag and declaration code based on content type
            # Each derivative row gets its appropriate declaration code
            if content_type == 'steel':
                flag = '232_Steel'
                # Steel derivatives always get declaration code 08
                dec_type_list.append(dec_type if dec_type else '08')
            elif content_type == 'aluminum':
                flag = '232_Aluminum'
                # Aluminum derivatives always get declaration code 07
                dec_type_list.append(dec_type if dec_type else '07')
            elif content_type == 'copper':
                flag = '232_Copper'
                # Copper derivatives get their declaration code
                dec_type_list.append(dec_type if dec_type else '11')
            elif content_type == 'wood':
                flag = '232_Wood'
                # Wood derivatives get their declaration code
                dec_type_list.append(dec_type if dec_type else '10')
            elif content_type == 'auto':
                flag = '232_Auto'
                # Auto derivatives get their declaration code
                dec_type_list.append(dec_type if dec_type else '')
            elif content_type == 'non_232':
                flag = 'Non_232'
                # Non-232 rows still need the declaration code from HTS lookup
                dec_type_list.append(dec_type)
            else:
                # Fallback for backward compatibility
                flag = f"232_{material}" if material else ''
                dec_type_list.append(dec_type)
            
            # Use imported country codes if available, otherwise fall back to MID-based default
            country_of_melt = r.get('country_of_melt', '')
            country_of_cast = r.get('country_of_cast', '')
            country_of_smelt = r.get('country_of_smelt', '')

            # Check for NaN/null values and use default if empty
            melt_code = country_of_melt if pd.notna(country_of_melt) and str(country_of_melt).strip() else melt
            cast_code = country_of_cast if pd.notna(country_of_cast) and str(country_of_cast).strip() else melt
            smelt_code = country_of_smelt if pd.notna(country_of_smelt) and str(country_of_smelt).strip() else melt

            country_melt_list.append(melt_code)
            country_cast_list.append(cast_code)
            prim_country_smelt_list.append(smelt_code)
            prim_smelt_flag_list.append(smelt_flag)
            flag_list.append(flag)

        df['DecTypeCd'] = dec_type_list
        df['CountryofMelt'] = country_melt_list
        df['CountryOfCast'] = country_cast_list
        df['PrimCountryOfSmelt'] = prim_country_smelt_list
        df['PrimSmeltFlag'] = prim_smelt_flag_list
        df['_232_flag'] = flag_list

        # Rename columns for preview
        df['Product No'] = df['part_number']
        df['ValueUSD'] = df['value_usd']

        # Include invoice_number if mapped (for split by invoice export feature)
        base_preview_cols = [
            'Product No','ValueUSD','HTSCode','MID','cbp_qty','DecTypeCd',
            'CountryofMelt','CountryOfCast','PrimCountryOfSmelt','PrimSmeltFlag',
            'SteelRatio','AluminumRatio','CopperRatio','WoodRatio','AutoRatio','NonSteelRatio','_232_flag','_not_in_db','Sec301_Exclusion_Tariff'
        ]
        preview_cols = base_preview_cols.copy()
        if 'invoice_number' in df.columns:
            preview_cols.append('invoice_number')
        preview_df = df[preview_cols].copy()
        self.on_done(preview_df, vr, None)
    
    def start_processing_with_editable_preview(self):
        """Open the CSV file in default editor for user to edit directly"""
        if not self.current_csv:
            return
        
        try:
            # Open the CSV file with the default system application
            import subprocess
            if sys.platform == 'win32':
                os.startfile(self.current_csv)
            elif sys.platform == 'darwin':  # macOS
                subprocess.run(['open', self.current_csv])
            else:  # linux
                subprocess.run(['xdg-open', self.current_csv])
            
            # Show message to user
            QMessageBox.information(
                self, 
                "Edit File", 
                f"Opening file for editing:\n{Path(self.current_csv).name}\n\n"
                "Edit the values, save the file, then return here.\n"
                "The CI Value input will be updated when you reload the file."
            )
            
            # Reload the file to get updated values
            self.reload_csv_values()
            
        except Exception as e:
            logger.error(f"Failed to open file: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open file:\n{e}")
    
    def reload_csv_values(self):
        """Reload CSV to recalculate total after external edits"""
        if not self.current_csv:
            return
        
        try:
            col_map = {v: k for k, v in self.shipment_mapping.items()}
            if Path(self.current_csv).suffix.lower() == ".xlsx":
                df = pd.read_excel(self.current_csv, dtype=str)
            else:
                df = pd.read_csv(self.current_csv, dtype=str)
            df = df.rename(columns=col_map)

            if 'value_usd' in df.columns:
                # Check for rows where value_usd is blank, empty, or zero
                original_count = len(df)
                df['value_usd'] = pd.to_numeric(df['value_usd'], errors='coerce')
                zero_rows_df = df[df['value_usd'].isna() | (df['value_usd'] == 0)]
                zero_count = len(zero_rows_df)
                
                removed_count = 0
                if zero_count > 0:
                    # Prompt user to confirm deletion
                    reply = QMessageBox.question(
                        self,
                        "Remove Zero Value Rows",
                        f"Found {zero_count} row(s) with blank or zero values.\n\n"
                        f"Do you want to remove these rows?\n\n"
                        f"• Yes: Remove rows and continue processing\n"
                        f"• No: Keep all rows and process as is",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes
                    )
                    
                    if reply == QMessageBox.Yes:
                        # Remove the zero value rows
                        df = df[df['value_usd'].notna() & (df['value_usd'] != 0)]
                        removed_count = original_count - len(df)
                        
                        # Save cleaned data back to file
                        if removed_count > 0:
                            # Rename back to original columns for saving
                            reverse_map = {k: v for k, v in self.shipment_mapping.items()}
                            save_df = df.rename(columns=reverse_map)
                            
                            if Path(self.current_csv).suffix.lower() == ".xlsx":
                                save_df.to_excel(self.current_csv, index=False)
                            else:
                                save_df.to_csv(self.current_csv, index=False)
                            
                            logger.info(f"Removed {removed_count} rows with blank/zero values")
                    else:
                        # User chose No - keep all rows
                        logger.info(f"User chose to keep {zero_count} row(s) with blank/zero values")
                
                # Calculate total
                total = df['value_usd'].sum()
                self.csv_total_value = round(total, 2)
                # Don't overwrite user's CI input - just update the check
                self.update_invoice_check()
                
                if removed_count > 0:
                    self.status.setText(f"File reloaded - Removed {removed_count} blank/zero rows")
                    self.status.setStyleSheet("background:#ff9800; color:white; font-weight:bold; padding:8px;")
                elif zero_count > 0:
                    self.status.setText(f"File reloaded - Kept {zero_count} blank/zero rows")
                    self.status.setStyleSheet("background:#2196F3; color:white; font-weight:bold; padding:8px;")
                else:
                    self.status.setText("File reloaded - Check invoice values")
                    self.status.setStyleSheet("background:#2196F3; color:white; font-weight:bold; padding:8px;")
        except Exception as e:
            logger.error(f"reload_csv_values failed: {e}")
            QMessageBox.warning(self, "Reload Error", f"Failed to reload values:\n{e}")

    def handle_invoice_diff(self, csv_sum, user_entered):
        # Display-only; enablement handled by update/check methods
        diff = abs(csv_sum - user_entered)
        threshold = 0.01
        if diff > threshold:
            self.invoice_check_label.setText(
                f": CSV = ${csv_sum:,.2f} | "
                f"Entered = ${user_entered:,.2f} | Diff = ${diff:,.2f}"
            )
            self.invoice_check_label.setStyleSheet("background:#A4262C; color:white; font-weight:bold; font-size:7pt; padding:3px;")
        else:
            self.invoice_check_label.setText(f"Match: ${csv_sum:,.2f}")
            self.invoice_check_label.setStyleSheet("background:#107C10; color:white; font-weight:bold; font-size:7pt; padding:3px;")

    def save_edited_values_and_process(self):
        if not hasattr(self, 'editable_invoice_df'):
            return

        try:
            updated = 0
            for row in range(self.missing_table.rowCount()):
                value_item = self.missing_table.item(row, 1)
                if value_item:
                    clean_val = value_item.text().replace('$', '').replace(',', '').strip()
                    try:
                        new_val = float(clean_val)
                        self.editable_invoice_df.at[row, 'value_usd'] = new_val
                        updated += 1
                    except:
                        self.editable_invoice_df.at[row, 'value_usd'] = 0.0

            # Save back to original file
            if self.editable_invoice_path.suffix.lower() == ".xlsx":
                self.editable_invoice_df.to_excel(self.editable_invoice_path, index=False)
            else:
                self.editable_invoice_df.to_csv(self.editable_invoice_path, index=False)

            # Update totals and UI
            new_total = float(self.editable_invoice_df['value_usd'].sum())
            self.ci_input.setText(f"{new_total:,.2f}")
            self.csv_total_value = round(new_total, 2)
            self.update_invoice_check()

            QMessageBox.information(self, "Success", 
                f"Updated {updated} line values!\n\n"
                f"New invoice total: ${new_total:,.2f}\n\n"
                "Starting processing...")

            # Hide button and reprocess
            self.save_and_process_btn.setVisible(False)
            self.start_processing()

        except Exception as e:
            logger.error(f"save_edited_values_and_process failed: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save edits:\n{e}")

    # ====================== MISSING DATA HANDLER ======================
    def log_missing_data_warning(self, missing_df):
        """Log missing data as warning but allow processing to continue"""
        # Note: missing_df has 'Part Number' column (with capital letters)
        part_col = 'Part Number' if 'Part Number' in missing_df.columns else 'part_number'
        parts_list = ", ".join(str(p) for p in missing_df[part_col].tolist()[:10])
        if len(missing_df) > 10:
            parts_list += f" ... and {len(missing_df) - 10} more"
        
        logger.warning(f"Found {len(missing_df)} parts with missing data: {parts_list}")
        self.status.setText(f"⚠ Warning: {len(missing_df)} parts have missing data - review in preview")
        self.status.setStyleSheet("background:#f7bfa1; color:white; font-weight:bold; padding:8px;")

    def on_worker_error(self, msg):
        self.progress.setVisible(False)
        self.process_btn.setEnabled(True)
        QMessageBox.critical(self, "Error", msg)
        self.status.setText("Error")
        self.status.setStyleSheet("background:#dd6e74; color:white; font-weight:bold;")

    def on_done(self, df, vr, fname):
        # Populate preview with editable Value column; export later when totals match
        self.progress.setVisible(False)
        self.last_processed_df = df.copy()
        self.last_output_filename = f"Upload_Sheet_{vr}_{datetime.now():%Y%m%d_%H%M}.xlsx"

        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)  # Disable sorting while populating
        self.table.setRowCount(len(df))
        has_232 = False
        for i, r in df.iterrows():
            flag = r.get('_232_flag', '')
            if flag:
                has_232 = True

            value_item = QTableWidgetItem(f"{r['ValueUSD']:,.2f}")
            value_item.setData(Qt.UserRole, r['ValueUSD'])

            # Get ratio values for display
            steel_ratio_val = r.get('SteelRatio', 0.0) or 0.0
            aluminum_ratio_val = r.get('AluminumRatio', 0.0) or 0.0
            copper_ratio_val = r.get('CopperRatio', 0.0) or 0.0
            wood_ratio_val = r.get('WoodRatio', 0.0) or 0.0
            auto_ratio_val = r.get('AutoRatio', 0.0) or 0.0
            non_steel_ratio_val = r.get('NonSteelRatio', 0.0) or 0.0
            is_232_row = steel_ratio_val > 0.0 or aluminum_ratio_val > 0.0 or copper_ratio_val > 0.0 or wood_ratio_val > 0.0 or auto_ratio_val > 0.0

            # Check if part is not in database - show "Not Found" in 232 Status column
            not_in_db = r.get('_not_in_db', False)
            status_display = "Not Found" if not_in_db else flag

            # Display percentages (empty for "Not Found" rows)
            if not_in_db:
                steel_display = ""
                aluminum_display = ""
                copper_display = ""
                wood_display = ""
                auto_display = ""
                non_steel_display = ""
            else:
                steel_display = f"{steel_ratio_val*100:.1f}%" if steel_ratio_val > 0 else ""
                aluminum_display = f"{aluminum_ratio_val*100:.1f}%" if aluminum_ratio_val > 0 else ""
                copper_display = f"{copper_ratio_val*100:.1f}%" if copper_ratio_val > 0 else ""
                wood_display = f"{wood_ratio_val*100:.1f}%" if wood_ratio_val > 0 else ""
                auto_display = f"{auto_ratio_val*100:.1f}%" if auto_ratio_val > 0 else ""
                non_steel_display = f"{non_steel_ratio_val*100:.1f}%" if non_steel_ratio_val > 0 else ""

            # Get Sec301 exclusion tariff value
            sec301_exclusion = str(r.get('Sec301_Exclusion_Tariff', '')).strip()
            has_sec301_exclusion = bool(sec301_exclusion and sec301_exclusion not in ['', 'nan', 'None'])

            # Create HTS item with tooltip if Sec301 exclusion exists
            hts_item = QTableWidgetItem(str(r.get('HTSCode','')))
            if has_sec301_exclusion:
                hts_item.setToolTip(f"Sec301 Exclusion Tariff: {sec301_exclusion}")

            product_no = str(r['Product No'])

            items = [
                QTableWidgetItem(product_no),
                value_item,
                hts_item,
                QTableWidgetItem(str(r.get('MID',''))),
                QTableWidgetItem(str(r.get('cbp_qty', ''))),
                QTableWidgetItem(str(r.get('DecTypeCd',''))),
                QTableWidgetItem(str(r.get('CountryofMelt',''))),
                QTableWidgetItem(str(r.get('CountryOfCast',''))),
                QTableWidgetItem(str(r.get('PrimCountryOfSmelt',''))),
                QTableWidgetItem(str(r.get('PrimSmeltFlag',''))),
                QTableWidgetItem(steel_display),
                QTableWidgetItem(aluminum_display),
                QTableWidgetItem(copper_display),
                QTableWidgetItem(wood_display),
                QTableWidgetItem(auto_display),
                QTableWidgetItem(non_steel_display),
                QTableWidgetItem(status_display)
            ]

            # Make all items editable except Steel%, Al%, Cu%, Wood%, Auto%, Non-232%, and 232 Status
            for idx, item in enumerate(items):
                if idx not in [10, 11, 12, 13, 14, 15, 16]:  # Not Steel%, Al%, Cu%, Wood%, Auto%, Non-232%, 232 Status
                    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)

            # Set font colors based on Section 232 material type
            row_color = self.get_preview_row_color(flag)
            for item in items:
                item.setForeground(row_color)
                f = item.font()
                f.setBold(False)
                item.setFont(f)
                item.setTextAlignment(Qt.AlignCenter)  # Center text in all columns

                # Add red border for rows with Sec301 exclusion
                if has_sec301_exclusion:
                    item.setData(Qt.UserRole + 1, 'sec301_exclusion')  # Mark for border styling

            for j, it in enumerate(items):
                self.table.setItem(i, j, it)

            # Apply background color to entire row if Sec301 exclusion exists
            if has_sec301_exclusion:
                # Get user-configured Sec301 background color
                sec301_bg_color = self.get_sec301_bg_color()
                for j in range(len(items)):
                    cell_item = self.table.item(i, j)
                    if cell_item:
                        # Set background color using saved preference
                        cell_item.setBackground(sec301_bg_color)

        self.table.setSortingEnabled(True)  # Re-enable sorting after populating
        self.table.blockSignals(False)
        self.table.itemChanged.connect(self.on_preview_value_edited)
        self.recalculate_total_and_check_match()
        self.apply_column_visibility()  # Apply saved column visibility settings

        # if has_232:
        #     self.status.setText("SECTION 232 ITEMS • EDIT VALUES • EXPORT WHEN READY")
        #     self.status.setStyleSheet("background:#A4262C; color:white; font-weight:bold; font-size:16pt; padding:10px;")
        # else:
        #     self.status.setText("Edit values • Export when total matches")
        #     self.status.setStyleSheet("font-size:14pt; padding:8px; background:#f0f0f0;")

    def setup_import_tab(self):
        from PyQt5.QtGui import QPalette

        layout = QVBoxLayout(self.tab_import)
        title = QLabel("<h2>Parts Import from CSV/Excel</h2><p>Drag & drop columns to map fields</p>")
        title.setAlignment(Qt.AlignCenter)
        # Use palette text color for consistent title styling
        app = QApplication.instance()
        palette = app.palette()
        text_color = palette.color(QPalette.Text)
        title.setStyleSheet(f"color: {text_color.name()};")
        layout.addWidget(title)

        # Buttons at top
        button_widget = QWidget()
        btn_layout = QHBoxLayout(button_widget)
        btn_load = QPushButton("Load CSV/Excel File")
        btn_load.setStyleSheet(self.get_button_style("info"))
        btn_load.clicked.connect(self.load_csv_for_import)
        btn_reset = QPushButton("Reset Mapping")
        btn_reset.setStyleSheet(self.get_button_style("danger"))
        btn_reset.clicked.connect(self.reset_import_mapping)
        btn_import = QPushButton("Import Now")
        btn_import.setFixedSize(100, 28)
        btn_import.setStyleSheet(self.get_button_style("success") + "QPushButton { font-size:10pt; padding:4px; }")
        btn_import.clicked.connect(self.start_parts_import)

        btn_update_sec301 = QPushButton("Update Sec301 Exclusion")
        btn_update_sec301.setStyleSheet(self.get_button_style("info"))
        btn_update_sec301.clicked.connect(self.update_sec301_single)

        btn_import_sec301 = QPushButton("Import Sec301 CSV")
        btn_import_sec301.setStyleSheet(self.get_button_style("info"))
        btn_import_sec301.clicked.connect(self.import_sec301_csv)

        btn_layout.addWidget(btn_load)
        btn_layout.addWidget(btn_reset)
        btn_layout.addWidget(btn_update_sec301)
        btn_layout.addWidget(btn_import_sec301)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_import)
        layout.addWidget(button_widget)

        # Main drag/drop area with scroll
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0,0,0,0)
        main_layout.setSpacing(20)

        left = QGroupBox("CSV/Excel Columns - Drag")
        left_outer_layout = QVBoxLayout()
        
        # Add scroll area for columns
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        scroll_widget = QWidget()
        left_layout = QVBoxLayout(scroll_widget)
        left_layout.setContentsMargins(5, 5, 5, 5)
        left_layout.setSpacing(5)
        self.drag_labels = []
        # Don't add stretch here - it will be added after labels are loaded
        
        scroll_area.setWidget(scroll_widget)
        left_outer_layout.addWidget(scroll_area)
        left.setLayout(left_outer_layout)
        
        # Store reference to left_layout for adding labels later
        self.import_left_layout = left_layout

        right = QGroupBox("Available Fields - Drop Here")
        right_outer_layout = QVBoxLayout()
        
        # Add scroll area for drop targets
        right_scroll_area = QScrollArea()
        right_scroll_area.setWidgetResizable(True)
        right_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        right_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        right_scroll_widget = QWidget()
        right_layout = QFormLayout(right_scroll_widget)
        right_layout.setLabelAlignment(Qt.AlignRight)
        right_layout.setContentsMargins(5, 5, 5, 5)
        
        self.import_targets = {}
        fields = {
            "part_number": "Part Number *",
            "hts_code": "HTS Code *",
            "mid": "MID *",
            "steel_ratio": "Steel Ratio *",
            "aluminum_ratio": "Aluminum Ratio",
            "copper_ratio": "Copper Ratio",
            "wood_ratio": "Wood Ratio",
            "auto_ratio": "Auto Ratio",
            "cbp_qty1": "CBP Qty1",
            "country_of_melt": "Country of Melt",
            "country_of_cast": "Country of Cast",
            "country_of_smelt": "Country of Smelt",
            "Sec301_Exclusion_Tariff": "Sec301 Exclusion Tariff",
            "client_code": "Client Code",
            "description": "Description",
            "country_origin": "Country of Origin"
        }
        drop_labels = {
            "steel_ratio": "Steel%",
            "aluminum_ratio": "Aluminum%",
            "copper_ratio": "Copper%",
            "wood_ratio": "Wood%",
            "auto_ratio": "Auto%"
        }
        for key, name in fields.items():
            drop_label = drop_labels.get(key)
            target = DropTarget(key, name, drop_label)
            target.dropped.connect(self.on_import_drop)
            is_required = "*" in name
            label_text = name.replace(" *", "")
            if is_required:
                label = QLabel(f"{label_text}: <span style='color:red;'>*</span>")
            else:
                label = QLabel(f"{label_text}:")
            right_layout.addRow(label, target)
            self.import_targets[key] = target
        
        right_scroll_area.setWidget(right_scroll_widget)
        right_outer_layout.addWidget(right_scroll_area)
        right.setLayout(right_outer_layout)

        main_layout.addWidget(left, 1)
        main_layout.addWidget(right, 2)
        layout.addWidget(main_widget, 1)
        self.import_widget = main_widget

        self.import_csv_path = None
        self.tab_import.setLayout(layout)

    def load_csv_for_import_from_path(self, path):
        """Load CSV/Excel file from a given path (used by drag-and-drop)"""
        if not path:
            return
        self.import_csv_path = path
        try:
            # Handle both CSV and Excel files
            if path.lower().endswith('.xlsx'):
                df = pd.read_excel(path, nrows=0, dtype=str)
            else:
                df = pd.read_csv(path, nrows=0, dtype=str)
            cols = list(df.columns)

            # Clear previous mappings when loading new file
            for target in self.import_targets.values():
                target.column_name = None
                target.setText(f"Drop {target.field_key} here")
                target.setProperty("occupied", False)
                target.style().unpolish(target)
                target.style().polish(target)

            # Clear existing labels
            for label in self.drag_labels:
                label.setParent(None)
                label.deleteLater()
            self.drag_labels = []
            
            # Add new labels directly to the layout
            for col in cols:
                lbl = DraggableLabel(str(col))
                self.import_left_layout.addWidget(lbl)
                self.drag_labels.append(lbl)

            # Add stretch at the end to push labels to the top
            self.import_left_layout.addStretch()

            # Try to restore saved mappings if they match columns in the new file
            if MAPPING_FILE.exists():
                try:
                    saved_mapping = json.loads(MAPPING_FILE.read_text())
                    # Restore mappings only if the column exists in the new file
                    for field_key, column_name in saved_mapping.items():
                        if column_name in cols and field_key in self.import_targets:
                            target = self.import_targets[field_key]
                            target.column_name = column_name
                            target.setText(f"{field_key}\n<- {column_name}")
                            target.setProperty("occupied", True)
                            target.style().unpolish(target)
                            target.style().polish(target)
                            logger.info(f"Restored mapping: {field_key} <- {column_name}")
                except Exception as e:
                    logger.warning(f"Failed to restore saved mappings: {e}")

            logger.info(f"Loaded CSV for import (drag-drop): {Path(path).name}")
            self.bottom_status.setText(f"CSV loaded: {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot read CSV:\n{e}")
    
    def load_csv_for_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select CSV/Excel", str(INPUT_DIR), "CSV/Excel Files (*.csv *.xlsx)")
        if not path: return
        self.load_csv_for_import_from_path(path)
    
    def load_csv_for_import_old(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select CSV/Excel", str(INPUT_DIR), "CSV/Excel Files (*.csv *.xlsx)")
        if not path: return
        self.import_csv_path = path
        try:
            # Handle both CSV and Excel files
            if path.lower().endswith('.xlsx'):
                df = pd.read_excel(path, nrows=0, dtype=str)
            else:
                df = pd.read_csv(path, nrows=0, dtype=str)
            cols = list(df.columns)
            for label in self.drag_labels:
                label.setParent(None)
            self.drag_labels = []
            left_layout = self.import_widget.layout().itemAt(0).widget().layout()
            for col in cols:
                lbl = DraggableLabel(str(col))
                left_layout.insertWidget(left_layout.count()-1, lbl)
                self.drag_labels.append(lbl)
            logger.info(f"Loaded CSV for import: {Path(path).name}")
            self.bottom_status.setText(f"CSV loaded: {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot read CSV:\n{e}")

    def reset_import_mapping(self):
        if QMessageBox.question(self, "Reset", "Clear all field mappings and column list?") != QMessageBox.Yes:
            return
        
        # Clear drop targets (right side)
        for target in self.import_targets.values():
            target.column_name = None
            target.setText(f"Drop {target.field_key} here")
            target.setProperty("occupied", False)
            target.style().unpolish(target); target.style().polish(target)
        
        # Clear drag labels (left side - CSV/Excel columns)
        for label in self.drag_labels:
            label.setParent(None)
            label.deleteLater()
        self.drag_labels = []
        
        # Clear the file path
        self.import_csv_path = None
        
        # Delete mapping file if it exists
        if MAPPING_FILE.exists():
            MAPPING_FILE.unlink()
        
        logger.info("Import mapping and column list reset")
        self.bottom_status.setText("Import mapping cleared")

    def on_import_drop(self, field_key, column_name):
        for k, t in self.import_targets.items():
            if t.column_name == column_name and k != field_key:
                t.column_name = None
                t.setText(f"Drop {t.field_key} here")
                t.setProperty("occupied", False)
                t.style().unpolish(t); t.style().polish(t)

        # Update the target that received the drop
        target = self.import_targets[field_key]
        target.column_name = column_name
        target.setText(f"{field_key}\n<- {column_name}")
        target.setProperty("occupied", True)
        target.style().unpolish(target); target.style().polish(target)

        self.current_mapping = getattr(self, 'current_mapping', {})
        self.current_mapping[field_key] = column_name
        MAPPING_FILE.write_text(json.dumps(self.current_mapping, indent=2))

    def start_parts_import(self):
        if not self.import_csv_path:
            QMessageBox.warning(self, "No File", "Load a CSV or Excel file first")
            return
        mapping = {k: t.column_name for k, t in self.import_targets.items() if t.column_name}
        required_fields = ['part_number','hts_code','mid','steel_ratio']
        missing = [f for f in required_fields if f not in mapping]
        if missing:
            fields = {
                'part_number': 'Part Number',
                'description': 'Description',
                'quantity': 'Quantity',
                'net_weight': 'Net Weight',
                'value_usd': 'Value (USD)',
                'hts_code': 'HTS Code',
                'mid': 'MID',
                'cbp_qty1': 'CBP_qty1',
                'steel_ratio': 'Steel Ratio',
                'aluminum_ratio': 'Aluminum Ratio',
                'copper_ratio': 'Copper Ratio',
                'wood_ratio': 'Wood Ratio',
                'auto_ratio': 'Auto Ratio',
                'non_steel_ratio': 'Non-Steel Ratio',
            }
        try:
            if self.import_csv_path.lower().endswith('.xlsx'):
                df = pd.read_excel(self.import_csv_path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_csv(self.import_csv_path, dtype=str, keep_default_na=False)
            df = df.fillna("").rename(columns=str.strip)
            col_map = {v: k for k, v in mapping.items()}
            df = df.rename(columns=col_map)
            required = ['part_number','hts_code','mid','steel_ratio']
            missing = [f for f in required if f not in df.columns]
            if missing:
                QMessageBox.critical(self, "Error", f"Missing required fields: {', '.join(missing)}")
                self.status.setText("Import failed")
                return
            # First pass: validate ratios and collect any rows with total > 1.0
            invalid_ratio_rows = []
            for idx, r in df.iterrows():
                part = str(r.get('part_number', '')).strip()
                if not part:
                    continue

                # Helper function to parse ratio values for validation
                def parse_ratio_val(value_str):
                    try:
                        if value_str:
                            ratio = float(value_str)
                            if ratio > 1.0:
                                ratio /= 100.0
                            return max(0.0, ratio)
                        return 0.0
                    except:
                        return 0.0

                steel_val = parse_ratio_val(str(r.get('steel_ratio', r.get('Sec 232 Content Ratio', r.get('Steel %', '')))).strip())
                aluminum_val = parse_ratio_val(str(r.get('aluminum_ratio', r.get('Aluminum %', ''))).strip())
                copper_val = parse_ratio_val(str(r.get('copper_ratio', r.get('Copper %', ''))).strip())
                wood_val = parse_ratio_val(str(r.get('wood_ratio', r.get('Wood %', ''))).strip())
                auto_val = parse_ratio_val(str(r.get('auto_ratio', r.get('Auto %', ''))).strip())
                non_steel_val = parse_ratio_val(str(r.get('non_steel_ratio', r.get('Non-Steel %', ''))).strip())

                total_ratio = steel_val + aluminum_val + copper_val + wood_val + auto_val + non_steel_val
                if total_ratio > 1.01:  # Allow small floating point tolerance
                    invalid_ratio_rows.append((part, total_ratio))

            # If there are invalid rows, reject the import
            if invalid_ratio_rows:
                msg = f"Import failed. The following {len(invalid_ratio_rows)} part(s) have total ratios exceeding 100%:\n\n"
                for part, total in invalid_ratio_rows[:15]:  # Show first 15
                    msg += f"  {part}: {total*100:.1f}%\n"
                if len(invalid_ratio_rows) > 15:
                    msg += f"  ... and {len(invalid_ratio_rows) - 15} more\n"
                msg += "\nPlease correct these rows in your import file and try again."

                QMessageBox.critical(self, "Invalid Ratios Detected", msg)
                self.status.setText("Import failed - invalid ratios")
                return

            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            updated = inserted = 0
            now = datetime.now().isoformat()
            for _, r in df.iterrows():
                part = str(r.get('part_number', '')).strip()
                if not part: continue
                desc = str(r.get('description', r.get('Description', ''))).strip()
                hts = str(r.get('hts_code', '')).strip()
                origin = str(r.get('country_origin', '')).strip().upper()[:2]
                mid = str(r.get('mid', '')).strip()
                # Get client_code if it was mapped, otherwise empty string
                client_code = str(r.get('client_code', '')).strip() if 'client_code' in df.columns else ""
                # Get cbp_qty1 if it was mapped, otherwise empty string
                cbp_qty1 = str(r.get('cbp_qty1', '')).strip() if 'cbp_qty1' in df.columns else ""
                # Get country codes if mapped, otherwise empty string
                country_of_melt = str(r.get('country_of_melt', '')).strip().upper()[:2] if 'country_of_melt' in df.columns else ""
                country_of_cast = str(r.get('country_of_cast', '')).strip().upper()[:2] if 'country_of_cast' in df.columns else ""
                country_of_smelt = str(r.get('country_of_smelt', '')).strip().upper()[:2] if 'country_of_smelt' in df.columns else ""
                # Get Sec301_Exclusion_Tariff if mapped, otherwise empty string
                sec301_exclusion = str(r.get('Sec301_Exclusion_Tariff', '')).strip() if 'Sec301_Exclusion_Tariff' in df.columns else ""

                # Helper function to parse ratio values
                def parse_ratio(value_str):
                    try:
                        if value_str:
                            ratio = float(value_str)
                            if ratio > 1.0: ratio /= 100.0
                            return max(0.0, min(1.0, ratio))
                        return 0.0
                    except:
                        return 0.0

                # Parse all ratio fields
                steel_str = str(r.get('steel_ratio', r.get('Sec 232 Content Ratio', r.get('Steel %', '')))).strip()
                steel_ratio = parse_ratio(steel_str)

                aluminum_str = str(r.get('aluminum_ratio', r.get('Aluminum %', ''))).strip()
                aluminum_ratio = parse_ratio(aluminum_str)

                copper_str = str(r.get('copper_ratio', r.get('Copper %', ''))).strip()
                copper_ratio = parse_ratio(copper_str)

                wood_str = str(r.get('wood_ratio', r.get('Wood %', ''))).strip()
                wood_ratio = parse_ratio(wood_str)

                auto_str = str(r.get('auto_ratio', r.get('Auto %', ''))).strip()
                auto_ratio = parse_ratio(auto_str)

                # Calculate non_steel_ratio as remainder (1.0 minus all 232 ratios)
                total_232 = steel_ratio + aluminum_ratio + copper_ratio + wood_ratio + auto_ratio
                non_steel_ratio = max(0.0, 1.0 - total_232)

                c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                          steel_ratio, non_steel_ratio, last_updated, cbp_qty1, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                          country_of_melt, country_of_cast, country_of_smelt, Sec301_Exclusion_Tariff)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                          ON CONFLICT(part_number) DO UPDATE SET
                          description=excluded.description, hts_code=excluded.hts_code,
                          country_origin=excluded.country_origin, mid=excluded.mid,
                          client_code=excluded.client_code, steel_ratio=excluded.steel_ratio,
                          non_steel_ratio=excluded.non_steel_ratio, last_updated=excluded.last_updated,
                          cbp_qty1=excluded.cbp_qty1, aluminum_ratio=excluded.aluminum_ratio,
                          copper_ratio=excluded.copper_ratio, wood_ratio=excluded.wood_ratio,
                          auto_ratio=excluded.auto_ratio, country_of_melt=excluded.country_of_melt,
                          country_of_cast=excluded.country_of_cast, country_of_smelt=excluded.country_of_smelt,
                          Sec301_Exclusion_Tariff=excluded.Sec301_Exclusion_Tariff""",
                          (part, desc, hts, origin, mid, client_code, steel_ratio, non_steel_ratio, now,
                           cbp_qty1, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                           country_of_melt, country_of_cast, country_of_smelt, sec301_exclusion))
                if c.rowcount:
                    inserted += 1 if conn.total_changes > updated+inserted else 0
                    updated += 1 if conn.total_changes == updated+inserted else 0
            conn.commit(); conn.close()
            QMessageBox.information(self, "Success", f"Imported!\nUpdated: {updated}\nInserted: {inserted}")
            
            # Only refresh parts table if Parts View tab has been initialized
            if hasattr(self, 'parts_table'):
                self.refresh_parts_table()
            
            self.load_available_mids()
            self.bottom_status.setText("Import complete")
        except Exception as e:
            logger.error(f"Import failed: {e}")
            QMessageBox.critical(self, "Error", f"Import failed: {str(e)}")
            self.status.setText("Import failed")

    def update_sec301_single(self):
        """Update Section 301 Exclusion Tariff for a single part number"""
        # Create dialog for input
        dialog = QDialog(self)
        dialog.setWindowTitle("Update Sec301 Exclusion Tariff")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout(dialog)

        # Instructions
        instructions = QLabel("Enter the part number and Section 301 Exclusion Tariff to update:")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        # Form layout
        form_layout = QFormLayout()

        part_number_input = QLineEdit()
        part_number_input.setPlaceholderText("e.g., PART12345")
        form_layout.addRow("Part Number:", part_number_input)

        sec301_input = QLineEdit()
        sec301_input.setPlaceholderText("e.g., 9903.88.15")
        form_layout.addRow("Sec301 Exclusion Tariff:", sec301_input)

        layout.addLayout(form_layout)

        # Buttons
        button_box = QHBoxLayout()
        btn_update = QPushButton("Update")
        btn_update.setStyleSheet(self.get_button_style("success"))
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setStyleSheet(self.get_button_style("danger"))

        btn_update.clicked.connect(dialog.accept)
        btn_cancel.clicked.connect(dialog.reject)

        button_box.addWidget(btn_update)
        button_box.addWidget(btn_cancel)
        layout.addLayout(button_box)

        # Show dialog
        if dialog.exec_() == QDialog.Accepted:
            part_number = part_number_input.text().strip()
            sec301_tariff = sec301_input.text().strip()

            if not part_number:
                QMessageBox.warning(self, "Input Required", "Please enter a part number.")
                return

            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()

                # Check if part exists
                c.execute("SELECT part_number FROM parts_master WHERE part_number=?", (part_number,))
                exists = c.fetchone()

                if not exists:
                    QMessageBox.warning(self, "Part Not Found",
                                      f"Part number '{part_number}' not found in database.\n\n"
                                      "Please import the part first using the regular import function.")
                    conn.close()
                    return

                # Update the Sec301_Exclusion_Tariff
                now = datetime.now().isoformat()
                c.execute("""UPDATE parts_master
                            SET Sec301_Exclusion_Tariff=?, last_updated=?
                            WHERE part_number=?""",
                         (sec301_tariff, now, part_number))
                conn.commit()
                conn.close()

                # Refresh parts table if visible
                if hasattr(self, 'parts_table'):
                    self.refresh_parts_table()

                QMessageBox.information(self, "Success",
                                      f"Updated Sec301 Exclusion Tariff for part '{part_number}'")
                logger.info(f"Updated Sec301 Exclusion Tariff for {part_number}: {sec301_tariff}")
                self.status.setText(f"Updated Sec301 for {part_number}")

            except Exception as e:
                logger.error(f"Failed to update Sec301 exclusion: {e}")
                QMessageBox.critical(self, "Error", f"Failed to update:\n{e}")

    def import_sec301_csv(self):
        """Import Section 301 Exclusion Tariffs from CSV file"""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Sec301 Exclusions CSV File",
            "",
            "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*.*)"
        )
        if not path:
            return

        try:
            # Read the file
            if path.lower().endswith('.xlsx'):
                df = pd.read_excel(path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_csv(path, dtype=str, keep_default_na=False)

            df = df.fillna("").rename(columns=str.strip)

            # Show column mapping dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Map CSV Columns")
            dialog.setMinimumWidth(500)
            layout = QVBoxLayout(dialog)

            instructions = QLabel(
                "<b>Map your CSV columns to the required fields:</b><br><br>"
                "Select which columns from your CSV contain the Part Number and Sec301 Exclusion Tariff."
            )
            instructions.setWordWrap(True)
            layout.addWidget(instructions)

            form_layout = QFormLayout()

            part_combo = QComboBox()
            part_combo.addItems(["-- Select Column --"] + list(df.columns))
            form_layout.addRow("Part Number Column:", part_combo)

            sec301_combo = QComboBox()
            sec301_combo.addItems(["-- Select Column --"] + list(df.columns))
            form_layout.addRow("Sec301 Exclusion Tariff Column:", sec301_combo)

            layout.addLayout(form_layout)

            # Buttons
            button_box = QHBoxLayout()
            btn_import = QPushButton("Import")
            btn_import.setStyleSheet(self.get_button_style("success"))
            btn_cancel = QPushButton("Cancel")
            btn_cancel.setStyleSheet(self.get_button_style("danger"))

            btn_import.clicked.connect(dialog.accept)
            btn_cancel.clicked.connect(dialog.reject)

            button_box.addWidget(btn_import)
            button_box.addWidget(btn_cancel)
            layout.addLayout(button_box)

            # Show dialog
            if dialog.exec_() == QDialog.Accepted:
                part_col = part_combo.currentText()
                sec301_col = sec301_combo.currentText()

                if part_col == "-- Select Column --" or sec301_col == "-- Select Column --":
                    QMessageBox.warning(self, "Selection Required",
                                      "Please select both columns before importing.")
                    return

                # Process the import
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                updated = 0
                not_found = []
                now = datetime.now().isoformat()

                for _, row in df.iterrows():
                    part_number = str(row.get(part_col, '')).strip()
                    sec301_tariff = str(row.get(sec301_col, '')).strip()

                    if not part_number:
                        continue

                    # Check if part exists
                    c.execute("SELECT part_number FROM parts_master WHERE part_number=?", (part_number,))
                    exists = c.fetchone()

                    if exists:
                        c.execute("""UPDATE parts_master
                                    SET Sec301_Exclusion_Tariff=?, last_updated=?
                                    WHERE part_number=?""",
                                 (sec301_tariff, now, part_number))
                        updated += 1
                    else:
                        not_found.append(part_number)

                conn.commit()
                conn.close()

                # Refresh parts table if visible
                if hasattr(self, 'parts_table'):
                    self.refresh_parts_table()

                # Show results
                result_msg = f"Updated {updated} part(s) with Sec301 Exclusion Tariffs."
                if not_found:
                    result_msg += f"\n\n{len(not_found)} part(s) not found in database:\n"
                    result_msg += "\n".join(not_found[:10])  # Show first 10
                    if len(not_found) > 10:
                        result_msg += f"\n... and {len(not_found) - 10} more"

                QMessageBox.information(self, "Import Complete", result_msg)
                logger.info(f"Imported Sec301 exclusions: {updated} updated, {len(not_found)} not found")
                self.status.setText(f"Imported {updated} Sec301 exclusions")

        except Exception as e:
            logger.error(f"Failed to import Sec301 CSV: {e}")
            QMessageBox.critical(self, "Error", f"Failed to import CSV:\n{e}")

    def setup_shipment_mapping_tab(self):
        logger.debug(f"setup_shipment_mapping_tab called - tab_shipment_map={self.tab_shipment_map}")
        layout = QVBoxLayout(self.tab_shipment_map)
        title = QLabel("<h2>Invoice Mapping Profiles</h2><p>Save and load column mappings</p>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Buttons at top - wrap in widget for proper rendering
        top_bar_widget = QWidget()
        top_bar = QHBoxLayout(top_bar_widget)
        self.profile_combo_map = QComboBox()
        self.profile_combo_map.setMinimumWidth(300)
        self.profile_combo_map.currentTextChanged.connect(self.load_selected_profile_full)
        top_bar.addWidget(QLabel("Saved Profiles:"))
        top_bar.addWidget(self.profile_combo_map)

        # Load profiles immediately after creating the combo box
        self.load_mapping_profiles()

        # Add header row input
        top_bar.addWidget(QLabel("Header Row:"))
        self.header_row_input = QLineEdit()
        self.header_row_input.setPlaceholderText("1")
        self.header_row_input.setMaximumWidth(50)
        self.header_row_input.setToolTip("Row number containing column headers (1 = first row, 2 = second row, etc.)")
        top_bar.addWidget(self.header_row_input)

        btn_save = QPushButton("Save Current Mapping As...")
        btn_save.setStyleSheet(self.get_button_style("success"))
        btn_save.clicked.connect(self.save_mapping_profile)
        btn_delete = QPushButton("Delete Profile")
        btn_delete.setStyleSheet(self.get_button_style("danger"))
        btn_delete.clicked.connect(self.delete_mapping_profile)
        btn_load_csv = QPushButton("Load Invoice File")
        btn_load_csv.setStyleSheet(self.get_button_style("info"))
        btn_load_csv.clicked.connect(self.load_csv_for_shipment_mapping)
        btn_reset = QPushButton("Reset Current")
        btn_reset.setStyleSheet(self.get_button_style("danger"))
        btn_reset.clicked.connect(self.reset_current_mapping)
        top_bar.addWidget(btn_load_csv)
        top_bar.addWidget(btn_reset)
        top_bar.addWidget(btn_save)
        top_bar.addWidget(btn_delete)
        top_bar.addStretch()
        layout.addWidget(top_bar_widget)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        self.shipment_widget = QWidget()
        self.shipment_layout = QHBoxLayout(self.shipment_widget)

        left = QGroupBox("Your CSV Columns - Drag")
        left_layout = QVBoxLayout()
        self.shipment_drag_labels = []
        left_layout.addStretch()
        left.setLayout(left_layout)

        right = QGroupBox("Required Fields - Drop")
        right_layout = QFormLayout()
        right_layout.setLabelAlignment(Qt.AlignRight)
        self.shipment_targets = {}
        required_fields = {
            "part_number": "Part Number *",
            "value_usd": "Value USD *"
        }
        for key, name in required_fields.items():
            target = DropTarget(key, name)
            target.dropped.connect(self.on_shipment_drop)
            right_layout.addRow(f"{name}:", target)
            self.shipment_targets[key] = target
        right.setLayout(right_layout)

        optional = QGroupBox("Optional Fields - Drop")
        optional_layout = QFormLayout()
        optional_layout.setLabelAlignment(Qt.AlignRight)
        optional_fields = {
            "invoice_number": "Invoice Number",
            "quantity": "Quantity",
            "hts_code": "HTS Code",
            "cbp_qty1": "CBP Qty1"
        }
        for key, name in optional_fields.items():
            target = DropTarget(key, name)
            target.dropped.connect(self.on_shipment_drop)
            optional_layout.addRow(f"{name}:", target)
            self.shipment_targets[key] = target
        optional.setLayout(optional_layout)

        self.shipment_layout.addWidget(left,1); self.shipment_layout.addWidget(right,1); self.shipment_layout.addWidget(optional,1)
        scroll_layout.addWidget(self.shipment_widget)

        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll, 1)
        self.tab_shipment_map.setLayout(layout)

    def setup_output_mapping_tab(self):
        """Setup the Output XLSX Mapping tab for customizing export column names and order"""
        layout = QVBoxLayout(self.tab_output_map)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)

        title = QLabel("<h2>Output XLSX Column Mapping</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Top bar with profile management
        top_bar_widget = QWidget()
        top_bar = QHBoxLayout(top_bar_widget)
        top_bar.setContentsMargins(0, 0, 0, 0)
        self.output_profile_combo = QComboBox()
        self.output_profile_combo.setMinimumWidth(250)
        self.output_profile_combo.currentTextChanged.connect(self.load_output_mapping_profile)
        top_bar.addWidget(QLabel("Saved Profiles:"))
        top_bar.addWidget(self.output_profile_combo)

        # Load existing profiles
        self.load_output_mapping_profiles()

        btn_reset_output = QPushButton("Reset to Default")
        btn_reset_output.setStyleSheet(self.get_button_style("info"))
        btn_reset_output.clicked.connect(self.reset_output_mapping)

        btn_save_output = QPushButton("Save Current Mapping As...")
        btn_save_output.setStyleSheet(self.get_button_style("success"))
        btn_save_output.clicked.connect(self.save_output_mapping_profile)

        btn_delete_output = QPushButton("Delete Profile")
        btn_delete_output.setStyleSheet(self.get_button_style("danger"))
        btn_delete_output.clicked.connect(self.delete_output_mapping_profile)

        top_bar.addWidget(btn_reset_output)
        top_bar.addWidget(btn_save_output)
        top_bar.addWidget(btn_delete_output)
        top_bar.addStretch()
        layout.addWidget(top_bar_widget)

        # Helper function to create export color picker button
        def create_export_color_button(config_key, default_color, label_text):
            """Create a color picker button for export with saved color (per-user setting)"""
            button = QPushButton()
            button.setFixedSize(80, 25)

            # Load saved color from per-user settings or use default
            saved_color = get_user_setting(config_key, default_color)

            button.setStyleSheet(f"background-color: {saved_color}; border: 1px solid #999;")

            def pick_color():
                current_color = get_user_setting(config_key, default_color)
                color = QColorDialog.getColor(QColor(current_color), self, f"Choose {label_text} Color")
                if color.isValid():
                    color_hex = color.name()
                    button.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999;")
                    set_user_setting(config_key, color_hex)
                    logger.info(f"Saved export color preference {config_key}: {color_hex}")
                    self.bottom_status.setText(f"{label_text} export color set to {color_hex}")

            button.clicked.connect(pick_color)
            return button

        # === COLORS AND OPTIONS SECTION (horizontal layout) ===
        options_widget = QWidget()
        options_layout = QHBoxLayout(options_widget)
        options_layout.setContentsMargins(0, 5, 0, 5)
        options_layout.setSpacing(15)

        # --- Left: Export Colors (Grid Layout) ---
        colors_group = QGroupBox("Export Colors")
        colors_grid = QGridLayout(colors_group)
        colors_grid.setSpacing(5)
        colors_grid.setContentsMargins(10, 10, 10, 10)

        # Load saved font color (per-user setting)
        self.output_font_color = get_user_setting('output_font_color', '#000000')

        # Font color
        self.output_font_color_btn = QPushButton()
        self.output_font_color_btn.setFixedSize(80, 25)
        self.output_font_color_btn.setStyleSheet(f"background-color: {self.output_font_color}; border: 1px solid #999;")
        self.output_font_color_btn.clicked.connect(self.pick_output_font_color)

        # Row 0: Default Font and Steel
        colors_grid.addWidget(QLabel("Default:"), 0, 0, Qt.AlignRight)
        colors_grid.addWidget(self.output_font_color_btn, 0, 1)
        colors_grid.addWidget(QLabel("Steel:"), 0, 2, Qt.AlignRight)
        colors_grid.addWidget(create_export_color_button('export_steel_color', '#4a4a4a', 'Steel'), 0, 3)

        # Row 1: Aluminum and Copper
        colors_grid.addWidget(QLabel("Aluminum:"), 1, 0, Qt.AlignRight)
        colors_grid.addWidget(create_export_color_button('export_aluminum_color', '#6495ED', 'Aluminum'), 1, 1)
        colors_grid.addWidget(QLabel("Copper:"), 1, 2, Qt.AlignRight)
        colors_grid.addWidget(create_export_color_button('export_copper_color', '#B87333', 'Copper'), 1, 3)

        # Row 2: Wood and Automotive
        colors_grid.addWidget(QLabel("Wood:"), 2, 0, Qt.AlignRight)
        colors_grid.addWidget(create_export_color_button('export_wood_color', '#8B4513', 'Wood'), 2, 1)
        colors_grid.addWidget(QLabel("Auto:"), 2, 2, Qt.AlignRight)
        colors_grid.addWidget(create_export_color_button('export_automotive_color', '#2F4F4F', 'Automotive'), 2, 3)

        # Row 3: Non-232
        colors_grid.addWidget(QLabel("Non-232:"), 3, 0, Qt.AlignRight)
        colors_grid.addWidget(create_export_color_button('export_non232_color', '#FF0000', 'Non-232'), 3, 1)

        options_layout.addWidget(colors_group)

        # --- Middle: Column Visibility ---
        visibility_group = QGroupBox("Column Visibility")
        visibility_layout = QGridLayout(visibility_group)
        visibility_layout.setSpacing(5)
        visibility_layout.setContentsMargins(10, 10, 10, 10)

        self.output_column_visibility = {}
        ratio_columns = ['SteelRatio', 'AluminumRatio', 'CopperRatio', 'WoodRatio', 'AutoRatio', 'NonSteelRatio']

        for idx, col in enumerate(ratio_columns):
            is_visible = True
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("SELECT value FROM app_config WHERE key = ?", (f'export_col_visible_{col}',))
                row = c.fetchone()
                conn.close()
                if row:
                    is_visible = row[0] == 'True'
            except:
                pass

            checkbox = QCheckBox(col.replace('Ratio', '%'))
            checkbox.setChecked(is_visible)
            checkbox.stateChanged.connect(lambda state, col=col: self.update_column_visibility(col, state))
            self.output_column_visibility[col] = checkbox
            # 2 columns layout
            visibility_layout.addWidget(checkbox, idx // 2, idx % 2)

        options_layout.addWidget(visibility_group)

        # --- Right: Export Options ---
        export_options_group = QGroupBox("Export Options")
        export_options_layout = QVBoxLayout(export_options_group)
        export_options_layout.setContentsMargins(10, 10, 10, 10)

        self.split_by_invoice = False
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", ('export_split_by_invoice',))
            row = c.fetchone()
            conn.close()
            if row:
                self.split_by_invoice = row[0] == 'True'
        except:
            pass

        self.split_by_invoice_checkbox = QCheckBox("Split by Invoice Number")
        self.split_by_invoice_checkbox.setChecked(self.split_by_invoice)
        self.split_by_invoice_checkbox.stateChanged.connect(self.update_split_by_invoice_setting)
        export_options_layout.addWidget(self.split_by_invoice_checkbox)

        split_note = QLabel("Creates separate files per invoice.\nRequires Invoice Number mapping.")
        split_note.setStyleSheet("color: gray; font-size: 9pt;")
        split_note.setWordWrap(True)
        export_options_layout.addWidget(split_note)
        export_options_layout.addStretch()

        options_layout.addWidget(export_options_group)
        options_layout.addStretch()

        layout.addWidget(options_widget)

        # === COLUMN MAPPING SECTION ===
        mapping_group = QGroupBox("Column Name Mapping")
        mapping_layout = QVBoxLayout(mapping_group)
        mapping_layout.setContentsMargins(10, 10, 10, 10)

        # Scrollable area for column mappings
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Create form for column name mappings
        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setLabelAlignment(Qt.AlignRight)

        # Default column mappings (internal_name: display_name)
        self.default_output_columns = {
            'Product No': 'Product No',
            'ValueUSD': 'ValueUSD',
            'HTSCode': 'HTSCode',
            'MID': 'MID',
            'CalcWtNet': 'CalcWtNet',
            'DecTypeCd': 'DecTypeCd',
            'CountryofMelt': 'CountryofMelt',
            'CountryOfCast': 'CountryOfCast',
            'PrimCountryOfSmelt': 'PrimCountryOfSmelt',
            'PrimSmeltFlag': 'PrimSmeltFlag',
            'SteelRatio': 'SteelRatio',
            'AluminumRatio': 'AluminumRatio',
            'CopperRatio': 'CopperRatio',
            'WoodRatio': 'WoodRatio',
            'AutoRatio': 'AutoRatio',
            'NonSteelRatio': 'NonSteelRatio',
            '232_Status': '232_Status'
        }

        # Initialize current mapping if not exists or is None
        if not hasattr(self, 'output_column_mapping') or self.output_column_mapping is None:
            self.output_column_mapping = self.default_output_columns.copy()

        # Create text inputs for each column
        self.output_column_inputs = {}
        for internal_name, display_name in self.output_column_mapping.items():
            line_edit = QLineEdit(display_name)
            line_edit.setMinimumWidth(200)
            line_edit.textChanged.connect(lambda text, key=internal_name: self.update_output_column_name(key, text))
            self.output_column_inputs[internal_name] = line_edit
            form_layout.addRow(f"{internal_name}:", line_edit)

        scroll_layout.addWidget(form_widget)
        scroll.setWidget(scroll_widget)
        mapping_layout.addWidget(scroll)
        layout.addWidget(mapping_group, 1)
        self.tab_output_map.setLayout(layout)

    def update_output_column_name(self, internal_name, new_name):
        """Update the output column mapping when user changes a name"""
        self.output_column_mapping[internal_name] = new_name

    def update_column_visibility(self, col_name, state):
        """Save column visibility setting to database"""
        is_visible = state == 2  # Qt.Checked = 2
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                      (f'export_col_visible_{col_name}', str(is_visible)))
            conn.commit()
            conn.close()
            logger.info(f"Column visibility updated: {col_name} = {is_visible}")
            self.bottom_status.setText(f"{col_name} export visibility: {'visible' if is_visible else 'hidden'}")
        except Exception as e:
            logger.error(f"Failed to save column visibility: {e}")

    def update_split_by_invoice_setting(self, state):
        """Save split by invoice setting to database"""
        self.split_by_invoice = state == 2  # Qt.Checked = 2
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)",
                      ('export_split_by_invoice', str(self.split_by_invoice)))
            conn.commit()
            conn.close()
            logger.info(f"Split by invoice setting updated: {self.split_by_invoice}")
            self.bottom_status.setText(f"Split by invoice: {'enabled' if self.split_by_invoice else 'disabled'}")
        except Exception as e:
            logger.error(f"Failed to save split by invoice setting: {e}")

    def pick_output_font_color(self):
        """Open color picker for output font color (per-user setting)"""
        color = QColorDialog.getColor(QColor(self.output_font_color), self, "Choose Export Font Color")
        if color.isValid():
            color_hex = color.name()
            self.output_font_color = color_hex
            if is_widget_valid(self.output_font_color_btn):
                self.output_font_color_btn.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999;")
            # Save to per-user settings
            set_user_setting('output_font_color', color_hex)
            logger.info(f"Saved output font color: {color_hex}")
            self.bottom_status.setText(f"Output font color set to {color_hex}")

    def reset_output_mapping(self):
        """Reset output column mapping to default values"""
        self.output_column_mapping = self.default_output_columns.copy()
        if hasattr(self, 'output_column_inputs'):
            for internal_name, line_edit in self.output_column_inputs.items():
                if is_widget_valid(line_edit):
                    line_edit.blockSignals(True)
                    line_edit.setText(self.default_output_columns[internal_name])
                    line_edit.blockSignals(False)
        self.bottom_status.setText("Output mapping reset to default")
        logger.info("Output column mapping reset to default")

    def load_output_mapping_profiles(self):
        """Load output mapping profiles from database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("SELECT profile_name FROM output_column_mappings ORDER BY created_date DESC", conn)
            conn.close()

            profile_names = df['profile_name'].tolist()

            # Update the Configuration dialog combo box (if still valid)
            if hasattr(self, 'output_profile_combo') and is_widget_valid(self.output_profile_combo):
                self.output_profile_combo.blockSignals(True)
                self.output_profile_combo.clear()
                self.output_profile_combo.addItem("-- Select Profile --")
                for name in profile_names:
                    self.output_profile_combo.addItem(name)
                self.output_profile_combo.blockSignals(False)

            # Update the quick access combo box on Process Shipment tab (if still valid)
            if hasattr(self, 'quick_export_profile_combo') and is_widget_valid(self.quick_export_profile_combo):
                current_selection = self.quick_export_profile_combo.currentText()
                self.quick_export_profile_combo.blockSignals(True)
                self.quick_export_profile_combo.clear()
                self.quick_export_profile_combo.addItem("-- Select Profile --")
                for name in profile_names:
                    self.quick_export_profile_combo.addItem(name)
                # Restore selection if it still exists
                idx = self.quick_export_profile_combo.findText(current_selection)
                if idx >= 0:
                    self.quick_export_profile_combo.setCurrentIndex(idx)
                self.quick_export_profile_combo.blockSignals(False)

            logger.info(f"Loaded {len(df)} output mapping profiles")
        except Exception as e:
            logger.error(f"Failed to load output mapping profiles: {e}")

    def on_quick_export_profile_changed(self, profile_name):
        """Handle export profile change from Process Shipment tab"""
        if not profile_name or profile_name == "-- Select Profile --":
            return

        # Load the profile (reuse existing method)
        self.load_output_mapping_profile(profile_name)

        # Sync the Configuration dialog combo if it exists and is still valid
        if hasattr(self, 'output_profile_combo') and is_widget_valid(self.output_profile_combo):
            self.output_profile_combo.blockSignals(True)
            self.output_profile_combo.setCurrentText(profile_name)
            self.output_profile_combo.blockSignals(False)

        self.bottom_status.setText(f"Export profile: {profile_name}")

    def load_output_mapping_profile(self, profile_name):
        """Load selected output mapping profile"""
        if not profile_name or profile_name == "-- Select Profile --":
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT mapping_json FROM output_column_mappings WHERE profile_name=?", (profile_name,))
            row = c.fetchone()
            conn.close()

            if row:
                profile_data = json.loads(row[0])

                # Check if this is the new format (with nested structure) or old format (just column mapping)
                if 'column_mapping' in profile_data:
                    # New format
                    self.output_column_mapping = profile_data.get('column_mapping', {})
                    column_visibility = profile_data.get('column_visibility', {})
                    split_by_invoice = profile_data.get('split_by_invoice', False)
                else:
                    # Old format - just column mapping
                    self.output_column_mapping = profile_data
                    column_visibility = {}
                    split_by_invoice = False

                # Update column name inputs (if Configuration dialog is still open)
                if hasattr(self, 'output_column_inputs'):
                    for internal_name, line_edit in self.output_column_inputs.items():
                        if is_widget_valid(line_edit):
                            line_edit.blockSignals(True)
                            line_edit.setText(self.output_column_mapping.get(internal_name, internal_name))
                            line_edit.blockSignals(False)

                # Update column visibility checkboxes (if Configuration dialog is still open)
                if hasattr(self, 'output_column_visibility'):
                    for col_name, checkbox in self.output_column_visibility.items():
                        if is_widget_valid(checkbox):
                            checkbox.blockSignals(True)
                            # Default to True if not specified in profile
                            is_visible = column_visibility.get(col_name, True)
                            checkbox.setChecked(is_visible)
                            checkbox.blockSignals(False)
                        # Save to database regardless of widget state
                        is_visible = column_visibility.get(col_name, True)
                        self.update_column_visibility(col_name, 2 if is_visible else 0)

                # Update split by invoice checkbox (if Configuration dialog is still open)
                if hasattr(self, 'split_by_invoice_checkbox') and is_widget_valid(self.split_by_invoice_checkbox):
                    self.split_by_invoice_checkbox.blockSignals(True)
                    self.split_by_invoice_checkbox.setChecked(split_by_invoice)
                    self.split_by_invoice_checkbox.blockSignals(False)
                self.split_by_invoice = split_by_invoice
                # Save to database regardless of widget state
                self.update_split_by_invoice_setting(2 if split_by_invoice else 0)

                # Sync the quick export profile combo on Process Shipment tab
                if hasattr(self, 'quick_export_profile_combo') and is_widget_valid(self.quick_export_profile_combo):
                    self.quick_export_profile_combo.blockSignals(True)
                    self.quick_export_profile_combo.setCurrentText(profile_name)
                    self.quick_export_profile_combo.blockSignals(False)

                self.bottom_status.setText(f"Loaded output mapping profile: {profile_name}")
                logger.info(f"Loaded output mapping profile: {profile_name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load profile:\n{e}")
            logger.error(f"Failed to load output mapping profile: {e}")

    def save_output_mapping_profile(self):
        """Save current output column mapping as a named profile"""
        name, ok = QInputDialog.getText(self, "Save Output Mapping Profile", "Enter profile name:")
        if not ok or not name.strip():
            return
        name = name.strip()

        # Check if profile exists
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT profile_name FROM output_column_mappings WHERE profile_name=?", (name,))
            exists = c.fetchone()

            if exists:
                reply = QMessageBox.question(self, "Overwrite?", f"Profile '{name}' exists. Overwrite?")
                if reply != QMessageBox.Yes:
                    conn.close()
                    return

            # Build column visibility dict from checkboxes
            column_visibility = {}
            if hasattr(self, 'output_column_visibility'):
                for col_name, checkbox in self.output_column_visibility.items():
                    column_visibility[col_name] = checkbox.isChecked()

            # Get split by invoice setting
            split_by_invoice = self.split_by_invoice if hasattr(self, 'split_by_invoice') else False

            # Save all settings in new format
            profile_data = {
                'column_mapping': self.output_column_mapping,
                'column_visibility': column_visibility,
                'split_by_invoice': split_by_invoice
            }
            mapping_str = json.dumps(profile_data)
            now = datetime.now().isoformat()

            c.execute("""INSERT OR REPLACE INTO output_column_mappings (profile_name, mapping_json, created_date)
                         VALUES (?,?,?)""", (name, mapping_str, now))
            conn.commit()
            conn.close()

            self.load_output_mapping_profiles()
            if is_widget_valid(self.output_profile_combo):
                self.output_profile_combo.setCurrentText(name)
            self.bottom_status.setText(f"Saved output mapping profile: {name}")
            logger.info(f"Saved output mapping profile: {name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save profile:\n{e}")
            logger.error(f"Failed to save output mapping profile: {e}")

    def delete_output_mapping_profile(self):
        """Delete selected output mapping profile"""
        if not is_widget_valid(self.output_profile_combo):
            return
        profile_name = self.output_profile_combo.currentText()
        if not profile_name or profile_name == "-- Select Profile --":
            QMessageBox.information(self, "No Profile Selected", "Please select a profile to delete.")
            return

        reply = QMessageBox.question(self, "Confirm Delete",
                                     f"Delete profile '{profile_name}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM output_column_mappings WHERE profile_name=?", (profile_name,))
            conn.commit()
            conn.close()

            self.load_output_mapping_profiles()
            self.bottom_status.setText(f"Deleted output mapping profile: {profile_name}")
            logger.info(f"Deleted output mapping profile: {profile_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete profile:\n{e}")
            logger.error(f"Failed to delete output mapping profile: {e}")

    def load_csv_for_shipment_mapping(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Invoice File",
            str(INPUT_DIR),
            "All Supported (*.csv *.xlsx);;CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        if not path: return
        try:
            # Get header row value from input field
            header_row = 0  # Default: first row is header
            if hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
                try:
                    header_row_value = int(self.header_row_input.text().strip())
                    # Convert from 1-based to 0-based indexing
                    header_row = max(0, header_row_value - 1)
                except ValueError:
                    header_row = 0

            # Determine file type and extract data accordingly
            file_ext = Path(path).suffix.lower()

            if file_ext == '.pdf':
                # Extract using pdfplumber table extraction
                df = self.extract_pdf_table(path)
                logger.info(f"PDF detected: {Path(path).name} - using table extraction")
                # Move PDF to processed folder after successful extraction
                self.move_pdf_to_processed(path)
            elif file_ext == '.xlsx':
                df = pd.read_excel(path, nrows=0, dtype=str, header=header_row)
            else:  # .csv
                df = pd.read_csv(path, nrows=0, dtype=str, header=header_row)

            cols = list(df.columns)

            # Clear existing labels
            for label in self.shipment_drag_labels:
                label.setParent(None)
            self.shipment_drag_labels = []

            # Add new labels from extracted columns
            left_layout = self.shipment_widget.layout().itemAt(0).widget().layout()
            for col in cols:
                lbl = DraggableLabel(str(col))
                left_layout.insertWidget(left_layout.count()-1, lbl)
                self.shipment_drag_labels.append(lbl)

            # Determine file type for status message
            file_type = "PDF" if file_ext == '.pdf' else ("Excel" if file_ext == '.xlsx' else "CSV")
            logger.info(f"{file_type} file loaded for mapping: {Path(path).name}")
            self.status.setText(f"{file_type} file loaded: {Path(path).name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot read file:\n{e}")
            logger.error(f"File loading failed: {str(e)}")

    def extract_pdf_table(self, pdf_path):
        """
        Extract tabular data from PDF invoices using pdfplumber.

        Attempts to find and extract the first valid table from the PDF.
        If no table found, tries text-based extraction as fallback.
        Returns a DataFrame with the extracted data.

        Args:
            pdf_path (str): Path to PDF file

        Returns:
            pd.DataFrame: DataFrame with extracted table data

        Raises:
            Exception: If PDF cannot be processed or no table found
        """
        raise Exception("PDF support for invoice extraction has been removed in this version. Please use CSV or Excel files.")

        # PDF support removed; all code below is unreachable

    def _extract_pdf_text_fallback(self, pdf_path):
        """
        Fallback method to extract text-based data from PDFs without tables.

        Attempts supplier-specific extraction (e.g., AROMATE invoices).
        Falls back to generic text extraction if no supplier pattern matches.

        Args:
            pdf_path (str): Path to PDF file

        Returns:
            pd.DataFrame: DataFrame with extracted data or text lines

        Raises:
            Exception: If PDF cannot be read
        """
        raise Exception("PDF support for invoice extraction has been removed in this version. Please use CSV or Excel files.")
        # PDF support removed; all code below is unreachable

    def _extract_aromate_invoice(self, text):
        """
        Extract AROMATE invoice data using regex pattern matching.

        Extracts SKU, quantity, unit price, and total price from AROMATE invoices.

        Args:
            text (str): Extracted PDF text

        Returns:
            pd.DataFrame: DataFrame with part_number, quantity, unit_price, total_price
        """
        import re

        # Pattern matches both formats:
        # Format 1: SKU# 1562485 76,080 PCS USD 0.6580 USD 50,060.64
        # Format 2: SKU# 2641486 15,120 PCS 0.7140 10,795.68
        pattern = r'SKU#\s*(\d+)\s+(\d+(?:,\d{3})*)\s+PCS\s+(?:USD\s+)?([\d.]+)\s+(?:USD\s+)?([\d,]+\.\d{2})'

        matches = re.findall(pattern, text)

        if not matches:
            logger.warning("AROMATE pattern not found, falling back to generic text extraction")
            all_text = [line.strip() for line in text.split('\n') if line.strip()]
            return pd.DataFrame({'text_line': all_text})

        # Convert matches to DataFrame
        data = []
        for sku, qty, unit_price, total_price in matches:
            data.append({
                'part_number': sku,
                'quantity': int(qty.replace(',', '')),
                'unit_price': float(unit_price),
                'total_price': float(total_price.replace(',', ''))
            })

        df = pd.DataFrame(data)
        logger.info(f"AROMATE invoice extraction successful: {len(df)} items extracted")
        return df

    def on_shipment_drop(self, field_key, column_name):
        # Check if shipment_targets is valid
        if hasattr(self, 'shipment_targets') and self.shipment_targets:
            try:
                for k, t in self.shipment_targets.items():
                    if t and hasattr(t, 'setText'):
                        if t.column_name == column_name and k != field_key:
                            t.column_name = None
                            t.setText(f"Drop {t.field_key} here")
                            t.setProperty("occupied", False)
                            t.style().unpolish(t); t.style().polish(t)
            except (RuntimeError, AttributeError):
                pass  # Widgets have been deleted

        self.shipment_mapping[field_key] = column_name
        SHIPMENT_MAPPING_FILE.write_text(json.dumps(self.shipment_mapping, indent=2))
        logger.info(f"Shipment mapping saved: {field_key} to {column_name}")

    def reset_current_mapping(self):
        self.shipment_mapping = {}

        # Clear drop targets (right side) - only if they exist and are valid
        if hasattr(self, 'shipment_targets') and self.shipment_targets:
            try:
                for target in self.shipment_targets.values():
                    if target and hasattr(target, 'setText'):
                        target.column_name = None
                        target.setText(f"Drop {target.field_key} here")
                        target.setProperty("occupied", False)
                        target.style().unpolish(target); target.style().polish(target)
            except (RuntimeError, AttributeError):
                pass  # Widgets have been deleted

        # Clear CSV columns drag labels (left side) - only if they exist and are valid
        if hasattr(self, 'shipment_drag_labels') and self.shipment_drag_labels:
            try:
                for label in self.shipment_drag_labels:
                    if label and hasattr(label, 'setParent'):
                        label.setParent(None)
                        label.deleteLater()
                self.shipment_drag_labels = []
            except (RuntimeError, AttributeError):
                pass  # Widgets have been deleted

        self.status.setText("Current mapping reset")

    def load_mapping_profiles(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("SELECT profile_name FROM mapping_profiles ORDER BY created_date DESC", conn)
            conn.close()
            
            # Update Process tab combo (tab 0 - always initialized)
            if hasattr(self, 'profile_combo'):
                self.profile_combo.blockSignals(True)
                self.profile_combo.clear()
                self.profile_combo.addItem("-- Select Profile --")
                for name in df['profile_name'].tolist():
                    self.profile_combo.addItem(name)
                self.profile_combo.blockSignals(False)
            
            # Update Invoice Mapping Profiles tab combo (tab 1 - may not be initialized yet)
            if hasattr(self, 'profile_combo_map'):
                self.profile_combo_map.blockSignals(True)
                self.profile_combo_map.clear()
                self.profile_combo_map.addItem("-- Select Profile --")
                for name in df['profile_name'].tolist():
                    self.profile_combo_map.addItem(name)
                self.profile_combo_map.blockSignals(False)
                
            logger.info(f"Loaded {len(df)} mapping profiles")
        except Exception as e:
            logger.error(f"Failed to load mapping profiles: {e}")

    def save_mapping_profile(self):
        name, ok = QInputDialog.getText(self, "Save Mapping Profile", "Enter profile name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        if self.profile_combo.findText(name) != -1:
            if QMessageBox.question(self, "Overwrite?", f"Profile '{name}' exists. Overwrite?") != QMessageBox.Yes:
                return

        mapping_str = json.dumps(self.shipment_mapping)

        # Get header row value from input field
        header_row_value = 1  # Default
        if hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
            try:
                header_row_value = int(self.header_row_input.text().strip())
            except ValueError:
                header_row_value = 1

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO mapping_profiles (profile_name, mapping_json, header_row) VALUES (?, ?, ?)",
                      (name, mapping_str, header_row_value))
            conn.commit()
            conn.close()
            self.load_mapping_profiles()
            # Only update the combo on the Invoice Mapping Profiles tab (where save button is)
            self.profile_combo_map.setCurrentText(name)
            logger.success(f"Mapping profile saved: {name} (header_row={header_row_value})")
            self.status.setText(f"Profile saved: {name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save failed: {e}")

    def load_selected_profile(self, name):
        if not name or name == "-- Select Profile --":
            return
        self.load_selected_profile_full(name)

    def load_selected_profile_full(self, name):
        if not name or name == "-- Select Profile --":
            # Clear all mappings and reset UI
            self.shipment_mapping = {}
            self.profile_header_row = 1  # Reset to default

            # Clear drop targets (only if they exist and are valid)
            if hasattr(self, 'shipment_targets') and self.shipment_targets:
                try:
                    for target in self.shipment_targets.values():
                        if target and hasattr(target, 'setText'):
                            target.column_name = None
                            target.setText(f"Drop {target.field_key.replace('_', ' ')} here")
                            target.setProperty("occupied", False)
                            target.style().unpolish(target)
                            target.style().polish(target)
                except (RuntimeError, AttributeError):
                    pass  # Widgets have been deleted

            # Clear draggable CSV columns (only if they exist and are valid)
            if hasattr(self, 'shipment_drag_labels') and self.shipment_drag_labels:
                try:
                    for label in self.shipment_drag_labels:
                        if label and hasattr(label, 'deleteLater'):
                            label.deleteLater()
                    self.shipment_drag_labels.clear()
                except (RuntimeError, AttributeError):
                    pass  # Widgets have been deleted

            # Reset header row to default (only if widget exists and is valid)
            if hasattr(self, 'header_row_input') and self.header_row_input:
                try:
                    if hasattr(self.header_row_input, 'setText'):
                        self.header_row_input.setText("1")
                except (RuntimeError, AttributeError):
                    pass  # Widget has been deleted

            self.bottom_status.setText("Profile cleared")
            return

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT mapping_json, header_row FROM mapping_profiles WHERE profile_name = ?", (name,))
            row = c.fetchone()
            conn.close()
            if row:
                self.shipment_mapping = json.loads(row[0])
                # Restore header row value
                header_row_value = row[1] if len(row) > 1 and row[1] is not None else 1
                # Store as instance variable for use in Process Shipments tab
                self.profile_header_row = header_row_value
                if hasattr(self, 'header_row_input') and self.header_row_input:
                    try:
                        if hasattr(self.header_row_input, 'setText'):
                            self.header_row_input.setText(str(header_row_value))
                    except (RuntimeError, AttributeError):
                        pass  # Widget has been deleted
                self.apply_current_mapping()
                logger.info(f"Profile loaded: {name} (header_row={header_row_value})")
                self.bottom_status.setText(f"Loaded profile: {name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Load failed: {e}")

    def delete_mapping_profile(self):
        # Get profile name from Invoice Mapping Profiles tab combo (where delete button is)
        name = self.profile_combo_map.currentText()
        if not name or name == "-- Select Profile --":
            return
        if QMessageBox.question(self, "Delete", f"Delete profile '{name}'?") != QMessageBox.Yes:
            return
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM mapping_profiles WHERE profile_name = ?", (name,))
            conn.commit()
            conn.close()
            self.load_mapping_profiles()
            # Reset both combos to default after deletion
            self.profile_combo.setCurrentIndex(0)
            self.profile_combo_map.setCurrentIndex(0)
            logger.info(f"Profile deleted: {name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Delete failed: {e}")

    def apply_current_mapping(self):
        # Check if shipment_targets is valid (not deleted after dialog close)
        if not hasattr(self, 'shipment_targets') or not self.shipment_targets:
            return

        # Check if the widgets in shipment_targets are still valid
        try:
            # Try to actually access the widget's properties to see if it's been deleted
            for target in self.shipment_targets.values():
                if not target:
                    return
                # Try to access the widget - this will raise RuntimeError if deleted
                _ = target.isVisible()
                break  # Just check the first one
        except (RuntimeError, AttributeError):
            # Widgets have been deleted
            logger.debug("apply_current_mapping: shipment_targets widgets have been deleted, skipping")
            return

        # Batch UI updates to prevent GUI freezing
        try:
            for key, target in self.shipment_targets.items():
                # Additional safety check per widget
                try:
                    _ = target.isVisible()
                except (RuntimeError, AttributeError):
                    logger.debug(f"apply_current_mapping: target widget for {key} has been deleted, skipping")
                    continue

                col = self.shipment_mapping.get(key)
                if col:
                    target.column_name = col
                    target.setText(f"{key}\n<- {col}")
                    target.setProperty("occupied", True)
                else:
                    target.column_name = None
                    target.setText(f"Drop {key.replace('_', ' ')} here")
                    target.setProperty("occupied", False)

            # Apply all style updates at once after setting properties
            for target in self.shipment_targets.values():
                try:
                    _ = target.isVisible()
                    target.style().unpolish(target)
                    target.style().polish(target)
                except (RuntimeError, AttributeError):
                    continue
        except (RuntimeError, AttributeError) as e:
            logger.debug(f"apply_current_mapping: Error during mapping update: {e}")
            return

    def setup_master_tab(self):
        layout = QVBoxLayout(self.tab_master)
        title = QLabel("<h2>Parts View - Click any cell to edit</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        edit_box = QHBoxLayout()
        btn_add = QPushButton("Add Row")
        btn_add.setStyleSheet(self.get_button_style("success"))
        btn_del = QPushButton("Delete Selected")
        btn_del.setStyleSheet(self.get_button_style("danger"))
        btn_save = QPushButton("Save Changes")
        btn_save.setStyleSheet(self.get_button_style("success"))
        btn_refresh = QPushButton("Refresh")
        btn_refresh.setStyleSheet(self.get_button_style("info"))
        btn_import_units = QPushButton("Import HTS Units")
        btn_import_units.setStyleSheet(self.get_button_style("secondary"))
        btn_import_units.setToolTip("Import CBP Qty1 units from HTS reference file")
        btn_add.clicked.connect(self.add_part_row)
        btn_del.clicked.connect(self.delete_selected_parts)
        btn_save.clicked.connect(self.save_parts_table)
        btn_refresh.clicked.connect(self.refresh_parts_table)
        btn_import_units.clicked.connect(self.import_hts_units)
        edit_box.addWidget(QLabel("Edit:"))
        edit_box.addWidget(btn_add); edit_box.addWidget(btn_del); edit_box.addWidget(btn_save); edit_box.addWidget(btn_refresh)
        edit_box.addWidget(btn_import_units)
        edit_box.addStretch()
        layout.addLayout(edit_box)

        # SQL Query Builder
        query_group = QGroupBox("SQL Query Builder")
        query_layout = QVBoxLayout()
        
        query_controls = QHBoxLayout()
        query_controls.addWidget(QLabel("SELECT * FROM parts_master WHERE"))
        
        self.query_field = QComboBox()
        self.query_field.addItems(["part_number", "description", "hts_code", "country_origin", "mid", "client_code", "steel_ratio", "aluminum_ratio", "copper_ratio", "wood_ratio", "non_steel_ratio", "cbp_qty1", "Sec301_Exclusion_Tariff"])
        query_controls.addWidget(self.query_field)
        
        self.query_operator = QComboBox()
        self.query_operator.addItems(["=", "LIKE", ">", "<", ">=", "<=", "!="])
        query_controls.addWidget(self.query_operator)
        
        self.query_value = QLineEdit()
        self.query_value.setPlaceholderText("Enter value...")
        self.query_value.setReadOnly(False)
        self.query_value.setEnabled(True)
        self.query_value.setStyleSheet(self.get_input_style())
        query_controls.addWidget(self.query_value, 1)
        
        btn_run_query = QPushButton("Run Query")
        btn_run_query.setStyleSheet(self.get_button_style("info"))
        btn_run_query.clicked.connect(self.run_custom_query)
        query_controls.addWidget(btn_run_query)
        
        btn_clear_query = QPushButton("Show All")
        btn_clear_query.setStyleSheet(self.get_button_style("default"))
        btn_clear_query.clicked.connect(self.refresh_parts_table)
        query_controls.addWidget(btn_clear_query)
        
        query_layout.addLayout(query_controls)
        
        # Custom SQL input
        custom_sql_layout = QHBoxLayout()
        custom_sql_layout.addWidget(QLabel("Custom SQL:"))
        self.custom_sql_input = QLineEdit()
        self.custom_sql_input.setPlaceholderText("SELECT * FROM parts_master WHERE ...")
        self.custom_sql_input.setReadOnly(False)
        self.custom_sql_input.setEnabled(True)
        self.custom_sql_input.setStyleSheet(self.get_input_style())
        custom_sql_layout.addWidget(self.custom_sql_input, 1)
        btn_run_custom = QPushButton("Execute")
        btn_run_custom.setStyleSheet(self.get_button_style("success"))
        btn_run_custom.clicked.connect(self.run_custom_sql)
        custom_sql_layout.addWidget(btn_run_custom)
        query_layout.addLayout(custom_sql_layout)
        
        self.query_result_label = QLabel("Ready")
        self.query_result_label.setStyleSheet("padding:5px; background:#f0f0f0;")
        query_layout.addWidget(self.query_result_label)
        
        query_group.setLayout(query_layout)
        layout.addWidget(query_group)

        search_box = QHBoxLayout()
        search_box.addWidget(QLabel("Quick Search:"))
        self.search_field_combo = QComboBox()
        self.search_field_combo.addItems(["All Fields","part_number","description","hts_code","country_origin","mid","client_code","steel_ratio","aluminum_ratio","copper_ratio","wood_ratio","auto_ratio","non_steel_ratio","cbp_qty1","Sec301_Exclusion_Tariff"])
        # Refocus search input after combo selection
        self.search_field_combo.currentIndexChanged.connect(lambda: self.search_input.setFocus())
        search_box.addWidget(self.search_field_combo)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Type to filter...")
        self.search_input.setReadOnly(False)
        self.search_input.setEnabled(True)
        self.search_input.setStyleSheet(self.get_input_style())
        self.search_input.textChanged.connect(self.filter_parts_table)
        search_box.addWidget(self.search_input, 1)
        layout.addLayout(search_box)

        table_box = QGroupBox("Parts Master Table")
        tl = QVBoxLayout()
        self.parts_table = QTableWidget()
        self.parts_table.setColumnCount(15)
        self.parts_table.setHorizontalHeaderLabels([
            "part_number", "description", "hts_code", "country_origin", "mid", "client_code", "steel_ratio", "aluminum_ratio", "copper_ratio", "wood_ratio", "auto_ratio", "non_steel_ratio", "cbp_qty1", "Sec301_Exclusion_Tariff", "updated_date"
        ])
        self.parts_table.setEditTriggers(QTableWidget.AllEditTriggers)
        self.parts_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.parts_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.parts_table.setSortingEnabled(False)  # Disabled for better performance
        tl.addWidget(self.parts_table)
        table_box.setLayout(tl)
        layout.addWidget(table_box, 1)

        self.refresh_parts_table()
        self.tab_master.setLayout(layout)

    def refresh_parts_table(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            # Use explicit column ordering to match header labels
            df = pd.read_sql("""
                SELECT part_number, description, hts_code, country_origin, mid, client_code,
                       steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio, non_steel_ratio,
                       cbp_qty1, Sec301_Exclusion_Tariff, last_updated as updated_date
                FROM parts_master ORDER BY part_number
            """, conn)
            conn.close()
            self.populate_parts_table(df)
            self.query_result_label.setText("Showing all parts")
            self.query_result_label.setStyleSheet("padding:5px; background:#f0f0f0;")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot load parts:\n{e}")

    def import_hts_units(self):
        """Import CBP Qty1 units from HTS reference file and update parts_master"""
        # Look for the reference file in Resources/References
        ref_file = BASE_DIR / "Resources" / "References" / "HTS_qty1.xlsx"
        
        if not ref_file.exists():
            # Allow user to select a file
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select HTS Units Reference File",
                str(BASE_DIR / "Resources" / "References"),
                "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*.*)"
            )
            if not file_path:
                return
            ref_file = Path(file_path)
        
        try:
            # Read the reference file
            if ref_file.suffix.lower() in ['.xlsx', '.xls']:
                df_ref = pd.read_excel(ref_file)
            else:
                df_ref = pd.read_csv(ref_file)
            
            # Expect columns: 'Tariff No' and 'Uom 1'
            if 'Tariff No' not in df_ref.columns or 'Uom 1' not in df_ref.columns:
                QMessageBox.warning(self, "Invalid File", 
                    "Reference file must have 'Tariff No' and 'Uom 1' columns.")
                return
            
            # Clean up HTS codes (remove dots for matching)
            def normalize_hts(hts):
                if pd.isna(hts):
                    return ""
                return str(hts).replace(".", "").strip()
            
            # Create lookup dictionary
            hts_units = {}
            for _, row in df_ref.iterrows():
                hts_code = normalize_hts(row['Tariff No'])
                unit = str(row['Uom 1']).strip() if pd.notna(row['Uom 1']) else ""
                if hts_code and unit:
                    hts_units[hts_code] = unit
            
            # Update database
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            
            # Get all parts with HTS codes
            c.execute("SELECT part_number, hts_code FROM parts_master WHERE hts_code IS NOT NULL AND hts_code != ''")
            parts = c.fetchall()
            
            updated = 0
            for part_number, hts_code in parts:
                normalized = normalize_hts(hts_code)
                if normalized in hts_units:
                    c.execute("UPDATE parts_master SET cbp_qty1=? WHERE part_number=?",
                              (hts_units[normalized], part_number))
                    updated += c.rowcount
            
            conn.commit()
            conn.close()
            
            # Refresh the table
            self.refresh_parts_table()
            
            QMessageBox.information(self, "Import Complete", 
                f"Updated {updated} parts with CBP Qty1 units.\n"
                f"Reference file had {len(hts_units)} HTS codes.")
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import HTS units:\n{e}")

    def add_part_row(self):
        row = self.parts_table.rowCount()
        self.parts_table.insertRow(row)
        self.parts_table.setItem(row, 0, QTableWidgetItem("NEW_PART"))

    def delete_selected_parts(self):
        rows = sorted(set(index.row() for index in self.parts_table.selectedIndexes()), reverse=True)
        if not rows:
            QMessageBox.information(self, "Info", "Select rows to delete")
            return
        if QMessageBox.question(self, "Confirm", f"Delete {len(rows)} parts?") != QMessageBox.Yes:
            return
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        deleted = 0
        for row in rows:
            part = self.parts_table.item(row, 0).text().strip()
            if part and part != "NEW_PART":
                c.execute("DELETE FROM parts_master WHERE part_number=?", (part,))
                deleted += c.rowcount
                self.parts_table.removeRow(row)
        conn.commit(); conn.close()
        QMessageBox.information(self, "Success", f"Deleted {deleted} parts")
        self.load_available_mids()

    def save_parts_table(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            now = datetime.now().isoformat()
            saved = 0
            for row in range(self.parts_table.rowCount()):
                items = [self.parts_table.item(row, col) for col in range(15)]
                if not items[0] or not items[0].text().strip(): continue
                part = items[0].text().strip()
                desc = items[1].text() if items[1] else ""
                hts = items[2].text() if items[2] else ""
                origin = (items[3].text() or "").upper()[:2]
                mid = items[4].text() if items[4] else ""
                client_code = items[5].text() if items[5] else ""
                try:
                    steel = float(items[6].text()) if items[6] and items[6].text() else 0.0
                    steel = max(0.0, min(1.0, steel))
                except:
                    steel = 0.0
                try:
                    aluminum = float(items[7].text()) if items[7] and items[7].text() else 0.0
                    aluminum = max(0.0, min(1.0, aluminum))
                except:
                    aluminum = 0.0
                try:
                    copper = float(items[8].text()) if items[8] and items[8].text() else 0.0
                    copper = max(0.0, min(1.0, copper))
                except:
                    copper = 0.0
                try:
                    wood = float(items[9].text()) if items[9] and items[9].text() else 0.0
                    wood = max(0.0, min(1.0, wood))
                except:
                    wood = 0.0
                try:
                    auto = float(items[10].text()) if items[10] and items[10].text() else 0.0
                    auto = max(0.0, min(1.0, auto))
                except:
                    auto = 0.0
                # Non-232 ratio is remainder after all Section 232 materials
                non_steel = max(0.0, 1.0 - steel - aluminum - copper - wood - auto)
                cbp_qty1 = items[12].text() if items[12] else ""
                sec301_exclusion = items[13].text() if items[13] else ""
                c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                          steel_ratio, non_steel_ratio, last_updated, cbp_qty1, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                          Sec301_Exclusion_Tariff)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                          ON CONFLICT(part_number) DO UPDATE SET
                          description=excluded.description, hts_code=excluded.hts_code,
                          country_origin=excluded.country_origin, mid=excluded.mid,
                          client_code=excluded.client_code, steel_ratio=excluded.steel_ratio,
                          non_steel_ratio=excluded.non_steel_ratio, last_updated=excluded.last_updated,
                          cbp_qty1=excluded.cbp_qty1, aluminum_ratio=excluded.aluminum_ratio,
                          copper_ratio=excluded.copper_ratio, wood_ratio=excluded.wood_ratio,
                          auto_ratio=excluded.auto_ratio, Sec301_Exclusion_Tariff=excluded.Sec301_Exclusion_Tariff""",
                          (part, desc, hts, origin, mid, client_code, steel, non_steel, now, cbp_qty1, aluminum, copper, wood, auto, sec301_exclusion))
                if c.rowcount: saved += 1
            conn.commit(); conn.close()
            QMessageBox.information(self, "Success", f"Saved {saved} parts!")
            self.bottom_status.setText("Database saved")
            self.load_available_mids()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save failed:\n{e}")

    def add_not_found_parts_to_db(self):
        """
        Add parts that were not found in the database during processing.
        Uses the part number and MID from the current preview table.
        Returns the count of parts added.
        """
        if self.last_processed_df is None:
            return 0

        # Get rows where _not_in_db is True
        not_found_df = self.last_processed_df[self.last_processed_df['_not_in_db'] == True]
        if not_found_df.empty:
            return 0

        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            now = datetime.now().isoformat()
            added_count = 0
            added_parts = []

            for idx, row in not_found_df.iterrows():
                part_number = str(row.get('Product No', '')).strip()
                if not part_number:
                    continue

                # Check if part already exists (might have been added manually since processing)
                c.execute("SELECT 1 FROM parts_master WHERE part_number = ?", (part_number,))
                if c.fetchone():
                    continue  # Part already exists, skip

                # Get MID from the table at the corresponding row index
                # Find the table row that matches this part number
                table_row = None
                for i in range(self.table.rowCount()):
                    item = self.table.item(i, 0)  # Column 0 is Product No
                    if item and item.text() == part_number:
                        table_row = i
                        break

                if table_row is None:
                    continue

                # Get values from the preview table
                mid = self.table.item(table_row, 3).text() if self.table.item(table_row, 3) else ""
                hts_code = self.table.item(table_row, 2).text() if self.table.item(table_row, 2) else ""

                # Insert the part with minimal information (part_number and MID)
                c.execute("""INSERT INTO parts_master (part_number, description, hts_code, country_origin, mid, client_code,
                          steel_ratio, non_steel_ratio, last_updated, cbp_qty1, aluminum_ratio, copper_ratio, wood_ratio, auto_ratio,
                          Sec301_Exclusion_Tariff)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                          (part_number, '', hts_code, '', mid, '', 0.0, 1.0, now, '', 0.0, 0.0, 0.0, 0.0, ''))

                if c.rowcount:
                    added_count += 1
                    added_parts.append(part_number)

            conn.commit()
            conn.close()

            if added_count > 0:
                logger.info(f"Added {added_count} new parts to database: {added_parts}")
                # Refresh the MID dropdown in case new MIDs were added
                self.load_available_mids()

            return added_count

        except Exception as e:
            logger.error(f"Failed to add not-found parts to database: {e}")
            return 0

    def filter_parts_table(self, text):
        text = text.lower().strip()
        if not text:
            for row in range(self.parts_table.rowCount()):
                self.parts_table.setRowHidden(row, False)
            return
        for row in range(self.parts_table.rowCount()):
            match = any(text in (self.parts_table.item(row, col).text() if self.parts_table.item(row, col) else "").lower() 
                       for col in range(self.parts_table.columnCount()))
            self.parts_table.setRowHidden(row, not match)

    def run_custom_query(self):
        """Execute SQL query builder query"""
        try:
            field = self.query_field.currentText()
            operator = self.query_operator.currentText()
            value = self.query_value.text().strip()
            
            if not value:
                QMessageBox.warning(self, "Query Error", "Please enter a value to search for.")
                return
            
            # Build WHERE clause
            if operator == "LIKE":
                where_clause = f"{field} LIKE ?"
                params = (f"%{value}%",)
            else:
                where_clause = f"{field} {operator} ?"
                params = (value,)
            
            sql = f"SELECT * FROM parts_master WHERE {where_clause} ORDER BY part_number"
            
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn, params=params)
            conn.close()
            
            self.populate_parts_table(df)
            self.query_result_label.setText(f"Query returned {len(df)} results")
            self.query_result_label.setStyleSheet("padding:5px; background:#107C10; color:white;")
            logger.info(f"Query executed: {sql} with params {params}")
            
        except Exception as e:
            logger.error(f"Query execution failed: {e}")
            self.query_result_label.setText(f"Query Error: {str(e)}")
            self.query_result_label.setStyleSheet("padding:5px; background:#A4262C; color:white;")
            QMessageBox.critical(self, "Query Error", f"Failed to execute query:\n{e}")

    def run_custom_sql(self):
        """Execute custom SQL query"""
        try:
            sql = self.custom_sql_input.text().strip()
            
            if not sql:
                QMessageBox.warning(self, "Query Error", "Please enter a SQL query.")
                return
            
            # Basic validation - must be SELECT and FROM parts_master
            if not sql.upper().startswith("SELECT"):
                QMessageBox.warning(self, "Query Error", "Only SELECT queries are allowed.")
                return
            
            if "parts_master" not in sql.lower():
                QMessageBox.warning(self, "Query Error", "Query must reference 'parts_master' table.")
                return
            
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql(sql, conn)
            conn.close()
            
            self.populate_parts_table(df)
            self.query_result_label.setText(f"Custom query returned {len(df)} results")
            self.query_result_label.setStyleSheet("padding:5px; background:#d4edda; color:#155724;")
            logger.info(f"Custom SQL executed: {sql}")
            
        except Exception as e:
            logger.error(f"Custom SQL execution failed: {e}")
            self.query_result_label.setText(f"SQL Error: {str(e)}")
            self.query_result_label.setStyleSheet("padding:5px; background:#A4262C; color:white;")
            QMessageBox.critical(self, "SQL Error", f"Failed to execute SQL:\n{e}")

    def populate_parts_table(self, df):
        """Populate the parts table with a dataframe"""
        self.parts_table.blockSignals(True)
        self.parts_table.setRowCount(len(df))
        # Map table column headers to dataframe columns
        # Headers: part_number, description, hts_code, country_origin, mid, client_code, 
        #          steel_ratio, aluminum_ratio, copper_ratio, wood_ratio, non_steel_ratio, 
        #          cbp_qty1, updated_date
        for i, row in df.iterrows():
            # Column 0: part_number
            self.parts_table.setItem(i, 0, QTableWidgetItem(str(row.get('part_number', '')) if pd.notna(row.get('part_number')) else ""))
            # Column 1: description
            self.parts_table.setItem(i, 1, QTableWidgetItem(str(row.get('description', '')) if pd.notna(row.get('description')) else ""))
            # Column 2: hts_code
            self.parts_table.setItem(i, 2, QTableWidgetItem(str(row.get('hts_code', '')) if pd.notna(row.get('hts_code')) else ""))
            # Column 3: country_origin
            self.parts_table.setItem(i, 3, QTableWidgetItem(str(row.get('country_origin', '')) if pd.notna(row.get('country_origin')) else ""))
            # Column 4: mid
            self.parts_table.setItem(i, 4, QTableWidgetItem(str(row.get('mid', '')) if pd.notna(row.get('mid')) else ""))
            # Column 5: client_code
            self.parts_table.setItem(i, 5, QTableWidgetItem(str(row.get('client_code', '')) if pd.notna(row.get('client_code')) else ""))
            # Column 6: steel_ratio
            self.parts_table.setItem(i, 6, QTableWidgetItem(str(row.get('steel_ratio', 0.0)) if pd.notna(row.get('steel_ratio')) else "0.0"))
            # Column 7: aluminum_ratio
            self.parts_table.setItem(i, 7, QTableWidgetItem(str(row.get('aluminum_ratio', 0.0)) if pd.notna(row.get('aluminum_ratio')) else "0.0"))
            # Column 8: copper_ratio
            self.parts_table.setItem(i, 8, QTableWidgetItem(str(row.get('copper_ratio', 0.0)) if pd.notna(row.get('copper_ratio')) else "0.0"))
            # Column 9: wood_ratio
            self.parts_table.setItem(i, 9, QTableWidgetItem(str(row.get('wood_ratio', 0.0)) if pd.notna(row.get('wood_ratio')) else "0.0"))
            # Column 10: auto_ratio
            self.parts_table.setItem(i, 10, QTableWidgetItem(str(row.get('auto_ratio', 0.0)) if pd.notna(row.get('auto_ratio')) else "0.0"))
            # Column 11: non_steel_ratio
            self.parts_table.setItem(i, 11, QTableWidgetItem(str(row.get('non_steel_ratio', 0.0)) if pd.notna(row.get('non_steel_ratio')) else "0.0"))
            # Column 12: cbp_qty1
            self.parts_table.setItem(i, 12, QTableWidgetItem(str(row.get('cbp_qty1', '')) if pd.notna(row.get('cbp_qty1')) else ""))
            # Column 13: Sec301_Exclusion_Tariff
            self.parts_table.setItem(i, 13, QTableWidgetItem(str(row.get('Sec301_Exclusion_Tariff', '')) if pd.notna(row.get('Sec301_Exclusion_Tariff')) else ""))
            # Column 14: updated_date
            self.parts_table.setItem(i, 14, QTableWidgetItem(str(row.get('updated_date', '')) if pd.notna(row.get('updated_date')) else ""))
        self.parts_table.blockSignals(False)

    # ...existing code...

    def setup_config_tab(self):
        layout = QVBoxLayout(self.tab_config)
        title = QLabel("<h2>Customs Configuration</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Buttons at top
        top_bar = QHBoxLayout()
        btn_import_excel = QPushButton("Import Section 232 Tariffs (CSV/Excel)")
        btn_import_excel.setStyleSheet(self.get_button_style("success"))
        btn_import_excel.clicked.connect(self.import_tariff_232)
        btn_refresh = QPushButton("Refresh View")
        btn_refresh.setStyleSheet(self.get_button_style("info"))
        btn_refresh.clicked.connect(self.refresh_tariff_view)
        top_bar.addWidget(btn_import_excel)
        top_bar.addWidget(btn_refresh)
        top_bar.addStretch()
        layout.addLayout(top_bar)

        # Section 232 Tariff Viewer
        group1 = QGroupBox("Section 232 Tariff List")
        g1_layout = QVBoxLayout()

        # Search and filter
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel("Filter:"))
        self.tariff_filter = QLineEdit()
        self.tariff_filter.setPlaceholderText("Search HTS code, classification, or chapter...")
        self.tariff_filter.setStyleSheet(self.get_input_style())
        self.tariff_filter.textChanged.connect(self.filter_tariff_table)
        filter_bar.addWidget(self.tariff_filter)

        self.tariff_material_filter = QComboBox()
        self.tariff_material_filter.addItems(["All", "Steel", "Aluminum", "Wood", "Copper"])
        self.tariff_material_filter.currentTextChanged.connect(self.filter_tariff_table)
        filter_bar.addWidget(QLabel("Material:"))
        filter_bar.addWidget(self.tariff_material_filter)
        self.bottom_status.setText("Loading Exported Files...")
        QApplication.processEvents()
        # ...existing code...
        self.bottom_status.setText("Ready")
        
        # Add color toggle checkbox
        self.tariff_color_toggle = QCheckBox("Color by Material")
        self.tariff_color_toggle.setChecked(False)  # Disabled by default
        self.tariff_color_toggle.stateChanged.connect(self.filter_tariff_table)
        filter_bar.addWidget(self.tariff_color_toggle)
        
        g1_layout.addLayout(filter_bar)
        
        # Table
        self.tariff_table = QTableWidget()
        self.tariff_table.setColumnCount(7)
        self.tariff_table.setHorizontalHeaderLabels(["HTS Code", "Material", "Classification", 
                       "Chapter", "Chapter Description", 
                       "Declaration", "Notes"])
        self.tariff_table.horizontalHeader().setStretchLastSection(True)
        self.tariff_table.setAlternatingRowColors(True)
        self.tariff_table.setStyleSheet("")
        self.tariff_table.setSortingEnabled(False)  # Disabled for better performance
        g1_layout.addWidget(self.tariff_table)
        
        # Count label
        self.tariff_count_label = QLabel("Total: 0 tariff codes")
        self.tariff_count_label.setStyleSheet("font-weight:bold; padding:5px;")
        g1_layout.addWidget(self.tariff_count_label)
        
        group1.setLayout(g1_layout)
        layout.addWidget(group1)

        self.tab_config.setLayout(layout)
        
        # Load initial data
        self.refresh_tariff_view()

    def import_tariff_232(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Section 232 Tariffs CSV", "", "CSV Files (*.csv);;Excel (*.xlsx)")
        if not path:
            return
        try:
            # Read file based on extension
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path, dtype=str, keep_default_na=False)
            else:
                df = pd.read_excel(path, header=0)
            
            df = df.fillna("")
            
            # Check if it's the new comprehensive format
            if 'HTS Code' in df.columns and 'Material' in df.columns and 'Classification' in df.columns:
                # New comprehensive CSV format with all columns
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM tariff_232")
                
                imported = 0
                for _, row in df.iterrows():
                    hts_code = str(row['HTS Code']).strip().replace(".", "")[:10]
                    material = str(row['Material']).strip()
                    classification = str(row['Classification']).strip()
                    chapter = str(row['Chapter']).strip()
                    chapter_desc = str(row['Chapter Description']).strip()
                    declaration = str(row['Declaration Required']).strip()
                    notes = str(row['Notes']).strip()
                    
                    if hts_code and material in ['Steel', 'Aluminum', 'Wood', 'Copper']:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (hts_code, material, classification, chapter, 
                                  chapter_desc, declaration, notes))
                        imported += 1
                
                conn.commit()
                conn.close()
                QMessageBox.information(self, "Success", 
                    f"Imported {imported} Section 232 tariff codes\n\n"
                    f"Format: Comprehensive 7-column CSV")
                logger.success(f"tariff_232 table updated with {imported} codes (comprehensive format)")
                self.status.setText(f"Section 232 list imported: {imported} codes")
            elif 'HTS Code' in df.columns and 'Material' in df.columns:
                # Simple CSV format (HTS Code, Material only)
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM tariff_232")
                
                imported = 0
                for _, row in df.iterrows():
                    hts_code = str(row['HTS Code']).strip().replace(".", "")[:10]
                    material = str(row['Material']).strip()
                    
                    if hts_code and material in ['Steel', 'Aluminum', 'Wood', 'Copper']:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (hts_code, material, '', '', '', '', ''))
                        imported += 1
                
                conn.commit()
                conn.close()
                QMessageBox.information(self, "Success", f"Imported {imported} Section 232 tariff codes")
                logger.success(f"tariff_232 table updated with {imported} codes")
                self.status.setText(f"Section 232 list imported: {imported} codes")
            else:
                # Legacy Excel format (2 columns: steel, aluminum)
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("DELETE FROM tariff_232")
                steel_codes = [str(x).replace(".", "")[:10] for x in df.iloc[1:, 0] if pd.notna(x) and str(x).strip()]
                alum_codes = [str(x).replace(".", "")[:10] for x in df.iloc[1:, 1] if pd.notna(x) and str(x).strip()]
                for code in steel_codes:
                    if code:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (code, 'Steel', '', '', '', '', ''))
                for code in alum_codes:
                    if code:
                        c.execute("""INSERT OR REPLACE INTO tariff_232 
                                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                 (code, 'Aluminum', '', '', '', '', ''))
                conn.commit()
                conn.close()
                imported = len(steel_codes) + len(alum_codes)
                QMessageBox.information(self, "Success", 
                    f"Imported {imported} 232 codes\n\n"
                    f"Format: Legacy Excel (2-column)")
                logger.success("tariff_232 table updated (legacy format)")
                self.status.setText("Section 232 list imported")
            
            self.refresh_tariff_view()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Import failed: {e}")
            logger.error(f"Section 232 import error: {e}")


    def refresh_tariff_view(self):
        """Load and display all tariff codes from database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("""SELECT hts_code, material, classification, chapter, 
                                       chapter_description, declaration_required, notes 
                                FROM tariff_232 
                                ORDER BY material, chapter, hts_code""", conn)
            conn.close()
            
            self.tariff_full_data = df  # Store for filtering
            self.filter_tariff_table()
            
        except Exception as e:
            logger.error(f"Refresh tariff view failed: {e}")
            self.tariff_table.setRowCount(0)
            self.tariff_count_label.setText("Error loading tariff data")

    def filter_tariff_table(self):
        """Filter and display tariff table based on search criteria"""
        try:
            if not hasattr(self, 'tariff_full_data') or self.tariff_full_data.empty:
                self.tariff_table.setRowCount(0)
                self.tariff_count_label.setText("No tariff data")
                return
            
            df = self.tariff_full_data.copy()
            
            # Apply text filter - search across HTS code, classification, and chapter description
            search_text = self.tariff_filter.text().strip().lower()
            if search_text:
                mask = (df['hts_code'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['classification'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['chapter_description'].astype(str).str.lower().str.contains(search_text, na=False))
                df = df[mask]
            
            # Apply material filter
            material_filter = self.tariff_material_filter.currentText()
            if material_filter != "All":
                df = df[df['material'] == material_filter]
            
            # Populate table
            self.tariff_table.setSortingEnabled(False)
            self.tariff_table.setRowCount(len(df))
            
            for row_idx, (_, row) in enumerate(df.iterrows()):
                # Create items for all 7 columns
                items = [
                    QTableWidgetItem(str(row['hts_code'])),
                    QTableWidgetItem(str(row['material'])),
                    QTableWidgetItem(str(row['classification'])),
                    QTableWidgetItem(str(row['chapter'])),
                    QTableWidgetItem(str(row['chapter_description'])),
                    QTableWidgetItem(str(row['declaration_required'])),
                    QTableWidgetItem(str(row['notes']))
                ]
                
                # Make all items read-only
                for item in items:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                
                # Color code by material (if toggle is enabled)
                if self.tariff_color_toggle.isChecked():
                    material_colors = {
                        'Steel': QColor('#e3f2fd'),      # Light blue
                        'Aluminum': QColor('#fff3e0'),   # Light orange
                        'Wood': QColor('#f1f8e9'),       # Light green
                        'Copper': QColor('#ffe0b2')      # Light copper/bronze
                    }
                    
                    material = row['material']
                    if material in material_colors:
                        bg_color = material_colors[material]
                        # Apply color to entire row for better visibility
                        for item in items:
                            item.setBackground(bg_color)
                
                # Add items to table
                for col_idx, item in enumerate(items):
                    self.tariff_table.setItem(row_idx, col_idx, item)
            
            self.tariff_table.setSortingEnabled(True)
            self.tariff_count_label.setText(f"Total: {len(df)} tariff codes (filtered from {len(self.tariff_full_data)})")
            
        except Exception as e:
            logger.error(f"Filter tariff table failed: {e}")
            logger.trace(traceback.format_exc())
            self.tariff_table.setRowCount(0)

    def setup_log_tab(self):
        layout = QVBoxLayout(self.tab_log)
        title = QLabel("<h2>Log View</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        log_box = QGroupBox("Real-time Log")
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        self.log_text.setContextMenuPolicy(Qt.CustomContextMenu)
        self.log_text.customContextMenuRequested.connect(self.log_context_menu)
        log_layout.addWidget(self.log_text)

        btn_layout = QHBoxLayout()
        btn_copy = QPushButton("Copy to Clipboard")
        btn_copy.setStyleSheet("background:#0078D7; color:white; font-weight:bold;")
        btn_copy.clicked.connect(self.copy_log_to_clipboard)
        btn_clear = QPushButton("Clear Log")
        btn_clear.clicked.connect(lambda: (self.log_text.clear(), logger.logs.clear()))
        btn_layout.addWidget(btn_copy)
        btn_layout.addWidget(btn_clear)
        btn_layout.addStretch()
        log_layout.addLayout(btn_layout)
        
        log_box.setLayout(log_layout)
        layout.addWidget(log_box)

        self.log_timer = QTimer()
        self.log_timer.timeout.connect(self.update_log)
        self.log_timer.start(500)
        self.tab_log.setLayout(layout)

    def setup_actions_tab(self):
        """Section 232 Actions Reference Tab - Chapter 99 tariff actions"""
        layout = QVBoxLayout(self.tab_actions)
        
        # Title
        title = QLabel("<h2>Section 232 Tariff Actions (Chapter 99)</h2>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Info box
        info_box = QGroupBox("Reference Information")
        info_layout = QVBoxLayout()
        info_text = QLabel(
            "This table contains the Chapter 99 tariff action codes and rates for Section 232 duties. "
            "These are the actual tariff numbers used in ACE for declaring Section 232 merchandise."
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)
        info_box.setLayout(info_layout)
        layout.addWidget(info_box)
        
        # Control buttons
        btn_layout = QHBoxLayout()
        btn_import = QPushButton("Import Actions CSV")
        btn_import.setStyleSheet(self.get_button_style("info"))
        btn_import.clicked.connect(self.import_actions_csv)
        btn_layout.addWidget(btn_import)

        btn_refresh = QPushButton("Refresh View")
        btn_refresh.setStyleSheet(self.get_button_style("default"))
        btn_refresh.clicked.connect(self.refresh_actions_view)
        btn_layout.addWidget(btn_refresh)

        # Edit mode toggle
        self.actions_edit_mode = False
        self.btn_edit_actions = QPushButton("Enable Edit Mode")
        self.btn_edit_actions.setStyleSheet(self.get_button_style("warning"))
        self.btn_edit_actions.clicked.connect(self.toggle_actions_edit_mode)
        btn_layout.addWidget(self.btn_edit_actions)

        # Save/Cancel buttons (hidden by default)
        self.btn_save_actions = QPushButton("Save Changes")
        self.btn_save_actions.setStyleSheet(self.get_button_style("success"))
        self.btn_save_actions.clicked.connect(self.save_actions_changes)
        self.btn_save_actions.setVisible(False)
        btn_layout.addWidget(self.btn_save_actions)

        self.btn_cancel_actions = QPushButton("Cancel")
        self.btn_cancel_actions.setStyleSheet(self.get_button_style("default"))
        self.btn_cancel_actions.clicked.connect(self.cancel_actions_edit)
        self.btn_cancel_actions.setVisible(False)
        btn_layout.addWidget(self.btn_cancel_actions)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Filter bar
        filter_bar = QHBoxLayout()
        self.actions_filter = QLineEdit()
        self.actions_filter.setPlaceholderText("Search tariff number, action, or description...")
        self.actions_filter.setStyleSheet(self.get_input_style())
        self.actions_filter.textChanged.connect(self.filter_actions_table)
        filter_bar.addWidget(self.actions_filter)
        
        self.actions_material_filter = QComboBox()
        self.actions_material_filter.addItems(["All", "Steel", "Aluminum", "Copper", "Wood"])
        self.actions_material_filter.currentTextChanged.connect(self.filter_actions_table)
        filter_bar.addWidget(QLabel("Commodity:"))
        filter_bar.addWidget(self.actions_material_filter)
        
        # Add color toggle checkbox
        self.actions_color_toggle = QCheckBox("Color by Material")
        self.actions_color_toggle.setChecked(False)  # Disabled by default
        self.actions_color_toggle.stateChanged.connect(self.filter_actions_table)
        filter_bar.addWidget(self.actions_color_toggle)
        
        layout.addLayout(filter_bar)
        
        # Table
        self.actions_table = QTableWidget()
        self.actions_table.setColumnCount(10)
        self.actions_table.setHorizontalHeaderLabels([
            "Tariff No", "Action", "Description", "Ad Valorem Rate",
            "Effective Date", "Expiration Date", "Specific Rate",
            "Additional Declaration", "Note", "Link"
        ])
        self.actions_table.horizontalHeader().setStretchLastSection(True)
        self.actions_table.setAlternatingRowColors(True)
        self.actions_table.setStyleSheet("")
        self.actions_table.setSortingEnabled(True)
        layout.addWidget(self.actions_table)
        
        # Count label
        self.actions_count_label = QLabel("Total: 0 actions")
        self.actions_count_label.setStyleSheet("font-weight:bold; padding:5px;")
        layout.addWidget(self.actions_count_label)
        
        self.tab_actions.setLayout(layout)
        
        # Load data
        self.refresh_actions_view()

    def import_actions_csv(self):
        """Import Section 232 Actions from CSV/TSV"""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Section 232 Actions File", "",
            "CSV/TSV Files (*.csv *.txt);;All Files (*.*)"
        )
        if not path:
            return
        
        try:
            # Try to detect the delimiter and read the file
            if path.lower().endswith('.txt'):
                # For .txt files, assume tab-delimited
                df = pd.read_csv(path, sep='\t', dtype=str, keep_default_na=False)
            else:
                # For .csv files, let pandas auto-detect
                df = pd.read_csv(path, dtype=str, keep_default_na=False)
            
            df = df.fillna("")
            
            # Normalize column names (strip whitespace)
            df.columns = df.columns.str.strip()
            
            # Check for required columns
            required_cols = ['Tariff No', 'Action']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                QMessageBox.warning(self, "Invalid Format", 
                    f"CSV is missing required columns: {', '.join(missing_cols)}\n\n"
                    f"Found columns: {', '.join(df.columns.tolist())}")
                return
            
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("DELETE FROM sec_232_actions")
            
            imported = 0
            for _, row in df.iterrows():
                tariff_no = str(row.get('Tariff No', '')).strip()
                action = str(row.get('Action', '')).strip()
                description = str(row.get('Tariff Description', '')).strip()
                advalorem = str(row.get('Advalorem Rate', '')).strip()
                effective = str(row.get('Effective Date', '')).strip()
                expiration = str(row.get('Expiration Date', '')).strip()
                specific = str(row.get('Specific Rate', '')).strip()
                declaration = str(row.get('Additional Declaration Required', '')).strip()
                note = str(row.get('Note', '')).strip()
                link = str(row.get('Link', '')).strip()
                
                if tariff_no and action:
                    c.execute("""INSERT OR REPLACE INTO sec_232_actions 
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                             (tariff_no, action, description, advalorem, effective, 
                              expiration, specific, declaration, note, link))
                    imported += 1
            
            conn.commit()
            conn.close()
            
            QMessageBox.information(self, "Success", 
                f"Imported {imported} Section 232 action records")
            logger.success(f"sec_232_actions table updated with {imported} records")
            self.status.setText(f"Section 232 actions imported: {imported} records")
            self.refresh_actions_view()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Import failed: {e}")
            logger.error(f"Section 232 actions import error: {e}")

    def refresh_actions_view(self):
        """Load and display all Section 232 actions from database"""
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("""SELECT tariff_no, action, description, advalorem_rate,
                                       effective_date, expiration_date, specific_rate,
                                       additional_declaration, note, link
                                FROM sec_232_actions
                                ORDER BY tariff_no""", conn)
            conn.close()
            
            self.actions_full_data = df
            self.filter_actions_table()
            
        except Exception as e:
            logger.error(f"Refresh actions view failed: {e}")
            self.actions_table.setRowCount(0)
            self.actions_count_label.setText("Error loading actions data")

    def filter_actions_table(self):
        """Filter and display actions table based on search criteria"""
        try:
            if not hasattr(self, 'actions_full_data') or self.actions_full_data.empty:
                self.actions_table.setRowCount(0)
                self.actions_count_label.setText("No actions data")
                return
            
            df = self.actions_full_data.copy()
            
            # Apply text filter
            search_text = self.actions_filter.text().strip().lower()
            if search_text:
                mask = (df['tariff_no'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['action'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['description'].astype(str).str.lower().str.contains(search_text, na=False) |
                       df['note'].astype(str).str.lower().str.contains(search_text, na=False))
                df = df[mask]
            
            # Apply material filter
            material_filter = self.actions_material_filter.currentText()
            if material_filter != "All":
                # Filter by action column containing material name
                df = df[df['action'].str.contains(material_filter, case=False, na=False)]
            
            # Populate table
            self.actions_table.setSortingEnabled(False)
            self.actions_table.setRowCount(len(df))
            
            for row_idx, (_, row) in enumerate(df.iterrows()):
                items = [
                    QTableWidgetItem(str(row['tariff_no'])),
                    QTableWidgetItem(str(row['action'])),
                    QTableWidgetItem(str(row['description'])),
                    QTableWidgetItem(str(row['advalorem_rate'])),
                    QTableWidgetItem(str(row['effective_date'])),
                    QTableWidgetItem(str(row['expiration_date'])),
                    QTableWidgetItem(str(row['specific_rate'])),
                    QTableWidgetItem(str(row['additional_declaration'])),
                    QTableWidgetItem(str(row['note'])),
                    QTableWidgetItem(str(row['link']))
                ]
                
                # Make items editable/read-only based on edit mode
                # Columns that can be edited: Action (1), Description (2), Note (8), Link (9)
                editable_columns = {1, 2, 8, 9}
                for col_idx, item in enumerate(items):
                    if col_idx in editable_columns and self.actions_edit_mode:
                        # Enable editing
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                    else:
                        # Read-only
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                # Apply color coding if toggle is enabled
                if self.actions_color_toggle.isChecked():
                    material_colors = {
                        'Steel': QColor('#e3f2fd'),
                        'Aluminum': QColor('#fff3e0'),
                        'Wood': QColor('#f1f8e9'),
                        'Copper': QColor('#ffe0b2')
                    }
                    
                    action_text = str(row['action']).upper()
                    bg_color = None
                    
                    if 'STEEL' in action_text:
                        bg_color = material_colors['Steel']
                    elif 'ALUMINUM' in action_text:
                        bg_color = material_colors['Aluminum']
                    elif 'COPPER' in action_text:
                        bg_color = material_colors['Copper']
                    elif 'WOOD' in action_text or 'LUMBER' in action_text or 'FURNITURE' in action_text:
                        bg_color = material_colors['Wood']
                    
                    # Apply background color to entire row
                    if bg_color:
                        for item in items:
                            item.setBackground(bg_color)
                
                # Highlight expired actions regardless of toggle state
                if 'EXPIRED' in str(row['note']).upper():
                    for item in items:
                        item.setForeground(QColor('#999999'))  # Gray out expired
                
                # Add items to table
                for col_idx, item in enumerate(items):
                    self.actions_table.setItem(row_idx, col_idx, item)
            
            self.actions_table.setSortingEnabled(True)
            self.actions_count_label.setText(
                f"Total: {len(df)} actions (filtered from {len(self.actions_full_data)})"
            )
            
        except Exception as e:
            logger.error(f"Filter actions table failed: {e}")
            self.actions_table.setRowCount(0)

    def toggle_actions_edit_mode(self):
        """Toggle edit mode for Section 232 Actions table"""
        self.actions_edit_mode = not self.actions_edit_mode

        if self.actions_edit_mode:
            # Entering edit mode
            self.btn_edit_actions.setText("Disable Edit Mode")
            self.btn_edit_actions.setStyleSheet(self.get_button_style("danger"))
            self.btn_save_actions.setVisible(True)
            self.btn_cancel_actions.setVisible(True)
            self.actions_filter.setEnabled(False)
            self.actions_material_filter.setEnabled(False)
            self.actions_color_toggle.setEnabled(False)

            # Store original data for cancel functionality
            if hasattr(self, 'actions_full_data'):
                self.actions_original_data = self.actions_full_data.copy()

            # Re-render table with editable cells
            self.filter_actions_table()
        else:
            # Exiting edit mode (cancel)
            self.cancel_actions_edit()

    def save_actions_changes(self):
        """Save changes made to Section 232 Actions table to database"""
        if not hasattr(self, 'actions_full_data'):
            QMessageBox.warning(self, "No Data", "No actions data to save")
            return

        try:
            # Collect all current table data
            updated_rows = []
            for row_idx in range(self.actions_table.rowCount()):
                row_data = []
                for col_idx in range(self.actions_table.columnCount()):
                    item = self.actions_table.item(row_idx, col_idx)
                    row_data.append(item.text() if item else "")
                updated_rows.append(row_data)

            # Create DataFrame from updated rows
            columns = ['tariff_no', 'action', 'description', 'advalorem_rate',
                      'effective_date', 'expiration_date', 'specific_rate',
                      'additional_declaration', 'note', 'link']
            df_updated = pd.DataFrame(updated_rows, columns=columns)

            # Update database
            conn = sqlite3.connect(str(DB_PATH))
            df_updated.to_sql('sec_232_actions', conn, if_exists='replace', index=False)
            conn.close()

            # Update internal data
            self.actions_full_data = df_updated

            # Exit edit mode
            self.actions_edit_mode = False
            self.btn_edit_actions.setText("Enable Edit Mode")
            self.btn_edit_actions.setStyleSheet(self.get_button_style("warning"))
            self.btn_save_actions.setVisible(False)
            self.btn_cancel_actions.setVisible(False)
            self.actions_filter.setEnabled(True)
            self.actions_material_filter.setEnabled(True)
            self.actions_color_toggle.setEnabled(True)

            # Refresh display
            self.filter_actions_table()

            QMessageBox.information(self, "Success", f"Saved {len(df_updated)} actions to database")
            logger.success(f"Saved {len(df_updated)} Section 232 actions to database")

        except Exception as e:
            logger.error(f"Save actions failed: {e}")
            QMessageBox.critical(self, "Save Failed", f"Error saving changes:\n{str(e)}")

    def cancel_actions_edit(self):
        """Cancel edit mode and discard changes"""
        self.actions_edit_mode = False
        self.btn_edit_actions.setText("Enable Edit Mode")
        self.btn_edit_actions.setStyleSheet(self.get_button_style("warning"))
        self.btn_save_actions.setVisible(False)
        self.btn_cancel_actions.setVisible(False)
        self.actions_filter.setEnabled(True)
        self.actions_material_filter.setEnabled(True)
        self.actions_color_toggle.setEnabled(True)

        # Restore original data if available
        if hasattr(self, 'actions_original_data'):
            self.actions_full_data = self.actions_original_data.copy()
            del self.actions_original_data

        # Refresh display
        self.filter_actions_table()

    def on_preview_value_edited(self, item):
        # Only update for Value column edits
        if item.column() != 1:
            return
        text = item.text().replace('$', '').replace(',', '').strip()
        try:
            new_val = float(text)
            if new_val < 0:
                raise ValueError()
            item.setData(Qt.UserRole, new_val)
            item.setText(f"{new_val:,.2f}")
        except Exception:
            old_val = item.data(Qt.UserRole) or 0.0
            item.setText(f"{old_val:,.2f}")
        self.recalculate_total_and_check_match()

    def add_preview_row(self):
        """Add a new empty row to the preview table"""
        if self.last_processed_df is None:
            QMessageBox.warning(self, "No Data", "Please process a shipment first before adding rows.")
            return
        
        # Disconnect signals while adding row
        self.table.blockSignals(True)
        
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Create default items for the new row
        default_mid = self.selected_mid or ""
        default_melt = str(default_mid)[:2] if default_mid else ""
        
        value_item = QTableWidgetItem("0.00")
        value_item.setData(Qt.UserRole, 0.0)
        value_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
        
        items = [
            QTableWidgetItem("NEW_PART"),  # Product No
            value_item,  # Value
            QTableWidgetItem(""),  # HTS
            QTableWidgetItem(default_mid),  # MID
            QTableWidgetItem("0.00"),  # cbp_qty
            QTableWidgetItem("CO"),  # Dec
            QTableWidgetItem(default_melt),  # Melt
            QTableWidgetItem(""),  # Cast
            QTableWidgetItem(""),  # Smelt
            QTableWidgetItem(""),  # Flag
            QTableWidgetItem("100.0%"),  # Steel%
            QTableWidgetItem(""),  # Al%
            QTableWidgetItem(""),  # Cu%
            QTableWidgetItem(""),  # Wood%
            QTableWidgetItem(""),  # Auto%
            QTableWidgetItem(""),  # Non-232%
            QTableWidgetItem("")  # 232 Status
        ]

        # Make all items editable except ratios and 232 status
        for i, item in enumerate(items):
            if i not in [10, 11, 12, 13, 14, 15, 16]:  # Not Steel%, Al%, Cu%, Wood%, Auto%, Non-232%, 232 Status
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self.table.setItem(row, i, item)
        
        self.table.blockSignals(False)
        self.recalculate_total_and_check_match()
        logger.info(f"Added new row at position {row + 1}")

    def delete_preview_row(self):
        """Delete selected row(s) from the preview table"""
        if self.last_processed_df is None:
            QMessageBox.warning(self, "No Data", "No preview data to delete.")
            return
        
        selected_rows = sorted(set(index.row() for index in self.table.selectedIndexes()), reverse=True)
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select row(s) to delete.")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete {len(selected_rows)} row(s)?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # Disconnect signals while deleting
        self.table.blockSignals(True)
        
        for row in selected_rows:
            self.table.removeRow(row)
            logger.info(f"Deleted row {row + 1}")
        
        self.table.blockSignals(False)
        self.recalculate_total_and_check_match()
        self.status.setText(f"Deleted {len(selected_rows)} row(s)")

    def copy_column_to_clipboard(self):
        """Copy selected column data to clipboard"""
        if self.last_processed_df is None:
            QMessageBox.warning(self, "No Data", "No preview data available.")
            return
        
        # Get selected cells
        selected = self.table.selectedIndexes()
        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select cells from a column to copy.")
            return
        
        # Determine if user selected a single column or multiple cells
        columns = set(index.column() for index in selected)
        
        if len(columns) == 1:
            # Single column selected - copy all values from that column
            col = list(columns)[0]
            column_data = []
            for row in range(self.table.rowCount()):
                item = self.table.item(row, col)
                if item:
                    # For Value column, use the stored float value
                    if col == 1:
                        value = item.data(Qt.UserRole)
                        column_data.append(str(value) if value is not None else "")
                    else:
                        column_data.append(item.text())
                else:
                    column_data.append("")
            
            # Copy to clipboard
            clipboard_text = "\n".join(column_data)
            QApplication.clipboard().setText(clipboard_text)
            
            # Get column name
            header = self.table.horizontalHeaderItem(col)
            col_name = header.text() if header else f"Column {col + 1}"
            QMessageBox.information(self, "Copied", f"Copied {len(column_data)} values from '{col_name}' to clipboard.")
            logger.info(f"Copied column '{col_name}' to clipboard ({len(column_data)} rows)")
        else:
            # Multiple columns or mixed selection - copy selected cells as tab-separated
            # Group by row
            by_row = {}
            for index in selected:
                row = index.row()
                col = index.column()
                if row not in by_row:
                    by_row[row] = {}
                by_row[row][col] = index
            
            # Build clipboard text
            rows_text = []
            for row in sorted(by_row.keys()):
                cells = []
                for col in sorted(by_row[row].keys()):
                    item = self.table.item(row, col)
                    if item:
                        if col == 1:  # Value column
                            value = item.data(Qt.UserRole)
                            cells.append(str(value) if value is not None else "")
                        else:
                            cells.append(item.text())
                    else:
                        cells.append("")
                rows_text.append("\t".join(cells))
            
            clipboard_text = "\n".join(rows_text)
            QApplication.clipboard().setText(clipboard_text)
            QMessageBox.information(self, "Copied", f"Copied {len(selected)} cells to clipboard (tab-separated).")
            logger.info(f"Copied {len(selected)} selected cells to clipboard")

    def select_column(self, column_index):
        """Select entire column when header is clicked"""
        self.table.clearSelection()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, column_index)
            if item:
                item.setSelected(True)

    def save_column_widths(self):
        """Save column widths to per-user settings for persistence"""
        try:
            widths = {}
            for col in range(self.table.columnCount()):
                header_text = self.table.horizontalHeaderItem(col).text()
                widths[header_text] = self.table.columnWidth(col)

            import json
            set_user_setting('column_widths', json.dumps(widths))
        except Exception as e:
            logger.debug(f"Could not save column widths: {e}")

    def load_column_widths(self):
        """Load saved column widths from per-user settings"""
        try:
            import json
            widths_json = get_user_setting('column_widths')

            if widths_json:
                widths = json.loads(widths_json)
                # Check if any width is 0 - if so, this is corrupted data, clear it
                has_zero_width = any(w == 0 for w in widths.values())
                if has_zero_width:
                    set_user_setting('column_widths', '')
                    logger.info("Cleared corrupted column widths (had 0-width columns)")
                else:
                    for col in range(self.table.columnCount()):
                        header_text = self.table.horizontalHeaderItem(col).text()
                        if header_text in widths and widths[header_text] > 20:  # Minimum 20px width
                            self.table.setColumnWidth(col, widths[header_text])
        except Exception as e:
            logger.debug(f"Could not load column widths: {e}")

    def recalculate_total_and_check_match(self):
        if self.last_processed_df is None:
            return
        total = 0.0
        for i in range(self.table.rowCount()):
            cell = self.table.item(i, 1)
            total += (cell.data(Qt.UserRole) or 0.0) if cell else 0.0
        
        # Don't update CI input - let user keep their target value
        # Just compare the preview total against the CI input
        ci_text = self.ci_input.text().replace(',', '').strip()
        try:
            target_value = float(ci_text) if ci_text else self.csv_total_value
        except:
            target_value = self.csv_total_value
        
        diff = abs(total - target_value)
        threshold = 0.01
        if diff <= threshold:
            self.process_btn.setEnabled(True)
            self.process_btn.setText("Export Worksheet")
            self.process_btn.setFocus()  # Keep focus on button so user can press Enter to export
            self.status.setText("VALUES MATCH → READY TO EXPORT")
            self.status.setStyleSheet("background:#107C10; color:white; font-weight:bold; font-size:16pt;")
        else:
            self.process_btn.setEnabled(False)
            self.process_btn.setText("Export Worksheet (Values Don't Match)")
            self.status.setText(f"Preview: ${total:,.2f} • Target: ${target_value:,.2f}")

    def _process_or_export(self):
        # If no preview yet, run processing; otherwise proceed to export
        if self.last_processed_df is None:
            self.start_processing()
        else:
            self.final_export()

    def _export_single_file(self, df_out, cols, filename, is_network, steel_mask, aluminum_mask, copper_mask, wood_mask, auto_mask, non232_mask, sec301_mask):
        """Export a single Excel file with formatting. Used by both regular export and split-by-invoice export."""
        from openpyxl.styles import PatternFill

        # Helper function to get export color from per-user settings
        def get_export_color(config_key, default_color):
            return get_user_setting(config_key, default_color)

        # Get user-selected font color from per-user settings
        font_color_hex = get_user_setting('output_font_color', '#000000')
        font_color_rgb = '00' + font_color_hex.lstrip('#').upper()

        # Get Section 232 material type colors
        steel_color = get_export_color('export_steel_color', '#4a4a4a')
        aluminum_color = get_export_color('export_aluminum_color', '#6495ED')
        copper_color = get_export_color('export_copper_color', '#B87333')
        wood_color = get_export_color('export_wood_color', '#8B4513')
        auto_color = get_export_color('export_automotive_color', '#2F4F4F')
        non232_color = get_export_color('export_non232_color', '#FF0000')

        # Create fonts for each material type
        steel_font = ExcelFont(name='Arial', size=11, color='00' + steel_color.lstrip('#').upper())
        aluminum_font = ExcelFont(name='Arial', size=11, color='00' + aluminum_color.lstrip('#').upper())
        copper_font = ExcelFont(name='Arial', size=11, color='00' + copper_color.lstrip('#').upper())
        wood_font = ExcelFont(name='Arial', size=11, color='00' + wood_color.lstrip('#').upper())
        auto_font = ExcelFont(name='Arial', size=11, color='00' + auto_color.lstrip('#').upper())
        non232_font = ExcelFont(name='Arial', size=11, color='00' + non232_color.lstrip('#').upper())
        default_font = ExcelFont(name='Arial', size=11, color=font_color_rgb)

        orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

        # Build index lists for each material type
        steel_indices = [i for i, val in enumerate(steel_mask.tolist()) if val]
        aluminum_indices = [i for i, val in enumerate(aluminum_mask.tolist()) if val]
        copper_indices = [i for i, val in enumerate(copper_mask.tolist()) if val]
        wood_indices = [i for i, val in enumerate(wood_mask.tolist()) if val]
        auto_indices = [i for i, val in enumerate(auto_mask.tolist()) if val]
        non232_indices = [i for i, val in enumerate(non232_mask.tolist()) if val]
        sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

        if is_network:
            # Write to temp file then copy
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                temp_path = Path(tmp.name)

            with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                df_out[cols].to_excel(writer, index=False)
                ws = next(iter(writer.sheets.values()))

                # Apply font to header row
                for col_idx in range(1, len(cols) + 1):
                    ws.cell(row=1, column=col_idx).font = ExcelFont(name='Arial', size=11, bold=True)

                # Apply font and background to data rows
                for row_idx in range(len(df_out)):
                    row_num = row_idx + 2
                    is_sec301 = row_idx in sec301_indices

                    if row_idx in steel_indices:
                        cell_font = steel_font
                    elif row_idx in aluminum_indices:
                        cell_font = aluminum_font
                    elif row_idx in copper_indices:
                        cell_font = copper_font
                    elif row_idx in wood_indices:
                        cell_font = wood_font
                    elif row_idx in auto_indices:
                        cell_font = auto_font
                    elif row_idx in non232_indices:
                        cell_font = non232_font
                    else:
                        cell_font = default_font

                    for col_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.font = cell_font
                        if is_sec301:
                            cell.fill = orange_fill

            # Copy to network location
            out = OUTPUT_DIR / filename
            shutil.copy2(temp_path, out)
            temp_path.unlink()
        else:
            # Direct write for local path
            out = OUTPUT_DIR / filename
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                df_out[cols].to_excel(writer, index=False)
                ws = next(iter(writer.sheets.values()))

                center_alignment = Alignment(horizontal="center", vertical="center")

                # Apply font to header row
                for col_idx in range(1, len(cols) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = default_font
                    cell.alignment = center_alignment

                # Apply font and background to data rows
                for row_num in range(2, len(df_out) + 2):
                    row_idx = row_num - 2
                    is_sec301 = row_idx in sec301_indices

                    if row_idx in steel_indices:
                        font_to_use = steel_font
                    elif row_idx in aluminum_indices:
                        font_to_use = aluminum_font
                    elif row_idx in copper_indices:
                        font_to_use = copper_font
                    elif row_idx in wood_indices:
                        font_to_use = wood_font
                    elif row_idx in auto_indices:
                        font_to_use = auto_font
                    elif row_idx in non232_indices:
                        font_to_use = non232_font
                    else:
                        font_to_use = default_font

                    for col_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.font = font_to_use
                        cell.alignment = center_alignment
                        if is_sec301:
                            cell.fill = orange_fill

                # Auto-size columns
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Exported: {out.name}")
        return out

    def final_export(self):
        if self.last_processed_df is None:
            return
        
        # Check if table has rows before attempting export
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Empty Preview", "No data to export. Please process a shipment file first.")
            return
            
        # Ensure totals match prior to export
        running_total = 0.0
        for i in range(self.table.rowCount()):
            cell = self.table.item(i, 1)
            running_total += (cell.data(Qt.UserRole) or 0.0) if cell else 0.0
        
        # Compare against CI input value (what user entered/approved)
        ci_text = self.ci_input.text().replace(',', '').strip()
        try:
            target_value = float(ci_text)
        except:
            target_value = self.csv_total_value
        
        if abs(running_total - target_value) > 0.05:
            QMessageBox.warning(self, "Totals Mismatch", f"Values don't match invoice total.\nPreview: ${running_total:,.2f}\nTarget: ${target_value:,.2f}")
            return

        out = OUTPUT_DIR / (self.last_output_filename or f"Upload_Sheet_{datetime.now():%Y%m%d_%H%M}.xlsx")
        
        # Rebuild DataFrame from current table state (handles added/deleted/edited rows)
        export_data = []
        for i in range(self.table.rowCount()):
            value_cell = self.table.item(i, 1)
            value = value_cell.data(Qt.UserRole) if value_cell else 0.0
            
            # Get ratio percentages as floats (handle empty values)
            # Column indices: 10=Steel%, 11=Al%, 12=Cu%, 13=Wood%, 14=Auto%, 15=Non-232%, 16=232 Status
            steel_text = self.table.item(i, 10).text() if self.table.item(i, 10) else ""
            aluminum_text = self.table.item(i, 11).text() if self.table.item(i, 11) else ""
            copper_text = self.table.item(i, 12).text() if self.table.item(i, 12) else ""
            wood_text = self.table.item(i, 13).text() if self.table.item(i, 13) else ""
            auto_text = self.table.item(i, 14).text() if self.table.item(i, 14) else ""

            # Parse percentages safely
            def parse_pct(text):
                if not text or text.strip() == '':
                    return 0.0
                try:
                    return float(text.replace('%', '').strip()) / 100.0
                except (ValueError, TypeError):
                    return 0.0

            steel_ratio = parse_pct(steel_text)
            aluminum_ratio = parse_pct(aluminum_text)
            copper_ratio = parse_pct(copper_text)
            wood_ratio = parse_pct(wood_text)
            auto_ratio = parse_pct(auto_text)

            # Get Non-232% ratio from column 15
            non_steel_text = self.table.item(i, 15).text() if self.table.item(i, 15) else ""
            non_steel_ratio = parse_pct(non_steel_text)

            # Get Sec301 exclusion data from last_processed_df if available
            sec301_exclusion = ""
            invoice_number = ""
            if self.last_processed_df is not None and i < len(self.last_processed_df):
                sec301_exclusion = str(self.last_processed_df.iloc[i].get('Sec301_Exclusion_Tariff', '')).strip()
                if sec301_exclusion in ['', 'nan', 'None']:
                    sec301_exclusion = ""
                # Get invoice number for split by invoice feature
                invoice_number = str(self.last_processed_df.iloc[i].get('invoice_number', '')).strip()
                if invoice_number in ['', 'nan', 'None']:
                    invoice_number = ""

            row_data = {
                'Product No': self.table.item(i, 0).text() if self.table.item(i, 0) else "",
                'ValueUSD': value,
                'HTSCode': self.table.item(i, 2).text() if self.table.item(i, 2) else "",
                'MID': self.table.item(i, 3).text() if self.table.item(i, 3) else "",
                'CalcWtNet': round(float(self.table.item(i, 4).text())) if self.table.item(i, 4) and self.table.item(i, 4).text() else 0,
                'DecTypeCd': self.table.item(i, 5).text() if self.table.item(i, 5) else "CO",
                'CountryofMelt': self.table.item(i, 6).text() if self.table.item(i, 6) else "",
                'CountryOfCast': self.table.item(i, 7).text() if self.table.item(i, 7) else "",
                'PrimCountryOfSmelt': self.table.item(i, 8).text() if self.table.item(i, 8) else "",
                'PrimSmeltFlag': self.table.item(i, 9).text() if self.table.item(i, 9) else "",
                'SteelRatio': steel_ratio,
                'AluminumRatio': aluminum_ratio,
                'CopperRatio': copper_ratio,
                'WoodRatio': wood_ratio,
                'AutoRatio': auto_ratio,
                'NonSteelRatio': non_steel_ratio,
                '_232_flag': self.table.item(i, 16).text() if self.table.item(i, 16) else "",  # Column 16 is 232_Status
                '_sec301_exclusion': sec301_exclusion,
                '_invoice_number': invoice_number
            }
            export_data.append(row_data)

        df_out = pd.DataFrame(export_data)

        # Build masks for each Section 232 material type BEFORE converting to percentage strings
        steel_mask = df_out['_232_flag'].fillna('').str.contains('232_Steel', case=False, na=False)
        aluminum_mask = df_out['_232_flag'].fillna('').str.contains('232_Aluminum', case=False, na=False)
        copper_mask = df_out['_232_flag'].fillna('').str.contains('232_Copper', case=False, na=False)
        wood_mask = df_out['_232_flag'].fillna('').str.contains('232_Wood', case=False, na=False)
        auto_mask = df_out['_232_flag'].fillna('').str.contains('232_Auto', case=False, na=False)
        non232_mask = df_out['_232_flag'].fillna('').str.contains('Non_232', case=False, na=False)

        # Build mask for Sec301 exclusion rows (for light orange background)
        sec301_mask = df_out['_sec301_exclusion'].fillna('').astype(str).str.strip().ne('') & \
                      ~df_out['_sec301_exclusion'].fillna('').astype(str).str.contains('nan|None', case=False, na=False)
        
        # Convert ratios to percentage strings for export
        df_out['SteelRatio'] = (df_out['SteelRatio'] * 100).round(1).astype(str) + "%"
        df_out['AluminumRatio'] = (df_out['AluminumRatio'] * 100).round(1).astype(str) + "%"
        df_out['CopperRatio'] = (df_out['CopperRatio'] * 100).round(1).astype(str) + "%"
        df_out['WoodRatio'] = (df_out['WoodRatio'] * 100).round(1).astype(str) + "%"
        df_out['AutoRatio'] = (df_out['AutoRatio'] * 100).round(1).astype(str) + "%"
        df_out['NonSteelRatio'] = (df_out['NonSteelRatio'] * 100).round(1).astype(str) + "%"
        df_out['232_Status'] = df_out['_232_flag'].fillna('')

        # Build initial columns list
        cols = ['Product No','ValueUSD','HTSCode','MID','CalcWtNet','DecTypeCd',
            'CountryofMelt','CountryOfCast','PrimCountryOfSmelt','PrimSmeltFlag']

        # Add ratio columns based on visibility settings
        ratio_columns = ['SteelRatio', 'AluminumRatio', 'CopperRatio', 'WoodRatio', 'AutoRatio', 'NonSteelRatio']
        for col in ratio_columns:
            # Check if column is visible (default to True if not set)
            is_visible = True
            try:
                conn = sqlite3.connect(str(DB_PATH))
                c = conn.cursor()
                c.execute("SELECT value FROM app_config WHERE key = ?", (f'export_col_visible_{col}',))
                row = c.fetchone()
                conn.close()
                if row:
                    is_visible = row[0] == 'True'
            except:
                pass

            if is_visible:
                cols.append(col)

        # Always add 232_Status at the end
        cols.append('232_Status')

        # Apply custom column mapping if set
        if hasattr(self, 'output_column_mapping') and self.output_column_mapping:
            # Create rename dictionary for columns that have custom names
            rename_dict = {}
            for col in cols:
                if col in self.output_column_mapping and self.output_column_mapping[col] != col:
                    rename_dict[col] = self.output_column_mapping[col]

            # Rename columns if custom names are defined
            if rename_dict:
                df_out = df_out.rename(columns=rename_dict)
                # Update cols list with new names
                cols = [self.output_column_mapping.get(col, col) for col in cols]
                logger.info(f"Applied custom column mapping: {rename_dict}")

        # Check if we should split by invoice number
        split_by_invoice = False
        try:
            conn = sqlite3.connect(str(DB_PATH))
            c = conn.cursor()
            c.execute("SELECT value FROM app_config WHERE key = ?", ('export_split_by_invoice',))
            row = c.fetchone()
            conn.close()
            if row:
                split_by_invoice = row[0] == 'True'
        except:
            pass

        # Get unique invoice numbers if splitting is enabled
        unique_invoices = []
        if split_by_invoice and '_invoice_number' in df_out.columns:
            unique_invoices = df_out['_invoice_number'].dropna().unique()
            unique_invoices = [inv for inv in unique_invoices if inv and str(inv).strip() not in ['', 'nan', 'None']]
            if len(unique_invoices) > 1:
                logger.info(f"Split by invoice enabled: Found {len(unique_invoices)} unique invoices")
            else:
                # Only one or no invoices, don't split
                unique_invoices = []

        try:
            t_start = time.time()

            # Show export progress indicator
            self.export_progress_widget.show()
            self.export_status_label.setText("Exporting:")
            self.export_progress_bar.setValue(0)
            QApplication.processEvents()
            
            # Check if output directory is on network (slower) or local
            output_str = str(OUTPUT_DIR)
            is_network = output_str.startswith('\\\\') or (len(output_str) > 1 and output_str[1] == ':' and not output_str.startswith('C:'))

            # Handle split by invoice if enabled
            if unique_invoices and len(unique_invoices) > 1:
                # Export multiple files, one per invoice
                exported_files = []
                total_invoices = len(unique_invoices)

                for idx, invoice_num in enumerate(unique_invoices):
                    # Filter dataframe for this invoice
                    invoice_df = df_out[df_out['_invoice_number'] == invoice_num].copy()

                    # Recalculate masks for this subset
                    inv_steel_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Steel', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_aluminum_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Aluminum', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_copper_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Copper', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_wood_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Wood', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_auto_mask = invoice_df['_232_flag'].fillna('').str.contains('232_Auto', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_non232_mask = invoice_df['_232_flag'].fillna('').str.contains('Non_232', case=False, na=False) if '_232_flag' in invoice_df.columns else pd.Series([False] * len(invoice_df))
                    inv_sec301_mask = invoice_df['_sec301_exclusion'].fillna('').astype(str).str.strip().ne('') & \
                                      ~invoice_df['_sec301_exclusion'].fillna('').astype(str).str.contains('nan|None', case=False, na=False) if '_sec301_exclusion' in invoice_df.columns else pd.Series([False] * len(invoice_df))

                    # Generate filename with invoice number and date
                    invoice_filename = f"{invoice_num}_{datetime.now():%Y%m%d}.xlsx"

                    # Update progress
                    progress_pct = int(10 + (idx / total_invoices) * 80)
                    self.export_progress_bar.setValue(progress_pct)
                    self.bottom_status.setText(f"Exporting invoice {idx + 1} of {total_invoices}: {invoice_num}")
                    QApplication.processEvents()

                    # Export this invoice's data
                    out_path = self._export_single_file(
                        invoice_df, cols, invoice_filename, is_network,
                        inv_steel_mask, inv_aluminum_mask, inv_copper_mask, inv_wood_mask,
                        inv_auto_mask, inv_non232_mask, inv_sec301_mask
                    )
                    exported_files.append(out_path.name)

                self.export_progress_bar.setValue(100)
                QApplication.processEvents()

                # Move processed CSV to Processed folder
                if self.current_csv and Path(self.current_csv).exists():
                    try:
                        source_file = Path(self.current_csv)
                        dest_file = PROCESSED_DIR / source_file.name
                        if dest_file.exists():
                            dest_file.unlink()
                        source_file.rename(dest_file)
                        logger.info(f"Moved processed file: {source_file.name} -> Processed/")
                        self.current_csv = None
                    except Exception as e:
                        logger.warning(f"Could not move CSV to Processed folder: {e}")

                self.refresh_exported_files()
                self.refresh_input_files()

                # Add "Not Found" parts to the database
                added_parts_count = self.add_not_found_parts_to_db()

                QTimer.singleShot(500, self.export_progress_widget.hide)

                # Build success message
                success_msg = f"Export complete!\nCreated {len(exported_files)} files:\n" + "\n".join(exported_files[:10])
                if len(exported_files) > 10:
                    success_msg += f"\n... and {len(exported_files) - 10} more"
                if added_parts_count > 0:
                    success_msg += f"\n\n{added_parts_count} new part(s) added to database."

                QMessageBox.information(self, "Success", success_msg)
                logger.success(f"Split export complete: {len(exported_files)} files created" + (f" ({added_parts_count} parts added to DB)" if added_parts_count > 0 else ""))
                self.clear_all()
                return

            # If network path, use local temp then copy (40x faster!)
            if is_network:
                self.bottom_status.setText("Generating export file...")
                self.export_progress_bar.setValue(10)
                QApplication.processEvents()
                
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    temp_path = Path(tmp.name)
                
                self.export_progress_bar.setValue(20)
                QApplication.processEvents()
                
                with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                    df_out[cols].to_excel(writer, index=False)
                    t_write = time.time() - t_start
                    
                    self.export_progress_bar.setValue(50)
                    QApplication.processEvents()
                    
                    # Apply Arial font to all cells and red font to non-steel rows
                    t_format_start = time.time()
                    ws = next(iter(writer.sheets.values()))

                    # Set Arial font for all cells (including header)
                    from openpyxl.styles import PatternFill

                    # Helper function to get export color from per-user settings
                    def get_export_color(config_key, default_color):
                        return get_user_setting(config_key, default_color)

                    # Get user-selected font color from per-user settings
                    font_color_hex = get_user_setting('output_font_color', '#000000')
                    font_color_rgb = '00' + font_color_hex.lstrip('#').upper()

                    # Get Section 232 material type colors
                    steel_color = get_export_color('export_steel_color', '#4a4a4a')
                    aluminum_color = get_export_color('export_aluminum_color', '#6495ED')
                    copper_color = get_export_color('export_copper_color', '#B87333')
                    wood_color = get_export_color('export_wood_color', '#8B4513')
                    auto_color = get_export_color('export_automotive_color', '#2F4F4F')
                    non232_color = get_export_color('export_non232_color', '#FF0000')

                    # Create fonts for each material type
                    steel_font = ExcelFont(name='Arial', size=11, color='00' + steel_color.lstrip('#').upper())
                    aluminum_font = ExcelFont(name='Arial', size=11, color='00' + aluminum_color.lstrip('#').upper())
                    copper_font = ExcelFont(name='Arial', size=11, color='00' + copper_color.lstrip('#').upper())
                    wood_font = ExcelFont(name='Arial', size=11, color='00' + wood_color.lstrip('#').upper())
                    auto_font = ExcelFont(name='Arial', size=11, color='00' + auto_color.lstrip('#').upper())
                    non232_font = ExcelFont(name='Arial', size=11, color='00' + non232_color.lstrip('#').upper())
                    default_font = ExcelFont(name='Arial', size=11, color=font_color_rgb)

                    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Light orange

                    # Apply font to header row
                    for col_idx in range(1, len(cols) + 1):
                        ws.cell(row=1, column=col_idx).font = ExcelFont(name='Arial', size=11, bold=True)

                    # Build index lists for each material type
                    steel_indices = [i for i, val in enumerate(steel_mask.tolist()) if val]
                    aluminum_indices = [i for i, val in enumerate(aluminum_mask.tolist()) if val]
                    copper_indices = [i for i, val in enumerate(copper_mask.tolist()) if val]
                    wood_indices = [i for i, val in enumerate(wood_mask.tolist()) if val]
                    auto_indices = [i for i, val in enumerate(auto_mask.tolist()) if val]
                    non232_indices = [i for i, val in enumerate(non232_mask.tolist()) if val]
                    sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

                    # Apply font and background to data rows
                    for row_idx in range(len(df_out)):
                        row_num = row_idx + 2
                        is_sec301 = row_idx in sec301_indices

                        # Determine font based on material type
                        if row_idx in steel_indices:
                            cell_font = steel_font
                        elif row_idx in aluminum_indices:
                            cell_font = aluminum_font
                        elif row_idx in copper_indices:
                            cell_font = copper_font
                        elif row_idx in wood_indices:
                            cell_font = wood_font
                        elif row_idx in auto_indices:
                            cell_font = auto_font
                        elif row_idx in non232_indices:
                            cell_font = non232_font
                        else:
                            cell_font = default_font

                        for col_idx in range(1, len(cols) + 1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.font = cell_font
                            # Apply light orange background for Sec301 exclusions
                            if is_sec301:
                                cell.fill = orange_fill
                    
                    t_format = time.time() - t_format_start
                
                self.export_progress_bar.setValue(70)
                QApplication.processEvents()
                
                # Copy to network location
                self.bottom_status.setText("Copying to network location...")
                self.export_status_label.setText("Copying:")
                t_copy_start = time.time()
                out = OUTPUT_DIR / (self.last_output_filename or f"Upload_Sheet_{datetime.now():%Y%m%d_%H%M}.xlsx")
                shutil.copy2(temp_path, out)
                temp_path.unlink()
                t_copy = time.time() - t_copy_start
                
                self.export_progress_bar.setValue(90)
                QApplication.processEvents()
                
                t_total = time.time() - t_start
                logger.info(f"Export timing - Write: {t_write:.2f}s, Format: {t_format:.2f}s, Copy: {t_copy:.2f}s, Total: {t_total:.2f}s")
            else:
                # Local path - direct write
                self.export_progress_bar.setValue(20)
                QApplication.processEvents()
                
                out = OUTPUT_DIR / (self.last_output_filename or f"Upload_Sheet_{datetime.now():%Y%m%d_%H%M}.xlsx")
                with pd.ExcelWriter(out, engine='openpyxl') as writer:
                    df_out[cols].to_excel(writer, index=False)
                    t_write = time.time() - t_start
                    
                    self.export_progress_bar.setValue(60)
                    QApplication.processEvents()
                    
                    # Apply formatting: Arial font, center alignment, auto-sized columns
                    t_format_start = time.time()
                    ws = next(iter(writer.sheets.values()))

                    # Create font, fill, and alignment styles
                    from openpyxl.styles import PatternFill

                    # Helper function to get export color from per-user settings
                    def get_export_color(config_key, default_color):
                        return get_user_setting(config_key, default_color)

                    # Get user-selected font color from per-user settings
                    font_color_hex = get_user_setting('output_font_color', '#000000')
                    font_color_rgb = '00' + font_color_hex.lstrip('#').upper()

                    # Get Section 232 material type colors
                    steel_color = get_export_color('export_steel_color', '#4a4a4a')
                    aluminum_color = get_export_color('export_aluminum_color', '#6495ED')
                    copper_color = get_export_color('export_copper_color', '#B87333')
                    wood_color = get_export_color('export_wood_color', '#8B4513')
                    auto_color = get_export_color('export_automotive_color', '#2F4F4F')
                    non232_color = get_export_color('export_non232_color', '#FF0000')

                    # Create fonts for each material type
                    steel_font = ExcelFont(name='Arial', color='00' + steel_color.lstrip('#').upper())
                    aluminum_font = ExcelFont(name='Arial', color='00' + aluminum_color.lstrip('#').upper())
                    copper_font = ExcelFont(name='Arial', color='00' + copper_color.lstrip('#').upper())
                    wood_font = ExcelFont(name='Arial', color='00' + wood_color.lstrip('#').upper())
                    auto_font = ExcelFont(name='Arial', color='00' + auto_color.lstrip('#').upper())
                    non232_font = ExcelFont(name='Arial', color='00' + non232_color.lstrip('#').upper())
                    normal_font = ExcelFont(name="Arial", color=font_color_rgb)

                    center_alignment = Alignment(horizontal="center", vertical="center")
                    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Light orange

                    # Build index lists for each material type
                    steel_indices = [i for i, val in enumerate(steel_mask.tolist()) if val]
                    aluminum_indices = [i for i, val in enumerate(aluminum_mask.tolist()) if val]
                    copper_indices = [i for i, val in enumerate(copper_mask.tolist()) if val]
                    wood_indices = [i for i, val in enumerate(wood_mask.tolist()) if val]
                    auto_indices = [i for i, val in enumerate(auto_mask.tolist()) if val]
                    non232_indices = [i for i, val in enumerate(non232_mask.tolist()) if val]
                    sec301_indices = [i for i, val in enumerate(sec301_mask.tolist()) if val]

                    # Apply font and background to data rows
                    for row_num in range(2, len(df_out) + 2):  # Start at 2 (after header)
                        row_idx = row_num - 2
                        is_sec301 = row_idx in sec301_indices

                        # Determine font based on material type
                        if row_idx in steel_indices:
                            font_to_use = steel_font
                        elif row_idx in aluminum_indices:
                            font_to_use = aluminum_font
                        elif row_idx in copper_indices:
                            font_to_use = copper_font
                        elif row_idx in wood_indices:
                            font_to_use = wood_font
                        elif row_idx in auto_indices:
                            font_to_use = auto_font
                        elif row_idx in non232_indices:
                            font_to_use = non232_font
                        else:
                            font_to_use = normal_font

                        for col_idx in range(1, len(cols) + 1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.font = font_to_use
                            cell.alignment = center_alignment
                            # Apply light orange background for Sec301 exclusions
                            if is_sec301:
                                cell.fill = orange_fill

                    # Apply Arial font and center alignment to header row
                    for col_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = normal_font
                        cell.alignment = center_alignment

                    # Auto-size columns based on content with padding
                    for col_idx, column in enumerate(ws.columns, 1):
                        max_length = 0
                        column_letter = ws.cell(row=1, column=col_idx).column_letter

                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass

                        # Add padding (2 extra characters) and set column width
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column_letter].width = adjusted_width

                    t_format = time.time() - t_format_start
                
                self.export_progress_bar.setValue(90)
                QApplication.processEvents()
                
                t_total = time.time() - t_start
                logger.info(f"Export timing - Write: {t_write:.2f}s, Format: {t_format:.2f}s, Total: {t_total:.2f}s")
            
            self.export_progress_bar.setValue(100)
            QApplication.processEvents()
            
            # Move processed CSV to Processed folder
            if self.current_csv and Path(self.current_csv).exists():
                try:
                    source_file = Path(self.current_csv)
                    dest_file = PROCESSED_DIR / source_file.name
                    
                    # If destination exists, remove it first
                    if dest_file.exists():
                        dest_file.unlink()
                    
                    source_file.rename(dest_file)
                    logger.info(f"Moved processed file: {source_file.name} -> Processed/")
                    self.current_csv = None
                except Exception as e:
                    logger.warning(f"Could not move CSV to Processed folder: {e}")
            
            self.refresh_exported_files()
            self.refresh_input_files()  # Refresh to show file moved

            # Add "Not Found" parts to the database
            added_parts_count = self.add_not_found_parts_to_db()

            # Hide progress indicator after brief delay
            QTimer.singleShot(500, self.export_progress_widget.hide)

            # Build success message
            success_msg = f"Export complete!\nSaved: {out.name}"
            if added_parts_count > 0:
                success_msg += f"\n\n{added_parts_count} new part(s) added to database."

            QMessageBox.information(self, "Success", success_msg)
            logger.success(f"Export complete: {out.name}" + (f" ({added_parts_count} parts added to DB)" if added_parts_count > 0 else ""))
        except Exception as e:
            self.export_progress_widget.hide()
            QMessageBox.critical(self, "Export Failed", str(e))
            return
        self.clear_all()

    def log_context_menu(self, pos):
        menu = QMenu()
        copy_action = menu.addAction("Copy")
        action = menu.exec_(self.log_text.mapToGlobal(pos))
        if action == copy_action:
            self.log_text.copy()

    def copy_log_to_clipboard(self):
        QApplication.clipboard().setText(self.log_text.toPlainText())
        QMessageBox.information(self, "Copied", "Log copied to clipboard!")

    def update_log(self):
        self.log_text.setPlainText(logger.get_logs())
        sb = self.log_text.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _install_preview_shortcuts(self):
        # Install Ctrl+B on the preview table to toggle bold on selected cells
        try:
            self._bold_shortcut = QShortcut(QKeySequence("Ctrl+B"), self.table)
            self._bold_shortcut.setContext(Qt.WidgetWithChildrenShortcut)
            self._bold_shortcut.activated.connect(self.toggle_preview_bold)
        except Exception:
            pass

    def toggle_preview_bold(self):
        items = self.table.selectedItems()
        if not items:
            return
        # Toggle based on the first selected item's current bold state
        target_bold = not items[0].font().bold()
        for it in items:
            f = it.font()
            f.setBold(target_bold)
            it.setFont(f)

    def load_available_mids(self):
        try:
            conn = sqlite3.connect(str(DB_PATH))
            df = pd.read_sql("SELECT DISTINCT mid FROM parts_master WHERE mid IS NOT NULL AND mid != '' ORDER BY mid", conn)
            conn.close()
            self.available_mids = df['mid'].tolist()
            if self.available_mids:
                self.mid_combo.clear()
                self.mid_combo.addItem("-- Select MID --")  # Placeholder item
                self.mid_combo.addItems(self.available_mids)
                self.mid_combo.setCurrentIndex(0)  # Start with placeholder
                self.selected_mid = ""  # No default selection
        except Exception as e:
            logger.error(f"MID load failed: {e}")

    def on_mid_changed(self, text):
        """Handle MID selection change"""
        if text and text != "-- Select MID --":
            self.selected_mid = text
        else:
            self.selected_mid = ""

    def update_export_invoice_total(self):
        """Update the invoice total display when a file is selected"""
        selected_items = self.exports_list.selectedItems()
        if not selected_items:
            self.export_invoice_total.setText("")
            return

        filename = selected_items[0].text()
        filepath = OUTPUT_DIR / filename

        try:
            # Read the Excel file and sum the ValueUSD column
            df = pd.read_excel(filepath, engine='openpyxl')

            # Try to find the value column (might be named differently based on mapping)
            value_col = None
            for col in df.columns:
                if 'value' in col.lower() or 'usd' in col.lower():
                    value_col = col
                    break

            if value_col:
                total = pd.to_numeric(df[value_col], errors='coerce').sum()
                self.export_invoice_total.setText(f"${total:,.2f}")
            else:
                self.export_invoice_total.setText("N/A")
        except Exception as e:
            logger.debug(f"Could not read invoice total from {filename}: {e}")
            self.export_invoice_total.setText("Error")

    def refresh_exported_files(self):
        """Load and display exported files from Output folder"""
        try:
            # Remember current selection and focus state
            current_item = self.exports_list.currentItem()
            current_file = current_item.text() if current_item else None
            had_focus = self.exports_list.hasFocus()

            # Block signals during refresh to prevent triggering events
            self.exports_list.blockSignals(True)
            self.exports_list.clear()
            if OUTPUT_DIR.exists():
                # Include all xlsx files (Upload_Sheet_* and split-by-invoice files)
                files = sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
                for f in files[:20]:  # Show last 20 files
                    self.exports_list.addItem(f.name)

            # Restore selection if the file still exists
            if current_file:
                items = self.exports_list.findItems(current_file, Qt.MatchExactly)
                if items:
                    self.exports_list.setCurrentItem(items[0])
            # If widget had focus but no previous selection, select first item
            elif had_focus and self.exports_list.count() > 0:
                self.exports_list.setCurrentRow(0)

            # Unblock signals after refresh complete
            self.exports_list.blockSignals(False)
        except Exception as e:
            logger.error(f"Refresh exports failed: {e}")
            # Make sure to unblock signals even on error
            self.exports_list.blockSignals(False)

    def refresh_input_files(self):
        """Load and display CSV/Excel files from Input folder"""
        try:
            # Remember current selection and focus state
            current_item = self.input_files_list.currentItem()
            current_file = current_item.text() if current_item else None
            had_focus = self.input_files_list.hasFocus()

            # Block signals during refresh to prevent triggering events
            self.input_files_list.blockSignals(True)
            self.input_files_list.clear()
            if INPUT_DIR.exists():
                # Combine CSV, XLSX, and XLS files
                csv_files = list(INPUT_DIR.glob("*.csv"))
                xlsx_files = list(INPUT_DIR.glob("*.xlsx"))
                xls_files = list(INPUT_DIR.glob("*.xls"))
                files = sorted(csv_files + xlsx_files + xls_files, key=lambda p: p.stat().st_mtime, reverse=True)
                for f in files[:50]:  # Show up to 50 files
                    self.input_files_list.addItem(f.name)
            
            # Restore selection if the file still exists
            if current_file:
                items = self.input_files_list.findItems(current_file, Qt.MatchExactly)
                if items:
                    self.input_files_list.setCurrentItem(items[0])
            # If widget had focus but no previous selection, select first item
            elif had_focus and self.input_files_list.count() > 0:
                self.input_files_list.setCurrentRow(0)

            # Unblock signals after refresh complete
            self.input_files_list.blockSignals(False)
        except Exception as e:
            logger.error(f"Refresh input files failed: {e}")
            # Make sure to unblock signals even on error
            self.input_files_list.blockSignals(False)

    def load_selected_input_file(self, item):
        """Load the selected CSV file from Input folder"""
        # Check if a map profile is selected first
        current_profile = self.profile_combo.currentText()
        if not current_profile or current_profile == "-- Select Profile --":
            QMessageBox.warning(self, "No Profile", "Please select a Map Profile first.")
            # Re-enable input fields after modal dialog
            self._enable_input_fields()
            return
        
        try:
            file_path = INPUT_DIR / item.text()
            if file_path.exists():
                self.current_csv = str(file_path)
                self.file_label.setText(file_path.name)
                
                # Clear previous processing state when loading new file
                self.last_processed_df = None
                self.table.setRowCount(0)

                # Get header row value from profile or input field
                header_row = 0  # Default: first row is header
                # First check if there's a profile header row loaded
                if hasattr(self, 'profile_header_row') and self.profile_header_row:
                    header_row = max(0, self.profile_header_row - 1)
                # Otherwise check input field (for Invoice Mapping Profiles tab)
                elif hasattr(self, 'header_row_input') and self.header_row_input.text().strip():
                    try:
                        header_row_value = int(self.header_row_input.text().strip())
                        header_row = max(0, header_row_value - 1)
                    except ValueError:
                        header_row = 0

                # Read total value - handle both CSV and Excel files
                col_map = {v: k for k, v in self.shipment_mapping.items()}
                if file_path.suffix.lower() == '.xlsx':
                    df = pd.read_excel(file_path, dtype=str, header=header_row)
                else:
                    df = pd.read_csv(file_path, dtype=str, header=header_row)

                # DEBUG: Log DataFrame info
                logger.info(f"[CSV TOTAL DEBUG] DataFrame columns after reading with header={header_row}: {df.columns.tolist()}")
                logger.info(f"[CSV TOTAL DEBUG] Shipment mapping: {self.shipment_mapping}")

                # Calculate total using original column name before renaming
                # Only sum rows that have a part number to exclude total/subtotal rows
                value_column = None
                part_number_column = None

                if 'value_usd' in self.shipment_mapping:
                    original_col_name = self.shipment_mapping['value_usd']
                    logger.info(f"[CSV TOTAL DEBUG] Looking for column '{original_col_name}' in DataFrame")
                    logger.info(f"[CSV TOTAL DEBUG] Column exists: {original_col_name in df.columns}")
                    if original_col_name in df.columns:
                        value_column = original_col_name
                        logger.info(f"[CSV TOTAL DEBUG] First 5 values in {original_col_name}: {df[original_col_name].head().tolist()}")
                else:
                    logger.warning("[CSV TOTAL DEBUG] 'value_usd' not found in shipment_mapping")

                # Get part number column to filter rows
                if 'part_number' in self.shipment_mapping:
                    part_number_col_name = self.shipment_mapping['part_number']
                    if part_number_col_name in df.columns:
                        part_number_column = part_number_col_name
                        logger.info(f"[CSV TOTAL DEBUG] Part number column: '{part_number_column}'")

                if value_column:
                    # Filter to only rows that have a part number (exclude total/subtotal rows)
                    if part_number_column:
                        df_filtered = df[df[part_number_column].notna() & (df[part_number_column].astype(str).str.strip() != '')]
                        logger.info(f"[CSV TOTAL DEBUG] Rows with part numbers: {len(df_filtered)} of {len(df)} total rows")
                        total = pd.to_numeric(df_filtered[value_column], errors='coerce').sum()
                    else:
                        # If no part number column, sum all rows (old behavior)
                        logger.warning("[CSV TOTAL DEBUG] No part_number column mapped, summing all rows")
                        total = pd.to_numeric(df[value_column], errors='coerce').sum()

                    self.csv_total_value = round(total, 2)
                    logger.info(f"[CSV TOTAL DEBUG] Calculated total: ${self.csv_total_value:,.2f}")

                    # Update the invoice check - this will display the total and control button state
                    self.update_invoice_check()
                else:
                    logger.warning("[CSV TOTAL DEBUG] No value column found, CSV total will be 0.00")

                # Rename columns for other uses
                df = df.rename(columns=col_map)

                self.bottom_status.setText(f"Loaded: {file_path.name}")
                logger.info(f"Loaded: {file_path.name}")

                # Ensure input fields remain editable after loading
                self._enable_input_fields()

                # CRITICAL FIX: Modal dialogs fix the keyboard lock by triggering Qt event processing
                # When "No Profile" dialog is dismissed, _enable_input_fields() is called
                # and the modal event loop refresh fixes keyboard routing
                # Simulate the same effect here by processing events + enabling fields
                QApplication.processEvents()  # Process any pending Qt events
                self._enable_input_fields()  # Re-enable fields (same as after dialog dismissal)

                # Force widget style refresh (same as reselecting profile)
                for widget in [self.ci_input, self.wt_input]:
                    widget.style().unpolish(widget)
                    widget.style().polish(widget)

                # Move keyboard focus to CI input
                self.input_files_list.clearFocus()  # Remove focus from list
                self.ci_input.setFocus(Qt.OtherFocusReason)  # Give focus to CI input
                logger.info(f"[FOCUS] ci_input.hasFocus()={self.ci_input.hasFocus()}")
        except Exception as e:
            logger.error(f"Load input file failed: {e}")
            self.status.setText(f"Error loading file: {e}")
            QMessageBox.warning(self, "Error", f"Could not load file:\n{e}")
            # Ensure fields stay enabled even after error
            self._enable_input_fields()

    def open_exported_file(self, item):
        """Open the selected exported file with user's preferred application"""
        try:
            file_path = OUTPUT_DIR / item.text()
            if file_path.exists():
                import subprocess
                if sys.platform == 'win32':
                    os.startfile(str(file_path))
                elif sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', str(file_path)])
                else:  # Linux and other Unix-like systems
                    # Check user preference for Excel viewer (per-user setting)
                    viewer_preference = get_user_setting('excel_viewer', 'System Default')

                    if viewer_preference == "Gnumeric":
                        subprocess.run(['gnumeric', str(file_path)])
                    else:
                        subprocess.run(['xdg-open', str(file_path)])
        except Exception as e:
            logger.error(f"Open file failed: {e}")
            QMessageBox.warning(self, "Error", f"Could not open file:\n{e}")

    def setup_auto_refresh(self):
        """Set up automatic refresh timers for file lists"""
        # Track last modification times to avoid unnecessary refreshes
        self.last_input_mtime = 0
        self.last_output_mtime = 0
        
        # Auto-refresh input files every 10 seconds
        self.input_refresh_timer = QTimer(self)
        self.input_refresh_timer.timeout.connect(self.refresh_input_files_light)
        self.input_refresh_timer.start(10000)  # 10000ms = 10 seconds
        
        # Auto-refresh exported files every 10 seconds
        self.export_refresh_timer = QTimer(self)
        self.export_refresh_timer.timeout.connect(self.refresh_exported_files_light)
        self.export_refresh_timer.start(10000)  # 10000ms = 10 seconds
        
        # Clean up old exported files every 30 minutes
        self.cleanup_timer = QTimer(self)
        self.cleanup_timer.timeout.connect(self.cleanup_old_exports)
        self.cleanup_timer.start(1800000)  # 1800000ms = 30 minutes
        
        # Run cleanup once on startup
        QTimer.singleShot(5000, self.cleanup_old_exports)  # Wait 5 seconds after startup
        
        logger.info("Auto-refresh enabled for file lists (10 second interval)")
    
    def refresh_input_files_light(self):
        """Lightweight refresh - only update if on Process Shipment tab"""
        try:
            # Only refresh if on Process Shipment tab (tab index 0)
            if self.tabs.currentIndex() != 0:
                return

            if not INPUT_DIR.exists():
                return

            # Always refresh - directory mtime is unreliable on network drives
            self.refresh_input_files()
        except:
            pass  # Silently ignore errors during auto-refresh

    def refresh_exported_files_light(self):
        """Lightweight refresh - only update if on Process Shipment tab"""
        try:
            # Only refresh if on Process Shipment tab (tab index 0)
            if self.tabs.currentIndex() != 0:
                return

            if not OUTPUT_DIR.exists():
                return

            # Always refresh - directory mtime is unreliable on network drives
            self.refresh_exported_files()
        except:
            pass  # Silently ignore errors during auto-refresh
    
    def cleanup_old_exports(self):
        """Move exported files older than 3 days to Output/Processed directory"""
        try:
            if not OUTPUT_DIR.exists():
                return
            
            # Ensure Output/Processed directory exists
            OUTPUT_PROCESSED_DIR.mkdir(exist_ok=True)
            
            from datetime import datetime, timedelta
            cutoff_date = datetime.now() - timedelta(days=3)
            moved_count = 0
            
            # Process all .xlsx files in Output directory
            for file_path in OUTPUT_DIR.glob("*.xlsx"):
                try:
                    # Skip if it's a directory
                    if file_path.is_dir():
                        continue
                    
                    # Get file modification time
                    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    
                    # Move if older than 3 days
                    if file_mtime < cutoff_date:
                        dest_path = OUTPUT_PROCESSED_DIR / file_path.name
                        
                        # Handle duplicate filenames
                        if dest_path.exists():
                            base_name = file_path.stem
                            ext = file_path.suffix
                            counter = 1
                            while dest_path.exists():
                                dest_path = OUTPUT_PROCESSED_DIR / f"{base_name}_{counter}{ext}"
                                counter += 1
                        
                        # Move the file
                        shutil.move(str(file_path), str(dest_path))
                        moved_count += 1
                        logger.info(f"Moved old export to Processed: {file_path.name}")
                except Exception as e:
                    logger.warning(f"Failed to move {file_path.name}: {e}")
            
            if moved_count > 0:
                logger.info(f"Cleanup: Moved {moved_count} exported file(s) older than 3 days to Output/Processed")
                # Refresh the exported files list if we're on the Process Shipment tab
                if self.tabs.currentIndex() == 0:
                    self.refresh_exported_files()
        except Exception as e:
            logger.error(f"Cleanup old exports failed: {e}")

if __name__ == "__main__":
    import traceback
    app = QApplication(sys.argv)
    try:
        # Theme will be set by apply_saved_theme() during initialization
        icon_path = TEMP_RESOURCES_DIR / "banner_bg.png"
        if not icon_path.exists():
            icon_path = TEMP_RESOURCES_DIR / "icon.ico"
        if icon_path.exists():
            app.setWindowIcon(QIcon(str(icon_path)))
        
        # Create and show splash screen
        splash_widget = QWidget()
        splash_widget.setFixedSize(500, 300)
        splash_widget.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        splash_widget.setAttribute(Qt.WA_TranslucentBackground)
        splash_layout = QVBoxLayout(splash_widget)
        splash_layout.setContentsMargins(0, 0, 0, 0)
        splash_container = QWidget()
        splash_container.setStyleSheet("""
            QWidget {
                background-color: #333333;
                border: 3px solid #0078D4;
                border-radius: 15px;
            }
        """)
        container_layout = QVBoxLayout(splash_container)
        container_layout.setContentsMargins(30, 30, 30, 30)
        container_layout.setSpacing(20)
        title_label = QLabel(f"<h1 style='color: #0078D4;'>{APP_NAME}</h1>")
        title_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(title_label)
        splash_message = QLabel("Initializing application\nPlease wait...")
        splash_message.setStyleSheet("color: #f3f3f3; font-size: 12pt; font-weight: bold;")
        splash_message.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(splash_message)
        splash_progress = QProgressBar()
        splash_progress.setRange(0, 100)
        splash_progress.setValue(0)
        splash_progress.setTextVisible(True)
        splash_progress.setFormat("%p%")
        splash_progress.setFixedHeight(20)
        splash_progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #555;
                border-radius: 5px;
                background-color: #1e1e1e;
                text-align: center;
                color: white;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
                border-radius: 3px;
            }
        """)
        container_layout.addWidget(splash_progress)
        splash_layout.addWidget(splash_container)
        splash_widget.show()
        screen_geo = app.desktop().availableGeometry()
        splash_widget.move(
            (screen_geo.width() - splash_widget.width()) // 2,
            (screen_geo.height() - splash_widget.height()) // 2
        )
        app.processEvents()
        
        logger.info("Application started")
        splash_message.setText("Creating main window...\nPlease wait...")
        splash_progress.setValue(10)
        app.processEvents()
        
        win = DerivativeMill()
        win.setWindowTitle(APP_NAME)
        splash_widget.close()
        win.show()

        def finish_initialization():
            win.status.setText("Initializing application...")
            win.load_config_paths()
            win.status.setText("Applying theme...")
            win.apply_saved_theme()
            win.status.setText("Loading MIDs...")
            win.load_available_mids()
            win.status.setText("Loading profiles...")
            win.load_mapping_profiles()
            win.status.setText("Loading export profiles...")
            win.load_output_mapping_profiles()
            win.status.setText("Scanning input files...")
            win.refresh_input_files()
            win.status.setText("Starting auto-refresh...")
            win.setup_auto_refresh()
            win.status.setText("Ready")
            # Final aggressive enable after all initialization
            QTimer.singleShot(0, win._enable_input_fields)
            QTimer.singleShot(100, win._enable_input_fields)
            QTimer.singleShot(500, win._enable_input_fields)
            QTimer.singleShot(1000, win._enable_input_fields)
            # Check for updates after a short delay (non-blocking)
            QTimer.singleShot(2000, win.check_for_updates_startup)

        # Start initialization after window is shown
        QTimer.singleShot(100, finish_initialization)
        sys.exit(app.exec_())
    except Exception as e:
        error_msg = f"Unhandled Exception:\n{str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        try:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(None, "Application Error", error_msg)
        except Exception:
            pass
        sys.exit(1)
